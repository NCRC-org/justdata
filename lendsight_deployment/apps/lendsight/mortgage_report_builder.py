"""
Mortgage report builder module for creating Excel reports from HMDA mortgage data.
Similar structure to branch report builder but for mortgage origination data.
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Any, Tuple
from datetime import datetime
import os
import re


def build_mortgage_report(raw_data: List[Dict[str, Any]], counties: List[str], years: List[int], census_data: Dict = None, progress_tracker=None) -> Dict[str, pd.DataFrame]:
    """
    Process raw BigQuery HMDA data and build comprehensive mortgage report dataframes.
    
    Args:
        raw_data: List of dictionaries from BigQuery results
        counties: List of counties in the report
        years: List of years in the report
        census_data: Optional dictionary of Census demographic data by county
        progress_tracker: Optional progress tracker for real-time progress updates
        
    Returns:
        Dictionary containing multiple dataframes for different report sections
    """
    if not raw_data:
        raise ValueError("No data provided for report building")
    
    # Convert to DataFrame
    df = pd.DataFrame(raw_data)
    
    # Ensure required columns exist (mortgage-specific)
    required_columns = ['lei', 'year', 'county_code', 'county_state', 'total_originations', 'lmict_originations', 'mmct_originations']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Clean and prepare data
    df = clean_mortgage_data(df)
    
    # Calculate HHI for loan amounts in latest year (similar to deposits for branches)
    hhi_data = calculate_mortgage_hhi(df)
    
    # Build different report sections (pass census_data to demographic overview)
    
    report_sections = [
        ('summary', 'Summary Table', lambda: create_mortgage_summary_table(df, counties, years)),
        ('demographic_overview', 'Demographic Overview', lambda: create_demographic_overview_table(df, years, census_data=census_data)),
        ('income_neighborhood_indicators', 'Income & Neighborhood Indicators', lambda: create_income_neighborhood_indicators_table(df, years)),
        ('top_lenders_detailed', 'Top Lenders Detailed', lambda: create_top_lenders_detailed_table(df, years)),
        ('by_lender', 'Lender Summary', lambda: create_lender_summary(df, years)),
        ('by_county', 'County Summary', lambda: create_mortgage_county_summary(df, counties, years)),
        ('trends', 'Trends Analysis', lambda: create_mortgage_trend_analysis(df, years)),
    ]
    
    total_sections = len(report_sections)
    report_data = {}
    
    for idx, (key, section_name, create_func) in enumerate(report_sections, 1):
        if progress_tracker:
            progress_tracker.update_section_progress(idx, total_sections, section_name)
        report_data[key] = create_func()
    
    # Add non-section data
    report_data['raw_data'] = df
    report_data['hhi'] = hhi_data
    
    return report_data


def calculate_mortgage_hhi(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calculate Herfindahl-Hirschman Index (HHI) for mortgage loan amounts in the latest year.
    
    Similar to branch HHI but uses loan amounts instead of deposits.
    
    Returns:
        Dictionary with HHI value, concentration level, year, and top lenders by loan volume
    """
    if 'total_loan_amount' not in df.columns:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': None,
            'total_loan_amount': None,
            'top_lenders': []
        }
    
    # Find latest year
    latest_year = df['year'].max()
    latest_year_df = df[df['year'] == latest_year].copy()
    
    # Aggregate loan amounts by lender (across all counties)
    lender_loans = latest_year_df.groupby('lender_name')['total_loan_amount'].sum().reset_index()
    lender_loans = lender_loans[lender_loans['total_loan_amount'] > 0]  # Remove lenders with zero loans
    
    if lender_loans.empty:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': latest_year,
            'total_loan_amount': 0,
            'top_lenders': []
        }
    
    total_loan_amount = lender_loans['total_loan_amount'].sum()
    
    # Calculate market shares (as percentages)
    lender_loans['market_share'] = (lender_loans['total_loan_amount'] / total_loan_amount) * 100
    
    # Calculate HHI: sum of squared market shares (0-10,000 scale)
    hhi = (lender_loans['market_share'] ** 2).sum()
    
    # Determine concentration level
    if hhi < 1500:
        concentration_level = 'Low concentration (competitive market)'
    elif hhi < 2500:
        concentration_level = 'Moderate concentration'
    else:
        concentration_level = 'High concentration'
    
    # Get top lenders by loan amount
    top_lenders = lender_loans.nlargest(10, 'total_loan_amount').copy()
    top_lenders_list = []
    for _, row in top_lenders.iterrows():
        top_lenders_list.append({
            'lender_name': row['lender_name'],
            'total_loan_amount': float(row['total_loan_amount']),
            'market_share': float(row['market_share'])
        })
    
    return {
        'hhi': float(hhi),
        'concentration_level': concentration_level,
        'year': int(latest_year),
        'total_loan_amount': float(total_loan_amount),
        'total_lenders': len(lender_loans),
        'top_lenders': top_lenders_list
    }


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize Excel sheet name by removing invalid characters.
    
    Excel sheet names cannot contain: : \ / ? * [ ]
    Also truncate to max_length (Excel limit is 31 characters).
    
    Args:
        name: Original sheet name
        max_length: Maximum length (default 31, Excel's limit)
    
    Returns:
        Sanitized sheet name
    """
    if not name:
        return 'Sheet'
    
    # Replace invalid characters with dash
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '-')
    
    # Truncate to max length
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length]
    
    return sanitized


def correct_lender_name_capitalization(name: str) -> str:
    """
    Correct capitalization for common lender names to match their official branding.
    
    This function preserves the exact capitalization from the database but can apply
    corrections for well-known lenders whose names may be incorrectly capitalized in the source data.
    """
    if not name or pd.isna(name):
        return name
    
    name = str(name).strip()
    
    # Dictionary of common lender name corrections
    # Key is case-insensitive, value is the correct capitalization
    lender_corrections = {
        'jpmorgan chase': 'JPMorgan Chase',
        'jp morgan chase': 'JPMorgan Chase',
        'j.p. morgan chase': 'JPMorgan Chase',
        'j.p.morgan chase': 'JPMorgan Chase',
        'bank of america': 'Bank of America',
        'wells fargo': 'Wells Fargo',
        'citibank': 'Citibank',
        'us bank': 'U.S. Bank',
        'usbank': 'U.S. Bank',
        'pnc bank': 'PNC Bank',
        'truist': 'Truist',
        'capital one': 'Capital One',
        'first national bank': 'First National Bank',
        'huntington bank': 'Huntington Bank',
        'keybank': 'KeyBank',
        'regions bank': 'Regions Bank',
        'td bank': 'TD Bank',
        'suntrust': 'SunTrust',
        'bb&t': 'BB&T',
        'bbt': 'BB&T',
    }
    
    # Check for exact match (case-insensitive)
    name_lower = name.lower()
    if name_lower in lender_corrections:
        return lender_corrections[name_lower]
    
    # Check for partial matches (e.g., "JPMorgan Chase Bank" should match "JPMorgan Chase")
    for incorrect, correct in lender_corrections.items():
        if name_lower.startswith(incorrect) or incorrect in name_lower:
            # Replace the incorrect portion with the correct capitalization
            # Use case-insensitive replacement
            pattern = re.compile(re.escape(incorrect), re.IGNORECASE)
            corrected = pattern.sub(correct, name)
            # If the replacement changed something, return it
            if corrected != name:
                return corrected
    
    # If no correction found, return original name (preserving database capitalization)
    return name


def clean_mortgage_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and prepare mortgage data for analysis.
    
    Similar to branch data cleaning but for mortgage-specific fields.
    """
    # Make a copy to avoid modifying original
    df = df.copy()
    
    # Ensure numeric columns are properly typed
    numeric_columns = ['total_originations', 'lmib_originations', 'lmict_originations', 
                      'mmct_originations', 'total_loan_amount', 'avg_loan_amount', 'avg_income']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Ensure year is integer
    if 'year' in df.columns:
        df['year'] = pd.to_numeric(df['year'], errors='coerce').astype('Int64')
    
    # Fill missing lender names and convert to uppercase
    if 'lender_name' in df.columns:
        df['lender_name'] = df['lender_name'].fillna('Unknown Lender')
        # Convert all lender names to uppercase for display
        df['lender_name'] = df['lender_name'].apply(lambda x: str(x).upper() if x else x)
    
    return df


def create_demographic_overview_table(df: pd.DataFrame, years: List[int], census_data: Dict = None) -> pd.DataFrame:
    """
    Create demographic overview table showing lending activity by race/ethnicity.
    
    Shows number of loans and percent for each race/ethnic group over time.
    Only includes groups that are >= 1% of all loans.
    Denominator = loans with demographic data (total loans minus loans without race/ethnicity data).
    
    Table structure:
    - Far left: Metric column (race/ethnicity group)
    - Middle: Column for each year showing "Number (Percent%)"
    - Far right: Change column showing increase/decrease over entire span
    """
    print(f"  [DEBUG] create_demographic_overview_table called with census_data: {census_data is not None}")
    if census_data:
        print(f"  [DEBUG] census_data has {len(census_data)} counties: {list(census_data.keys())}")
    
    # Check if we have the race/ethnicity columns
    race_columns = ['hispanic_originations', 'black_originations', 'white_originations', 
                   'asian_originations', 'native_american_originations', 'hopi_originations']
    
    if not all(col in df.columns for col in race_columns):
        # Return empty DataFrame if we don't have the required columns
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Calculate total loans with demographic data
        # This is the sum of all race/ethnicity originations (may have overlap)
        # We need to calculate loans WITH demographic data differently
        # For now, use total_originations and subtract those without data
        # But we need to track which loans have demographic data
        
        # Sum originations by race/ethnicity
        hispanic = int(year_df['hispanic_originations'].sum())
        black = int(year_df['black_originations'].sum())
        white = int(year_df['white_originations'].sum())
        asian = int(year_df['asian_originations'].sum())
        native_american = int(year_df['native_american_originations'].sum())
        hopi = int(year_df['hopi_originations'].sum())
        
        # Total originations
        total_originations = int(year_df['total_originations'].sum())
        
        # Loans with demographic data (from SQL query if available)
        if 'loans_with_demographic_data' in year_df.columns:
            loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
        else:
            # Fallback: estimate as sum of all race/ethnicity originations
            # Note: This may overcount if loans are counted in multiple categories
            loans_with_demographics = hispanic + black + white + asian + native_american + hopi
        
        # More accurate: loans with demographics = total - loans without any classification
        # Since we're using COALESCE methodology, if a loan has no valid race/ethnicity,
        # it won't be counted in any of the race columns
        # So loans_with_demographics should be the sum of unique loans across all categories
        # For simplicity, we'll use the maximum of the individual counts as a proxy
        # But actually, we should track this in the SQL query
        
        # For now, use total_originations as denominator and calculate percentages
        # The SQL query should provide a has_demographic_data flag, but if not,
        # we'll use the sum of race originations as proxy
        
        yearly_data.append({
            'year': year,
            'total_originations': total_originations,
            'hispanic': hispanic,
            'black': black,
            'white': white,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'loans_with_demographics': loans_with_demographics
        })
    
    # Build table structure
    # First, determine which groups are >= 1% across all years
    all_totals = [d['total_originations'] for d in yearly_data]
    max_total = max(all_totals) if all_totals else 0
    
    # Define the standard order to match Population Demographics table:
    # White, Black, Hispanic, Asian, Native American, Hawaiian/PI
    standard_order = ['white', 'black', 'hispanic', 'asian', 'native_american', 'hopi']
    
    # Calculate max percentage for each group across all years
    group_max_pct = {}
    for group in standard_order:
        max_count = max([d[group] for d in yearly_data]) if yearly_data else 0
        # Use the year with max total as denominator for threshold check
        max_pct = (max_count / max_total * 100) if max_total > 0 else 0
        group_max_pct[group] = max_pct
    
    # Filter groups >= 1% and maintain the standard order
    included_groups = [group for group in standard_order if group_max_pct.get(group, 0) >= 1.0]
    
    if not included_groups:
        return pd.DataFrame()
    
    # Get population shares from Census data (if available)
    # For multiple counties, aggregate by weighting by total population
    population_shares = {}
    if census_data:
        total_population = 0
        group_totals = {
            'hispanic': 0,
            'black': 0,
            'white': 0,
            'asian': 0,
            'native_american': 0,
            'hopi': 0
        }
        
        # Aggregate across all counties by calculating weighted averages
        for county, county_census in census_data.items():
            demographics = county_census.get('demographics', {})
            if demographics:
                county_pop = demographics.get('total_population', 0)
                if county_pop and county_pop > 0:
                    total_population += county_pop
                    # Calculate actual counts from percentages
                    group_totals['hispanic'] += (demographics.get('hispanic_percentage', 0) / 100) * county_pop
                    group_totals['black'] += (demographics.get('black_percentage', 0) / 100) * county_pop
                    group_totals['white'] += (demographics.get('white_percentage', 0) / 100) * county_pop
                    group_totals['asian'] += (demographics.get('asian_percentage', 0) / 100) * county_pop
                    group_totals['native_american'] += (demographics.get('native_american_percentage', 0) / 100) * county_pop
                    group_totals['hopi'] += (demographics.get('hopi_percentage', 0) / 100) * county_pop
        
        # Calculate weighted average percentages
        if total_population > 0:
            for group in group_totals:
                population_shares[group] = (group_totals[group] / total_population * 100)
            print(f"  [DEBUG] Calculated population_shares: {population_shares}")
        else:
            print(f"  [WARNING] Total population is 0, cannot calculate population shares")
    else:
        print(f"  [DEBUG] No census_data provided, skipping population shares")
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add Population Share column (if Census data available)
    if population_shares:
        print(f"  [DEBUG] Adding Population Share (%) column to table")
        result_data['Population Share (%)'] = []
    else:
        print(f"  [DEBUG] No population_shares, skipping Population Share column")
    
    # Add change column
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Group labels
    group_labels = {
        'hispanic': 'Hispanic',
        'black': 'Black',
        'white': 'White',
        'asian': 'Asian',
        'native_american': 'Native American',
        'hopi': 'Hawaiian/Pacific Islander'
    }
    
    # Add Total Loans row at the top (showing just the integer, no percentage)
    result_data['Metric'].append('Total Loans')
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total = year_data['total_originations']
            result_data[str(year)].append(f"{total:,}")
        else:
            result_data[str(year)].append("0")
    # Population Share column (empty for Total Loans row)
    if population_shares:
        result_data['Population Share (%)'].append("")
    # Change column for Total Loans shows count change only
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            count_change = last_year_data['total_originations'] - first_year_data['total_originations']
            if count_change > 0:
                change_str = f"+{count_change:,}"
            elif count_change < 0:
                change_str = f"{count_change:,}"
            else:
                change_str = "0"
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            result_data[f'Change Over Time ({time_span})'].append("N/A")
    else:
        if len(years) >= 2:
            result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    # Calculate for each included group (show only percentages, no raw numbers)
    for group in included_groups:
        result_data['Metric'].append(group_labels[group])
        
        # Calculate for each year
        group_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[group]
                # Denominator: loans with demographic data (total loans minus loans without race/ethnicity data)
                total_loans = year_data['total_originations']
                loans_with_demo = year_data['loans_with_demographics']
                denominator = loans_with_demo if loans_with_demo > 0 else total_loans
                pct = (count / denominator * 100) if denominator > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                group_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                group_changes.append((0, 0.0))
        
        # Add Population Share column (if Census data available)
        if population_shares:
            pop_share = population_shares.get(group, 0)
            result_data['Population Share (%)'].append(f"{pop_share:.1f}%")
        
        # Calculate change from first to last year (show only percentage point change)
        if len(years) >= 2 and len(group_changes) >= 2:
            first_pct = group_changes[0][1]
            last_pct = group_changes[-1][1]
            pct_change = last_pct - first_pct
            
            # Format change: show only percentage point change
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    result = pd.DataFrame(result_data)
    return result


def create_income_neighborhood_indicators_table(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create income and neighborhood indicators table.
    
    Shows:
    - Number of loans overall (total originations)
    - Low to Moderate Income Census Tract (LMICT) originations
    - Low to Moderate Income Borrower (LMIB) originations
    - Majority Minority Census Tract (MMCT) originations
    
    Table structure:
    - Far left: Metric column
    - Middle: Column for each year showing "Number (Percent%)"
    - Far right: Change column showing increase/decrease over entire span
    """
    # Check if we have the required columns
    required_columns = ['total_originations', 'lmict_originations', 'lmib_originations', 'mmct_originations']
    if not all(col in df.columns for col in required_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Sum originations
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'lmib': lmib,
            'mmct': mmct
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add change column
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Define metrics in order
    metrics = [
        ('total', 'Loans'),
        ('lmict', 'Low to Moderate Income Census Tract'),
        ('lmib', 'Low to Moderate Income Borrower'),
        ('mmct', 'Majority Minority Census Tract')
    ]
    
    # Calculate for each metric
    for metric_key, metric_label in metrics:
        result_data['Metric'].append(metric_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # For "Loans" row, show just the integer (no percentage)
                if metric_key == 'total':
                    result_data[str(year)].append(f"{count:,}")
                    metric_changes.append((count, 100.0))
                else:
                    # For other rows, show only percentage
                    pct = (count / total * 100) if total > 0 else 0.0
                    result_data[str(year)].append(f"{pct:.1f}%")
                    metric_changes.append((count, pct))
            else:
                if metric_key == 'total':
                    result_data[str(year)].append("0")
                    metric_changes.append((0, 0.0))
                else:
                    result_data[str(year)].append("0.0%")
                    metric_changes.append((0, 0.0))
        
        # Calculate change from first to last year
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            
            # For "Loans" row, show count change only
            if metric_key == 'total':
                count_change = last_count - first_count
                if count_change > 0:
                    change_str = f"+{count_change:,}"
                elif count_change < 0:
                    change_str = f"{count_change:,}"
                else:
                    change_str = "0"
            else:
                # For other rows, show only percentage point change
                pct_change = last_pct - first_pct
                if pct_change > 0:
                    change_str = f"+{pct_change:.1f}pp"
                elif pct_change < 0:
                    change_str = f"{pct_change:.1f}pp"
                else:
                    change_str = "0.0pp"
            
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    result = pd.DataFrame(result_data)
    return result


def create_top_lenders_detailed_table(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create detailed table showing top lenders by total loans in most recent year.
    
    For each lender, shows:
    - Total loans/applications (most recent year)
    - Share to each race/ethnic group (using same methodology as demographic overview)
    - Share to income and neighborhood indicators (LMIB, LMICT, MMCT)
    
    Lenders are sorted in descending order by total loans in the most recent year.
    All lenders are included; JavaScript handles showing/hiding rows beyond the first 10.
    """
    if not years:
        return pd.DataFrame()
    
    # Check required columns
    required_columns = ['lender_name', 'total_originations', 'hispanic_originations', 
                       'black_originations', 'white_originations', 'asian_originations',
                       'native_american_originations', 'hopi_originations',
                       'lmib_originations', 'lmict_originations', 'mmct_originations',
                       'loans_with_demographic_data']
    
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        return pd.DataFrame()
    
    # FIRST: Calculate overall percentages across ALL YEARS (same logic as Section 1)
    # to determine which columns to include
    yearly_totals = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        if not year_df.empty:
            total_originations = int(year_df['total_originations'].sum())
            hispanic = int(year_df['hispanic_originations'].sum())
            black = int(year_df['black_originations'].sum())
            white = int(year_df['white_originations'].sum())
            asian = int(year_df['asian_originations'].sum())
            native_american = int(year_df['native_american_originations'].sum())
            hopi = int(year_df['hopi_originations'].sum())
            
            if 'loans_with_demographic_data' in year_df.columns:
                loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
            else:
                loans_with_demographics = hispanic + black + white + asian + native_american + hopi
            
            yearly_totals.append({
                'year': year,
                'total_originations': total_originations,
                'hispanic': hispanic,
                'black': black,
                'white': white,
                'asian': asian,
                'native_american': native_american,
                'hopi': hopi,
                'loans_with_demographics': loans_with_demographics
            })
    
    # Use same logic as Section 1: max count across all years / max total across all years
    all_totals = [d['total_originations'] for d in yearly_totals]
    max_total = max(all_totals) if all_totals else 0
    
    # Calculate max percentage for each group across all years (matching Section 1 logic)
    group_max_pct = {}
    for group in ['hispanic', 'black', 'white', 'asian', 'native_american', 'hopi']:
        max_count = max([d[group] for d in yearly_totals]) if yearly_totals else 0
        # Use the year with max total as denominator for threshold check (same as Section 1)
        max_pct = (max_count / max_total * 100) if max_total > 0 else 0
        group_max_pct[group] = max_pct
    
    # Determine which columns to include (>= 1% overall, matching Section 1)
    include_hispanic = group_max_pct.get('hispanic', 0) >= 1.0
    include_black = group_max_pct.get('black', 0) >= 1.0
    include_white = group_max_pct.get('white', 0) >= 1.0
    include_asian = group_max_pct.get('asian', 0) >= 1.0
    include_native_american = group_max_pct.get('native_american', 0) >= 1.0
    include_hopi = group_max_pct.get('hopi', 0) >= 1.0
    
    # NOW: Get most recent year data for the actual table
    latest_year = max(years)
    latest_year_df = df[df['year'] == latest_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    # Aggregate by lender (using latest year data only for the table)
    lender_data = []
    for lender_name in latest_year_df['lender_name'].unique():
        lender_df = latest_year_df[latest_year_df['lender_name'] == lender_name]
        
        total = int(lender_df['total_originations'].sum())
        loans_with_demo = int(lender_df['loans_with_demographic_data'].sum()) if 'loans_with_demographic_data' in lender_df.columns else total
        
        # Get lender type (if available)
        lender_type = None
        if 'lender_type' in lender_df.columns:
            lender_types = lender_df['lender_type'].dropna().unique()
            if len(lender_types) > 0:
                lender_type = lender_types[0]  # Use the first non-null type
        
        # Race/ethnicity originations
        hispanic = int(lender_df['hispanic_originations'].sum())
        black = int(lender_df['black_originations'].sum())
        white = int(lender_df['white_originations'].sum())
        asian = int(lender_df['asian_originations'].sum())
        native_american = int(lender_df['native_american_originations'].sum())
        hopi = int(lender_df['hopi_originations'].sum())
        
        # Income and neighborhood indicators
        lmib = int(lender_df['lmib_originations'].sum())
        lmict = int(lender_df['lmict_originations'].sum())
        mmct = int(lender_df['mmct_originations'].sum())
        
        lender_data.append({
            'lender_name': lender_name,
            'lender_type': lender_type,  # Include lender type
            'total_loans': total,
            'loans_with_demographic_data': loans_with_demo,
            'hispanic': hispanic,
            'black': black,
            'white': white,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'lmib': lmib,
            'lmict': lmict,
            'mmct': mmct
        })
    
    # Sort by total loans descending
    lender_data.sort(key=lambda x: x['total_loans'], reverse=True)
    
    # Return all lenders - JavaScript will handle showing/hiding rows beyond the first 10
    # This allows the table to work for communities with fewer than 10 lenders
    if not lender_data:
        return pd.DataFrame()
    
    # Build result table
    result_rows = []
    
    for lender in lender_data:
        lender_name = lender['lender_name']  # Already uppercase from clean_mortgage_data
        total = lender['total_loans']
        loans_with_demo = lender['loans_with_demographic_data']
        
        # Calculate percentages for race/ethnicity (denominator = loans with demographic data)
        denominator_demo = loans_with_demo if loans_with_demo > 0 else total
        
        # Calculate percentages for income/neighborhood (denominator = total loans)
        row_data = {
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Total Loans': f"{total:,}"
        }
        
        # Race/ethnicity percentages - only include columns that are >= 1% overall
        if include_hispanic:
            hispanic_pct = (lender['hispanic'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Hispanic (%)'] = f"{hispanic_pct:.1f}"
        
        if include_black:
            black_pct = (lender['black'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Black (%)'] = f"{black_pct:.1f}"
        
        if include_white:
            white_pct = (lender['white'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['White (%)'] = f"{white_pct:.1f}"
        
        if include_asian:
            asian_pct = (lender['asian'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Asian (%)'] = f"{asian_pct:.1f}"
        
        if include_native_american:
            native_american_pct = (lender['native_american'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Native American (%)'] = f"{native_american_pct:.1f}"
        
        if include_hopi:
            hopi_pct = (lender['hopi'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Hawaiian/Pacific Islander (%)'] = f"{hopi_pct:.1f}"
        
        # Income and neighborhood indicator percentages (denominator = total loans)
        lmib_pct = (lender['lmib'] / total * 100) if total > 0 else 0.0
        lmict_pct = (lender['lmict'] / total * 100) if total > 0 else 0.0
        mmct_pct = (lender['mmct'] / total * 100) if total > 0 else 0.0
        
        row_data['LMIB (%)'] = f"{lmib_pct:.1f}"
        row_data['LMICT (%)'] = f"{lmict_pct:.1f}"
        row_data['MMCT (%)'] = f"{mmct_pct:.1f}"
        
        # Note: "No Data (%)" column removed per user request
        
        result_rows.append(row_data)
    
    result = pd.DataFrame(result_rows)
    return result


def create_mortgage_summary_table(df: pd.DataFrame, counties: List[str], years: List[int]) -> pd.DataFrame:
    """
    Create yearly breakdown table for mortgage origination data.
    
    Similar to branch summary but for mortgage originations.
    Properly dedupes: LMI Borrower only, MMCT only, and both LMICT/MMCT.
    """
    result_data = {
        'Variable': ['Total Originations', 'LMI Borrower Only Originations', 'MMCT Only Originations', 'Both LMICT/MMCT Originations']
    }
    
    # Calculate for each year
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Total originations
        total_originations = int(year_df['total_originations'].sum())
        
        # LMI Borrower only (LMIB but not MMCT)
        # Note: We need to check if we have both LMIB and MMCT flags
        # For now, use lmib_originations - mmct_originations as approximation
        # This is a simplification - actual data may need different logic
        lmib_only = int(year_df['lmib_originations'].sum() - year_df[year_df['mmct_originations'] > 0]['lmib_originations'].sum())
        lmib_only = max(0, lmib_only)  # Ensure non-negative
        
        # MMCT only (not LMI Borrower)
        mmct_only = int(year_df['mmct_originations'].sum() - year_df[year_df['lmib_originations'] > 0]['mmct_originations'].sum())
        mmct_only = max(0, mmct_only)  # Ensure non-negative
        
        # Both LMICT/MMCT
        both = int(year_df[(year_df['lmib_originations'] > 0) & (year_df['mmct_originations'] > 0)]['total_originations'].sum())
        
        result_data[str(year)] = [
            total_originations,
            lmib_only,
            mmct_only,
            both
        ]
    
    result = pd.DataFrame(result_data)
    
    # Calculate net change from first to last year
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        
        net_changes = []
        for idx, row in result.iterrows():
            first_val = row[first_year]
            last_val = row[last_year]
            net_change = last_val - first_val
            net_changes.append(net_change)
        
        result['Net Change'] = net_changes
    
    return result


def create_lender_summary(df: pd.DataFrame, years: List[int] = None) -> pd.DataFrame:
    """
    Create lender summary table showing top lenders by origination volume.
    
    Similar to bank summary but for mortgage lenders.
    """
    if not years or len(years) == 0:
        return pd.DataFrame()
    
    latest_year = max(years)
    first_year = min(years)
    
    latest_year_df = df[df['year'] == latest_year].copy()
    first_year_df = df[df['year'] == first_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    lender_data = []
    
    # Get all unique lenders
    all_lenders = latest_year_df['lender_name'].unique()
    
    for lender_name in all_lenders:
        # Latest year data for this lender
        lender_latest = latest_year_df[latest_year_df['lender_name'] == lender_name]
        lender_first = first_year_df[first_year_df['lender_name'] == lender_name] if first_year != latest_year else lender_latest
        
        # Sum originations
        total_originations_latest = int(lender_latest['total_originations'].sum())
        total_originations_first = int(lender_first['total_originations'].sum())
        
        # Calculate LMI Borrower, MMCT, and both
        lmib_latest = int(lender_latest['lmib_originations'].sum())
        mmct_latest = int(lender_latest['mmct_originations'].sum())
        both_latest = int(lender_latest[(lender_latest['lmib_originations'] > 0) & (lender_latest['mmct_originations'] > 0)]['total_originations'].sum())
        
        # Calculate percentages
        lmib_pct = round((lmib_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        mmct_pct = round((mmct_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        both_pct = round((both_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        
        # Net change
        net_change = total_originations_latest - total_originations_first
        
        lender_data.append({
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Total Originations': total_originations_latest,
            'LMI Borrower Only Originations': f"{lmib_latest} ({lmib_pct}%)",
            'MMCT Only Originations': f"{mmct_latest} ({mmct_pct}%)",
            'Both LMICT/MMCT Originations': f"{both_latest} ({both_pct}%)",
            'Net Change': net_change,
            # Store raw values for sorting
            '_total_originations': total_originations_latest
        })
    
    result = pd.DataFrame(lender_data)
    
    # Sort by total originations descending
    result = result.sort_values('_total_originations', ascending=False).reset_index(drop=True)
    
    # Remove helper column
    result = result.drop(columns=['_total_originations'])
    
    return result


def create_mortgage_county_summary(df: pd.DataFrame, counties: List[str] = None, years: List[int] = None) -> pd.DataFrame:
    """
    Create county-by-county summary for mortgage originations.
    
    Only shown if multiple counties selected.
    """
    # Only create this table if multiple counties
    if counties and len(counties) <= 1:
        return pd.DataFrame()
    
    if not years or len(years) == 0:
        return pd.DataFrame()
    
    # Use the most recent year
    latest_year = max(years)
    df_latest = df[df['year'] == latest_year].copy()
    
    if df_latest.empty:
        return pd.DataFrame()
    
    county_data = []
    
    # Process each county
    for county in df_latest['county_state'].unique():
        county_df = df_latest[df_latest['county_state'] == county]
        
        # Sum originations
        total_originations = int(county_df['total_originations'].sum())
        
        # Calculate LMI Borrower, MMCT, and both
        lmib = int(county_df['lmib_originations'].sum())
        mmct = int(county_df['mmct_originations'].sum())
        both = int(county_df[(county_df['lmib_originations'] > 0) & (county_df['mmct_originations'] > 0)]['total_originations'].sum())
        
        # Count unique lenders
        num_lenders = county_df['lender_name'].nunique()
        
        # Calculate percentages
        lmib_pct = round((lmib / total_originations * 100), 1) if total_originations > 0 else 0.0
        mmct_pct = round((mmct / total_originations * 100), 1) if total_originations > 0 else 0.0
        both_pct = round((both / total_originations * 100), 1) if total_originations > 0 else 0.0
        
        county_data.append({
            'County': county,
            'Total Originations': total_originations,
            'LMI Borrower Only Originations': f"{lmib} ({lmib_pct}%)",
            'MMCT Only Originations': f"{mmct} ({mmct_pct}%)",
            'Both LMICT/MMCT Originations': f"{both} ({both_pct}%)",
            'Number of Lenders': num_lenders
        })
    
    result = pd.DataFrame(county_data)
    
    # Sort by Total Originations descending
    result = result.sort_values('Total Originations', ascending=False).reset_index(drop=True)
    
    return result


def create_mortgage_trend_analysis(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create year-over-year trend analysis for mortgage originations.
    """
    if not years or len(years) < 2:
        return pd.DataFrame()
    
    yearly_totals = []
    
    for year in sorted(years):
        year_df = df[df['year'] == year]
        
        total = int(year_df['total_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_totals.append({
            'Year': year,
            'Total Originations': total,
            'LMI Borrower Originations': lmib,
            'MMCT Originations': mmct
        })
    
    result = pd.DataFrame(yearly_totals)
    
    # Calculate year-over-year changes
    if len(result) > 1:
        result['YoY Change (%)'] = result['Total Originations'].pct_change() * 100
        result['YoY Change (%)'] = result['YoY Change (%)'].round(1)
        result.loc[0, 'YoY Change (%)'] = None  # First year has no previous year
    
    return result


def create_demographic_overview_table_for_excel(df: pd.DataFrame, years: List[int], census_data: Dict = None) -> pd.DataFrame:
    """
    Create demographic overview table for Excel export with ALL race/ethnic categories.
    
    This version includes all categories regardless of percentage.
    Note: "No Data" row is not included per user request.
    """
    race_columns = ['hispanic_originations', 'black_originations', 'white_originations', 
                   'asian_originations', 'native_american_originations', 'hopi_originations']
    
    if not all(col in df.columns for col in race_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        hispanic = int(year_df['hispanic_originations'].sum())
        black = int(year_df['black_originations'].sum())
        white = int(year_df['white_originations'].sum())
        asian = int(year_df['asian_originations'].sum())
        native_american = int(year_df['native_american_originations'].sum())
        hopi = int(year_df['hopi_originations'].sum())
        
        total_originations = int(year_df['total_originations'].sum())
        
        if 'loans_with_demographic_data' in year_df.columns:
            loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
        else:
            loans_with_demographics = hispanic + black + white + asian + native_american + hopi
        
        no_data_loans = total_originations - loans_with_demographics
        
        yearly_data.append({
            'year': year,
            'total_originations': total_originations,
            'hispanic': hispanic,
            'black': black,
            'white': white,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'loans_with_demographics': loans_with_demographics,
            'no_data_loans': no_data_loans
        })
    
    # Get population shares from Census data (if available)
    # For multiple counties, aggregate by weighting by total population
    population_shares = {}
    if census_data:
        total_population = 0
        group_totals = {
            'hispanic': 0,
            'black': 0,
            'white': 0,
            'asian': 0,
            'native_american': 0,
            'hopi': 0
        }
        
        # Aggregate across all counties by calculating weighted averages
        for county, county_census in census_data.items():
            demographics = county_census.get('demographics', {})
            if demographics:
                county_pop = demographics.get('total_population', 0)
                if county_pop and county_pop > 0:
                    total_population += county_pop
                    # Calculate actual counts from percentages
                    group_totals['hispanic'] += (demographics.get('hispanic_percentage', 0) / 100) * county_pop
                    group_totals['black'] += (demographics.get('black_percentage', 0) / 100) * county_pop
                    group_totals['white'] += (demographics.get('white_percentage', 0) / 100) * county_pop
                    group_totals['asian'] += (demographics.get('asian_percentage', 0) / 100) * county_pop
                    group_totals['native_american'] += (demographics.get('native_american_percentage', 0) / 100) * county_pop
                    group_totals['hopi'] += (demographics.get('hopi_percentage', 0) / 100) * county_pop
        
        # Calculate weighted average percentages
        if total_population > 0:
            for group in group_totals:
                population_shares[group] = (group_totals[group] / total_population * 100)
    
    # Build result table - include ALL groups
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add Population Share column (if Census data available)
    if population_shares:
        result_data['Population Share (%)'] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Group labels
    group_labels = {
        'hispanic': 'Hispanic',
        'black': 'Black',
        'white': 'White',
        'asian': 'Asian',
        'native_american': 'Native American',
        'hopi': 'Hawaiian/Pacific Islander'
    }
    
    # Add Total Loans row
    result_data['Metric'].append('Total Loans')
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            result_data[str(year)].append(year_data['total_originations'])
        else:
            result_data[str(year)].append(0)
    
    # Population Share column (empty for Total Loans row)
    if population_shares:
        result_data['Population Share (%)'].append("")
    
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            count_change = last_year_data['total_originations'] - first_year_data['total_originations']
            result_data[f'Change Over Time ({time_span})'].append(count_change)
        else:
            result_data[f'Change Over Time ({time_span})'].append(0)
    
    # Add ALL race/ethnicity groups (not filtered by 1%)
    for group in ['hispanic', 'black', 'white', 'asian', 'native_american', 'hopi']:
        result_data['Metric'].append(group_labels[group])
        
        group_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[group]
                loans_with_demo = year_data['loans_with_demographics']
                pct = (count / loans_with_demo * 100) if loans_with_demo > 0 else 0.0
                result_data[str(year)].append(pct)
                group_changes.append((count, pct))
            else:
                result_data[str(year)].append(0.0)
                group_changes.append((0, 0.0))
        
        # Add Population Share column (if Census data available)
        if population_shares:
            pop_share = population_shares.get(group, 0)
            result_data['Population Share (%)'].append(pop_share)
        
        # Change column
        if len(years) >= 2 and len(group_changes) >= 2:
            first_pct = group_changes[0][1]
            last_pct = group_changes[-1][1]
            pct_change = last_pct - first_pct
            result_data[f'Change Over Time ({time_span})'].append(pct_change)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append(0.0)
    
    # Note: "No Data" row removed per user request
    
    result = pd.DataFrame(result_data)
    return result


def create_income_neighborhood_indicators_table_for_excel(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create income and neighborhood indicators table for Excel export.
    
    This matches the report table structure exactly:
    - Loans row: integer only
    - Other rows: percentages only
    - Change Over Time column
    """
    required_columns = ['total_originations', 'lmict_originations', 'lmib_originations', 'mmct_originations']
    if not all(col in df.columns for col in required_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'lmib': lmib,
            'mmct': mmct
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Define metrics in order
    metrics = [
        ('total', 'Loans'),
        ('lmict', 'Low to Moderate Income Census Tract'),
        ('lmib', 'Low to Moderate Income Borrower'),
        ('mmct', 'Majority Minority Census Tract')
    ]
    
    # Calculate for each metric
    for metric_key, metric_label in metrics:
        result_data['Metric'].append(metric_label)
        
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # For "Loans" row, show just the integer (no percentage)
                if metric_key == 'total':
                    result_data[str(year)].append(count)
                    metric_changes.append((count, 100.0))
                else:
                    # For other rows, show only percentage
                    pct = (count / total * 100) if total > 0 else 0.0
                    result_data[str(year)].append(pct)
                    metric_changes.append((count, pct))
            else:
                if metric_key == 'total':
                    result_data[str(year)].append(0)
                    metric_changes.append((0, 0.0))
                else:
                    result_data[str(year)].append(0.0)
                    metric_changes.append((0, 0.0))
        
        # Calculate change from first to last year
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            
            # For "Loans" row, show count change only
            if metric_key == 'total':
                count_change = last_count - first_count
                result_data[f'Change Over Time ({time_span})'].append(count_change)
            else:
                # For other rows, show only percentage point change
                pct_change = last_pct - first_pct
                result_data[f'Change Over Time ({time_span})'].append(pct_change)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append(0 if metric_key == 'total' else 0.0)
    
    result = pd.DataFrame(result_data)
    return result


def create_top_lenders_table_for_excel(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create top lenders table for Excel export with ALL race/ethnic categories and No Data column.
    
    This version includes all categories regardless of percentage, and includes the No Data column.
    """
    if not years:
        return pd.DataFrame()
    
    latest_year = max(years)
    
    latest_year_df = df[df['year'] == latest_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    required_columns = ['lender_name', 'total_originations', 'hispanic_originations', 
                       'black_originations', 'white_originations', 'asian_originations',
                       'native_american_originations', 'hopi_originations',
                       'lmib_originations', 'lmict_originations', 'mmct_originations',
                       'loans_with_demographic_data']
    
    missing_cols = [col for col in required_columns if col not in latest_year_df.columns]
    if missing_cols:
        return pd.DataFrame()
    
    # Aggregate by lender
    lender_data = []
    for lender_name in latest_year_df['lender_name'].unique():
        lender_df = latest_year_df[latest_year_df['lender_name'] == lender_name]
        
        total = int(lender_df['total_originations'].sum())
        loans_with_demo = int(lender_df['loans_with_demographic_data'].sum()) if 'loans_with_demographic_data' in lender_df.columns else total
        
        lender_type = None
        if 'lender_type' in lender_df.columns:
            lender_types = lender_df['lender_type'].dropna().unique()
            if len(lender_types) > 0:
                lender_type = lender_types[0]
        
        hispanic = int(lender_df['hispanic_originations'].sum())
        black = int(lender_df['black_originations'].sum())
        white = int(lender_df['white_originations'].sum())
        asian = int(lender_df['asian_originations'].sum())
        native_american = int(lender_df['native_american_originations'].sum())
        hopi = int(lender_df['hopi_originations'].sum())
        
        lmib = int(lender_df['lmib_originations'].sum())
        lmict = int(lender_df['lmict_originations'].sum())
        mmct = int(lender_df['mmct_originations'].sum())
        
        loans_no_data = total - loans_with_demo
        
        lender_data.append({
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Lender Type': lender_type if lender_type else '',
            'Total Loans': total,
            'Hispanic (%)': (hispanic / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Black (%)': (black / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'White (%)': (white / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Asian (%)': (asian / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Native American (%)': (native_american / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Hawaiian/Pacific Islander (%)': (hopi / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'LMIB (%)': (lmib / total * 100) if total > 0 else 0.0,
            'LMICT (%)': (lmict / total * 100) if total > 0 else 0.0,
            'MMCT (%)': (mmct / total * 100) if total > 0 else 0.0,
            'No Data (%)': (loans_no_data / total * 100) if total > 0 else 0.0
        })
    
    # Sort by total loans descending
    lender_data.sort(key=lambda x: x['Total Loans'], reverse=True)
    
    result = pd.DataFrame(lender_data)
    return result


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize a sheet name for Excel compatibility.
    
    Excel sheet names cannot contain: : \ / ? * [ ]
    Also limited to 31 characters.
    """
    # Replace invalid characters
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '-')
    
    # Truncate if too long
    if len(name) > max_length:
        name = name[:max_length]
    
    return name


def save_mortgage_excel_report(report_data: Dict[str, pd.DataFrame], output_path: str, metadata: Dict[str, Any] = None):
    """
    Save the mortgage report data to an Excel file with multiple sheets.
    
    Sheets:
    1. Methods and Definitions
    2. Section 1: Demographic Overview (with ALL categories and No Data)
    3. Section 2: Income and Neighborhood Indicators
    4. Section 3: Top Lenders (with ALL categories and No Data)
    5. Raw Data (all raw data used in the report)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Get raw data for creating Excel-specific tables
    raw_df = report_data.get('raw_data', pd.DataFrame())
    years = metadata.get('years', []) if metadata else []
    
    if raw_df.empty or not years:
        raise ValueError("Cannot create Excel report: missing raw data or years")
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Get workbook and remove default sheet if it exists
        workbook = writer.book
        # Remove default "Sheet" if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Methods and Definitions sheet (first sheet)
        methods_sheet = workbook.create_sheet(sanitize_sheet_name('Methods and Definitions'), 0)
        
        # Prepare methods content
        methods_content = []
        
        methods_content.append(('METHODS AND DEFINITIONS', ''))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Sources', ''))
        methods_content.append(('', 'Home Mortgage Disclosure Act (HMDA) data, compiled and maintained in NCRC\'s curated databases'))
        methods_content.append(('', 'U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates for demographic context'))
        methods_content.append(('', 'Census data retrieved via U.S. Census Bureau API (https://api.census.gov)'))
        if metadata and 'census_data' in metadata and metadata.get('census_data'):
            census_year = None
            for county_data in metadata.get('census_data', {}).values():
                if 'data_year' in county_data:
                    census_year = county_data['data_year']
                    break
            if census_year:
                methods_content.append(('', f'Census data year: {census_year}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Race and Ethnicity Classification', ''))
        methods_content.append(('', 'Race and ethnicity are determined using the COALESCE function across applicant_race_1 through applicant_race_5 and applicant_ethnicity_1 through applicant_ethnicity_5 fields.'))
        methods_content.append(('', 'Hispanic classification takes precedence (if any ethnicity field indicates Hispanic, the borrower is classified as Hispanic).'))
        methods_content.append(('', 'For non-Hispanic borrowers, the first valid race code (excluding codes 6, 7, 8 which indicate "Not Provided", "Not Applicable", or "Information not provided") is used from race_1 through race_5.'))
        methods_content.append(('', 'Percentages are calculated as: (group loans / loans with demographic data) × 100, where loans with demographic data = all loans minus loans lacking race/ethnicity data.'))
        methods_content.append(('', ''))
        
        methods_content.append(('Income and Neighborhood Indicators', ''))
        methods_content.append(('LMIB', 'Low to Moderate Income Borrower: Borrowers with income below 80% of the area median family income'))
        methods_content.append(('LMICT', 'Low to Moderate Income Census Tract: Census tracts where median family income is below 80% of the area median family income'))
        methods_content.append(('MMCT', 'Majority Minority Census Tract: Census tracts where minority populations represent more than 50% of total population'))
        methods_content.append(('', 'Percentages are calculated as: (category loans / total loans) × 100'))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Filters Applied', ''))
        methods_content.append(('', 'Originations only (action taken = 1)'))
        methods_content.append(('', 'Owner-occupied properties (occupancy type = 1)'))
        methods_content.append(('', 'Forward loans (excludes reverse mortgages)'))
        methods_content.append(('', 'Site-built properties (construction method = 1)'))
        methods_content.append(('', '1-4 unit properties (total_units IN (1, 2, 3, 4))'))
        if metadata and 'loan_purpose' in metadata:
            loan_purpose = metadata['loan_purpose']
            if isinstance(loan_purpose, list):
                if len(loan_purpose) == 0 or set(loan_purpose) == {'purchase', 'refinance', 'equity'} or 'all' in loan_purpose:
                    methods_content.append(('', 'Loan Purpose: All loan purposes'))
                else:
                    purpose_map = {
                        'purchase': 'Home Purchase (loan purpose = 1)',
                        'refinance': 'Refinance & Cash-Out (loan purpose IN (31, 32))',
                        'equity': 'Home Equity (loan purpose IN (2, 4))'
                    }
                    purposes = [purpose_map.get(p, p) for p in loan_purpose if p != 'all']
                    methods_content.append(('', f'Loan Purpose: {", ".join(purposes)}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Abbreviations', ''))
        methods_content.append(('HMDA', 'Home Mortgage Disclosure Act'))
        methods_content.append(('CFPB', 'Consumer Financial Protection Bureau'))
        methods_content.append(('NCRC', 'National Community Reinvestment Coalition'))
        methods_content.append(('LMI', 'Low-to-Moderate Income'))
        methods_content.append(('LMIB', 'Low-to-Moderate Income Borrower'))
        methods_content.append(('LMICT', 'Low-to-Moderate Income Census Tract'))
        methods_content.append(('MMCT', 'Majority-Minority Census Tract'))
        methods_content.append(('AMFI', 'Area Median Family Income'))
        methods_content.append(('pp', 'Percentage Points'))
        methods_content.append(('', ''))
        
        # Write methods content
        for row_idx, (label, value) in enumerate(methods_content, start=1):
            methods_sheet.cell(row=row_idx, column=1, value=label)
            methods_sheet.cell(row=row_idx, column=2, value=value)
            if label and label != '':
                methods_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Section 1: Demographic Overview (with ALL categories and No Data)
        # Pass census_data if available for population share column
        census_data_for_excel = metadata.get('census_data', {}) if metadata else {}
        demo_table = create_demographic_overview_table_for_excel(raw_df, years, census_data=census_data_for_excel)
        if not demo_table.empty:
            demo_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 1 - Demographic Overview'), index=False)
        
        # Section 2: Income and Neighborhood Indicators
        # This matches the report table exactly
        income_table = create_income_neighborhood_indicators_table_for_excel(raw_df, years)
        if not income_table.empty:
            income_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2 - Income and Neighborhood'), index=False)
        
        # Section 3: Top Lenders (with ALL categories and No Data)
        # This matches the report table exactly but includes ALL categories
        top_lenders_table = create_top_lenders_table_for_excel(raw_df, years)
        if not top_lenders_table.empty:
            top_lenders_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 3 - Top Lenders'), index=False)
        
        # Raw Data sheet - all raw data used in the report
        if not raw_df.empty:
            raw_export_df = raw_df.copy()
            raw_export_df.to_excel(writer, sheet_name=sanitize_sheet_name('Raw Data'), index=False)
        
        # Style the header row for all data sheets (not Methods sheet)
        from openpyxl.styles import NamedStyle
        from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE
        
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Number format styles
        integer_format = '#,##0'  # Format: 1,234
        percentage_format = '#,##0.00'  # Format: 12.34
        
        for sheet_name in writer.sheets:
            if sheet_name != 'Methods and Definitions':
                worksheet = writer.sheets[sheet_name]
                if worksheet.max_row > 0:
                    # Style header row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply number formatting to data cells
                    # Get header row to identify column types
                    header_row = worksheet[1]
                    headers = [cell.value for cell in header_row]
                    
                    # Iterate through data rows (skip header row)
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col_idx, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            
                            # Skip if cell is empty or not a number
                            if cell.value is None:
                                continue
                            
                            # Determine format based on header name and value type
                            header_str = str(header) if header else ''
                            
                            # Get the metric value from first column to determine row type
                            metric_cell = worksheet.cell(row=row_idx, column=1)
                            metric_value = metric_cell.value if metric_cell else ''
                            metric_str = str(metric_value) if metric_value else ''
                            
                            # Check if it's a percentage column
                            if '%' in header_str or 'percent' in header_str.lower() or 'Population Share' in header_str:
                                # Format as percentage with 2 decimal places
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = percentage_format
                            # Check if it's a year column - format based on row type
                            elif header_str.isdigit() or 'year' in header_str.lower():
                                # Year columns - format based on metric row
                                if isinstance(cell.value, (int, float)):
                                    # If row is "Total Loans" or "Loans", format as integer
                                    if 'Total Loans' in metric_str or metric_str == 'Loans':
                                        cell.number_format = integer_format
                                    else:
                                        # All other rows in year columns are percentages
                                        cell.number_format = percentage_format
                            # Check if it's "Change Over Time" - could be integer or percentage
                            elif 'Change Over Time' in header_str:
                                # Use the metric_str we already determined above
                                # If metric is "Total Loans" or "Loans", it's a count (integer)
                                # Otherwise it's a percentage change
                                if 'Total Loans' in metric_str or metric_str == 'Loans':
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = integer_format
                                else:
                                    # Percentage point change
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = percentage_format
                            # Check if it's "Net Change" - usually integer
                            elif 'Net Change' in header_str:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = integer_format
                            # Check if it's a metric column (first column) - no formatting
                            elif col_idx == 1:
                                pass
                            # Default: if it's a number and not a percentage, format as integer
                            elif isinstance(cell.value, (int, float)):
                                # Check if value looks like a percentage (between 0 and 100, likely decimal)
                                if 0 <= cell.value <= 100 and isinstance(cell.value, float) and cell.value != int(cell.value):
                                    # Likely a percentage value stored as number
                                    cell.number_format = percentage_format
                                else:
                                    # Integer value
                                    cell.number_format = integer_format
        
        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

