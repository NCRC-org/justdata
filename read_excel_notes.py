#!/usr/bin/env python3
"""Read notes from existing Excel file."""

import openpyxl
from pathlib import Path

excel_path = Path(r'C:\Users\edite\Downloads\1071_Analysis_Tables_20251125_102440.xlsx')

if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets:", wb.sheetnames)
    
    if 'Notes' in wb.sheetnames:
        notes_sheet = wb['Notes']
        print("\nNotes content:")
        print("=" * 80)
        for row in notes_sheet.iter_rows():
            values = [cell.value for cell in row if cell.value]
            if values:
                print(" | ".join(str(v) for v in values))
    else:
        print("\nNo Notes sheet found")
else:
    print(f"File not found: {excel_path}")

