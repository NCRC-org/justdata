#!/usr/bin/env python3
"""
Excel generator for Data Explorer Lender Analysis.
Adapted from MergerMeter Excel generator to show Subject Lender vs Peers
with data broken down by CBSA and county.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Optional, List, Any
from datetime import datetime
from collections import defaultdict
from shared.utils.bigquery_client import get_bigquery_client, execute_query
from .config import DataExplorerConfig


def _get_cbsa_name_from_code(cbsa_code: str, cbsa_name_cache: Dict[str, str] = None) -> str:
    """Look up CBSA name from code, using cache if provided."""
    if cbsa_name_cache is None:
        cbsa_name_cache = {}
    
    cbsa_code_str = str(cbsa_code).strip()
    
    # Check cache first
    if cbsa_code_str in cbsa_name_cache:
        return cbsa_name_cache[cbsa_code_str]
    
    # If code is empty or invalid, return fallback
    if not cbsa_code_str or cbsa_code_str.lower() in ['nan', 'none', '']:
        return f"CBSA {cbsa_code_str}" if cbsa_code_str else "Non-MSA"
    
    # Try to look up from BigQuery
    try:
        query = f"""
        SELECT DISTINCT cbsa as cbsa_name
        FROM `{DataExplorerConfig.PROJECT_ID}.geo.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) = '{cbsa_code_str}'
        LIMIT 1
        """
        results = execute_query(query)
        if results and len(results) > 0:
            cbsa_name = str(results[0].get('cbsa_name', '')).strip()
            if cbsa_name and cbsa_name.lower() not in ['nan', 'none', '']:
                cbsa_name_cache[cbsa_code_str] = cbsa_name
                return cbsa_name
    except Exception as e:
        print(f"    Warning: Could not look up CBSA name for {cbsa_code_str}: {e}")
    
    # Fallback
    if cbsa_code_str == '99999' or cbsa_code_str == '':
        return "Non-MSA"
    return f"CBSA {cbsa_code_str}"


def create_lender_analysis_excel(
    output_path: Path,
    subject_lender_name: str,
    raw_results: Dict[str, Dict[str, List[Dict[str, Any]]]],
    assessment_areas: Optional[Dict] = None,
    metadata: Optional[Dict] = None
):
    """
    Create Excel workbook with lender analysis data using MergerMeter-style format.
    
    Shows data broken down by CBSA and county, with Subject Lender vs Peers comparison.
    
    Args:
        output_path: Path to save Excel file
        subject_lender_name: Subject lender name
        raw_results: Dictionary with structure:
            {
                'hmda': {'subject': [...], 'peer': [...]},
                'sb': {'subject': [...], 'peer': [...]},
                'branches': {'subject': [...], 'peer': [...]}
            }
        assessment_areas: Assessment area information (counties with CBSA info)
        metadata: Additional metadata (years, filters, etc.)
    """
    
    # Convert raw results to DataFrames
    hmda_subject_df = None
    hmda_peer_df = None
    sb_subject_df = None
    sb_peer_df = None
    branch_subject_df = None
    branch_peer_df = None
    
    if raw_results.get('hmda', {}).get('subject'):
        hmda_subject_df = pd.DataFrame(raw_results['hmda']['subject'])
    if raw_results.get('hmda', {}).get('peer'):
        hmda_peer_df = pd.DataFrame(raw_results['hmda']['peer'])
    
    if raw_results.get('sb', {}).get('subject'):
        sb_subject_df = pd.DataFrame(raw_results['sb']['subject'])
    if raw_results.get('sb', {}).get('peer'):
        sb_peer_df = pd.DataFrame(raw_results['sb']['peer'])
    
    if raw_results.get('branches', {}).get('subject'):
        branch_subject_df = pd.DataFrame(raw_results['branches']['subject'])
    if raw_results.get('branches', {}).get('peer'):
        branch_peer_df = pd.DataFrame(raw_results['branches']['peer'])
    
    # Get years from metadata
    hmda_years = metadata.get('hmda_years', []) if metadata else []
    sb_years = metadata.get('sb_years', []) if metadata else []
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Assessment Areas Sheet
    if assessment_areas:
        create_assessment_areas_sheet(
            wb, subject_lender_name, assessment_areas,
            header_fill, header_font, border_thin
        )
    
    # 2. Mortgage Data Sheet
    if hmda_subject_df is not None and not hmda_subject_df.empty:
        create_mortgage_sheet(
            wb, subject_lender_name, hmda_subject_df, hmda_peer_df,
            header_fill, header_font, border_thin
        )
    
    # 3. Small Business Sheet
    if sb_subject_df is not None and not sb_subject_df.empty:
        create_sb_sheet(
            wb, subject_lender_name, sb_subject_df, sb_peer_df,
            header_fill, header_font, border_thin
        )
    
    # 4. Branch Sheet
    if branch_subject_df is not None and not branch_subject_df.empty:
        create_branch_sheet(
            wb, subject_lender_name, branch_subject_df, branch_peer_df,
            header_fill, header_font, border_thin
        )
    
    # 5. Notes/Methodology Sheet
    create_notes_sheet(wb, subject_lender_name, hmda_years, sb_years, metadata)
    
    # Save workbook
    wb.save(output_path)
    print(f"\n[OK] Lender Analysis Excel workbook saved to: {output_path}")


def create_assessment_areas_sheet(wb, lender_name, assessment_areas_data,
                                  header_fill, header_font, border_thin):
    """Create Assessment Areas sheet showing all counties."""
    ws = wb.create_sheet("Assessment Areas")
    
    # Header
    ws['A1'] = f'{lender_name} Assessment Areas'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Column headers (removed "Assessment Area" column B)
    headers = ['State', 'CBSA Name', 'CBSA Code', 'County, State', 'GEOID5']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get counties
    counties = assessment_areas_data.get('counties', [])
    if not counties or not isinstance(counties, list):
        counties = []
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Populate data
    row = 3
    if counties:
        # Group by state and assessment area
        state_groups = defaultdict(lambda: defaultdict(list))
        for county in counties:
            state = county.get('state_name', county.get('state', '')) or 'Unknown'
            # Use CBSA name as Assessment Area name
            cbsa_code = str(county.get('cbsa_code', ''))
            cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
            state_groups[state][cbsa_name].append(county)
        
        for state in sorted(state_groups.keys()):
            for aa in sorted(state_groups[state].keys()):
                for county in state_groups[state][aa]:
                    ws.cell(row, 1).value = state
                    cbsa_code = str(county.get('cbsa_code', ''))
                    cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
                    ws.cell(row, 2).value = cbsa_name  # CBSA Name (column B removed, so this is now column 2)
                    ws.cell(row, 3).value = cbsa_code   # CBSA Code
                    county_name = county.get('county_name', '')
                    state_name = county.get('state_name', county.get('state', ''))
                    ws.cell(row, 4).value = f"{county_name}, {state_name}" if county_name and state_name else (county_name or state_name)
                    ws.cell(row, 5).value = str(county.get('geoid5', ''))
                    row += 1
    
    # Auto-adjust column widths (5 columns now, not 6)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created Assessment Areas sheet: {len(counties)} counties")


def create_mortgage_sheet(wb, lender_name, subject_data, peer_data,
                          header_fill, header_font, border_thin):
    """Create Mortgage Data sheet with Grand Total and CBSA breakdown."""
    sheet_name = f"{lender_name} Mortgage Data"
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['CBSA Name', 'Metric', 'Subject Lender', 'Peers', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No mortgage data for {lender_name}")
        return
    
    # Metrics in order
    metrics = [
        'Loans', 'LMICT%', 'LMIB%', 'LMIB$', 'MMCT%',
        'MINB%', 'Asian%', 'Black%', 'Native American%', 'HoPI%', 'Hispanic%'
    ]
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    # Need to handle different column name variations
    grand_total_agg = {}
    if 'total_loans' in subject_data.columns:
        grand_total_agg['total_loans'] = subject_data['total_loans'].sum()
    if 'lmict_loans' in subject_data.columns:
        grand_total_agg['lmict_loans'] = subject_data['lmict_loans'].sum()
    if 'lmib_loans' in subject_data.columns:
        grand_total_agg['lmib_loans'] = subject_data['lmib_loans'].sum()
    if 'lmib_amount' in subject_data.columns:
        grand_total_agg['lmib_amount'] = subject_data['lmib_amount'].sum()
    if 'mmct_loans' in subject_data.columns:
        grand_total_agg['mmct_loans'] = subject_data['mmct_loans'].sum()
    if 'minb_loans' in subject_data.columns:
        grand_total_agg['minb_loans'] = subject_data['minb_loans'].sum()
    if 'asian_loans' in subject_data.columns:
        grand_total_agg['asian_loans'] = subject_data['asian_loans'].sum()
    if 'black_loans' in subject_data.columns:
        grand_total_agg['black_loans'] = subject_data['black_loans'].sum()
    if 'native_american_loans' in subject_data.columns:
        grand_total_agg['native_american_loans'] = subject_data['native_american_loans'].sum()
    if 'hopi_loans' in subject_data.columns:
        grand_total_agg['hopi_loans'] = subject_data['hopi_loans'].sum()
    if 'hispanic_loans' in subject_data.columns:
        grand_total_agg['hispanic_loans'] = subject_data['hispanic_loans'].sum()
    
    grand_total_loans = grand_total_agg.get('total_loans', 0)
    
    # Calculate peer grand total if available
    peer_grand_total_agg = {}
    peer_grand_total_loans = 0
    if peer_data is not None and not peer_data.empty:
        if 'total_loans' in peer_data.columns:
            peer_grand_total_agg['total_loans'] = peer_data['total_loans'].sum()
        if 'lmict_loans' in peer_data.columns:
            peer_grand_total_agg['lmict_loans'] = peer_data['lmict_loans'].sum()
        if 'lmib_loans' in peer_data.columns:
            peer_grand_total_agg['lmib_loans'] = peer_data['lmib_loans'].sum()
        if 'lmib_amount' in peer_data.columns:
            peer_grand_total_agg['lmib_amount'] = peer_data['lmib_amount'].sum()
        if 'mmct_loans' in peer_data.columns:
            peer_grand_total_agg['mmct_loans'] = peer_data['mmct_loans'].sum()
        if 'minb_loans' in peer_data.columns:
            peer_grand_total_agg['minb_loans'] = peer_data['minb_loans'].sum()
        if 'asian_loans' in peer_data.columns:
            peer_grand_total_agg['asian_loans'] = peer_data['asian_loans'].sum()
        if 'black_loans' in peer_data.columns:
            peer_grand_total_agg['black_loans'] = peer_data['black_loans'].sum()
        if 'native_american_loans' in peer_data.columns:
            peer_grand_total_agg['native_american_loans'] = peer_data['native_american_loans'].sum()
        if 'hopi_loans' in peer_data.columns:
            peer_grand_total_agg['hopi_loans'] = peer_data['hopi_loans'].sum()
        if 'hispanic_loans' in peer_data.columns:
            peer_grand_total_agg['hispanic_loans'] = peer_data['hispanic_loans'].sum()
        peer_grand_total_loans = peer_grand_total_agg.get('total_loans', 0)
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'Loans':
            ws.cell(row, 3).value = int(grand_total_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'LMICT%':
            pct = (grand_total_agg.get('lmict_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB%':
            pct = (grand_total_agg.get('lmib_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB$':
            ws.cell(row, 3).value = float(grand_total_agg.get('lmib_amount', 0))
            ws.cell(row, 3).number_format = '$#,##0'
        elif metric == 'MMCT%':
            pct = (grand_total_agg.get('mmct_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric.endswith('%'):
            metric_map = {
                'MINB%': 'minb_loans',
                'Asian%': 'asian_loans',
                'Black%': 'black_loans',
                'Native American%': 'native_american_loans',
                'HoPI%': 'hopi_loans',
                'Hispanic%': 'hispanic_loans'
            }
            field = metric_map.get(metric, '')
            if field:
                pct = (grand_total_agg.get(field, 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
        
        # Peer Grand Total
        if peer_grand_total_agg:
            if metric == 'Loans':
                ws.cell(row, 4).value = int(peer_grand_total_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'LMICT%':
                pct = (peer_grand_total_agg.get('lmict_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB%':
                pct = (peer_grand_total_agg.get('lmib_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB$':
                ws.cell(row, 4).value = float(peer_grand_total_agg.get('lmib_amount', 0))
                ws.cell(row, 4).number_format = '$#,##0'
            elif metric == 'MMCT%':
                pct = (peer_grand_total_agg.get('mmct_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric.endswith('%'):
                metric_map = {
                    'MINB%': 'minb_loans',
                    'Asian%': 'asian_loans',
                    'Black%': 'black_loans',
                    'Native American%': 'native_american_loans',
                    'HoPI%': 'hopi_loans',
                    'Hispanic%': 'hispanic_loans'
                }
                field = metric_map.get(metric, '')
                if field:
                    pct = (peer_grand_total_agg.get(field, 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                    ws.cell(row, 4).value = pct / 100
                    ws.cell(row, 4).number_format = '0.00%'
        
        # Difference for percentages
        if metric.endswith('%') and metric != 'LMIB$':
            ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group by CBSA and aggregate
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        
        # Sort by total loans
        cbsa_totals = {}
        for cbsa_code, group_data in grouped:
            total = group_data['total_loans'].sum() if 'total_loans' in group_data.columns and not group_data.empty else 0
            cbsa_totals[cbsa_code] = total
        
        sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
        
        # Process each CBSA
        for group_key, group_data in sorted_cbsas:
            cbsa_code = str(group_key)
            
            # Aggregate data
            agg = {}
            for col in ['total_loans', 'lmict_loans', 'lmib_loans', 'lmib_amount', 'mmct_loans',
                       'minb_loans', 'asian_loans', 'black_loans', 'native_american_loans',
                       'hopi_loans', 'hispanic_loans']:
                if col in group_data.columns:
                    agg[col] = group_data[col].sum()
            
            total_loans = agg.get('total_loans', 0)
            
            # Get CBSA name
            cbsa_name = None
            if 'cbsa_name' in group_data.columns and not group_data.empty:
                cbsa_name_val = group_data['cbsa_name'].iloc[0]
                if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                    cbsa_name = str(cbsa_name_val).strip()
            
            if not cbsa_name:
                cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
            
            # Write data for each metric
            for metric in metrics:
                ws.cell(row, 1).value = cbsa_name
                ws.cell(row, 2).value = metric
                
                # Subject data
                if metric == 'Loans':
                    ws.cell(row, 3).value = int(total_loans)
                    ws.cell(row, 3).number_format = '#,##0'
                elif metric == 'LMICT%':
                    pct = (agg.get('lmict_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                    ws.cell(row, 3).value = pct / 100
                    ws.cell(row, 3).number_format = '0.00%'
                elif metric == 'LMIB%':
                    pct = (agg.get('lmib_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                    ws.cell(row, 3).value = pct / 100
                    ws.cell(row, 3).number_format = '0.00%'
                elif metric == 'LMIB$':
                    ws.cell(row, 3).value = float(agg.get('lmib_amount', 0))
                    ws.cell(row, 3).number_format = '$#,##0'
                elif metric == 'MMCT%':
                    pct = (agg.get('mmct_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                    ws.cell(row, 3).value = pct / 100
                    ws.cell(row, 3).number_format = '0.00%'
                elif metric.endswith('%'):
                    metric_map = {
                        'MINB%': 'minb_loans',
                        'Asian%': 'asian_loans',
                        'Black%': 'black_loans',
                        'Native American%': 'native_american_loans',
                        'HoPI%': 'hopi_loans',
                        'Hispanic%': 'hispanic_loans'
                    }
                    field = metric_map.get(metric, '')
                    if field:
                        pct = (agg.get(field, 0) / total_loans * 100) if total_loans > 0 else 0
                        ws.cell(row, 3).value = pct / 100
                        ws.cell(row, 3).number_format = '0.00%'
                
                # Peer data
                if peer_data is not None and not peer_data.empty and 'cbsa_code' in peer_data.columns:
                    peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                    if not peer_group.empty:
                        peer_agg = {}
                        for col in ['total_loans', 'lmict_loans', 'lmib_loans', 'lmib_amount', 'mmct_loans',
                                   'minb_loans', 'asian_loans', 'black_loans', 'native_american_loans',
                                   'hopi_loans', 'hispanic_loans']:
                            if col in peer_group.columns:
                                peer_agg[col] = peer_group[col].sum()
                        
                        peer_total = peer_agg.get('total_loans', 0)
                        
                        if metric == 'Loans':
                            ws.cell(row, 4).value = int(peer_total)
                            ws.cell(row, 4).number_format = '#,##0'
                        elif metric == 'LMICT%':
                            pct = (peer_agg.get('lmict_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                            ws.cell(row, 4).value = pct / 100
                            ws.cell(row, 4).number_format = '0.00%'
                        elif metric == 'LMIB%':
                            pct = (peer_agg.get('lmib_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                            ws.cell(row, 4).value = pct / 100
                            ws.cell(row, 4).number_format = '0.00%'
                        elif metric == 'LMIB$':
                            ws.cell(row, 4).value = float(peer_agg.get('lmib_amount', 0))
                            ws.cell(row, 4).number_format = '$#,##0'
                        elif metric == 'MMCT%':
                            pct = (peer_agg.get('mmct_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                            ws.cell(row, 4).value = pct / 100
                            ws.cell(row, 4).number_format = '0.00%'
                        elif metric.endswith('%'):
                            metric_map = {
                                'MINB%': 'minb_loans',
                                'Asian%': 'asian_loans',
                                'Black%': 'black_loans',
                                'Native American%': 'native_american_loans',
                                'HoPI%': 'hopi_loans',
                                'Hispanic%': 'hispanic_loans'
                            }
                            field = metric_map.get(metric, '')
                            if field:
                                pct = (peer_agg.get(field, 0) / peer_total * 100) if peer_total > 0 else 0
                                ws.cell(row, 4).value = pct / 100
                                ws.cell(row, 4).number_format = '0.00%'
                
                # Difference
                if metric.endswith('%') and metric != 'LMIB$':
                    ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
                    ws.cell(row, 5).number_format = '0.00%'
                
                row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_sb_sheet(wb, lender_name, subject_data, peer_data,
                   header_fill, header_font, border_thin):
    """Create Small Business sheet with Grand Total and CBSA breakdown."""
    sheet_name = f"{lender_name} SB Lending"
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['CBSA Name', 'Metric', 'Subject Lender', 'Peers', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No SB data for {lender_name}")
        return
    
    metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount', 'Loans Rev Under $1m', 'Avg Loan Amt for RUM SB']
    
    # Handle column name variations
    sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in subject_data.columns else 'sb_loans_count'
    lmict_count_col = 'lmict_count' if 'lmict_count' in subject_data.columns else 'lmict_loans_count'
    loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in subject_data.columns else 'loans_rev_under_1m'
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total
    grand_total_sb_loans = subject_data[sb_loans_col].sum() if sb_loans_col in subject_data.columns else 0
    grand_total_lmict = subject_data[lmict_count_col].sum() if lmict_count_col in subject_data.columns else 0
    grand_total_rev_under_1m = subject_data[loans_rev_col].sum() if loans_rev_col in subject_data.columns else 0
    
    # Calculate weighted averages for Grand Total
    if 'lmict_loans_amount' in subject_data.columns:
        grand_total_avg_lmict = subject_data['lmict_loans_amount'].sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    elif 'avg_sb_lmict_loan_amount' in subject_data.columns:
        grand_total_avg_lmict = (subject_data[lmict_count_col] * subject_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    else:
        grand_total_avg_lmict = 0
    
    if 'amount_rev_under_1m' in subject_data.columns:
        grand_total_avg_rev = subject_data['amount_rev_under_1m'].sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    elif 'avg_loan_amt_rum_sb' in subject_data.columns:
        grand_total_avg_rev = (subject_data[loans_rev_col] * subject_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    else:
        grand_total_avg_rev = 0
    
    # Calculate peer grand totals if available
    peer_grand_total_sb_loans = 0
    peer_grand_total_lmict = 0
    peer_grand_total_rev_under_1m = 0
    peer_grand_total_avg_lmict = 0
    peer_grand_total_avg_rev = 0
    if peer_data is not None and not peer_data.empty:
        peer_sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in peer_data.columns else 'sb_loans_count'
        peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_data.columns else 'lmict_loans_count'
        peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_data.columns else 'loans_rev_under_1m'
        
        peer_grand_total_sb_loans = peer_data[peer_sb_loans_col].sum() if peer_sb_loans_col in peer_data.columns else 0
        peer_grand_total_lmict = peer_data[peer_lmict_count_col].sum() if peer_lmict_count_col in peer_data.columns else 0
        peer_grand_total_rev_under_1m = peer_data[peer_loans_rev_col].sum() if peer_loans_rev_col in peer_data.columns else 0
        
        if 'lmict_loans_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = peer_data['lmict_loans_amount'].sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        elif 'avg_sb_lmict_loan_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = (peer_data[peer_lmict_count_col] * peer_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        else:
            peer_grand_total_avg_lmict = 0
        
        if 'amount_rev_under_1m' in peer_data.columns:
            peer_grand_total_avg_rev = peer_data['amount_rev_under_1m'].sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        elif 'avg_loan_amt_rum_sb' in peer_data.columns:
            peer_grand_total_avg_rev = (peer_data[peer_loans_rev_col] * peer_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        else:
            peer_grand_total_avg_rev = 0
    
    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'SB Loans':
            ws.cell(row, 3).value = int(grand_total_sb_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == '#LMICT':
            ws.cell(row, 3).value = int(grand_total_lmict)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg SB LMICT Loan Amount':
            ws.cell(row, 3).value = float(grand_total_avg_lmict)
            ws.cell(row, 3).number_format = '$#,##0'
        elif metric == 'Loans Rev Under $1m':
            ws.cell(row, 3).value = int(grand_total_rev_under_1m)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg Loan Amt for RUM SB':
            ws.cell(row, 3).value = float(grand_total_avg_rev)
            ws.cell(row, 3).number_format = '$#,##0'
        
        # Peer Grand Total
        if peer_data is not None and not peer_data.empty:
            if metric == 'SB Loans':
                ws.cell(row, 4).value = int(peer_grand_total_sb_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == '#LMICT':
                ws.cell(row, 4).value = int(peer_grand_total_lmict)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg SB LMICT Loan Amount':
                ws.cell(row, 4).value = float(peer_grand_total_avg_lmict)
                ws.cell(row, 4).number_format = '$#,##0'
            elif metric == 'Loans Rev Under $1m':
                ws.cell(row, 4).value = int(peer_grand_total_rev_under_1m)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg Loan Amt for RUM SB':
                ws.cell(row, 4).value = float(peer_grand_total_avg_rev)
                ws.cell(row, 4).number_format = '$#,##0'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group by CBSA and aggregate
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        
        # Sort by total SB loans
        cbsa_totals = {}
        for cbsa_code, group_data in grouped:
            total = group_data[sb_loans_col].sum() if sb_loans_col in group_data.columns and not group_data.empty else 0
            cbsa_totals[cbsa_code] = total
        
        sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
        
        # Process each CBSA
        for group_key, group_data in sorted_cbsas:
            cbsa_code = str(group_key)
            
            # Aggregate data
            agg = {}
            if sb_loans_col in group_data.columns:
                agg['sb_loans'] = group_data[sb_loans_col].sum()
            if lmict_count_col in group_data.columns:
                agg['lmict'] = group_data[lmict_count_col].sum()
            if loans_rev_col in group_data.columns:
                agg['rev_under_1m'] = group_data[loans_rev_col].sum()
            
            # Calculate averages
            if 'lmict_loans_amount' in group_data.columns:
                agg['avg_lmict'] = group_data['lmict_loans_amount'].sum() / agg.get('lmict', 1) if agg.get('lmict', 0) > 0 else 0
            elif 'avg_sb_lmict_loan_amount' in group_data.columns:
                agg['avg_lmict'] = (group_data[lmict_count_col] * group_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / agg.get('lmict', 1) if agg.get('lmict', 0) > 0 else 0
            else:
                agg['avg_lmict'] = 0
            
            if 'amount_rev_under_1m' in group_data.columns:
                agg['avg_rev'] = group_data['amount_rev_under_1m'].sum() / agg.get('rev_under_1m', 1) if agg.get('rev_under_1m', 0) > 0 else 0
            elif 'avg_loan_amt_rum_sb' in group_data.columns:
                agg['avg_rev'] = (group_data[loans_rev_col] * group_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / agg.get('rev_under_1m', 1) if agg.get('rev_under_1m', 0) > 0 else 0
            else:
                agg['avg_rev'] = 0
            
            # Get CBSA name
            cbsa_name = None
            if 'cbsa_name' in group_data.columns and not group_data.empty:
                cbsa_name_val = group_data['cbsa_name'].iloc[0]
                if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                    cbsa_name = str(cbsa_name_val).strip()
            
            if not cbsa_name:
                cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
            
            # Write data for each metric
            for metric in metrics:
                ws.cell(row, 1).value = cbsa_name
                ws.cell(row, 2).value = metric
                
                # Subject data
                if metric == 'SB Loans':
                    ws.cell(row, 3).value = int(agg.get('sb_loans', 0))
                    ws.cell(row, 3).number_format = '#,##0'
                elif metric == '#LMICT':
                    ws.cell(row, 3).value = int(agg.get('lmict', 0))
                    ws.cell(row, 3).number_format = '#,##0'
                elif metric == 'Avg SB LMICT Loan Amount':
                    ws.cell(row, 3).value = float(agg.get('avg_lmict', 0))
                    ws.cell(row, 3).number_format = '$#,##0'
                elif metric == 'Loans Rev Under $1m':
                    ws.cell(row, 3).value = int(agg.get('rev_under_1m', 0))
                    ws.cell(row, 3).number_format = '#,##0'
                elif metric == 'Avg Loan Amt for RUM SB':
                    ws.cell(row, 3).value = float(agg.get('avg_rev', 0))
                    ws.cell(row, 3).number_format = '$#,##0'
                
                # Peer data
                if peer_data is not None and not peer_data.empty and 'cbsa_code' in peer_data.columns:
                    peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                    if not peer_group.empty:
                        peer_agg = {}
                        if peer_sb_loans_col in peer_group.columns:
                            peer_agg['sb_loans'] = peer_group[peer_sb_loans_col].sum()
                        if peer_lmict_count_col in peer_group.columns:
                            peer_agg['lmict'] = peer_group[peer_lmict_count_col].sum()
                        if peer_loans_rev_col in peer_group.columns:
                            peer_agg['rev_under_1m'] = peer_group[peer_loans_rev_col].sum()
                        
                        # Calculate peer averages
                        if 'lmict_loans_amount' in peer_group.columns:
                            peer_agg['avg_lmict'] = peer_group['lmict_loans_amount'].sum() / peer_agg.get('lmict', 1) if peer_agg.get('lmict', 0) > 0 else 0
                        elif 'avg_sb_lmict_loan_amount' in peer_group.columns:
                            peer_agg['avg_lmict'] = (peer_group[peer_lmict_count_col] * peer_group['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_agg.get('lmict', 1) if peer_agg.get('lmict', 0) > 0 else 0
                        else:
                            peer_agg['avg_lmict'] = 0
                        
                        if 'amount_rev_under_1m' in peer_group.columns:
                            peer_agg['avg_rev'] = peer_group['amount_rev_under_1m'].sum() / peer_agg.get('rev_under_1m', 1) if peer_agg.get('rev_under_1m', 0) > 0 else 0
                        elif 'avg_loan_amt_rum_sb' in peer_group.columns:
                            peer_agg['avg_rev'] = (peer_group[peer_loans_rev_col] * peer_group['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_agg.get('rev_under_1m', 1) if peer_agg.get('rev_under_1m', 0) > 0 else 0
                        else:
                            peer_agg['avg_rev'] = 0
                        
                        if metric == 'SB Loans':
                            ws.cell(row, 4).value = int(peer_agg.get('sb_loans', 0))
                            ws.cell(row, 4).number_format = '#,##0'
                        elif metric == '#LMICT':
                            ws.cell(row, 4).value = int(peer_agg.get('lmict', 0))
                            ws.cell(row, 4).number_format = '#,##0'
                        elif metric == 'Avg SB LMICT Loan Amount':
                            ws.cell(row, 4).value = float(peer_agg.get('avg_lmict', 0))
                            ws.cell(row, 4).number_format = '$#,##0'
                        elif metric == 'Loans Rev Under $1m':
                            ws.cell(row, 4).value = int(peer_agg.get('rev_under_1m', 0))
                            ws.cell(row, 4).number_format = '#,##0'
                        elif metric == 'Avg Loan Amt for RUM SB':
                            ws.cell(row, 4).value = float(peer_agg.get('avg_rev', 0))
                            ws.cell(row, 4).number_format = '$#,##0'
                
                row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_branch_sheet(wb, lender_name, subject_data, peer_data,
                        header_fill, header_font, border_thin):
    """Create Branch sheet with Grand Total and CBSA breakdown."""
    sheet_name = f"{lender_name} Branches"
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['CBSA Name', 'Metric', 'Subject Lender', 'Peers', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No branch data for {lender_name}")
        return
    
    metrics = ['Branches', 'Deposits', 'Avg Deposits per Branch']
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total
    grand_total_branches = subject_data['total_branches'].sum() if 'total_branches' in subject_data.columns else 0
    grand_total_deposits = subject_data['total_deposits'].sum() if 'total_deposits' in subject_data.columns else 0
    grand_total_avg_deposits = grand_total_deposits / grand_total_branches if grand_total_branches > 0 else 0
    
    # Calculate peer grand totals if available
    peer_grand_total_branches = 0
    peer_grand_total_deposits = 0
    peer_grand_total_avg_deposits = 0
    if peer_data is not None and not peer_data.empty:
        peer_grand_total_branches = peer_data['total_branches'].sum() if 'total_branches' in peer_data.columns else 0
        peer_grand_total_deposits = peer_data['total_deposits'].sum() if 'total_deposits' in peer_data.columns else 0
        peer_grand_total_avg_deposits = peer_grand_total_deposits / peer_grand_total_branches if peer_grand_total_branches > 0 else 0
    
    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'Branches':
            ws.cell(row, 3).value = int(grand_total_branches)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Deposits':
            ws.cell(row, 3).value = float(grand_total_deposits)
            ws.cell(row, 3).number_format = '$#,##0'
        elif metric == 'Avg Deposits per Branch':
            ws.cell(row, 3).value = float(grand_total_avg_deposits)
            ws.cell(row, 3).number_format = '$#,##0'
        
        # Peer Grand Total
        if peer_data is not None and not peer_data.empty:
            if metric == 'Branches':
                ws.cell(row, 4).value = int(peer_grand_total_branches)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Deposits':
                ws.cell(row, 4).value = float(peer_grand_total_deposits)
                ws.cell(row, 4).number_format = '$#,##0'
            elif metric == 'Avg Deposits per Branch':
                ws.cell(row, 4).value = float(peer_grand_total_avg_deposits)
                ws.cell(row, 4).number_format = '$#,##0'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group by CBSA and aggregate
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        
        # Sort by total branches
        cbsa_totals = {}
        for cbsa_code, group_data in grouped:
            total = group_data['total_branches'].sum() if 'total_branches' in group_data.columns and not group_data.empty else 0
            cbsa_totals[cbsa_code] = total
        
        sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
        
        # Process each CBSA
        for group_key, group_data in sorted_cbsas:
            cbsa_code = str(group_key)
            
            # Aggregate data
            total_branches = group_data['total_branches'].sum() if 'total_branches' in group_data.columns else 0
            total_deposits = group_data['total_deposits'].sum() if 'total_deposits' in group_data.columns else 0
            avg_deposits = total_deposits / total_branches if total_branches > 0 else 0
            
            # Get CBSA name
            cbsa_name = None
            if 'cbsa_name' in group_data.columns and not group_data.empty:
                cbsa_name_val = group_data['cbsa_name'].iloc[0]
                if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                    cbsa_name = str(cbsa_name_val).strip()
            
            if not cbsa_name:
                cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
            
            # Write data for each metric
            for metric in metrics:
                ws.cell(row, 1).value = cbsa_name
                ws.cell(row, 2).value = metric
                
                # Subject data
                if metric == 'Branches':
                    ws.cell(row, 3).value = int(total_branches)
                    ws.cell(row, 3).number_format = '#,##0'
                elif metric == 'Deposits':
                    ws.cell(row, 3).value = float(total_deposits)
                    ws.cell(row, 3).number_format = '$#,##0'
                elif metric == 'Avg Deposits per Branch':
                    ws.cell(row, 3).value = float(avg_deposits)
                    ws.cell(row, 3).number_format = '$#,##0'
                
                # Peer data
                if peer_data is not None and not peer_data.empty and 'cbsa_code' in peer_data.columns:
                    peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                    if not peer_group.empty:
                        peer_total_branches = peer_group['total_branches'].sum() if 'total_branches' in peer_group.columns else 0
                        peer_total_deposits = peer_group['total_deposits'].sum() if 'total_deposits' in peer_group.columns else 0
                        peer_avg_deposits = peer_total_deposits / peer_total_branches if peer_total_branches > 0 else 0
                        
                        if metric == 'Branches':
                            ws.cell(row, 4).value = int(peer_total_branches)
                            ws.cell(row, 4).number_format = '#,##0'
                        elif metric == 'Deposits':
                            ws.cell(row, 4).value = float(peer_total_deposits)
                            ws.cell(row, 4).number_format = '$#,##0'
                        elif metric == 'Avg Deposits per Branch':
                            ws.cell(row, 4).value = float(peer_avg_deposits)
                            ws.cell(row, 4).number_format = '$#,##0'
                
                row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_notes_sheet(wb, lender_name, hmda_years, sb_years, metadata):
    """Create Notes/Methodology sheet."""
    ws = wb.create_sheet("Notes")
    
    row = 1
    ws.cell(row, 1).value = "METHODS AND DEFINITIONS"
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2
    
    ws.cell(row, 1).value = "Subject Lender:"
    ws.cell(row, 2).value = lender_name
    row += 1
    
    if hmda_years:
        ws.cell(row, 1).value = "HMDA Years:"
        ws.cell(row, 2).value = ", ".join(map(str, sorted(hmda_years)))
        row += 1
    
    if sb_years:
        ws.cell(row, 1).value = "Small Business Years:"
        ws.cell(row, 2).value = ", ".join(map(str, sorted(sb_years)))
        row += 1
    
    ws.cell(row, 1).value = "Generated:"
    ws.cell(row, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 2
    
    ws.cell(row, 1).value = "Peer Comparison:"
    ws.cell(row, 2).value = "Peer lenders are identified based on similar lending volume (50%-200% of subject lender volume)."
    row += 1
    ws.cell(row, 2).value = "Peer statistics represent averages across all peer lenders in the selected geography."
    row += 2
    
    ws.cell(row, 1).value = "Data Sources:"
    ws.cell(row, 2).value = "HMDA, Small Business Lending (CRA), and FDIC Summary of Deposits (SOD) data compiled and maintained in NCRC's curated BigQuery databases."
    row += 2
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

