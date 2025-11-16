"""
Report builder module for creating Excel reports from BigQuery data.
Handles data processing, trend analysis, and Excel export.
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Any, Tuple
from datetime import datetime
import os


def build_report(raw_data: List[Dict[str, Any]], counties: List[str], years: List[int]) -> Dict[str, pd.DataFrame]:
    """
    Process raw BigQuery data and build comprehensive report dataframes.
    
    Args:
        raw_data: List of dictionaries from BigQuery results
        counties: List of counties in the report
        years: List of years in the report
        
    Returns:
        Dictionary containing multiple dataframes for different report sections
    """
    if not raw_data:
        raise ValueError("No data provided for report building")
    
    # Convert to DataFrame
    df = pd.DataFrame(raw_data)
    
    
    # Ensure required columns exist
    required_columns = ['bank_name', 'year', 'geoid5', 'county_state', 'total_branches', 'lmict', 'mmct']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Clean and prepare data
    df = clean_data(df)
    
    # Calculate HHI for deposits in latest year
    hhi_data = calculate_hhi(df)
    
    # Build different report sections
    report_data = {
        'summary': create_summary_table(df, counties, years),
        'by_bank': create_bank_summary(df, years),  # Pass years for net change calculation
        'by_county': create_county_summary(df, counties, years),
        'trends': create_trend_analysis(df, years),
        'raw_data': df,
        'hhi': hhi_data
    }
    
    return report_data


def calculate_hhi(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calculate Herfindahl-Hirschman Index (HHI) for bank deposits in the latest year.
    
    HHI Interpretation:
    - HHI < 1,500: Low concentration (competitive market)
    - HHI 1,500-2,500: Moderate concentration
    - HHI > 2,500: High concentration
    
    Returns:
        Dictionary with HHI value, concentration level, year, and top banks by deposits
    """
    if 'total_deposits' not in df.columns:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': None,
            'total_deposits': None,
            'top_banks': []
        }
    
    # Find latest year
    latest_year = df['year'].max()
    latest_year_df = df[df['year'] == latest_year].copy()
    
    # Aggregate deposits by bank (across all counties)
    bank_deposits = latest_year_df.groupby('bank_name')['total_deposits'].sum().reset_index()
    bank_deposits = bank_deposits[bank_deposits['total_deposits'] > 0]  # Remove banks with zero deposits
    
    if bank_deposits.empty:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': latest_year,
            'total_deposits': 0,
            'top_banks': []
        }
    
    total_deposits = bank_deposits['total_deposits'].sum()
    
    # Calculate market shares (as percentages)
    bank_deposits['market_share'] = (bank_deposits['total_deposits'] / total_deposits) * 100
    
    # Calculate HHI: sum of squared market shares (0-10,000 scale)
    hhi = (bank_deposits['market_share'] ** 2).sum()
    
    # Determine concentration level
    if hhi < 1500:
        concentration_level = 'Low concentration (competitive market)'
    elif hhi < 2500:
        concentration_level = 'Moderate concentration'
    else:
        concentration_level = 'High concentration'
    
    # Get top banks by deposits
    top_banks = bank_deposits.nlargest(10, 'total_deposits')[['bank_name', 'total_deposits', 'market_share']].to_dict('records')
    
    return {
        'hhi': round(hhi, 2),
        'concentration_level': concentration_level,
        'year': int(latest_year),
        'total_deposits': int(total_deposits),
        'top_banks': top_banks,
        'total_banks': len(bank_deposits)
    }


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and prepare the raw data."""
    # Convert numeric columns
    numeric_columns = ['total_branches', 'lmict', 'mmct']
    if 'total_deposits' in df.columns:
        numeric_columns.append('total_deposits')
    
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Map lmict/mmct to br_lmi/br_minority for compatibility with existing code
    if 'lmict' in df.columns:
        df['br_lmi'] = df['lmict'].astype(int)
    if 'mmct' in df.columns:
        df['br_minority'] = df['mmct'].astype(int)
    
    # Ensure uninumbr exists (it should come from SQL query)
    if 'uninumbr' not in df.columns:
        # If missing, try to create a unique identifier from geoid5 + bank_name + year
        # This is a fallback, but ideally uninumbr should come from the query
        if 'geoid5' in df.columns and 'bank_name' in df.columns and 'year' in df.columns:
            df['uninumbr'] = df['geoid5'].astype(str) + '_' + df['bank_name'].astype(str) + '_' + df['year'].astype(str)
        else:
            raise ValueError("Missing required columns: uninumbr (and cannot create fallback)")
    
    # Convert year to integer
    df['year'] = pd.to_numeric(df['year'], errors='coerce').fillna(0).astype(int)
    
    # Clean bank names - remove "National Association" and variations
    if 'bank_name' in df.columns:
        df['bank_name'] = df['bank_name'].str.strip()
        # Remove "National Association" and variations (case-insensitive)
        df['bank_name'] = df['bank_name'].str.replace(r'\s+National\s+Association\s*$', '', case=False, regex=True)
        df['bank_name'] = df['bank_name'].str.replace(r'\s+National\s+Assoc\.?\s*$', '', case=False, regex=True)
        df['bank_name'] = df['bank_name'].str.replace(r'\s+N\.?A\.?\s*$', '', case=False, regex=True)
        # Clean up any extra whitespace
    df['bank_name'] = df['bank_name'].str.strip()
    
    # Remove rows with invalid data
    df = df[df['year'] > 0]
    if 'total_branches' in df.columns:
        df = df[df['total_branches'] >= 0]
    
    return df


def create_summary_table(df: pd.DataFrame, counties: List[str], years: List[int]) -> pd.DataFrame:
    """Create a high-level summary table with years as columns and variables as rows.
    Properly dedupes: LMI only, MMCT only, and both LMICT/MMCT."""
    
    # Need to work with unique branches (uninumbr) to properly dedupe
    if 'uninumbr' not in df.columns or 'br_lmi' not in df.columns or 'br_minority' not in df.columns:
        raise ValueError("Missing required columns: uninumbr, br_lmi, br_minority")
    
    result_data = {
        'Variable': ['Total Branches', 'LMI Only Branches', 'MMCT Only Branches', 'Both LMICT/MMCT Branches']
    }
    
    # Calculate for each year
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Count unique total branches
        total_unique = year_df['uninumbr'].nunique()
    
        # Count unique branches that are LMI only (LMI=1, MMCT=0)
        lmi_only = year_df[(year_df['br_lmi'] == 1) & (year_df['br_minority'] == 0)]['uninumbr'].nunique()
        
        # Count unique branches that are MMCT only (LMI=0, MMCT=1)
        mmct_only = year_df[(year_df['br_lmi'] == 0) & (year_df['br_minority'] == 1)]['uninumbr'].nunique()
        
        # Count unique branches that are both (LMI=1, MMCT=1)
        both = year_df[(year_df['br_lmi'] == 1) & (year_df['br_minority'] == 1)]['uninumbr'].nunique()
        
        result_data[str(year)] = [
            int(total_unique),
            int(lmi_only),
            int(mmct_only),
            int(both)
        ]
    
    result = pd.DataFrame(result_data)
    
    # Calculate net change from first to last year
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        
        net_changes = []
        for idx, row in result.iterrows():
            first_val = row[first_year]
            last_val = row[last_year]
            net_change = last_val - first_val
            net_changes.append(net_change)
        
        result['Net Change'] = net_changes
    
    return result


def create_bank_summary(df: pd.DataFrame, years: List[int] = None) -> pd.DataFrame:
    """Create summary by bank for the latest year with proper deduplication.
    Returns all banks (for export) but marks top 10 for display.
    Includes net change from first to last year."""
    
    if years is None:
        years = sorted(df['year'].unique().tolist())
    
    latest_year = max(years)
    first_year = min(years)
    
    # Filter to latest year for current counts
    latest_year_df = df[df['year'] == latest_year].copy()
    
    # Filter to first year for net change calculation
    first_year_df = df[df['year'] == first_year].copy() if first_year != latest_year else latest_year_df.copy()
    
    # Need unique branches per bank
    if 'uninumbr' not in df.columns or 'br_lmi' not in df.columns or 'br_minority' not in df.columns:
        raise ValueError("Missing required columns: uninumbr, br_lmi, br_minority")
    
    bank_data = []
    
    # Get all unique banks
    all_banks = latest_year_df['bank_name'].unique()
    
    for bank_name in all_banks:
        # Latest year data for this bank
        bank_latest = latest_year_df[latest_year_df['bank_name'] == bank_name]
        bank_first = first_year_df[first_year_df['bank_name'] == bank_name] if first_year != latest_year else bank_latest
        
        # Count unique branches
        total_branches_latest = bank_latest['uninumbr'].nunique()
        total_branches_first = bank_first['uninumbr'].nunique()
        
        # Dedupe: LMI only, MMCT only, both
        lmi_only_latest = bank_latest[(bank_latest['br_lmi'] == 1) & (bank_latest['br_minority'] == 0)]['uninumbr'].nunique()
        mmct_only_latest = bank_latest[(bank_latest['br_lmi'] == 0) & (bank_latest['br_minority'] == 1)]['uninumbr'].nunique()
        both_latest = bank_latest[(bank_latest['br_lmi'] == 1) & (bank_latest['br_minority'] == 1)]['uninumbr'].nunique()
        
        # Calculate percentages
        lmi_only_pct = round((lmi_only_latest / total_branches_latest * 100), 1) if total_branches_latest > 0 else 0.0
        mmct_only_pct = round((mmct_only_latest / total_branches_latest * 100), 1) if total_branches_latest > 0 else 0.0
        both_pct = round((both_latest / total_branches_latest * 100), 1) if total_branches_latest > 0 else 0.0
        
        # Net change
        net_change = total_branches_latest - total_branches_first
        
        bank_data.append({
            'Bank Name': bank_name,
            'Total Branches': total_branches_latest,
            'LMI Only Branches': f"{lmi_only_latest} ({lmi_only_pct}%)",
            'MMCT Only Branches': f"{mmct_only_latest} ({mmct_only_pct}%)",
            'Both LMICT/MMCT Branches': f"{both_latest} ({both_pct}%)",
            'Net Change': net_change,
            # Store raw values for sorting
            '_total_branches': total_branches_latest,
            '_lmi_only': lmi_only_latest,
            '_mmct_only': mmct_only_latest,
            '_both': both_latest
        })
    
    result = pd.DataFrame(bank_data)
    
    # Sort by total branches descending
    result = result.sort_values('_total_branches', ascending=False).reset_index(drop=True)
    
    # Mark top 10 (for display purposes, but keep all for export)
    result['_is_top_10'] = result.index < 10
    
    # Drop helper columns
    result = result.drop(columns=['_total_branches', '_lmi_only', '_mmct_only', '_both', '_is_top_10'])
    
    return result


def create_county_summary(df: pd.DataFrame, counties: List[str] = None, years: List[int] = None) -> pd.DataFrame:
    """Create summary by county for the most recent year. Returns empty DataFrame if only one county.
    Includes deduped LMI Only, MMCT Only, and Both LMICT/MMCT columns."""
    # Only create this table if multiple counties
    if counties and len(counties) <= 1:
        return pd.DataFrame()
    
    if not years or len(years) == 0:
        return pd.DataFrame()
    
    # Use the most recent year instead of hardcoded 2024
    latest_year = max(years)
    df_latest = df[df['year'] == latest_year].copy()
    
    if df_latest.empty:
        return pd.DataFrame()
    
    # Need to work with unique branches (uninumbr) to properly dedupe
    if 'uninumbr' not in df_latest.columns or 'br_lmi' not in df_latest.columns or 'br_minority' not in df_latest.columns:
        raise ValueError("Missing required columns: uninumbr, br_lmi, br_minority")
    
    county_data = []
    
    # Process each county
    for county in df_latest['county_state'].unique():
        county_df = df_latest[df_latest['county_state'] == county]
        
        # Count unique total branches
        total_unique = county_df['uninumbr'].nunique()
        
        # Dedupe: LMI only, MMCT only, both
        lmi_only = county_df[(county_df['br_lmi'] == 1) & (county_df['br_minority'] == 0)]['uninumbr'].nunique()
        mmct_only = county_df[(county_df['br_lmi'] == 0) & (county_df['br_minority'] == 1)]['uninumbr'].nunique()
        both = county_df[(county_df['br_lmi'] == 1) & (county_df['br_minority'] == 1)]['uninumbr'].nunique()
        
        # Count unique banks
        num_banks = county_df['bank_name'].nunique()
        
        # Calculate percentages
        lmi_only_pct = round((lmi_only / total_unique * 100), 1) if total_unique > 0 else 0.0
        mmct_only_pct = round((mmct_only / total_unique * 100), 1) if total_unique > 0 else 0.0
        both_pct = round((both / total_unique * 100), 1) if total_unique > 0 else 0.0
        
        county_data.append({
            'County': county,
            'Total Branches': total_unique,
            'LMI Only Branches': f"{lmi_only} ({lmi_only_pct}%)",
            'MMCT Only Branches': f"{mmct_only} ({mmct_only_pct}%)",
            'Both LMICT/MMCT Branches': f"{both} ({both_pct}%)",
            'Number of Banks': num_banks
        })
    
    result = pd.DataFrame(county_data)
    
    # Sort by Total Branches descending
    result = result.sort_values('Total Branches', ascending=False).reset_index(drop=True)
    
    return result


def create_trend_analysis(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """Create trend analysis across years."""
    if len(years) < 2:
        # Not enough years for trend analysis
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_totals = df.groupby('year').agg({
        'total_branches': 'sum',
        'lmict': 'sum',
        'mmct': 'sum',
        'bank_name': 'nunique'
    }).reset_index()
    
    # Calculate year-over-year changes
    yearly_totals['Total Branches YoY %'] = yearly_totals['total_branches'].pct_change() * 100
    yearly_totals['LMI Branches YoY %'] = yearly_totals['lmict'].pct_change() * 100
    yearly_totals['Minority Branches YoY %'] = yearly_totals['mmct'].pct_change() * 100
    
    # Calculate percentages
    yearly_totals['LMI %'] = (yearly_totals['lmict'] / yearly_totals['total_branches'] * 100).round(1)
    yearly_totals['Minority %'] = (yearly_totals['mmct'] / yearly_totals['total_branches'] * 100).round(1)
    
    # Rename columns
    yearly_totals.columns = ['Year', 'Total Branches', 'LMI Branches', 'Minority Branches', 'Number of Banks', 
                           'Total Branches YoY %', 'LMI Branches YoY %', 'Minority Branches YoY %', 'LMI %', 'Minority %']
    
    return yearly_totals.round(1)


def save_excel_report(report_data: Dict[str, pd.DataFrame], output_path: str, metadata: Dict[str, Any] = None):
    """
    Save the report data to an Excel file with multiple sheets.
    
    Args:
        report_data: Dictionary containing dataframes for different report sections
        output_path: Path where the Excel file should be saved
        metadata: Optional dictionary with report metadata (counties, years, etc.)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Write report tables matching the report sections
        # Section 1: Yearly Breakdown
        if 'summary' in report_data and not report_data['summary'].empty:
            report_data['summary'].to_excel(writer, sheet_name='Section 1: Yearly Breakdown', index=False)
        
        # Section 2: Analysis by Bank
        if 'by_bank' in report_data and not report_data['by_bank'].empty:
            report_data['by_bank'].to_excel(writer, sheet_name='Section 2: Analysis by Bank', index=False)
        
        # Section 3: County by County Analysis (only if multiple counties)
        if 'by_county' in report_data and not report_data['by_county'].empty:
            # Check if multiple counties were selected
            multiple_counties = False
            if metadata and 'counties' in metadata:
                counties = metadata['counties']
                if isinstance(counties, list) and len(counties) > 1:
                    multiple_counties = True
            
            if multiple_counties:
                report_data['by_county'].to_excel(writer, sheet_name='Section 3: County by County', index=False)
        
        # Raw Data sheet
        if 'raw_data' in report_data and not report_data['raw_data'].empty:
            report_data['raw_data'].to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Create Notes sheet with metadata
        workbook = writer.book
        notes_sheet = workbook.create_sheet('Notes', 0)  # Insert as first sheet
        
        # Prepare notes content
        notes_content = []
        
        # Report Information
        notes_content.append(('Report Information', ''))
        notes_content.append(('', ''))
        if metadata:
            if 'generated_at' in metadata:
                from datetime import datetime
                try:
                    gen_date = datetime.fromisoformat(metadata['generated_at'].replace('Z', '+00:00'))
                    notes_content.append(('Report Generated:', gen_date.strftime('%Y-%m-%d %H:%M:%S UTC')))
                except:
                    notes_content.append(('Report Generated:', metadata.get('generated_at', 'Not available')))
            else:
                notes_content.append(('Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')))
        else:
            notes_content.append(('Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')))
        notes_content.append(('', ''))
        
        # Dates Analyzed
        notes_content.append(('Dates Analyzed', ''))
        notes_content.append(('', ''))
        if metadata and 'years' in metadata:
            years = metadata['years']
            if isinstance(years, list) and years:
                year_range = f"{min(years)} to {max(years)}"
                notes_content.append(('Year Range:', year_range))
                notes_content.append(('Years Included:', ', '.join(map(str, sorted(years)))))
            else:
                notes_content.append(('Year Range:', str(years)))
        else:
            notes_content.append(('Year Range:', 'Not specified'))
        notes_content.append(('', ''))
        
        # Locations Analyzed
        notes_content.append(('Locations Analyzed', ''))
        notes_content.append(('', ''))
        if metadata and 'counties' in metadata:
            counties = metadata['counties']
            if isinstance(counties, list):
                notes_content.append(('Counties:', ', '.join(counties)))
                notes_content.append(('Number of Counties:', len(counties)))
            else:
                notes_content.append(('Counties:', str(counties)))
        else:
            notes_content.append(('Counties:', 'Not specified'))
        notes_content.append(('', ''))
        
        # Data Filters
        notes_content.append(('Data Filters', ''))
        notes_content.append(('', ''))
        notes_content.append(('Branch Service Types:', 'All branch service types are included in the analysis'))
        notes_content.append(('Data Source:', 'FDIC Summary of Deposits (SOD) data'))
        notes_content.append(('Geographic Level:', 'County level (using geoid5: state FIPS + county FIPS)'))
        notes_content.append(('', ''))
        
        # Data Sources
        notes_content.append(('Data Sources', ''))
        notes_content.append(('', ''))
        notes_content.append(('Primary Source:', 'FDIC Summary of Deposits (SOD)'))
        notes_content.append(('Data Tables:', 'branches.sod, branches.sod_legacy, branches.sod25 (from NCRC datasets)'))
        notes_content.append(('Geographic Crosswalk:', 'geo.cbsa_to_county (for county-state mapping)'))
        notes_content.append(('Data Quality:', 'All data has been downloaded, cleaned, and tested by NCRC Research staff'))
        notes_content.append(('', ''))
        
        # Calculations and Methods
        notes_content.append(('Calculations and Methods', ''))
        notes_content.append(('', ''))
        notes_content.append(('Branch Deduplication:', 'Each branch (uninumbr) appears only once per year. LMI Only, MMCT Only, and Both LMICT/MMCT categories are mutually exclusive.'))
        notes_content.append(('Net Change Calculation:', 'Difference between the first year and final year of the analysis period. Positive values indicate growth, negative values indicate decline.'))
        notes_content.append(('HHI (Herfindahl-Hirschman Index):', 'Calculated using deposit shares of all banks in the study area for the most recent year. HHI < 1,500 = Low concentration; HHI 1,500-2,500 = Moderate concentration; HHI > 2,500 = High concentration.'))
        notes_content.append(('LMI Classification:', 'Branches located in Low-to-Moderate Income (LMI) census tracts'))
        notes_content.append(('MMCT Classification:', 'Branches located in Majority-Minority Census Tracts (MMCT)'))
        notes_content.append(('Deposit Aggregation:', 'Total deposits are summed across all branches for each bank in each county/year'))
        notes_content.append(('', ''))
        
        # Important Notes
        notes_content.append(('Important Notes', ''))
        notes_content.append(('', ''))
        notes_content.append(('Census Boundary Changes:', 'The 2020 Census boundaries took effect in 2022. This resulted in a dramatic increase in the number of middle- and upper-income majority-minority census tracts nationally. Therefore, a dramatic increase in majority-minority branch locations is expected between 2021 and 2022.'))
        notes_content.append(('Bank Name Cleaning:', 'Variations of "National Association" are removed from bank names for consistency.'))
        notes_content.append(('Data Completeness:', 'All branch service types are included in the analysis. The Raw Data sheet contains all available fields for each branch record.'))
        notes_content.append(('', ''))
        
        # Record Count
        if metadata and 'total_records' in metadata:
            notes_content.append(('Data Summary', ''))
            notes_content.append(('', ''))
            notes_content.append(('Total Records:', metadata['total_records']))
            notes_content.append(('', ''))
        
        # Write notes to sheet
        for row_idx, (label, value) in enumerate(notes_content, start=1):
            notes_sheet.cell(row=row_idx, column=1, value=label)
            notes_sheet.cell(row=row_idx, column=2, value=value)
            
            # Style header rows (non-empty labels in first column)
            if label and value == '':
                notes_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
                notes_sheet.cell(row=row_idx, column=1).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Adjust column widths for Notes sheet
        notes_sheet.column_dimensions['A'].width = 35
        notes_sheet.column_dimensions['B'].width = 80
        
        # Auto-adjust column widths for all other sheets
        for sheet_name in writer.sheets:
            if sheet_name != 'Notes':
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_report_metadata(counties: List[str], years: List[int], record_count: int) -> Dict[str, Any]:
    """Generate metadata for the report."""
    return {
        'generated_at': datetime.now().isoformat(),
        'counties': counties,
        'years': years,
        'total_records': record_count,
        'report_type': 'FDIC Bank Branch Analysis'
    }
