"""
Excel generator for MergerMeter analysis output.

This module provides the create_merger_excel function that generates
standardized Excel reports for bank merger analysis.

Updated format to match NCRC MergerMeter expected output:
- Notes sheet with methodology and data sources
- Assessment Areas with Bank Name, CBSA Name, CBSA Code, State, County Name, County Code, State Code
- Mortgage Goals and SB Goals sheets with formulas
- Data sheets named "{SHORT_NAME} MORTGAGE DATA", "{SHORT_NAME} SB DATA", "{SHORT_NAME} BRANCH DATA"
- Percentages stored as raw decimals, Difference shows "--" for non-percentage rows
"""

from pathlib import Path
import pandas as pd
from typing import Dict, Optional, List
import logging
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# NCRC brand colors
NCRC_BLUE = "034EA0"
HEADER_FILL = PatternFill(start_color=NCRC_BLUE, end_color=NCRC_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _get_short_bank_name(bank_name: str) -> str:
    """
    Extract a short identifier from bank name for sheet naming.
    E.g., "FIFTH THIRD BANK" -> "FIFTH", "COMERICA BANK" -> "COMERICA"
    """
    if not bank_name:
        return "BANK"
    # Common patterns to remove
    remove_patterns = [" BANK", " NATIONAL ASSOCIATION", " N.A.", ", N.A.", " NA"]
    short_name = bank_name.upper()
    for pattern in remove_patterns:
        short_name = short_name.replace(pattern.upper(), "")
    # Take first word if still too long
    parts = short_name.strip().split()
    if len(parts) > 0:
        # For names like "FIFTH THIRD", keep both words
        if len(parts) >= 2 and len(parts[0]) + len(parts[1]) <= 15:
            return f"{parts[0]} {parts[1]}".strip()
        return parts[0].strip()
    return short_name[:15].strip()


def _create_empty_mortgage_sheet(wb: Workbook, bank_name: str, lei: str):
    """Create a mortgage data sheet with explanatory note when no HMDA data exists."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} MORTGAGE DATA"

    # Truncate sheet name if too long (Excel max is 31 chars)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(1, 1, f"{bank_name} - Mortgage Lending Data (HMDA)")
    title_cell.font = Font(bold=True, size=14, color=NCRC_BLUE)
    title_cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:F3')
    note_cell = ws.cell(3, 1, f"No HMDA mortgage data found for {bank_name}")
    note_cell.font = Font(bold=True, size=12, color="C00000")  # Red for emphasis
    note_cell.alignment = Alignment(horizontal="left")

    # Details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"Bank Name: {bank_name}").font = Font(size=11)
    ws.cell(7, 1, f"LEI: {lei if lei else 'Not provided'}").font = Font(size=11)

    # Explanation
    ws.merge_cells('A9:F9')
    ws.cell(9, 1, "This may occur if:").font = Font(bold=True, size=11)
    ws.cell(10, 1, "• The bank does not have HMDA-reportable mortgage lending activity").font = Font(size=11, italic=True)
    ws.cell(11, 1, "• The LEI does not match any records in the HMDA database").font = Font(size=11, italic=True)
    ws.cell(12, 1, "• The bank's lending is outside the specified assessment areas").font = Font(size=11, italic=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 60

    logger.info(f"Created empty mortgage sheet for {bank_name} (LEI: {lei})")


def _create_empty_sb_sheet(wb: Workbook, bank_name: str, sb_id: str):
    """Create a small business data sheet with explanatory note when no SB data exists."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} SB DATA"

    # Truncate sheet name if too long (Excel max is 31 chars)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(1, 1, f"{bank_name} - Small Business Lending Data (CRA)")
    title_cell.font = Font(bold=True, size=14, color=NCRC_BLUE)
    title_cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:F3')
    note_cell = ws.cell(3, 1, f"No CRA small business lending data found for {bank_name}")
    note_cell.font = Font(bold=True, size=12, color="C00000")  # Red for emphasis
    note_cell.alignment = Alignment(horizontal="left")

    # Details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"Bank Name: {bank_name}").font = Font(size=11)
    ws.cell(7, 1, f"SB Respondent ID: {sb_id if sb_id else 'Not provided'}").font = Font(size=11)

    # Explanation
    ws.merge_cells('A9:F9')
    ws.cell(9, 1, "This may occur if:").font = Font(bold=True, size=11)
    ws.cell(10, 1, "• The bank is not required to report CRA small business data").font = Font(size=11, italic=True)
    ws.cell(11, 1, "• The SB Respondent ID does not match any records in the CRA database").font = Font(size=11, italic=True)
    ws.cell(12, 1, "• The bank's lending is outside the specified assessment areas").font = Font(size=11, italic=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 60

    logger.info(f"Created empty SB sheet for {bank_name} (SB ID: {sb_id})")


def create_merger_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    assessment_areas_data: Optional[Dict] = None,
    mortgage_goals_data: Optional[Dict] = None,
    sb_goals_data: Optional[pd.DataFrame] = None,
    bank_a_mortgage_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_data: Optional[pd.DataFrame] = None,
    bank_a_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_a_sb_data: Optional[pd.DataFrame] = None,
    bank_b_sb_data: Optional[pd.DataFrame] = None,
    bank_a_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_b_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_a_branch_data: Optional[pd.DataFrame] = None,
    bank_b_branch_data: Optional[pd.DataFrame] = None,
    years_hmda: Optional[List[int]] = None,
    years_sb: Optional[List[int]] = None,
    baseline_years_hmda: Optional[List[int]] = None,
    baseline_years_sb: Optional[List[int]] = None,
    bank_a_lei: str = "",
    bank_b_lei: str = "",
    bank_a_rssd: str = "",
    bank_b_rssd: str = "",
    bank_a_sb_id: str = "",
    bank_b_sb_id: str = "",
    loan_purpose: str = "",
    action_taken: str = "",
    occupancy_type: str = "",
    total_units: str = "",
    construction_method: str = "",
    not_reverse: str = ""
):
    """
    Create Excel workbook with merger analysis data.

    Args:
        output_path: Path to save the Excel file
        bank_a_name: Name of the acquirer bank
        bank_b_name: Name of the target bank
        assessment_areas_data: Dict with 'acquirer' and 'target' keys containing counties
        mortgage_goals_data: Optional mortgage goals data
        sb_goals_data: Optional small business goals data
        bank_a_mortgage_data: Transformed mortgage data for acquirer
        bank_b_mortgage_data: Transformed mortgage data for target
        bank_a_mortgage_peer_data: Peer mortgage data for acquirer (if separate)
        bank_b_mortgage_peer_data: Peer mortgage data for target (if separate)
        bank_a_sb_data: Transformed SB data for acquirer
        bank_b_sb_data: Transformed SB data for target
        bank_a_sb_peer_data: Peer SB data for acquirer (if separate)
        bank_b_sb_peer_data: Peer SB data for target (if separate)
        bank_a_branch_data: Transformed branch data for acquirer
        bank_b_branch_data: Transformed branch data for target
        years_hmda: List of HMDA years included
        years_sb: List of SB years included
        bank_a_lei: LEI for acquirer
        bank_b_lei: LEI for target
        bank_a_rssd: RSSD ID for acquirer
        bank_b_rssd: RSSD ID for target
        bank_a_sb_id: SB ID for acquirer
        bank_b_sb_id: SB ID for target
        loan_purpose: HMDA loan purpose filter
        action_taken: HMDA action taken filter
        occupancy_type: HMDA occupancy type filter
        total_units: HMDA total units filter
        construction_method: HMDA construction method filter
        not_reverse: HMDA reverse mortgage filter
    """
    logger.info(f"Creating merger Excel report: {output_path}")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create Notes sheet (formerly Summary)
    _create_notes_sheet(
        wb, bank_a_name, bank_b_name,
        years_hmda, years_sb,
        bank_a_lei, bank_b_lei,
        bank_a_rssd, bank_b_rssd,
        bank_a_sb_id, bank_b_sb_id,
        loan_purpose, action_taken,
        occupancy_type, total_units,
        construction_method, not_reverse
    )

    # Create Assessment Areas sheet
    if assessment_areas_data:
        _create_assessment_areas_sheet(
            wb, bank_a_name, bank_b_name, assessment_areas_data
        )

    # Create Mortgage Goals sheet if data provided (using baseline years)
    if mortgage_goals_data:
        goals_hmda_years = baseline_years_hmda if baseline_years_hmda else years_hmda
        _create_mortgage_goals_sheet(wb, mortgage_goals_data, goals_hmda_years)

    # Create SB Goals sheet if data provided (using baseline years)
    if sb_goals_data is not None and not sb_goals_data.empty:
        goals_sb_years = baseline_years_sb if baseline_years_sb else years_sb
        _create_sb_goals_sheet(wb, sb_goals_data, goals_sb_years)

    # Create Mortgage DATA sheets - always create sheets with explanatory note if empty
    if bank_a_mortgage_data is not None and not bank_a_mortgage_data.empty:
        _create_mortgage_data_sheet(
            wb, bank_a_name, bank_a_mortgage_data, bank_a_mortgage_peer_data
        )
    else:
        _create_empty_mortgage_sheet(wb, bank_a_name, bank_a_lei)

    if bank_b_mortgage_data is not None and not bank_b_mortgage_data.empty:
        _create_mortgage_data_sheet(
            wb, bank_b_name, bank_b_mortgage_data, bank_b_mortgage_peer_data
        )
    else:
        _create_empty_mortgage_sheet(wb, bank_b_name, bank_b_lei)

    # Create SB DATA sheets - always create sheets with explanatory note if empty
    if bank_a_sb_data is not None and not bank_a_sb_data.empty:
        _create_sb_data_sheet(
            wb, bank_a_name, bank_a_sb_data, bank_a_sb_peer_data
        )
    else:
        _create_empty_sb_sheet(wb, bank_a_name, bank_a_sb_id)

    if bank_b_sb_data is not None and not bank_b_sb_data.empty:
        _create_sb_data_sheet(
            wb, bank_b_name, bank_b_sb_data, bank_b_sb_peer_data
        )
    else:
        _create_empty_sb_sheet(wb, bank_b_name, bank_b_sb_id)

    # Create Branch DATA sheets
    if bank_a_branch_data is not None and not bank_a_branch_data.empty:
        _create_branch_data_sheet(wb, bank_a_name, bank_a_branch_data)

    if bank_b_branch_data is not None and not bank_b_branch_data.empty:
        _create_branch_data_sheet(wb, bank_b_name, bank_b_branch_data)

    # Save workbook
    wb.save(output_path)
    logger.info(f"Merger Excel report saved to: {output_path}")


def _create_notes_sheet(
    wb: Workbook,
    bank_a_name: str,
    bank_b_name: str,
    years_hmda: Optional[List[int]],
    years_sb: Optional[List[int]],
    bank_a_lei: str,
    bank_b_lei: str,
    bank_a_rssd: str,
    bank_b_rssd: str,
    bank_a_sb_id: str,
    bank_b_sb_id: str,
    loan_purpose: str,
    action_taken: str,
    occupancy_type: str,
    total_units: str,
    construction_method: str,
    not_reverse: str
):
    """Create Notes sheet with methodology and data sources."""
    ws = wb.create_sheet("Notes", 0)

    row = 1
    # Title
    ws.cell(row, 1, "Community Benefits Merger Analysis - Methodology")
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2

    # Bank Information section
    ws.cell(row, 1, "Bank Information")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Acquirer Bank info
    ws.cell(row, 1, f"Acquirer Bank: {bank_a_name}")
    ws.cell(row, 2, f"LEI: {bank_a_lei}" if bank_a_lei else "LEI: N/A")
    ws.cell(row, 3, f"RSSD: {bank_a_rssd}" if bank_a_rssd else "RSSD: N/A")
    ws.cell(row, 4, f"SB Respondent ID: {bank_a_sb_id}" if bank_a_sb_id else "SB Respondent ID: N/A")
    row += 1

    # Target Bank info
    ws.cell(row, 1, f"Target Bank: {bank_b_name}")
    ws.cell(row, 2, f"LEI: {bank_b_lei}" if bank_b_lei else "LEI: N/A")
    ws.cell(row, 3, f"RSSD: {bank_b_rssd}" if bank_b_rssd else "RSSD: N/A")
    ws.cell(row, 4, f"SB Respondent ID: {bank_b_sb_id}" if bank_b_sb_id else "SB Respondent ID: N/A")
    row += 2

    # Data Sources section
    ws.cell(row, 1, "Data Sources")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "HMDA Data:")
    ws.cell(row, 2, "https://ffiec.cfpb.gov/data-publication/snapshot-national-loan-level-dataset")
    row += 1

    ws.cell(row, 1, "Small Business Lending Data:")
    ws.cell(row, 2, "https://www.ffiec.gov/cra/sbl/default.aspx")
    row += 1

    ws.cell(row, 1, "Branch Data (SOD25):")
    ws.cell(row, 2, "https://www.fdic.gov/resources/bankers/summary-of-deposits/")
    row += 2

    # Methodology section
    ws.cell(row, 1, "Methodology")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    hmda_years_str = ", ".join(map(str, years_hmda)) if years_hmda else "N/A"
    sb_years_str = ", ".join(map(str, years_sb)) if years_sb else "N/A"

    ws.cell(row, 1, f"HMDA Years Analyzed: {hmda_years_str}")
    row += 1
    ws.cell(row, 1, f"Small Business Years Analyzed: {sb_years_str}")
    row += 1
    ws.cell(row, 1, "Branch Data Year: 2025")
    row += 2

    # Filters Applied section
    ws.cell(row, 1, "Filters Applied")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Build filter description
    filter_parts = []
    if loan_purpose:
        purpose_map = {'1': 'Home Purchase', '2': 'Home Improvement', '3': 'Refinance',
                       '31': 'Cash-Out Refinance', '32': 'No Cash-out Refinance'}
        filter_parts.append(f"Loan Purpose: {purpose_map.get(loan_purpose, loan_purpose)}")
    if action_taken:
        action_map = {'1': 'Loan Originated'}
        filter_parts.append(f"Action Taken: {action_map.get(action_taken, action_taken)}")
    if occupancy_type:
        filter_parts.append(f"Occupancy Type: Owner-occupied")
    if total_units:
        filter_parts.append(f"Total Units: 1 unit, 2 units, 3 units, 4 units")
    if construction_method:
        filter_parts.append(f"Construction Method: Site-built, Manufactured Home")
    if not_reverse:
        filter_parts.append(f"Reverse Mortgage: Not Reverse Mortgage")

    ws.cell(row, 1, "HMDA Filters:")
    ws.cell(row, 2, "; ".join(filter_parts) if filter_parts else "All loans")
    row += 1

    ws.cell(row, 1, "Peer Definition:")
    ws.cell(row, 2, "Lenders making 50% to 200% of subject bank's application volume in a given year in a CBSA (for mortgage and SB data). For branch data, subject bank compared to all other banks in the CBSA in 2025.")
    row += 2

    # Calculations section
    ws.cell(row, 1, "Calculations and Formulas")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "All percentage calculations performed in BigQuery before export to Excel.")
    row += 1

    ws.cell(row, 1, "Mortgage Goals Sheet:")
    ws.cell(row, 2, "Grand Total rows include data ONLY from states where both banks have assessment areas.")
    row += 1

    ws.cell(row, 1, "SB Goals Sheet:")
    ws.cell(row, 2, "Grand Total rows include data ONLY from states where both banks have assessment areas.")
    row += 1

    ws.cell(row, 1, "Baseline Years for Goals Calculations:")
    ws.cell(row, 2, f"Baseline formulas use data from years {hmda_years_str} (2-year average).")
    row += 1

    ws.cell(row, 1, "Excel Format:")
    ws.cell(row, 2, "Null/missing values displayed as '--'. All calculations use raw decimal percentages.")
    row += 2

    # Race/Ethnicity Classification
    ws.cell(row, 1, "Race/Ethnicity Classification:")
    ws.cell(row, 2, "NCRC hierarchical methodology for HMDA data.")
    row += 2

    # Application Version
    from justdata.shared.utils.versions import get_version
    ws.cell(row, 1, "Application Version")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, "MergerMeter Version:")
    ws.cell(row, 2, get_version('mergermeter'))
    row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30


def _create_assessment_areas_sheet(
    wb: Workbook,
    bank_a_name: str,
    bank_b_name: str,
    assessment_areas_data: Dict
):
    """Create assessment areas sheet with Bank Name, CBSA Name, CBSA Code, State, County Name, County Code, State Code."""
    ws = wb.create_sheet("Assessment Areas")

    # Headers
    headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State', 'County Name', 'County Code', 'State Code']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Get counties
    acquirer_counties = assessment_areas_data.get('acquirer', {}).get('counties', [])
    target_counties = assessment_areas_data.get('target', {}).get('counties', [])

    row = 2

    # Write Bank A (Target) counties first, then Bank B (Acquirer)
    # Based on example file, target bank data comes first
    for county in target_counties:
        ws.cell(row, 1, bank_b_name)  # Bank Name
        ws.cell(row, 2, county.get('cbsa_name', ''))  # CBSA Name
        ws.cell(row, 3, str(county.get('cbsa_code', '')))  # CBSA Code
        ws.cell(row, 4, county.get('state_name', ''))  # State
        ws.cell(row, 5, county.get('county_name', ''))  # County Name
        # County Code is the full GEOID5 or just county portion
        geoid5 = str(county.get('geoid5', ''))
        ws.cell(row, 6, geoid5)  # County Code (GEOID5)
        # State Code is the 2-digit state FIPS
        state_code = geoid5[:2] if len(geoid5) >= 2 else county.get('state_code', '')
        ws.cell(row, 7, state_code)  # State Code
        row += 1

    for county in acquirer_counties:
        ws.cell(row, 1, bank_a_name)  # Bank Name
        ws.cell(row, 2, county.get('cbsa_name', ''))  # CBSA Name
        ws.cell(row, 3, str(county.get('cbsa_code', '')))  # CBSA Code
        ws.cell(row, 4, county.get('state_name', ''))  # State
        ws.cell(row, 5, county.get('county_name', ''))  # County Name
        geoid5 = str(county.get('geoid5', ''))
        ws.cell(row, 6, geoid5)  # County Code (GEOID5)
        state_code = geoid5[:2] if len(geoid5) >= 2 else county.get('state_code', '')
        ws.cell(row, 7, state_code)  # State Code
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    logger.info(f"Created Assessment Areas sheet: {len(acquirer_counties)} acquirer, {len(target_counties)} target counties")


def _create_mortgage_goals_sheet(wb: Workbook, mortgage_goals_data: Dict, years_hmda: Optional[List[int]]):
    """Create Mortgage Goals sheet with state breakdowns and formulas."""
    ws = wb.create_sheet("Mortgage Goals")

    # Headers
    headers = ['State', 'Metric', 'Home Purchase', 'Refinance and Cash-Out Refi',
               'Home Improvement and Home Equity', 'Total', 'HP Goal', 'Refi Goal',
               'HI Goal', 'NCRC Proposal', 'Baseline', 'Total Increase in LMIB Lending']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Row 2: Multipliers
    multipliers = [None, None, 1.1, 1.5, 1, None, 5, 5, 5, None, 5, None]
    for col_idx, val in enumerate(multipliers, 1):
        if val is not None:
            ws.cell(2, col_idx, val)

    # Row 3+ will be populated with state data and formulas
    # This is a template - actual data will come from mortgage_goals_data
    row = 3

    if mortgage_goals_data:
        # Process by state
        states_data = mortgage_goals_data.get('by_state', {})
        grand_totals = mortgage_goals_data.get('grand_total', {})

        # Write Grand Total first
        metrics = ['Loans', '~LMICT', '~LMIB', 'LMIB$', '~MMCT', '~MINB',
                   '~Asian', '~Black', '~Native American', '~HoPI', '~Hispanic']

        for metric in metrics:
            ws.cell(row, 1, 'Grand Total')
            ws.cell(row, 2, metric)

            # Get values from grand_totals
            hp_val = grand_totals.get(f'{metric}_hp', grand_totals.get('home_purchase', {}).get(metric, 0))
            refi_val = grand_totals.get(f'{metric}_refi', grand_totals.get('refinance', {}).get(metric, 0))
            hi_val = grand_totals.get(f'{metric}_hi', grand_totals.get('home_improvement', {}).get(metric, 0))

            ws.cell(row, 3, hp_val)  # Home Purchase
            ws.cell(row, 4, refi_val)  # Refinance
            ws.cell(row, 5, hi_val)  # Home Improvement

            # Total formula
            ws.cell(row, 6, f'=SUM(C{row}:E{row})')

            # Goal formulas
            ws.cell(row, 7, f'=IFERROR(((C{row}/2)*C$2)*G$2,0)')  # HP Goal
            ws.cell(row, 8, f'=IFERROR(((D{row}/2)*D$2)*H$2,0)')  # Refi Goal
            ws.cell(row, 9, f'=IFERROR(((E{row}/2)*E$2)*I$2,0)')  # HI Goal
            ws.cell(row, 10, f'=SUM(G{row}:I{row})')  # NCRC Proposal

            # Baseline and Total Increase only for LMIB$ row
            if metric == 'LMIB$':
                ws.cell(row, 11, f'=IFERROR((F{row}/2)*K$2,0)')
                ws.cell(row, 12, f'=J{row}-K{row}')

            # Apply number formatting based on metric
            if metric == 'LMIB$':
                # Dollar format for LMIB$
                for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                    ws.cell(row, col).number_format = '$#,##0'
            else:
                # Thousands format for loan counts
                for col in [3, 4, 5, 6, 7, 8, 9, 10]:
                    ws.cell(row, col).number_format = '#,##0'

            row += 1

        # Write each state's data
        for state_name, state_data in sorted(states_data.items()):
            for metric in metrics:
                ws.cell(row, 1, state_name)
                ws.cell(row, 2, metric)

                hp_val = state_data.get(f'{metric}_hp', state_data.get('home_purchase', {}).get(metric, 0))
                refi_val = state_data.get(f'{metric}_refi', state_data.get('refinance', {}).get(metric, 0))
                hi_val = state_data.get(f'{metric}_hi', state_data.get('home_improvement', {}).get(metric, 0))

                ws.cell(row, 3, hp_val)
                ws.cell(row, 4, refi_val)
                ws.cell(row, 5, hi_val)

                ws.cell(row, 6, f'=SUM(C{row}:E{row})')
                ws.cell(row, 7, f'=IFERROR(((C{row}/2)*C$2)*G$2,0)')
                ws.cell(row, 8, f'=IFERROR(((D{row}/2)*D$2)*H$2,0)')
                ws.cell(row, 9, f'=IFERROR(((E{row}/2)*E$2)*I$2,0)')
                ws.cell(row, 10, f'=SUM(G{row}:I{row})')

                if metric == 'LMIB$':
                    ws.cell(row, 11, f'=IFERROR((F{row}/2)*K$2,0)')
                    ws.cell(row, 12, f'=J{row}-K{row}')

                # Apply number formatting based on metric
                if metric == 'LMIB$':
                    # Dollar format for LMIB$
                    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                        ws.cell(row, col).number_format = '$#,##0'
                else:
                    # Thousands format for loan counts
                    for col in [3, 4, 5, 6, 7, 8, 9, 10]:
                        ws.cell(row, col).number_format = '#,##0'

                row += 1

    # Adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    logger.info(f"Created Mortgage Goals sheet")


def _create_sb_goals_sheet(wb: Workbook, sb_goals_data: pd.DataFrame, years_sb: Optional[List[int]]):
    """Create SB Goals sheet with state breakdowns and formulas."""
    ws = wb.create_sheet("SB Goals")

    # Headers
    headers = ['State', 'Metric', 'Base Value', 'NCRC Proposal', 'Baseline', 'Total Increase']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Row 2: Multipliers
    multipliers = [None, None, 1.2, 5, 5, None]
    for col_idx, val in enumerate(multipliers, 1):
        if val is not None:
            ws.cell(2, col_idx, val)

    row = 3

    if sb_goals_data is not None and not sb_goals_data.empty:
        # Metrics for SB Goals
        metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount',
                   'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']

        # Group by state
        if 'state_name' in sb_goals_data.columns:
            # Write Grand Total first
            for metric in metrics:
                ws.cell(row, 1, 'Grand Total')
                ws.cell(row, 2, metric)

                # Calculate grand total value based on metric
                if metric == 'SB Loans':
                    val = sb_goals_data['sb_loans_total'].sum() if 'sb_loans_total' in sb_goals_data.columns else 0
                elif metric == '#LMICT':
                    val = sb_goals_data['lmict_count'].sum() if 'lmict_count' in sb_goals_data.columns else 0
                elif metric == 'Avg SB LMICT Loan Amount':
                    lmict_count = sb_goals_data['lmict_count'].sum() if 'lmict_count' in sb_goals_data.columns else 0
                    lmict_amount = sb_goals_data['lmict_loans_amount'].sum() if 'lmict_loans_amount' in sb_goals_data.columns else 0
                    val = lmict_amount / lmict_count if lmict_count > 0 else 0
                elif metric == 'Loans Rev Under $1m':
                    val = sb_goals_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in sb_goals_data.columns else 0
                else:  # Avg Loan Amt for <$1M GAR SB
                    rev_count = sb_goals_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in sb_goals_data.columns else 0
                    rev_amount = sb_goals_data['amount_rev_under_1m'].sum() if 'amount_rev_under_1m' in sb_goals_data.columns else 0
                    val = rev_amount / rev_count if rev_count > 0 else 0

                ws.cell(row, 3, val)  # Base Value
                ws.cell(row, 4, f'=IFERROR(((C{row}/2)*C$2)*D$2,0)')  # NCRC Proposal

                # Baseline and Total Increase only for avg loan amount metrics
                if 'Avg' in metric:
                    ws.cell(row, 5, f'=IFERROR((C{row}/2)*E$2,0)')
                    ws.cell(row, 6, f'=D{row}-E{row}')

                # Apply number formatting based on metric
                if 'Avg' in metric or '$' in metric:
                    # Dollar format for average amounts
                    for col in [3, 4, 5, 6]:
                        ws.cell(row, col).number_format = '$#,##0'
                else:
                    # Thousands format for loan counts
                    for col in [3, 4]:
                        ws.cell(row, col).number_format = '#,##0'

                row += 1

            # Write by state
            for state_name, state_data in sb_goals_data.groupby('state_name'):
                for metric in metrics:
                    ws.cell(row, 1, state_name)
                    ws.cell(row, 2, metric)

                    if metric == 'SB Loans':
                        val = state_data['sb_loans_total'].sum() if 'sb_loans_total' in state_data.columns else 0
                    elif metric == '#LMICT':
                        val = state_data['lmict_count'].sum() if 'lmict_count' in state_data.columns else 0
                    elif metric == 'Avg SB LMICT Loan Amount':
                        lmict_count = state_data['lmict_count'].sum() if 'lmict_count' in state_data.columns else 0
                        lmict_amount = state_data['lmict_loans_amount'].sum() if 'lmict_loans_amount' in state_data.columns else 0
                        val = lmict_amount / lmict_count if lmict_count > 0 else 0
                    elif metric == 'Loans Rev Under $1m':
                        val = state_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in state_data.columns else 0
                    else:
                        rev_count = state_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in state_data.columns else 0
                        rev_amount = state_data['amount_rev_under_1m'].sum() if 'amount_rev_under_1m' in state_data.columns else 0
                        val = rev_amount / rev_count if rev_count > 0 else 0

                    ws.cell(row, 3, val)
                    ws.cell(row, 4, f'=IFERROR(((C{row}/2)*C$2)*D$2,0)')

                    if 'Avg' in metric:
                        ws.cell(row, 5, f'=IFERROR((C{row}/2)*E$2,0)')
                        ws.cell(row, 6, f'=D{row}-E{row}')

                    # Apply number formatting based on metric
                    if 'Avg' in metric or '$' in metric:
                        # Dollar format for average amounts
                        for col in [3, 4, 5, 6]:
                            ws.cell(row, col).number_format = '$#,##0'
                    else:
                        # Thousands format for loan counts
                        for col in [3, 4]:
                            ws.cell(row, col).number_format = '#,##0'

                    row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    logger.info(f"Created SB Goals sheet")


def _create_mortgage_data_sheet(
    wb: Workbook,
    bank_name: str,
    subject_data: pd.DataFrame,
    peer_data: Optional[pd.DataFrame] = None
):
    """Create mortgage data sheet with HMDA metrics - named {SHORT_NAME} MORTGAGE DATA."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} MORTGAGE DATA"
    ws = wb.create_sheet(sheet_name)

    # Headers - use "Assessment Area" instead of "CBSA Name"
    headers = ['Assessment Area', 'Metric', 'Bank', 'Peer', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')

    if subject_data is None or subject_data.empty:
        logger.warning(f"No mortgage data for {bank_name}")
        return

    # Metrics - match expected format
    metrics = [
        'Loans', 'LMICT%', 'LMIB%', 'LMIB$', 'MMCT%',
        'MINB%', 'Asian%', 'Black%', 'Native American%', 'HoPI%', 'Hispanic%'
    ]

    # Calculate Grand Total
    grand_total = _calculate_mortgage_grand_total(subject_data)
    peer_grand_total = _calculate_mortgage_grand_total(peer_data) if peer_data is not None else None

    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1, 'Grand Total').font = Font(bold=True)
        ws.cell(row, 2, metric)
        _write_mortgage_metric_new_format(ws, row, 3, metric, grand_total)
        if peer_grand_total:
            _write_mortgage_metric_new_format(ws, row, 4, metric, peer_grand_total)
        # Difference column - show "--" for Loans and LMIB$, otherwise calculate
        if metric in ['Loans', 'LMIB$']:
            ws.cell(row, 5, '--')
        else:
            # For percentage metrics, show the difference as percentage
            diff_cell = ws.cell(row, 5, f'=IFERROR(C{row}-D{row},0)')
            diff_cell.number_format = '0.00%'
        row += 1

    # Write CBSA-level data
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        # Sort by total loans (descending)
        sorted_groups = sorted(grouped, key=lambda x: x[1]['total_loans'].sum() if 'total_loans' in x[1].columns else 0, reverse=True)

        for cbsa_code, group_data in sorted_groups:
            cbsa_name = _get_cbsa_name(group_data)
            cbsa_total = _calculate_mortgage_cbsa_total(group_data)
            peer_cbsa_total = None
            if peer_data is not None and 'cbsa_code' in peer_data.columns:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    peer_cbsa_total = _calculate_mortgage_cbsa_total(peer_group)

            for metric in metrics:
                ws.cell(row, 1, cbsa_name)
                ws.cell(row, 2, metric)
                _write_mortgage_metric_new_format(ws, row, 3, metric, cbsa_total)
                if peer_cbsa_total:
                    _write_mortgage_metric_new_format(ws, row, 4, metric, peer_cbsa_total)
                # Difference column
                if metric in ['Loans', 'LMIB$']:
                    ws.cell(row, 5, '--')
                else:
                    diff_cell = ws.cell(row, 5, f'=IFERROR(C{row}-D{row},0)')
                    diff_cell.number_format = '0.00%'
                row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

    logger.info(f"Created mortgage DATA sheet for {bank_name}: {row - 2} rows")


def _create_sb_data_sheet(
    wb: Workbook,
    bank_name: str,
    subject_data: pd.DataFrame,
    peer_data: Optional[pd.DataFrame] = None
):
    """Create small business lending data sheet - named {SHORT_NAME} SB DATA."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} SB DATA"
    ws = wb.create_sheet(sheet_name)

    # Headers
    headers = ['Assessment Area', 'Metric', 'Bank', 'Peer', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')

    if subject_data is None or subject_data.empty:
        logger.warning(f"No SB data for {bank_name}")
        return

    # Metrics - match expected format
    metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount', 'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']

    # Determine column names
    sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in subject_data.columns else 'sb_loans_count'
    lmict_col = 'lmict_count' if 'lmict_count' in subject_data.columns else 'lmict_loans_count'
    rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in subject_data.columns else 'loans_rev_under_1m'

    # Calculate Grand Total
    grand_total = _calculate_sb_grand_total(subject_data, sb_loans_col, lmict_col, rev_col)
    peer_grand_total = None
    if peer_data is not None and not peer_data.empty:
        peer_sb_col = 'sb_loans_total' if 'sb_loans_total' in peer_data.columns else 'sb_loans_count'
        peer_lmict_col = 'lmict_count' if 'lmict_count' in peer_data.columns else 'lmict_loans_count'
        peer_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_data.columns else 'loans_rev_under_1m'
        peer_grand_total = _calculate_sb_grand_total(peer_data, peer_sb_col, peer_lmict_col, peer_rev_col)

    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1, 'Grand Total').font = Font(bold=True)
        ws.cell(row, 2, metric)
        _write_sb_metric_new_format(ws, row, 3, metric, grand_total)
        if peer_grand_total:
            _write_sb_metric_new_format(ws, row, 4, metric, peer_grand_total)
        # Difference column - show "--" for SB Loans, calculate difference for others
        if metric == 'SB Loans':
            ws.cell(row, 5, '--')
        else:
            diff_cell = ws.cell(row, 5, f'=IFERROR(C{row}-D{row},0)')
            # Apply dollar format for average amount metrics
            if 'Avg' in metric:
                diff_cell.number_format = '$#,##0'
            else:
                diff_cell.number_format = '#,##0'
        row += 1

    # Write CBSA-level data
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        for cbsa_code, group_data in grouped:
            cbsa_name = _get_cbsa_name(group_data)
            cbsa_total = _calculate_sb_cbsa_total(group_data, sb_loans_col, lmict_col, rev_col)
            peer_cbsa_total = None
            if peer_data is not None and 'cbsa_code' in peer_data.columns:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    peer_sb_col = 'sb_loans_total' if 'sb_loans_total' in peer_group.columns else 'sb_loans_count'
                    peer_lmict_col = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                    peer_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                    peer_cbsa_total = _calculate_sb_cbsa_total(peer_group, peer_sb_col, peer_lmict_col, peer_rev_col)

            for metric in metrics:
                ws.cell(row, 1, cbsa_name)
                ws.cell(row, 2, metric)
                _write_sb_metric_new_format(ws, row, 3, metric, cbsa_total)
                if peer_cbsa_total:
                    _write_sb_metric_new_format(ws, row, 4, metric, peer_cbsa_total)
                if metric == 'SB Loans':
                    ws.cell(row, 5, '--')
                else:
                    diff_cell = ws.cell(row, 5, f'=IFERROR(C{row}-D{row},0)')
                    if 'Avg' in metric:
                        diff_cell.number_format = '$#,##0'
                    else:
                        diff_cell.number_format = '#,##0'
                row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

    logger.info(f"Created SB DATA sheet for {bank_name}: {row - 2} rows")


def _create_branch_data_sheet(wb: Workbook, bank_name: str, branch_data: pd.DataFrame):
    """Create branch data sheet - named {SHORT_NAME} BRANCH DATA."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} BRANCH DATA"
    ws = wb.create_sheet(sheet_name)

    # Headers - use "Other" instead of "Other/Market"
    headers = ['Assessment Area', 'Metric', 'Bank', 'Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')

    if branch_data is None or branch_data.empty:
        logger.warning(f"No branch data for {bank_name}")
        return

    metrics = ['Branches', 'LMICT', 'MMCT']

    # Determine column names
    total_col = 'total_branches' if 'total_branches' in branch_data.columns else 'subject_total_branches'
    lmict_col = 'branches_in_lmict' if 'branches_in_lmict' in branch_data.columns else None
    mmct_col = 'branches_in_mmct' if 'branches_in_mmct' in branch_data.columns else None
    other_total_col = 'other_total_branches' if 'other_total_branches' in branch_data.columns else 'market_total_branches'

    # Calculate Grand Total
    grand_total_subject = branch_data[total_col].sum() if total_col in branch_data.columns else 0
    grand_total_other = branch_data[other_total_col].sum() if other_total_col in branch_data.columns else 0
    grand_lmict = branch_data[lmict_col].sum() if lmict_col and lmict_col in branch_data.columns else 0
    grand_mmct = branch_data[mmct_col].sum() if mmct_col and mmct_col in branch_data.columns else 0
    other_lmict_col = 'other_branches_in_lmict' if 'other_branches_in_lmict' in branch_data.columns else 'market_branches_in_lmict'
    other_mmct_col = 'other_branches_in_mmct' if 'other_branches_in_mmct' in branch_data.columns else 'market_branches_in_mmct'
    grand_other_lmict = branch_data[other_lmict_col].sum() if other_lmict_col in branch_data.columns else 0
    grand_other_mmct = branch_data[other_mmct_col].sum() if other_mmct_col in branch_data.columns else 0

    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1, 'Grand Total').font = Font(bold=True)
        ws.cell(row, 2, metric)

        if metric == 'Branches':
            ws.cell(row, 3, int(grand_total_subject))
            ws.cell(row, 4, int(grand_total_other))
            ws.cell(row, 5, '--')  # No difference for Branches count
        elif metric == 'LMICT':
            ws.cell(row, 3, int(grand_lmict))
            ws.cell(row, 4, int(grand_other_lmict))
            # Difference as percentage point difference
            branches_row = row - 1
            ws.cell(row, 5, f'=IFERROR((C{row}/C{branches_row})-(D{row}/D{branches_row}),0)')
            ws.cell(row, 5).number_format = '0.00%'
        elif metric == 'MMCT':
            ws.cell(row, 3, int(grand_mmct))
            ws.cell(row, 4, int(grand_other_mmct))
            branches_row = row - 2
            ws.cell(row, 5, f'=IFERROR((C{row}/C{branches_row})-(D{row}/D{branches_row}),0)')
            ws.cell(row, 5).number_format = '0.00%'
        row += 1

    # Write CBSA-level data
    if 'cbsa_code' in branch_data.columns:
        grouped = branch_data.groupby('cbsa_code')
        for cbsa_code, group_data in grouped:
            cbsa_name = _get_cbsa_name(group_data)
            subject_total = group_data[total_col].sum() if total_col in group_data.columns else 0
            other_total = group_data[other_total_col].sum() if other_total_col in group_data.columns else 0
            subject_lmict = group_data[lmict_col].sum() if lmict_col and lmict_col in group_data.columns else 0
            subject_mmct = group_data[mmct_col].sum() if mmct_col and mmct_col in group_data.columns else 0
            other_lmict = group_data[other_lmict_col].sum() if other_lmict_col in group_data.columns else 0
            other_mmct = group_data[other_mmct_col].sum() if other_mmct_col in group_data.columns else 0

            for metric in metrics:
                ws.cell(row, 1, cbsa_name)
                ws.cell(row, 2, metric)

                if metric == 'Branches':
                    ws.cell(row, 3, int(subject_total))
                    ws.cell(row, 4, int(other_total))
                    ws.cell(row, 5, '--')
                elif metric == 'LMICT':
                    ws.cell(row, 3, int(subject_lmict))
                    ws.cell(row, 4, int(other_lmict))
                    branches_row = row - 1
                    ws.cell(row, 5, f'=IFERROR((C{row}/C{branches_row})-(D{row}/D{branches_row}),0)')
                    ws.cell(row, 5).number_format = '0.00%'
                elif metric == 'MMCT':
                    ws.cell(row, 3, int(subject_mmct))
                    ws.cell(row, 4, int(other_mmct))
                    branches_row = row - 2
                    ws.cell(row, 5, f'=IFERROR((C{row}/C{branches_row})-(D{row}/D{branches_row}),0)')
                    ws.cell(row, 5).number_format = '0.00%'
                row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    logger.info(f"Created branch DATA sheet for {bank_name}: {row - 2} rows")


# Helper functions

def _get_cbsa_name(group_data: pd.DataFrame) -> str:
    """Extract CBSA name from group data."""
    if 'cbsa_name' in group_data.columns and not group_data.empty:
        cbsa_name = group_data['cbsa_name'].iloc[0]
        if cbsa_name and str(cbsa_name).lower() not in ['nan', 'none', '']:
            return str(cbsa_name).strip()
    if 'cbsa_code' in group_data.columns and not group_data.empty:
        return f"CBSA {group_data['cbsa_code'].iloc[0]}"
    return "Unknown CBSA"


def _calculate_mortgage_grand_total(df: pd.DataFrame) -> Dict:
    """Calculate grand totals for mortgage data."""
    if df is None or df.empty:
        return {}

    return {
        'total_loans': df['total_loans'].sum() if 'total_loans' in df.columns else 0,
        'lmict_loans': df['lmict_loans'].sum() if 'lmict_loans' in df.columns else 0,
        'lmib_loans': df['lmib_loans'].sum() if 'lmib_loans' in df.columns else 0,
        'lmib_amount': df['lmib_amount'].sum() if 'lmib_amount' in df.columns else 0,
        'mmct_loans': df['mmct_loans'].sum() if 'mmct_loans' in df.columns else 0,
        'minb_loans': df['minb_loans'].sum() if 'minb_loans' in df.columns else 0,
        'asian_loans': df['asian_loans'].sum() if 'asian_loans' in df.columns else 0,
        'black_loans': df['black_loans'].sum() if 'black_loans' in df.columns else 0,
        'native_american_loans': df['native_american_loans'].sum() if 'native_american_loans' in df.columns else 0,
        'hopi_loans': df['hopi_loans'].sum() if 'hopi_loans' in df.columns else 0,
        'hispanic_loans': df['hispanic_loans'].sum() if 'hispanic_loans' in df.columns else 0,
    }


def _calculate_mortgage_cbsa_total(group_data: pd.DataFrame) -> Dict:
    """Calculate totals for a single CBSA in mortgage data."""
    return _calculate_mortgage_grand_total(group_data)


def _write_mortgage_metric_new_format(ws, row: int, col: int, metric: str, totals: Dict):
    """Write a mortgage metric value to cell with proper number formatting."""
    if not totals:
        return

    total_loans = totals.get('total_loans', 0)
    cell = ws.cell(row, col)

    if metric == 'Loans':
        cell.value = int(total_loans)
        cell.number_format = '#,##0'  # Thousands with commas, no decimals
    elif metric == 'LMICT%':
        pct = (totals.get('lmict_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'  # Percentage with 2 decimals
    elif metric == 'LMIB%':
        pct = (totals.get('lmib_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'LMIB$':
        cell.value = int(totals.get('lmib_amount', 0))
        cell.number_format = '$#,##0'  # Dollar format
    elif metric == 'MMCT%':
        pct = (totals.get('mmct_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'MINB%':
        pct = (totals.get('minb_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'Asian%':
        pct = (totals.get('asian_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'Black%':
        pct = (totals.get('black_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'Native American%':
        pct = (totals.get('native_american_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'HoPI%':
        pct = (totals.get('hopi_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'
    elif metric == 'Hispanic%':
        pct = (totals.get('hispanic_loans', 0) / total_loans) if total_loans > 0 else 0
        cell.value = pct
        cell.number_format = '0.00%'


def _calculate_sb_grand_total(df: pd.DataFrame, sb_col: str, lmict_col: str, rev_col: str) -> Dict:
    """Calculate grand totals for SB data."""
    if df is None or df.empty:
        return {}

    sb_total = df[sb_col].sum() if sb_col in df.columns else 0
    lmict_total = df[lmict_col].sum() if lmict_col in df.columns else 0
    rev_total = df[rev_col].sum() if rev_col in df.columns else 0

    # Calculate averages
    avg_lmict = 0
    if lmict_total > 0:
        if 'lmict_loans_amount' in df.columns:
            avg_lmict = df['lmict_loans_amount'].sum() / lmict_total
        elif 'avg_sb_lmict_loan_amount' in df.columns:
            avg_lmict = (df[lmict_col] * df['avg_sb_lmict_loan_amount'].fillna(0)).sum() / lmict_total

    avg_rev = 0
    if rev_total > 0:
        if 'amount_rev_under_1m' in df.columns:
            avg_rev = df['amount_rev_under_1m'].sum() / rev_total
        elif 'avg_loan_amt_rum_sb' in df.columns:
            avg_rev = (df[rev_col] * df['avg_loan_amt_rum_sb'].fillna(0)).sum() / rev_total

    return {
        'sb_loans_total': sb_total,
        'lmict_count': lmict_total,
        'loans_rev_under_1m_count': rev_total,
        'avg_sb_lmict_loan_amount': avg_lmict,
        'avg_loan_amt_rum_sb': avg_rev,
    }


def _calculate_sb_cbsa_total(group_data: pd.DataFrame, sb_col: str, lmict_col: str, rev_col: str) -> Dict:
    """Calculate totals for a single CBSA in SB data."""
    return _calculate_sb_grand_total(group_data, sb_col, lmict_col, rev_col)


def _write_sb_metric_new_format(ws, row: int, col: int, metric: str, totals: Dict):
    """Write a SB metric value to cell with proper number formatting."""
    if not totals:
        return

    cell = ws.cell(row, col)

    if metric == 'SB Loans':
        cell.value = int(totals.get('sb_loans_total', 0))
        cell.number_format = '#,##0'  # Thousands with commas
    elif metric == '#LMICT':
        cell.value = int(totals.get('lmict_count', 0))
        cell.number_format = '#,##0'  # Thousands with commas
    elif metric == 'Avg SB LMICT Loan Amount':
        cell.value = float(totals.get('avg_sb_lmict_loan_amount', 0))
        cell.number_format = '$#,##0'  # Dollar format
    elif metric == 'Loans Rev Under $1m':
        cell.value = int(totals.get('loans_rev_under_1m_count', 0))
        cell.number_format = '#,##0'  # Thousands with commas
    elif metric == 'Avg Loan Amt for <$1M GAR SB':
        cell.value = float(totals.get('avg_loan_amt_rum_sb', 0))
        cell.number_format = '$#,##0'  # Dollar format


# Keep old function signatures for backward compatibility
def _write_mortgage_metric(ws, row: int, col: int, metric: str, totals: Dict):
    """Write a mortgage metric value to cell - backward compatible version."""
    _write_mortgage_metric_new_format(ws, row, col, metric, totals)


def _write_sb_metric(ws, row: int, col: int, metric: str, totals: Dict):
    """Write a SB metric value to cell - backward compatible version."""
    _write_sb_metric_new_format(ws, row, col, metric, totals)
