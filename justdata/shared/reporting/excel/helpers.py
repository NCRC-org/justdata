"""Shared helpers for the merger Excel generator."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# NCRC brand colors
NCRC_BLUE = "034EA0"
HEADER_FILL = PatternFill(start_color=NCRC_BLUE, end_color=NCRC_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _normalize_year_column(df: pd.DataFrame, year_col: str) -> pd.DataFrame:
    """Normalize year column values to plain integer strings.

    Handles int (2023), float (2023.0), string ('2023'), and string ('2023.0')
    by converting all to '2023'. This prevents type mismatches when filtering
    DataFrames by year, since BigQuery may return year as int, float, or string
    depending on the query path and whether NaN values coerced the column to float.
    """
    if df is None or df.empty or year_col not in df.columns:
        return df
    df = df.copy()
    df[year_col] = df[year_col].apply(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() != '' else ''
    )
    return df


def _get_short_bank_name(bank_name: str) -> str:
    """
    Extract a short identifier from bank name for sheet naming.
    E.g., "FIFTH THIRD BANK" -> "FIFTH", "COMERICA BANK" -> "COMERICA"
    """
    if not bank_name:
        return "BANK"
    # Common patterns to remove
    remove_patterns = [" BANK", " NATIONAL ASSOCIATION", " N.A.", ", N.A.", " NA"]
    short_name = bank_name.upper()
    for pattern in remove_patterns:
        short_name = short_name.replace(pattern.upper(), "")
    # Take first word if still too long
    parts = short_name.strip().split()
    if len(parts) > 0:
        # For names like "FIFTH THIRD", keep both words
        if len(parts) >= 2 and len(parts[0]) + len(parts[1]) <= 15:
            return f"{parts[0]} {parts[1]}".strip()
        return parts[0].strip()
    return short_name[:15].strip()


def _create_empty_mortgage_sheet(wb: Workbook, bank_name: str, lei: str):
    """Create a mortgage data sheet with explanatory note when no HMDA data exists."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} MORTGAGE DATA"

    # Truncate sheet name if too long (Excel max is 31 chars)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(1, 1, f"{bank_name} - Mortgage Lending Data (HMDA)")
    title_cell.font = Font(bold=True, size=14, color=NCRC_BLUE)
    title_cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:F3')
    note_cell = ws.cell(3, 1, f"No HMDA mortgage data found for {bank_name}")
    note_cell.font = Font(bold=True, size=12, color="C00000")  # Red for emphasis
    note_cell.alignment = Alignment(horizontal="left")

    # Details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"Bank Name: {bank_name}").font = Font(size=11)
    ws.cell(7, 1, f"LEI: {lei if lei else 'Not provided'}").font = Font(size=11)

    # Explanation
    ws.merge_cells('A9:F9')
    ws.cell(9, 1, "This may occur if:").font = Font(bold=True, size=11)
    ws.cell(10, 1, "• The bank does not have HMDA-reportable mortgage lending activity").font = Font(size=11, italic=True)
    ws.cell(11, 1, "• The LEI does not match any records in the HMDA database").font = Font(size=11, italic=True)
    ws.cell(12, 1, "• The bank's lending is outside the specified assessment areas").font = Font(size=11, italic=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 60

    logger.info(f"Created empty mortgage sheet for {bank_name} (LEI: {lei})")


def _create_empty_sb_sheet(wb: Workbook, bank_name: str, sb_id: str):
    """Create a small business data sheet with explanatory note when no SB data exists."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} SB DATA"

    # Truncate sheet name if too long (Excel max is 31 chars)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(sheet_name)

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(1, 1, f"{bank_name} - Small Business Lending Data (CRA)")
    title_cell.font = Font(bold=True, size=14, color=NCRC_BLUE)
    title_cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:F3')
    note_cell = ws.cell(3, 1, f"No CRA small business lending data found for {bank_name}")
    note_cell.font = Font(bold=True, size=12, color="C00000")  # Red for emphasis
    note_cell.alignment = Alignment(horizontal="left")

    # Details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"Bank Name: {bank_name}").font = Font(size=11)
    ws.cell(7, 1, f"SB Respondent ID: {sb_id if sb_id else 'Not provided'}").font = Font(size=11)

    # Explanation
    ws.merge_cells('A9:F9')
    ws.cell(9, 1, "This may occur if:").font = Font(bold=True, size=11)
    ws.cell(10, 1, "• The bank is not required to report CRA small business data").font = Font(size=11, italic=True)
    ws.cell(11, 1, "• The SB Respondent ID does not match any records in the CRA database").font = Font(size=11, italic=True)
    ws.cell(12, 1, "• The bank's lending is outside the specified assessment areas").font = Font(size=11, italic=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 60

    logger.info(f"Created empty SB sheet for {bank_name} (SB ID: {sb_id})")


