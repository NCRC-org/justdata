"""Shared merger Excel report generator (create_merger_excel)."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from pathlib import Path

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
)
from justdata.shared.reporting.excel.worksheets.assessment_areas import (
    _create_assessment_areas_sheet,
)
from justdata.shared.reporting.excel.worksheets.branch_data import (
    _create_branch_data_sheet,
)
from justdata.shared.reporting.excel.worksheets.mortgage_data import (
    _create_mortgage_data_sheet,
)
from justdata.shared.reporting.excel.worksheets.mortgage_goals import (
    _create_mortgage_goals_sheet,
)
from justdata.shared.reporting.excel.worksheets.notes import _create_notes_sheet
from justdata.shared.reporting.excel.worksheets.sb_data import _create_sb_data_sheet
from justdata.shared.reporting.excel.worksheets.sb_goals import _create_sb_goals_sheet

logger = logging.getLogger(__name__)


def create_merger_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    assessment_areas_data: Optional[Dict] = None,
    mortgage_goals_data: Optional[Dict] = None,
    sb_goals_data: Optional[pd.DataFrame] = None,
    bank_a_mortgage_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_data: Optional[pd.DataFrame] = None,
    bank_a_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_b_mortgage_peer_data: Optional[pd.DataFrame] = None,
    bank_a_sb_data: Optional[pd.DataFrame] = None,
    bank_b_sb_data: Optional[pd.DataFrame] = None,
    bank_a_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_b_sb_peer_data: Optional[pd.DataFrame] = None,
    bank_a_branch_data: Optional[pd.DataFrame] = None,
    bank_b_branch_data: Optional[pd.DataFrame] = None,
    years_hmda: Optional[List[int]] = None,
    years_sb: Optional[List[int]] = None,
    baseline_years_hmda: Optional[List[int]] = None,
    baseline_years_sb: Optional[List[int]] = None,
    bank_a_lei: str = "",
    bank_b_lei: str = "",
    bank_a_rssd: str = "",
    bank_b_rssd: str = "",
    bank_a_sb_id: str = "",
    bank_b_sb_id: str = "",
    loan_purpose: str = "",
    action_taken: str = "",
    occupancy_type: str = "",
    total_units: str = "",
    construction_method: str = "",
    not_reverse: str = "",
    single_bank_mode: bool = False
):
    """
    Create Excel workbook with merger analysis data.

    Args:
        output_path: Path to save the Excel file
        bank_a_name: Name of the acquirer bank
        bank_b_name: Name of the target bank
        assessment_areas_data: Dict with 'acquirer' and 'target' keys containing counties
        mortgage_goals_data: Optional mortgage goals data
        sb_goals_data: Optional small business goals data
        bank_a_mortgage_data: Transformed mortgage data for acquirer
        bank_b_mortgage_data: Transformed mortgage data for target
        bank_a_mortgage_peer_data: Peer mortgage data for acquirer (if separate)
        bank_b_mortgage_peer_data: Peer mortgage data for target (if separate)
        bank_a_sb_data: Transformed SB data for acquirer
        bank_b_sb_data: Transformed SB data for target
        bank_a_sb_peer_data: Peer SB data for acquirer (if separate)
        bank_b_sb_peer_data: Peer SB data for target (if separate)
        bank_a_branch_data: Transformed branch data for acquirer
        bank_b_branch_data: Transformed branch data for target
        years_hmda: List of HMDA years included
        years_sb: List of SB years included
        bank_a_lei: LEI for acquirer
        bank_b_lei: LEI for target
        bank_a_rssd: RSSD ID for acquirer
        bank_b_rssd: RSSD ID for target
        bank_a_sb_id: SB ID for acquirer
        bank_b_sb_id: SB ID for target
        loan_purpose: HMDA loan purpose filter
        action_taken: HMDA action taken filter
        occupancy_type: HMDA occupancy type filter
        total_units: HMDA total units filter
        construction_method: HMDA construction method filter
        not_reverse: HMDA reverse mortgage filter
        single_bank_mode: Whether this is a single-bank analysis
    """
    logger.info(f"Creating merger Excel report: {output_path}")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create Notes sheet (formerly Summary)
    _create_notes_sheet(
        wb, bank_a_name, bank_b_name,
        years_hmda, years_sb,
        bank_a_lei, bank_b_lei,
        bank_a_rssd, bank_b_rssd,
        bank_a_sb_id, bank_b_sb_id,
        loan_purpose, action_taken,
        occupancy_type, total_units,
        construction_method, not_reverse,
        single_bank_mode=single_bank_mode
    )

    # Create Assessment Areas sheet
    if assessment_areas_data:
        _create_assessment_areas_sheet(
            wb, bank_a_name, bank_b_name, assessment_areas_data
        )

    # Create Mortgage Goals sheet if data provided (using baseline years)
    if mortgage_goals_data:
        goals_hmda_years = baseline_years_hmda if baseline_years_hmda else years_hmda
        _create_mortgage_goals_sheet(wb, mortgage_goals_data, goals_hmda_years)

    # Create SB Goals sheet if data provided (using baseline years)
    if sb_goals_data is not None and not sb_goals_data.empty:
        goals_sb_years = baseline_years_sb if baseline_years_sb else years_sb
        _create_sb_goals_sheet(wb, sb_goals_data, goals_sb_years)

    # Create Mortgage DATA sheets - always create sheets with explanatory note if empty
    if bank_a_mortgage_data is not None and not bank_a_mortgage_data.empty:
        _create_mortgage_data_sheet(
            wb, bank_a_name, bank_a_mortgage_data, bank_a_mortgage_peer_data,
            years=years_hmda
        )
    else:
        _create_empty_mortgage_sheet(wb, bank_a_name, bank_a_lei)

    if bank_b_mortgage_data is not None and not bank_b_mortgage_data.empty:
        _create_mortgage_data_sheet(
            wb, bank_b_name, bank_b_mortgage_data, bank_b_mortgage_peer_data,
            years=years_hmda
        )
    else:
        _create_empty_mortgage_sheet(wb, bank_b_name, bank_b_lei)

    # Create SB DATA sheets - always create sheets with explanatory note if empty
    if bank_a_sb_data is not None and not bank_a_sb_data.empty:
        _create_sb_data_sheet(
            wb, bank_a_name, bank_a_sb_data, bank_a_sb_peer_data,
            years=years_sb
        )
    else:
        _create_empty_sb_sheet(wb, bank_a_name, bank_a_sb_id)

    if bank_b_sb_data is not None and not bank_b_sb_data.empty:
        _create_sb_data_sheet(
            wb, bank_b_name, bank_b_sb_data, bank_b_sb_peer_data,
            years=years_sb
        )
    else:
        _create_empty_sb_sheet(wb, bank_b_name, bank_b_sb_id)

    # Create Branch DATA sheets
    if bank_a_branch_data is not None and not bank_a_branch_data.empty:
        _create_branch_data_sheet(wb, bank_a_name, bank_a_branch_data)

    if bank_b_branch_data is not None and not bank_b_branch_data.empty:
        _create_branch_data_sheet(wb, bank_b_name, bank_b_branch_data)

    # Save workbook
    wb.save(output_path)
    logger.info(f"Merger Excel report saved to: {output_path}")


