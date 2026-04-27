"""Branch data worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_branch_data_sheet(wb: Workbook, bank_name: str, branch_data: pd.DataFrame):
    """Create branch data sheet - named {SHORT_NAME} BRANCH DATA."""
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} BRANCH DATA"
    ws = wb.create_sheet(sheet_name)

    # Headers - use "Other" instead of "Other/Market"
    headers = ['Assessment Area', 'Metric', 'Bank', 'Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')

    if branch_data is None or branch_data.empty:
        logger.warning(f"No branch data for {bank_name}")
        return

    metrics = ['Branches', 'LMICT', 'MMCT']

    # Determine column names
    total_col = 'total_branches' if 'total_branches' in branch_data.columns else 'subject_total_branches'
    lmict_col = 'branches_in_lmict' if 'branches_in_lmict' in branch_data.columns else None
    mmct_col = 'branches_in_mmct' if 'branches_in_mmct' in branch_data.columns else None
    other_total_col = 'other_total_branches' if 'other_total_branches' in branch_data.columns else 'market_total_branches'

    # Calculate Grand Total
    grand_total_subject = branch_data[total_col].sum() if total_col in branch_data.columns else 0
    grand_total_other = branch_data[other_total_col].sum() if other_total_col in branch_data.columns else 0
    grand_lmict = branch_data[lmict_col].sum() if lmict_col and lmict_col in branch_data.columns else 0
    grand_mmct = branch_data[mmct_col].sum() if mmct_col and mmct_col in branch_data.columns else 0
    other_lmict_col = 'other_branches_in_lmict' if 'other_branches_in_lmict' in branch_data.columns else 'market_branches_in_lmict'
    other_mmct_col = 'other_branches_in_mmct' if 'other_branches_in_mmct' in branch_data.columns else 'market_branches_in_mmct'
    grand_other_lmict = branch_data[other_lmict_col].sum() if other_lmict_col in branch_data.columns else 0
    grand_other_mmct = branch_data[other_mmct_col].sum() if other_mmct_col in branch_data.columns else 0

    # Write Grand Total section
    row = 2
    for metric in metrics:
        ws.cell(row, 1, 'Grand Total').font = Font(bold=True)
        ws.cell(row, 2, metric)

        if metric == 'Branches':
            ws.cell(row, 3, int(grand_total_subject))
            ws.cell(row, 4, int(grand_total_other))
            ws.cell(row, 5, '--')  # No difference for Branches count
        elif metric == 'LMICT':
            ws.cell(row, 3, int(grand_lmict))
            ws.cell(row, 4, int(grand_other_lmict))
            # Difference as percentage point difference (computed)
            diff_val = _compute_branch_pct_diff(grand_lmict, grand_total_subject, grand_other_lmict, grand_total_other)
            ws.cell(row, 5, diff_val)
            ws.cell(row, 5).number_format = '0.00%'
        elif metric == 'MMCT':
            ws.cell(row, 3, int(grand_mmct))
            ws.cell(row, 4, int(grand_other_mmct))
            diff_val = _compute_branch_pct_diff(grand_mmct, grand_total_subject, grand_other_mmct, grand_total_other)
            ws.cell(row, 5, diff_val)
            ws.cell(row, 5).number_format = '0.00%'
        row += 1

    # Write CBSA-level data
    if 'cbsa_code' in branch_data.columns:
        grouped = branch_data.groupby('cbsa_code')
        for cbsa_code, group_data in grouped:
            cbsa_name = _get_cbsa_name(group_data)
            subject_total = group_data[total_col].sum() if total_col in group_data.columns else 0
            other_total = group_data[other_total_col].sum() if other_total_col in group_data.columns else 0
            subject_lmict = group_data[lmict_col].sum() if lmict_col and lmict_col in group_data.columns else 0
            subject_mmct = group_data[mmct_col].sum() if mmct_col and mmct_col in group_data.columns else 0
            other_lmict = group_data[other_lmict_col].sum() if other_lmict_col in group_data.columns else 0
            other_mmct = group_data[other_mmct_col].sum() if other_mmct_col in group_data.columns else 0

            for metric in metrics:
                ws.cell(row, 1, cbsa_name)
                ws.cell(row, 2, metric)

                if metric == 'Branches':
                    ws.cell(row, 3, int(subject_total))
                    ws.cell(row, 4, int(other_total))
                    ws.cell(row, 5, '--')
                elif metric == 'LMICT':
                    ws.cell(row, 3, int(subject_lmict))
                    ws.cell(row, 4, int(other_lmict))
                    diff_val = _compute_branch_pct_diff(subject_lmict, subject_total, other_lmict, other_total)
                    ws.cell(row, 5, diff_val)
                    ws.cell(row, 5).number_format = '0.00%'
                elif metric == 'MMCT':
                    ws.cell(row, 3, int(subject_mmct))
                    ws.cell(row, 4, int(other_mmct))
                    diff_val = _compute_branch_pct_diff(subject_mmct, subject_total, other_mmct, other_total)
                    ws.cell(row, 5, diff_val)
                    ws.cell(row, 5).number_format = '0.00%'
                row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    logger.info(f"Created branch DATA sheet for {bank_name}: {row - 2} rows")


# Helper functions for computing difference values in Python
# (replaces Excel formulas so SheetJS preview and Excel "Enable Editing" both work)

