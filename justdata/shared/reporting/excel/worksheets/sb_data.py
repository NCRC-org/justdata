"""Small business data worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_sb_data_sheet(
    wb: Workbook,
    bank_name: str,
    subject_data: pd.DataFrame,
    peer_data: Optional[pd.DataFrame] = None,
    years: Optional[List] = None
):
    """Create small business lending data sheet - named {SHORT_NAME} SB DATA.

    If multiple years are provided and the data contains a year column,
    creates year-by-year column groups (Bank/Peer/Diff per year) plus a Total group.
    Falls back to single-column layout for single year or missing year data.
    """
    short_name = _get_short_bank_name(bank_name)
    sheet_name = f"{short_name} SB DATA"
    ws = wb.create_sheet(sheet_name)

    if subject_data is None or subject_data.empty:
        logger.warning(f"No SB data for {bank_name}")
        headers = ['Assessment Area', 'Metric', 'Bank', 'Peer', 'Difference']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')
        return

    # Metrics - match expected format
    metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount', 'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']

    # Determine column names for subject
    sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in subject_data.columns else 'sb_loans_count'
    lmict_col = 'lmict_count' if 'lmict_count' in subject_data.columns else 'lmict_loans_count'
    rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in subject_data.columns else 'loans_rev_under_1m'

    # Determine column names for peer
    peer_sb_col = sb_loans_col
    peer_lmict_col = lmict_col
    peer_rev_col = rev_col
    if peer_data is not None and not peer_data.empty:
        peer_sb_col = 'sb_loans_total' if 'sb_loans_total' in peer_data.columns else 'sb_loans_count'
        peer_lmict_col = 'lmict_count' if 'lmict_count' in peer_data.columns else 'lmict_loans_count'
        peer_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_data.columns else 'loans_rev_under_1m'

    # Determine year column
    year_col = None
    if 'year' in subject_data.columns:
        year_col = 'year'
    elif 'activity_year' in subject_data.columns:
        year_col = 'activity_year'

    sorted_years = sorted([str(y) for y in years]) if years else []
    use_yearly = len(sorted_years) >= 2 and year_col is not None

    # Normalize year column to plain strings (handles int/float/string mismatches)
    if use_yearly:
        raw_years = subject_data[year_col].unique().tolist() if year_col in subject_data.columns else []
        raw_types = list(set(type(v).__name__ for v in raw_years))
        print(f"[DEBUG] SB {bank_name}: year_col='{year_col}', raw year values={raw_years}, types={raw_types}")
        logger.info(f"SB {bank_name}: year_col='{year_col}', raw year values={raw_years}, types={raw_types}")

        subject_data = _normalize_year_column(subject_data, year_col)
        peer_data = _normalize_year_column(peer_data, year_col) if peer_data is not None else peer_data

        norm_years = subject_data[year_col].unique().tolist() if year_col in subject_data.columns else []
        print(f"[DEBUG] SB {bank_name}: normalized years={norm_years}, sorted_years={sorted_years}")
        logger.info(f"SB {bank_name}: normalized years={norm_years}, sorted_years={sorted_years}")

    if not use_yearly:
        # --- Original single-column layout ---
        headers = ['Assessment Area', 'Metric', 'Bank', 'Peer', 'Difference']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')

        grand_total = _calculate_sb_grand_total(subject_data, sb_loans_col, lmict_col, rev_col)
        peer_grand_total = None
        if peer_data is not None and not peer_data.empty:
            peer_grand_total = _calculate_sb_grand_total(peer_data, peer_sb_col, peer_lmict_col, peer_rev_col)

        row = 2
        for metric in metrics:
            ws.cell(row, 1, 'Grand Total').font = Font(bold=True)
            ws.cell(row, 2, metric)
            _write_sb_metric_new_format(ws, row, 3, metric, grand_total)
            if peer_grand_total:
                _write_sb_metric_new_format(ws, row, 4, metric, peer_grand_total)
            if metric == 'SB Loans':
                ws.cell(row, 5, '--')
            elif metric in ['#LMICT', 'Loans Rev Under $1m']:
                diff_val = _compute_sb_ratio_diff(metric, grand_total, peer_grand_total)
                diff_cell = ws.cell(row, 5, diff_val)
                diff_cell.number_format = '0.00%'
            elif 'Avg' in metric:
                diff_val = _compute_sb_avg_diff(metric, grand_total, peer_grand_total)
                diff_cell = ws.cell(row, 5, diff_val)
                diff_cell.number_format = '$#,##0'
            row += 1

        if 'cbsa_code' in subject_data.columns:
            grouped = subject_data.groupby('cbsa_code')
            for cbsa_code, group_data in grouped:
                cbsa_name = _get_cbsa_name(group_data)
                cbsa_total = _calculate_sb_cbsa_total(group_data, sb_loans_col, lmict_col, rev_col)
                peer_cbsa_total = None
                if peer_data is not None and 'cbsa_code' in peer_data.columns:
                    peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                    if not peer_group.empty:
                        p_sb = 'sb_loans_total' if 'sb_loans_total' in peer_group.columns else 'sb_loans_count'
                        p_lm = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                        p_rv = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                        peer_cbsa_total = _calculate_sb_cbsa_total(peer_group, p_sb, p_lm, p_rv)
                for metric in metrics:
                    ws.cell(row, 1, cbsa_name)
                    ws.cell(row, 2, metric)
                    _write_sb_metric_new_format(ws, row, 3, metric, cbsa_total)
                    if peer_cbsa_total:
                        _write_sb_metric_new_format(ws, row, 4, metric, peer_cbsa_total)
                    if metric == 'SB Loans':
                        ws.cell(row, 5, '--')
                    elif metric in ['#LMICT', 'Loans Rev Under $1m']:
                        diff_val = _compute_sb_ratio_diff(metric, cbsa_total, peer_cbsa_total)
                        diff_cell = ws.cell(row, 5, diff_val)
                        diff_cell.number_format = '0.00%'
                    elif 'Avg' in metric:
                        diff_val = _compute_sb_avg_diff(metric, cbsa_total, peer_cbsa_total)
                        diff_cell = ws.cell(row, 5, diff_val)
                        diff_cell.number_format = '$#,##0'
                    row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        for col in ['C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
        logger.info(f"Created SB DATA sheet for {bank_name}: {row - 2} rows")
        return

    # --- Year-by-year layout ---
    num_years = len(sorted_years)

    # Two-row header
    for col_idx, header in enumerate(['Assessment Area', 'Metric'], 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, col_idx).fill = HEADER_FILL
        ws.cell(2, col_idx).border = BORDER_THIN

    col = 3
    for year in sorted_years:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        cell = ws.cell(1, col)
        cell.value = str(year)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for c_off in range(3):
            ws.cell(1, col + c_off).border = BORDER_THIN
        for i, sub in enumerate(['Bank', 'Peer', 'Diff']):
            cell = ws.cell(2, col + i)
            cell.value = sub
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 3

    # Total group header
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
    cell = ws.cell(1, col)
    cell.value = 'Total'
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for c_off in range(3):
        ws.cell(1, col + c_off).border = BORDER_THIN
    for i, sub in enumerate(['Bank', 'Peer', 'Diff']):
        cell = ws.cell(2, col + i)
        cell.value = sub
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_sb_section(ws, area_name, subj_df, peer_df, start_row, is_grand=False):
        """Write one SB assessment area section with year-by-year columns."""
        r = start_row
        sb_loans_row = r  # First metric is always SB Loans

        # Pre-compute totals per year and overall
        yr_totals = {}
        for yr in sorted_years:
            yr_str = str(yr)
            yr_s = subj_df[subj_df[year_col].astype(str) == yr_str] if not subj_df.empty else pd.DataFrame()
            yr_p = peer_df[peer_df[year_col].astype(str) == yr_str] if peer_df is not None and not peer_df.empty else pd.DataFrame()
            yr_totals[yr_str] = (
                _calculate_sb_grand_total(yr_s, sb_loans_col, lmict_col, rev_col) if not yr_s.empty else {},
                _calculate_sb_grand_total(yr_p, peer_sb_col, peer_lmict_col, peer_rev_col) if not yr_p.empty else {}
            )

        overall_subj = _calculate_sb_grand_total(subj_df, sb_loans_col, lmict_col, rev_col)
        overall_peer = _calculate_sb_grand_total(peer_df, peer_sb_col, peer_lmict_col, peer_rev_col) if peer_df is not None and not peer_df.empty else {}

        for metric in metrics:
            ws.cell(r, 1, area_name)
            if is_grand:
                ws.cell(r, 1).font = Font(bold=True)
            ws.cell(r, 2, metric)

            c = 3
            for yr in sorted_years:
                yr_str = str(yr)
                s_tot, p_tot = yr_totals[yr_str]

                _write_sb_metric_new_format(ws, r, c, metric, s_tot)
                if p_tot:
                    _write_sb_metric_new_format(ws, r, c + 1, metric, p_tot)

                if metric == 'SB Loans':
                    ws.cell(r, c + 2, '--')
                elif metric in ['#LMICT', 'Loans Rev Under $1m']:
                    diff_val = _compute_sb_ratio_diff(metric, s_tot, p_tot)
                    diff_cell = ws.cell(r, c + 2, diff_val)
                    diff_cell.number_format = '0.00%'
                elif 'Avg' in metric:
                    diff_val = _compute_sb_avg_diff(metric, s_tot, p_tot)
                    diff_cell = ws.cell(r, c + 2, diff_val)
                    diff_cell.number_format = '$#,##0'

                c += 3

            # Total column group (all years combined)
            _write_sb_metric_new_format(ws, r, c, metric, overall_subj)
            if overall_peer:
                _write_sb_metric_new_format(ws, r, c + 1, metric, overall_peer)

            if metric == 'SB Loans':
                ws.cell(r, c + 2, '--')
            elif metric in ['#LMICT', 'Loans Rev Under $1m']:
                diff_val = _compute_sb_ratio_diff(metric, overall_subj, overall_peer)
                diff_cell = ws.cell(r, c + 2, diff_val)
                diff_cell.number_format = '0.00%'
            elif 'Avg' in metric:
                diff_val = _compute_sb_avg_diff(metric, overall_subj, overall_peer)
                diff_cell = ws.cell(r, c + 2, diff_val)
                diff_cell.number_format = '$#,##0'

            r += 1
        return r

    # Grand Total section
    row = 3  # Start after 2-row header
    row = _write_sb_section(ws, 'Grand Total', subject_data, peer_data, row, is_grand=True)

    # CBSA-level sections
    if 'cbsa_code' in subject_data.columns:
        grouped = subject_data.groupby('cbsa_code')
        for cbsa_code, group_data in grouped:
            cbsa_name = _get_cbsa_name(group_data)
            peer_group = None
            if peer_data is not None and 'cbsa_code' in peer_data.columns:
                pg = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not pg.empty:
                    peer_group = pg
            row = _write_sb_section(ws, cbsa_name, group_data, peer_group, row)

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    total_data_cols = (num_years + 1) * 3
    for i in range(total_data_cols):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    logger.info(f"Created SB DATA sheet for {bank_name} (year-by-year): {row - 3} rows, {num_years} years")


