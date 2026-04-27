"""Assessment areas worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_assessment_areas_sheet(
    wb: Workbook,
    bank_a_name: str,
    bank_b_name: str,
    assessment_areas_data: Dict
):
    """Create assessment areas sheet with Bank Name, CBSA Name, CBSA Code, State, County Name, County Code, State Code."""
    ws = wb.create_sheet("Assessment Areas")

    # Headers
    headers = ['Bank Name', 'CBSA Name', 'CBSA Code', 'State', 'County Name', 'County Code', 'State Code']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Get counties
    acquirer_counties = assessment_areas_data.get('acquirer', {}).get('counties', [])
    target_counties = assessment_areas_data.get('target', {}).get('counties', [])

    row = 2

    # Write Bank A (Target) counties first, then Bank B (Acquirer)
    # Based on example file, target bank data comes first
    for county in target_counties:
        ws.cell(row, 1, bank_b_name)  # Bank Name
        ws.cell(row, 2, county.get('cbsa_name', ''))  # CBSA Name
        ws.cell(row, 3, str(county.get('cbsa_code', '')))  # CBSA Code
        ws.cell(row, 4, county.get('state_name', ''))  # State
        ws.cell(row, 5, county.get('county_name', ''))  # County Name
        # County Code is the full GEOID5 or just county portion
        geoid5 = str(county.get('geoid5', ''))
        ws.cell(row, 6, geoid5)  # County Code (GEOID5)
        # State Code is the 2-digit state FIPS
        state_code = geoid5[:2] if len(geoid5) >= 2 else county.get('state_code', '')
        ws.cell(row, 7, state_code)  # State Code
        row += 1

    for county in acquirer_counties:
        ws.cell(row, 1, bank_a_name)  # Bank Name
        ws.cell(row, 2, county.get('cbsa_name', ''))  # CBSA Name
        ws.cell(row, 3, str(county.get('cbsa_code', '')))  # CBSA Code
        ws.cell(row, 4, county.get('state_name', ''))  # State
        ws.cell(row, 5, county.get('county_name', ''))  # County Name
        geoid5 = str(county.get('geoid5', ''))
        ws.cell(row, 6, geoid5)  # County Code (GEOID5)
        state_code = geoid5[:2] if len(geoid5) >= 2 else county.get('state_code', '')
        ws.cell(row, 7, state_code)  # State Code
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    logger.info(f"Created Assessment Areas sheet: {len(acquirer_counties)} acquirer, {len(target_counties)} target counties")


