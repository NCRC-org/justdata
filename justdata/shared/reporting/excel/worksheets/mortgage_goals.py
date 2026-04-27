"""Mortgage goals worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_mortgage_goals_sheet(wb: Workbook, mortgage_goals_data: Dict, years_hmda: Optional[List[int]]):
    """Create Mortgage Goals sheet with state breakdowns and formulas."""
    ws = wb.create_sheet("Mortgage Goals")

    # Headers
    headers = ['State', 'Metric', 'Home Purchase', 'Refinance and Cash-Out Refi',
               'Home Improvement and Home Equity', 'Total', 'HP Goal', 'Refi Goal',
               'HI Goal', 'NCRC Proposal', 'Baseline', 'Total Increase in LMIB Lending']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Row 2: Multipliers
    multipliers = [None, None, 1.1, 1.5, 1, None, 5, 5, 5, None, 5, None]
    for col_idx, val in enumerate(multipliers, 1):
        if val is not None:
            ws.cell(2, col_idx, val)

    # Row 3+ will be populated with state data and formulas
    # This is a template - actual data will come from mortgage_goals_data
    row = 3

    if mortgage_goals_data:
        # Process by state
        states_data = mortgage_goals_data.get('by_state', {})
        grand_totals = mortgage_goals_data.get('grand_total', {})

        # Write Grand Total first
        metrics = ['Loans', '~LMICT', '~LMIB', 'LMIB$', '~MMCT', '~MINB',
                   '~Asian', '~Black', '~Native American', '~HoPI', '~Hispanic']

        for metric in metrics:
            ws.cell(row, 1, 'Grand Total')
            ws.cell(row, 2, metric)

            # Get values from grand_totals
            hp_val = grand_totals.get(f'{metric}_hp', grand_totals.get('home_purchase', {}).get(metric, 0))
            refi_val = grand_totals.get(f'{metric}_refi', grand_totals.get('refinance', {}).get(metric, 0))
            hi_val = grand_totals.get(f'{metric}_hi', grand_totals.get('home_equity', {}).get(metric, 0))

            ws.cell(row, 3, hp_val)  # Home Purchase
            ws.cell(row, 4, refi_val)  # Refinance
            ws.cell(row, 5, hi_val)  # Home Improvement

            # Total (computed)
            total_val = (hp_val or 0) + (refi_val or 0) + (hi_val or 0)
            ws.cell(row, 6, total_val)

            # Goal values (computed) - uses multipliers from row 2
            # Multipliers: C2=1.1, D2=1.5, E2=1, G2=5, H2=5, I2=5, K2=5
            hp_goal = ((hp_val or 0) / 2) * 1.1 * 5 if hp_val else 0
            refi_goal = ((refi_val or 0) / 2) * 1.5 * 5 if refi_val else 0
            hi_goal = ((hi_val or 0) / 2) * 1 * 5 if hi_val else 0
            ws.cell(row, 7, hp_goal)   # HP Goal
            ws.cell(row, 8, refi_goal)  # Refi Goal
            ws.cell(row, 9, hi_goal)   # HI Goal
            ws.cell(row, 10, hp_goal + refi_goal + hi_goal)  # NCRC Proposal

            # Baseline and Total Increase only for LMIB$ row
            if metric == 'LMIB$':
                baseline = (total_val / 2) * 5 if total_val else 0
                ws.cell(row, 11, baseline)
                ws.cell(row, 12, (hp_goal + refi_goal + hi_goal) - baseline)

            # Apply number formatting based on metric
            if metric == 'LMIB$':
                # Dollar format for LMIB$
                for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                    ws.cell(row, col).number_format = '$#,##0'
            else:
                # Thousands format for loan counts
                for col in [3, 4, 5, 6, 7, 8, 9, 10]:
                    ws.cell(row, col).number_format = '#,##0'

            row += 1

        # Write each state's data
        for state_name, state_data in sorted(states_data.items()):
            for metric in metrics:
                ws.cell(row, 1, state_name)
                ws.cell(row, 2, metric)

                hp_val = state_data.get(f'{metric}_hp', state_data.get('home_purchase', {}).get(metric, 0))
                refi_val = state_data.get(f'{metric}_refi', state_data.get('refinance', {}).get(metric, 0))
                hi_val = state_data.get(f'{metric}_hi', state_data.get('home_equity', {}).get(metric, 0))

                ws.cell(row, 3, hp_val)
                ws.cell(row, 4, refi_val)
                ws.cell(row, 5, hi_val)

                # Computed values using same multipliers as grand total
                total_val = (hp_val or 0) + (refi_val or 0) + (hi_val or 0)
                ws.cell(row, 6, total_val)
                hp_goal = ((hp_val or 0) / 2) * 1.1 * 5 if hp_val else 0
                refi_goal = ((refi_val or 0) / 2) * 1.5 * 5 if refi_val else 0
                hi_goal = ((hi_val or 0) / 2) * 1 * 5 if hi_val else 0
                ws.cell(row, 7, hp_goal)
                ws.cell(row, 8, refi_goal)
                ws.cell(row, 9, hi_goal)
                ws.cell(row, 10, hp_goal + refi_goal + hi_goal)

                if metric == 'LMIB$':
                    baseline = (total_val / 2) * 5 if total_val else 0
                    ws.cell(row, 11, baseline)
                    ws.cell(row, 12, (hp_goal + refi_goal + hi_goal) - baseline)

                # Apply number formatting based on metric
                if metric == 'LMIB$':
                    # Dollar format for LMIB$
                    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
                        ws.cell(row, col).number_format = '$#,##0'
                else:
                    # Thousands format for loan counts
                    for col in [3, 4, 5, 6, 7, 8, 9, 10]:
                        ws.cell(row, col).number_format = '#,##0'

                row += 1

    # Adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    logger.info(f"Created Mortgage Goals sheet")


