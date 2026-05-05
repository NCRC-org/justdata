"""Notes / cover worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_notes_sheet(
    wb: Workbook,
    bank_a_name: str,
    bank_b_name: str,
    years_hmda: Optional[List[int]],
    years_sb: Optional[List[int]],
    bank_a_lei: str,
    bank_b_lei: str,
    bank_a_rssd: str,
    bank_b_rssd: str,
    bank_a_sb_id: str,
    bank_b_sb_id: str,
    loan_purpose: str,
    action_taken: str,
    occupancy_type: str,
    total_units: str,
    construction_method: str,
    not_reverse: str,
    single_bank_mode: bool = False
):
    """Create Notes sheet with methodology and data sources."""
    ws = wb.create_sheet("Notes", 0)

    row = 1
    # Title
    ws.cell(row, 1, "Community Benefits Merger Analysis - Methodology")
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2

    # Bank Information section
    ws.cell(row, 1, "Bank Information")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Acquirer Bank info
    ws.cell(row, 1, f"Acquirer Bank: {bank_a_name}")
    ws.cell(row, 2, f"LEI: {bank_a_lei}" if bank_a_lei else "LEI: N/A")
    ws.cell(row, 3, f"RSSD: {bank_a_rssd}" if bank_a_rssd else "RSSD: N/A")
    ws.cell(row, 4, f"SB Respondent ID: {bank_a_sb_id}" if bank_a_sb_id else "SB Respondent ID: N/A")
    row += 1

    # Target Bank info
    ws.cell(row, 1, f"Target Bank: {bank_b_name}")
    ws.cell(row, 2, f"LEI: {bank_b_lei}" if bank_b_lei else "LEI: N/A")
    ws.cell(row, 3, f"RSSD: {bank_b_rssd}" if bank_b_rssd else "RSSD: N/A")
    ws.cell(row, 4, f"SB Respondent ID: {bank_b_sb_id}" if bank_b_sb_id else "SB Respondent ID: N/A")
    row += 2

    # Data Sources section
    ws.cell(row, 1, "Data Sources")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "HMDA Data:")
    ws.cell(row, 2, "https://ffiec.cfpb.gov/data-publication/snapshot-national-loan-level-dataset")
    row += 1

    ws.cell(row, 1, "Small Business Lending Data:")
    ws.cell(row, 2, "https://www.ffiec.gov/cra/sbl/default.aspx")
    row += 1

    ws.cell(row, 1, "Branch Data (SOD25):")
    ws.cell(row, 2, "https://www.fdic.gov/resources/bankers/summary-of-deposits/")
    row += 2

    # Methodology section
    ws.cell(row, 1, "Methodology")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    hmda_years_str = ", ".join(map(str, years_hmda)) if years_hmda else "N/A"
    sb_years_str = ", ".join(map(str, years_sb)) if years_sb else "N/A"

    ws.cell(row, 1, f"HMDA Years Analyzed: {hmda_years_str}")
    row += 1
    ws.cell(row, 1, f"Small Business Years Analyzed: {sb_years_str}")
    row += 1
    ws.cell(row, 1, "Branch Data Year: 2025")
    row += 2

    # Filters Applied section
    ws.cell(row, 1, "Filters Applied")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Build filter description
    filter_parts = []
    if loan_purpose:
        purpose_map = {'1': 'Home Purchase', '2': 'Home Improvement', '3': 'Refinance',
                       '31': 'Cash-Out Refinance', '32': 'No Cash-out Refinance'}
        filter_parts.append(f"Loan Purpose: {purpose_map.get(loan_purpose, loan_purpose)}")
    if action_taken:
        action_map = {'1': 'Loan Originated'}
        filter_parts.append(f"Action Taken: {action_map.get(action_taken, action_taken)}")
    if occupancy_type:
        filter_parts.append(f"Occupancy Type: Owner-occupied")
    if total_units:
        filter_parts.append(f"Total Units: 1 unit, 2 units, 3 units, 4 units")
    if construction_method:
        filter_parts.append(f"Construction Method: Site-built, Manufactured Home")
    if not_reverse:
        filter_parts.append(f"Reverse Mortgage: Not Reverse Mortgage")

    ws.cell(row, 1, "HMDA Filters:")
    ws.cell(row, 2, "; ".join(filter_parts) if filter_parts else "All loans")
    row += 1

    ws.cell(row, 1, "Peer Definition:")
    ws.cell(row, 2, "Lenders making 50% to 200% of subject bank's application volume in a given year in a CBSA (for mortgage and SB data). For branch data, subject bank compared to all other banks in the CBSA in 2025.")
    row += 2

    # Calculations section
    ws.cell(row, 1, "Calculations and Formulas")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "All percentage calculations performed in BigQuery before export to Excel.")
    row += 1

    ws.cell(row, 1, "Mortgage Goals Sheet:")
    if single_bank_mode:
        ws.cell(row, 2, "Grand Total rows include data from all states where the bank has assessment areas.")
    else:
        ws.cell(row, 2, "Grand Total rows include data from all states where EITHER bank has assessment areas.")
    row += 1

    ws.cell(row, 1, "SB Goals Sheet:")
    if single_bank_mode:
        ws.cell(row, 2, "Grand Total rows include data from all states where the bank has assessment areas.")
    else:
        ws.cell(row, 2, "Grand Total rows include data from all states where EITHER bank has assessment areas.")
    row += 1

    ws.cell(row, 1, "Baseline Years for Goals Calculations:")
    ws.cell(row, 2, f"Baseline formulas use data from years {hmda_years_str} (2-year average).")
    row += 2

    # Small Business Methodology Section
    ws.cell(row, 1, "Small Business Lending Methodology")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "Data Source:")
    ws.cell(row, 2, "FFIEC CRA Small Business Loan Data (loans under $1 million to businesses with revenues under $1 million)")
    row += 1

    ws.cell(row, 1, "Goal Calculation Formula:")
    row += 1

    ws.cell(row, 1, "  1. Baseline:")
    ws.cell(row, 2, "Sum of loans/amounts from the baseline years (e.g., 2022-2023)")
    row += 1

    ws.cell(row, 1, "  2. Annualize:")
    ws.cell(row, 2, "Baseline / Number of years = Annual average")
    row += 1

    ws.cell(row, 1, "  3. Apply Improvement:")
    ws.cell(row, 2, "Annual average x (1 + Improvement %) = Target annual amount")
    row += 1

    ws.cell(row, 1, "  4. Calculate Goal:")
    ws.cell(row, 2, "Target annual amount x Agreement length (years) = Goal")
    row += 1

    ws.cell(row, 1, "Example Calculation:")
    ws.cell(row, 2, "If baseline = 1,000 loans over 2 years, improvement = 5%, agreement = 5 years:")
    row += 1

    ws.cell(row, 1, "")
    ws.cell(row, 2, "Annual avg = 1,000/2 = 500 loans/year")
    row += 1

    ws.cell(row, 1, "")
    ws.cell(row, 2, "With 5% increase = 500 x 1.05 = 525 loans/year")
    row += 1

    ws.cell(row, 1, "")
    ws.cell(row, 2, "5-Year Goal = 525 x 5 = 2,625 loans")
    row += 1

    ws.cell(row, 1, "SB Metrics Included:")
    row += 1

    ws.cell(row, 1, "  - Total SB Loans:")
    ws.cell(row, 2, "All small business loans originated in assessment areas")
    row += 1

    ws.cell(row, 1, "  - SB LMI Tract Loans:")
    ws.cell(row, 2, "Loans in low-to-moderate income census tracts (tract income <80% of area median)")
    row += 1

    ws.cell(row, 1, "  - SB Minority Tract Loans:")
    ws.cell(row, 2, "Loans in census tracts where minority population exceeds 50%")
    row += 1

    ws.cell(row, 1, "Important Caveats:")
    row += 1

    ws.cell(row, 1, "  - Only banks with assets > $1.564B are required to report CRA SB data")
    row += 1

    ws.cell(row, 1, "  - SB data is reported by Respondent ID (RSSD), which may differ from the HMDA LEI")
    row += 1

    if single_bank_mode:
        ws.cell(row, 1, "  - Goals are calculated for all assessment areas where the bank has operations")
    else:
        ws.cell(row, 1, "  - Goals are calculated for assessment areas where EITHER bank has operations")
    row += 2

    # SB LMICT Estimation Methodology
    ws.cell(row, 1, "Note on Small Business LMI Tract Lending Estimates")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row, 1, "Small business lending data, as reported to federal regulators, is provided at the county level in three loan size")
    row += 1
    ws.cell(row, 1, "buckets: under $100,000, $100,000-$250,000, and $250,000-$1,000,000. For each bucket, only the number of loans and the")
    row += 1
    ws.cell(row, 1, "total dollar amount are disclosed - individual loan amounts are not available. The number of loans made in low-to-moderate")
    row += 1
    ws.cell(row, 1, "income (LMI) census tracts is reported as a single count across all buckets, without a dollar amount breakdown.")
    row += 2

    ws.cell(row, 1, "To estimate LMI tract lending amounts, this report uses the following method:")
    row += 1
    ws.cell(row, 1, "  1. Calculate the LMI tract share: LMI Tract Loan Count / Total Loan Count")
    row += 1
    ws.cell(row, 1, "  2. Multiply: LMI Tract Share x Total Amount (all buckets) = Estimated LMI Tract Amount")
    row += 2

    ws.cell(row, 1, "Example: If a county has 1,000 total small business loans totaling $50 million, and 200 of those loans are in LMI tracts,")
    row += 1
    ws.cell(row, 1, "then: LMI Share = 200 / 1,000 = 20%, and Estimated LMI Amount = 20% x $50M = $10 million.")
    row += 2

    ws.cell(row, 1, "This calculation is performed per county, then summed to the CBSA level. This is an approximation. Because individual loan")
    row += 1
    ws.cell(row, 1, "amounts within each bucket are unknown (e.g., 5 loans in the $100,000-$250,000 bucket could each be $100,000 or each be")
    row += 1
    ws.cell(row, 1, "$250,000), and because LMI tract loans may not be evenly distributed across buckets, the estimated dollar amount may differ")
    row += 1
    ws.cell(row, 1, "from calculations that use a different order of operations or methodology.")
    row += 2

    ws.cell(row, 1, "Excel Format:")
    ws.cell(row, 2, "Null/missing values displayed as '--'. All calculations use raw decimal percentages.")
    row += 2

    # Race/Ethnicity Classification
    ws.cell(row, 1, "Race/Ethnicity Classification:")
    ws.cell(row, 2, "NCRC hierarchical methodology for HMDA data.")
    row += 2

    # Application Version
    from justdata.shared.utils.versions import get_version
    ws.cell(row, 1, "Application Version")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, "MergerMeter Version:")
    ws.cell(row, 2, get_version('mergermeter'))
    row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30


