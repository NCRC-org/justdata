"""Small business goals worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.shared.reporting.excel.helpers import (
    BORDER_THIN,
    HEADER_FILL,
    HEADER_FONT,
    NCRC_BLUE,
)

from justdata.shared.reporting.excel.computations import (
    _calculate_mortgage_cbsa_total,
    _calculate_mortgage_grand_total,
    _calculate_sb_cbsa_total,
    _calculate_sb_grand_total,
    _compute_branch_pct_diff,
    _compute_mortgage_pct_diff,
    _compute_sb_avg_diff,
    _compute_sb_ratio_diff,
    _get_cbsa_name,
    _write_mortgage_metric,
    _write_mortgage_metric_new_format,
    _write_sb_metric,
    _write_sb_metric_new_format,
)
from justdata.shared.reporting.excel.helpers import (
    _create_empty_mortgage_sheet,
    _create_empty_sb_sheet,
    _get_short_bank_name,
    _normalize_year_column,
)

logger = logging.getLogger(__name__)


def _create_sb_goals_sheet(wb: Workbook, sb_goals_data: pd.DataFrame, years_sb: Optional[List[int]]):
    """Create SB Goals sheet with state breakdowns and formulas."""
    ws = wb.create_sheet("SB Goals")

    # Headers
    headers = ['State', 'Metric', 'Base Value', 'NCRC Proposal', 'Baseline', 'Total Increase']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    # Row 2: Multipliers
    multipliers = [None, None, 1.2, 5, 5, None]
    for col_idx, val in enumerate(multipliers, 1):
        if val is not None:
            ws.cell(2, col_idx, val)

    row = 3

    if sb_goals_data is not None and not sb_goals_data.empty:
        # Metrics for SB Goals
        metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount',
                   'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']

        # Group by state
        if 'state_name' in sb_goals_data.columns:
            # Write Grand Total first
            for metric in metrics:
                ws.cell(row, 1, 'Grand Total')
                ws.cell(row, 2, metric)

                # Calculate grand total value based on metric
                if metric == 'SB Loans':
                    val = sb_goals_data['sb_loans_total'].sum() if 'sb_loans_total' in sb_goals_data.columns else 0
                elif metric == '#LMICT':
                    val = sb_goals_data['lmict_count'].sum() if 'lmict_count' in sb_goals_data.columns else 0
                elif metric == 'Avg SB LMICT Loan Amount':
                    lmict_count = sb_goals_data['lmict_count'].sum() if 'lmict_count' in sb_goals_data.columns else 0
                    lmict_amount = sb_goals_data['lmict_loans_amount'].sum() if 'lmict_loans_amount' in sb_goals_data.columns else 0
                    val = lmict_amount / lmict_count if lmict_count > 0 else 0
                elif metric == 'Loans Rev Under $1m':
                    val = sb_goals_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in sb_goals_data.columns else 0
                else:  # Avg Loan Amt for <$1M GAR SB
                    rev_count = sb_goals_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in sb_goals_data.columns else 0
                    rev_amount = sb_goals_data['amount_rev_under_1m'].sum() if 'amount_rev_under_1m' in sb_goals_data.columns else 0
                    val = rev_amount / rev_count if rev_count > 0 else 0

                ws.cell(row, 3, val)  # Base Value
                # NCRC Proposal (computed) - multipliers: C2=1.2, D2=5, E2=5
                ncrc_proposal = ((val / 2) * 1.2 * 5) if val else 0
                ws.cell(row, 4, ncrc_proposal)  # NCRC Proposal

                # Baseline and Total Increase only for avg loan amount metrics
                if 'Avg' in metric:
                    baseline = (val / 2) * 5 if val else 0
                    ws.cell(row, 5, baseline)
                    ws.cell(row, 6, ncrc_proposal - baseline)

                # Apply number formatting based on metric
                # Note: Only 'Avg' metrics are dollar amounts. "Loans Rev Under $1m" is a count
                # (the $1m refers to the revenue threshold, not the format)
                if 'Avg' in metric:
                    # Dollar format for average amounts only
                    for col in [3, 4, 5, 6]:
                        ws.cell(row, col).number_format = '$#,##0'
                else:
                    # Thousands format for loan counts
                    for col in [3, 4]:
                        ws.cell(row, col).number_format = '#,##0'

                row += 1

            # Write by state
            for state_name, state_data in sb_goals_data.groupby('state_name'):
                for metric in metrics:
                    ws.cell(row, 1, state_name)
                    ws.cell(row, 2, metric)

                    if metric == 'SB Loans':
                        val = state_data['sb_loans_total'].sum() if 'sb_loans_total' in state_data.columns else 0
                    elif metric == '#LMICT':
                        val = state_data['lmict_count'].sum() if 'lmict_count' in state_data.columns else 0
                    elif metric == 'Avg SB LMICT Loan Amount':
                        lmict_count = state_data['lmict_count'].sum() if 'lmict_count' in state_data.columns else 0
                        lmict_amount = state_data['lmict_loans_amount'].sum() if 'lmict_loans_amount' in state_data.columns else 0
                        val = lmict_amount / lmict_count if lmict_count > 0 else 0
                    elif metric == 'Loans Rev Under $1m':
                        val = state_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in state_data.columns else 0
                    else:
                        rev_count = state_data['loans_rev_under_1m_count'].sum() if 'loans_rev_under_1m_count' in state_data.columns else 0
                        rev_amount = state_data['amount_rev_under_1m'].sum() if 'amount_rev_under_1m' in state_data.columns else 0
                        val = rev_amount / rev_count if rev_count > 0 else 0

                    ws.cell(row, 3, val)
                    # NCRC Proposal (computed)
                    ncrc_proposal = ((val / 2) * 1.2 * 5) if val else 0
                    ws.cell(row, 4, ncrc_proposal)

                    if 'Avg' in metric:
                        baseline = (val / 2) * 5 if val else 0
                        ws.cell(row, 5, baseline)
                        ws.cell(row, 6, ncrc_proposal - baseline)

                    # Apply number formatting based on metric
                    # Note: Only 'Avg' metrics are dollar amounts. "Loans Rev Under $1m" is a count
                    if 'Avg' in metric:
                        # Dollar format for average amounts only
                        for col in [3, 4, 5, 6]:
                            ws.cell(row, col).number_format = '$#,##0'
                    else:
                        # Thousands format for loan counts
                        for col in [3, 4]:
                            ws.cell(row, col).number_format = '#,##0'

                    row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    logger.info(f"Created SB Goals sheet")


