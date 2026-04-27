"""Mortgage-report Excel workbook writer."""
import re
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import numpy as np

from datetime import datetime
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from justdata.apps.lendsight.report_builder.formatting import sanitize_sheet_name
from justdata.apps.lendsight.report_builder.excel_export.demographic import (
    create_demographic_overview_table_for_excel,
    create_population_demographics_table_for_excel,
)
from justdata.apps.lendsight.report_builder.excel_export.top_lenders import (
    create_top_lenders_table_for_excel,
    create_top_lenders_by_year_table_for_excel,
)

def save_mortgage_excel_report(report_data: Dict[str, pd.DataFrame], output_path: str, metadata: Dict[str, Any] = None):
    """
    Save the mortgage report data to an Excel file with multiple sheets.
    
    Sheets:
    1. Methods and Definitions
    2. Population Demographics (Census data across time periods)
    3. Section 1: Demographic Overview (with ALL categories and No Data)
    4. Section 2: Income and Neighborhood Indicators
    5. Section 3: Top Lenders (with ALL categories and No Data)
    6. Raw Data (all raw data used in the report)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Get raw data for creating Excel-specific tables
    raw_df = report_data.get('raw_data', pd.DataFrame())
    years = metadata.get('years', []) if metadata else []

    # Check if raw_data is available (it's not stored in cache due to size)
    if raw_df.empty:
        # For cached results, raw_data is not available
        # Create a minimal Excel with just the summary tables
        print("[WARNING] raw_data not available - creating Excel from summary tables only")
        raw_df = None  # Signal that we don't have raw data

    if not years:
        raise ValueError("Cannot create Excel report: missing years information")
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Get workbook and remove default sheet if it exists
        workbook = writer.book
        # Remove default "Sheet" if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Methods and Definitions sheet (first sheet)
        methods_sheet = workbook.create_sheet(sanitize_sheet_name('Methods and Definitions'), 0)
        
        # Prepare methods content
        methods_content = []
        
        methods_content.append(('METHODS AND DEFINITIONS', ''))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Sources', ''))
        methods_content.append(('', 'Home Mortgage Disclosure Act (HMDA) via FFIEC, accessed through NCRC\'s BigQuery data warehouse'))
        methods_content.append(('', 'U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates for demographic context'))
        methods_content.append(('', 'Census data retrieved via U.S. Census Bureau API (https://api.census.gov)'))
        if metadata and 'census_data' in metadata and metadata.get('census_data'):
            census_year = None
            for county_data in metadata.get('census_data', {}).values():
                if 'data_year' in county_data:
                    census_year = county_data['data_year']
                    break
            if census_year:
                methods_content.append(('', f'Census data year: {census_year}'))
        methods_content.append(('', ''))

        methods_content.append(('Report Configuration', ''))
        if metadata and metadata.get('counties'):
            counties_list = metadata['counties']
            if isinstance(counties_list, list):
                methods_content.append(('', f'Geography: {", ".join(str(c) for c in counties_list)}'))
            else:
                methods_content.append(('', f'Geography: {counties_list}'))
        if metadata and metadata.get('years'):
            report_years = metadata['years']
            if isinstance(report_years, list) and len(report_years) > 0:
                methods_content.append(('', f'Years Analyzed: {min(report_years)}-{max(report_years)}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Race and Ethnicity Classification', ''))
        methods_content.append(('', 'Race and ethnicity are determined using the COALESCE function across applicant_race_1 through applicant_race_5 and applicant_ethnicity_1 through applicant_ethnicity_5 fields.'))
        methods_content.append(('', 'Hispanic classification takes precedence (if any ethnicity field indicates Hispanic, the borrower is classified as Hispanic).'))
        methods_content.append(('', 'For non-Hispanic borrowers, the first valid race code (excluding codes 6, 7, 8 which indicate "Not Provided", "Not Applicable", or "Information not provided") is used from race_1 through race_5.'))
        methods_content.append(('', 'Percentages are calculated as: (group loans / loans with demographic data) × 100, where loans with demographic data = all loans minus loans lacking race/ethnicity data.'))
        methods_content.append(('', ''))
        
        methods_content.append(('Income and Neighborhood Indicators', ''))
        methods_content.append(('LMIB', 'Low to Moderate Income Borrower: Borrowers with income below 80% of the area median family income'))
        methods_content.append(('LMICT', 'Low to Moderate Income Census Tract: Census tracts where median family income is below 80% of the area median family income'))
        methods_content.append(('MMCT', 'Majority Minority Census Tract: Census tracts where minority populations represent more than 50% of total population'))
        methods_content.append(('', 'Percentages are calculated as: (category loans / total loans) × 100'))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Filters Applied', ''))
        methods_content.append(('', 'Originations only (action taken = 1)'))
        methods_content.append(('', 'Owner-occupied properties (occupancy type = 1)'))
        methods_content.append(('', 'Forward loans (excludes reverse mortgages)'))
        methods_content.append(('', 'Site-built properties (construction method = 1)'))
        methods_content.append(('', '1-4 unit properties (total_units IN (1, 2, 3, 4))'))
        if metadata and 'loan_purpose' in metadata:
            loan_purpose = metadata['loan_purpose']
            if isinstance(loan_purpose, list):
                if len(loan_purpose) == 0 or set(loan_purpose) == {'purchase', 'refinance', 'equity'} or 'all' in loan_purpose:
                    methods_content.append(('', 'Loan Purpose: All loan purposes'))
                else:
                    purpose_map = {
                        'purchase': 'Home Purchase (loan purpose = 1)',
                        'refinance': 'Refinance & Cash-Out (loan purpose IN (31, 32))',
                        'equity': 'Home Equity (loan purpose IN (2, 4))'
                    }
                    purposes = [purpose_map.get(p, p) for p in loan_purpose if p != 'all']
                    methods_content.append(('', f'Loan Purpose: {", ".join(purposes)}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Abbreviations', ''))
        methods_content.append(('HMDA', 'Home Mortgage Disclosure Act'))
        methods_content.append(('CFPB', 'Consumer Financial Protection Bureau'))
        methods_content.append(('NCRC', 'National Community Reinvestment Coalition'))
        methods_content.append(('LMI', 'Low-to-Moderate Income'))
        methods_content.append(('LMIB', 'Low-to-Moderate Income Borrower'))
        methods_content.append(('LMICT', 'Low-to-Moderate Income Census Tract'))
        methods_content.append(('MMCT', 'Majority-Minority Census Tract'))
        methods_content.append(('AMFI', 'Area Median Family Income'))
        methods_content.append(('pp', 'Percentage Points'))
        methods_content.append(('', ''))

        methods_content.append(('Disclaimer', ''))
        methods_content.append(('', 'This report was generated by NCRC\'s JustData LendSight application using publicly available HMDA data.'))
        methods_content.append(('', 'Data accuracy is dependent on lender reporting to the FFIEC.'))
        methods_content.append(('', 'This analysis is provided for informational purposes and does not constitute legal or financial advice.'))
        methods_content.append(('', ''))

        methods_content.append(('', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'))
        from justdata.apps.lendsight.version import __version__ as lendsight_version
        methods_content.append(('', f'Platform: JustData LendSight v{lendsight_version}'))
        
        # Write methods content
        for row_idx, (label, value) in enumerate(methods_content, start=1):
            methods_sheet.cell(row=row_idx, column=1, value=label)
            methods_sheet.cell(row=row_idx, column=2, value=value)
            if label and label != '':
                methods_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Population Demographics table (from Census data)
        census_data_for_excel = metadata.get('census_data', {}) if metadata else {}
        population_demo_table = create_population_demographics_table_for_excel(census_data=census_data_for_excel)
        if not population_demo_table.empty:
            population_demo_table.to_excel(writer, sheet_name=sanitize_sheet_name('Population Demographics'), index=False)
        
        # Section 1: Demographic Overview (with ALL categories and No Data)
        # Pass census_data if available for population share column
        if raw_df is not None:
            demo_table = create_demographic_overview_table_for_excel(raw_df, years, census_data=census_data_for_excel)
            if not demo_table.empty:
                demo_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 1 - Demographic Overview'), index=False)
        else:
            # Use pre-computed demographic_overview table from report_data
            demographic_overview = report_data.get('demographic_overview', pd.DataFrame())
            if isinstance(demographic_overview, list):
                demographic_overview = pd.DataFrame(demographic_overview)
            if not demographic_overview.empty:
                demographic_overview.to_excel(writer, sheet_name=sanitize_sheet_name('Section 1 - Demographic Overview'), index=False)

        # Section 2: Income and Neighborhood Indicators (3 tables)
        # Table 1: Lending to Income Borrowers
        income_borrowers = report_data.get('income_borrowers', pd.DataFrame())
        if isinstance(income_borrowers, list):
            income_borrowers = pd.DataFrame(income_borrowers)
        if not income_borrowers.empty:
            income_borrowers.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2a - Income Borrowers'), index=False)

        # Table 2: Lending to Census Tracts by Income
        income_tracts = report_data.get('income_tracts', pd.DataFrame())
        if isinstance(income_tracts, list):
            income_tracts = pd.DataFrame(income_tracts)
        if not income_tracts.empty:
            income_tracts.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2b - Income Tracts'), index=False)

        # Table 3: Lending to Census Tracts by Minority Population
        minority_tracts = report_data.get('minority_tracts', pd.DataFrame())
        if isinstance(minority_tracts, list):
            minority_tracts = pd.DataFrame(minority_tracts)
        if not minority_tracts.empty:
            minority_tracts.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2c - Minority Tracts'), index=False)
        
        # Section 3: Top Lenders (with ALL categories and No Data)
        # This matches the report table exactly but includes ALL categories
        if raw_df is not None:
            top_lenders_table = create_top_lenders_table_for_excel(raw_df, years)
            if not top_lenders_table.empty:
                top_lenders_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 3 - Top Lenders'), index=False)

            # Section 3b: Top 10 Lenders Over Time (2020-2024) - Excel Export Only
            export_years = [y for y in years if 2020 <= y <= 2024]
            if export_years:
                top_lenders_by_year_table = create_top_lenders_by_year_table_for_excel(raw_df, export_years, top_n=10)
                if not top_lenders_by_year_table.empty:
                    top_lenders_by_year_table.to_excel(writer, sheet_name=sanitize_sheet_name('Top 10 Lenders Over Time'), index=False)

            # Raw Data sheet - all raw data used in the report
            if not raw_df.empty:
                raw_export_df = raw_df.copy()
                raw_export_df.to_excel(writer, sheet_name=sanitize_sheet_name('Raw Data'), index=False)
        else:
            # Use pre-computed top_lenders_detailed table from report_data if available
            top_lenders_detailed = report_data.get('top_lenders_detailed', pd.DataFrame())
            if isinstance(top_lenders_detailed, list):
                top_lenders_detailed = pd.DataFrame(top_lenders_detailed)
            if not top_lenders_detailed.empty:
                top_lenders_detailed.to_excel(writer, sheet_name=sanitize_sheet_name('Section 3 - Top Lenders'), index=False)
        
        # Style the header row for all data sheets (not Methods sheet)
        from openpyxl.styles import NamedStyle
        from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE
        
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Number format styles
        integer_format = '#,##0'  # Format: 1,234
        percentage_format = '#,##0.00'  # Format: 12.34
        
        for sheet_name in writer.sheets:
            if sheet_name != 'Methods and Definitions':
                worksheet = writer.sheets[sheet_name]
                if worksheet.max_row > 0:
                    # Style header row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply number formatting to data cells
                    # Get header row to identify column types
                    header_row = worksheet[1]
                    headers = [cell.value for cell in header_row]
                    
                    # Iterate through data rows (skip header row)
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col_idx, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            
                            # Skip if cell is empty or not a number
                            if cell.value is None:
                                continue
                            
                            # Determine format based on header name and value type
                            header_str = str(header) if header else ''
                            
                            # Get the metric value from first column to determine row type
                            metric_cell = worksheet.cell(row=row_idx, column=1)
                            metric_value = metric_cell.value if metric_cell else ''
                            metric_str = str(metric_value) if metric_value else ''
                            
                            # Special handling for Population Demographics sheet
                            if sheet_name == 'Population Demographics':
                                if isinstance(cell.value, (int, float)):
                                    # Total Population row: format as integer
                                    if 'Total Population' in metric_str:
                                        cell.number_format = integer_format
                                    # All other rows (demographic percentages): format as percentage
                                    elif '(%)' in metric_str:
                                        cell.number_format = percentage_format
                            # Check if it's a percentage column
                            elif '%' in header_str or 'percent' in header_str.lower() or 'Population Share' in header_str:
                                # Format as percentage with 2 decimal places
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = percentage_format
                            # Check if it's a year column - format based on row type
                            elif header_str.isdigit() or 'year' in header_str.lower():
                                # Year columns - format based on metric row
                                if isinstance(cell.value, (int, float)):
                                    # If row is "Total Loans" or "Loans", format as integer
                                    if 'Total Loans' in metric_str or metric_str == 'Loans':
                                        cell.number_format = integer_format
                                    else:
                                        # All other rows in year columns are percentages
                                        cell.number_format = percentage_format
                            # Check if it's "Change Over Time" - could be integer or percentage
                            elif 'Change Over Time' in header_str:
                                # Use the metric_str we already determined above
                                # If metric is "Total Loans" or "Loans", it's a count (integer)
                                # Otherwise it's a percentage change
                                if 'Total Loans' in metric_str or metric_str == 'Loans':
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = integer_format
                                else:
                                    # Percentage point change
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = percentage_format
                            # Check if it's "Net Change" - usually integer
                            elif 'Net Change' in header_str:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = integer_format
                            # Check if it's a metric column (first column) - no formatting
                            elif col_idx == 1:
                                pass
                            # Default: if it's a number and not a percentage, format as integer
                            elif isinstance(cell.value, (int, float)):
                                # Check if value looks like a percentage (between 0 and 100, likely decimal)
                                if 0 <= cell.value <= 100 and isinstance(cell.value, float) and cell.value != int(cell.value):
                                    # Likely a percentage value stored as number
                                    cell.number_format = percentage_format
                                else:
                                    # Integer value
                                    cell.number_format = integer_format
        
        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

