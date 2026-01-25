"""
Mortgage report builder module for creating Excel reports from HMDA mortgage data.
Similar structure to branch report builder but for mortgage origination data.
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import os
import re
import requests


def build_mortgage_report(raw_data: List[Dict[str, Any]], counties: List[str], years: List[int], census_data: Dict = None, hud_data: Dict[str, Dict[str, float]] = None, progress_tracker=None) -> Dict[str, pd.DataFrame]:
    """
    Process raw BigQuery HMDA data and build comprehensive mortgage report dataframes.
    
    Args:
        raw_data: List of dictionaries from BigQuery results
        counties: List of counties in the report
        years: List of years in the report
        census_data: Optional dictionary of Census demographic data by county
        hud_data: Optional dictionary mapping GEOID5 to HUD income distribution data
        progress_tracker: Optional progress tracker for real-time progress updates
        
    Returns:
        Dictionary containing multiple dataframes for different report sections
    """
    if not raw_data:
        raise ValueError("No data provided for report building")
    
    # Convert to DataFrame
    df = pd.DataFrame(raw_data)
    
    # Ensure required columns exist (mortgage-specific)
    required_columns = ['lei', 'year', 'county_code', 'county_state', 'total_originations', 'lmict_originations', 'mmct_originations']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    
    # Clean and prepare data
    df = clean_mortgage_data(df)
    
    # Calculate HHI for loan amounts in latest year (similar to deposits for branches)
    hhi_data = calculate_mortgage_hhi(df)
    
    # Build different report sections (pass census_data to demographic overview)
    
    report_sections = [
        ('summary', 'Summary Table', lambda: create_mortgage_summary_table(df, counties, years)),
        ('demographic_overview', 'Demographic Overview', lambda: create_demographic_overview_table(df, years, census_data=census_data)),
        ('income_borrowers', 'Income Borrowers', lambda: create_income_borrowers_table(df, years, hud_data=hud_data)),
        ('income_tracts', 'Income Tracts', lambda: create_income_tracts_table(df, years, hud_data=hud_data, census_data=census_data)),
        ('minority_tracts', 'Minority Tracts', lambda: create_minority_tracts_table(df, years, census_data=census_data)),
        # REMOVED: income_neighborhood_tracts and income_neighborhood_indicators - consolidated into the three tables above
        ('top_lenders_detailed', 'Top Lenders Detailed', lambda: create_top_lenders_detailed_table(df, years)),
        ('market_concentration', 'Market Concentration', lambda: create_market_concentration_table(df, years, metadata={'counties': counties, 'years': years})),
        ('by_lender', 'Lender Summary', lambda: create_lender_summary(df, years)),
        ('by_county', 'County Summary', lambda: create_mortgage_county_summary(df, counties, years)),
        ('trends', 'Trends Analysis', lambda: create_mortgage_trend_analysis(df, years)),
    ]
    
    total_sections = len(report_sections)
    report_data = {}
    
    for idx, (key, section_name, create_func) in enumerate(report_sections, 1):
        if progress_tracker:
            progress_tracker.update_section_progress(idx, total_sections, section_name)
        report_data[key] = create_func()
    
    # Add non-section data
    report_data['raw_data'] = df
    report_data['hhi'] = hhi_data
    
    return report_data


def calculate_mortgage_hhi(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calculate Herfindahl-Hirschman Index (HHI) for mortgage loan amounts in the latest year.
    
    Similar to branch HHI but uses loan amounts instead of deposits.
    
    Returns:
        Dictionary with HHI value, concentration level, year, and top lenders by loan volume
    """
    if 'total_loan_amount' not in df.columns:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': None,
            'total_loan_amount': None,
            'top_lenders': []
        }
    
    # Find latest year
    latest_year = df['year'].max()
    latest_year_df = df[df['year'] == latest_year].copy()
    
    # Aggregate loan amounts by lender (across all counties)
    lender_loans = latest_year_df.groupby('lender_name')['total_loan_amount'].sum().reset_index()
    lender_loans = lender_loans[lender_loans['total_loan_amount'] > 0]  # Remove lenders with zero loans
    
    if lender_loans.empty:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': latest_year,
            'total_loan_amount': 0,
            'top_lenders': []
        }
    
    total_loan_amount = lender_loans['total_loan_amount'].sum()
    
    # Calculate market shares (as percentages)
    lender_loans['market_share'] = (lender_loans['total_loan_amount'] / total_loan_amount) * 100
    
    # Calculate HHI: sum of squared market shares (0-10,000 scale)
    hhi = (lender_loans['market_share'] ** 2).sum()
    
    # Determine concentration level
    if hhi < 1500:
        concentration_level = 'Low concentration (competitive market)'
    elif hhi < 2500:
        concentration_level = 'Moderate concentration'
    else:
        concentration_level = 'High concentration'
    
    # Get top lenders by loan amount
    top_lenders = lender_loans.nlargest(10, 'total_loan_amount').copy()
    top_lenders_list = []
    for _, row in top_lenders.iterrows():
        top_lenders_list.append({
            'lender_name': row['lender_name'],
            'total_loan_amount': float(row['total_loan_amount']),
            'market_share': float(row['market_share'])
        })
    
    return {
        'hhi': float(hhi),
        'concentration_level': concentration_level,
        'year': int(latest_year),
        'total_loan_amount': float(total_loan_amount),
        'total_lenders': len(lender_loans),
        'top_lenders': top_lenders_list
    }


def create_market_concentration_table(df: pd.DataFrame, years: List[int], metadata: Dict = None) -> List[Dict[str, Any]]:
    """
    Create market concentration table showing HHI (Herfindahl-Hirschman Index) by year and loan purpose.
    
    This function calculates HHI for:
    - All Loans (combined)
    - Home purchase loans (loan_purpose = '1')
    - Refinance loans (loan_purpose IN ('31', '32'))
    - Home equity loans (loan_purpose IN ('2', '4'))
    
    Args:
        df: DataFrame with mortgage data (must have 'year', 'lender_name', 'total_loan_amount', 'loan_purpose')
        years: List of years to calculate HHI for
        metadata: Optional metadata dict
    
    Returns:
        List of dictionaries with HHI data by year and loan purpose category
    """
    # Define loan purpose filters
    loan_purpose_filters = {
        'All Loans': None,  # No filter, use all data
        'Home Purchase': ['1'],
        'Refinance': ['31', '32'],
        'Home Equity': ['2', '4']
    }
    
    hhi_results = []
    
    # Check if loan_purpose column exists
    has_loan_purpose = 'loan_purpose' in df.columns
    
    for purpose_label, purpose_codes in loan_purpose_filters.items():
        purpose_hhi_data = {'Loan Purpose': purpose_label}
        
        # Filter DataFrame by loan purpose if codes are specified
        if purpose_codes is None:
            # All Loans - use full DataFrame
            filtered_df = df.copy()
        elif has_loan_purpose:
            # Filter by loan purpose codes
            filtered_df = df[df['loan_purpose'].isin(purpose_codes)].copy()
        else:
            # loan_purpose column not available - can only calculate for All Loans
            if purpose_label == 'All Loans':
                filtered_df = df.copy()
            else:
                # Cannot calculate for specific loan purposes without loan_purpose column
                # Fill with None for all years
                for year in sorted(years):
                    purpose_hhi_data[year] = None
                hhi_results.append(purpose_hhi_data)
                continue
        
        # Calculate HHI for each year
        for year in sorted(years):
            year_df = filtered_df[filtered_df['year'] == year].copy()
            hhi_value = calculate_mortgage_hhi_for_year(year_df, year)
            purpose_hhi_data[year] = hhi_value['hhi'] if hhi_value['hhi'] is not None else None
        
        hhi_results.append(purpose_hhi_data)
    
    return hhi_results


def calculate_mortgage_hhi_for_year(df: pd.DataFrame, year: int) -> Dict[str, Any]:
    """
    Calculate Herfindahl-Hirschman Index (HHI) for mortgage loan amounts for a specific year.
    
    Args:
        df: DataFrame with mortgage data for the year (must have 'lender_name', 'total_loan_amount')
        year: Year being calculated (for metadata)
    
    Returns:
        Dictionary with HHI value, concentration level, year, and metadata
    """
    if 'total_loan_amount' not in df.columns or 'lender_name' not in df.columns:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': year,
            'total_loan_amount': None,
            'top_lenders': []
        }
    
    # Aggregate loan amounts by lender (across all counties)
    lender_loans = df.groupby('lender_name')['total_loan_amount'].sum().reset_index()
    lender_loans = lender_loans[lender_loans['total_loan_amount'] > 0]  # Remove lenders with zero loans
    
    if lender_loans.empty:
        return {
            'hhi': None,
            'concentration_level': 'Not Available',
            'year': year,
            'total_loan_amount': 0,
            'top_lenders': []
        }
    
    total_loan_amount = lender_loans['total_loan_amount'].sum()
    
    # Calculate market shares (as percentages)
    lender_loans['market_share'] = (lender_loans['total_loan_amount'] / total_loan_amount) * 100
    
    # Calculate HHI: sum of squared market shares (0-10,000 scale)
    hhi = (lender_loans['market_share'] ** 2).sum()
    
    # Determine concentration level
    if hhi < 1500:
        concentration_level = 'Low concentration (competitive market)'
    elif hhi < 2500:
        concentration_level = 'Moderate concentration'
    else:
        concentration_level = 'High concentration'
    
    return {
        'hhi': float(hhi),
        'concentration_level': concentration_level,
        'year': int(year),
        'total_loan_amount': float(total_loan_amount),
        'total_lenders': len(lender_loans)
    }


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize Excel sheet name by removing invalid characters.
    
    Excel sheet names cannot contain: : \ / ? * [ ]
    Also truncate to max_length (Excel limit is 31 characters).
    
    Args:
        name: Original sheet name
        max_length: Maximum length (default 31, Excel's limit)
    
    Returns:
        Sanitized sheet name
    """
    if not name:
        return 'Sheet'
    
    # Replace invalid characters with dash
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '-')
    
    # Truncate to max length
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length]
    
    return sanitized


def abbreviate_long_lender_name(name: str, max_length: int = 30) -> str:
    """
    Abbreviate lender names that exceed max_length characters.
    
    Applies common word abbreviations first, then truncates if still too long.
    Preserves important parts of the name (usually the beginning).
    
    Args:
        name: Lender name to abbreviate
        max_length: Maximum allowed length (default 30)
    
    Returns:
        Abbreviated name if original was too long, otherwise original name
    """
    if not name or pd.isna(name):
        return name
    
    name = str(name).strip()
    
    # If already short enough, return as-is
    if len(name) <= max_length:
        return name
    
    # Common word abbreviations (case-insensitive)
    # Format: (full_word, abbreviation)
    abbreviations = [
        ('MORTGAGE', 'MTG'),
        ('GROUP', 'GRP'),
        ('CORPORATION', 'CORP'),
        ('COMPANY', 'CO'),
        ('INCORPORATED', 'INC'),
        ('ASSOCIATES', 'ASSOC'),
        ('ASSOCIATION', 'ASSOC'),
        ('FINANCIAL', 'FINL'),
        ('SERVICES', 'SVCS'),
        ('SERVICE', 'SVC'),
        ('BANKING', 'BKG'),
        ('CREDIT', 'CR'),
        ('FEDERAL', 'FED'),
        ('NATIONAL', 'NATL'),
        ('INTERNATIONAL', 'INTL'),
        ('AMERICA', 'AMER'),
        ('AMERICAN', 'AMER'),
        ('MANAGEMENT', 'MGMT'),
        ('INVESTMENT', 'INV'),
        ('INVESTMENTS', 'INV'),
        ('HOLDINGS', 'HLDGS'),
        ('HOLDING', 'HLDG'),
        ('ENTERPRISES', 'ENT'),
        ('ENTERPRISE', 'ENT'),
    ]
    
    # Apply abbreviations (case-insensitive)
    abbreviated = name
    for full_word, abbrev in abbreviations:
        # Use word boundaries to avoid partial matches
        pattern = re.compile(r'\b' + re.escape(full_word) + r'\b', re.IGNORECASE)
        abbreviated = pattern.sub(abbrev, abbreviated)
    
    # If abbreviations helped, check length again
    if len(abbreviated) <= max_length:
        return abbreviated
    
    # Still too long - intelligently truncate
    # Strategy: Keep the beginning, remove middle words if needed, keep important ending words
    words = abbreviated.split()
    
    # If we have words, try to keep first part and last word
    if len(words) > 1:
        # Keep first word(s) and last word, remove middle words
        first_word = words[0]
        last_word = words[-1]
        
        # Try: First word + Last word
        candidate = f"{first_word} {last_word}"
        if len(candidate) <= max_length:
            return candidate
        
        # Try: First word only (if it's not too long)
        if len(first_word) <= max_length:
            return first_word
        
        # Last resort: Truncate first word
        if len(first_word) > max_length - 4:
            return first_word[:max_length - 3] + "..."
    
    # Fallback: Simple truncation with ellipsis
    if len(abbreviated) > max_length:
        return abbreviated[:max_length - 3] + "..."
    
    return abbreviated


def correct_lender_name_capitalization(name: str) -> str:
    """
    Correct capitalization for common lender names to match their official branding.
    
    This function preserves the exact capitalization from the database but can apply
    corrections for well-known lenders whose names may be incorrectly capitalized in the source data.
    """
    if not name or pd.isna(name):
        return name
    
    name = str(name).strip()
    
    # Dictionary of common lender name corrections
    # Key is case-insensitive, value is the correct capitalization
    lender_corrections = {
        'jpmorgan chase': 'JPMorgan Chase',
        'jp morgan chase': 'JPMorgan Chase',
        'j.p. morgan chase': 'JPMorgan Chase',
        'j.p.morgan chase': 'JPMorgan Chase',
        'bank of america': 'Bank of America',
        'wells fargo': 'Wells Fargo',
        'citibank': 'Citibank',
        'us bank': 'U.S. Bank',
        'usbank': 'U.S. Bank',
        'pnc bank': 'PNC Bank',
        'truist': 'Truist',
        'capital one': 'Capital One',
        'first national bank': 'First National Bank',
        'huntington bank': 'Huntington Bank',
        'keybank': 'KeyBank',
        'regions bank': 'Regions Bank',
        'td bank': 'TD Bank',
        'suntrust': 'SunTrust',
        'bb&t': 'BB&T',
        'bbt': 'BB&T',
    }
    
    # Check for exact match (case-insensitive)
    name_lower = name.lower()
    if name_lower in lender_corrections:
        return lender_corrections[name_lower]
    
    # Check for partial matches (e.g., "JPMorgan Chase Bank" should match "JPMorgan Chase")
    for incorrect, correct in lender_corrections.items():
        if name_lower.startswith(incorrect) or incorrect in name_lower:
            # Replace the incorrect portion with the correct capitalization
            # Use case-insensitive replacement
            pattern = re.compile(re.escape(incorrect), re.IGNORECASE)
            corrected = pattern.sub(correct, name)
            # If the replacement changed something, return it
            if corrected != name:
                return corrected
    
    # If no correction found, return original name (preserving database capitalization)
    return name


def clean_mortgage_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and prepare mortgage data for analysis.
    
    Similar to branch data cleaning but for mortgage-specific fields.
    """
    # Make a copy to avoid modifying original
    df = df.copy()
    
    # Ensure numeric columns are properly typed
    numeric_columns = ['total_originations', 'lmib_originations', 'lmict_originations', 
                      'mmct_originations', 'total_loan_amount', 'avg_loan_amount', 'avg_income']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Ensure year is integer
    if 'year' in df.columns:
        df['year'] = pd.to_numeric(df['year'], errors='coerce').astype('Int64')
    
    # Fill missing lender names, convert to uppercase, and abbreviate long names
    if 'lender_name' in df.columns:
        df['lender_name'] = df['lender_name'].fillna('Unknown Lender')
        # Convert all lender names to uppercase for display
        df['lender_name'] = df['lender_name'].apply(lambda x: str(x).upper() if x else x)
        # Abbreviate names longer than 30 characters
        df['lender_name'] = df['lender_name'].apply(lambda x: abbreviate_long_lender_name(x, max_length=30) if x else x)
    
    return df


def create_demographic_overview_table(df: pd.DataFrame, years: List[int], census_data: Dict = None) -> pd.DataFrame:
    """
    Create demographic overview table showing lending activity by race/ethnicity.
    
    Shows number of loans and percent for each race/ethnic group over time.
    Only includes groups that are >= 1% of all loans.
    Denominator = loans with demographic data (total loans minus loans without race/ethnicity data).
    
    Table structure:
    - Far left: Metric column (race/ethnicity group)
    - Middle: Column for each year showing "Number (Percent%)"
    - Far right: Change column showing increase/decrease over entire span
    """
    print(f"  [DEBUG] create_demographic_overview_table called with census_data: {census_data is not None}")
    if census_data:
        print(f"  [DEBUG] census_data has {len(census_data)} counties: {list(census_data.keys())}")
    
    # Check if we have the race/ethnicity columns
    race_columns = ['hispanic_originations', 'black_originations', 'white_originations', 
                   'asian_originations', 'native_american_originations', 'hopi_originations',
                   'multi_racial_originations']
    
    if not all(col in df.columns for col in race_columns):
        # Return empty DataFrame if we don't have the required columns
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Calculate total loans with demographic data
        # This is the sum of all race/ethnicity originations (may have overlap)
        # We need to calculate loans WITH demographic data differently
        # For now, use total_originations and subtract those without data
        # But we need to track which loans have demographic data
        
        # Sum originations by race/ethnicity
        hispanic = int(year_df['hispanic_originations'].sum())
        black = int(year_df['black_originations'].sum())
        white = int(year_df['white_originations'].sum())
        asian = int(year_df['asian_originations'].sum())
        native_american = int(year_df['native_american_originations'].sum())
        hopi = int(year_df['hopi_originations'].sum())
        multi_racial = int(year_df['multi_racial_originations'].sum()) if 'multi_racial_originations' in year_df.columns else 0
        
        # Total originations
        total_originations = int(year_df['total_originations'].sum())
        
        # Loans with demographic data (from SQL query if available)
        if 'loans_with_demographic_data' in year_df.columns:
            loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
        else:
            # Fallback: estimate as sum of all race/ethnicity originations
            # Note: This may overcount if loans are counted in multiple categories
            loans_with_demographics = hispanic + black + white + asian + native_american + hopi
        
        # More accurate: loans with demographics = total - loans without any classification
        # Since we're using COALESCE methodology, if a loan has no valid race/ethnicity,
        # it won't be counted in any of the race columns
        # So loans_with_demographics should be the sum of unique loans across all categories
        # For simplicity, we'll use the maximum of the individual counts as a proxy
        # But actually, we should track this in the SQL query
        
        # For now, use total_originations as denominator and calculate percentages
        # The SQL query should provide a has_demographic_data flag, but if not,
        # we'll use the sum of race originations as proxy
        
        yearly_data.append({
            'year': year,
            'total_originations': total_originations,
            'hispanic': hispanic,
            'black': black,
            'white': white,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'multi_racial': multi_racial,
            'loans_with_demographics': loans_with_demographics
        })
    
    # Build table structure
    # First, determine which groups are >= 1% across all years
    # Use total_originations (all applications) as denominator for threshold calculation
    all_totals = [d['total_originations'] for d in yearly_data]
    max_total = max(all_totals) if all_totals else 0
    
    # Define the standard order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
    standard_order = ['white', 'hispanic', 'black', 'asian', 'native_american', 'hopi', 'multi_racial']
    
    # Calculate max percentage for each group across all years
    group_max_pct = {}
    group_total_share = {}  # Total share across all years for sorting
    for group in standard_order:
        max_count = max([d[group] for d in yearly_data]) if yearly_data else 0
        # Use the year with max total_originations as denominator for threshold check
        max_pct = (max_count / max_total * 100) if max_total > 0 else 0
        group_max_pct[group] = max_pct
        
        # Calculate total share (average across all years) for sorting
        # Use loans_with_demographics for percentage calculation, but total_originations for threshold
        total_count = sum([d[group] for d in yearly_data]) if yearly_data else 0
        total_loans_all_years = sum([d['loans_with_demographics'] for d in yearly_data]) if yearly_data else 0
        avg_share = (total_count / total_loans_all_years * 100) if total_loans_all_years > 0 else 0
        group_total_share[group] = avg_share
    
    # Filter groups >= 1% and sort by total share (descending), then maintain standard order for ties
    included_groups = [group for group in standard_order if group_max_pct.get(group, 0) >= 1.0]
    # Sort by total share descending, but maintain standard order for groups with same share
    included_groups.sort(key=lambda g: (-group_total_share.get(g, 0), standard_order.index(g)))
    
    if not included_groups:
        return pd.DataFrame()
    
    # Get population shares from Census data (if available)
    # Use most recent census data: ACS if available, otherwise 2020 Census
    # For multiple counties, aggregate by weighting by total population
    population_shares = {}
    if census_data:
        total_population = 0
        group_totals = {
            'hispanic': 0,
            'black': 0,
            'white': 0,
            'asian': 0,
            'native_american': 0,
            'hopi': 0,
            'multi_racial': 0
        }
        
        # Aggregate across all counties by calculating weighted averages
        for county, county_census in census_data.items():
            demographics = None
            
            # Check for new structure (time_periods) - prefer ACS, fallback to 2020 Census
            if 'time_periods' in county_census:
                time_periods = county_census.get('time_periods', {})
                # Prefer ACS (most recent), fallback to 2020 Census
                if 'acs' in time_periods and time_periods['acs'].get('demographics'):
                    demographics = time_periods['acs'].get('demographics', {})
                    print(f"  [DEBUG] Using ACS data for {county}")
                elif 'census2020' in time_periods and time_periods['census2020'].get('demographics'):
                    demographics = time_periods['census2020'].get('demographics', {})
                    print(f"  [DEBUG] Using 2020 Census data for {county}")
            # Fallback to old structure (demographics at top level)
            elif 'demographics' in county_census:
                demographics = county_census.get('demographics', {})
                print(f"  [DEBUG] Using legacy demographics structure for {county}")
            
            if demographics:
                county_pop = demographics.get('total_population', 0)
                if county_pop and county_pop > 0:
                    total_population += county_pop
                    # Calculate actual counts from percentages
                    group_totals['hispanic'] += (demographics.get('hispanic_percentage', 0) / 100) * county_pop
                    group_totals['black'] += (demographics.get('black_percentage', 0) / 100) * county_pop
                    group_totals['white'] += (demographics.get('white_percentage', 0) / 100) * county_pop
                    group_totals['asian'] += (demographics.get('asian_percentage', 0) / 100) * county_pop
                    group_totals['native_american'] += (demographics.get('native_american_percentage', 0) / 100) * county_pop
                    group_totals['hopi'] += (demographics.get('hopi_percentage', 0) / 100) * county_pop
                    group_totals['multi_racial'] += (demographics.get('multi_racial_percentage', 0) / 100) * county_pop
        
        # Calculate weighted average percentages
        if total_population > 0:
            for group in group_totals:
                population_shares[group] = (group_totals[group] / total_population * 100)
            print(f"  [DEBUG] Calculated population_shares: {population_shares}")
        else:
            print(f"  [WARNING] Total population is 0, cannot calculate population shares")
    else:
        print(f"  [DEBUG] No census_data provided, skipping population shares")
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add Population Share column (if Census data available)
    if population_shares:
        print(f"  [DEBUG] Adding Population Share (%) column to table")
        result_data['Population Share (%)'] = []
    else:
        print(f"  [DEBUG] No population_shares, skipping Population Share column")
    
    # Add change column
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Group labels
    group_labels = {
        'hispanic': 'Hispanic',
        'black': 'Black',
        'white': 'White',
        'asian': 'Asian',
        'native_american': 'Native American',
        'hopi': 'Hawaiian/Pacific Islander',
        'multi_racial': 'Multi-Racial'
    }
    
    # Add Total Loans row at the top (showing just the integer, no percentage)
    result_data['Metric'].append('Total Loans')
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total = year_data['total_originations']
            result_data[str(year)].append(f"{total:,}")
        else:
            result_data[str(year)].append("0")
    # Population Share column (empty for Total Loans row)
    if population_shares:
        result_data['Population Share (%)'].append("")
    # Change column for Total Loans shows percentage change from first to last year
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            first_total = first_year_data['total_originations']
            last_total = last_year_data['total_originations']
            if first_total > 0:
                pct_change = ((last_total - first_total) / first_total) * 100
                if pct_change > 0:
                    change_str = f"+{pct_change:.1f}%"
                elif pct_change < 0:
                    change_str = f"{pct_change:.1f}%"
                else:
                    change_str = "0.0%"
            else:
                change_str = "N/A"
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            result_data[f'Change Over Time ({time_span})'].append("N/A")
    else:
        if len(years) >= 2:
            result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    # Calculate for each included group (show only percentages, no raw numbers)
    for group in included_groups:
        result_data['Metric'].append(group_labels[group])
        
        # Calculate for each year
        group_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data.get(group, 0)
                # Denominator: loans with demographic data (total loans minus loans without race/ethnicity data)
                total_loans = year_data['total_originations']
                loans_with_demo = year_data['loans_with_demographics']
                denominator = loans_with_demo if loans_with_demo > 0 else total_loans
                pct = (count / denominator * 100) if denominator > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                group_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                group_changes.append((0, 0.0))
        
        # Add Population Share column (if Census data available)
        if population_shares:
            pop_share = population_shares.get(group, 0)
            result_data['Population Share (%)'].append(f"{pop_share:.1f}%")
        
        # Calculate change from first to last year (show only percentage point change)
        if len(years) >= 2 and len(group_changes) >= 2:
            first_pct = group_changes[0][1]
            last_pct = group_changes[-1][1]
            pct_change = last_pct - first_pct
            
            # Format change: show only percentage point change
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    result = pd.DataFrame(result_data)
    return result


def create_income_neighborhood_indicators_table(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create income and neighborhood indicators table.
    
    Shows:
    - Number of loans overall (total originations)
    - Low to Moderate Income Census Tract (LMICT) originations
    - Low to Moderate Income Borrower (LMIB) originations
    - Majority Minority Census Tract (MMCT) originations
    
    Table structure:
    - Far left: Metric column
    - Middle: Column for each year showing "Number (Percent%)"
    - Far right: Change column showing increase/decrease over entire span
    """
    # Check if we have the required columns
    required_columns = ['total_originations', 'lmict_originations', 'lmib_originations', 'mmct_originations']
    if not all(col in df.columns for col in required_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Sum originations
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'lmib': lmib,
            'mmct': mmct
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add change column
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Define metrics in order
    metrics = [
        ('total', 'Loans'),
        ('lmict', 'Low to Moderate Income Census Tract'),
        ('lmib', 'Low to Moderate Income Borrower'),
        ('mmct', 'Majority Minority Census Tract')
    ]
    
    # Calculate for each metric
    for metric_key, metric_label in metrics:
        result_data['Metric'].append(metric_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # For "Loans" row, show just the integer (no percentage)
                if metric_key == 'total':
                    result_data[str(year)].append(f"{count:,}")
                    metric_changes.append((count, 100.0))
                else:
                    # For other rows, show only percentage
                    pct = (count / total * 100) if total > 0 else 0.0
                    result_data[str(year)].append(f"{pct:.1f}%")
                    metric_changes.append((count, pct))
            else:
                if metric_key == 'total':
                    result_data[str(year)].append("0")
                    metric_changes.append((0, 0.0))
                else:
                    result_data[str(year)].append("0.0%")
                    metric_changes.append((0, 0.0))
        
        # Calculate change from first to last year
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            
            # For "Loans" row, show percentage change from first to last year
            if metric_key == 'total':
                if first_count > 0:
                    pct_change = ((last_count - first_count) / first_count) * 100
                    if pct_change > 0:
                        change_str = f"+{pct_change:.1f}%"
                    elif pct_change < 0:
                        change_str = f"{pct_change:.1f}%"
                    else:
                        change_str = "0.0%"
                else:
                    change_str = "N/A"
            else:
                # For other rows, show only percentage point change
                pct_change = last_pct - first_pct
                if pct_change > 0:
                    change_str = f"+{pct_change:.1f}pp"
                elif pct_change < 0:
                    change_str = f"{pct_change:.1f}pp"
                else:
                    change_str = "0.0pp"
            
            result_data[f'Change Over Time ({time_span})'].append(change_str)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append("N/A")
    
    result = pd.DataFrame(result_data)
    return result


def calculate_minority_quartiles(df: pd.DataFrame) -> Dict[str, float]:
    """
    Calculate minority population quartile thresholds from tract-level data.
    
    Args:
        df: DataFrame with tract-level data including 'tract_minority_population_percent'
    
    Returns:
        Dictionary with quartile thresholds: {'q25': float, 'q50': float, 'q75': float}
    """
    if 'tract_minority_population_percent' not in df.columns:
        return {'q25': 0.0, 'q50': 0.0, 'q75': 0.0}
    
    # Get unique tracts with their minority percentages
    tract_minority = df[df['tract_minority_population_percent'].notna()].copy()
    if tract_minority.empty:
        return {'q25': 0.0, 'q50': 0.0, 'q75': 0.0}
    
    # Get unique tract minority percentages (one per tract)
    unique_tracts = tract_minority.groupby('tract_code')['tract_minority_population_percent'].first()
    minority_pcts = unique_tracts.tolist()
    
    if not minority_pcts:
        return {'q25': 0.0, 'q50': 0.0, 'q75': 0.0}
    
    # Calculate quartiles
    q25 = float(np.percentile(minority_pcts, 25))
    q50 = float(np.percentile(minority_pcts, 50))
    q75 = float(np.percentile(minority_pcts, 75))
    
    return {'q25': q25, 'q50': q50, 'q75': q75}


def classify_tract_minority_quartile(minority_pct: float, quartiles: Dict[str, float]) -> str:
    """
    Classify a tract's minority percentage into a quartile.
    
    Args:
        minority_pct: Minority population percentage for the tract
        quartiles: Dictionary with quartile thresholds
    
    Returns:
        Quartile label: 'low', 'moderate', 'middle', or 'high'
    """
    if pd.isna(minority_pct):
        return 'unknown'
    
    if minority_pct <= quartiles['q25']:
        return 'low'
    elif minority_pct <= quartiles['q50']:
        return 'moderate'
    elif minority_pct <= quartiles['q75']:
        return 'middle'
    else:
        return 'high'


def create_income_borrowers_table(df: pd.DataFrame, years: List[int], hud_data: Dict[str, Dict[str, float]] = None) -> pd.DataFrame:
    """
    Create Table 1: Lending to Income Borrowers.
    
    Shows:
    - Low to Moderate Income Borrowers (aggregate)
    - Low Income Borrowers
    - Moderate Income Borrowers
    - Middle Income Borrowers
    - Upper Income Borrowers
    
    For borrower income rows, shows percentage of borrowers in that income category.
    Population Share column uses HUD low/mod income distribution data.
    
    Args:
        df: DataFrame with loan data
        years: List of years
        hud_data: Dictionary mapping GEOID5 to HUD income distribution data (for Population Share)
    """
    # Check required columns
    required_cols = ['total_originations', 'lmib_originations', 'low_income_borrower_originations',
                     'moderate_income_borrower_originations', 'middle_income_borrower_originations',
                     'upper_income_borrower_originations', 'geoid5']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns for income borrowers table: {missing_cols}")
        return pd.DataFrame()
    
    # Aggregate by year and county
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Aggregate by county
        county_totals = {}
        for geoid5 in year_df['geoid5'].unique():
            county_df = year_df[year_df['geoid5'] == geoid5]
            county_totals[geoid5] = {
                'total': int(county_df['total_originations'].sum()),
                'lmib': int(county_df['lmib_originations'].sum()),
                'low': int(county_df['low_income_borrower_originations'].sum()),
                'moderate': int(county_df['moderate_income_borrower_originations'].sum()),
                'middle': int(county_df['middle_income_borrower_originations'].sum()),
                'upper': int(county_df['upper_income_borrower_originations'].sum()),
            }
        
        # Sum across all counties
        total = sum(ct['total'] for ct in county_totals.values())
        lmib = sum(ct['lmib'] for ct in county_totals.values())
        low = sum(ct['low'] for ct in county_totals.values())
        moderate = sum(ct['moderate'] for ct in county_totals.values())
        middle = sum(ct['middle'] for ct in county_totals.values())
        upper = sum(ct['upper'] for ct in county_totals.values())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmib': lmib,
            'low': low,
            'moderate': moderate,
            'middle': middle,
            'upper': upper
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Get HUD population percentages (aggregate across counties)
    hud_low_mod_pct = 0.0
    hud_low_pct = 0.0
    hud_moderate_pct = 0.0
    hud_middle_pct = 0.0
    hud_upper_pct = 0.0
    total_persons = 0

    if hud_data:
        # Aggregate HUD data across all counties in the report
        geoids = df['geoid5'].unique()
        print(f"  [DEBUG] Table 1 - HUD data available with {len(hud_data)} counties")
        print(f"  [DEBUG] Table 1 - HUD data keys (sample): {list(hud_data.keys())[:5]}")
        print(f"  [DEBUG] Table 1 - DataFrame geoids (sample): {[str(g).zfill(5) for g in list(geoids)[:5]]}")

        # Check for matches between DataFrame geoids and HUD data
        matched_geoids = [str(geoid).zfill(5) for geoid in geoids if str(geoid).zfill(5) in hud_data]
        print(f"  [DEBUG] Table 1 - Matched geoids: {len(matched_geoids)} out of {len(geoids)}: {matched_geoids}")

        total_persons = sum(hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0) for geoid in geoids)
        print(f"  [DEBUG] Table 1 - Total persons from HUD data: {total_persons:,}")
        if total_persons > 0:
            hud_low_mod_pct = sum(hud_data.get(str(geoid).zfill(5), {}).get('low_mod_income_pct', 0) * 
                                  hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0) 
                                  for geoid in geoids) / total_persons
            hud_low_pct = sum(hud_data.get(str(geoid).zfill(5), {}).get('low_income_pct', 0) * 
                             hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0) 
                             for geoid in geoids) / total_persons
            hud_moderate_pct = sum(hud_data.get(str(geoid).zfill(5), {}).get('moderate_income_pct', 0) * 
                                  hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0) 
                                  for geoid in geoids) / total_persons
            hud_middle_pct = sum(hud_data.get(str(geoid).zfill(5), {}).get('middle_income_pct', 0) * 
                               hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0) 
                               for geoid in geoids) / total_persons
            hud_upper_pct = sum(hud_data.get(str(geoid).zfill(5), {}).get('upper_income_pct', 0) *
                               hud_data.get(str(geoid).zfill(5), {}).get('total_persons', 0)
                               for geoid in geoids) / total_persons
            print(f"  [DEBUG] Table 1 - HUD percentages: low={hud_low_pct:.1f}%, mod={hud_moderate_pct:.1f}%, mid={hud_middle_pct:.1f}%, upper={hud_upper_pct:.1f}%")
    else:
        print(f"  [WARNING] Table 1 - No HUD data provided, Population Share column will not be added")

    # Add Population Share column (if HUD data available)
    if hud_data and total_persons > 0:
        result_data['Population Share (%)'] = []

    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data['Change'] = []
    
    # Define metrics in order
    # Note: For 'lmib', we'll calculate it as low + moderate to ensure mathematical consistency
    metrics = [
        ('lmib', 'Low to Moderate Income Borrowers', hud_low_mod_pct, True),  # True = calculate as low + moderate
        ('low', 'Low Income Borrowers', hud_low_pct, False),
        ('moderate', 'Moderate Income Borrowers', hud_moderate_pct, False),
        ('middle', 'Middle Income Borrowers', hud_middle_pct, False),
        ('upper', 'Upper Income Borrowers', hud_upper_pct, False),
    ]
    
    # Calculate for each metric
    for metric_key, metric_label, hud_pct, is_calculated in metrics:
        result_data['Metric'].append(metric_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                # For LMI, calculate as Low + Moderate to ensure math is correct
                if is_calculated and metric_key == 'lmib':
                    count = year_data['low'] + year_data['moderate']
                else:
                    count = year_data[metric_key]
                total = year_data['total']
                # Show percentage of loans
                pct = (count / total * 100) if total > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                metric_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                metric_changes.append((0, 0.0))
        
        # Add Population Share column
        if hud_data and total_persons > 0:
            if hud_pct is not None and hud_pct > 0:
                result_data['Population Share (%)'].append(f"{hud_pct:.1f}%")
            else:
                result_data['Population Share (%)'].append("")
        
        # Calculate percentage change
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            pct_change = last_pct - first_pct
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data['Change'].append(change_str)
        else:
            if len(years) >= 2:
                result_data['Change'].append("N/A")
    
    # Add "Total Loans" row as the first row (before all other metrics)
    total_loans_row = {'Metric': 'Total Loans'}
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total_loans_row[str(year)] = f"{year_data['total']:,}"
        else:
            total_loans_row[str(year)] = "0"
    
    # Add Population Share column for Total Loans (empty)
    if hud_data and total_persons > 0:
        total_loans_row['Population Share (%)'] = ""
    
    # Add Change column for Total Loans (percentage change from first to last year)
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            first_total = first_year_data['total']
            last_total = last_year_data['total']
            if first_total > 0:
                pct_change = ((last_total - first_total) / first_total) * 100
                if pct_change > 0:
                    total_loans_row['Change'] = f"+{pct_change:.1f}%"
                elif pct_change < 0:
                    total_loans_row['Change'] = f"{pct_change:.1f}%"
                else:
                    total_loans_row['Change'] = "0.0%"
            else:
                total_loans_row['Change'] = "N/A"
        else:
            total_loans_row['Change'] = "N/A"

    # Insert Total Loans row at the beginning
    result = pd.DataFrame(result_data)
    total_loans_df = pd.DataFrame([total_loans_row])
    result = pd.concat([total_loans_df, result], ignore_index=True)

    return result


def create_income_neighborhood_tracts_table(df: pd.DataFrame, years: List[int], 
                                            hud_data: Dict[str, Dict[str, float]] = None,
                                            census_data: Dict = None) -> pd.DataFrame:
    """
    Create Table 2: Lending to Census Tracts.
    
    Shows:
    - Low to Moderate Income Census Tracts (aggregate)
    - Low Income Census Tracts
    - Moderate Income Census Tracts
    - Middle Income Census Tracts
    - Upper Income Census Tracts
    - Majority Minority Census Tracts (aggregate)
    - Low Minority Census Tracts (bottom quartile)
    - Moderate Minority Census Tracts (2nd quartile)
    - Middle Minority Census Tracts (3rd quartile)
    - High Minority Census Tracts (top quartile)
    
    Args:
        df: DataFrame with loan data
        years: List of years
        hud_data: Dictionary mapping GEOID5 to HUD income distribution data
        census_data: Dictionary with census demographic data for population shares
    """
    # Check required columns
    required_cols = ['total_originations', 'lmict_originations', 'low_income_tract_originations',
                     'moderate_income_tract_originations', 'middle_income_tract_originations',
                     'upper_income_tract_originations', 'mmct_originations', 
                     'tract_minority_population_percent', 'tract_code']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns for income/neighborhood tracts table: {missing_cols}")
        return pd.DataFrame()
    
    # Calculate minority quartiles
    quartiles = calculate_minority_quartiles(df)
    
    # Classify each row's tract into a quartile
    df['minority_quartile'] = df['tract_minority_population_percent'].apply(
        lambda x: classify_tract_minority_quartile(x, quartiles)
    )
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Income tract categories
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        low_tract = int(year_df['low_income_tract_originations'].sum())
        moderate_tract = int(year_df['moderate_income_tract_originations'].sum())
        middle_tract = int(year_df['middle_income_tract_originations'].sum())
        upper_tract = int(year_df['upper_income_tract_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        # Minority quartile categories
        low_minority = int(year_df[year_df['minority_quartile'] == 'low']['total_originations'].sum())
        moderate_minority = int(year_df[year_df['minority_quartile'] == 'moderate']['total_originations'].sum())
        middle_minority = int(year_df[year_df['minority_quartile'] == 'middle']['total_originations'].sum())
        high_minority = int(year_df[year_df['minority_quartile'] == 'high']['total_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'low_tract': low_tract,
            'moderate_tract': moderate_tract,
            'middle_tract': middle_tract,
            'upper_tract': upper_tract,
            'mmct': mmct,
            'low_minority': low_minority,
            'moderate_minority': moderate_minority,
            'middle_minority': middle_minority,
            'high_minority': high_minority
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data['Change'] = []
    
    # Format quartile ranges for labels
    q25_str = f"{quartiles['q25']:.1f}%"
    q50_str = f"{quartiles['q50']:.1f}%"
    q75_str = f"{quartiles['q75']:.1f}%"
    
    # Define metrics in order
    metrics = [
        ('lmict', 'Low to Moderate Income Census Tracts', None),
        ('low_tract', 'Low Income Census Tracts', None),
        ('moderate_tract', 'Moderate Income Census Tracts', None),
        ('middle_tract', 'Middle Income Census Tracts', None),
        ('upper_tract', 'Upper Income Census Tracts', None),
        ('mmct', 'Majority Minority Census Tracts', None),
        ('low_minority', f'Low Minority Census Tracts (0-{q25_str})', None),
        ('moderate_minority', f'Moderate Minority Census Tracts ({q25_str}-{q50_str})', None),
        ('middle_minority', f'Middle Minority Census Tracts ({q50_str}-{q75_str})', None),
        ('high_minority', f'High Minority Census Tracts ({q75_str}-100%)', None),
    ]
    
    # Calculate for each metric
    for metric_key, metric_label, _ in metrics:
        result_data['Metric'].append(metric_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # Show percentage of loans
                pct = (count / total * 100) if total > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                metric_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                metric_changes.append((0, 0.0))
        
        # Calculate percentage change
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            pct_change = last_pct - first_pct
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data['Change'].append(change_str)
        else:
            if len(years) >= 2:
                result_data['Change'].append("N/A")
    
    # Add "Total Loans" row as the first row (before all other metrics)
    total_loans_row = {'Metric': 'Total Loans'}
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total_loans_row[str(year)] = f"{year_data['total']:,}"
        else:
            total_loans_row[str(year)] = "0"
    
    # Add Change column for Total Loans (percentage change from first to last year)
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            first_total = first_year_data['total']
            last_total = last_year_data['total']
            if first_total > 0:
                pct_change = ((last_total - first_total) / first_total) * 100
                if pct_change > 0:
                    total_loans_row['Change'] = f"+{pct_change:.1f}%"
                elif pct_change < 0:
                    total_loans_row['Change'] = f"{pct_change:.1f}%"
                else:
                    total_loans_row['Change'] = "0.0%"
            else:
                total_loans_row['Change'] = "N/A"
        else:
            total_loans_row['Change'] = "N/A"
    
    # Insert Total Loans row at the beginning
    result = pd.DataFrame(result_data)
    total_loans_df = pd.DataFrame([total_loans_row])
    result = pd.concat([total_loans_df, result], ignore_index=True)
    
    return result


def create_income_tracts_table(df: pd.DataFrame, years: List[int],
                                hud_data: Dict[str, Dict[str, float]] = None,
                                census_data: Dict = None) -> pd.DataFrame:
    """
    Create Table 2: Lending to Census Tracts by Income Level.

    Shows:
    - Low to Moderate Income Census Tracts (aggregate)
    - Low Income Census Tracts
    - Moderate Income Census Tracts
    - Middle Income Census Tracts
    - Upper Income Census Tracts

    Population Share column uses ACS tract-level data (percentage of census tracts in each income category).

    Args:
        df: DataFrame with loan data (must include tract_code and tract_to_msa_income_percentage columns)
        years: List of years
        hud_data: Dictionary mapping GEOID5 to HUD income distribution data (not used for Population Share)
        census_data: Dictionary with census demographic data (not used for Population Share)
    """
    # Check required columns
    required_cols = ['total_originations', 'lmict_originations', 'low_income_tract_originations',
                     'moderate_income_tract_originations', 'middle_income_tract_originations',
                     'upper_income_tract_originations', 'geoid5']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns for income tracts table: {missing_cols}")
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Income tract categories
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        low_tract = int(year_df['low_income_tract_originations'].sum())
        moderate_tract = int(year_df['moderate_income_tract_originations'].sum())
        middle_tract = int(year_df['middle_income_tract_originations'].sum())
        upper_tract = int(year_df['upper_income_tract_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'low_tract': low_tract,
            'moderate_tract': moderate_tract,
            'middle_tract': middle_tract,
            'upper_tract': upper_tract
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Calculate tract income shares from ACS tract-level data
    # For Table 2 (Neighborhood Income), Population Share represents the percentage of census tracts
    # (not population) in each income category, calculated from the tract data in the DataFrame
    tract_income_shares = {}
    if 'tract_code' in df.columns:
        if 'tract_to_msa_income_percentage' in df.columns:
            # Get unique tracts and their income classifications
            unique_tracts = df[['tract_code', 'tract_to_msa_income_percentage']].drop_duplicates()
            total_tracts = len(unique_tracts)

            if total_tracts > 0:
                # Classify tracts by income category based on tract_to_msa_income_percentage
                # Low: <= 50%, Moderate: >50% and <=80%, Middle: >80% and <=120%, Upper: >120%
                # Handle NaN/null values
                unique_tracts_clean = unique_tracts.dropna(subset=['tract_to_msa_income_percentage'])
                if len(unique_tracts_clean) > 0:
                    low_tracts = len(unique_tracts_clean[unique_tracts_clean['tract_to_msa_income_percentage'] <= 50])
                    moderate_tracts = len(unique_tracts_clean[(unique_tracts_clean['tract_to_msa_income_percentage'] > 50) &
                                                           (unique_tracts_clean['tract_to_msa_income_percentage'] <= 80)])
                    middle_tracts = len(unique_tracts_clean[(unique_tracts_clean['tract_to_msa_income_percentage'] > 80) &
                                                          (unique_tracts_clean['tract_to_msa_income_percentage'] <= 120)])
                    upper_tracts = len(unique_tracts_clean[unique_tracts_clean['tract_to_msa_income_percentage'] > 120])

                    total_valid_tracts = len(unique_tracts_clean)

                    # Calculate percentages of tracts in each category
                    tract_income_shares['low'] = (low_tracts / total_valid_tracts * 100) if total_valid_tracts > 0 else 0
                    tract_income_shares['moderate'] = (moderate_tracts / total_valid_tracts * 100) if total_valid_tracts > 0 else 0
                    tract_income_shares['lmict'] = tract_income_shares['low'] + tract_income_shares['moderate']
                    tract_income_shares['middle'] = (middle_tracts / total_valid_tracts * 100) if total_valid_tracts > 0 else 0
                    tract_income_shares['upper'] = (upper_tracts / total_valid_tracts * 100) if total_valid_tracts > 0 else 0
                    print(f"  [DEBUG] Calculated tract income shares from ACS data: {tract_income_shares}")
        else:
            print(f"  [WARNING] tract_to_msa_income_percentage column not found in DataFrame. Available columns: {list(df.columns)}")
            print(f"  [WARNING] Population Share column will not be added to Table 2")

    # Always add Population Share column if we have tract_income_shares data
    if tract_income_shares:
        result_data['Population Share (%)'] = []
        print(f"  [DEBUG] Added Population Share (%) column to Table 2")
    else:
        print(f"  [WARNING] No tract_income_shares calculated - Population Share column will not be added")
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data['Change'] = []
    
    # Check if low income tracts exist
    # Use multiple checks: tract income share, loan data, or tract classification
    has_low_income_tracts = False

    # First check: If tract income shares show > 0, tracts definitely exist
    if tract_income_shares and tract_income_shares.get('low', 0) > 0:
        has_low_income_tracts = True
        print(f"  [DEBUG] Low income tracts detected via tract income share: {tract_income_shares.get('low', 0):.1f}%")
    # Second check: If there are any loans to low-income tracts in any year
    elif any(d['low_tract'] > 0 for d in yearly_data):
        has_low_income_tracts = True
        print(f"  [DEBUG] Low income tracts detected via loan data")
    # Third check: If tract classification data shows tracts with <= 50% MSA income
    elif 'tract_code' in df.columns and 'tract_to_msa_income_percentage' in df.columns:
        unique_tracts = df[['tract_code', 'tract_to_msa_income_percentage']].drop_duplicates()
        has_low_income_tracts = len(unique_tracts[unique_tracts['tract_to_msa_income_percentage'] <= 50]) > 0
        if has_low_income_tracts:
            print(f"  [DEBUG] Low income tracts detected via tract classification")

    if not has_low_income_tracts:
        print(f"  [DEBUG] No low income tracts detected - will add asterisk")

    # Define metrics in order
    # Note: For 'lmict', we'll calculate it as low_tract + moderate_tract to ensure mathematical consistency
    metrics = [
        ('lmict', 'Low to Moderate Income Census Tracts', tract_income_shares.get('lmict'), False, True),  # True = calculate as low + moderate
        ('low_tract', 'Low Income Census Tracts', tract_income_shares.get('low'), not has_low_income_tracts, False),
        ('moderate_tract', 'Moderate Income Census Tracts', tract_income_shares.get('moderate'), False, False),
        ('middle_tract', 'Middle Income Census Tracts', tract_income_shares.get('middle'), False, False),
        ('upper_tract', 'Upper Income Census Tracts', tract_income_shares.get('upper'), False, False),
    ]
    
    # Calculate for each metric
    for metric_key, metric_label, pop_share, needs_asterisk, is_calculated in metrics:
        # Add asterisk if this category doesn't exist
        display_label = metric_label + ('*' if needs_asterisk else '')
        result_data['Metric'].append(display_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                # For LMICT, calculate as Low + Moderate to ensure math is correct
                if is_calculated and metric_key == 'lmict':
                    count = year_data['low_tract'] + year_data['moderate_tract']
                else:
                    count = year_data[metric_key]
                total = year_data['total']
                # Show percentage of loans
                pct = (count / total * 100) if total > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                metric_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                metric_changes.append((0, 0.0))
        
        # Add Population Share column
        if tract_income_shares:
            if pop_share is not None:
                result_data['Population Share (%)'].append(f"{pop_share:.1f}%")
            else:
                result_data['Population Share (%)'].append("")
        
        # Calculate percentage change
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            pct_change = last_pct - first_pct
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data['Change'].append(change_str)
        else:
            if len(years) >= 2:
                result_data['Change'].append("N/A")
    
    # Add "Total Loans" row as the first row (before all other metrics)
    total_loans_row = {'Metric': 'Total Loans'}
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total_loans_row[str(year)] = f"{year_data['total']:,}"
        else:
            total_loans_row[str(year)] = "0"
    
    # Add Population Share column for Total Loans (empty)
    if tract_income_shares:
        total_loans_row['Population Share (%)'] = ""
    
    # Add Change column for Total Loans (percentage change from first to last year)
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            first_total = first_year_data['total']
            last_total = last_year_data['total']
            if first_total > 0:
                pct_change = ((last_total - first_total) / first_total) * 100
                if pct_change > 0:
                    total_loans_row['Change'] = f"+{pct_change:.1f}%"
                elif pct_change < 0:
                    total_loans_row['Change'] = f"{pct_change:.1f}%"
                else:
                    total_loans_row['Change'] = "0.0%"
            else:
                total_loans_row['Change'] = "N/A"
        else:
            total_loans_row['Change'] = "N/A"
    
    # Insert Total Loans row at the beginning
    result = pd.DataFrame(result_data)
    total_loans_df = pd.DataFrame([total_loans_row])
    result = pd.concat([total_loans_df, result], ignore_index=True)
    
    return result


def get_tract_population_data_for_counties(df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Fetch tract-level population data from Census API for all counties in the dataframe.
    
    Returns:
        Dictionary mapping tract_code to {total_population, minority_percentage}
    """
    import os
    
    api_key = os.getenv('CENSUS_API_KEY')
    if not api_key:
        print("Warning: CENSUS_API_KEY not set. Cannot fetch tract-level population data.")
        return {}
    
    # Get unique state/county combinations from dataframe
    if 'state_fips' not in df.columns or 'county_fips' not in df.columns:
        print("Warning: state_fips or county_fips not in dataframe. Cannot fetch tract data.")
        return {}
    
    # Get unique state/county pairs
    state_county_pairs = df[['state_fips', 'county_fips']].drop_duplicates()
    
    tract_data = {}
    
    for _, row in state_county_pairs.iterrows():
        state_fips = str(int(row['state_fips'])).zfill(2)
        county_fips = str(int(row['county_fips'])).zfill(3)
        
        try:
            # Use 2022 ACS 5-year estimates (most recent available)
            acs_year = "2022"
            url = f"https://api.census.gov/data/{acs_year}/acs/acs5"
            params = {
                'get': 'NAME,B01003_001E,B03002_003E,tract',
                'for': f'tract:*',
                'in': f'state:{state_fips} county:{county_fips}',
                'key': api_key
            }
            
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if len(data) < 2:
                continue
            
            headers = data[0]
            name_idx = headers.index('NAME')
            total_pop_idx = headers.index('B01003_001E')
            white_non_hisp_idx = headers.index('B03002_003E')
            tract_idx = headers.index('tract')
            
            for row_data in data[1:]:
                tract_code_str = row_data[tract_idx].zfill(6)
                # Create full tract code: state (2) + county (3) + tract (6) = 11 digits
                full_tract_code = f"{state_fips}{county_fips}{tract_code_str}"
                
                try:
                    total_pop = float(row_data[total_pop_idx]) if row_data[total_pop_idx] not in ['-888888888', '-666666666', '-999999999', 'null', 'None', ''] else 0
                    white_non_hisp = float(row_data[white_non_hisp_idx]) if row_data[white_non_hisp_idx] not in ['-888888888', '-666666666', '-999999999', 'null', 'None', ''] else 0
                    
                    if total_pop > 0:
                        minority_pop = total_pop - white_non_hisp
                        minority_pct = (minority_pop / total_pop) * 100 if total_pop > 0 else 0
                        
                        tract_data[full_tract_code] = {
                            'total_population': total_pop,
                            'minority_percentage': minority_pct
                        }
                except (ValueError, IndexError):
                    continue
                    
        except Exception as e:
            print(f"Warning: Error fetching tract data for {state_fips}/{county_fips}: {e}")
            continue
    
    return tract_data


def create_minority_tracts_table(df: pd.DataFrame, years: List[int], 
                                  census_data: Dict = None) -> pd.DataFrame:
    """
    Create Table 3: Lending to Census Tracts by Minority Population.
    
    Shows:
    - Majority Minority Census Tracts (aggregate)
    - Low Minority Census Tracts (bottom quartile)
    - Moderate Minority Census Tracts (2nd quartile)
    - Middle Minority Census Tracts (3rd quartile)
    - High Minority Census Tracts (top quartile)
    
    Args:
        df: DataFrame with loan data
        years: List of years
        census_data: Dictionary with census demographic data for population shares
    """
    # Check required columns
    required_cols = ['total_originations', 'mmct_originations', 
                     'tract_minority_population_percent', 'tract_code']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns for minority tracts table: {missing_cols}")
        return pd.DataFrame()
    
    # Calculate minority quartiles
    quartiles = calculate_minority_quartiles(df)
    
    # Classify each row's tract into a quartile
    df['minority_quartile'] = df['tract_minority_population_percent'].apply(
        lambda x: classify_tract_minority_quartile(x, quartiles)
    )
    
    # Get tract-level population data from Census API
    tract_pop_data = get_tract_population_data_for_counties(df)
    
    # Calculate population shares based on actual population in tracts
    unique_tracts = df[['tract_code', 'minority_quartile', 'tract_minority_population_percent']].drop_duplicates()
    total_tracts = len(unique_tracts)
    quartile_shares = {}
    
    # Calculate MMCT population share using actual population data
    if tract_pop_data and len(tract_pop_data) > 0:
        total_population = 0
        mmct_population = 0
        quartile_populations = {'low': 0, 'moderate': 0, 'middle': 0, 'high': 0}
        
        # Get state and county FIPS from dataframe to construct full tract codes
        if 'state_fips' in df.columns and 'county_fips' in df.columns:
            # Create a mapping of tract_code to state/county for constructing full GEOID
            tract_to_fips = df[['tract_code', 'state_fips', 'county_fips']].drop_duplicates()
            tract_fips_map = {}
            for _, row in tract_to_fips.iterrows():
                tract_code_short = str(row['tract_code']).zfill(6)  # 6-digit tract code
                state_fips = str(int(row['state_fips'])).zfill(2)
                county_fips = str(int(row['county_fips'])).zfill(3)
                full_tract_code = f"{state_fips}{county_fips}{tract_code_short}"  # 11-digit GEOID
                tract_fips_map[tract_code_short] = full_tract_code
        
        for _, tract_row in unique_tracts.iterrows():
            tract_code_short = str(tract_row['tract_code']).zfill(6)  # 6-digit tract code
            tract_minority_pct = tract_row['tract_minority_population_percent']
            quartile = tract_row['minority_quartile']
            
            # Construct full 11-digit tract code for lookup
            full_tract_code = tract_fips_map.get(tract_code_short) if 'tract_fips_map' in locals() else None
            
            if full_tract_code and full_tract_code in tract_pop_data:
                pop = tract_pop_data[full_tract_code]['total_population']
                total_population += pop
                
                # Check if MMCT (>= 50% minority)
                if tract_minority_pct >= 50:
                    mmct_population += pop
                
                # Add to quartile population
                if quartile in quartile_populations:
                    quartile_populations[quartile] += pop
        
        # Calculate population shares
        if total_population > 0:
            quartile_shares['mmct'] = (mmct_population / total_population) * 100
            quartile_shares['low'] = (quartile_populations['low'] / total_population) * 100
            quartile_shares['moderate'] = (quartile_populations['moderate'] / total_population) * 100
            quartile_shares['middle'] = (quartile_populations['middle'] / total_population) * 100
            quartile_shares['high'] = (quartile_populations['high'] / total_population) * 100
        else:
            # Fallback to tract distribution if no population data
            if total_tracts > 0:
                mmct_tracts = unique_tracts[unique_tracts['tract_minority_population_percent'] >= 50]
                quartile_shares['mmct'] = (len(mmct_tracts) / total_tracts * 100) if total_tracts > 0 else 0
                quartile_shares['low'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'low']) / total_tracts * 100
                quartile_shares['moderate'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'moderate']) / total_tracts * 100
                quartile_shares['middle'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'middle']) / total_tracts * 100
                quartile_shares['high'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'high']) / total_tracts * 100
    else:
        # Fallback to tract distribution if Census API data not available
        if total_tracts > 0:
            mmct_tracts = unique_tracts[unique_tracts['tract_minority_population_percent'] >= 50]
            quartile_shares['mmct'] = (len(mmct_tracts) / total_tracts * 100) if total_tracts > 0 else 0
            quartile_shares['low'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'low']) / total_tracts * 100
            quartile_shares['moderate'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'moderate']) / total_tracts * 100
            quartile_shares['middle'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'middle']) / total_tracts * 100
            quartile_shares['high'] = len(unique_tracts[unique_tracts['minority_quartile'] == 'high']) / total_tracts * 100
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        total = int(year_df['total_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        # Minority quartile categories
        low_minority = int(year_df[year_df['minority_quartile'] == 'low']['total_originations'].sum())
        moderate_minority = int(year_df[year_df['minority_quartile'] == 'moderate']['total_originations'].sum())
        middle_minority = int(year_df[year_df['minority_quartile'] == 'middle']['total_originations'].sum())
        high_minority = int(year_df[year_df['minority_quartile'] == 'high']['total_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'mmct': mmct,
            'low_minority': low_minority,
            'moderate_minority': moderate_minority,
            'middle_minority': middle_minority,
            'high_minority': high_minority
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add Population Share column (if quartile shares calculated)
    if quartile_shares:
        result_data['Population Share (%)'] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data['Change'] = []
    
    # Format quartile ranges for labels
    q25_str = f"{quartiles['q25']:.1f}%"
    q50_str = f"{quartiles['q50']:.1f}%"
    q75_str = f"{quartiles['q75']:.1f}%"
    
    # Check if majority minority tracts exist (if mmct count is 0 and no tracts have >= 50% minority)
    has_mmct_tracts = False
    if 'tract_code' in df.columns and 'tract_minority_population_percent' in df.columns:
        unique_tracts = df[['tract_code', 'tract_minority_population_percent']].drop_duplicates()
        has_mmct_tracts = len(unique_tracts[unique_tracts['tract_minority_population_percent'] >= 50]) > 0
    
    # Define metrics in order
    metrics = [
        ('mmct', 'Majority Minority Census Tracts', quartile_shares.get('mmct'), not has_mmct_tracts),
        ('low_minority', f'Low Minority Census Tracts (0-{q25_str})', quartile_shares.get('low'), False),
        ('moderate_minority', f'Moderate Minority Census Tracts ({q25_str}-{q50_str})', quartile_shares.get('moderate'), False),
        ('middle_minority', f'Middle Minority Census Tracts ({q50_str}-{q75_str})', quartile_shares.get('middle'), False),
        ('high_minority', f'High Minority Census Tracts ({q75_str}-100%)', quartile_shares.get('high'), False),
    ]
    
    # Calculate for each metric
    for metric_key, metric_label, pop_share, needs_asterisk in metrics:
        # Add asterisk if this category doesn't exist
        display_label = metric_label + ('*' if needs_asterisk else '')
        result_data['Metric'].append(display_label)
        
        # Calculate for each year
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # Show percentage of loans
                pct = (count / total * 100) if total > 0 else 0.0
                result_data[str(year)].append(f"{pct:.1f}%")
                metric_changes.append((count, pct))
            else:
                result_data[str(year)].append("0.0%")
                metric_changes.append((0, 0.0))
        
        # Add Population Share column
        if quartile_shares:
            if pop_share is not None:
                result_data['Population Share (%)'].append(f"{pop_share:.1f}%")
            else:
                result_data['Population Share (%)'].append("")
        
        # Calculate percentage change
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            pct_change = last_pct - first_pct
            if pct_change > 0:
                change_str = f"+{pct_change:.1f}pp"
            elif pct_change < 0:
                change_str = f"{pct_change:.1f}pp"
            else:
                change_str = "0.0pp"
            result_data['Change'].append(change_str)
        else:
            if len(years) >= 2:
                result_data['Change'].append("N/A")
    
    # Add "Total Loans" row as the first row (before all other metrics)
    total_loans_row = {'Metric': 'Total Loans'}
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            total_loans_row[str(year)] = f"{year_data['total']:,}"
        else:
            total_loans_row[str(year)] = "0"
    
    # Add Population Share column for Total Loans (empty)
    if quartile_shares:
        total_loans_row['Population Share (%)'] = ""
    
    # Add Change column for Total Loans (percentage change from first to last year)
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            first_total = first_year_data['total']
            last_total = last_year_data['total']
            if first_total > 0:
                pct_change = ((last_total - first_total) / first_total) * 100
                if pct_change > 0:
                    total_loans_row['Change'] = f"+{pct_change:.1f}%"
                elif pct_change < 0:
                    total_loans_row['Change'] = f"{pct_change:.1f}%"
                else:
                    total_loans_row['Change'] = "0.0%"
            else:
                total_loans_row['Change'] = "N/A"
        else:
            total_loans_row['Change'] = "N/A"
    
    # Insert Total Loans row at the beginning
    result = pd.DataFrame(result_data)
    total_loans_df = pd.DataFrame([total_loans_row])
    result = pd.concat([total_loans_df, result], ignore_index=True)
    
    return result


def map_lender_type(lender_type: str) -> str:
    """
    Map lender type from lenders18 table to simplified display name.
    
    Args:
        lender_type: Original lender type from lenders18.type_name
    
    Returns:
        Simplified lender type: 'Bank', 'Mortgage', or 'Credit Union'
    """
    if not lender_type:
        return ''
    
    lender_type_lower = str(lender_type).lower()
    
    if 'bank' in lender_type_lower or 'affiliate' in lender_type_lower:
        return 'Bank'
    elif 'mortgage' in lender_type_lower:
        return 'Mortgage'
    elif 'credit union' in lender_type_lower or 'credit' in lender_type_lower:
        return 'Credit Union'
    else:
        return lender_type  # Return original if no match


def create_top_lenders_detailed_table(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create detailed table showing top lenders by total loans in most recent year.
    
    For each lender, shows:
    - Lender Type (Bank, Mortgage, or Credit Union)
    - Total loans/applications (most recent year)
    - Share to each race/ethnic group (using same methodology as demographic overview)
    - Share to income and neighborhood indicators (LMIB, LMICT, MMCT)
    
    Lenders are sorted in descending order by total loans in the most recent year.
    All lenders are included; JavaScript handles showing/hiding rows beyond the first 10.
    """
    if not years:
        return pd.DataFrame()
    
    # Check required columns
    required_columns = ['lender_name', 'total_originations', 'hispanic_originations', 
                       'black_originations', 'white_originations', 'asian_originations',
                       'native_american_originations', 'hopi_originations', 'multi_racial_originations',
                       'lmib_originations', 'lmict_originations', 'mmct_originations',
                       'loans_with_demographic_data']
    
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        return pd.DataFrame()
    
    # FIRST: Calculate overall percentages across ALL YEARS (same logic as Section 1)
    # to determine which columns to include
    yearly_totals = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        if not year_df.empty:
            total_originations = int(year_df['total_originations'].sum())
            hispanic = int(year_df['hispanic_originations'].sum())
            black = int(year_df['black_originations'].sum())
            white = int(year_df['white_originations'].sum())
            asian = int(year_df['asian_originations'].sum())
            native_american = int(year_df['native_american_originations'].sum())
            hopi = int(year_df['hopi_originations'].sum())
            multi_racial = int(year_df['multi_racial_originations'].sum()) if 'multi_racial_originations' in year_df.columns else 0
            
            if 'loans_with_demographic_data' in year_df.columns:
                loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
            else:
                loans_with_demographics = hispanic + black + white + asian + native_american + hopi
            
            yearly_totals.append({
                'year': year,
                'total_originations': total_originations,
                'hispanic': hispanic,
                'black': black,
                'white': white,
                'asian': asian,
                'native_american': native_american,
                'hopi': hopi,
                'multi_racial': multi_racial,
                'loans_with_demographics': loans_with_demographics
            })
    
    # Use same logic as Section 1: max count across all years / max total across all years
    all_totals = [d['total_originations'] for d in yearly_totals]
    max_total = max(all_totals) if all_totals else 0
    
    # Calculate max percentage for each group across all years (matching Section 1 logic)
    # Order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
    group_max_pct = {}
    group_total_share = {}  # For sorting by total share
    for group in ['white', 'hispanic', 'black', 'asian', 'native_american', 'hopi', 'multi_racial']:
        max_count = max([d[group] for d in yearly_totals]) if yearly_totals else 0
        # Use the year with max total as denominator for threshold check (same as Section 1)
        max_pct = (max_count / max_total * 100) if max_total > 0 else 0
        group_max_pct[group] = max_pct
        
        # Calculate total share for sorting
        total_count = sum([d[group] for d in yearly_totals]) if yearly_totals else 0
        total_loans_all_years = sum([d['loans_with_demographics'] for d in yearly_totals]) if yearly_totals else 0
        avg_share = (total_count / total_loans_all_years * 100) if total_loans_all_years > 0 else 0
        group_total_share[group] = avg_share
    
    # Determine which columns to include (>= 1% overall, matching Section 1)
    # Order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
    include_white = group_max_pct.get('white', 0) >= 1.0
    include_hispanic = group_max_pct.get('hispanic', 0) >= 1.0
    include_black = group_max_pct.get('black', 0) >= 1.0
    include_asian = group_max_pct.get('asian', 0) >= 1.0
    include_native_american = group_max_pct.get('native_american', 0) >= 1.0
    include_hopi = group_max_pct.get('hopi', 0) >= 1.0
    # Always include multi-racial if it has any data (even if < 1%)
    include_multi_racial = group_max_pct.get('multi_racial', 0) > 0
    
    # NOW: Get most recent year data for the actual table
    latest_year = max(years)
    latest_year_df = df[df['year'] == latest_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    # Aggregate by lender (using latest year data only for the table)
    lender_data = []
    for lender_name in latest_year_df['lender_name'].unique():
        lender_df = latest_year_df[latest_year_df['lender_name'] == lender_name]
        
        total = int(lender_df['total_originations'].sum())
        loans_with_demo = int(lender_df['loans_with_demographic_data'].sum()) if 'loans_with_demographic_data' in lender_df.columns else total
        
        # Get lender type (if available)
        lender_type = None
        if 'lender_type' in lender_df.columns:
            lender_types = lender_df['lender_type'].dropna().unique()
            if len(lender_types) > 0:
                lender_type = lender_types[0]  # Use the first non-null type
        
        # Race/ethnicity originations
        white = int(lender_df['white_originations'].sum())
        hispanic = int(lender_df['hispanic_originations'].sum())
        black = int(lender_df['black_originations'].sum())
        asian = int(lender_df['asian_originations'].sum())
        native_american = int(lender_df['native_american_originations'].sum())
        hopi = int(lender_df['hopi_originations'].sum())
        multi_racial = int(lender_df['multi_racial_originations'].sum()) if 'multi_racial_originations' in lender_df.columns else 0
        
        # Income and neighborhood indicators
        lmib = int(lender_df['lmib_originations'].sum())
        lmict = int(lender_df['lmict_originations'].sum())
        mmct = int(lender_df['mmct_originations'].sum())
        
        lender_data.append({
            'lender_name': lender_name,
            'lender_type': lender_type,  # Include lender type
            'total_loans': total,
            'loans_with_demographic_data': loans_with_demo,
            'white': white,
            'hispanic': hispanic,
            'black': black,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'multi_racial': multi_racial,
            'lmib': lmib,
            'lmict': lmict,
            'mmct': mmct
        })
    
    # Sort by total loans descending
    lender_data.sort(key=lambda x: x['total_loans'], reverse=True)
    
    # Return all lenders - JavaScript will handle showing/hiding rows beyond the first 10
    # This allows the table to work for communities with fewer than 10 lenders
    if not lender_data:
        return pd.DataFrame()
    
    # Build result table
    result_rows = []
    
    for lender in lender_data:
        lender_name = lender['lender_name']  # Already uppercase from clean_mortgage_data
        total = lender['total_loans']
        loans_with_demo = lender['loans_with_demographic_data']
        
        # Calculate percentages for race/ethnicity (denominator = loans with demographic data)
        denominator_demo = loans_with_demo if loans_with_demo > 0 else total
        
        # Map lender type to simplified name
        mapped_lender_type = map_lender_type(lender['lender_type']) if lender['lender_type'] else ''
        
        # Calculate percentages for income/neighborhood (denominator = total loans)
        row_data = {
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Lender Type': mapped_lender_type,
            'Total Loans': f"{total:,}"
        }
        
        # Race/ethnicity percentages - only include columns that are >= 1% overall
        # Order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
        if include_white:
            white_pct = (lender['white'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['White (%)'] = f"{white_pct:.1f}"
        
        if include_hispanic:
            hispanic_pct = (lender['hispanic'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Hispanic (%)'] = f"{hispanic_pct:.1f}"
        
        if include_black:
            black_pct = (lender['black'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Black (%)'] = f"{black_pct:.1f}"
        
        if include_asian:
            asian_pct = (lender['asian'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Asian (%)'] = f"{asian_pct:.1f}"
        
        if include_native_american:
            native_american_pct = (lender['native_american'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Native American (%)'] = f"{native_american_pct:.1f}"
        
        if include_hopi:
            hopi_pct = (lender['hopi'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Hawaiian/Pacific Islander (%)'] = f"{hopi_pct:.1f}"
        
        if include_multi_racial:
            multi_racial_pct = (lender['multi_racial'] / denominator_demo * 100) if denominator_demo > 0 else 0.0
            row_data['Multi-Racial (%)'] = f"{multi_racial_pct:.1f}"
        
        # Income and neighborhood indicator percentages (denominator = total loans)
        lmib_pct = (lender['lmib'] / total * 100) if total > 0 else 0.0
        lmict_pct = (lender['lmict'] / total * 100) if total > 0 else 0.0
        mmct_pct = (lender['mmct'] / total * 100) if total > 0 else 0.0
        
        row_data['LMIB (%)'] = f"{lmib_pct:.1f}"
        row_data['LMICT (%)'] = f"{lmict_pct:.1f}"
        row_data['MMCT (%)'] = f"{mmct_pct:.1f}"
        
        # Note: "No Data (%)" column removed per user request
        
        result_rows.append(row_data)
    
    result = pd.DataFrame(result_rows)
    return result


def create_mortgage_summary_table(df: pd.DataFrame, counties: List[str], years: List[int]) -> pd.DataFrame:
    """
    Create yearly breakdown table for mortgage origination data.
    
    Similar to branch summary but for mortgage originations.
    Properly dedupes: LMI Borrower only, MMCT only, and both LMICT/MMCT.
    """
    result_data = {
        'Variable': ['Total Originations', 'LMI Borrower Only Originations', 'MMCT Only Originations', 'Both LMICT/MMCT Originations']
    }
    
    # Calculate for each year
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        # Total originations
        total_originations = int(year_df['total_originations'].sum())
        
        # LMI Borrower only (LMIB but not MMCT)
        # Note: We need to check if we have both LMIB and MMCT flags
        # For now, use lmib_originations - mmct_originations as approximation
        # This is a simplification - actual data may need different logic
        lmib_only = int(year_df['lmib_originations'].sum() - year_df[year_df['mmct_originations'] > 0]['lmib_originations'].sum())
        lmib_only = max(0, lmib_only)  # Ensure non-negative
        
        # MMCT only (not LMI Borrower)
        mmct_only = int(year_df['mmct_originations'].sum() - year_df[year_df['lmib_originations'] > 0]['mmct_originations'].sum())
        mmct_only = max(0, mmct_only)  # Ensure non-negative
        
        # Both LMICT/MMCT
        both = int(year_df[(year_df['lmib_originations'] > 0) & (year_df['mmct_originations'] > 0)]['total_originations'].sum())
        
        result_data[str(year)] = [
            total_originations,
            lmib_only,
            mmct_only,
            both
        ]
    
    result = pd.DataFrame(result_data)
    
    # Calculate net change from first to last year
    if len(years) >= 2:
        first_year = str(min(years))
        last_year = str(max(years))
        
        net_changes = []
        for idx, row in result.iterrows():
            first_val = row[first_year]
            last_val = row[last_year]
            net_change = last_val - first_val
            net_changes.append(net_change)
        
        result['Net Change'] = net_changes
    
    return result


def create_lender_summary(df: pd.DataFrame, years: List[int] = None) -> pd.DataFrame:
    """
    Create lender summary table showing top lenders by origination volume.
    
    Similar to bank summary but for mortgage lenders.
    """
    if not years or len(years) == 0:
        return pd.DataFrame()
    
    latest_year = max(years)
    first_year = min(years)
    
    latest_year_df = df[df['year'] == latest_year].copy()
    first_year_df = df[df['year'] == first_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    lender_data = []
    
    # Get all unique lenders
    all_lenders = latest_year_df['lender_name'].unique()
    
    for lender_name in all_lenders:
        # Latest year data for this lender
        lender_latest = latest_year_df[latest_year_df['lender_name'] == lender_name]
        lender_first = first_year_df[first_year_df['lender_name'] == lender_name] if first_year != latest_year else lender_latest
        
        # Sum originations
        total_originations_latest = int(lender_latest['total_originations'].sum())
        total_originations_first = int(lender_first['total_originations'].sum())
        
        # Calculate LMI Borrower, MMCT, and both
        lmib_latest = int(lender_latest['lmib_originations'].sum())
        mmct_latest = int(lender_latest['mmct_originations'].sum())
        both_latest = int(lender_latest[(lender_latest['lmib_originations'] > 0) & (lender_latest['mmct_originations'] > 0)]['total_originations'].sum())
        
        # Calculate percentages
        lmib_pct = round((lmib_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        mmct_pct = round((mmct_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        both_pct = round((both_latest / total_originations_latest * 100), 1) if total_originations_latest > 0 else 0.0
        
        # Net change
        net_change = total_originations_latest - total_originations_first
        
        lender_data.append({
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Total Originations': total_originations_latest,
            'LMI Borrower Only Originations': f"{lmib_latest} ({lmib_pct}%)",
            'MMCT Only Originations': f"{mmct_latest} ({mmct_pct}%)",
            'Both LMICT/MMCT Originations': f"{both_latest} ({both_pct}%)",
            'Net Change': net_change,
            # Store raw values for sorting
            '_total_originations': total_originations_latest
        })
    
    result = pd.DataFrame(lender_data)
    
    # Sort by total originations descending
    result = result.sort_values('_total_originations', ascending=False).reset_index(drop=True)
    
    # Remove helper column
    result = result.drop(columns=['_total_originations'])
    
    return result


def create_mortgage_county_summary(df: pd.DataFrame, counties: List[str] = None, years: List[int] = None) -> pd.DataFrame:
    """
    Create county-by-county summary for mortgage originations.
    
    Only shown if multiple counties selected.
    """
    # Only create this table if multiple counties
    if counties and len(counties) <= 1:
        return pd.DataFrame()
    
    if not years or len(years) == 0:
        return pd.DataFrame()
    
    # Use the most recent year
    latest_year = max(years)
    df_latest = df[df['year'] == latest_year].copy()
    
    if df_latest.empty:
        return pd.DataFrame()
    
    county_data = []
    
    # Process each county
    for county in df_latest['county_state'].unique():
        county_df = df_latest[df_latest['county_state'] == county]
        
        # Sum originations
        total_originations = int(county_df['total_originations'].sum())
        
        # Calculate LMI Borrower, MMCT, and both
        lmib = int(county_df['lmib_originations'].sum())
        mmct = int(county_df['mmct_originations'].sum())
        both = int(county_df[(county_df['lmib_originations'] > 0) & (county_df['mmct_originations'] > 0)]['total_originations'].sum())
        
        # Count unique lenders
        num_lenders = county_df['lender_name'].nunique()
        
        # Calculate percentages
        lmib_pct = round((lmib / total_originations * 100), 1) if total_originations > 0 else 0.0
        mmct_pct = round((mmct / total_originations * 100), 1) if total_originations > 0 else 0.0
        both_pct = round((both / total_originations * 100), 1) if total_originations > 0 else 0.0
        
        county_data.append({
            'County': county,
            'Total Originations': total_originations,
            'LMI Borrower Only Originations': f"{lmib} ({lmib_pct}%)",
            'MMCT Only Originations': f"{mmct} ({mmct_pct}%)",
            'Both LMICT/MMCT Originations': f"{both} ({both_pct}%)",
            'Number of Lenders': num_lenders
        })
    
    result = pd.DataFrame(county_data)
    
    # Sort by Total Originations descending
    result = result.sort_values('Total Originations', ascending=False).reset_index(drop=True)
    
    return result


def create_mortgage_trend_analysis(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create year-over-year trend analysis for mortgage originations.
    """
    if not years or len(years) < 2:
        return pd.DataFrame()
    
    yearly_totals = []
    
    for year in sorted(years):
        year_df = df[df['year'] == year]
        
        total = int(year_df['total_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_totals.append({
            'Year': year,
            'Total Originations': total,
            'LMI Borrower Originations': lmib,
            'MMCT Originations': mmct
        })
    
    result = pd.DataFrame(yearly_totals)
    
    # Calculate year-over-year changes
    if len(result) > 1:
        result['YoY Change (%)'] = result['Total Originations'].pct_change() * 100
        result['YoY Change (%)'] = result['YoY Change (%)'].round(1)
        result.loc[0, 'YoY Change (%)'] = None  # First year has no previous year
    
    return result


def create_demographic_overview_table_for_excel(df: pd.DataFrame, years: List[int], census_data: Dict = None) -> pd.DataFrame:
    """
    Create demographic overview table for Excel export with ALL race/ethnic categories.
    
    This version includes all categories regardless of percentage.
    Note: "No Data" row is not included per user request.
    """
    race_columns = ['hispanic_originations', 'black_originations', 'white_originations', 
                   'asian_originations', 'native_american_originations', 'hopi_originations',
                   'multi_racial_originations']
    
    if not all(col in df.columns for col in race_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        hispanic = int(year_df['hispanic_originations'].sum())
        black = int(year_df['black_originations'].sum())
        white = int(year_df['white_originations'].sum())
        asian = int(year_df['asian_originations'].sum())
        native_american = int(year_df['native_american_originations'].sum())
        hopi = int(year_df['hopi_originations'].sum())
        multi_racial = int(year_df['multi_racial_originations'].sum()) if 'multi_racial_originations' in year_df.columns else 0
        
        total_originations = int(year_df['total_originations'].sum())
        
        if 'loans_with_demographic_data' in year_df.columns:
            loans_with_demographics = int(year_df['loans_with_demographic_data'].sum())
        else:
            loans_with_demographics = hispanic + black + white + asian + native_american + hopi
        
        no_data_loans = total_originations - loans_with_demographics
        
        yearly_data.append({
            'year': year,
            'total_originations': total_originations,
            'hispanic': hispanic,
            'black': black,
            'white': white,
            'asian': asian,
            'native_american': native_american,
            'hopi': hopi,
            'multi_racial': multi_racial,
            'loans_with_demographics': loans_with_demographics,
            'no_data_loans': no_data_loans
        })
    
    # Get population shares from Census data (if available)
    # For multiple counties, aggregate by weighting by total population
    population_shares = {}
    if census_data:
        total_population = 0
        group_totals = {
            'hispanic': 0,
            'black': 0,
            'white': 0,
            'asian': 0,
            'native_american': 0,
            'hopi': 0,
            'multi_racial': 0
        }
        
        # Aggregate across all counties by calculating weighted averages
        for county, county_census in census_data.items():
            demographics = county_census.get('demographics', {})
            if demographics:
                county_pop = demographics.get('total_population', 0)
                if county_pop and county_pop > 0:
                    total_population += county_pop
                    # Calculate actual counts from percentages
                    group_totals['hispanic'] += (demographics.get('hispanic_percentage', 0) / 100) * county_pop
                    group_totals['black'] += (demographics.get('black_percentage', 0) / 100) * county_pop
                    group_totals['white'] += (demographics.get('white_percentage', 0) / 100) * county_pop
                    group_totals['asian'] += (demographics.get('asian_percentage', 0) / 100) * county_pop
                    group_totals['native_american'] += (demographics.get('native_american_percentage', 0) / 100) * county_pop
                    group_totals['hopi'] += (demographics.get('hopi_percentage', 0) / 100) * county_pop
                    group_totals['multi_racial'] += (demographics.get('multi_racial_percentage', 0) / 100) * county_pop
        
        # Calculate weighted average percentages
        if total_population > 0:
            for group in group_totals:
                population_shares[group] = (group_totals[group] / total_population * 100)
    
    # Build result table - include ALL groups
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add Population Share column (if Census data available)
    if population_shares:
        result_data['Population Share (%)'] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Group labels
    group_labels = {
        'hispanic': 'Hispanic',
        'black': 'Black',
        'white': 'White',
        'asian': 'Asian',
        'native_american': 'Native American',
        'hopi': 'Hawaiian/Pacific Islander',
        'multi_racial': 'Multi-Racial'
    }
    
    # Add Total Loans row
    result_data['Metric'].append('Total Loans')
    for year in sorted(years):
        year_data = next((d for d in yearly_data if d['year'] == year), None)
        if year_data:
            result_data[str(year)].append(year_data['total_originations'])
        else:
            result_data[str(year)].append(0)
    
    # Population Share column (empty for Total Loans row)
    if population_shares:
        result_data['Population Share (%)'].append("")
    
    if len(years) >= 2:
        first_year_data = next((d for d in yearly_data if d['year'] == min(years)), None)
        last_year_data = next((d for d in yearly_data if d['year'] == max(years)), None)
        if first_year_data and last_year_data:
            count_change = last_year_data['total_originations'] - first_year_data['total_originations']
            result_data[f'Change Over Time ({time_span})'].append(count_change)
        else:
            result_data[f'Change Over Time ({time_span})'].append(0)
    
    # Add ALL race/ethnicity groups (not filtered by 1%)
    # Order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
    for group in ['white', 'hispanic', 'black', 'asian', 'native_american', 'hopi', 'multi_racial']:
        result_data['Metric'].append(group_labels[group])
        
        group_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data.get(group, 0)
                loans_with_demo = year_data['loans_with_demographics']
                pct = (count / loans_with_demo * 100) if loans_with_demo > 0 else 0.0
                result_data[str(year)].append(pct)
                group_changes.append((count, pct))
            else:
                result_data[str(year)].append(0.0)
                group_changes.append((0, 0.0))
        
        # Add Population Share column (if Census data available)
        if population_shares:
            pop_share = population_shares.get(group, 0)
            result_data['Population Share (%)'].append(pop_share)
        
        # Change column
        if len(years) >= 2 and len(group_changes) >= 2:
            first_pct = group_changes[0][1]
            last_pct = group_changes[-1][1]
            pct_change = last_pct - first_pct
            result_data[f'Change Over Time ({time_span})'].append(pct_change)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append(0.0)
    
    # Note: "No Data" row removed per user request
    
    result = pd.DataFrame(result_data)
    return result


def create_income_neighborhood_indicators_table_for_excel(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create income and neighborhood indicators table for Excel export.
    
    This matches the report table structure exactly:
    - Loans row: integer only
    - Other rows: percentages only
    - Change Over Time column
    """
    required_columns = ['total_originations', 'lmict_originations', 'lmib_originations', 'mmct_originations']
    if not all(col in df.columns for col in required_columns):
        return pd.DataFrame()
    
    # Aggregate by year
    yearly_data = []
    for year in sorted(years):
        year_df = df[df['year'] == year].copy()
        
        total = int(year_df['total_originations'].sum())
        lmict = int(year_df['lmict_originations'].sum())
        lmib = int(year_df['lmib_originations'].sum())
        mmct = int(year_df['mmct_originations'].sum())
        
        yearly_data.append({
            'year': year,
            'total': total,
            'lmict': lmict,
            'lmib': lmib,
            'mmct': mmct
        })
    
    # Build result table
    result_data = {'Metric': []}
    
    # Add year columns
    for year in sorted(years):
        result_data[str(year)] = []
    
    # Add change column
    if len(years) >= 2:
        time_span = f"{min(years)}-{max(years)}"
        result_data[f'Change Over Time ({time_span})'] = []
    
    # Define metrics in order
    metrics = [
        ('total', 'Loans'),
        ('lmict', 'Low to Moderate Income Census Tract'),
        ('lmib', 'Low to Moderate Income Borrower'),
        ('mmct', 'Majority Minority Census Tract')
    ]
    
    # Calculate for each metric
    for metric_key, metric_label in metrics:
        result_data['Metric'].append(metric_label)
        
        metric_changes = []
        for year in sorted(years):
            year_data = next((d for d in yearly_data if d['year'] == year), None)
            if year_data:
                count = year_data[metric_key]
                total = year_data['total']
                # For "Loans" row, show just the integer (no percentage)
                if metric_key == 'total':
                    result_data[str(year)].append(count)
                    metric_changes.append((count, 100.0))
                else:
                    # For other rows, show only percentage
                    pct = (count / total * 100) if total > 0 else 0.0
                    result_data[str(year)].append(pct)
                    metric_changes.append((count, pct))
            else:
                if metric_key == 'total':
                    result_data[str(year)].append(0)
                    metric_changes.append((0, 0.0))
                else:
                    result_data[str(year)].append(0.0)
                    metric_changes.append((0, 0.0))
        
        # Calculate change from first to last year
        if len(years) >= 2 and len(metric_changes) >= 2:
            first_count, first_pct = metric_changes[0]
            last_count, last_pct = metric_changes[-1]
            
            # For "Loans" row, show count change only
            if metric_key == 'total':
                count_change = last_count - first_count
                result_data[f'Change Over Time ({time_span})'].append(count_change)
            else:
                # For other rows, show only percentage point change
                pct_change = last_pct - first_pct
                result_data[f'Change Over Time ({time_span})'].append(pct_change)
        else:
            if len(years) >= 2:
                result_data[f'Change Over Time ({time_span})'].append(0 if metric_key == 'total' else 0.0)
    
    result = pd.DataFrame(result_data)
    return result


def create_top_lenders_table_for_excel(df: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    """
    Create top lenders table for Excel export with ALL race/ethnic categories and No Data column.
    
    This version includes all categories regardless of percentage, and includes the No Data column.
    """
    if not years:
        return pd.DataFrame()
    
    latest_year = max(years)
    
    latest_year_df = df[df['year'] == latest_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    required_columns = ['lender_name', 'total_originations', 'hispanic_originations', 
                       'black_originations', 'white_originations', 'asian_originations',
                       'native_american_originations', 'hopi_originations', 'multi_racial_originations',
                       'lmib_originations', 'lmict_originations', 'mmct_originations',
                       'loans_with_demographic_data']
    
    missing_cols = [col for col in required_columns if col not in latest_year_df.columns]
    if missing_cols:
        return pd.DataFrame()
    
    # Aggregate by lender
    lender_data = []
    for lender_name in latest_year_df['lender_name'].unique():
        lender_df = latest_year_df[latest_year_df['lender_name'] == lender_name]
        
        total = int(lender_df['total_originations'].sum())
        loans_with_demo = int(lender_df['loans_with_demographic_data'].sum()) if 'loans_with_demographic_data' in lender_df.columns else total
        
        lender_type = None
        if 'lender_type' in lender_df.columns:
            lender_types = lender_df['lender_type'].dropna().unique()
            if len(lender_types) > 0:
                lender_type = lender_types[0]
        
        hispanic = int(lender_df['hispanic_originations'].sum())
        black = int(lender_df['black_originations'].sum())
        white = int(lender_df['white_originations'].sum())
        asian = int(lender_df['asian_originations'].sum())
        multi_racial = int(lender_df['multi_racial_originations'].sum()) if 'multi_racial_originations' in lender_df.columns else 0
        native_american = int(lender_df['native_american_originations'].sum())
        hopi = int(lender_df['hopi_originations'].sum())
        
        lmib = int(lender_df['lmib_originations'].sum())
        lmict = int(lender_df['lmict_originations'].sum())
        mmct = int(lender_df['mmct_originations'].sum())
        
        loans_no_data = total - loans_with_demo
        
        lender_data.append({
            'Lender Name': lender_name,  # Already uppercase from clean_mortgage_data
            'Lender Type': lender_type if lender_type else '',
            'Total Loans': total,
            'Hispanic (%)': (hispanic / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Black (%)': (black / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'White (%)': (white / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Asian (%)': (asian / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Native American (%)': (native_american / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Hawaiian/Pacific Islander (%)': (hopi / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'Multi-Racial (%)': (multi_racial / loans_with_demo * 100) if loans_with_demo > 0 else 0.0,
            'LMIB (%)': (lmib / total * 100) if total > 0 else 0.0,
            'LMICT (%)': (lmict / total * 100) if total > 0 else 0.0,
            'MMCT (%)': (mmct / total * 100) if total > 0 else 0.0,
            'No Data (%)': (loans_no_data / total * 100) if total > 0 else 0.0
        })
    
    # Sort by total loans descending
    lender_data.sort(key=lambda x: x['Total Loans'], reverse=True)
    
    result = pd.DataFrame(lender_data)
    return result


def create_top_lenders_by_year_table_for_excel(df: pd.DataFrame, years: List[int], top_n: int = 10) -> pd.DataFrame:
    """
    Create top N lenders by year table for Excel export (2020-2024).
    Shows loan counts and amounts for top lenders across years.
    
    Args:
        df: DataFrame with raw mortgage data
        years: List of years to include (should be 2020-2024)
        top_n: Number of top lenders to include (default 10)
    
    Returns:
        DataFrame with columns: Lender Name, [Year] Loans, [Year] Amount ($000s) for each year
    """
    if df.empty or not years:
        return pd.DataFrame()
    
    # Get top N lenders from latest year
    latest_year = max(years)
    latest_year_df = df[df['year'] == latest_year].copy()
    
    if latest_year_df.empty:
        return pd.DataFrame()
    
    # Aggregate by lender for latest year to get top N
    latest_year_lenders = latest_year_df.groupby('lender_name').agg({
        'total_originations': 'sum'
    }).reset_index()
    latest_year_lenders = latest_year_lenders.sort_values('total_originations', ascending=False)
    top_lender_names = latest_year_lenders.head(top_n)['lender_name'].tolist()
    
    # Build result data
    result_data = []
    
    # Header row
    header = ['Lender Name']
    for year in sorted(years):
        header.append(f'{year} Loans')
    for year in sorted(years):
        header.append(f'{year} Amount ($000s)')
    result_data.append(header)
    
    # Data rows for each top lender
    for lender_name in top_lender_names:
        row = [lender_name]
        
        # Add loan counts for each year
        for year in sorted(years):
            year_df = df[(df['year'] == year) & (df['lender_name'] == lender_name)]
            total_loans = int(year_df['total_originations'].sum()) if not year_df.empty else 0
            row.append(total_loans)
        
        # Add loan amounts for each year (in thousands)
        for year in sorted(years):
            year_df = df[(df['year'] == year) & (df['lender_name'] == lender_name)]
            # Try different column names for loan amount
            if not year_df.empty:
                if 'total_loan_amount' in year_df.columns:
                    total_amount = year_df['total_loan_amount'].sum() / 1000
                elif 'loan_amount' in year_df.columns:
                    total_amount = year_df['loan_amount'].sum() / 1000
                else:
                    total_amount = 0
            else:
                total_amount = 0
            row.append(total_amount)
        
        result_data.append(row)
    
    result = pd.DataFrame(result_data[1:], columns=result_data[0])
    return result


def create_population_demographics_table_for_excel(census_data: Dict = None) -> pd.DataFrame:
    """
    Create population demographics table from Census data for Excel export.
    
    This table shows total population and racial/ethnic composition across time periods
    (2010 Census, 2020 Census, and most recent ACS 5-year estimates).
    
    For multiple counties, demographics are aggregated using weighted averages
    based on each county's population.
    
    Args:
        census_data: Dictionary of Census data by county name, with structure:
            {
                'County Name': {
                    'time_periods': {
                        'census2010': {
                            'year': '2010 Census',
                            'demographics': {
                                'total_population': int,
                                'white_percentage': float,
                                'black_percentage': float,
                                ...
                            }
                        },
                        ...
                    }
                }
            }
    
    Returns:
        DataFrame with rows for Total Population and each demographic group,
        and columns for each time period.
    """
    if not census_data or len(census_data) == 0:
        return pd.DataFrame()
    
    def aggregate_time_period(time_period_key: str):
        """Aggregate demographics across counties for a specific time period using weighted averages."""
        total_population = 0
        white_sum = 0
        black_sum = 0
        asian_sum = 0
        native_am_sum = 0
        hopi_sum = 0
        multi_racial_sum = 0
        hispanic_sum = 0
        
        for county_name, county_data in census_data.items():
            time_periods = county_data.get('time_periods', {})
            period_data = time_periods.get(time_period_key)
            
            if period_data and period_data.get('demographics'):
                demographics = period_data['demographics']
                pop = demographics.get('total_population', 0)
                
                if pop > 0:
                    total_population += pop
                    
                    # Calculate weighted counts: (percentage * population) / 100
                    white_sum += (demographics.get('white_percentage', 0) * pop) / 100
                    black_sum += (demographics.get('black_percentage', 0) * pop) / 100
                    asian_sum += (demographics.get('asian_percentage', 0) * pop) / 100
                    native_am_sum += (demographics.get('native_american_percentage', 0) * pop) / 100
                    hopi_sum += (demographics.get('hopi_percentage', 0) * pop) / 100
                    multi_racial_sum += (demographics.get('multi_racial_percentage', 0) * pop) / 100
                    hispanic_sum += (demographics.get('hispanic_percentage', 0) * pop) / 100
        
        if total_population == 0:
            return None
        
        # Calculate weighted average percentages
        return {
            'total_population': total_population,
            'white_percentage': (white_sum / total_population) * 100,
            'black_percentage': (black_sum / total_population) * 100,
            'hispanic_percentage': (hispanic_sum / total_population) * 100,
            'asian_percentage': (asian_sum / total_population) * 100,
            'native_american_percentage': (native_am_sum / total_population) * 100,
            'hopi_percentage': (hopi_sum / total_population) * 100,
            'multi_racial_percentage': (multi_racial_sum / total_population) * 100
        }
    
    # Aggregate data for each time period
    census2010_data = aggregate_time_period('census2010')
    census2020_data = aggregate_time_period('census2020')
    acs_data = aggregate_time_period('acs')
    
    # Determine which time periods we have data for and get labels
    time_periods = []
    
    if census2010_data:
        time_periods.append({
            'key': 'census2010',
            'label': '2010 Census',
            'data': census2010_data
        })
    
    if census2020_data:
        time_periods.append({
            'key': 'census2020',
            'label': '2020 Census',
            'data': census2020_data
        })
    
    if acs_data:
        # Get the ACS year label from the first county
        first_county = list(census_data.values())[0]
        acs_period = first_county.get('time_periods', {}).get('acs', {})
        acs_label = acs_period.get('year', '2024 ACS')
        time_periods.append({
            'key': 'acs',
            'label': acs_label,
            'data': acs_data
        })
    
    if len(time_periods) == 0:
        return pd.DataFrame()
    
    # Determine visible race groups (show if >= 1% in any time period)
    all_percentages = {
        'white': [],
        'black': [],
        'hispanic': [],
        'asian': [],
        'native_american': [],
        'hopi': [],
        'multi_racial': []
    }
    
    for period in time_periods:
        if period['data']:
            all_percentages['white'].append(period['data'].get('white_percentage', 0))
            all_percentages['black'].append(period['data'].get('black_percentage', 0))
            all_percentages['hispanic'].append(period['data'].get('hispanic_percentage', 0))
            all_percentages['asian'].append(period['data'].get('asian_percentage', 0))
            all_percentages['native_american'].append(period['data'].get('native_american_percentage', 0))
            all_percentages['hopi'].append(period['data'].get('hopi_percentage', 0))
            all_percentages['multi_racial'].append(period['data'].get('multi_racial_percentage', 0))
    
    # Define race groups in desired order: White, Hispanic, Black, Asian, Native American, HoPI, Multi-Racial
    race_groups = [
        {'name': 'White (%)', 'key': 'white', 'percentages': all_percentages['white']},
        {'name': 'Hispanic (%)', 'key': 'hispanic', 'percentages': all_percentages['hispanic']},
        {'name': 'Black (%)', 'key': 'black', 'percentages': all_percentages['black']},
        {'name': 'Asian (%)', 'key': 'asian', 'percentages': all_percentages['asian']},
        {'name': 'Native American (%)', 'key': 'native_american', 'percentages': all_percentages['native_american']},
        {'name': 'Hawaiian/PI (%)', 'key': 'hopi', 'percentages': all_percentages['hopi']},
        {'name': 'Multi-Racial (%)', 'key': 'multi_racial', 'percentages': all_percentages['multi_racial']}
    ]
    
    # Filter out groups that are less than 1% in all time periods
    visible_groups = [g for g in race_groups if any(pct >= 1.0 for pct in g['percentages'])]
    
    # Build the table data
    table_data = []
    
    # Total Population row
    total_pop_row = {'Demographic': 'Total Population'}
    for period in time_periods:
        total_pop_row[period['label']] = period['data'].get('total_population', 0)
    table_data.append(total_pop_row)
    
    # Add rows for visible race groups
    for group in visible_groups:
        group_row = {'Demographic': group['name']}
        for period in time_periods:
            pct = period['data'].get(f"{group['key']}_percentage", 0)
            group_row[period['label']] = pct
        table_data.append(group_row)
    
    # Create DataFrame
    if not table_data:
        return pd.DataFrame()
    
    df = pd.DataFrame(table_data)
    return df


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize a sheet name for Excel compatibility.
    
    Excel sheet names cannot contain: : \ / ? * [ ]
    Also limited to 31 characters.
    """
    # Replace invalid characters
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '-')
    
    # Truncate if too long
    if len(name) > max_length:
        name = name[:max_length]
    
    return name


def save_mortgage_excel_report(report_data: Dict[str, pd.DataFrame], output_path: str, metadata: Dict[str, Any] = None):
    """
    Save the mortgage report data to an Excel file with multiple sheets.
    
    Sheets:
    1. Methods and Definitions
    2. Population Demographics (Census data across time periods)
    3. Section 1: Demographic Overview (with ALL categories and No Data)
    4. Section 2: Income and Neighborhood Indicators
    5. Section 3: Top Lenders (with ALL categories and No Data)
    6. Raw Data (all raw data used in the report)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Get raw data for creating Excel-specific tables
    raw_df = report_data.get('raw_data', pd.DataFrame())
    years = metadata.get('years', []) if metadata else []

    # Check if raw_data is available (it's not stored in cache due to size)
    if raw_df.empty:
        # For cached results, raw_data is not available
        # Create a minimal Excel with just the summary tables
        print("[WARNING] raw_data not available - creating Excel from summary tables only")
        raw_df = None  # Signal that we don't have raw data

    if not years:
        raise ValueError("Cannot create Excel report: missing years information")
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Get workbook and remove default sheet if it exists
        workbook = writer.book
        # Remove default "Sheet" if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Methods and Definitions sheet (first sheet)
        methods_sheet = workbook.create_sheet(sanitize_sheet_name('Methods and Definitions'), 0)
        
        # Prepare methods content
        methods_content = []
        
        methods_content.append(('METHODS AND DEFINITIONS', ''))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Sources', ''))
        methods_content.append(('', 'Home Mortgage Disclosure Act (HMDA) data, compiled and maintained in NCRC\'s curated databases'))
        methods_content.append(('', 'U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates for demographic context'))
        methods_content.append(('', 'Census data retrieved via U.S. Census Bureau API (https://api.census.gov)'))
        if metadata and 'census_data' in metadata and metadata.get('census_data'):
            census_year = None
            for county_data in metadata.get('census_data', {}).values():
                if 'data_year' in county_data:
                    census_year = county_data['data_year']
                    break
            if census_year:
                methods_content.append(('', f'Census data year: {census_year}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Race and Ethnicity Classification', ''))
        methods_content.append(('', 'Race and ethnicity are determined using the COALESCE function across applicant_race_1 through applicant_race_5 and applicant_ethnicity_1 through applicant_ethnicity_5 fields.'))
        methods_content.append(('', 'Hispanic classification takes precedence (if any ethnicity field indicates Hispanic, the borrower is classified as Hispanic).'))
        methods_content.append(('', 'For non-Hispanic borrowers, the first valid race code (excluding codes 6, 7, 8 which indicate "Not Provided", "Not Applicable", or "Information not provided") is used from race_1 through race_5.'))
        methods_content.append(('', 'Percentages are calculated as: (group loans / loans with demographic data) × 100, where loans with demographic data = all loans minus loans lacking race/ethnicity data.'))
        methods_content.append(('', ''))
        
        methods_content.append(('Income and Neighborhood Indicators', ''))
        methods_content.append(('LMIB', 'Low to Moderate Income Borrower: Borrowers with income below 80% of the area median family income'))
        methods_content.append(('LMICT', 'Low to Moderate Income Census Tract: Census tracts where median family income is below 80% of the area median family income'))
        methods_content.append(('MMCT', 'Majority Minority Census Tract: Census tracts where minority populations represent more than 50% of total population'))
        methods_content.append(('', 'Percentages are calculated as: (category loans / total loans) × 100'))
        methods_content.append(('', ''))
        
        methods_content.append(('Data Filters Applied', ''))
        methods_content.append(('', 'Originations only (action taken = 1)'))
        methods_content.append(('', 'Owner-occupied properties (occupancy type = 1)'))
        methods_content.append(('', 'Forward loans (excludes reverse mortgages)'))
        methods_content.append(('', 'Site-built properties (construction method = 1)'))
        methods_content.append(('', '1-4 unit properties (total_units IN (1, 2, 3, 4))'))
        if metadata and 'loan_purpose' in metadata:
            loan_purpose = metadata['loan_purpose']
            if isinstance(loan_purpose, list):
                if len(loan_purpose) == 0 or set(loan_purpose) == {'purchase', 'refinance', 'equity'} or 'all' in loan_purpose:
                    methods_content.append(('', 'Loan Purpose: All loan purposes'))
                else:
                    purpose_map = {
                        'purchase': 'Home Purchase (loan purpose = 1)',
                        'refinance': 'Refinance & Cash-Out (loan purpose IN (31, 32))',
                        'equity': 'Home Equity (loan purpose IN (2, 4))'
                    }
                    purposes = [purpose_map.get(p, p) for p in loan_purpose if p != 'all']
                    methods_content.append(('', f'Loan Purpose: {", ".join(purposes)}'))
        methods_content.append(('', ''))
        
        methods_content.append(('Abbreviations', ''))
        methods_content.append(('HMDA', 'Home Mortgage Disclosure Act'))
        methods_content.append(('CFPB', 'Consumer Financial Protection Bureau'))
        methods_content.append(('NCRC', 'National Community Reinvestment Coalition'))
        methods_content.append(('LMI', 'Low-to-Moderate Income'))
        methods_content.append(('LMIB', 'Low-to-Moderate Income Borrower'))
        methods_content.append(('LMICT', 'Low-to-Moderate Income Census Tract'))
        methods_content.append(('MMCT', 'Majority-Minority Census Tract'))
        methods_content.append(('AMFI', 'Area Median Family Income'))
        methods_content.append(('pp', 'Percentage Points'))
        methods_content.append(('', ''))
        
        # Write methods content
        for row_idx, (label, value) in enumerate(methods_content, start=1):
            methods_sheet.cell(row=row_idx, column=1, value=label)
            methods_sheet.cell(row=row_idx, column=2, value=value)
            if label and label != '':
                methods_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Population Demographics table (from Census data)
        census_data_for_excel = metadata.get('census_data', {}) if metadata else {}
        population_demo_table = create_population_demographics_table_for_excel(census_data=census_data_for_excel)
        if not population_demo_table.empty:
            population_demo_table.to_excel(writer, sheet_name=sanitize_sheet_name('Population Demographics'), index=False)
        
        # Section 1: Demographic Overview (with ALL categories and No Data)
        # Pass census_data if available for population share column
        if raw_df is not None:
            demo_table = create_demographic_overview_table_for_excel(raw_df, years, census_data=census_data_for_excel)
            if not demo_table.empty:
                demo_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 1 - Demographic Overview'), index=False)
        else:
            # Use pre-computed demographic_overview table from report_data
            demographic_overview = report_data.get('demographic_overview', pd.DataFrame())
            if isinstance(demographic_overview, list):
                demographic_overview = pd.DataFrame(demographic_overview)
            if not demographic_overview.empty:
                demographic_overview.to_excel(writer, sheet_name=sanitize_sheet_name('Section 1 - Demographic Overview'), index=False)

        # Section 2: Income and Neighborhood Indicators (3 tables)
        # Table 1: Lending to Income Borrowers
        income_borrowers = report_data.get('income_borrowers', pd.DataFrame())
        if isinstance(income_borrowers, list):
            income_borrowers = pd.DataFrame(income_borrowers)
        if not income_borrowers.empty:
            income_borrowers.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2a - Income Borrowers'), index=False)

        # Table 2: Lending to Census Tracts by Income
        income_tracts = report_data.get('income_tracts', pd.DataFrame())
        if isinstance(income_tracts, list):
            income_tracts = pd.DataFrame(income_tracts)
        if not income_tracts.empty:
            income_tracts.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2b - Income Tracts'), index=False)

        # Table 3: Lending to Census Tracts by Minority Population
        minority_tracts = report_data.get('minority_tracts', pd.DataFrame())
        if isinstance(minority_tracts, list):
            minority_tracts = pd.DataFrame(minority_tracts)
        if not minority_tracts.empty:
            minority_tracts.to_excel(writer, sheet_name=sanitize_sheet_name('Section 2c - Minority Tracts'), index=False)
        
        # Section 3: Top Lenders (with ALL categories and No Data)
        # This matches the report table exactly but includes ALL categories
        if raw_df is not None:
            top_lenders_table = create_top_lenders_table_for_excel(raw_df, years)
            if not top_lenders_table.empty:
                top_lenders_table.to_excel(writer, sheet_name=sanitize_sheet_name('Section 3 - Top Lenders'), index=False)

            # Section 3b: Top 10 Lenders Over Time (2020-2024) - Excel Export Only
            export_years = [y for y in years if 2020 <= y <= 2024]
            if export_years:
                top_lenders_by_year_table = create_top_lenders_by_year_table_for_excel(raw_df, export_years, top_n=10)
                if not top_lenders_by_year_table.empty:
                    top_lenders_by_year_table.to_excel(writer, sheet_name=sanitize_sheet_name('Top 10 Lenders Over Time'), index=False)

            # Raw Data sheet - all raw data used in the report
            if not raw_df.empty:
                raw_export_df = raw_df.copy()
                raw_export_df.to_excel(writer, sheet_name=sanitize_sheet_name('Raw Data'), index=False)
        else:
            # Use pre-computed top_lenders_detailed table from report_data if available
            top_lenders_detailed = report_data.get('top_lenders_detailed', pd.DataFrame())
            if isinstance(top_lenders_detailed, list):
                top_lenders_detailed = pd.DataFrame(top_lenders_detailed)
            if not top_lenders_detailed.empty:
                top_lenders_detailed.to_excel(writer, sheet_name=sanitize_sheet_name('Section 3 - Top Lenders'), index=False)
        
        # Style the header row for all data sheets (not Methods sheet)
        from openpyxl.styles import NamedStyle
        from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_PERCENTAGE
        
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Number format styles
        integer_format = '#,##0'  # Format: 1,234
        percentage_format = '#,##0.00'  # Format: 12.34
        
        for sheet_name in writer.sheets:
            if sheet_name != 'Methods and Definitions':
                worksheet = writer.sheets[sheet_name]
                if worksheet.max_row > 0:
                    # Style header row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply number formatting to data cells
                    # Get header row to identify column types
                    header_row = worksheet[1]
                    headers = [cell.value for cell in header_row]
                    
                    # Iterate through data rows (skip header row)
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col_idx, header in enumerate(headers, start=1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            
                            # Skip if cell is empty or not a number
                            if cell.value is None:
                                continue
                            
                            # Determine format based on header name and value type
                            header_str = str(header) if header else ''
                            
                            # Get the metric value from first column to determine row type
                            metric_cell = worksheet.cell(row=row_idx, column=1)
                            metric_value = metric_cell.value if metric_cell else ''
                            metric_str = str(metric_value) if metric_value else ''
                            
                            # Special handling for Population Demographics sheet
                            if sheet_name == 'Population Demographics':
                                if isinstance(cell.value, (int, float)):
                                    # Total Population row: format as integer
                                    if 'Total Population' in metric_str:
                                        cell.number_format = integer_format
                                    # All other rows (demographic percentages): format as percentage
                                    elif '(%)' in metric_str:
                                        cell.number_format = percentage_format
                            # Check if it's a percentage column
                            elif '%' in header_str or 'percent' in header_str.lower() or 'Population Share' in header_str:
                                # Format as percentage with 2 decimal places
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = percentage_format
                            # Check if it's a year column - format based on row type
                            elif header_str.isdigit() or 'year' in header_str.lower():
                                # Year columns - format based on metric row
                                if isinstance(cell.value, (int, float)):
                                    # If row is "Total Loans" or "Loans", format as integer
                                    if 'Total Loans' in metric_str or metric_str == 'Loans':
                                        cell.number_format = integer_format
                                    else:
                                        # All other rows in year columns are percentages
                                        cell.number_format = percentage_format
                            # Check if it's "Change Over Time" - could be integer or percentage
                            elif 'Change Over Time' in header_str:
                                # Use the metric_str we already determined above
                                # If metric is "Total Loans" or "Loans", it's a count (integer)
                                # Otherwise it's a percentage change
                                if 'Total Loans' in metric_str or metric_str == 'Loans':
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = integer_format
                                else:
                                    # Percentage point change
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = percentage_format
                            # Check if it's "Net Change" - usually integer
                            elif 'Net Change' in header_str:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = integer_format
                            # Check if it's a metric column (first column) - no formatting
                            elif col_idx == 1:
                                pass
                            # Default: if it's a number and not a percentage, format as integer
                            elif isinstance(cell.value, (int, float)):
                                # Check if value looks like a percentage (between 0 and 100, likely decimal)
                                if 0 <= cell.value <= 100 and isinstance(cell.value, float) and cell.value != int(cell.value):
                                    # Likely a percentage value stored as number
                                    cell.number_format = percentage_format
                                else:
                                    # Integer value
                                    cell.number_format = integer_format
        
        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

