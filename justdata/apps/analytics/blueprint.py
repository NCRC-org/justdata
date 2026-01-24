"""
Flask Blueprint for Analytics application.

Provides admin-only routes for viewing user activity patterns,
research activity maps, and coalition-building opportunities.
"""

from flask import Blueprint, jsonify, request, render_template, Response
from justdata.main.auth import login_required, admin_required, staff_required
import io
from datetime import datetime

from .bigquery_client import (
    get_user_locations,
    get_research_activity,
    get_lender_interest,
    get_coalition_opportunities,
    get_summary,
    get_user_activity_timeline,
    sync_new_events
)

# Create blueprint
analytics_bp = Blueprint(
    'analytics',
    __name__,
    url_prefix='/analytics',
    template_folder='templates',
    static_folder='static',
    static_url_path='/analytics/static'
)

# Base template context
BASE_CONTEXT = {
    'app_name': 'Analytics',
    'app_description': 'Internal staff analytics for JustData usage patterns and coalition opportunities',
    'app_subtitle': 'Staff Analytics Tool'
}


def get_context(breadcrumb_items=None):
    """Get template context with optional breadcrumbs."""
    ctx = BASE_CONTEXT.copy()
    if breadcrumb_items:
        ctx['breadcrumb_items'] = breadcrumb_items
    return ctx


# ============================================
# Page Routes
# ============================================

@analytics_bp.route('/')
@login_required
@staff_required
def dashboard():
    """Main analytics dashboard."""
    # Trigger incremental sync on dashboard load (rate-limited to once per hour)
    try:
        sync_result = sync_new_events()
        if sync_result.get('synced_count', 0) > 0:
            print(f"[INFO] Analytics: Synced {sync_result['synced_count']} new events on dashboard load")
    except Exception as e:
        print(f"[WARN] Analytics: Sync check failed: {e}")

    breadcrumbs = [{'name': 'Analytics', 'url': '/analytics'}]
    return render_template('analytics/dashboard.html', **get_context(breadcrumbs))


@analytics_bp.route('/user-map')
@login_required
@staff_required
def user_map():
    """User locations map view."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Report Locations', 'url': '/analytics/user-map'}
    ]
    return render_template('analytics/user_map.html', **get_context(breadcrumbs))


@analytics_bp.route('/research-map')
@login_required
@staff_required
def research_map():
    """Research activity map view."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Research Activity', 'url': '/analytics/research-map'}
    ]
    return render_template('analytics/research_map.html', **get_context(breadcrumbs))


@analytics_bp.route('/lender-map')
@login_required
@staff_required
def lender_map():
    """Lender interest map view."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Lender Interest', 'url': '/analytics/lender-map'}
    ]
    return render_template('analytics/lender_map.html', **get_context(breadcrumbs))


@analytics_bp.route('/coalitions')
@login_required
@staff_required
def coalitions():
    """Coalition opportunities table view."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Coalitions', 'url': '/analytics/coalitions'}
    ]
    return render_template('analytics/coalitions.html', **get_context(breadcrumbs))


@analytics_bp.route('/users')
@login_required
@staff_required
def users():
    """Users tab view."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Users', 'url': '/analytics/users'}
    ]
    return render_template('analytics/users.html', **get_context(breadcrumbs))


# ============================================
# API Routes
# ============================================

@analytics_bp.route('/api/summary')
@login_required
@staff_required
def api_summary():
    """Get summary metrics for dashboard."""
    try:
        days = request.args.get('days', 90, type=int)
        data = get_summary(days=days)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/user-locations')
@login_required
@staff_required
def api_user_locations():
    """Get user location clusters."""
    try:
        days = request.args.get('days', 90, type=int)
        state = request.args.get('state', None)
        # User type filtering
        user_types = request.args.getlist('user_types')
        exclude_user_types = request.args.getlist('exclude_user_types')
        data = get_user_locations(
            days=days,
            state=state,
            user_types=user_types if user_types else None,
            exclude_user_types=exclude_user_types if exclude_user_types else None
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/research-activity')
@login_required
@staff_required
def api_research_activity():
    """Get research activity by county."""
    try:
        days = request.args.get('days', 90, type=int)
        app = request.args.get('app', None)
        state = request.args.get('state', None)
        # User type filtering
        user_types = request.args.getlist('user_types')
        exclude_user_types = request.args.getlist('exclude_user_types')
        data = get_research_activity(
            days=days,
            app=app,
            state=state,
            user_types=user_types if user_types else None,
            exclude_user_types=exclude_user_types if exclude_user_types else None
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/lender-interest')
@login_required
@staff_required
def api_lender_interest():
    """Get lender interest data."""
    try:
        days = request.args.get('days', 90, type=int)
        min_users = request.args.get('min_users', 1, type=int)
        # User type filtering
        user_types = request.args.getlist('user_types')
        exclude_user_types = request.args.getlist('exclude_user_types')
        data = get_lender_interest(
            days=days,
            min_users=min_users,
            user_types=user_types if user_types else None,
            exclude_user_types=exclude_user_types if exclude_user_types else None
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/coalition-opportunities')
@login_required
@staff_required
def api_coalition_opportunities():
    """Get coalition-building opportunities."""
    try:
        days = request.args.get('days', 90, type=int)
        min_users = request.args.get('min_users', 3, type=int)
        entity_type = request.args.get('entity_type', None)
        state = request.args.get('state', None)
        data = get_coalition_opportunities(
            days=days,
            min_users=min_users,
            entity_type=entity_type,
            state=state
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/entity-users')
@login_required
@staff_required
def api_entity_users():
    """Get users researching a specific entity (county or lender)."""
    try:
        from .bigquery_client import get_entity_users
        entity_type = request.args.get('entity_type')
        entity_id = request.args.get('entity_id')
        days = request.args.get('days', 90, type=int)

        if not entity_type or not entity_id:
            return jsonify({'success': False, 'error': 'entity_type and entity_id are required'}), 400

        data = get_entity_users(
            entity_type=entity_type,
            entity_id=entity_id,
            days=days
        )
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/timeline')
@login_required
@staff_required
def api_timeline():
    """Get activity timeline data."""
    try:
        days = request.args.get('days', 30, type=int)
        data = get_user_activity_timeline(days=days)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/lender-interest/<lender_id>')
@login_required
@staff_required
def lender_detail(lender_id):
    """Lender detail page showing reports and researchers for a specific lender."""
    breadcrumbs = [
        {'name': 'Analytics', 'url': '/analytics'},
        {'name': 'Lender Interest', 'url': '/analytics/lender-map'},
        {'name': 'Lender Details', 'url': f'/analytics/lender-interest/{lender_id}'}
    ]
    return render_template('analytics/lender_detail.html', lender_id=lender_id, **get_context(breadcrumbs))


@analytics_bp.route('/api/lender-detail/<lender_id>')
@login_required
@staff_required
def api_lender_detail(lender_id):
    """Get detailed data for a specific lender including reports and researchers."""
    try:
        from .bigquery_client import get_lender_detail
        days = request.args.get('days', 90, type=int)
        data = get_lender_detail(lender_id=lender_id, days=days)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/users')
@login_required
@staff_required
def api_users():
    """Get list of users with activity summary."""
    try:
        from .bigquery_client import get_users
        days = request.args.get('days', 90, type=int)
        search = request.args.get('search', None)
        data = get_users(days=days, search=search)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/users/<user_id>/activity')
@login_required
@staff_required
def api_user_activity(user_id):
    """Get detailed activity for a specific user."""
    try:
        from .bigquery_client import get_user_activity
        days = request.args.get('days', 90, type=int)
        data = get_user_activity(user_id=user_id, days=days)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/api/user-types')
@login_required
@staff_required
def api_user_types():
    """Get list of distinct user types for filtering."""
    try:
        from .bigquery_client import get_bigquery_client, EVENTS_TABLE
        client = get_bigquery_client()
        query = f"""
            SELECT DISTINCT user_type
            FROM `{EVENTS_TABLE}`
            WHERE user_type IS NOT NULL
            ORDER BY user_type
        """
        results = client.query(query).result()
        user_types = [row.user_type for row in results if row.user_type]
        return jsonify({'success': True, 'data': user_types})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@analytics_bp.route('/health')
def health():
    """Health check endpoint."""
    return jsonify({'status': 'healthy', 'app': 'analytics'})


@analytics_bp.route('/api/export')
@login_required
@staff_required
def api_export():
    """
    Export analytics data to Excel with multiple joinable sheets.

    Sheets:
    - Events: All events (fact table) with user_id, county_fips, lender_id
    - Users: User profiles with user_id as key
    - Counties: County summary with county_fips as key
    - Lenders: Lender summary with lender_id as key
    - Coalitions: Coalition opportunities

    All sheets are joinable via common keys for AI/offline analysis.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from .bigquery_client import (
            get_bigquery_client, EVENTS_TABLE, get_users,
            get_coalition_opportunities, get_entity_users
        )

        days = request.args.get('days', 90, type=int)

        # Create workbook
        wb = Workbook()

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3D5C", end_color="1E3D5C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header(ws):
            """Apply header styling to first row."""
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def auto_width(ws):
            """Auto-adjust column widths."""
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Get BigQuery client
        client = get_bigquery_client()

        # ============================================
        # Sheet 1: Events (Fact Table)
        # ============================================
        ws_events = wb.active
        ws_events.title = "Events"

        # Query all events
        cutoff_clause = f"AND event_timestamp >= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL {days} DAY)" if days > 0 else ""
        events_query = f"""
            SELECT
                event_id,
                event_timestamp,
                event_name AS app_name,
                user_id,
                county_fips,
                county_name,
                state,
                lender_id,
                lender_name
            FROM `{EVENTS_TABLE}`
            WHERE 1=1 {cutoff_clause}
            ORDER BY event_timestamp DESC
            LIMIT 10000
        """
        events_results = list(client.query(events_query).result())

        # Write headers
        event_headers = ["event_id", "event_timestamp", "app_name", "user_id",
                        "county_fips", "county_name", "state", "lender_id", "lender_name"]
        ws_events.append(event_headers)
        style_header(ws_events)

        # Write data
        for row in events_results:
            ws_events.append([
                row.event_id,
                row.event_timestamp.isoformat() if row.event_timestamp else None,
                row.app_name,
                row.user_id,
                row.county_fips,
                row.county_name,
                row.state,
                row.lender_id,
                row.lender_name
            ])

        auto_width(ws_events)

        # ============================================
        # Sheet 2: Users (Dimension Table)
        # ============================================
        ws_users = wb.create_sheet("Users")

        # Query unique users with enrichment from Firestore
        users_data = get_users(days=days)

        user_headers = ["user_id", "user_email", "user_type", "organization_name",
                       "total_reports", "counties_researched", "lenders_researched",
                       "first_activity", "last_activity"]
        ws_users.append(user_headers)
        style_header(ws_users)

        for user in users_data:
            ws_users.append([
                user.get('user_id'),
                user.get('user_email'),
                user.get('user_type'),
                user.get('organization_name'),
                user.get('total_reports', 0),
                user.get('counties_researched', 0),
                user.get('lenders_researched', 0),
                user.get('first_activity'),
                user.get('last_activity')
            ])

        auto_width(ws_users)

        # ============================================
        # Sheet 3: Counties (Dimension Table)
        # ============================================
        ws_counties = wb.create_sheet("Counties")

        counties_query = f"""
            SELECT
                county_fips,
                county_name,
                state,
                COUNT(*) as total_reports,
                COUNT(DISTINCT user_id) as unique_users,
                MAX(event_timestamp) as last_activity
            FROM `{EVENTS_TABLE}`
            WHERE county_fips IS NOT NULL {cutoff_clause}
            GROUP BY county_fips, county_name, state
            ORDER BY total_reports DESC
        """
        counties_results = list(client.query(counties_query).result())

        county_headers = ["county_fips", "county_name", "state", "total_reports",
                         "unique_users", "last_activity"]
        ws_counties.append(county_headers)
        style_header(ws_counties)

        for row in counties_results:
            ws_counties.append([
                row.county_fips,
                row.county_name,
                row.state,
                row.total_reports,
                row.unique_users,
                row.last_activity.isoformat() if row.last_activity else None
            ])

        auto_width(ws_counties)

        # ============================================
        # Sheet 4: Lenders (Dimension Table)
        # ============================================
        ws_lenders = wb.create_sheet("Lenders")

        lenders_query = f"""
            SELECT
                lender_id,
                lender_name,
                COUNT(*) as total_reports,
                COUNT(DISTINCT user_id) as unique_users,
                COUNT(DISTINCT state) as states_researching,
                MAX(event_timestamp) as last_activity
            FROM `{EVENTS_TABLE}`
            WHERE lender_id IS NOT NULL {cutoff_clause}
            GROUP BY lender_id, lender_name
            ORDER BY total_reports DESC
        """
        lenders_results = list(client.query(lenders_query).result())

        lender_headers = ["lender_id", "lender_name", "total_reports",
                         "unique_users", "states_researching", "last_activity"]
        ws_lenders.append(lender_headers)
        style_header(ws_lenders)

        for row in lenders_results:
            ws_lenders.append([
                row.lender_id,
                row.lender_name,
                row.total_reports,
                row.unique_users,
                row.states_researching,
                row.last_activity.isoformat() if row.last_activity else None
            ])

        auto_width(ws_lenders)

        # ============================================
        # Sheet 5: Coalitions (Analysis Table)
        # ============================================
        ws_coalitions = wb.create_sheet("Coalitions")

        coalitions_data = get_coalition_opportunities(days=days, min_users=2)

        coalition_headers = ["entity_type", "entity_id", "entity_name", "unique_users",
                            "unique_organizations", "organizations", "researcher_states"]
        ws_coalitions.append(coalition_headers)
        style_header(ws_coalitions)

        for item in coalitions_data:
            orgs = item.get('organizations', [])
            states = item.get('researcher_states', [])
            ws_coalitions.append([
                item.get('entity_type'),
                item.get('entity_id'),
                item.get('entity_name'),
                item.get('unique_users', 0),
                item.get('unique_organizations', 0),
                "; ".join(orgs) if orgs else None,
                "; ".join(states) if states else None
            ])

        auto_width(ws_coalitions)

        # ============================================
        # Sheet 6: Data Dictionary
        # ============================================
        ws_dict = wb.create_sheet("Data Dictionary")

        dict_headers = ["Sheet", "Column", "Description", "Join Key"]
        ws_dict.append(dict_headers)
        style_header(ws_dict)

        dictionary = [
            ("Events", "event_id", "Unique identifier for each report generation event", ""),
            ("Events", "event_timestamp", "When the report was generated (ISO 8601)", ""),
            ("Events", "app_name", "Application used (lendsight_report, bizsight_report, etc.)", ""),
            ("Events", "user_id", "Firebase user ID", "→ Users.user_id"),
            ("Events", "county_fips", "5-digit FIPS code for county researched", "→ Counties.county_fips"),
            ("Events", "county_name", "Name of county researched", ""),
            ("Events", "state", "State abbreviation", ""),
            ("Events", "lender_id", "LEI or unique lender identifier", "→ Lenders.lender_id"),
            ("Events", "lender_name", "Name of lender researched", ""),
            ("Users", "user_id", "Firebase user ID (primary key)", ""),
            ("Users", "user_email", "User email address", ""),
            ("Users", "user_type", "User category (member, registered, institutional, etc.)", ""),
            ("Users", "organization_name", "User's organization", ""),
            ("Users", "total_reports", "Total reports generated by this user", ""),
            ("Users", "counties_researched", "Number of unique counties researched", ""),
            ("Users", "lenders_researched", "Number of unique lenders researched", ""),
            ("Counties", "county_fips", "5-digit FIPS code (primary key)", ""),
            ("Counties", "county_name", "County name", ""),
            ("Counties", "state", "State abbreviation", ""),
            ("Counties", "total_reports", "Total reports for this county", ""),
            ("Counties", "unique_users", "Number of unique users researching this county", ""),
            ("Lenders", "lender_id", "LEI or unique identifier (primary key)", ""),
            ("Lenders", "lender_name", "Lender name", ""),
            ("Lenders", "total_reports", "Total reports for this lender", ""),
            ("Lenders", "unique_users", "Number of unique users researching this lender", ""),
            ("Lenders", "states_researching", "Number of states with researchers interested", ""),
            ("Coalitions", "entity_type", "Type of entity (county or lender)", ""),
            ("Coalitions", "entity_id", "county_fips or lender_id", "→ Counties or Lenders"),
            ("Coalitions", "unique_users", "Users researching this entity (coalition potential)", ""),
            ("Coalitions", "organizations", "Organizations involved (semicolon-separated)", ""),
        ]

        for row in dictionary:
            ws_dict.append(row)

        auto_width(ws_dict)

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        date_str = datetime.now().strftime("%Y-%m-%d")
        period_str = f"{days}d" if days > 0 else "all-time"
        filename = f"analytics-export-{period_str}-{date_str}.xlsx"

        return Response(
            buffer.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
