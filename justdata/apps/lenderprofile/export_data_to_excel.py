#!/usr/bin/env python3
"""
Export LenderProfile data to Excel with separate sheets for each data source.
Run this script to generate an Excel file with all the data used in reports.
"""

import os
import sys
import json
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

# Load environment variables
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), '.env'))

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def collect_data(lender_name: str) -> dict:
    """Collect all data for a lender."""
    from justdata.apps.lenderprofile.processors.identifier_resolver import IdentifierResolver
    from justdata.apps.lenderprofile.processors.data_collector import DataCollector

    # First resolve identifiers
    resolver = IdentifierResolver()
    print(f"Resolving identifiers for: {lender_name}")
    identifiers = resolver.resolve_by_name(lender_name)
    print(f"Identifiers: {identifiers}")

    # Then collect all data
    collector = DataCollector()
    print(f"Collecting data for: {lender_name}")
    data = collector.collect_all_data(identifiers, lender_name)
    return data


def export_to_excel(data: dict, output_path: str):
    """Export data to Excel with separate sheets."""

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Institution Info
    ws = wb.create_sheet("Institution")
    inst_data = []

    # Flatten institution info
    identifiers = data.get('identifiers', {})
    inst_data.append(["Field", "Value"])
    inst_data.append(["Name", data.get('institution_name', '')])
    inst_data.append(["FDIC Cert", identifiers.get('fdic_cert', '')])
    inst_data.append(["RSSD ID", identifiers.get('rssd_id', '')])
    inst_data.append(["LEI", identifiers.get('lei', '')])
    inst_data.append(["Ticker", identifiers.get('ticker', '')])
    inst_data.append(["SEC CIK", identifiers.get('cik', '')])

    for row in inst_data:
        ws.append(row)

    # 2. FDIC Financials
    ws = wb.create_sheet("FDIC Financials")
    fdic_data = data.get('fdic_financials', {}).get('data', [])
    if fdic_data:
        df = pd.DataFrame(fdic_data)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    else:
        ws.append(["No FDIC financial data available"])

    # 3. Branches
    ws = wb.create_sheet("Branches")
    branch_data = data.get('branch_network', {})
    if branch_data:
        # Summary
        ws.append(["Summary"])
        ws.append(["Total Branches", branch_data.get('current_count', 0)])
        ws.append(["States", len(branch_data.get('by_state', {}))])
        ws.append([])

        # By State
        ws.append(["State", "Branch Count"])
        for state, count in sorted(branch_data.get('by_state', {}).items(), key=lambda x: -x[1]):
            ws.append([state, count])
    else:
        ws.append(["No branch data available"])

    # 4. HMDA Lending
    ws = wb.create_sheet("HMDA Lending")
    hmda_data = data.get('hmda_footprint', {})
    if hmda_data:
        # By Year
        by_year = hmda_data.get('by_year', {})
        ws.append(["Year", "Applications", "Originations", "Amount ($000)"])
        for year, year_data in sorted(by_year.items()):
            if isinstance(year_data, dict):
                ws.append([year, year_data.get('applications', 0), year_data.get('originations', 0), year_data.get('amount', 0)])
        ws.append([])

        # By Purpose - handle various data structures
        ws.append(["Loan Purpose Summary"])
        by_purpose = hmda_data.get('by_purpose_year', {})
        try:
            purposes = set()
            for year_data in by_purpose.values():
                if isinstance(year_data, dict):
                    purposes.update(year_data.keys())

            if purposes:
                ws.append(["Year"] + list(purposes))
                for year in sorted(by_purpose.keys()):
                    row = [year]
                    year_data = by_purpose[year]
                    for purpose in purposes:
                        if isinstance(year_data, dict):
                            purpose_data = year_data.get(purpose, {})
                            if isinstance(purpose_data, dict):
                                row.append(purpose_data.get('originations', 0))
                            else:
                                row.append(purpose_data if isinstance(purpose_data, (int, float)) else 0)
                        else:
                            row.append(0)
                    ws.append(row)
        except Exception as e:
            ws.append([f"Error parsing loan purpose data: {e}"])
    else:
        ws.append(["No HMDA data available"])

    # 5. Small Business Lending (CRA)
    ws = wb.create_sheet("SB Lending")
    sb_data = data.get('sb_lending', {})
    if sb_data and sb_data.get('has_data'):
        yearly = sb_data.get('yearly_lending', [])
        if yearly:
            df = pd.DataFrame(yearly)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
    else:
        ws.append(["No small business lending data available"])

    # 6. CFPB Complaints
    ws = wb.create_sheet("CFPB Complaints")
    cfpb_data = data.get('cfpb_complaints', {})
    if cfpb_data and cfpb_data.get('total', 0) > 0:
        ws.append(["Total Complaints", cfpb_data.get('total', 0)])
        ws.append([])

        # Trends by year
        trends = cfpb_data.get('trends', {})
        by_year = trends.get('by_year', {})
        if by_year:
            ws.append(["Year", "Complaints"])
            for year, count in sorted(by_year.items()):
                ws.append([year, count])
        ws.append([])

        # Top categories
        categories = cfpb_data.get('aggregations', {}).get('products', [])
        if categories:
            ws.append(["Product Category", "Count"])
            for cat in categories[:10]:
                if isinstance(cat, dict):
                    ws.append([cat.get('name', ''), cat.get('count', 0)])
    else:
        ws.append(["No CFPB complaint data available"])

    # 7. SEC Filings
    ws = wb.create_sheet("SEC Filings")
    sec_data = data.get('sec', {})
    if sec_data:
        ws.append(["CIK", sec_data.get('cik', '')])
        ws.append(["Company Name", sec_data.get('company_name', '')])
        ws.append(["Ticker", sec_data.get('ticker', '')])
        ws.append([])

        filings = sec_data.get('filings', {})
        for filing_type in ['10k', '10q', '8k', 'def14a']:
            type_filings = filings.get(filing_type, [])
            if type_filings:
                ws.append([f"{filing_type.upper()} Filings"])
                ws.append(["Form", "Filed Date", "Accession Number"])
                for f in type_filings[:10]:
                    if isinstance(f, dict):
                        ws.append([f.get('form', ''), f.get('filed', ''), f.get('accession', '')])
                    else:
                        ws.append([str(f)])
                ws.append([])
    else:
        ws.append(["No SEC data available"])

    # 8. SEC Topics (NCRC-relevant mentions)
    ws = wb.create_sheet("SEC Topics")
    sec_topics = data.get('sec_topics', {})
    if sec_topics:
        ws.append(["Topic", "Mention Count", "Sample Text"])
        for topic_id, topic_data in sec_topics.items():
            if isinstance(topic_data, dict) and topic_data.get('count', 0) > 0:
                mentions = topic_data.get('mentions', [])
                sample = ''
                if mentions and isinstance(mentions[0], dict):
                    sample = mentions[0].get('text', '')[:200]
                elif mentions:
                    sample = str(mentions[0])[:200]
                ws.append([topic_data.get('name', topic_id), topic_data.get('count', 0), sample])
    else:
        ws.append(["No SEC topic data available"])

    # 9. Executive Compensation
    ws = wb.create_sheet("Executive Comp")
    exec_data = data.get('sec_parsed', {}).get('proxy', {}).get('executive_compensation', [])
    if exec_data:
        ws.append(["Name", "Title", "Salary", "Bonus", "Stock Awards", "Total"])
        for ex in exec_data:
            if isinstance(ex, dict):
                ws.append([
                    ex.get('name', ''),
                    ex.get('title', ''),
                    ex.get('salary', 0),
                    ex.get('bonus', 0),
                    ex.get('stock_awards', 0),
                    ex.get('total', 0)
                ])
    else:
        ws.append(["No executive compensation data available"])

    # 10. News
    ws = wb.create_sheet("News")
    news_data = data.get('news', {}).get('articles', [])
    if news_data:
        ws.append(["Date", "Source", "Title", "URL"])
        for article in news_data[:50]:
            if isinstance(article, dict):
                ws.append([
                    article.get('publishedAt', ''),
                    article.get('source', {}).get('name', ''),
                    article.get('title', ''),
                    article.get('url', '')
                ])
    else:
        ws.append(["No news articles available"])

    # 11. Congressional Trading
    ws = wb.create_sheet("Congressional Trading")
    trading_data = data.get('congressional_trading', {}).get('trades', [])
    if trading_data:
        ws.append(["Date", "Representative", "Transaction", "Amount"])
        for trade in trading_data[:50]:
            if isinstance(trade, dict):
                ws.append([
                    trade.get('transaction_date', ''),
                    trade.get('representative', ''),
                    trade.get('transaction_type', ''),
                    trade.get('amount', '')
                ])
    else:
        ws.append(["No congressional trading data available"])

    # 12. GLEIF Corporate Info
    ws = wb.create_sheet("Corporate Structure")
    gleif_data = data.get('gleif', {})
    if gleif_data:
        entity = gleif_data.get('entity', {})
        if isinstance(entity, dict):
            entity_info = entity.get('entity', {})
            ws.append(["Legal Name", entity_info.get('legalName', {}).get('name', '')])
            ws.append(["Status", entity_info.get('status', '')])
            ws.append(["Jurisdiction", entity_info.get('jurisdiction', '')])

            hq = entity_info.get('headquartersAddress', {})
            ws.append(["HQ City", hq.get('city', '')])
            ws.append(["HQ State", hq.get('region', '')])
            ws.append([])

        parent = gleif_data.get('parent', {})
        if parent:
            ws.append(["Parent Company"])
            parent_attrs = parent.get('attributes', {})
            parent_entity = parent_attrs.get('entity', {})
            ws.append(["Parent Name", parent_entity.get('legalName', {}).get('name', '')])
            ws.append(["Parent LEI", parent_attrs.get('lei', '')])
    else:
        ws.append(["No corporate structure data available"])

    # Save
    wb.save(output_path)
    print(f"Data exported to: {output_path}")
    return output_path


if __name__ == "__main__":
    lender_name = "Fifth Third Bank"

    # Collect data
    data = collect_data(lender_name)

    # Export to Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"C:/Code/JustData/data/{lender_name.replace(' ', '_')}_data_{timestamp}.xlsx"

    # Ensure data directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    export_to_excel(data, output_path)

    # Also save raw JSON for reference
    json_path = output_path.replace('.xlsx', '.json')
    with open(json_path, 'w') as f:
        # Convert non-serializable types
        def default_serializer(obj):
            if hasattr(obj, 'isoformat'):
                return obj.isoformat()
            elif hasattr(obj, 'tolist'):
                return obj.tolist()
            elif hasattr(obj, '__dict__'):
                return str(obj)
            return str(obj)

        json.dump(data, f, indent=2, default=default_serializer)
    print(f"JSON data saved to: {json_path}")
