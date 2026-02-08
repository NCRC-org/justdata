"""
Excel generator for MergerMeter analysis output.

Uses the shared Excel generator from justdata.shared.reporting.merger_excel_generator
which provides a simplified, standardized format for merger analysis reports.

This module acts as a wrapper that:
1. Transforms query results using shared.reporting.merger_data_transformer
2. Builds assessment_areas_data dictionary from enriched counties
3. Calls the shared generator to create the Excel workbook
"""

from pathlib import Path
import pandas as pd
from typing import Dict, Optional, List
import logging
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from justdata.apps.mergermeter.config import PROJECT_ID
from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query

# Import shared modules
try:
    from justdata.shared.reporting.merger_data_transformer import (
        transform_mortgage_data,
        transform_sb_data,
        transform_branch_data
    )
    from justdata.shared.reporting.merger_excel_generator import create_merger_excel as shared_create_merger_excel
    SHARED_GENERATOR_AVAILABLE = True
except ImportError as e:
    SHARED_GENERATOR_AVAILABLE = False
    logging.warning(f"Shared generator not available: {e}. Falling back to legacy implementation.")

logger = logging.getLogger(__name__)


def _get_cbsa_name_from_code(cbsa_code: str, cbsa_name_cache: Dict[str, str] = None) -> str:
    """Look up CBSA name from code, using cache if provided."""
    if cbsa_name_cache is None:
        cbsa_name_cache = {}
    
    cbsa_code_str = str(cbsa_code).strip()
    
    # Check cache first
    if cbsa_code_str in cbsa_name_cache:
        return cbsa_name_cache[cbsa_code_str]
    
    # If code is empty or invalid, return fallback
    if not cbsa_code_str or cbsa_code_str.lower() in ['nan', 'none', '']:
        return f"CBSA {cbsa_code_str}" if cbsa_code_str else "Non-MSA"
    
    # Try to look up from BigQuery
    try:
        client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')
        query = f"""
        SELECT DISTINCT cbsa as cbsa_name
        FROM `{PROJECT_ID}.shared.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) = '{cbsa_code_str}'
        LIMIT 1
        """
        results = execute_query(client, query)
        if results and len(results) > 0:
            cbsa_name = str(results[0].get('cbsa_name', '')).strip()
            if cbsa_name and cbsa_name.lower() not in ['nan', 'none', '']:
                cbsa_name_cache[cbsa_code_str] = cbsa_name
                return cbsa_name
    except Exception as e:
        print(f"    Warning: Could not look up CBSA name for {cbsa_code_str}: {e}")
    
    # Fallback
    if cbsa_code_str == '99999' or cbsa_code_str == '':
        return "Non-MSA"
    return f"CBSA {cbsa_code_str}"


def _transform_mortgage_goals_data(mortgage_goals_data: Optional[Dict]) -> Optional[Dict]:
    """
    Transform mortgage_goals_data from the query format to the expected Excel format.

    Input format (from app.py):
    {
        'home_purchase': DataFrame([{state_name, total_loans, lmict_loans, ...}]),
        'refinance': DataFrame([...]),
        'home_improvement': DataFrame([...])
    }

    Output format (expected by merger_excel_generator):
    {
        'by_state': {
            'Illinois': {
                'home_purchase': {'Loans': 100, '~LMICT': 50, ...},
                'refinance': {...},
                'home_improvement': {...}
            },
            ...
        },
        'grand_total': {
            'home_purchase': {'Loans': 1000, ...},
            'refinance': {...},
            'home_improvement': {...}
        }
    }
    """
    if not mortgage_goals_data:
        return None

    # Column mapping from DataFrame columns to expected metric names
    column_to_metric = {
        'total_loans': 'Loans',
        'lmict_loans': '~LMICT',
        'lmib_loans': '~LMIB',
        'lmib_amount': 'LMIB$',
        'mmct_loans': '~MMCT',
        'minb_loans': '~MINB',
        'asian_loans': '~Asian',
        'black_loans': '~Black',
        'native_american_loans': '~Native American',
        'hopi_loans': '~HoPI',
        'hispanic_loans': '~Hispanic'
    }

    # Loan type mapping
    loan_type_map = {
        'home_purchase': 'home_purchase',
        'refinance': 'refinance',
        'home_improvement': 'home_improvement',
        'home_equity': 'home_equity'  # Keep home_equity as home_equity (codes 2+4)
    }

    by_state = {}
    grand_total = {}

    for loan_type, df in mortgage_goals_data.items():
        mapped_loan_type = loan_type_map.get(loan_type, loan_type)

        if df is None or (hasattr(df, 'empty') and df.empty):
            continue

        if not isinstance(df, pd.DataFrame):
            continue

        # Calculate grand totals for this loan type
        loan_type_totals = {}
        for col, metric in column_to_metric.items():
            if col in df.columns:
                loan_type_totals[metric] = int(df[col].sum()) if pd.notna(df[col].sum()) else 0

        if loan_type_totals:
            grand_total[mapped_loan_type] = loan_type_totals

        # Process each state
        if 'state_name' in df.columns:
            for _, row in df.iterrows():
                state_name = row.get('state_name', '')
                if not state_name or pd.isna(state_name):
                    continue

                if state_name not in by_state:
                    by_state[state_name] = {}

                state_metrics = {}
                for col, metric in column_to_metric.items():
                    if col in df.columns:
                        val = row.get(col, 0)
                        state_metrics[metric] = int(val) if pd.notna(val) else 0

                if state_metrics:
                    by_state[state_name][mapped_loan_type] = state_metrics

    if not by_state and not grand_total:
        return None

    return {
        'by_state': by_state,
        'grand_total': grand_total
    }


def create_merger_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame] = None,
    bank_a_hmda_peer: Optional[pd.DataFrame] = None,
    bank_b_hmda_subject: Optional[pd.DataFrame] = None,
    bank_b_hmda_peer: Optional[pd.DataFrame] = None,
    bank_a_sb_subject: Optional[pd.DataFrame] = None,
    bank_a_sb_peer: Optional[pd.DataFrame] = None,
    bank_b_sb_subject: Optional[pd.DataFrame] = None,
    bank_b_sb_peer: Optional[pd.DataFrame] = None,
    bank_a_branch: Optional[pd.DataFrame] = None,
    bank_b_branch: Optional[pd.DataFrame] = None,
    hhi_data: Optional[pd.DataFrame] = None,
    assessment_areas: Optional[Dict] = None,
    metadata: Optional[Dict] = None,
    mortgage_goals_data: Optional[Dict] = None,
    sb_goals_data: Optional[pd.DataFrame] = None
):
    """
    Create Excel workbook with merger analysis data using shared generator.
    
    This function:
    1. Transforms query results using shared transformer
    2. Builds assessment_areas_data dictionary
    3. Calls shared Excel generator
    
    Args:
        output_path: Path to save Excel file
        bank_a_name: Acquirer bank name
        bank_b_name: Target bank name
        All data DataFrames for HMDA, SB, Branch, and HHI
        assessment_areas: Assessment area information (dict with 'acquirer' and 'target' keys)
        metadata: Additional metadata (years, loan purpose, LEI, RSSD, etc.)
    """
    
    if not SHARED_GENERATOR_AVAILABLE:
        logger.error("Shared generator not available. Cannot create Excel file.")
        raise ImportError("Shared Excel generator modules not found. Please ensure shared.reporting modules are available.")
    
    # Extract metadata
    hmda_years = metadata.get('hmda_years', []) if metadata else []
    sb_years = metadata.get('sb_years', []) if metadata else []
    baseline_hmda_years = metadata.get('baseline_hmda_years', []) if metadata else []
    baseline_sb_years = metadata.get('baseline_sb_years', []) if metadata else []
    bank_a_lei = metadata.get('acquirer_lei', '') if metadata else ''
    bank_b_lei = metadata.get('target_lei', '') if metadata else ''
    bank_a_rssd = metadata.get('acquirer_rssd', '') if metadata else ''
    bank_b_rssd = metadata.get('target_rssd', '') if metadata else ''
    bank_a_sb_id = metadata.get('acquirer_sb_id', '') if metadata else ''
    bank_b_sb_id = metadata.get('target_sb_id', '') if metadata else ''
    
    # Extract filter parameters
    loan_purpose = metadata.get('loan_purpose', '') if metadata else ''
    action_taken = metadata.get('action_taken', '') if metadata else ''
    occupancy_type = metadata.get('occupancy_type', '') if metadata else ''
    total_units = metadata.get('total_units', '') if metadata else ''
    construction_method = metadata.get('construction_method', '') if metadata else ''
    not_reverse = metadata.get('not_reverse', '') if metadata else ''
    
    # Build assessment_areas_data dictionary from assessment_areas
    # Expected format: {'acquirer': {'counties': [...]}, 'target': {'counties': [...]}}
    assessment_areas_data = {}
    if assessment_areas:
        # Get enriched counties from assessment_areas dict
        acquirer_counties = assessment_areas.get('acquirer', {}).get('counties', [])
        target_counties = assessment_areas.get('target', {}).get('counties', [])
        
        assessment_areas_data = {
            'acquirer': {'counties': acquirer_counties},
            'target': {'counties': target_counties}
        }
    else:
        # Empty structure if no assessment areas provided
        assessment_areas_data = {
            'acquirer': {'counties': []},
            'target': {'counties': []}
        }
    
    # Build CBSA name map and required CBSAs from assessment areas
    # Also build reverse map from CBSA names to codes for matching
    cbsa_name_map = {}
    cbsa_name_to_code_map = {}  # Reverse mapping: name -> code
    bank_a_required_cbsas = set()
    bank_b_required_cbsas = set()
    
    def normalize_cbsa_for_matching(code_or_name):
        """Normalize CBSA code or name for matching (handles special cases)."""
        if not code_or_name:
            return None
        s = str(code_or_name).strip()
        # Handle special cases
        if s == '0' or s.lower() == 'nan' or s == '--':
            return None
        # Normalize "Rural *" to "NON-METRO-*"
        if s.lower().startswith('rural '):
            state = s[6:].strip()  # Remove "Rural " prefix
            return f'NON-METRO-{state}'
        # Normalize "State Non-Metro Area" to "NON-METRO-State"
        if ' non-metro' in s.lower() or ' non-msa' in s.lower():
            state = s.split()[0]  # Get first word (state name)
            return f'NON-METRO-{state}'
        return s
    
    for county in assessment_areas_data.get('acquirer', {}).get('counties', []):
        cbsa_code = str(county.get('cbsa_code', '')).strip()
        cbsa_name = county.get('cbsa_name', '').strip()
        state_name = county.get('state_name', '').strip()

        # Handle rural counties (CBSA code 99999) - these are valid, not invalid
        if cbsa_code == '99999':
            # Add the literal '99999' code since HMDA queries return this for rural areas
            bank_a_required_cbsas.add('99999')
            # Also add state-specific rural identifier for better matching
            if state_name:
                rural_name = f"Rural {state_name}"
                normalized_name = normalize_cbsa_for_matching(rural_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = rural_name
                    cbsa_name_map[normalized_name] = rural_name
            elif cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = cbsa_name
            continue

        # Skip truly invalid CBSA codes (but not 99999 which is valid for rural)
        if cbsa_code in ['0', '--', ''] or cbsa_code.lower() == 'nan':
            # If code is invalid but name exists, use name
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map[normalized_name] = cbsa_name
            continue
        
        # Normalize code
        normalized_code = normalize_cbsa_for_matching(cbsa_code)
        if normalized_code:
            bank_a_required_cbsas.add(normalized_code)
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                cbsa_name_map[normalized_code] = cbsa_name
                # Also map name to code for reverse lookup
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    cbsa_name_to_code_map[normalized_name] = normalized_code
        
        # Also add by name if code is missing but name exists
        if not normalized_code and cbsa_name and cbsa_name != '0':
            normalized_name = normalize_cbsa_for_matching(cbsa_name)
            if normalized_name:
                bank_a_required_cbsas.add(normalized_name)  # Use name as key if no code
                cbsa_name_map[normalized_name] = cbsa_name
    
    for county in assessment_areas_data.get('target', {}).get('counties', []):
        cbsa_code = str(county.get('cbsa_code', '')).strip()
        cbsa_name = county.get('cbsa_name', '').strip()
        state_name = county.get('state_name', '').strip()

        # Handle rural counties (CBSA code 99999) - these are valid, not invalid
        if cbsa_code == '99999':
            # Add the literal '99999' code since HMDA queries return this for rural areas
            bank_b_required_cbsas.add('99999')
            # Also add state-specific rural identifier for better matching
            if state_name:
                rural_name = f"Rural {state_name}"
                normalized_name = normalize_cbsa_for_matching(rural_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = rural_name
                    cbsa_name_map[normalized_name] = rural_name
            elif cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = cbsa_name
            continue

        # Skip truly invalid CBSA codes (but not 99999 which is valid for rural)
        if cbsa_code in ['0', '--', ''] or cbsa_code.lower() == 'nan':
            # If code is invalid but name exists, use name
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map[normalized_name] = cbsa_name
            continue
        
        # Normalize code
        normalized_code = normalize_cbsa_for_matching(cbsa_code)
        if normalized_code:
            bank_b_required_cbsas.add(normalized_code)
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                cbsa_name_map[normalized_code] = cbsa_name
                # Also map name to code for reverse lookup
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    cbsa_name_to_code_map[normalized_name] = normalized_code
        
        # Also add by name if code is missing but name exists
        if not normalized_code and cbsa_name and cbsa_name != '0':
            normalized_name = normalize_cbsa_for_matching(cbsa_name)
            if normalized_name:
                bank_b_required_cbsas.add(normalized_name)  # Use name as key if no code
                cbsa_name_map[normalized_name] = cbsa_name
    
    # Also extract CBSA names from query results if available
    # Filter out invalid CBSA codes like '0', '99999', etc.
    for df in [bank_a_hmda_subject, bank_a_hmda_peer, bank_b_hmda_subject, bank_b_hmda_peer,
                bank_a_sb_subject, bank_a_sb_peer, bank_b_sb_subject, bank_b_sb_peer,
                bank_a_branch, bank_b_branch]:
        if df is not None and not df.empty and 'cbsa_code' in df.columns and 'cbsa_name' in df.columns:
            for _, row in df.iterrows():
                cbsa_code = str(row.get('cbsa_code', '')).strip()
                cbsa_name = str(row.get('cbsa_name', '')).strip()
                # Filter out invalid CBSA codes
                if cbsa_code and cbsa_code != '--' and cbsa_code.lower() != 'nan' and cbsa_code != '0' and cbsa_code != '99999':
                    if cbsa_name and cbsa_name != '--' and cbsa_name.lower() not in ['nan', 'none', ''] and cbsa_name != '0':
                        cbsa_name_map[cbsa_code] = cbsa_name
    
    logger.info(f"Built CBSA name map with {len(cbsa_name_map)} entries")
    logger.info(f"Bank A required CBSAs: {len(bank_a_required_cbsas)}")
    logger.info(f"Bank B required CBSAs: {len(bank_b_required_cbsas)}")

    # Transform data using shared transformer
    # Ensure DataFrames are not None (use empty DataFrame instead)
    bank_a_mortgage_transformed = transform_mortgage_data(
        subject_df=bank_a_hmda_subject if bank_a_hmda_subject is not None and not bank_a_hmda_subject.empty else pd.DataFrame(),
        peer_df=bank_a_hmda_peer if bank_a_hmda_peer is not None and not bank_a_hmda_peer.empty else pd.DataFrame(),
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )

    # Transform peer data separately for Excel sheet display
    bank_a_mortgage_peer_transformed = transform_mortgage_data(
        subject_df=bank_a_hmda_peer if bank_a_hmda_peer is not None and not bank_a_hmda_peer.empty else pd.DataFrame(),
        peer_df=pd.DataFrame(),  # No peer-of-peer
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )

    bank_b_mortgage_transformed = transform_mortgage_data(
        subject_df=bank_b_hmda_subject if bank_b_hmda_subject is not None and not bank_b_hmda_subject.empty else pd.DataFrame(),
        peer_df=bank_b_hmda_peer if bank_b_hmda_peer is not None and not bank_b_hmda_peer.empty else pd.DataFrame(),
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )

    # Transform peer data separately for Excel sheet display
    bank_b_mortgage_peer_transformed = transform_mortgage_data(
        subject_df=bank_b_hmda_peer if bank_b_hmda_peer is not None and not bank_b_hmda_peer.empty else pd.DataFrame(),
        peer_df=pd.DataFrame(),  # No peer-of-peer
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )
    
    # Rename SB columns to match transformer expectations
    # MergerMeter uses: sb_loans_count, lmict_loans_count, loans_rev_under_1m, lmict_loans_amount, amount_rev_under_1m
    # Transformer expects: sb_loans_total, lmict_count, loans_rev_under_1m_count, avg_sb_lmict_loan_amount, avg_loan_amt_rum_sb
    def rename_sb_columns(df):
        """Rename SB DataFrame columns to match transformer expectations."""
        if df is None or df.empty:
            return pd.DataFrame()  # Return empty DataFrame with correct structure
        
        df = df.copy()
        
        # Log original columns for debugging
        print(f"[DEBUG] rename_sb_columns - Original columns: {list(df.columns)}")
        logger.debug(f"Original SB DataFrame columns: {list(df.columns)}")
        
        # Rename columns - check if they exist first
        rename_map = {}
        if 'sb_loans_count' in df.columns:
            rename_map['sb_loans_count'] = 'sb_loans_total'
            print(f"[DEBUG] Will rename sb_loans_count -> sb_loans_total")
        elif 'sb_loans_total' not in df.columns:
            warning_msg = "Neither sb_loans_count nor sb_loans_total found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if 'lmict_loans_count' in df.columns:
            rename_map['lmict_loans_count'] = 'lmict_count'
            print(f"[DEBUG] Will rename lmict_loans_count -> lmict_count")
        elif 'lmict_count' not in df.columns:
            warning_msg = "Neither lmict_loans_count nor lmict_count found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if 'loans_rev_under_1m' in df.columns:
            rename_map['loans_rev_under_1m'] = 'loans_rev_under_1m_count'
            print(f"[DEBUG] Will rename loans_rev_under_1m -> loans_rev_under_1m_count")
        elif 'loans_rev_under_1m_count' not in df.columns:
            warning_msg = "Neither loans_rev_under_1m nor loans_rev_under_1m_count found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if rename_map:
            df = df.rename(columns=rename_map)
        
        # Calculate averages from raw amounts and counts (before aggregation)
        # The transformer will aggregate across years, so we calculate per-row averages first
        if 'lmict_loans_amount' in df.columns and 'lmict_count' in df.columns:
            # Calculate avg_sb_lmict_loan_amount per row
            df['avg_sb_lmict_loan_amount'] = df.apply(
                lambda row: row['lmict_loans_amount'] / row['lmict_count'] 
                if pd.notna(row['lmict_count']) and row['lmict_count'] > 0 else None, axis=1
            )
        elif 'avg_sb_lmict_loan_amount' not in df.columns:
            # If we don't have the data to calculate, set to None
            logger.warning("Cannot calculate avg_sb_lmict_loan_amount - missing required columns")
            df['avg_sb_lmict_loan_amount'] = None
        
        if 'amount_rev_under_1m' in df.columns and 'loans_rev_under_1m_count' in df.columns:
            # Calculate avg_loan_amt_rum_sb per row
            df['avg_loan_amt_rum_sb'] = df.apply(
                lambda row: row['amount_rev_under_1m'] / row['loans_rev_under_1m_count']
                if pd.notna(row['loans_rev_under_1m_count']) and row['loans_rev_under_1m_count'] > 0 else None, axis=1
            )
        elif 'avg_loan_amt_rum_sb' not in df.columns:
            # If we don't have the data to calculate, set to None
            logger.warning("Cannot calculate avg_loan_amt_rum_sb - missing required columns")
            df['avg_loan_amt_rum_sb'] = None
        
        print(f"[DEBUG] rename_sb_columns - Final columns: {list(df.columns)}")
        logger.debug(f"Renamed SB DataFrame columns: {list(df.columns)}")
        
        # Verify required columns exist
        required_cols = ['sb_loans_total', 'lmict_count', 'loans_rev_under_1m_count']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            error_msg = f"Missing required columns after rename: {missing_cols}. Available columns: {list(df.columns)}"
            print(f"[ERROR] {error_msg}")
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        return df
    
    try:
        # Log original column names for debugging
        if bank_a_sb_subject is not None and not bank_a_sb_subject.empty:
            print(f"[DEBUG] Bank A SB subject original columns: {list(bank_a_sb_subject.columns)}")
            logger.info(f"Bank A SB subject original columns: {list(bank_a_sb_subject.columns)}")
        if bank_a_sb_peer is not None and not bank_a_sb_peer.empty:
            print(f"[DEBUG] Bank A SB peer original columns: {list(bank_a_sb_peer.columns)}")
            logger.info(f"Bank A SB peer original columns: {list(bank_a_sb_peer.columns)}")
        if bank_b_sb_subject is not None and not bank_b_sb_subject.empty:
            print(f"[DEBUG] Bank B SB subject original columns: {list(bank_b_sb_subject.columns)}")
            logger.info(f"Bank B SB subject original columns: {list(bank_b_sb_subject.columns)}")
        if bank_b_sb_peer is not None and not bank_b_sb_peer.empty:
            print(f"[DEBUG] Bank B SB peer original columns: {list(bank_b_sb_peer.columns)}")
            logger.info(f"Bank B SB peer original columns: {list(bank_b_sb_peer.columns)}")
        
        bank_a_sb_subject_renamed = rename_sb_columns(
            bank_a_sb_subject.copy() if bank_a_sb_subject is not None and not bank_a_sb_subject.empty else pd.DataFrame()
        )
        bank_a_sb_peer_renamed = rename_sb_columns(
            bank_a_sb_peer.copy() if bank_a_sb_peer is not None and not bank_a_sb_peer.empty else pd.DataFrame()
        )
        bank_b_sb_subject_renamed = rename_sb_columns(
            bank_b_sb_subject.copy() if bank_b_sb_subject is not None and not bank_b_sb_subject.empty else pd.DataFrame()
        )
        bank_b_sb_peer_renamed = rename_sb_columns(
            bank_b_sb_peer.copy() if bank_b_sb_peer is not None and not bank_b_sb_peer.empty else pd.DataFrame()
        )
        
        # Log column names after rename for debugging
        if not bank_a_sb_subject_renamed.empty:
            print(f"[DEBUG] Bank A SB subject columns after rename: {list(bank_a_sb_subject_renamed.columns)}")
            logger.info(f"Bank A SB subject columns after rename: {list(bank_a_sb_subject_renamed.columns)}")
        else:
            print(f"[DEBUG] Bank A SB subject DataFrame is empty after rename")
        if not bank_a_sb_peer_renamed.empty:
            print(f"[DEBUG] Bank A SB peer columns after rename: {list(bank_a_sb_peer_renamed.columns)}")
            logger.info(f"Bank A SB peer columns after rename: {list(bank_a_sb_peer_renamed.columns)}")
        else:
            print(f"[DEBUG] Bank A SB peer DataFrame is empty after rename")
        
        bank_a_sb_transformed = transform_sb_data(
            subject_df=bank_a_sb_subject_renamed,
            peer_df=bank_a_sb_peer_renamed,
            cbsa_name_map=cbsa_name_map,
            required_cbsas=bank_a_required_cbsas
        )
        
        bank_b_sb_transformed = transform_sb_data(
            subject_df=bank_b_sb_subject_renamed,
            peer_df=bank_b_sb_peer_renamed,
            cbsa_name_map=cbsa_name_map,
            required_cbsas=bank_b_required_cbsas
        )
    except Exception as e:
        error_msg = f"Error transforming SB data: {e}"
        print(f"[ERROR] {error_msg}")
        print(f"[ERROR] Exception type: {type(e).__name__}")
        
        # Print column information for all SB DataFrames
        print("\n[ERROR] SB DataFrame Column Information:")
        if bank_a_sb_subject is not None:
            print(f"  Bank A SB Subject: {list(bank_a_sb_subject.columns) if not bank_a_sb_subject.empty else 'EMPTY'}")
        else:
            print(f"  Bank A SB Subject: None")
        if bank_a_sb_peer is not None:
            print(f"  Bank A SB Peer: {list(bank_a_sb_peer.columns) if not bank_a_sb_peer.empty else 'EMPTY'}")
        else:
            print(f"  Bank A SB Peer: None")
        if bank_b_sb_subject is not None:
            print(f"  Bank B SB Subject: {list(bank_b_sb_subject.columns) if not bank_b_sb_subject.empty else 'EMPTY'}")
        else:
            print(f"  Bank B SB Subject: None")
        if bank_b_sb_peer is not None:
            print(f"  Bank B SB Peer: {list(bank_b_sb_peer.columns) if not bank_b_sb_peer.empty else 'EMPTY'}")
        else:
            print(f"  Bank B SB Peer: None")
        
        logger.error(error_msg, exc_info=True)
        # Return empty DataFrames if transformation fails
        bank_a_sb_transformed = pd.DataFrame(columns=['assessment_area', 'metric', 'bank_value', 'peer_value', 'difference'])
        bank_b_sb_transformed = pd.DataFrame(columns=['assessment_area', 'metric', 'bank_value', 'peer_value', 'difference'])
        bank_a_sb_peer_renamed = pd.DataFrame()
        bank_b_sb_peer_renamed = pd.DataFrame()
    
    # Transform branch data - need to rename columns to match transformer expectations
    # MergerMeter uses: market_total_branches, market_branches_in_lmict, market_branches_in_mmct
    # Transformer expects: other_total_branches, other_branches_in_lmict, other_branches_in_mmct
    bank_a_branch_for_transform = bank_a_branch.copy() if bank_a_branch is not None and not bank_a_branch.empty else pd.DataFrame()
    if not bank_a_branch_for_transform.empty:
        # Rename market columns to other columns
        rename_map = {
            'market_total_branches': 'other_total_branches',
            'market_branches_in_lmict': 'other_branches_in_lmict',
            'market_branches_in_mmct': 'other_branches_in_mmct'
        }
        bank_a_branch_for_transform = bank_a_branch_for_transform.rename(columns=rename_map)
    
    bank_b_branch_for_transform = bank_b_branch.copy() if bank_b_branch is not None and not bank_b_branch.empty else pd.DataFrame()
    if not bank_b_branch_for_transform.empty:
        # Rename market columns to other columns
        rename_map = {
            'market_total_branches': 'other_total_branches',
            'market_branches_in_lmict': 'other_branches_in_lmict',
            'market_branches_in_mmct': 'other_branches_in_mmct'
        }
        bank_b_branch_for_transform = bank_b_branch_for_transform.rename(columns=rename_map)
    
    bank_a_branch_transformed = transform_branch_data(
        branch_df=bank_a_branch_for_transform,
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )
    
    bank_b_branch_transformed = transform_branch_data(
        branch_df=bank_b_branch_for_transform,
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )
    
    logger.info("Data transformation complete. Calling shared Excel generator...")

    # Transform mortgage goals data to expected format
    transformed_mortgage_goals = _transform_mortgage_goals_data(mortgage_goals_data)
    if transformed_mortgage_goals:
        logger.info(f"Transformed mortgage_goals_data: {len(transformed_mortgage_goals.get('by_state', {}))} states, grand_total has {len(transformed_mortgage_goals.get('grand_total', {}))} loan types")

    # Determine if single bank mode from metadata
    single_bank_mode = metadata.get('single_bank_mode', False) if metadata else False

    # Call shared generator
    shared_create_merger_excel(
        output_path=output_path,
        bank_a_name=bank_a_name,
        bank_b_name=bank_b_name,
        assessment_areas_data=assessment_areas_data,
        mortgage_goals_data=transformed_mortgage_goals,
        sb_goals_data=sb_goals_data if sb_goals_data is not None and not sb_goals_data.empty else None,
        bank_a_mortgage_data=bank_a_mortgage_transformed,
        bank_b_mortgage_data=bank_b_mortgage_transformed,
        bank_a_mortgage_peer_data=bank_a_mortgage_peer_transformed,
        bank_b_mortgage_peer_data=bank_b_mortgage_peer_transformed,
        bank_a_sb_data=bank_a_sb_transformed,
        bank_b_sb_data=bank_b_sb_transformed,
        bank_a_sb_peer_data=bank_a_sb_peer_renamed,
        bank_b_sb_peer_data=bank_b_sb_peer_renamed,
        bank_a_branch_data=bank_a_branch_transformed,
        bank_b_branch_data=bank_b_branch_transformed,
        years_hmda=hmda_years,
        years_sb=sb_years,
        baseline_years_hmda=baseline_hmda_years,
        baseline_years_sb=baseline_sb_years,
        bank_a_lei=bank_a_lei,
        bank_b_lei=bank_b_lei,
        bank_a_rssd=bank_a_rssd,
        bank_b_rssd=bank_b_rssd,
        bank_a_sb_id=bank_a_sb_id,
        bank_b_sb_id=bank_b_sb_id,
        loan_purpose=loan_purpose,
        action_taken=action_taken,
        occupancy_type=occupancy_type,
        total_units=total_units,
        construction_method=construction_method,
        not_reverse=not_reverse,
        single_bank_mode=single_bank_mode
    )
    
    logger.info(f"Excel workbook saved to: {output_path}")
    print(f"\n[OK] Excel workbook saved to: {output_path}")
    
    # Add HHI sheet - always create it, with explanatory note if empty
    try:
        wb = load_workbook(output_path)
        if hhi_data is not None and not hhi_data.empty:
            logger.info(f"[HHI] HHI data provided - Shape: {hhi_data.shape}, Columns: {list(hhi_data.columns)}")
            _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
            logger.info("Added HHI Analysis sheet with data")
        else:
            logger.warning(f"[HHI] No HHI data - creating sheet with explanatory note")
            _add_empty_hhi_sheet(wb, bank_a_name, bank_b_name, bank_a_rssd, bank_b_rssd)
            logger.info("Added HHI Analysis sheet with no-data explanation")
        wb.save(output_path)
    except Exception as e:
        logger.error(f"Could not add HHI sheet: {e}")
        import traceback
        traceback.print_exc()


def _add_empty_hhi_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str, bank_a_rssd: str, bank_b_rssd: str):
    """Add HHI Analysis sheet with explanatory note when no overlapping counties exist."""

    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))

    # Title style
    title_font = Font(bold=True, size=14, color="034EA0")
    note_font = Font(size=11, italic=True)
    info_font = Font(size=11)

    # Title
    ws.merge_cells('A1:G1')
    cell = ws.cell(1, 1, "HHI Analysis - Deposit Market Concentration")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:G3')
    note_cell = ws.cell(3, 1, "No HHI analysis available - Banks do not have any overlapping counties with branch locations.")
    note_cell.font = Font(size=12, bold=True, color="C00000")  # Red color for emphasis
    note_cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Bank details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"• {bank_a_name}").font = info_font
    ws.cell(6, 3, f"RSSD: {bank_a_rssd if bank_a_rssd else 'N/A'}").font = info_font
    ws.cell(7, 1, f"• {bank_b_name}").font = info_font
    ws.cell(7, 3, f"RSSD: {bank_b_rssd if bank_b_rssd else 'N/A'}").font = info_font

    # Additional explanation
    ws.merge_cells('A9:G9')
    ws.cell(9, 1, "HHI (Herfindahl-Hirschman Index) analysis requires both banks to have branch locations in at least one common county.").font = note_font
    ws.merge_cells('A10:G10')
    ws.cell(10, 1, "Since these banks operate in separate geographic markets, deposit concentration analysis cannot be performed.").font = note_font

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    logger.info(f"[HHI] Created empty HHI sheet with explanatory note")


def _add_hhi_sheet(wb: Workbook, hhi_data: pd.DataFrame, bank_a_name: str, bank_b_name: str):
    """Add HHI Analysis sheet to workbook for counties where both banks have branches."""

    logger.info(f"[HHI] Adding HHI sheet - Input data shape: {hhi_data.shape}, Columns: {list(hhi_data.columns)}")
    
    # HHI calculator already filters to only counties where both banks have branches
    # So we can use all the data directly
    hhi_filtered = hhi_data.copy()
    
    if hhi_filtered.empty:
        logger.warning(f"[HHI] No counties with overlapping branches for HHI analysis")
        return
    
    logger.info(f"[HHI] Writing {len(hhi_filtered)} rows to Excel sheet")
    
    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))
    
    # Header style - NCRC blue (#034ea0)
    header_fill = PatternFill(start_color="034EA0", end_color="034EA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Title
    row = 1
    ws.merge_cells(f'A{row}:{get_column_letter(len(hhi_filtered.columns))}{row}')
    cell = ws.cell(row, 1, f"HHI Analysis - Counties with {bank_a_name} and {bank_b_name} Branches")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")
    row += 2
    
    # Headers
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        cell = ws.cell(row, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row += 1
    
    # Data rows
    for df_row_idx, df_row in hhi_filtered.iterrows():
        for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
            value = df_row[col_name]
            # Handle NaN/None values - show as empty string instead of "-"
            if pd.isna(value) or value is None:
                value = ""
            # Debug: Log first row to see what we're writing
            if row == 3 and col_idx == 1:
                logger.info(f"[HHI] Writing first data row - Column '{col_name}': {value} (type: {type(value)})")
            cell = ws.cell(row, col_idx, value)
            
            # Format numbers (only if value is not empty)
            if value != "" and pd.api.types.is_numeric_dtype(hhi_filtered[col_name]):
                if 'hhi' in col_name.lower():
                    cell.number_format = '#,##0'  # HHI is typically an integer
                elif 'deposits' in col_name.lower() or 'share' in col_name.lower():
                    if 'share' in col_name.lower() or 'percent' in col_name.lower():
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
            
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Adjust column widths
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        col_letter = get_column_letter(col_idx)
        try:
            col_max_len = hhi_filtered[col_name].astype(str).str.len().max() if not hhi_filtered[col_name].empty else 10
        except (ValueError, AttributeError):
            col_max_len = 10
        max_length = max(len(str(col_name)), col_max_len)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def create_simple_assessment_areas_sheet(wb, bank_a_name, bank_b_name, assessment_areas_data,
                                        header_fill, header_font, border_thin):
    """Create Assessment Areas sheet as simple side-by-side tables."""
    ws = wb.create_sheet("Assessment Areas")
    
    # Headers
    ws['A1'] = f'{bank_a_name} Assessment Areas'
    ws['A1'].font = Font(bold=True, size=12)
    ws['G1'] = f'{bank_b_name} Assessment Areas'
    ws['G1'].font = Font(bold=True, size=12)
    
    # Column headers for Bank A
    headers_a = ['State', 'Assessment Area', 'CBSA Name', 'CBSA Code', 'County, State', 'GEOID5']
    for col_idx, header in enumerate(headers_a, 1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers for Bank B
    headers_b = ['State', 'Assessment Area', 'CBSA Name', 'CBSA Code', 'County, State', 'GEOID5']
    for col_idx, header in enumerate(headers_b, 7):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get counties and assessment area mappings
    bank_a_key = 'acquirer' if 'acquirer' in assessment_areas_data else 'pnc'
    bank_a_data = assessment_areas_data.get(bank_a_key, {})
    counties_a = bank_a_data.get('mapped_counties', bank_a_data.get('counties', []))
    if not counties_a or not isinstance(counties_a, list):
        if isinstance(assessment_areas_data.get(bank_a_key), list):
            counties_a = assessment_areas_data.get(bank_a_key)
    
    bank_b_key = 'target' if 'target' in assessment_areas_data else 'firstbank'
    bank_b_data = assessment_areas_data.get(bank_b_key, {})
    counties_b = bank_b_data.get('mapped_counties', bank_b_data.get('counties', []))
    if not counties_b or not isinstance(counties_b, list):
        if isinstance(assessment_areas_data.get(bank_b_key), list):
            counties_b = assessment_areas_data.get(bank_b_key)
    
    # Get assessment area names from metadata if available
    # Build maps from GEOID5 to assessment area name for both banks
    bank_a_aa_map = {}  # Map GEOID5 to assessment area name
    bank_b_aa_map = {}  # Map GEOID5 to assessment area name
    
    if 'assessment_areas' in assessment_areas_data:
        aa_list = assessment_areas_data.get('assessment_areas', [])
        # Split into acquirer and target assessment areas
        # We'll match counties to assessment areas by checking if county GEOID5s match
        for aa_info in aa_list:
            aa_name = aa_info.get('name') or aa_info.get('cbsa_name') or aa_info.get('assessment_area') or 'Unnamed Assessment Area'
            counties_list = aa_info.get('counties', [])
            for county_item in counties_list:
                geoid5 = None
                if isinstance(county_item, dict):
                    geoid5 = str(county_item.get('geoid5', '')).zfill(5)
                    if not geoid5 or len(geoid5) != 5:
                        # Try to build from state_code and county_code
                        state_code = county_item.get('state_code') or county_item.get('state_fips')
                        county_code = county_item.get('county_code') or county_item.get('county_fips')
                        if state_code and county_code:
                            geoid5 = str(state_code).zfill(2) + str(county_code).zfill(3)
                
                if geoid5 and len(geoid5) == 5:
                    # Check if this GEOID5 is in bank A or bank B counties
                    for county in counties_a:
                        county_geoid5 = str(county.get('geoid5', '')).zfill(5)
                        if county_geoid5 == geoid5:
                            bank_a_aa_map[geoid5] = aa_name
                            break
                    
                    for county in counties_b:
                        county_geoid5 = str(county.get('geoid5', '')).zfill(5)
                        if county_geoid5 == geoid5:
                            bank_b_aa_map[geoid5] = aa_name
                            break
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Populate Bank A data
    row = 3
    if counties_a:
        # Group by state and assessment area
        state_groups = defaultdict(lambda: defaultdict(list))
        for county in counties_a:
            state = county.get('state_name', county.get('state', '')) or 'Unknown'
            # Try to get assessment area name from map
            geoid5 = str(county.get('geoid5', '')).zfill(5)
            aa = bank_a_aa_map.get(geoid5) or county.get('assessment_area') or county.get('aa_name') or 'Unnamed Assessment Area'
            state_groups[state][aa].append(county)
        
        for state in sorted(state_groups.keys()):
            for aa in sorted(state_groups[state].keys()):
                for county in state_groups[state][aa]:
                    ws.cell(row, 1).value = state
                    # Use CBSA name as Assessment Area name
                    cbsa_code = str(county.get('cbsa_code', ''))
                    cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
                    ws.cell(row, 2).value = cbsa_name  # Assessment Area = CBSA Name
                    ws.cell(row, 3).value = cbsa_name
                    ws.cell(row, 4).value = cbsa_code
                    county_name = county.get('county_name', '')
                    state_name = county.get('state_name', county.get('state', ''))
                    ws.cell(row, 5).value = f"{county_name}, {state_name}" if county_name and state_name else (county_name or state_name)
                    ws.cell(row, 6).value = str(county.get('geoid5', ''))
                    row += 1
    
    # Populate Bank B data
    row = 3
    if counties_b:
        state_groups = defaultdict(lambda: defaultdict(list))
        for county in counties_b:
            state = county.get('state_name', county.get('state', '')) or 'Unknown'
            # Try to get assessment area name from map
            geoid5 = str(county.get('geoid5', '')).zfill(5)
            aa = bank_b_aa_map.get(geoid5) or county.get('assessment_area') or county.get('aa_name') or 'Unnamed Assessment Area'
            state_groups[state][aa].append(county)
        
        for state in sorted(state_groups.keys()):
            for aa in sorted(state_groups[state].keys()):
                for county in state_groups[state][aa]:
                    ws.cell(row, 7).value = state
                    # Use CBSA name as Assessment Area name
                    cbsa_code = str(county.get('cbsa_code', ''))
                    cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
                    ws.cell(row, 8).value = cbsa_name  # Assessment Area = CBSA Name
                    ws.cell(row, 9).value = cbsa_name
                    ws.cell(row, 10).value = cbsa_code
                    county_name = county.get('county_name', '')
                    state_name = county.get('state_name', county.get('state', ''))
                    ws.cell(row, 11).value = f"{county_name}, {state_name}" if county_name and state_name else (county_name or state_name)
                    ws.cell(row, 12).value = str(county.get('geoid5', ''))
                    row += 1
    
    # Auto-adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created Assessment Areas sheet: {len(counties_a) if counties_a else 0} counties for {bank_a_name}, {len(counties_b) if counties_b else 0} counties for {bank_b_name}")


def create_simple_mortgage_sheet(wb, bank_name, subject_data, peer_data,
                                header_fill, header_font, border_thin, denominator_type):
    """Create Mortgage Data sheet as simple table with Grand Total at top."""
    sheet_name = f"{bank_name} Mortgage Data"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed CBSA Code column)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Peer/Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No mortgage data for {bank_name}")
        return
    
    # Metrics in order
    metrics = [
        'Loans', 'LMICT%', 'LMIB%', 'LMIB$', 'MMCT%',
        'MINB%', 'Asian%', 'Black%', 'Native American%', 'HoPI%', 'Hispanic%'
    ]
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    grand_total_agg = subject_data.agg({
        'total_loans': 'sum',
        'lmict_loans': 'sum',
        'lmib_loans': 'sum',
        'lmib_amount': 'sum',
        'mmct_loans': 'sum',
        'minb_loans': 'sum',
        'asian_loans': 'sum',
        'black_loans': 'sum',
        'native_american_loans': 'sum',
        'hopi_loans': 'sum',
        'hispanic_loans': 'sum'
    }).to_dict()
    
    grand_total_loans = grand_total_agg.get('total_loans', 0)
    
    # Calculate peer grand total if available
    peer_grand_total_agg = None
    peer_grand_total_loans = 0
    if peer_data is not None and not peer_data.empty:
        peer_grand_total_agg = peer_data.agg({
            'total_loans': 'sum',
            'lmict_loans': 'sum',
            'lmib_loans': 'sum',
            'lmib_amount': 'sum',
            'mmct_loans': 'sum',
            'minb_loans': 'sum',
            'asian_loans': 'sum',
            'black_loans': 'sum',
            'native_american_loans': 'sum',
            'hopi_loans': 'sum',
            'hispanic_loans': 'sum'
        }).to_dict()
        peer_grand_total_loans = peer_grand_total_agg.get('total_loans', 0)
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'Loans':
            ws.cell(row, 3).value = int(grand_total_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'LMICT%':
            pct = (grand_total_agg.get('lmict_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB%':
            pct = (grand_total_agg.get('lmib_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB$':
            ws.cell(row, 3).value = float(grand_total_agg.get('lmib_amount', 0))
            ws.cell(row, 3).number_format = '$#,##0'
        elif metric == 'MMCT%':
            pct = (grand_total_agg.get('mmct_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric.endswith('%'):
            metric_map = {
                'MINB%': 'minb_loans',
                'Asian%': 'asian_loans',
                'Black%': 'black_loans',
                'Native American%': 'native_american_loans',
                'HoPI%': 'hopi_loans',
                'Hispanic%': 'hispanic_loans'
            }
            field = metric_map.get(metric, '')
            if field:
                pct = (grand_total_agg.get(field, 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
        
        # Peer Grand Total
        if peer_grand_total_agg:
            if metric == 'Loans':
                ws.cell(row, 4).value = int(peer_grand_total_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'LMICT%':
                pct = (peer_grand_total_agg.get('lmict_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB%':
                pct = (peer_grand_total_agg.get('lmib_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB$':
                ws.cell(row, 4).value = float(peer_grand_total_agg.get('lmib_amount', 0))
                ws.cell(row, 4).number_format = '$#,##0'
            elif metric == 'MMCT%':
                pct = (peer_grand_total_agg.get('mmct_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric.endswith('%'):
                metric_map = {
                    'MINB%': 'minb_loans',
                    'Asian%': 'asian_loans',
                    'Black%': 'black_loans',
                    'Native American%': 'native_american_loans',
                    'HoPI%': 'hopi_loans',
                    'Hispanic%': 'hispanic_loans'
                }
                field = metric_map.get(metric, '')
                if field:
                    pct = (peer_grand_total_agg.get(field, 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                    ws.cell(row, 4).value = pct / 100
                    ws.cell(row, 4).number_format = '0.00%'
        
        # Difference for percentages
        if metric.endswith('%') and metric != 'LMIB$':
            ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group by CBSA and aggregate
    grouped = subject_data.groupby('cbsa_code')
    
    # Sort by total loans
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data['total_loans'].sum() if not group_data.empty else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    # Process each CBSA
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Aggregate data - only use total_loans as denominator (no loans_with_demographic_data)
        agg = group_data.agg({
            'total_loans': 'sum',
            'lmict_loans': 'sum',
            'lmib_loans': 'sum',
            'lmib_amount': 'sum',
            'mmct_loans': 'sum',
            'minb_loans': 'sum',
            'asian_loans': 'sum',
            'black_loans': 'sum',
            'native_american_loans': 'sum',
            'hopi_loans': 'sum',
            'hispanic_loans': 'sum'
        }).to_dict()
        
        total_loans = agg.get('total_loans', 0)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        # Always use total_loans as denominator (no loans_with_demographic_data)
        demo_denominator = total_loans
        
        # Write data for each metric
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Subject data (column 3 after removing CBSA Code)
            if metric == 'Loans':
                ws.cell(row, 3).value = int(total_loans)
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'LMICT%':
                pct = (agg.get('lmict_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric == 'LMIB%':
                pct = (agg.get('lmib_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric == 'LMIB$':
                ws.cell(row, 3).value = float(agg.get('lmib_amount', 0))
                ws.cell(row, 3).number_format = '$#,##0'
            elif metric == 'MMCT%':
                pct = (agg.get('mmct_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric.endswith('%'):
                # Race/ethnicity percentages
                metric_map = {
                    'MINB%': 'minb_loans',
                    'Asian%': 'asian_loans',
                    'Black%': 'black_loans',
                    'Native American%': 'native_american_loans',
                    'HoPI%': 'hopi_loans',
                    'Hispanic%': 'hispanic_loans'
                }
                field = metric_map.get(metric, '')
                if field:
                    pct = (agg.get(field, 0) / demo_denominator * 100) if demo_denominator > 0 else 0
                    ws.cell(row, 3).value = pct / 100
                    ws.cell(row, 3).number_format = '0.00%'
            
            # Peer data
            if peer_data is not None and not peer_data.empty:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    peer_agg = peer_group.agg({
                        'total_loans': 'sum',
                        'lmict_loans': 'sum',
                        'lmib_loans': 'sum',
                        'lmib_amount': 'sum',
                        'mmct_loans': 'sum',
                        'minb_loans': 'sum',
                        'asian_loans': 'sum',
                        'black_loans': 'sum',
                        'native_american_loans': 'sum',
                        'hopi_loans': 'sum',
                        'hispanic_loans': 'sum'
                    }).to_dict()
                    
                    peer_total = peer_agg.get('total_loans', 0)
                    
                    if metric == 'Loans':
                        ws.cell(row, 4).value = int(peer_total)
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'LMICT%':
                        pct = (peer_agg.get('lmict_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric == 'LMIB%':
                        pct = (peer_agg.get('lmib_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric == 'LMIB$':
                        ws.cell(row, 4).value = float(peer_agg.get('lmib_amount', 0))
                        ws.cell(row, 4).number_format = '$#,##0'
                    elif metric == 'MMCT%':
                        pct = (peer_agg.get('mmct_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric.endswith('%'):
                        metric_map = {
                            'MINB%': 'minb_loans',
                            'Asian%': 'asian_loans',
                            'Black%': 'black_loans',
                            'Native American%': 'native_american_loans',
                            'HoPI%': 'hopi_loans',
                            'Hispanic%': 'hispanic_loans'
                        }
                        field = metric_map.get(metric, '')
                        if field:
                            pct = (peer_agg.get(field, 0) / peer_total * 100) if peer_total > 0 else 0
                            ws.cell(row, 4).value = pct / 100
                            ws.cell(row, 4).number_format = '0.00%'
            
            # Difference (for percentages only) - columns shifted: C=Subject, D=Peer, E=Difference
            if metric.endswith('%') and metric != 'LMIB$':
                ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust column widths (removed CBSA Code column)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_simple_sb_sheet(wb, bank_name, subject_data, peer_data,
                          header_fill, header_font, border_thin):
    """Create Small Business sheet as simple table with Grand Total at top (no State column)."""
    sheet_name = f"{bank_name} SB Lending"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed State and CBSA Code columns)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Peer/Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No SB data for {bank_name}")
        return
    
    metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount', 'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']
    
    # Handle both column names: sb_loans_total (from merger report queries) and sb_loans_count (from MergerMeter queries)
    sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in subject_data.columns else 'sb_loans_count'
    lmict_count_col = 'lmict_count' if 'lmict_count' in subject_data.columns else 'lmict_loans_count'
    loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in subject_data.columns else 'loans_rev_under_1m'
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    grand_total_sb_loans = subject_data[sb_loans_col].sum() if sb_loans_col in subject_data.columns else 0
    grand_total_lmict = subject_data[lmict_count_col].sum() if lmict_count_col in subject_data.columns else 0
    grand_total_rev_under_1m = subject_data[loans_rev_col].sum() if loans_rev_col in subject_data.columns else 0
    
    # Calculate weighted averages for Grand Total
    if 'lmict_loans_amount' in subject_data.columns:
        grand_total_avg_lmict = subject_data['lmict_loans_amount'].sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    elif 'avg_sb_lmict_loan_amount' in subject_data.columns:
        grand_total_avg_lmict = (subject_data[lmict_count_col] * subject_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    else:
        grand_total_avg_lmict = 0
    
    if 'amount_rev_under_1m' in subject_data.columns:
        grand_total_avg_rev = subject_data['amount_rev_under_1m'].sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    elif 'avg_loan_amt_rum_sb' in subject_data.columns:
        grand_total_avg_rev = (subject_data[loans_rev_col] * subject_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    else:
        grand_total_avg_rev = 0
    
    # Calculate peer grand totals if available
    peer_grand_total_sb_loans = 0
    peer_grand_total_lmict = 0
    peer_grand_total_rev_under_1m = 0
    peer_grand_total_avg_lmict = 0
    peer_grand_total_avg_rev = 0
    if peer_data is not None and not peer_data.empty:
        peer_sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in peer_data.columns else 'sb_loans_count'
        peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_data.columns else 'lmict_loans_count'
        peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_data.columns else 'loans_rev_under_1m'
        
        peer_grand_total_sb_loans = peer_data[peer_sb_loans_col].sum() if peer_sb_loans_col in peer_data.columns else 0
        peer_grand_total_lmict = peer_data[peer_lmict_count_col].sum() if peer_lmict_count_col in peer_data.columns else 0
        peer_grand_total_rev_under_1m = peer_data[peer_loans_rev_col].sum() if peer_loans_rev_col in peer_data.columns else 0
        
        if 'lmict_loans_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = peer_data['lmict_loans_amount'].sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        elif 'avg_sb_lmict_loan_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = (peer_data[peer_lmict_count_col] * peer_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        else:
            peer_grand_total_avg_lmict = 0
        
        if 'amount_rev_under_1m' in peer_data.columns:
            peer_grand_total_avg_rev = peer_data['amount_rev_under_1m'].sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        elif 'avg_loan_amt_rum_sb' in peer_data.columns:
            peer_grand_total_avg_rev = (peer_data[peer_loans_rev_col] * peer_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        else:
            peer_grand_total_avg_rev = 0
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'SB Loans':
            ws.cell(row, 3).value = int(grand_total_sb_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == '#LMICT':
            ws.cell(row, 3).value = int(grand_total_lmict)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg SB LMICT Loan Amount':
            ws.cell(row, 3).value = float(grand_total_avg_lmict)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Loans Rev Under $1m':
            ws.cell(row, 3).value = int(grand_total_rev_under_1m)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg Loan Amt for <$1M GAR SB':
            ws.cell(row, 3).value = float(grand_total_avg_rev)
            ws.cell(row, 3).number_format = '#,##0'
        
        # Peer Grand Total
        if peer_data is not None and not peer_data.empty:
            if metric == 'SB Loans':
                ws.cell(row, 4).value = int(peer_grand_total_sb_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == '#LMICT':
                ws.cell(row, 4).value = int(peer_grand_total_lmict)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg SB LMICT Loan Amount':
                ws.cell(row, 4).value = float(peer_grand_total_avg_lmict)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Loans Rev Under $1m':
                ws.cell(row, 4).value = int(peer_grand_total_rev_under_1m)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg Loan Amt for <$1M GAR SB':
                ws.cell(row, 4).value = float(peer_grand_total_avg_rev)
                ws.cell(row, 4).number_format = '#,##0'
        
        # Difference formulas
        if metric in ['#LMICT', 'Loans Rev Under $1m']:
            sb_loans_row = row - metrics.index(metric)
            ws.cell(row, 5).value = f'=IFERROR((C{row}/C{sb_loans_row})-(D{row}/D{sb_loans_row}),0)'
            ws.cell(row, 5).number_format = '0.00%'
        elif metric in ['Avg SB LMICT Loan Amount', 'Avg Loan Amt for <$1M GAR SB']:
            ws.cell(row, 5).value = f'=IFERROR((C{row}/D{row})-1,0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group and sort by CBSA
    grouped = subject_data.groupby('cbsa_code')
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data[sb_loans_col].sum() if not group_data.empty and sb_loans_col in group_data.columns else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        # Aggregate - handle both column name formats
        # MergerMeter uses: sb_loans_count, lmict_loans_count, loans_rev_under_1m
        # Merger report uses: sb_loans_total, lmict_count, loans_rev_under_1m_count
        sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in group_data.columns else 'sb_loans_count'
        lmict_count_col = 'lmict_count' if 'lmict_count' in group_data.columns else 'lmict_loans_count'
        loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in group_data.columns else 'loans_rev_under_1m'
        
        agg = {
            'sb_loans_total': group_data[sb_loans_col].sum() if sb_loans_col in group_data.columns else 0,
            'lmict_count': group_data[lmict_count_col].sum() if lmict_count_col in group_data.columns else 0,
            'loans_rev_under_1m_count': group_data[loans_rev_col].sum() if loans_rev_col in group_data.columns else 0,
        }
        
        # Weighted averages - handle different column names
        # MergerMeter uses: lmict_loans_count, lmict_loans_amount, loans_rev_under_1m, amount_rev_under_1m
        # Merger report uses: lmict_count, avg_sb_lmict_loan_amount, loans_rev_under_1m_count, avg_loan_amt_rum_sb
        lmict_count_col = 'lmict_count' if 'lmict_count' in group_data.columns else 'lmict_loans_count'
        loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in group_data.columns else 'loans_rev_under_1m'
        
        if agg['lmict_count'] > 0:
            # Check if we have pre-calculated average or need to calculate from amounts
            if 'avg_sb_lmict_loan_amount' in group_data.columns:
                agg['avg_sb_lmict_loan_amount'] = (group_data[lmict_count_col] * group_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / agg['lmict_count']
            elif 'lmict_loans_amount' in group_data.columns:
                agg['avg_sb_lmict_loan_amount'] = group_data['lmict_loans_amount'].sum() / agg['lmict_count'] if agg['lmict_count'] > 0 else 0
            else:
                agg['avg_sb_lmict_loan_amount'] = 0
        else:
            agg['avg_sb_lmict_loan_amount'] = 0
        
        if agg['loans_rev_under_1m_count'] > 0:
            # Check if we have pre-calculated average or need to calculate from amounts
            if 'avg_loan_amt_rum_sb' in group_data.columns:
                agg['avg_loan_amt_rum_sb'] = (group_data[loans_rev_col] * group_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / agg['loans_rev_under_1m_count']
            elif 'amount_rev_under_1m' in group_data.columns:
                agg['avg_loan_amt_rum_sb'] = group_data['amount_rev_under_1m'].sum() / agg['loans_rev_under_1m_count'] if agg['loans_rev_under_1m_count'] > 0 else 0
            else:
                agg['avg_loan_amt_rum_sb'] = 0
        else:
            agg['avg_loan_amt_rum_sb'] = 0
        
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Subject data (column 3 after removing State and CBSA Code)
            if metric == 'SB Loans':
                ws.cell(row, 3).value = int(agg['sb_loans_total'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == '#LMICT':
                # Column C: Loan units (count), not percentage
                ws.cell(row, 3).value = int(agg['lmict_count'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Avg SB LMICT Loan Amount':
                ws.cell(row, 3).value = float(agg['avg_sb_lmict_loan_amount'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Loans Rev Under $1m':
                # Column C: Loan units (count), not percentage
                ws.cell(row, 3).value = int(agg['loans_rev_under_1m_count'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Avg Loan Amt for <$1M GAR SB':
                ws.cell(row, 3).value = float(agg['avg_loan_amt_rum_sb'])
                ws.cell(row, 3).number_format = '#,##0'
            
            # Peer data
            if peer_data is not None and not peer_data.empty:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    # Handle both column name formats for peer data too
                    peer_sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in peer_group.columns else 'sb_loans_count'
                    peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                    peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                    
                    peer_agg = {
                        'sb_loans_total': peer_group[peer_sb_loans_col].sum() if peer_sb_loans_col in peer_group.columns else 0,
                        'lmict_count': peer_group[peer_lmict_count_col].sum() if peer_lmict_count_col in peer_group.columns else 0,
                        'loans_rev_under_1m_count': peer_group[peer_loans_rev_col].sum() if peer_loans_rev_col in peer_group.columns else 0,
                    }
                    
                    # Handle different column names for peer data too
                    peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                    peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                    
                    if peer_agg['lmict_count'] > 0:
                        if 'avg_sb_lmict_loan_amount' in peer_group.columns:
                            peer_agg['avg_sb_lmict_loan_amount'] = (peer_group[peer_lmict_count_col] * peer_group['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_agg['lmict_count']
                        elif 'lmict_loans_amount' in peer_group.columns:
                            peer_agg['avg_sb_lmict_loan_amount'] = peer_group['lmict_loans_amount'].sum() / peer_agg['lmict_count'] if peer_agg['lmict_count'] > 0 else 0
                        else:
                            peer_agg['avg_sb_lmict_loan_amount'] = 0
                    else:
                        peer_agg['avg_sb_lmict_loan_amount'] = 0
                    
                    if peer_agg['loans_rev_under_1m_count'] > 0:
                        if 'avg_loan_amt_rum_sb' in peer_group.columns:
                            peer_agg['avg_loan_amt_rum_sb'] = (peer_group[peer_loans_rev_col] * peer_group['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_agg['loans_rev_under_1m_count']
                        elif 'amount_rev_under_1m' in peer_group.columns:
                            peer_agg['avg_loan_amt_rum_sb'] = peer_group['amount_rev_under_1m'].sum() / peer_agg['loans_rev_under_1m_count'] if peer_agg['loans_rev_under_1m_count'] > 0 else 0
                        else:
                            peer_agg['avg_loan_amt_rum_sb'] = 0
                    else:
                        peer_agg['avg_loan_amt_rum_sb'] = 0
                    
                    if metric == 'SB Loans':
                        ws.cell(row, 4).value = int(peer_agg['sb_loans_total'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == '#LMICT':
                        # Column D: Loan units (count), not percentage
                        ws.cell(row, 4).value = int(peer_agg['lmict_count'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Avg SB LMICT Loan Amount':
                        ws.cell(row, 4).value = float(peer_agg['avg_sb_lmict_loan_amount'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Loans Rev Under $1m':
                        # Column D: Loan units (count), not percentage
                        ws.cell(row, 4).value = int(peer_agg['loans_rev_under_1m_count'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Avg Loan Amt for <$1M GAR SB':
                        ws.cell(row, 4).value = float(peer_agg['avg_loan_amt_rum_sb'])
                        ws.cell(row, 4).number_format = '#,##0'
            
            # Difference formulas (columns shifted: C=Subject, D=Peer, E=Difference)
            # For #LMICT and Loans Rev Under $1m: Columns C and D are in loan units (counts),
            # Column E calculates percentage difference: (Subject% - Peer%) where % = metric/SB Loans
            if metric in ['#LMICT', 'Loans Rev Under $1m']:
                sb_loans_row = row - metrics.index(metric)
                ws.cell(row, 5).value = f'=IFERROR((C{row}/C{sb_loans_row})-(D{row}/D{sb_loans_row}),0)'
                ws.cell(row, 5).number_format = '0.00%'
            elif metric in ['Avg SB LMICT Loan Amount', 'Avg Loan Amt for <$1M GAR SB']:
                ws.cell(row, 5).value = f'=IFERROR((C{row}/D{row})-1,0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust widths (removed CBSA Code column)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_simple_branch_sheet(wb, bank_name, branch_data,
                               header_fill, header_font, border_thin):
    """Create Branch sheet as simple table with Grand Total at top."""
    sheet_name = f"{bank_name} Branches"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed CBSA Code column)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Other/Market', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if branch_data is None or branch_data.empty:
        print(f"    No branch data for {bank_name}")
        return
    
    metrics = ['Branches', 'LMICT', 'MMCT']
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    # Branch query returns: total_branches, branches_in_lmict, pct_lmict, branches_in_mmct, pct_mmct
    # Map to expected column names
    total_branches_col = 'total_branches' if 'total_branches' in branch_data.columns else 'subject_total_branches'
    lmict_col = 'branches_in_lmict' if 'branches_in_lmict' in branch_data.columns else 'subject_pct_lmict'
    pct_lmict_col = 'pct_lmict' if 'pct_lmict' in branch_data.columns else 'subject_pct_lmict'
    mmct_col = 'branches_in_mmct' if 'branches_in_mmct' in branch_data.columns else 'subject_pct_mmct'
    pct_mmct_col = 'pct_mmct' if 'pct_mmct' in branch_data.columns else 'subject_pct_mmct'
    
    grand_total_subject = branch_data[total_branches_col].sum() if total_branches_col in branch_data.columns else 0
    market_total_col = 'market_total_branches' if 'market_total_branches' in branch_data.columns else None
    grand_total_other = branch_data[market_total_col].sum() if market_total_col and market_total_col in branch_data.columns else 0
    
    # Calculate LMICT and MMCT counts for Grand Total
    if grand_total_subject > 0:
        if lmict_col in branch_data.columns:
            # Use count column if available
            subject_lmict_count = branch_data[lmict_col].sum()
            subject_lmict_pct = (subject_lmict_count / grand_total_subject * 100) if grand_total_subject > 0 else 0
        elif pct_lmict_col in branch_data.columns:
            # Use weighted average of percentages
            subject_lmict_pct = (branch_data[pct_lmict_col] * branch_data[total_branches_col]).sum() / grand_total_subject
        else:
            subject_lmict_pct = 0
        
        if mmct_col in branch_data.columns:
            # Use count column if available
            subject_mmct_count = branch_data[mmct_col].sum()
            subject_mmct_pct = (subject_mmct_count / grand_total_subject * 100) if grand_total_subject > 0 else 0
        elif pct_mmct_col in branch_data.columns:
            # Use weighted average of percentages
            subject_mmct_pct = (branch_data[pct_mmct_col] * branch_data[total_branches_col]).sum() / grand_total_subject
        else:
            subject_mmct_pct = 0
    else:
        subject_lmict_pct = 0
        subject_mmct_pct = 0
    
    if grand_total_other > 0:
        if 'market_branches_in_lmict' in branch_data.columns:
            other_lmict_count = branch_data['market_branches_in_lmict'].sum()
            other_lmict_pct = (other_lmict_count / grand_total_other * 100) if grand_total_other > 0 else 0
        elif 'market_pct_lmict' in branch_data.columns:
            other_lmict_pct = (branch_data['market_pct_lmict'] * branch_data[market_total_col]).sum() / grand_total_other
        else:
            other_lmict_pct = 0
        
        if 'market_branches_in_mmct' in branch_data.columns:
            other_mmct_count = branch_data['market_branches_in_mmct'].sum()
            other_mmct_pct = (other_mmct_count / grand_total_other * 100) if grand_total_other > 0 else 0
        elif 'market_pct_mmct' in branch_data.columns:
            other_mmct_pct = (branch_data['market_pct_mmct'] * branch_data[market_total_col]).sum() / grand_total_other
        else:
            other_mmct_pct = 0
    else:
        other_lmict_pct = 0
        other_mmct_pct = 0
    
    grand_total_subject_lmict = int(round((subject_lmict_pct / 100.0) * grand_total_subject))
    grand_total_subject_mmct = int(round((subject_mmct_pct / 100.0) * grand_total_subject))
    grand_total_other_lmict = int(round((other_lmict_pct / 100.0) * grand_total_other))
    grand_total_other_mmct = int(round((other_mmct_pct / 100.0) * grand_total_other))
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        if metric == 'Branches':
            ws.cell(row, 3).value = int(grand_total_subject)
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = int(grand_total_other)
            ws.cell(row, 4).number_format = '0'
        elif metric == 'LMICT':
            ws.cell(row, 3).value = grand_total_subject_lmict
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = grand_total_other_lmict
            ws.cell(row, 4).number_format = '0'
            branches_row = row - 1
            ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
            ws.cell(row, 5).number_format = '0.00%'
        elif metric == 'MMCT':
            ws.cell(row, 3).value = grand_total_subject_mmct
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = grand_total_other_mmct
            ws.cell(row, 4).number_format = '0'
            branches_row = row - 2
            ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group and sort by CBSA
    grouped = branch_data.groupby('cbsa_code')
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data[total_branches_col].sum() if total_branches_col in group_data.columns else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        subject_total = group_data[total_branches_col].sum() if total_branches_col in group_data.columns else 0
        other_total = group_data[market_total_col].sum() if market_total_col and market_total_col in group_data.columns else 0
        
        # Calculate LMICT and MMCT counts for subject
        if lmict_col in group_data.columns:
            subject_lmict_count = int(group_data[lmict_col].sum())
        elif pct_lmict_col in group_data.columns:
            subject_lmict_pct_val = group_data[pct_lmict_col].iloc[0] if not group_data.empty else 0
            subject_lmict_count = int(round((subject_lmict_pct_val / 100.0) * subject_total)) if subject_total > 0 else 0
        else:
            subject_lmict_count = 0
        
        if mmct_col in group_data.columns:
            subject_mmct_count = int(group_data[mmct_col].sum())
        elif pct_mmct_col in group_data.columns:
            subject_mmct_pct_val = group_data[pct_mmct_col].iloc[0] if not group_data.empty else 0
            subject_mmct_count = int(round((subject_mmct_pct_val / 100.0) * subject_total)) if subject_total > 0 else 0
        else:
            subject_mmct_count = 0
        
        # Calculate LMICT and MMCT counts for market
        if 'market_branches_in_lmict' in group_data.columns:
            other_lmict_count = int(group_data['market_branches_in_lmict'].sum())
        elif 'market_pct_lmict' in group_data.columns:
            other_lmict_pct_val = group_data['market_pct_lmict'].iloc[0] if not group_data.empty else 0
            other_lmict_count = int(round((other_lmict_pct_val / 100.0) * other_total)) if other_total > 0 else 0
        else:
            other_lmict_count = 0
        
        if 'market_branches_in_mmct' in group_data.columns:
            other_mmct_count = int(group_data['market_branches_in_mmct'].sum())
        elif 'market_pct_mmct' in group_data.columns:
            other_mmct_pct_val = group_data['market_pct_mmct'].iloc[0] if not group_data.empty else 0
            other_mmct_count = int(round((other_mmct_pct_val / 100.0) * other_total)) if other_total > 0 else 0
        else:
            other_mmct_count = 0
        
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Columns shifted: C=Subject, D=Other, E=Difference
            if metric == 'Branches':
                ws.cell(row, 3).value = int(subject_total)
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = int(other_total)
                ws.cell(row, 4).number_format = '0'
            elif metric == 'LMICT':
                ws.cell(row, 3).value = subject_lmict_count
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = other_lmict_count
                ws.cell(row, 4).number_format = '0'
                # Difference: percentage difference (C=Subject, D=Other)
                branches_row = row - 1
                ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
                ws.cell(row, 5).number_format = '0.00%'
            elif metric == 'MMCT':
                ws.cell(row, 3).value = subject_mmct_count
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = other_mmct_count
                ws.cell(row, 4).number_format = '0'
                # Difference: percentage difference (C=Subject, D=Other)
                branches_row = row - 2
                ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


def create_simple_notes_sheet(wb, bank_a_name, bank_b_name, years_hmda, years_sb):
    """Create simple Notes/Methodology sheet."""
    ws = wb.create_sheet("Notes")
    
    ws['A1'] = "Data Sources and Timeframes"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws.cell(row, 1).value = "1. Mortgage Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from HMDA (Home Mortgage Disclosure Act) data"
    row += 1
    ws.cell(row, 2).value = f"Covers years: {', '.join(years_hmda) if years_hmda else 'N/A'}"
    row += 2
    
    ws.cell(row, 1).value = "2. Small Business Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from Section 1071 Small Business Lending data"
    row += 1
    ws.cell(row, 2).value = f"Covers years: {', '.join(years_sb) if years_sb else 'N/A'}"
    row += 2
    
    ws.cell(row, 1).value = "3. Branch Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from FDIC Summary of Deposits data"
    row += 2
    
    ws.cell(row, 1).value = f"4. Banks Analyzed:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Acquirer: {bank_a_name}"
    row += 1
    ws.cell(row, 2).value = f"Target: {bank_b_name}"
    row += 2
    
    ws.cell(row, 1).value = "Generated:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Auto-adjust width
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    
    print("  Created Notes sheet")


def _add_hhi_sheet(wb: Workbook, hhi_data: pd.DataFrame, bank_a_name: str, bank_b_name: str):
    """Add HHI Analysis sheet to workbook for counties where both banks have branches."""
    
    # Filter to only counties where both banks have branches
    if 'bank_a_deposits' in hhi_data.columns and 'bank_b_deposits' in hhi_data.columns:
        hhi_filtered = hhi_data[
            (hhi_data['bank_a_deposits'] > 0) & 
            (hhi_data['bank_b_deposits'] > 0)
        ].copy()
    else:
        # If columns don't exist, use all data
        hhi_filtered = hhi_data.copy()
    
    if hhi_filtered.empty:
        print("  No counties with overlapping branches for HHI analysis")
        return
    
    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))
    
    # Header style - NCRC blue (#034ea0)
    header_fill = PatternFill(start_color="034EA0", end_color="034EA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Title
    row = 1
    ws.merge_cells(f'A{row}:{get_column_letter(len(hhi_filtered.columns))}{row}')
    cell = ws.cell(row, 1, f"HHI Analysis - Counties with {bank_a_name} and {bank_b_name} Branches")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")
    row += 2
    
    # Headers
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        cell = ws.cell(row, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row += 1
    
    # Data rows
    for df_row_idx, df_row in hhi_filtered.iterrows():
        for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
            value = df_row[col_name]
            cell = ws.cell(row, col_idx, value)
            
            # Format numbers
            if pd.api.types.is_numeric_dtype(hhi_filtered[col_name]):
                if 'hhi' in col_name.lower():
                    cell.number_format = '#,##0'  # HHI is typically an integer
                elif 'deposits' in col_name.lower() or 'share' in col_name.lower():
                    if 'share' in col_name.lower() or 'percent' in col_name.lower():
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
            
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Adjust column widths
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        col_letter = get_column_letter(col_idx)
        try:
            col_max_len = hhi_filtered[col_name].astype(str).str.len().max() if not hhi_filtered[col_name].empty else 10
        except (ValueError, AttributeError):
            col_max_len = 10
        max_length = max(len(str(col_name)), col_max_len)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    print("  Added HHI Analysis sheet")
