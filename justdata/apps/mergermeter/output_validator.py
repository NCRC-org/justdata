"""
MergerMeter Output Validator

Inspects a generated Excel workbook for data quality issues
before delivery to the user. Adds a warnings sheet if issues are found.
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def validate_workbook(wb: Workbook) -> list:
    """
    Validate a MergerMeter Excel workbook for data quality issues.

    Args:
        wb: openpyxl Workbook object (already generated)

    Returns:
        List of warning dicts: [{'sheet': str, 'row': int, 'issue': str, 'severity': str}]
        Empty list = no issues found.
    """
    warnings = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 2:
            continue

        # === Check 1: Raw CBSA codes ===
        seen_cbsa_codes = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=False):
            for cell in row:
                if (cell.value and isinstance(cell.value, str)
                    and re.match(r'^CBSA \d{4,5}$', cell.value.strip())):
                    code = cell.value.strip()
                    if code not in seen_cbsa_codes:
                        seen_cbsa_codes.add(code)
                        warnings.append({
                            'sheet': sheet_name,
                            'row': cell.row,
                            'issue': f'Unresolved CBSA code "{code}" — metro area name missing',
                            'severity': 'warning'
                        })

        # === Check 2: Missing peer data in data sheets ===
        if 'MORTGAGE DATA' in sheet_name or 'SB DATA' in sheet_name:
            areas_with_no_peers = []
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
                if len(row) < 4:
                    continue
                area_cell = row[0]
                metric_cell = row[1]
                bank_cell = row[2]
                peer_cell = row[3]

                if (metric_cell.value == 'Loans'
                    and area_cell.value
                    and area_cell.value != 'Grand Total'):

                    has_bank = (bank_cell.value is not None
                               and bank_cell.value != 0
                               and bank_cell.value != '')
                    has_peer = (peer_cell.value is not None
                               and peer_cell.value != 0
                               and peer_cell.value != '')

                    if has_bank and not has_peer:
                        areas_with_no_peers.append(area_cell.value)
                        warnings.append({
                            'sheet': sheet_name,
                            'row': area_cell.row,
                            'issue': f'No peer data for {area_cell.value} '
                                     f'(bank has {bank_cell.value} loans)',
                            'severity': 'critical'
                        })

            if len(areas_with_no_peers) > 3:
                warnings.append({
                    'sheet': sheet_name,
                    'row': 0,
                    'issue': f'{len(areas_with_no_peers)} assessment areas have no peer data. '
                             f'Peer comparison is incomplete for this bank.',
                    'severity': 'critical'
                })

        # === Check 3: Completely empty data sheets ===
        if ('MORTGAGE DATA' in sheet_name or 'SB DATA' in sheet_name
            or 'BRANCH DATA' in sheet_name):
            data_rows = ws.max_row - 2
            if data_rows <= 0:
                warnings.append({
                    'sheet': sheet_name,
                    'row': 0,
                    'issue': f'Sheet is completely empty — no data rows found',
                    'severity': 'critical'
                })

        # === Check 4: Year coverage gaps ===
        if 'MORTGAGE DATA' in sheet_name or 'SB DATA' in sheet_name:
            for row in ws.iter_rows(min_row=3, max_row=min(15, ws.max_row), values_only=False):
                if len(row) >= 7 and row[0].value == 'Grand Total' and row[1].value == 'Loans':
                    yr1_bank = row[2].value if len(row) > 2 else None
                    yr2_bank = row[5].value if len(row) > 5 else None

                    if yr1_bank in (None, 0, '') and yr2_bank not in (None, 0, ''):
                        warnings.append({
                            'sheet': sheet_name,
                            'row': row[0].row,
                            'issue': 'First year has zero data but second year has data — '
                                     'possible data gap or identifier mismatch',
                            'severity': 'warning'
                        })
                    elif yr2_bank in (None, 0, '') and yr1_bank not in (None, 0, ''):
                        warnings.append({
                            'sheet': sheet_name,
                            'row': row[0].row,
                            'issue': 'Second year has zero data but first year has data — '
                                     'possible data gap',
                            'severity': 'warning'
                        })
                    break

    return warnings


def add_warnings_sheet(wb: Workbook, warnings: list):
    """
    Add a Validation Warnings sheet to the workbook at position 1 (after Notes).
    """
    if not warnings:
        return

    ws = wb.create_sheet("Validation Warnings", 1)

    header_font = Font(bold=True, size=14, color="CC0000")
    subheader_font = Font(size=11, italic=True, color="666666")
    col_header_font = Font(bold=True, size=11)
    col_header_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    critical_font = Font(color="CC0000", bold=True)
    warning_font = Font(color="856404")
    thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))

    critical_count = sum(1 for w in warnings if w.get('severity') == 'critical')

    ws.cell(row=1, column=1, value="MergerMeter Data Quality Report").font = header_font
    ws.merge_cells('A1:D1')

    summary_text = f"{len(warnings)} issue(s) found"
    if critical_count > 0:
        summary_text += f" ({critical_count} critical)"
    summary_text += ". Review before using this data for CBA negotiations."
    ws.cell(row=2, column=1, value=summary_text).font = subheader_font
    ws.merge_cells('A2:D2')

    headers = ['#', 'Severity', 'Sheet', 'Issue']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill

    for i, warning in enumerate(warnings, 1):
        row_num = 4 + i
        ws.cell(row=row_num, column=1, value=i)

        severity = warning.get('severity', 'warning').upper()
        sev_cell = ws.cell(row=row_num, column=2, value=severity)
        sev_cell.font = critical_font if severity == 'CRITICAL' else warning_font

        ws.cell(row=row_num, column=3, value=warning.get('sheet', ''))
        ws.cell(row=row_num, column=4, value=warning.get('issue', ''))

        for col in range(1, 5):
            ws.cell(row=row_num, column=col).border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 80
