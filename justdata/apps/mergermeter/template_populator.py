"""
Template population functions for MergerMeter Excel reports.

This module handles populating the Excel template with data, following the exact
structure: Grand Total at row 2, then CBSA sections with merged cells in column A,
and formulas copied down as data is added.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Alignment, PatternFill
from pathlib import Path
import pandas as pd
from typing import Dict, Optional, List, Tuple
import re
from justdata.apps.mergermeter.config import PROJECT_ID


def populate_template_from_file(
    template_path: Path,
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame] = None,
    bank_a_hmda_peer: Optional[pd.DataFrame] = None,
    bank_b_hmda_subject: Optional[pd.DataFrame] = None,
    bank_b_hmda_peer: Optional[pd.DataFrame] = None,
    bank_a_sb_subject: Optional[pd.DataFrame] = None,
    bank_a_sb_peer: Optional[pd.DataFrame] = None,
    bank_b_sb_subject: Optional[pd.DataFrame] = None,
    bank_b_sb_peer: Optional[pd.DataFrame] = None,
    bank_a_branch: Optional[pd.DataFrame] = None,
    bank_b_branch: Optional[pd.DataFrame] = None,
    hhi_data: Optional[pd.DataFrame] = None,
    assessment_areas: Optional[Dict] = None,
    metadata: Optional[Dict] = None
):
    """
    Load template and populate with data following the template structure.
    
    Structure:
    - Row 1: Headers
    - Row 2: Grand Total (sum of all assessment areas)
    - Row 3+: Assessment area sections with CBSA names in column A (merged cells)
    - Formulas are copied down as data is added
    """
    print("Loading template and populating data...")
    wb = load_workbook(template_path, data_only=False)
    
    # Remove ALL cell colors from ALL sheets first - be very aggressive
    no_fill = PatternFill(fill_type=None)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Remove all fills from all cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill:
                    cell.fill = no_fill
        # Also clear any conditional formatting that might add colors
        if hasattr(ws, 'conditional_formatting'):
            ws.conditional_formatting = {}
    
    # Rename sheets to match bank names
    _rename_sheets(wb, bank_a_name, bank_b_name)
    
    # Populate each sheet type
    if bank_a_hmda_subject is not None or bank_a_hmda_peer is not None:
        _populate_mortgage_sheet(
            wb, f"{bank_a_name} Mortgage Data (%)", 
            bank_a_hmda_subject, bank_a_hmda_peer, assessment_areas
        )
    
    if bank_b_hmda_subject is not None or bank_b_hmda_peer is not None:
        _populate_mortgage_sheet(
            wb, f"{bank_b_name} Mortgage Data (%)",
            bank_b_hmda_subject, bank_b_hmda_peer, assessment_areas
        )
    
    if bank_a_sb_subject is not None or bank_a_sb_peer is not None:
        _populate_sb_sheet(
            wb, f"Banks A SB Lending",
            bank_a_sb_subject, bank_a_sb_peer, assessment_areas, bank_a_name
        )
    
    if bank_b_sb_subject is not None or bank_b_sb_peer is not None:
        _populate_sb_sheet(
            wb, f"Bank B SB Lending",
            bank_b_sb_subject, bank_b_sb_peer, assessment_areas, bank_b_name
        )
    
    if bank_a_branch is not None:
        _populate_branch_sheet(
            wb, f"{bank_a_name} Branches",
            bank_a_branch, assessment_areas
        )
    
    if bank_b_branch is not None:
        _populate_branch_sheet(
            wb, f"{bank_b_name} Branches",
            bank_b_branch, assessment_areas
        )
    
    if assessment_areas:
        _populate_assessment_areas_sheet(wb, assessment_areas, bank_a_name, bank_b_name)
    
    if hhi_data is not None and not hhi_data.empty:
        _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
    
    # Delete and recreate Notes sheet with all metadata
    _populate_notes_sheet(wb, bank_a_name, bank_b_name, metadata, assessment_areas)

    # FORMATTING PASS: Apply number formats to all data sheets
    _apply_number_formatting(wb, bank_a_name, bank_b_name)

    # FINAL PASS: Remove ALL cell colors one more time before saving
    no_fill = PatternFill(fill_type=None)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill:
                    cell.fill = no_fill
    
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


def _rename_sheets(wb, bank_a_name: str, bank_b_name: str):
    """Rename template sheets to match actual bank names."""
    for sheet_name in list(wb.sheetnames):
        new_name = sheet_name
        if 'PNC Bank' in sheet_name or 'PNC BANK' in sheet_name:
            new_name = sheet_name.replace('PNC Bank', bank_a_name).replace('PNC BANK', bank_a_name)
        elif 'FirstBank' in sheet_name or 'First Bank' in sheet_name:
            new_name = sheet_name.replace('FirstBank', bank_b_name).replace('First Bank', bank_b_name)
        elif 'Bank A' in sheet_name:
            new_name = sheet_name.replace('Bank A', bank_a_name)
        elif 'Bank B' in sheet_name:
            new_name = sheet_name.replace('Bank B', bank_b_name)
        elif 'Atlantic Union' in sheet_name:
            new_name = sheet_name.replace('Atlantic Union', bank_a_name)
        elif 'Sandy Spring' in sheet_name:
            new_name = sheet_name.replace('Sandy Spring', bank_b_name)
        
        if new_name != sheet_name:
            try:
                wb[sheet_name].title = new_name
            except:
                pass  # Sheet name might already exist


def _populate_mortgage_sheet(
    wb, sheet_name: str,
    subject_data: Optional[pd.DataFrame],
    peer_data: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict]
):
    """Populate mortgage data sheet starting at row 2 with Grand Total, then CBSA sections."""
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping...")
        return
    
    ws = wb[sheet_name]
    print(f"  Populating {sheet_name}...")
    
    # Row 1: Headers (already in template)
    # Row 2: Grand Total - populate with sum of all data
    # Row 3+: Assessment area sections
    
    # Aggregate data across all years and CBSAs for Grand Total
    if subject_data is not None and not subject_data.empty:
        # Sum across all rows (data is already aggregated by year and CBSA)
        grand_total_loans = subject_data.get('total_loans', pd.Series()).sum() if 'total_loans' in subject_data.columns else 0
        grand_lmict_loans = subject_data.get('lmict_loans', pd.Series()).sum() if 'lmict_loans' in subject_data.columns else 0
        grand_lmib_loans = subject_data.get('lmib_loans', pd.Series()).sum() if 'lmib_loans' in subject_data.columns else 0
        grand_lmib_amount = subject_data.get('lmib_amount', pd.Series()).sum() if 'lmib_amount' in subject_data.columns else 0
        grand_mmct_loans = subject_data.get('mmct_loans', pd.Series()).sum() if 'mmct_loans' in subject_data.columns else 0
        grand_minb_loans = subject_data.get('minb_loans', pd.Series()).sum() if 'minb_loans' in subject_data.columns else 0
    
    # Find where to start inserting data (after Grand Total section)
    # Template has Grand Total at row 2, then metrics rows 3-12, then Assessment Area 1 at row 13
    # We'll insert data starting at row 2 for Grand Total, then add CBSA sections
    
    # Group data by CBSA
    if subject_data is not None and not subject_data.empty:
        cbsa_groups = _group_by_cbsa(subject_data, assessment_areas)
        
        # Populate Grand Total at row 2
        _populate_mortgage_grand_total(ws, subject_data, peer_data, row=2)
        
        # Populate each CBSA section
        current_row = 13  # Start after Grand Total section (rows 2-12)
        print(f"  Found {len(cbsa_groups)} CBSAs to populate")
        for cbsa_name, cbsa_data in cbsa_groups.items():
            print(f"    Populating CBSA: {cbsa_name} at row {current_row}")
            current_row = _populate_mortgage_cbsa_section(
                ws, cbsa_name, cbsa_data, peer_data, current_row
            )
            current_row += 1  # Add spacing between sections


def _populate_mortgage_grand_total(
    ws, subject_data: pd.DataFrame, peer_data: Optional[pd.DataFrame], row: int
):
    """Populate Grand Total row (row 2) with aggregated data."""
    # Column A: "Grand Total" (already in template)
    # Column B: "Loans" label (already in template)
    # Column C: Subject bank total loans
    # Column D: Peer bank total loans
    # Column E: Difference formula (already in template, will reference C2-D2)
    
    if subject_data is not None and not subject_data.empty:
        total_loans = subject_data.get('total_loans', pd.Series()).sum() if 'total_loans' in subject_data.columns else 0
        _safe_set_cell_value(ws, row, 3, int(total_loans))  # Column C
    
    if peer_data is not None and not peer_data.empty:
        peer_total = peer_data.get('total_loans', pd.Series()).sum() if 'total_loans' in peer_data.columns else 0
        _safe_set_cell_value(ws, row, 4, int(peer_total))  # Column D
    
    # Populate metrics rows (3-12) with percentages and amounts
    # All percentages use total_loans as denominator
    if subject_data is not None and not subject_data.empty:
        total_loans = subject_data.get('total_loans', pd.Series()).sum()
        
        # Row 3: LMICT%
        lmict_loans = subject_data.get('lmict_loans', pd.Series()).sum()
        lmict_pct = (lmict_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 1, 3, lmict_pct)
        
        # Row 4: LMIB%
        lmib_loans = subject_data.get('lmib_loans', pd.Series()).sum()
        lmib_pct = (lmib_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 2, 3, lmib_pct)
        
        # Row 5: LMIB$ (amount, not percentage)
        lmib_amount = subject_data.get('lmib_amount', pd.Series()).sum()
        _safe_set_cell_value(ws, row + 3, 3, lmib_amount)
        
        # Row 6: MMCT%
        mmct_loans = subject_data.get('mmct_loans', pd.Series()).sum()
        mmct_pct = (mmct_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 4, 3, mmct_pct)
        
        # Row 7: MINB%
        minb_loans = subject_data.get('minb_loans', pd.Series()).sum()
        minb_pct = (minb_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 5, 3, minb_pct)
        
        # Row 8: Asian% (using total loans as denominator)
        asian_loans = subject_data.get('asian_loans', pd.Series()).sum()
        asian_pct = (asian_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 6, 3, asian_pct)
        
        # Row 9: Black% (using total loans as denominator)
        black_loans = subject_data.get('black_loans', pd.Series()).sum()
        black_pct = (black_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 7, 3, black_pct)
        
        # Row 10: Native American% (using total loans as denominator)
        native_american_loans = subject_data.get('native_american_loans', pd.Series()).sum()
        native_american_pct = (native_american_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 8, 3, native_american_pct)
        
        # Row 11: HoPI% (using total loans as denominator)
        hopi_loans = subject_data.get('hopi_loans', pd.Series()).sum()
        hopi_pct = (hopi_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 9, 3, hopi_pct)
        
        # Row 12: Hispanic% (using total loans as denominator)
        hispanic_loans = subject_data.get('hispanic_loans', pd.Series()).sum()
        hispanic_pct = (hispanic_loans / total_loans * 100) if total_loans > 0 else 0
        _safe_set_cell_value(ws, row + 10, 3, hispanic_pct)
    
    # Populate peer data (Column D)
    if peer_data is not None and not peer_data.empty:
        peer_total = peer_data.get('total_loans', pd.Series()).sum()
        
        # Row 3: LMICT%
        peer_lmict = peer_data.get('lmict_loans', pd.Series()).sum()
        peer_lmict_pct = (peer_lmict / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 1, 4, peer_lmict_pct)
        
        # Row 4: LMIB%
        peer_lmib = peer_data.get('lmib_loans', pd.Series()).sum()
        peer_lmib_pct = (peer_lmib / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 2, 4, peer_lmib_pct)
        
        # Row 5: LMIB$
        peer_lmib_amount = peer_data.get('lmib_amount', pd.Series()).sum()
        _safe_set_cell_value(ws, row + 3, 4, peer_lmib_amount)
        
        # Row 6: MMCT%
        peer_mmct = peer_data.get('mmct_loans', pd.Series()).sum()
        peer_mmct_pct = (peer_mmct / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 4, 4, peer_mmct_pct)
        
        # Row 7: MINB%
        peer_minb = peer_data.get('minb_loans', pd.Series()).sum()
        peer_minb_pct = (peer_minb / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 5, 4, peer_minb_pct)
        
        # Row 8: Asian%
        peer_asian = peer_data.get('asian_loans', pd.Series()).sum()
        peer_asian_pct = (peer_asian / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 6, 4, peer_asian_pct)
        
        # Row 9: Black%
        peer_black = peer_data.get('black_loans', pd.Series()).sum()
        peer_black_pct = (peer_black / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 7, 4, peer_black_pct)
        
        # Row 10: Native American%
        peer_native_american = peer_data.get('native_american_loans', pd.Series()).sum()
        peer_native_american_pct = (peer_native_american / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 8, 4, peer_native_american_pct)
        
        # Row 11: HoPI%
        peer_hopi = peer_data.get('hopi_loans', pd.Series()).sum()
        peer_hopi_pct = (peer_hopi / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 9, 4, peer_hopi_pct)
        
        # Row 12: Hispanic%
        peer_hispanic = peer_data.get('hispanic_loans', pd.Series()).sum()
        peer_hispanic_pct = (peer_hispanic / peer_total * 100) if peer_total > 0 else 0
        _safe_set_cell_value(ws, row + 10, 4, peer_hispanic_pct)
    
    # Formulas in column E are already in template and will calculate differences


def _populate_mortgage_cbsa_section(
    ws, cbsa_name: str, cbsa_data: pd.DataFrame,
    peer_data: Optional[pd.DataFrame], start_row: int
) -> int:
    """
    Populate a CBSA section in mortgage sheet.
    Returns the next available row after this section.
    """
    # Row 1: CBSA name in column A (merged across multiple rows)
    # Row 2: "Loans" label in column B, total in column C
    # Rows 3-12: Metrics with formulas
    
    # Insert CBSA name in column A (first row only, no merging)
    section_end_row = start_row + 11  # 12 rows per section (1 header + 11 metric rows)
    
    # Set CBSA name in first row only
    _safe_set_cell_value(ws, start_row, 1, cbsa_name)
    
    # Row 2: "Loans" label and totals
    loans_row = start_row + 1
    _safe_set_cell_value(ws, loans_row, 2, "Loans")  # Column B: "Loans" label
    
    # Fill down all metric labels in column B from template (rows 3-12)
    # Get metric labels from template Grand Total section (rows 3-12)
    template_metric_labels = []
    for template_row in range(3, 13):  # Rows 3-12 in template
        label_cell = ws.cell(template_row, 2)  # Column B
        if label_cell.value:
            template_metric_labels.append(label_cell.value)
    
    # Fill down metric labels for this CBSA section
    for i, label in enumerate(template_metric_labels):
        metric_row = loans_row + 1 + i
        if metric_row <= section_end_row:
            _safe_set_cell_value(ws, metric_row, 2, label)
    
    # Calculate totals for this CBSA (aggregate across years if multiple)
    if not cbsa_data.empty:
        cbsa_total = cbsa_data.get('total_loans', pd.Series()).sum() if 'total_loans' in cbsa_data.columns else 0
        _safe_set_cell_value(ws, loans_row, 3, int(cbsa_total))  # Column C: Subject bank
        
        # Populate metrics for this CBSA (rows 3-12)
        # Populate metrics for this CBSA (rows 3-12) - use safe_set_cell_value
        # Row 3: LMICT%
        lmict_loans = cbsa_data.get('lmict_loans', pd.Series()).sum()
        lmict_pct = (lmict_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 1, 3, lmict_pct)
        
        # Row 4: LMIB%
        lmib_loans = cbsa_data.get('lmib_loans', pd.Series()).sum()
        lmib_pct = (lmib_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 2, 3, lmib_pct)
        
        # Row 5: LMIB$
        lmib_amount = cbsa_data.get('lmib_amount', pd.Series()).sum()
        _safe_set_cell_value(ws, loans_row + 3, 3, lmib_amount)
        
        # Row 6: MMCT%
        mmct_loans = cbsa_data.get('mmct_loans', pd.Series()).sum()
        mmct_pct = (mmct_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 4, 3, mmct_pct)
        
        # Row 7: MINB%
        minb_loans = cbsa_data.get('minb_loans', pd.Series()).sum()
        minb_pct = (minb_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 5, 3, minb_pct)
        
        # Row 8: Asian% (using total loans as denominator)
        asian_loans = cbsa_data.get('asian_loans', pd.Series()).sum()
        asian_pct = (asian_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 6, 3, asian_pct)
        
        # Row 9: Black% (using total loans as denominator)
        black_loans = cbsa_data.get('black_loans', pd.Series()).sum()
        black_pct = (black_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 7, 3, black_pct)
        
        # Row 10: Native American% (using total loans as denominator)
        native_american_loans = cbsa_data.get('native_american_loans', pd.Series()).sum()
        native_american_pct = (native_american_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 8, 3, native_american_pct)
        
        # Row 11: HoPI% (using total loans as denominator)
        hopi_loans = cbsa_data.get('hopi_loans', pd.Series()).sum()
        hopi_pct = (hopi_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 9, 3, hopi_pct)
        
        # Row 12: Hispanic% (using total loans as denominator)
        hispanic_loans = cbsa_data.get('hispanic_loans', pd.Series()).sum()
        hispanic_pct = (hispanic_loans / cbsa_total * 100) if cbsa_total > 0 else 0
        _safe_set_cell_value(ws, loans_row + 10, 3, hispanic_pct)
    
    # Get peer data for this CBSA
    if peer_data is not None and not peer_data.empty:
        # Match CBSA in peer data
        peer_cbsa_data = _match_cbsa_in_data(peer_data, cbsa_name)
        if not peer_cbsa_data.empty:
            peer_total = peer_cbsa_data.get('total_loans', pd.Series()).sum()
            _safe_set_cell_value(ws, loans_row, 4, int(peer_total))  # Column D: Peer bank
            
            # Populate peer metrics - use safe_set_cell_value
            peer_lmict = peer_cbsa_data.get('lmict_loans', pd.Series()).sum()
            peer_lmict_pct = (peer_lmict / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 1, 4, peer_lmict_pct)
            
            peer_lmib = peer_cbsa_data.get('lmib_loans', pd.Series()).sum()
            peer_lmib_pct = (peer_lmib / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 2, 4, peer_lmib_pct)
            
            peer_lmib_amount = peer_cbsa_data.get('lmib_amount', pd.Series()).sum()
            _safe_set_cell_value(ws, loans_row + 3, 4, peer_lmib_amount)
            
            peer_mmct = peer_cbsa_data.get('mmct_loans', pd.Series()).sum()
            peer_mmct_pct = (peer_mmct / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 4, 4, peer_mmct_pct)
            
            peer_minb = peer_cbsa_data.get('minb_loans', pd.Series()).sum()
            peer_minb_pct = (peer_minb / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 5, 4, peer_minb_pct)
            
            peer_asian = peer_cbsa_data.get('asian_loans', pd.Series()).sum()
            peer_asian_pct = (peer_asian / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 6, 4, peer_asian_pct)
            
            peer_black = peer_cbsa_data.get('black_loans', pd.Series()).sum()
            peer_black_pct = (peer_black / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 7, 4, peer_black_pct)
            
            peer_native_american = peer_cbsa_data.get('native_american_loans', pd.Series()).sum()
            peer_native_american_pct = (peer_native_american / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 8, 4, peer_native_american_pct)
            
            peer_hopi = peer_cbsa_data.get('hopi_loans', pd.Series()).sum()
            peer_hopi_pct = (peer_hopi / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 9, 4, peer_hopi_pct)
            
            peer_hispanic = peer_cbsa_data.get('hispanic_loans', pd.Series()).sum()
            peer_hispanic_pct = (peer_hispanic / peer_total * 100) if peer_total > 0 else 0
            _safe_set_cell_value(ws, loans_row + 10, 4, peer_hispanic_pct)
    
    # Set formulas for ALL rows including "Loans" row (Column E: Difference)
    # Use simple difference formula to avoid circular references: =C{row}-D{row}
    # Set formula for "Loans" row first
    loans_formula = f"=IFERROR(C{loans_row}-D{loans_row},\"\")"
    _safe_set_cell_value(ws, loans_row, 5, loans_formula)
    
    # Set formulas for all metric rows (rows 3-12, which is loans_row+1 through loans_row+10)
    metrics_start = loans_row + 1
    for metric_row in range(metrics_start, section_end_row + 1):
        new_formula = f"=IFERROR(C{metric_row}-D{metric_row},\"\")"
        _safe_set_cell_value(ws, metric_row, 5, new_formula)
    
    return section_end_row


def _populate_sb_sheet(
    wb, sheet_name: str,
    subject_data: Optional[pd.DataFrame],
    peer_data: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict],
    bank_name: str
):
    """Populate small business lending sheet with state/county/CBSA structure."""
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping...")
        return
    
    ws = wb[sheet_name]
    print(f"  Populating {sheet_name}...")
    
    # Structure:
    # Row 1: Headers (State | County | CBSA | Data columns)
    # Row 2: Grand Total
    # Row 3+: State/Assessment Area sections
    
    # Populate Grand Total at row 2
    if subject_data is not None and not subject_data.empty:
        _populate_sb_grand_total(ws, subject_data, peer_data, row=2)
    
    # Group by state/assessment area
    if subject_data is not None and not subject_data.empty:
        state_groups = _group_sb_by_state(subject_data, assessment_areas)
        
        current_row = 7  # Start after Grand Total section
        for state_name, state_data in state_groups.items():
            current_row = _populate_sb_state_section(
                ws, state_name, state_data, peer_data, current_row
            )
            current_row += 1


def _populate_sb_grand_total(ws, subject_data: pd.DataFrame, peer_data: Optional[pd.DataFrame], row: int):
    """Populate Grand Total for SB lending."""
    # Column A: "Grand Total"
    # Column C: "SB Loans" label
    # Column D: Subject bank total
    # Column E: Peer bank total
    # Column F: Difference formula
    
    _safe_set_cell_value(ws, row, 1, "Grand Total")
    _safe_set_cell_value(ws, row, 3, "SB Loans")
    
    if not subject_data.empty:
        total_loans = subject_data.get('loan_count', pd.Series()).sum() if 'loan_count' in subject_data.columns else 0
        _safe_set_cell_value(ws, row, 4, total_loans)
    
    if peer_data is not None and not peer_data.empty:
        peer_total = peer_data.get('loan_count', pd.Series()).sum() if 'loan_count' in peer_data.columns else 0
        _safe_set_cell_value(ws, row, 5, peer_total)


def _populate_sb_state_section(
    ws, state_name: str, state_data: pd.DataFrame,
    peer_data: Optional[pd.DataFrame], start_row: int
) -> int:
    """Populate a state/assessment area section in SB sheet."""
    # Column A: State name (merged)
    # Column B: Assessment Area name (merged)
    # Column C: Metric labels
    # Column D: Subject bank data
    # Column E: Peer bank data
    # Column F: Difference formulas
    
    # Group by assessment area within state
    aa_groups = _group_sb_by_assessment_area(state_data)
    
    section_end_row = start_row
    for aa_name, aa_data in aa_groups.items():
        # Assessment area header row
        aa_row = section_end_row + 1
        _safe_set_cell_value(ws, aa_row, 1, state_name)
        _safe_set_cell_value(ws, aa_row, 2, aa_name)
        _safe_set_cell_value(ws, aa_row, 3, "SB Loans")
        
        # Calculate totals
        if not aa_data.empty:
            aa_total = aa_data.get('loan_count', pd.Series()).sum() if 'loan_count' in aa_data.columns else 0
            _safe_set_cell_value(ws, aa_row, 4, aa_total)
        
        # Peer data
        if peer_data is not None and not peer_data.empty:
            peer_aa_data = _match_cbsa_in_data(peer_data, aa_name)
            if not peer_aa_data.empty:
                peer_total = peer_aa_data.get('loan_count', pd.Series()).sum() if 'loan_count' in peer_aa_data.columns else 0
                _safe_set_cell_value(ws, aa_row, 5, peer_total)
        
        # Metrics rows with formulas - use simple difference to avoid circular references
        metrics_start = aa_row + 1
        for i in range(5):  # 5 metric rows
            metric_row = metrics_start + i
            # Use simple difference formula to avoid circular references
            new_formula = f"=IFERROR(D{metric_row}-E{metric_row},\"\")"
            _safe_set_cell_value(ws, metric_row, 6, new_formula)
        
        section_end_row = metrics_start + 4
    
    # Set state name in first row only (no merging)
    if section_end_row > start_row:
        _safe_set_cell_value(ws, start_row + 1, 1, state_name)
    
    return section_end_row


def _populate_branch_sheet(
    wb, sheet_name: str,
    branch_data: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict]
):
    """Populate branch sheet with assessment area names (CBSA names), no state columns."""
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping...")
        return
    
    ws = wb[sheet_name]
    print(f"  Populating {sheet_name}...")
    
    # Structure:
    # Row 1: Headers
    # Row 2: Grand Total
    # Row 3+: Assessment area sections (CBSA names in column A, merged)
    
    if branch_data is not None and not branch_data.empty:
        # Populate Grand Total
        _populate_branch_grand_total(ws, branch_data, row=2)
        
        # Group by CBSA
        cbsa_groups = _group_by_cbsa(branch_data, assessment_areas)
        
        current_row = 5  # Start after Grand Total section
        for cbsa_name, cbsa_branch_data in cbsa_groups.items():
            current_row = _populate_branch_cbsa_section(
                ws, cbsa_name, cbsa_branch_data, current_row
            )
            current_row += 1


def _populate_branch_grand_total(ws, branch_data: pd.DataFrame, row: int):
    """Populate Grand Total for branch data."""
    _safe_set_cell_value(ws, row, 1, "Grand Total")
    _safe_set_cell_value(ws, row, 2, "Branches")
    
    total_branches = branch_data.get('total_branches', pd.Series()).sum() if 'total_branches' in branch_data.columns else len(branch_data)
    _safe_set_cell_value(ws, row, 3, int(total_branches))


def _populate_branch_cbsa_section(
    ws, cbsa_name: str, cbsa_data: pd.DataFrame, start_row: int
) -> int:
    """Populate a CBSA section in branch sheet."""
    # Column A: CBSA name (merged)
    # Column B: "Branches" label
    # Column C: Branch count
    # Column D: Peer data
    # Column E: Difference formulas
    
    section_end_row = start_row + 2  # 3 rows per section
    
    # Set CBSA name in first row only (no merging)
    _safe_set_cell_value(ws, start_row, 1, cbsa_name)
    
    _safe_set_cell_value(ws, start_row, 2, "Branches")
    
    if not cbsa_data.empty:
        branch_count = cbsa_data.get('total_branches', pd.Series()).sum() if 'total_branches' in cbsa_data.columns else len(cbsa_data)
        _safe_set_cell_value(ws, start_row, 3, int(branch_count))  # Whole number
    
    # Copy formulas - use simple difference to avoid circular references
    new_formula = f"=IFERROR(C{start_row}-D{start_row},\"\")"
    _safe_set_cell_value(ws, start_row, 5, new_formula)
    
    return section_end_row


def _populate_assessment_areas_sheet(
    wb, assessment_areas: Dict, bank_a_name: str, bank_b_name: str
):
    """Populate Assessment Areas sheet with CBSA names and county lists, removing cell colors."""
    sheet_name = "Assessment Areas"
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping...")
        return
    
    ws = wb[sheet_name]
    print(f"  Populating {sheet_name}...")
    
    # Remove all cell colors from the sheet
    no_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type:
                cell.fill = no_fill
    
    # Structure (based on template):
    # Row 1: Headers - "@PNC AA | County State | | @FirstBank AA | County State |"
    # Row 2: "CBSA Name | County A | | CBSA Name | County A |"
    # Row 3+: " | County B | | | County B |"
    
    # Get assessment areas for each bank
    bank_a_aas = assessment_areas.get('bank_a', {})
    bank_b_aas = assessment_areas.get('bank_b', {})
    
    # Get BigQuery client for CBSA name lookups
    from justdata.shared.utils.bigquery_client import get_bigquery_client
    client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

    # Populate Bank A side (columns A and B)
    row = 2  # Start at row 2 (row 1 is header)
    for cbsa_code_or_name, counties in bank_a_aas.items():
        # Convert CBSA code to name if needed
        if str(cbsa_code_or_name).isdigit() or (cbsa_code_or_name and len(str(cbsa_code_or_name)) <= 5 and 'non-msa' not in str(cbsa_code_or_name).lower()):
            cbsa_name = _get_cbsa_name_from_code(str(cbsa_code_or_name), client)
        else:
            cbsa_name = str(cbsa_code_or_name)
        
        # Set CBSA name in column A (merge if needed based on template)
        _safe_set_cell_value(ws, row, 1, cbsa_name)  # Column A: CBSA name
        
        # List counties in column B (format: "County, State")
        if isinstance(counties, list):
            for i, county in enumerate(counties):
                # County might be a dict with county_name and state_name, or a string
                if isinstance(county, dict):
                    county_str = f"{county.get('county_name', '')}, {county.get('state_name', '')}"
                else:
                    county_str = str(county)
                _safe_set_cell_value(ws, row + i, 2, county_str)
            # CBSA name is only in first row (no merging)
            row += len(counties)
        else:
            county_str = str(counties) if not isinstance(counties, dict) else f"{counties.get('county_name', '')}, {counties.get('state_name', '')}"
            _safe_set_cell_value(ws, row, 2, county_str)
            row += 1
        row += 1  # Spacing
    
    # Populate Bank B side (columns D and E)
    row = 2
    for cbsa_code_or_name, counties in bank_b_aas.items():
        # Convert CBSA code to name if needed
        if str(cbsa_code_or_name).isdigit() or (cbsa_code_or_name and len(str(cbsa_code_or_name)) <= 5 and 'non-msa' not in str(cbsa_code_or_name).lower()):
            cbsa_name = _get_cbsa_name_from_code(str(cbsa_code_or_name), client)
        else:
            cbsa_name = str(cbsa_code_or_name)
        
        _safe_set_cell_value(ws, row, 4, cbsa_name)  # Column D: CBSA name
        
        if isinstance(counties, list):
            for i, county in enumerate(counties):
                if isinstance(county, dict):
                    county_str = f"{county.get('county_name', '')}, {county.get('state_name', '')}"
                else:
                    county_str = str(county)
                _safe_set_cell_value(ws, row + i, 5, county_str)
            # CBSA name is only in first row (no merging)
            row += len(counties)
        else:
            county_str = str(counties) if not isinstance(counties, dict) else f"{counties.get('county_name', '')}, {counties.get('state_name', '')}"
            _safe_set_cell_value(ws, row, 5, county_str)
            row += 1
        row += 1


def _add_hhi_sheet(wb, hhi_data: pd.DataFrame, bank_a_name: str, bank_b_name: str):
    """Add HHI Analysis sheet if it doesn't exist, or populate existing one."""
    sheet_name = "HHI Analysis"
    
    # Delete existing sheet if it exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    # Create new sheet
    ws = wb.create_sheet(sheet_name)
    print(f"  Populating {sheet_name}...")
    
    # Add headers
    _safe_set_cell_value(ws, 1, 1, "County, State")
    _safe_set_cell_value(ws, 1, 2, "GEOID5")
    _safe_set_cell_value(ws, 1, 3, "Pre-Merger HHI")
    _safe_set_cell_value(ws, 1, 4, "Post-Merger HHI")
    _safe_set_cell_value(ws, 1, 5, "HHI Change")
    _safe_set_cell_value(ws, 1, 6, "Pre-Merger Concentration")
    _safe_set_cell_value(ws, 1, 7, "Post-Merger Concentration")
    
    # Populate data
    if hhi_data is not None and not hhi_data.empty:
        for idx, row_data in hhi_data.iterrows():
            excel_row = idx + 2  # Start at row 2
            county_state = row_data.get('County, State', row_data.get('county_state', ''))
            geoid5 = row_data.get('GEOID5', row_data.get('geoid5', ''))
            pre_hhi = row_data.get('Pre-Merger HHI', row_data.get('pre_merger_hhi', 0))
            post_hhi = row_data.get('Post-Merger HHI', row_data.get('post_merger_hhi', 0))
            hhi_change = row_data.get('HHI Change', row_data.get('hhi_change', 0))
            pre_conc = row_data.get('Pre-Merger Concentration', row_data.get('pre_merger_concentration', ''))
            post_conc = row_data.get('Post-Merger Concentration', row_data.get('post_merger_concentration', ''))
            
            _safe_set_cell_value(ws, excel_row, 1, county_state)
            _safe_set_cell_value(ws, excel_row, 2, geoid5)
            _safe_set_cell_value(ws, excel_row, 3, pre_hhi)
            _safe_set_cell_value(ws, excel_row, 4, post_hhi)
            _safe_set_cell_value(ws, excel_row, 5, hhi_change)
            _safe_set_cell_value(ws, excel_row, 6, pre_conc)
            _safe_set_cell_value(ws, excel_row, 7, post_conc)
    else:
        print(f"  Warning: No HHI data provided for {sheet_name}")


def _populate_notes_sheet(
    wb, bank_a_name: str, bank_b_name: str,
    metadata: Optional[Dict], assessment_areas: Optional[Dict]
):
    """Delete existing Notes sheet and create new one with all metadata, filters, sources, and methods."""
    sheet_name = "Notes"
    
    # Delete existing Notes sheet if it exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    # Create new Notes sheet
    ws = wb.create_sheet(sheet_name, 0)  # Insert at beginning
    
    row = 1
    
    # Header
    _safe_set_cell_value(ws, row, 1, "Data Sources and Timeframes:")
    row += 2
    
    # 1. Mortgage Data Section
    ws.cell(row, 1, "1")
    ws.cell(row, 2, "Mortgage Data:")
    row += 1
    
    ws.cell(row, 2, "Sourced from HMDA (Home Mortgage Disclosure Act) database")
    row += 1
    
    # HMDA Years
    hmda_years = metadata.get('hmda_years', []) if metadata else []
    if isinstance(hmda_years, str):
        hmda_years = [y.strip() for y in hmda_years.split(',')]
    years_str = ', '.join(str(y) for y in hmda_years) if hmda_years else 'Not specified'
    ws.cell(row, 2, f"Covers {years_str}")
    row += 1
    
    ws.cell(row, 2, "Focuses on originations for owner-occupied, 1-4 unit, site-built properties")
    row += 1
    
    # Bank A Mortgage Info
    ws.cell(row, 4, "Respondent Name")
    ws.cell(row, 5, "LEI (Lenders18)")
    ws.cell(row, 6, "Respondent City")
    ws.cell(row, 7, "Respondent State")
    ws.cell(row, 8, "Type Name")
    ws.cell(row, 9, "Parent Name")
    row += 1
    
    acquirer_lei = metadata.get('acquirer_lei', '') if metadata else ''
    acquirer_city = metadata.get('acquirer_city', '') if metadata else ''
    acquirer_state = metadata.get('acquirer_state', '') if metadata else ''
    acquirer_type = metadata.get('acquirer_type', 'Bank or Affiliate') if metadata else 'Bank or Affiliate'
    acquirer_parent = metadata.get('acquirer_parent', '') if metadata else ''
    
    ws.cell(row, 4, bank_a_name)
    ws.cell(row, 5, acquirer_lei)
    ws.cell(row, 6, acquirer_city)
    ws.cell(row, 7, acquirer_state)
    ws.cell(row, 8, acquirer_type)
    ws.cell(row, 9, acquirer_parent)
    row += 1
    
    # Bank B Mortgage Info
    target_lei = metadata.get('target_lei', '') if metadata else ''
    target_city = metadata.get('target_city', '') if metadata else ''
    target_state = metadata.get('target_state', '') if metadata else ''
    target_type = metadata.get('target_type', 'Bank or Affiliate') if metadata else 'Bank or Affiliate'
    target_parent = metadata.get('target_parent', '') if metadata else ''
    
    ws.cell(row, 4, bank_b_name)
    ws.cell(row, 5, target_lei)
    ws.cell(row, 6, target_city)
    ws.cell(row, 7, target_state)
    ws.cell(row, 8, target_type)
    ws.cell(row, 9, target_parent)
    row += 2
    
    # HMDA Filters
    ws.cell(row, 2, "HMDA Filters Applied:")
    row += 1
    
    loan_purpose = metadata.get('loan_purpose', 'All') if metadata else 'All'
    action_taken = metadata.get('action_taken', '1') if metadata else '1'
    occupancy_type = metadata.get('occupancy_type', '1') if metadata else '1'
    total_units = metadata.get('total_units', '1-4') if metadata else '1-4'
    construction_method = metadata.get('construction_method', '1') if metadata else '1'
    not_reverse = metadata.get('not_reverse', '1') if metadata else '1'
    
    loan_purpose_map = {
        '1': 'Home Purchase',
        '2': 'Home Improvement',
        '3': 'Refinance',
        '31': 'Cash-out Refinance',
        '32': 'No Cash-out Refinance',
        '4': 'Other Purpose',
        '5': 'Not Applicable'
    }
    action_taken_map = {
        '1': 'Loan Originated',
        '2': 'Application Approved but Not Accepted',
        '3': 'Application Denied',
        '4': 'Application Withdrawn',
        '5': 'File Closed for Incompleteness',
        '6': 'Loan Purchased',
        '7': 'Preapproval Request Denied',
        '8': 'Preapproval Request Approved but Not Accepted'
    }
    
    ws.cell(row, 2, f"  - Loan Purpose: {loan_purpose_map.get(loan_purpose, loan_purpose)}")
    row += 1
    ws.cell(row, 2, f"  - Action Taken: {action_taken_map.get(action_taken, action_taken)} (Originations only)")
    row += 1
    ws.cell(row, 2, f"  - Occupancy Type: {occupancy_type} (Owner-occupied)")
    row += 1
    ws.cell(row, 2, f"  - Total Units: {total_units}")
    row += 1
    ws.cell(row, 2, f"  - Construction Method: {construction_method} (Site-built)")
    row += 1
    ws.cell(row, 2, f"  - Reverse Mortgage: {not_reverse} (Excluded)")
    row += 2
    
    # Peer Selection Method
    ws.cell(row, 2, "Peer Bank Selection Method:")
    row += 1
    ws.cell(row, 2, "  - 50%-200% Volume Rule: Banks with lending volume between 50% and 200% of subject bank's volume")
    row += 1
    ws.cell(row, 2, "  - Same assessment areas and time period")
    row += 2
    
    # 2. Small Business Data Section
    ws.cell(row, 1, "2")
    ws.cell(row, 2, "Small Business Data:")
    row += 1
    
    ws.cell(row, 2, "Sourced from CFPB Small Business Lending database")
    row += 1
    
    # SB Years
    sb_years = metadata.get('sb_years', []) if metadata else []
    if isinstance(sb_years, str):
        sb_years = [y.strip() for y in sb_years.split(',')]
    sb_years_str = ', '.join(str(y) for y in sb_years) if sb_years else 'Not specified'
    ws.cell(row, 2, f"Covers {sb_years_str}")
    row += 1
    
    # Bank A SB Info
    ws.cell(row, 4, "SB Lender")
    ws.cell(row, 5, "SB Lender City")
    ws.cell(row, 6, "SB Lender State")
    ws.cell(row, 7, "SB Resid")
    ws.cell(row, 8, "SB RSSD")
    row += 1
    
    acquirer_sb_id = metadata.get('acquirer_sb_id', '') if metadata else ''
    acquirer_rssd = metadata.get('acquirer_rssd', '') if metadata else ''
    
    ws.cell(row, 4, bank_a_name)
    ws.cell(row, 5, acquirer_city)
    ws.cell(row, 6, acquirer_state)
    ws.cell(row, 7, acquirer_sb_id)
    ws.cell(row, 8, acquirer_rssd)
    row += 1
    
    # Bank B SB Info
    target_sb_id = metadata.get('target_sb_id', '') if metadata else ''
    target_rssd = metadata.get('target_rssd', '') if metadata else ''
    
    ws.cell(row, 4, bank_b_name)
    ws.cell(row, 5, target_city)
    ws.cell(row, 6, target_state)
    ws.cell(row, 7, target_sb_id)
    ws.cell(row, 8, target_rssd)
    row += 2
    
    # 3. Branch Data Section
    ws.cell(row, 1, "3")
    ws.cell(row, 2, "Branch Data:")
    row += 1
    
    ws.cell(row, 2, "Sourced from FDIC Summary of Deposits (SOD) database")
    row += 1
    ws.cell(row, 2, "Year: 2025 (most recent available)")
    row += 1
    
    # Branch counts
    if assessment_areas:
        bank_a_aas = assessment_areas.get('bank_a', {})
        bank_b_aas = assessment_areas.get('bank_b', {})
        ws.cell(row, 2, f"Bank A Assessment Areas: {len(bank_a_aas)}")
        row += 1
        ws.cell(row, 2, f"Bank B Assessment Areas: {len(bank_b_aas)}")
        row += 1
    row += 1
    
    # 4. HHI Analysis Section
    ws.cell(row, 1, "4")
    ws.cell(row, 2, "HHI (Herfindahl-Hirschman Index) Analysis:")
    row += 1
    
    ws.cell(row, 2, "Sourced from FDIC Summary of Deposits (SOD) database")
    row += 1
    ws.cell(row, 2, "Calculates market concentration before and after merger")
    row += 1
    ws.cell(row, 2, "HHI thresholds: <1500 (competitive), 1500-2500 (moderate), >2500 (highly concentrated)")
    row += 2
    
    # 5. Assessment Areas
    ws.cell(row, 1, "5")
    ws.cell(row, 2, "Assessment Areas:")
    row += 1
    
    if assessment_areas:
        bank_a_aas = assessment_areas.get('bank_a', {})
        bank_b_aas = assessment_areas.get('bank_b', {})
        
        ws.cell(row, 2, f"Bank A ({bank_a_name}): {len(bank_a_aas)} assessment areas")
        row += 1
        for cbsa_name, counties in list(bank_a_aas.items())[:5]:  # Show first 5
            county_count = len(counties) if isinstance(counties, list) else 1
            ws.cell(row, 3, f"  - {cbsa_name} ({county_count} counties)")
            row += 1
        if len(bank_a_aas) > 5:
            ws.cell(row, 3, f"  ... and {len(bank_a_aas) - 5} more")
            row += 1
        
        row += 1
        ws.cell(row, 2, f"Bank B ({bank_b_name}): {len(bank_b_aas)} assessment areas")
        row += 1
        for cbsa_name, counties in list(bank_b_aas.items())[:5]:  # Show first 5
            county_count = len(counties) if isinstance(counties, list) else 1
            ws.cell(row, 3, f"  - {cbsa_name} ({county_count} counties)")
            row += 1
        if len(bank_b_aas) > 5:
            ws.cell(row, 3, f"  ... and {len(bank_b_aas) - 5} more")
            row += 1
    
    row += 2
    
    # 6. Methodology Notes
    ws.cell(row, 1, "6")
    ws.cell(row, 2, "Methodology Notes:")
    row += 1
    
    ws.cell(row, 2, "- Race/Ethnicity classification uses COALESCE methodology")
    row += 1
    ws.cell(row, 2, "- First checks for Hispanic ethnicity, then uses first valid race code for non-Hispanic applicants")
    row += 1
    ws.cell(row, 2, "- Race percentages use total loans as denominator")
    row += 1
    ws.cell(row, 2, "- LMICT: Low-to-Moderate Income Census Tract (≤80% of MSA median income)")
    row += 1
    ws.cell(row, 2, "- LMIB: Low-to-Moderate Income Borrower (≤80% of MSA median income)")
    row += 1
    ws.cell(row, 2, "- MMCT: Majority-Minority Census Tract (>50% minority population)")
    row += 1
    ws.cell(row, 2, "- MINB: Minority Borrower (Hispanic, Black, Asian, Native American, or HoPI)")
    row += 1


# Helper functions

def _get_cbsa_name_from_code(cbsa_code: str, client=None) -> str:
    """Get CBSA name from CBSA code using BigQuery."""
    if not cbsa_code or cbsa_code == 'N/A' or 'non-msa' in str(cbsa_code).lower():
        return str(cbsa_code)  # Return as-is for Non-MSA
    
    try:
        from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query

        if client is None:
            client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

        query = f"""
        SELECT DISTINCT cbsa as cbsa_name
        FROM `{PROJECT_ID}.geo.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) = '{cbsa_code}'
        LIMIT 1
        """
        results = execute_query(client, query)
        if results and results[0].get('cbsa_name'):
            return results[0]['cbsa_name']
        return str(cbsa_code)  # Fallback to code if name not found
    except Exception as e:
        print(f"  Warning: Could not look up CBSA name for {cbsa_code}: {e}")
        return str(cbsa_code)


def _group_by_cbsa(data: pd.DataFrame, assessment_areas: Optional[Dict]) -> Dict[str, pd.DataFrame]:
    """Group data by CBSA name, converting codes to names."""
    groups = {}
    
    # Get BigQuery client for CBSA name lookups
    from justdata.shared.utils.bigquery_client import get_bigquery_client
    client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

    if 'cbsa_name' in data.columns:
        for cbsa_name, group in data.groupby('cbsa_name'):
            # If it's actually a code, convert to name
            if str(cbsa_name).isdigit() or (cbsa_name and 'non-msa' not in str(cbsa_name).lower() and len(str(cbsa_name)) <= 5):
                cbsa_name = _get_cbsa_name_from_code(str(cbsa_name), client)
            groups[cbsa_name] = group
    elif 'cbsa_code' in data.columns:
        for cbsa_code, group in data.groupby('cbsa_code'):
            # Convert code to name
            cbsa_name = _get_cbsa_name_from_code(str(cbsa_code), client)
            groups[cbsa_name] = group
    
    return groups


def _group_sb_by_state(data: pd.DataFrame, assessment_areas: Optional[Dict]) -> Dict[str, pd.DataFrame]:
    """Group SB data by state."""
    groups = {}
    
    if 'state_name' in data.columns:
        for state_name, group in data.groupby('state_name'):
            groups[state_name] = group
    elif 'state_code' in data.columns:
        for state_code, group in data.groupby('state_code'):
            groups[str(state_code)] = group
    
    return groups


def _group_sb_by_assessment_area(data: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Group SB data by assessment area (CBSA)."""
    return _group_by_cbsa(data, None)


def _match_cbsa_in_data(data: pd.DataFrame, cbsa_name: str) -> pd.DataFrame:
    """Find matching CBSA data in peer data."""
    if 'cbsa_name' in data.columns:
        return data[data['cbsa_name'] == cbsa_name]
    elif 'cbsa_code' in data.columns:
        return data[data['cbsa_code'] == cbsa_name]
    return pd.DataFrame()


def _apply_number_formatting(wb, bank_a_name: str, bank_b_name: str):
    """
    Apply number formatting to all data sheets after population.

    Format rules:
    - Row 2 (Loans): '#,##0' (whole number with thousands separator)
    - Rows with '%' metrics: '0.00' (2 decimal places - values are already *100)
    - Rows with '$' metrics: '$#,##0' (currency)
    - Difference columns (E): '0.00' or percentage as appropriate
    - Branch difference columns: '0.00%' (but values need to be in decimal form)
    """
    # Define metric patterns and their formats
    # Note: Percentages are stored as 5.5 meaning 5.5%, so format as '0.00'
    PERCENTAGE_ROWS = ['LMICT%', 'LMIB%', 'MMCT%', 'MINB%', 'Asian%', 'Black%',
                       'Native American%', 'HoPI%', 'Hispanic%']
    CURRENCY_ROWS = ['LMIB$']
    LOAN_COUNT_ROWS = ['Loans', 'SB Loans', 'Branches']

    # Get sheet name patterns for mortgage data sheets
    short_name_a = bank_a_name.split()[0] if bank_a_name else 'Bank A'
    short_name_b = bank_b_name.split()[0] if bank_b_name else 'Bank B'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Check if this is a mortgage or SB data sheet
        is_mortgage_sheet = 'MORTGAGE' in sheet_name.upper() and 'GOALS' not in sheet_name.upper()
        is_sb_sheet = 'SB' in sheet_name.upper() or 'SMALL BUSINESS' in sheet_name.upper()
        is_branch_sheet = 'BRANCH' in sheet_name.upper()

        if not (is_mortgage_sheet or is_sb_sheet or is_branch_sheet):
            continue

        # Iterate through data rows (start from row 2)
        for row in range(2, ws.max_row + 1):
            # Get metric name from column B
            metric_cell = ws.cell(row, 2)
            metric_name = str(metric_cell.value).strip() if metric_cell.value else ''

            # Apply formatting to columns C, D, E (Subject, Peer, Difference)
            for col in [3, 4, 5]:
                cell = ws.cell(row, col)
                if cell.value is None or cell.value == '':
                    continue

                # Skip if it's a formula (starts with =)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Apply format based on metric type for formula cells too
                    if any(pct in metric_name for pct in PERCENTAGE_ROWS):
                        cell.number_format = '0.00'
                    elif any(curr in metric_name for curr in CURRENCY_ROWS):
                        cell.number_format = '$#,##0'
                    elif any(cnt in metric_name for cnt in LOAN_COUNT_ROWS):
                        cell.number_format = '#,##0'
                    elif is_branch_sheet and col == 5:
                        # Branch difference column should be percentage
                        cell.number_format = '0.00%'
                    continue

                # Apply format based on metric type
                if any(pct in metric_name for pct in PERCENTAGE_ROWS):
                    cell.number_format = '0.00'
                elif any(curr in metric_name for curr in CURRENCY_ROWS):
                    cell.number_format = '$#,##0'
                elif any(cnt in metric_name for cnt in LOAN_COUNT_ROWS):
                    cell.number_format = '#,##0'
                elif is_branch_sheet:
                    if col == 5 and metric_name not in ['Branches']:
                        # Branch LMICT/MMCT difference column - percentage
                        cell.number_format = '0.00%'
                    elif col in [3, 4]:
                        cell.number_format = '#,##0'
                else:
                    # Default: numeric with 2 decimals
                    if isinstance(cell.value, (int, float)):
                        if abs(cell.value) > 100:
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '0.00'


def _unmerge_cells_in_range(ws, range_str: str):
    """Unmerge any merged cells that overlap with the given range."""
    try:
        # Get all merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)
        
        # Parse the target range
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        
        # Find and unmerge overlapping ranges
        ranges_to_unmerge = []
        for merged_range in merged_ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged_range))
            # Check if ranges overlap
            if not (max_col < m_min_col or min_col > m_max_col or 
                   max_row < m_min_row or min_row > m_max_row):
                ranges_to_unmerge.append(merged_range)
        
        # Unmerge overlapping ranges
        for merged_range in ranges_to_unmerge:
            ws.unmerge_cells(str(merged_range))
    except Exception as e:
        # If unmerging fails, continue - it might not be necessary
        pass


def _safe_set_cell_value(ws, row: int, col: int, value, number_format: str = None):
    """Safely set a cell value, handling merged cells and optional number formatting.

    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        value: Value to set
        number_format: Optional Excel number format string (e.g., '#,##0', '0.00%', '$#,##0')
    """
    try:
        cell = ws.cell(row, col)
        # Check if this cell is part of a merged range
        for merged_range in list(ws.merged_cells.ranges):
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged_range))
            if (m_min_col <= col <= m_max_col and m_min_row <= row <= m_max_row):
                # This cell is in a merged range - unmerge first
                ws.unmerge_cells(str(merged_range))
                break

        # Now set the value
        cell.value = value

        # Apply number format if specified
        if number_format:
            cell.number_format = number_format
    except Exception as e:
        # Fallback: try direct assignment
        try:
            ws.cell(row, col).value = value
            if number_format:
                ws.cell(row, col).number_format = number_format
        except:
            pass


def _adjust_formula_rows(
    formula: str, template_row: int, target_row: int, reference_row: int, base_row: int
) -> str:
    """
    Adjust formula row references, avoiding circular references.
    
    Args:
        formula: Original formula string
        template_row: Row number in template where formula was
        target_row: Row number where formula should be placed
        reference_row: Row number that should be referenced (e.g., loans row)
        base_row: Base row for calculations (e.g., Grand Total row 2)
    """
    # For difference formulas, use simple =C{row}-D{row} pattern
    # Avoid circular references by ensuring formula doesn't reference itself
    
    # Check if formula creates a circular reference (references the same row it's in)
    # If so, use a simpler formula
    if f"{target_row}" in formula and f"C{target_row}" in formula:
        # Potential circular reference - use simple difference
        return f"=IFERROR(C{target_row}-D{target_row},\"\")"
    
    # Calculate offsets
    row_offset = target_row - template_row
    
    # Replace row numbers in formula
    # Pattern: C3 -> C{target_row}, but C2 (base) -> C{reference_row}
    def replace_row(match):
        col_letter = match.group(1)
        old_row = int(match.group(2))
        
        # Avoid circular reference - don't reference the same row
        if old_row == template_row and target_row == template_row:
            # This would create a circular reference
            return f"{col_letter}{reference_row}"
        
        if old_row == base_row:
            # Reference to base row should point to reference_row
            return f"{col_letter}{reference_row}"
        else:
            # Other rows adjusted by offset
            new_row = old_row + row_offset
            # Ensure we don't create circular reference
            if new_row == target_row and col_letter in ['C', 'D']:
                # Would reference same row - use reference_row instead
                return f"{col_letter}{reference_row}"
            return f"{col_letter}{new_row}"
    
    # Match column letters followed by row numbers
    pattern = r'([A-Z]+)(\d+)'
    new_formula = re.sub(pattern, replace_row, formula)
    
    return new_formula

