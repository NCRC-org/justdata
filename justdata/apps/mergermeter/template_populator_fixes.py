"""
Additional helper functions for template population fixes.
"""

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, Optional
import pandas as pd
from justdata.apps.mergermeter.config import PROJECT_ID


def get_cbsa_name_from_code(cbsa_code: str, client=None) -> str:
    """Get CBSA name from CBSA code using BigQuery."""
    if not cbsa_code or cbsa_code == 'N/A' or 'non-msa' in str(cbsa_code).lower():
        return str(cbsa_code)  # Return as-is for Non-MSA
    
    try:
        from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query

        if client is None:
            client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

        query = f"""
        SELECT DISTINCT cbsa as cbsa_name
        FROM `{PROJECT_ID}.geo.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) = '{cbsa_code}'
        LIMIT 1
        """
        results = execute_query(client, query)
        if results and results[0].get('cbsa_name'):
            return results[0]['cbsa_name']
        return str(cbsa_code)  # Fallback to code if name not found
    except Exception as e:
        print(f"  Warning: Could not look up CBSA name for {cbsa_code}: {e}")
        return str(cbsa_code)


def remove_cell_colors_from_sheet(ws):
    """Remove all cell fill colors from a worksheet."""
    no_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type:
                cell.fill = no_fill

