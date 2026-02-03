"""
Post-processing functions for Excel formatting after generation.
Handles Notes sheet, Assessment Areas, branch formatting, and sheet name cleanup.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import pandas as pd
from typing import Dict, Optional
from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query
from justdata.apps.mergermeter.config import PROJECT_ID


def post_process_excel(
    excel_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame],
    bank_b_hmda_subject: Optional[pd.DataFrame],
    bank_a_hmda_peer: Optional[pd.DataFrame],
    bank_b_hmda_peer: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict],
    metadata: Optional[Dict]
):
    """
    Post-process Excel file to apply all formatting updates:
    1. Fix sheet names (remove "1" suffix)
    2. Update Notes sheet columns A, B, C with mortgage metrics
    3. Update Assessment Areas sheet columns B and F with CBSA names
    4. Format branch sheets columns D and E as whole numbers
    5. Remove duplicate rows from branch sheets
    6. Update header colors to NCRC blue (#034ea0)
    """
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        # 1. Fix sheet names - remove "1" suffix
        _fix_sheet_names(wb, bank_a_name, bank_b_name)
        
        # 2. REMOVED: Update header colors - user wants NO colors at all
        # _update_header_colors(wb)
        
        # 3. Update Notes sheet columns A, B, C with mortgage metrics
        _update_notes_sheet(wb, bank_a_hmda_subject, bank_b_hmda_subject, 
                          bank_a_hmda_peer, bank_b_hmda_peer)
        
        # 4. Update Assessment Areas sheet columns B and F with CBSA names
        if assessment_areas:
            _update_assessment_areas_cbsa_names(wb, assessment_areas)
        
        # 5. Format branch sheets
        _format_branch_sheets(wb)
        
        # 6. Remove state columns from data sheets (keep only on goal setting sheet)
        _remove_state_columns_from_data_sheets(wb)
        
        # 7. Remove grand total cell colors
        _remove_grand_total_colors(wb)
        
        # 8. FINAL PASS: Remove ALL cell colors from ALL sheets (user wants NO colors)
        _remove_all_cell_colors(wb)
        
        wb.save(excel_path)
        print("  Post-processed Excel file: applied all formatting updates")
        
    except Exception as e:
        print(f"  Warning: Could not post-process Excel file: {e}")
        import traceback
        traceback.print_exc()


def _fix_sheet_names(wb, bank_a_name: str, bank_b_name: str):
    """Fix sheet names by removing '1' suffix and ensuring proper spacing."""
    for sheet_name in list(wb.sheetnames):
        ws = wb[sheet_name]
        new_name = sheet_name
        
        # Remove trailing "1" if it exists
        if new_name.endswith('1') and len(new_name) > 1:
            # Check if it's actually a duplicate (e.g., "Sheet1" vs "Sheet")
            base_name = new_name[:-1]
            if base_name not in [s for s in wb.sheetnames if s != sheet_name]:
                new_name = base_name
        
        # Fix spacing issues (e.g., "PNC BANKMortgage Data1" -> "PNC BANK Mortgage Data")
        # Add space before capital letters if missing
        import re
        new_name = re.sub(r'([a-z])([A-Z])', r'\1 \2', new_name)
        new_name = re.sub(r'([A-Z])([A-Z][a-z])', r'\1 \2', new_name)
        
        # Remove trailing "1" again after spacing fix
        if new_name.endswith(' 1') or new_name.endswith('1'):
            new_name = new_name.rstrip(' 1').rstrip('1')
        
        # Replace bank name placeholders
        if 'PNC Bank' in new_name or 'PNC BANK' in new_name:
            new_name = new_name.replace('PNC Bank', bank_a_name).replace('PNC BANK', bank_a_name.upper())
        if 'FirstBank' in new_name or 'First Bank' in new_name:
            new_name = new_name.replace('FirstBank', bank_b_name).replace('First Bank', bank_b_name)
        if 'Bank A' in new_name:
            new_name = new_name.replace('Bank A', bank_a_name)
        if 'Bank B' in new_name:
            new_name = new_name.replace('Bank B', bank_b_name)
        
        # Clean up multiple spaces
        new_name = ' '.join(new_name.split())
        
        if new_name != sheet_name:
            try:
                ws.title = new_name
            except:
                # If name already exists, try with a unique suffix
                counter = 2
                while f"{new_name}{counter}" in wb.sheetnames:
                    counter += 1
                ws.title = f"{new_name}{counter}"


def _update_header_colors(wb):
    """Update header colors to NCRC blue (#034ea0)."""
    ncrc_blue = PatternFill(start_color="034EA0", end_color="034EA0", fill_type="solid")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Update header row colors (usually row 1 or 2)
        for row_idx in [1, 2]:
            if row_idx <= ws.max_row:
                for col_idx in range(1, min(ws.max_column + 1, 50)):  # Limit to first 50 columns
                    cell = ws.cell(row_idx, col_idx)
                    # Check if it's a header cell (has fill color or is bold)
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        # Update to NCRC blue
                        cell.fill = ncrc_blue
                    elif cell.font and cell.font.bold:
                        # Also update bold cells in header rows
                        cell.fill = ncrc_blue


def _update_notes_sheet(wb, bank_a_hmda_subject, bank_b_hmda_subject, bank_a_hmda_peer, bank_b_hmda_peer):
    """Update Notes sheet columns A, B, C with mortgage metrics."""
    notes_sheet = None
    for sheet_name in wb.sheetnames:
        if 'note' in sheet_name.lower():
            notes_sheet = wb[sheet_name]
            break
    
    if not notes_sheet:
        return
    
    try:
        # Calculate aggregate metrics from HMDA data
        def calculate_metrics(hmda_df):
            if hmda_df is None or hmda_df.empty:
                return {
                    'loans': 0,
                    'lmict_pct': 0.0,
                    'lmib_pct': 0.0,
                    'lmib_amount': 0,
                    'mmct_pct': 0.0,
                    'minb_pct': 0.0,
                }
            
            # Find relevant columns (case-insensitive)
            total_loans_col = None
            lmict_loans_col = None
            lmib_loans_col = None
            lmib_amount_col = None
            mmct_loans_col = None
            minb_loans_col = None
            
            for col in hmda_df.columns:
                col_lower = str(col).lower()
                if 'total' in col_lower and 'loan' in col_lower and 'count' in col_lower:
                    total_loans_col = col
                elif 'lmict' in col_lower and 'loan' in col_lower:
                    lmict_loans_col = col
                elif 'lmib' in col_lower and 'loan' in col_lower:
                    lmib_loans_col = col
                elif 'lmib' in col_lower and ('amount' in col_lower or 'amt' in col_lower):
                    lmib_amount_col = col
                elif 'mmct' in col_lower and 'loan' in col_lower:
                    mmct_loans_col = col
                elif 'minb' in col_lower and 'loan' in col_lower:
                    minb_loans_col = col
            
            total_loans = int(hmda_df[total_loans_col].sum()) if total_loans_col else 0
            lmict_loans = int(hmda_df[lmict_loans_col].sum()) if lmict_loans_col else 0
            lmib_loans = int(hmda_df[lmib_loans_col].sum()) if lmib_loans_col else 0
            lmib_amount = float(hmda_df[lmib_amount_col].sum()) if lmib_amount_col else 0.0
            mmct_loans = int(hmda_df[mmct_loans_col].sum()) if mmct_loans_col else 0
            minb_loans = int(hmda_df[minb_loans_col].sum()) if minb_loans_col else 0
            
            lmict_pct = (lmict_loans / total_loans * 100) if total_loans > 0 else 0.0
            lmib_pct = (lmib_loans / total_loans * 100) if total_loans > 0 else 0.0
            mmct_pct = (mmct_loans / total_loans * 100) if total_loans > 0 else 0.0
            minb_pct = (minb_loans / total_loans * 100) if total_loans > 0 else 0.0
            
            return {
                'loans': total_loans,
                'lmict_pct': lmict_pct,
                'lmib_pct': lmib_pct,
                'lmib_amount': lmib_amount,
                'mmct_pct': mmct_pct,
                'minb_pct': minb_pct,
            }
        
        bank_a_metrics = calculate_metrics(bank_a_hmda_subject)
        bank_b_metrics = calculate_metrics(bank_b_hmda_subject)
        
        # Find rows with metric labels and update values
        metric_mapping = {
            'Loans': ('loans', None),
            'LMICT%': ('lmict_pct', '0.00%'),
            'LMIB%': ('lmib_pct', '0.00%'),
            'LMIB$': ('lmib_amount', '$#,##0'),
            'MMCT%': ('mmct_pct', '0.00%'),
            'MINB%': ('minb_pct', '0.00%'),
        }
        
        for row_idx in range(1, min(notes_sheet.max_row + 1, 100)):  # Check first 100 rows
            col_a_value = str(notes_sheet.cell(row_idx, 1).value or '').strip()
            
            for metric_name, (metric_key, number_format) in metric_mapping.items():
                if metric_name.lower() in col_a_value.lower():
                    # Update Bank A value (Column B)
                    bank_a_value = bank_a_metrics.get(metric_key, 0)
                    cell_b = notes_sheet.cell(row_idx, 2)
                    if number_format:
                        if '%' in number_format:
                            cell_b.value = bank_a_value / 100
                            cell_b.number_format = number_format
                        elif '$' in number_format:
                            cell_b.value = bank_a_value
                            cell_b.number_format = number_format
                        else:
                            cell_b.value = bank_a_value
                    else:
                        cell_b.value = bank_a_value
                    
                    # Update Bank B value (Column C)
                    bank_b_value = bank_b_metrics.get(metric_key, 0)
                    cell_c = notes_sheet.cell(row_idx, 3)
                    if number_format:
                        if '%' in number_format:
                            cell_c.value = bank_b_value / 100
                            cell_c.number_format = number_format
                        elif '$' in number_format:
                            cell_c.value = bank_b_value
                            cell_c.number_format = number_format
                        else:
                            cell_c.value = bank_b_value
                    else:
                        cell_c.value = bank_b_value
                    
                    break
        
    except Exception as e:
        print(f"  Warning: Could not update Notes sheet: {e}")


def _update_assessment_areas_cbsa_names(wb, assessment_areas):
    """Update Assessment Areas sheet columns B and F with CBSA names."""
    aa_sheet = None
    for sheet_name in wb.sheetnames:
        if 'assessment' in sheet_name.lower() and 'area' in sheet_name.lower():
            aa_sheet = wb[sheet_name]
            break
    
    if not aa_sheet:
        return

    try:
        client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

        # Find CBSA code column (usually column B or C)
        cbsa_code_col = None
        cbsa_name_col = None
        
        # Check header row
        for col_idx in range(1, min(aa_sheet.max_column + 1, 10)):
            header = str(aa_sheet.cell(1, col_idx).value or '').strip().lower()
            if 'cbsa' in header and 'code' in header:
                cbsa_code_col = col_idx
            elif 'cbsa' in header and ('name' in header or 'cbsa' == header):
                cbsa_name_col = col_idx
        
        # Default: column B for code, column F for name
        if not cbsa_code_col:
            cbsa_code_col = 2  # Column B
        if not cbsa_name_col:
            cbsa_name_col = 6  # Column F
        
        # Update CBSA names
        for row_idx in range(2, aa_sheet.max_row + 1):  # Start from row 2 (skip header)
            cbsa_code_cell = aa_sheet.cell(row_idx, cbsa_code_col)
            cbsa_name_cell = aa_sheet.cell(row_idx, cbsa_name_col)
            
            cbsa_code = str(cbsa_code_cell.value or '').strip()
            current_name = str(cbsa_name_cell.value or '').strip()
            
            # If already has a name and it's not empty, skip
            if current_name and current_name != 'N/A' and current_name != '':
                continue
            
            if cbsa_code and cbsa_code != 'N/A':
                # Check if it's a Non-MSA code (contains "Non-MSA")
                if 'non-msa' in cbsa_code.lower():
                    cbsa_name_cell.value = cbsa_code  # Use the code as name for Non-MSA
                elif cbsa_code.isdigit():
                    # Query CBSA name
                    query = f"""
                    SELECT DISTINCT cbsa as cbsa_name
                    FROM `{PROJECT_ID}.shared.cbsa_to_county`
                    WHERE CAST(cbsa_code AS STRING) = '{cbsa_code}'
                    LIMIT 1
                    """
                    results = execute_query(client, query)
                    if results:
                        cbsa_name = results[0].get('cbsa_name', '')
                        if cbsa_name:
                            cbsa_name_cell.value = cbsa_name
                else:
                    # If it's already a name (like "Wisconsin Non-MSA"), use it
                    cbsa_name_cell.value = cbsa_code
        
    except Exception as e:
        print(f"  Warning: Could not update Assessment Areas CBSA names: {e}")


def _format_branch_sheets(wb):
    """Format branch sheets: columns D and E as whole numbers, remove duplicates."""
    for sheet_name in wb.sheetnames:
        if 'branch' not in sheet_name.lower():
            continue
        
        ws = wb[sheet_name]
        
        try:
            # Format columns D and E as whole numbers
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                for col_idx in [4, 5]:  # Columns D and E
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None:
                        try:
                            # Convert to integer if it's a number
                            if isinstance(cell.value, (int, float)):
                                cell.value = int(cell.value)
                                cell.number_format = '#,##0'
                        except (ValueError, TypeError):
                            pass  # Keep as is if can't convert
            
            # Remove duplicate rows (e.g., repeated "Colorado Non-MSA" with missing data)
            # Look for rows where the first column is the same as previous rows
            rows_to_delete = []
            last_first_col_value = None
            consecutive_duplicates = []
            
            for row_idx in range(2, ws.max_row + 1):
                first_col_value = str(ws.cell(row_idx, 1).value or '').strip()
                
                # Check if this is a duplicate of the previous row
                if first_col_value == last_first_col_value and first_col_value:
                    consecutive_duplicates.append(row_idx)
                else:
                    # If we had consecutive duplicates, check if they have missing data
                    if consecutive_duplicates:
                        # Check if the first duplicate row has data
                        first_dup_row = consecutive_duplicates[0]
                        has_data = False
                        for col_idx in range(2, min(ws.max_column + 1, 10)):
                            cell_value = ws.cell(first_dup_row, col_idx).value
                            if cell_value and str(cell_value).strip() and str(cell_value) != '0':
                                has_data = True
                                break
                        
                        # If first row has data, delete the rest
                        if has_data:
                            rows_to_delete.extend(consecutive_duplicates[1:])
                        # If first row has no data, keep the first and delete the rest
                        else:
                            rows_to_delete.extend(consecutive_duplicates[1:])
                    
                    consecutive_duplicates = []
                    last_first_col_value = first_col_value
            
            # Handle remaining duplicates at the end
            if consecutive_duplicates:
                if len(consecutive_duplicates) > 1:
                    rows_to_delete.extend(consecutive_duplicates[1:])
            
            # Delete rows in reverse order to maintain indices
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)
            
            if rows_to_delete:
                print(f"  Removed {len(rows_to_delete)} duplicate rows from {sheet_name}")
        
        except Exception as e:
            print(f"  Warning: Could not format branch sheet {sheet_name}: {e}")


def _remove_state_columns_from_data_sheets(wb):
    """Remove state columns from mortgage, small business, and branch sheets. Keep states only on goal setting sheet."""
    state_indicators = ['state', 'state code', 'state name', 'state abbreviation']
    state_abbrevs = ['AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 
                     'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ',
                     'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT',
                     'VA', 'WA', 'WV', 'WI', 'WY', 'DC']
    
    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        
        # Skip goal setting sheet
        if 'goal' in sheet_lower and 'setting' in sheet_lower:
            continue
        
        # Only process mortgage, small business, and branch sheets
        if not any(keyword in sheet_lower for keyword in ['mortgage', 'small business', 'branch']):
            continue
        
        ws = wb[sheet_name]
        
        try:
            # Find state column by checking header row
            state_col_idx = None
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                header = str(ws.cell(1, col_idx).value or '').strip().lower()
                if any(indicator in header for indicator in state_indicators):
                    state_col_idx = col_idx
                    break
            
            # If not found by header, check if column A contains state data
            if not state_col_idx:
                # Check first few data rows of column A
                is_state_col = False
                for row_idx in range(2, min(6, ws.max_row + 1)):
                    cell_value = str(ws.cell(row_idx, 1).value or '').strip()
                    if cell_value.upper() in state_abbrevs or len(cell_value) == 2:
                        is_state_col = True
                        break
                
                if is_state_col:
                    state_col_idx = 1
            
            # Remove the state column if found
            if state_col_idx:
                print(f"  Removing state column (column {state_col_idx}) from {sheet_name}")
                ws.delete_cols(state_col_idx, 1)
        
        except Exception as e:
            print(f"  Warning: Could not remove state column from {sheet_name}: {e}")


def _remove_grand_total_colors(wb):
    """Remove cell colors from grand total rows across all sheets."""
    no_fill = PatternFill(fill_type=None)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        try:
            # Look for rows containing "Grand Total" or "Total" in the first few columns
            for row_idx in range(1, ws.max_row + 1):
                # Check first 3 columns for "Grand Total" or "Total" text
                is_grand_total = False
                for col_idx in range(1, min(4, ws.max_column + 1)):
                    cell_value = str(ws.cell(row_idx, col_idx).value or '').strip().lower()
                    if 'grand total' in cell_value or (col_idx == 1 and 'total' in cell_value and len(cell_value) < 20):
                        is_grand_total = True
                        break
                
                if is_grand_total:
                    # Remove fill color from all cells in this row
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row_idx, col_idx)
                        if cell.fill and cell.fill.fill_type:
                            cell.fill = no_fill
        
        except Exception as e:
            print(f"  Warning: Could not remove grand total colors from {sheet_name}: {e}")


def _remove_all_cell_colors(wb):
    """Remove ALL cell colors from ALL sheets - final pass to ensure no colors remain."""
    no_fill = PatternFill(fill_type=None)
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            # Remove all fills from all cells
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill:
                        cell.fill = no_fill
            # Also clear any conditional formatting that might add colors
            if hasattr(ws, 'conditional_formatting'):
                ws.conditional_formatting = {}
        except Exception as e:
            print(f"  Warning: Could not remove all colors from {sheet_name}: {e}")

