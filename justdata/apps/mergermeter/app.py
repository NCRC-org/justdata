#!/usr/bin/env python3
"""
MergerMeter Flask web application - Two-bank merger impact analyzer.
"""

from flask import render_template, request, jsonify, send_file, session, Response
import os
import sys
import tempfile
import zipfile
from datetime import datetime
import uuid
import threading
import time
import json
from typing import List, Dict
from pathlib import Path
from werkzeug.middleware.proxy_fix import ProxyFix

# Add repo root to path for shared modules
REPO_ROOT = Path(__file__).parent.parent.parent.absolute()
sys.path.insert(0, str(REPO_ROOT))

from justdata.shared.web.app_factory import create_app, register_standard_routes
from justdata.shared.utils.progress_tracker import get_progress, update_progress, create_progress_tracker
from justdata.shared.utils.unified_env import ensure_unified_env_loaded, get_unified_config

# Use absolute imports from repo root (like other apps) - avoids issues with gunicorn
from justdata.apps.mergermeter.config import TEMPLATES_DIR, STATIC_DIR, OUTPUT_DIR, PROJECT_ID
from justdata.apps.mergermeter.version import __version__

# GCS storage for persistent file storage across Cloud Run instances
from justdata.shared.utils.gcs_storage import upload_json, download_json

# Load unified environment configuration (primary method - works for both local and Render)
ensure_unified_env_loaded(verbose=True)
config = get_unified_config(verbose=True)
print(f"[MergerMeter] Environment: {'LOCAL' if config['IS_LOCAL'] else 'PRODUCTION (Render)'}")
print(f"[MergerMeter] Shared config loaded from: {config.get('SHARED_ENV_FILE', 'Environment variables')}")


def _import_local_module(module_name, *attributes):
    """
    Import a local module using file-based import to avoid relative import issues.
    Returns the module or a tuple of attributes if specified.
    If only one attribute is requested, returns it directly (not in a tuple).
    """
    import sys
    from pathlib import Path
    import importlib.util
    
    module_path = Path(__file__).parent / f'{module_name}.py'
    if module_path.exists():
        spec = importlib.util.spec_from_file_location(module_name, module_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        if attributes:
            result = tuple(getattr(module, attr) for attr in attributes)
            # If only one attribute, return it directly (not as a tuple)
            return result[0] if len(result) == 1 else result
        return module
    else:
        # Fallback: try relative import
        try:
            module = __import__(f'.{module_name}', fromlist=[*attributes] if attributes else [], package=__package__ or 'mergermeter')
            if attributes:
                result = tuple(getattr(module, attr) for attr in attributes)
                # If only one attribute, return it directly (not as a tuple)
                return result[0] if len(result) == 1 else result
            return module
        except (ImportError, AttributeError):
            # Last resort: try absolute import
            module = __import__(module_name, fromlist=[*attributes] if attributes else [])
            if attributes:
                result = tuple(getattr(module, attr) for attr in attributes)
                # If only one attribute, return it directly (not as a tuple)
                return result[0] if len(result) == 1 else result
            return module


# Create the Flask app
app = create_app(
    'mergermeter',
    template_folder=TEMPLATES_DIR,
    static_folder=STATIC_DIR
)

# Configure shared templates (needed for standalone running)
from jinja2 import ChoiceLoader, FileSystemLoader
SHARED_TEMPLATES_DIR = REPO_ROOT / 'shared' / 'web' / 'templates'
app.jinja_loader = ChoiceLoader([
    FileSystemLoader(str(TEMPLATES_DIR)),       # App templates first
    FileSystemLoader(str(SHARED_TEMPLATES_DIR)), # Shared templates second
])

# Add ProxyFix for proper request handling behind Render's proxy
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# Set maximum file upload size to 10MB (plenty for JSON files)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

# Note: /health endpoint is already registered by create_app() in app_factory.py


def index():
    """Main page with the analysis form"""
    return render_template('analysis_template.html',
                         version=__version__,
                         app_name='MergerMeter',
                         breadcrumb_items=[{'name': 'MergerMeter', 'url': '/mergermeter'}])


def report():
    """Report display page"""
    # Check if job_id is in URL parameters (fallback if session doesn't persist)
    job_id_from_url = request.args.get('job_id')
    if job_id_from_url and not session.get('job_id'):
        session['job_id'] = job_id_from_url
    return render_template('mergermeter_report.html',
                         version=__version__,
                         app_name='MergerMeter',
                         breadcrumb_items=[
                             {'name': 'MergerMeter', 'url': '/mergermeter'},
                             {'name': 'Report', 'url': '/mergermeter/report'}
                         ])


def progress_handler(job_id):
    """Progress tracking endpoint using Server-Sent Events"""
    def event_stream():
        last_percent = -1
        last_step = ""
        keepalive_counter = 0
        max_keepalive = 20  # Send keepalive every 10 seconds (20 * 0.5s)
        
        # Maximum stream duration: 5 minutes (600 seconds) to prevent memory issues
        max_duration = 300  # 5 minutes
        start_time = time.time()
        max_iterations = int(max_duration / 0.5)  # Maximum iterations based on sleep time
        iteration_count = 0
        
        try:
            # Send initial connection message
            yield f": connected\n\n"
            
            while iteration_count < max_iterations:
                try:
                    # Check if we've exceeded maximum duration
                    elapsed = time.time() - start_time
                    if elapsed >= max_duration:
                        print(f"Progress stream timeout for {job_id} after {elapsed:.1f}s")
                        yield f"data: {{\"percent\": {last_percent}, \"step\": \"Connection timeout - please refresh\", \"done\": true, \"error\": \"Stream timeout after 5 minutes\"}}\n\n"
                        break
                    
                    progress = get_progress(job_id)
                    if not progress:
                        # If no progress found, send default
                        progress = {'percent': 0, 'step': 'Starting...', 'done': False, 'error': None}
                    
                    percent = progress.get("percent", 0)
                    step = progress.get("step", "Starting...")
                    done = progress.get("done", False)
                    error = progress.get("error", None)
                    
                    # Escape step message for JSON
                    step_escaped = step.replace('"', '\\"').replace('\n', '\\n').replace('\r', '')
                    
                    # Send update if percent, step, done, or error changed
                    if percent != last_percent or step != last_step or done or error:
                        yield f"data: {{\"percent\": {percent}, \"step\": \"{step_escaped}\", \"done\": {str(done).lower()}, \"error\": {json.dumps(error) if error else 'null'}}}\n\n"
                        last_percent = percent
                        last_step = step
                        keepalive_counter = 0
                    
                    if done or error:
                        # Small delay to ensure final message is sent
                        time.sleep(0.1)
                        break
                    
                    # Send keepalive comment periodically to keep connection alive
                    keepalive_counter += 1
                    if keepalive_counter >= max_keepalive:
                        yield f": keepalive\n\n"
                        keepalive_counter = 0
                    
                    time.sleep(0.5)
                    iteration_count += 1
                    
                except GeneratorExit:
                    # Client disconnected
                    print(f"Client disconnected from progress stream for {job_id}")
                    break
                except (BrokenPipeError, ConnectionResetError, OSError) as e:
                    # Client connection issues - exit gracefully
                    print(f"Connection error in progress stream for {job_id}: {e}")
                    break
                except Exception as e:
                    # Log error but continue trying (with limit)
                    print(f"Error in progress stream for {job_id}: {e}")
                    import traceback
                    traceback.print_exc()
                    try:
                        yield f"data: {{\"percent\": {last_percent}, \"step\": \"Error reading progress...\", \"done\": false, \"error\": null}}\n\n"
                    except:
                        break
                    time.sleep(1)
                    iteration_count += 1
                    
        except GeneratorExit:
            # Client disconnected normally
            print(f"Progress stream closed for {job_id}")
        except (BrokenPipeError, ConnectionResetError, OSError) as e:
            # Connection errors - exit silently
            print(f"Connection closed for {job_id}: {e}")
        except Exception as e:
            # Final error - send error message and close
            print(f"Fatal error in progress stream for {job_id}: {e}")
            import traceback
            traceback.print_exc()
            try:
                yield f"data: {{\"percent\": 0, \"step\": \"Connection error\", \"done\": true, \"error\": \"Progress tracking error: {str(e)}\"}}\n\n"
            except:
                pass
    
    response = Response(event_stream(), mimetype="text/event-stream")
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'  # Disable buffering for nginx
    response.headers['Connection'] = 'keep-alive'
    return response


def analyze():
    """Handle analysis request - returns immediately, runs analysis in background thread"""
    try:
        # Get job ID for progress tracking
        job_id = request.form.get('job_id') or str(uuid.uuid4())
        session['job_id'] = job_id
        
        # Capture all form data BEFORE starting thread (Flask request context is thread-local)
        print(f"[DEBUG] Form keys received: {list(request.form.keys())}")
        print(f"[DEBUG] use_national_data from form: '{request.form.get('use_national_data', 'NOT_FOUND')}'")
        form_data = {
            'acquirer_lei': request.form.get('acquirer_lei', '').strip(),
            'acquirer_rssd': request.form.get('acquirer_rssd', '').strip(),
            'acquirer_sb_id': request.form.get('acquirer_sb_id', '').strip(),
            'target_lei': request.form.get('target_lei', '').strip(),
            'target_rssd': request.form.get('target_rssd', '').strip(),
            'target_sb_id': request.form.get('target_sb_id', '').strip(),
            'acquirer_name': request.form.get('acquirer_name', 'Bank A').strip(),
            'target_name': request.form.get('target_name', 'Bank B').strip(),
            'acquirer_assessment_areas': request.form.get('acquirer_assessment_areas', '[]'),
            'target_assessment_areas': request.form.get('target_assessment_areas', '[]'),
            'single_bank_mode': request.form.get('single_bank_mode', '0'),
            'use_national_data': request.form.get('use_national_data', '0'),
            'loan_purpose': request.form.get('loan_purpose', ''),
            'hmda_years': request.form.get('hmda_years', '2023,2024'),
            'sb_years': request.form.get('sb_years', '2023,2024'),
            'action_taken': request.form.get('action_taken', '1'),
            'occupancy_type': request.form.get('occupancy_type', '1'),
            'total_units': request.form.get('total_units', '1-4'),
            'construction_method': request.form.get('construction_method', '1,2'),  # Default: both site-built and manufactured homes
            'not_reverse': request.form.get('not_reverse', '1')
        }
        
        # Initialize progress
        update_progress(job_id, {'percent': 0, 'step': 'Initializing analysis...', 'done': False, 'error': None})
        
        # Run analysis in background thread so server can respond to progress requests
        def run_analysis():
            # Push application context for background thread
            with app.app_context():
                try:
                    _perform_analysis(job_id, form_data)
                except Exception as e:
                    import traceback
                    error_msg = str(e)
                    traceback.print_exc()
                    update_progress(job_id, {'percent': 0, 'step': 'Error occurred', 'done': True, 'error': error_msg})
        
        thread = threading.Thread(target=run_analysis, daemon=True)
        thread.start()
        
        # Return immediately with job_id
        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'Analysis started'
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': error_msg
        }), 500


def _perform_analysis(job_id, form_data):
    """Perform the actual analysis (runs in background thread)"""
    try:
        # Create progress tracker
        progress_tracker = create_progress_tracker(job_id)
        update_progress(job_id, {'percent': 0, 'step': 'Initializing analysis...', 'done': False, 'error': None})
        
        # Get form data from captured dict (handle None values)
        acquirer_lei = (form_data.get('acquirer_lei') or '').strip()
        acquirer_rssd = (form_data.get('acquirer_rssd') or '').strip()
        acquirer_sb_id = (form_data.get('acquirer_sb_id') or '').strip()
        target_lei = (form_data.get('target_lei') or '').strip()
        target_rssd = (form_data.get('target_rssd') or '').strip()
        target_sb_id = (form_data.get('target_sb_id') or '').strip()
        
        # Get bank names (should already be loaded)
        acquirer_name = (form_data.get('acquirer_name') or 'Bank A').strip()
        target_name = (form_data.get('target_name') or 'Bank B').strip()
        
        # Get assessment areas (from form data)
        acquirer_aa_json = form_data.get('acquirer_assessment_areas') or '[]'
        target_aa_json = form_data.get('target_assessment_areas') or '[]'
        
        try:
            acquirer_aa_data = json.loads(acquirer_aa_json) if acquirer_aa_json and isinstance(acquirer_aa_json, str) else []
            target_aa_data = json.loads(target_aa_json) if target_aa_json and isinstance(target_aa_json, str) else []
        except (json.JSONDecodeError, TypeError, AttributeError) as e:
            print(f"Error parsing assessment area JSON: {e}")
            acquirer_aa_data = []
            target_aa_data = []

        # Check if using national level data
        use_national_data_raw = form_data.get('use_national_data', '0')
        use_national_data = use_national_data_raw == '1'
        print(f"[DEBUG] use_national_data_raw = '{use_national_data_raw}', use_national_data = {use_national_data}")

        # Check if single bank mode (no Bank B)
        single_bank_mode_raw = form_data.get('single_bank_mode', '0')
        single_bank_mode = single_bank_mode_raw == '1'
        print(f"[DEBUG] single_bank_mode = {single_bank_mode}")

        # Clear target data if in single bank mode
        if single_bank_mode:
            target_lei = ''
            target_rssd = ''
            target_sb_id = ''
            target_name = ''
            target_aa_data = []

        update_progress(job_id, {'percent': 5, 'step': 'Parsing assessment areas and counties...', 'done': False, 'error': None})
        
        # Extract counties from assessment areas
        # Also handle MSA codes that might be in the counties list
        # IMPORTANT: Preserve assessment area names and ensure counties are properly extracted
        import re
        
        acquirer_counties = []
        acquirer_msa_codes = []
        acquirer_aa_with_names = []  # Keep track of assessment areas with their names
        
        for aa in acquirer_aa_data:
            if isinstance(aa, dict):
                # Try multiple possible keys for assessment area name
                aa_name = aa.get('cbsa_name') or aa.get('name') or aa.get('assessment_area') or aa.get('aa_name') or 'Unnamed Assessment Area'
                counties_list = aa.get('counties') or aa.get('county_list') or aa.get('county') or []
                
                # If counties is a string, try to split it
                if isinstance(counties_list, str):
                    counties_list = [c.strip() for c in counties_list.split(',') if c.strip()]
                
                # Process each county/MSA code in the list
                aa_counties = []
                aa_msa_codes = []
                
                for item in counties_list:
                    if not item:
                        continue
                    
                    # Handle new format: dict with state_code and county_code
                    if isinstance(item, dict):
                        # New format: {"state_code": "12", "county_code": "057"} or {"geoid5": "12057"}
                        aa_counties.append(item)
                        acquirer_counties.append(item)
                    else:
                        # Legacy format: string
                        item_str = str(item).strip()
                        if not item_str:
                            continue
                        
                        # Check if it's an MSA code pattern
                        msa_match = re.search(r'MSA\s+(\d+(?:\s*,\s*\d+)*)', item_str, re.IGNORECASE)
                        if msa_match:
                            msa_numbers = [code.strip() for code in msa_match.group(1).split(',')]
                            aa_msa_codes.extend(msa_numbers)
                            acquirer_msa_codes.extend(msa_numbers)
                        else:
                            aa_counties.append(item_str)
                            acquirer_counties.append(item_str)
                
                # Store assessment area with its counties for reference (even if empty, to track all AAs)
                acquirer_aa_with_names.append({
                    'name': aa_name,
                    'counties': aa_counties,
                    'msa_codes': aa_msa_codes
                })
        
        target_counties = []
        target_msa_codes = []
        target_aa_with_names = []  # Keep track of assessment areas with their names
        
        for aa in target_aa_data:
            if isinstance(aa, dict):
                # Try multiple possible keys for assessment area name
                aa_name = aa.get('cbsa_name') or aa.get('name') or aa.get('assessment_area') or aa.get('aa_name') or 'Unnamed Assessment Area'
                counties_list = aa.get('counties') or aa.get('county_list') or aa.get('county') or []
                
                # If counties is a string, try to split it
                if isinstance(counties_list, str):
                    counties_list = [c.strip() for c in counties_list.split(',') if c.strip()]
                
                # Process each county/MSA code in the list
                aa_counties = []
                aa_msa_codes = []
                
                for item in counties_list:
                    if not item:
                        continue
                    
                    # Handle new format: dict with state_code and county_code
                    if isinstance(item, dict):
                        # New format: {"state_code": "12", "county_code": "057"} or {"geoid5": "12057"}
                        aa_counties.append(item)
                        target_counties.append(item)
                    else:
                        # Legacy format: string
                        item_str = str(item).strip()
                        if not item_str:
                            continue
                        
                        # Check if it's an MSA code pattern
                        msa_match = re.search(r'MSA\s+(\d+(?:\s*,\s*\d+)*)', item_str, re.IGNORECASE)
                        if msa_match:
                            msa_numbers = [code.strip() for code in msa_match.group(1).split(',')]
                            aa_msa_codes.extend(msa_numbers)
                            target_msa_codes.extend(msa_numbers)
                        else:
                            aa_counties.append(item_str)
                            target_counties.append(item_str)
                
                # Store assessment area with its counties for reference (even if empty, to track all AAs)
                target_aa_with_names.append({
                    'name': aa_name,
                    'counties': aa_counties,
                    'msa_codes': aa_msa_codes
                })
        
        # Look up counties for MSA codes
        if acquirer_msa_codes:
            msa_counties_map = get_counties_by_msa_codes(list(set(acquirer_msa_codes)))
            for msa_data in msa_counties_map.values():
                counties_from_msa = msa_data.get('counties', [])
                acquirer_counties.extend(counties_from_msa)
                # Also add to the corresponding assessment areas
                for aa in acquirer_aa_with_names:
                    if aa['msa_codes']:
                        # Check if any of this AA's MSA codes match
                        for msa_code, msa_info in msa_counties_map.items():
                            if msa_code in aa['msa_codes']:
                                aa['counties'].extend(msa_info.get('counties', []))
        
        if target_msa_codes:
            msa_counties_map = get_counties_by_msa_codes(list(set(target_msa_codes)))
            for msa_data in msa_counties_map.values():
                counties_from_msa = msa_data.get('counties', [])
                target_counties.extend(counties_from_msa)
                # Also add to the corresponding assessment areas
                for aa in target_aa_with_names:
                    if aa['msa_codes']:
                        # Check if any of this AA's MSA codes match
                        for msa_code, msa_info in msa_counties_map.items():
                            if msa_code in aa['msa_codes']:
                                aa['counties'].extend(msa_info.get('counties', []))
        
        # Remove duplicates - handle both strings and dicts
        def deduplicate_counties(counties_list):
            """Remove duplicates from a list that may contain strings or dicts"""
            seen = set()
            unique = []
            for county in counties_list:
                if isinstance(county, dict):
                    # Convert dict to hashable tuple for deduplication
                    # Use state_code + county_code or geoid5 as key
                    if 'geoid5' in county:
                        key = ('geoid5', str(county['geoid5']).zfill(5))
                    elif 'state_code' in county and 'county_code' in county:
                        key = ('codes', str(county['state_code']).zfill(2), str(county['county_code']).zfill(3))
                    else:
                        # Fallback: use string representation
                        key = ('dict', str(sorted(county.items())))
                    
                    if key not in seen:
                        seen.add(key)
                        unique.append(county)
                else:
                    # String format - use directly
                    if county not in seen:
                        seen.add(county)
                        unique.append(county)
            return unique
        
        acquirer_counties = deduplicate_counties(acquirer_counties)
        target_counties = deduplicate_counties(target_counties)
        
        # Expand MSA names to counties (if user provided MSA names instead of county lists)
        detect_and_expand_msa_names = _import_local_module('county_mapper', 'detect_and_expand_msa_names')
        
        update_progress(job_id, {'percent': 8, 'step': 'Checking for MSA names and expanding to counties...', 'done': False, 'error': None})
        
        # Create progress callback for MSA expansion
        def msa_progress_callback(progress_data):
            update_progress(job_id, progress_data)
        
        acquirer_counties = detect_and_expand_msa_names(acquirer_counties, progress_callback=msa_progress_callback)
        target_counties = detect_and_expand_msa_names(target_counties, progress_callback=msa_progress_callback)
        
        # Remove duplicates again after expansion
        acquirer_counties = deduplicate_counties(acquirer_counties)
        target_counties = deduplicate_counties(target_counties)
        
        # Get filters from form_data (not request.form - we're in a background thread)
        loan_purpose = form_data.get('loan_purpose') or ''

        # Get HMDA filter values (defaults for multi-select: comma-separated values)
        action_taken = form_data.get('action_taken', '1')
        occupancy_type = form_data.get('occupancy_type', '1')
        total_units = form_data.get('total_units', '1,2,3,4')  # Default to 1-4 units
        construction_method = form_data.get('construction_method', '1,2')  # Default: both site-built and manufactured homes
        not_reverse = form_data.get('not_reverse', '1')
        peer_group = form_data.get('peer_group', 'volume_50_200')  # Peer group selection

        # Parse analysis year ranges (for analysis sheets)
        hmda_start_year = int(form_data.get('hmda_start_year') or 2023)
        hmda_end_year = int(form_data.get('hmda_end_year') or 2024)
        sb_start_year = int(form_data.get('sb_start_year') or 2023)
        sb_end_year = int(form_data.get('sb_end_year') or 2024)

        # Parse baseline year ranges (for goals sheets)
        baseline_hmda_start_year = int(form_data.get('baseline_hmda_start_year') or 2023)
        baseline_hmda_end_year = int(form_data.get('baseline_hmda_end_year') or 2024)
        baseline_sb_start_year = int(form_data.get('baseline_sb_start_year') or 2023)
        baseline_sb_end_year = int(form_data.get('baseline_sb_end_year') or 2024)

        # Generate year lists from ranges
        hmda_years = [str(y) for y in range(hmda_start_year, hmda_end_year + 1)]
        sb_years = [str(y) for y in range(sb_start_year, sb_end_year + 1)]
        baseline_hmda_years = [str(y) for y in range(baseline_hmda_start_year, baseline_hmda_end_year + 1)]
        baseline_sb_years = [str(y) for y in range(baseline_sb_start_year, baseline_sb_end_year + 1)]
        
        update_progress(job_id, {'percent': 10, 'step': 'Mapping counties to GEOIDs...', 'done': False, 'error': None})

        # Map counties to GEOIDs and enrich with metadata
        map_counties_to_geoids, enrich_counties_with_metadata = _import_local_module('county_mapper', 'map_counties_to_geoids', 'enrich_counties_with_metadata')

        # If using national data, get all US counties from BigQuery
        if use_national_data:
            update_progress(job_id, {'percent': 10, 'step': 'Loading all US counties for national analysis...', 'done': False, 'error': None})
            from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query
            client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

            # Query all US county GEOIDs from the cbsa_to_county table
            all_counties_query = """
            SELECT DISTINCT CAST(geoid5 AS STRING) as geoid5
            FROM `justdata-ncrc.shared.cbsa_to_county`
            WHERE geoid5 IS NOT NULL
            ORDER BY geoid5
            """
            try:
                results = execute_query(client, all_counties_query)
                national_geoids = [row['geoid5'].zfill(5) for row in results] if results else []
                print(f"[DEBUG] Loaded {len(national_geoids)} counties for national analysis")

                # Use national geoids for both banks
                acquirer_geoids = national_geoids
                target_geoids = national_geoids
                acquirer_unmapped = []
                target_unmapped = []

                # Create a single "National" assessment area
                acquirer_counties_enriched = [{'county_name': 'All US Counties', 'state_name': 'National', 'geoid5': g} for g in national_geoids[:5]] + [{'county_name': f'...and {len(national_geoids) - 5} more', 'state_name': '', 'geoid5': ''}]
                target_counties_enriched = acquirer_counties_enriched
                acquirer_aa_with_names = [{'cbsa_name': 'National', 'counties': national_geoids}]
                target_aa_with_names = [{'cbsa_name': 'National', 'counties': national_geoids}]

            except Exception as e:
                print(f"[ERROR] Failed to load national counties: {e}")
                update_progress(job_id, {
                    'percent': 0,
                    'step': 'Error loading national data',
                    'done': True,
                    'error': f'Failed to load national counties: {str(e)}'
                })
                return
        else:
            acquirer_geoids, acquirer_unmapped = map_counties_to_geoids(acquirer_counties)
            target_geoids, target_unmapped = map_counties_to_geoids(target_counties)

            # Enrich counties with full metadata (state_name, county_name, geoid5, cbsa_code, cbsa_name)
            acquirer_counties_enriched = enrich_counties_with_metadata(acquirer_counties, acquirer_geoids)
            target_counties_enriched = enrich_counties_with_metadata(target_counties, target_geoids)

        # Combine all GEOIDs for HHI calculation
        all_geoids = list(set(acquirer_geoids + target_geoids))
        
        if not acquirer_geoids and not target_geoids and not use_national_data:
            update_progress(job_id, {
                'percent': 0,
                'step': 'Error: No valid counties found',
                'done': True,
                'error': 'No valid counties found in assessment areas. Please check your county names.'
            })
            return
        
        update_progress(job_id, {'percent': 15, 'step': 'Querying HMDA data for Bank A...', 'done': False, 'error': None})
        
        # Query HMDA data
        build_hmda_subject_query, build_hmda_peer_query, build_sb_subject_query, build_sb_peer_query, build_branch_query, build_branch_market_query, build_branch_details_query = _import_local_module(
            'query_builders', 
            'build_hmda_subject_query', 'build_hmda_peer_query',
            'build_sb_subject_query', 'build_sb_peer_query',
            'build_branch_query', 'build_branch_market_query', 'build_branch_details_query'
        )
        from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query
        import pandas as pd

        client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

        # Bank A HMDA Subject
        bank_a_hmda_subject = pd.DataFrame()
        if acquirer_lei and acquirer_geoids:
            query = build_hmda_subject_query(
                acquirer_lei, acquirer_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_a_hmda_subject = pd.DataFrame(results)

        update_progress(job_id, {'percent': 25, 'step': 'Querying HMDA peer data for Bank A...', 'done': False, 'error': None})

        # Bank A HMDA Peer
        bank_a_hmda_peer = pd.DataFrame()
        if acquirer_lei and acquirer_geoids:
            query = build_hmda_peer_query(
                acquirer_lei, acquirer_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse, peer_group
            )
            results = execute_query(client, query)
            if results:
                bank_a_hmda_peer = pd.DataFrame(results)

        update_progress(job_id, {'percent': 35, 'step': 'Querying HMDA data for Bank B...', 'done': False, 'error': None})

        # Bank B HMDA Subject
        bank_b_hmda_subject = pd.DataFrame()
        if target_lei and target_geoids:
            query = build_hmda_subject_query(
                target_lei, target_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_b_hmda_subject = pd.DataFrame(results)

        update_progress(job_id, {'percent': 45, 'step': 'Querying HMDA peer data for Bank B...', 'done': False, 'error': None})

        # Bank B HMDA Peer
        bank_b_hmda_peer = pd.DataFrame()
        if target_lei and target_geoids:
            query = build_hmda_peer_query(
                target_lei, target_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse, peer_group
            )
            results = execute_query(client, query)
            if results:
                bank_b_hmda_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 55, 'step': 'Querying Small Business data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Small Business Subject
        bank_a_sb_subject = pd.DataFrame()
        if acquirer_sb_id and acquirer_geoids:
            query = build_sb_subject_query(acquirer_sb_id, acquirer_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_a_sb_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 65, 'step': 'Querying Small Business peer data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Small Business Peer
        bank_a_sb_peer = pd.DataFrame()
        if acquirer_sb_id and acquirer_geoids:
            query = build_sb_peer_query(acquirer_sb_id, acquirer_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_a_sb_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 75, 'step': 'Querying Small Business data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Small Business Subject
        bank_b_sb_subject = pd.DataFrame()
        if target_sb_id and target_geoids:
            query = build_sb_subject_query(target_sb_id, target_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_b_sb_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 85, 'step': 'Querying Small Business peer data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Small Business Peer
        bank_b_sb_peer = pd.DataFrame()
        if target_sb_id and target_geoids:
            query = build_sb_peer_query(target_sb_id, target_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_b_sb_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 87, 'step': 'Querying branch data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Branch Data (aggregated) - subject and market
        bank_a_branch = pd.DataFrame()
        bank_a_branch_details = pd.DataFrame()
        if acquirer_rssd and acquirer_geoids:
            # Get subject bank aggregated branch data
            query = build_branch_query(acquirer_rssd, acquirer_geoids, year=2025)
            results = execute_query(client, query)
            subject_branch = pd.DataFrame(results) if results else pd.DataFrame()
            
            # Get market (all other banks) aggregated branch data
            query_market = build_branch_market_query(acquirer_rssd, acquirer_geoids, year=2025)
            results_market = execute_query(client, query_market)
            market_branch = pd.DataFrame(results_market) if results_market else pd.DataFrame()
            
            # Merge subject and market data by CBSA
            if not subject_branch.empty and not market_branch.empty:
                # Merge on cbsa_code
                bank_a_branch = subject_branch.merge(
                    market_branch,
                    on='cbsa_code',
                    how='outer',
                    suffixes=('_subject', '_market')
                )
                # Rename columns to match expected format
                bank_a_branch = bank_a_branch.rename(columns={
                    'total_branches_subject': 'total_branches',
                    'branches_in_lmict_subject': 'branches_in_lmict',
                    'pct_lmict_subject': 'pct_lmict',
                    'branches_in_mmct_subject': 'branches_in_mmct',
                    'pct_mmct_subject': 'pct_mmct',
                    'cbsa_name_subject': 'cbsa_name',
                    'total_branches_market': 'market_total_branches',
                    'branches_in_lmict_market': 'market_branches_in_lmict',
                    'pct_lmict_market': 'market_pct_lmict',
                    'branches_in_mmct_market': 'market_branches_in_mmct',
                    'pct_mmct_market': 'market_pct_mmct'
                })
                # Fill missing values with 0
                bank_a_branch = bank_a_branch.fillna(0)
            elif not subject_branch.empty:
                bank_a_branch = subject_branch.copy()
                # Add empty market columns
                bank_a_branch['market_total_branches'] = 0
                bank_a_branch['market_branches_in_lmict'] = 0
                bank_a_branch['market_pct_lmict'] = 0
                bank_a_branch['market_branches_in_mmct'] = 0
                bank_a_branch['market_pct_mmct'] = 0
            elif not market_branch.empty:
                bank_a_branch = market_branch.copy()
                # Add empty subject columns
                bank_a_branch['total_branches'] = 0
                bank_a_branch['branches_in_lmict'] = 0
                bank_a_branch['pct_lmict'] = 0
                bank_a_branch['branches_in_mmct'] = 0
                bank_a_branch['pct_mmct'] = 0
            
            # Get individual branch details
            query_details = build_branch_details_query(acquirer_rssd, acquirer_geoids, year=2025)
            results_details = execute_query(client, query_details)
            if results_details:
                bank_a_branch_details = pd.DataFrame(results_details)
        
        update_progress(job_id, {'percent': 88, 'step': 'Querying branch data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Branch Data (aggregated) - subject and market
        bank_b_branch = pd.DataFrame()
        bank_b_branch_details = pd.DataFrame()
        if target_rssd and target_geoids:
            # Get subject bank aggregated branch data
            query = build_branch_query(target_rssd, target_geoids, year=2025)
            results = execute_query(client, query)
            subject_branch = pd.DataFrame(results) if results else pd.DataFrame()
            
            # Get market (all other banks) aggregated branch data
            query_market = build_branch_market_query(target_rssd, target_geoids, year=2025)
            results_market = execute_query(client, query_market)
            market_branch = pd.DataFrame(results_market) if results_market else pd.DataFrame()
            
            # Merge subject and market data by CBSA
            if not subject_branch.empty and not market_branch.empty:
                # Merge on cbsa_code
                bank_b_branch = subject_branch.merge(
                    market_branch,
                    on='cbsa_code',
                    how='outer',
                    suffixes=('_subject', '_market')
                )
                # Rename columns to match expected format
                bank_b_branch = bank_b_branch.rename(columns={
                    'total_branches_subject': 'total_branches',
                    'branches_in_lmict_subject': 'branches_in_lmict',
                    'pct_lmict_subject': 'pct_lmict',
                    'branches_in_mmct_subject': 'branches_in_mmct',
                    'pct_mmct_subject': 'pct_mmct',
                    'cbsa_name_subject': 'cbsa_name',
                    'total_branches_market': 'market_total_branches',
                    'branches_in_lmict_market': 'market_branches_in_lmict',
                    'pct_lmict_market': 'market_pct_lmict',
                    'branches_in_mmct_market': 'market_branches_in_mmct',
                    'pct_mmct_market': 'market_pct_mmct'
                })
                # Fill missing values with 0
                bank_b_branch = bank_b_branch.fillna(0)
            elif not subject_branch.empty:
                bank_b_branch = subject_branch.copy()
                # Add empty market columns
                bank_b_branch['market_total_branches'] = 0
                bank_b_branch['market_branches_in_lmict'] = 0
                bank_b_branch['market_pct_lmict'] = 0
                bank_b_branch['market_branches_in_mmct'] = 0
                bank_b_branch['market_pct_mmct'] = 0
            elif not market_branch.empty:
                bank_b_branch = market_branch.copy()
                # Add empty subject columns
                bank_b_branch['total_branches'] = 0
                bank_b_branch['branches_in_lmict'] = 0
                bank_b_branch['pct_lmict'] = 0
                bank_b_branch['branches_in_mmct'] = 0
                bank_b_branch['pct_mmct'] = 0
            
            # Get individual branch details
            query_details = build_branch_details_query(target_rssd, target_geoids, year=2025)
            results_details = execute_query(client, query_details)
            if results_details:
                bank_b_branch_details = pd.DataFrame(results_details)
        
        update_progress(job_id, {'percent': 90, 'step': 'Calculating HHI...', 'done': False, 'error': None})
        
        # Calculate HHI
        hhi_df = pd.DataFrame()
        print(f"[HHI] Starting HHI calculation - RSSDs: {acquirer_rssd}, {target_rssd}, Counties: {len(all_geoids) if all_geoids else 0}")
        if acquirer_rssd and target_rssd and all_geoids:
            calculate_hhi_by_county = _import_local_module('hhi_calculator', 'calculate_hhi_by_county')
            try:
                hhi_df = calculate_hhi_by_county(
                    county_geoids=all_geoids,
                    acquirer_rssd=acquirer_rssd,
                    target_rssd=target_rssd,
                    year=2025
                )
                print(f"[HHI] HHI calculation completed - {len(hhi_df)} counties with data")
            except Exception as e:
                print(f"[HHI] Error calculating HHI: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"[HHI] Skipping HHI calculation - Missing: RSSDs={not (acquirer_rssd and target_rssd)}, Counties={not all_geoids}")
        
        update_progress(job_id, {'percent': 92, 'step': 'Querying Mortgage Goals data...', 'done': False, 'error': None})
        
        # Query HMDA data for Mortgage Goals (by loan purpose type, aggregated by state)
        mortgage_goals_data = {}
        if (acquirer_lei or target_lei) and all_geoids:
            # Get state mapping from counties
            geoid5_list = "', '".join([str(g).zfill(5) for g in all_geoids])
            state_query = f"""
            SELECT DISTINCT 
                LPAD(CAST(geoid5 AS STRING), 5, '0') as geoid5,
                State as state_name
            FROM `justdata-ncrc.shared.cbsa_to_county`
            WHERE LPAD(CAST(geoid5 AS STRING), 5, '0') IN ('{geoid5_list}')
            """
            state_results = execute_query(client, state_query)
            state_map = {row['geoid5']: row.get('state_name', '') for row in state_results} if state_results else {}
            
            # Query for each loan purpose type
            loan_purpose_map = {
                'home_purchase': '1',
                'refinance': '31,32',
                'home_equity': '2,4'
            }
            
            for loan_type, loan_purpose_filter in loan_purpose_map.items():
                combined_dfs = []
                
                # Build county-level query for mortgage goals (not CBSA-level)
                # This allows us to properly break down multi-state CBSAs by county, then aggregate by state
                def build_county_level_hmda_query(lei, geoids, years, loan_purpose, action_taken, occupancy_type, total_units, construction_method, not_reverse):
                    """Build HMDA query aggregated by county (GEOID5) instead of CBSA"""
                    geoid5_list = "', '".join([str(g).zfill(5) for g in geoids])
                    years_list = "', '".join([str(y) for y in years])
                    
                    # Build filters
                    loan_purpose_filter_str = ""
                    if loan_purpose:
                        if ',' in loan_purpose:
                            purposes = [p.strip() for p in loan_purpose.split(',')]
                            purpose_list = "', '".join(purposes)
                            loan_purpose_filter_str = f"AND h.loan_purpose IN ('{purpose_list}')"
                        else:
                            loan_purpose_filter_str = f"AND h.loan_purpose = '{loan_purpose.strip()}'"
                    
                    action_taken_filter = ""
                    if action_taken:
                        if ',' in action_taken:
                            actions = [a.strip() for a in action_taken.split(',')]
                            action_list = "', '".join(actions)
                            action_taken_filter = f"AND h.action_taken IN ('{action_list}')"
                        else:
                            action_taken_filter = f"AND h.action_taken = '{action_taken.strip()}'"
                    
                    occupancy_filter = ""
                    if occupancy_type:
                        if ',' in occupancy_type:
                            occupancies = [o.strip() for o in occupancy_type.split(',')]
                            occupancy_list = "', '".join(occupancies)
                            occupancy_filter = f"AND h.occupancy_type IN ('{occupancy_list}')"
                        else:
                            occupancy_filter = f"AND h.occupancy_type = '{occupancy_type.strip()}'"
                    
                    units_filter = ""
                    if total_units:
                        if ',' in total_units:
                            units = [u.strip() for u in total_units.split(',')]
                            units_list = "', '".join(units)
                            units_filter = f"AND h.total_units IN ('{units_list}')"
                        else:
                            units_filter = f"AND h.total_units = '{total_units.strip()}'"
                    
                    construction_filter = ""
                    if construction_method:
                        if ',' in construction_method:
                            constructions = [c.strip() for c in construction_method.split(',')]
                            construction_list = "', '".join(constructions)
                            construction_filter = f"AND h.construction_method IN ('{construction_list}')"
                        else:
                            construction_filter = f"AND h.construction_method = '{construction_method.strip()}'"
                    
                    reverse_filter = ""
                    if not_reverse:
                        if ',' in not_reverse:
                            reverse_values = [r.strip() for r in not_reverse.split(',')]
                            if '1' in reverse_values and '2' not in reverse_values:
                                reverse_filter = "AND h.reverse_mortgage != '1'"
                            elif '2' in reverse_values and '1' not in reverse_values:
                                reverse_filter = "AND h.reverse_mortgage = '1'"
                        else:
                            if not_reverse == '1':
                                reverse_filter = "AND h.reverse_mortgage != '1'"
                            elif not_reverse == '2':
                                reverse_filter = "AND h.reverse_mortgage = '1'"
                    
                    # Get state name from county crosswalk
                    query = f"""
                    WITH county_state_map AS (
                        SELECT DISTINCT
                            LPAD(CAST(geoid5 AS STRING), 5, '0') as geoid5,
                            State as state_name
                        FROM `justdata-ncrc.shared.cbsa_to_county`
                        WHERE LPAD(CAST(geoid5 AS STRING), 5, '0') IN ('{geoid5_list}')
                    ),
                    filtered_hmda AS (
                        SELECT 
                            h.geoid5,
                            h.loan_amount,
                            -- Use pre-computed boolean flags from de_hmda (convert BOOL to INT)
                            CASE WHEN h.is_lmict THEN 1 ELSE 0 END as is_lmict,
                            CASE WHEN h.is_lmib THEN 1 ELSE 0 END as is_lmib,
                            CASE WHEN h.is_mmct THEN 1 ELSE 0 END as is_mmct,
                            CASE WHEN h.is_hispanic THEN 1 ELSE 0 END as is_hispanic,
                            CASE WHEN h.is_black THEN 1 ELSE 0 END as is_black,
                            CASE WHEN h.is_asian THEN 1 ELSE 0 END as is_asian,
                            CASE WHEN h.is_native_american THEN 1 ELSE 0 END as is_native_american,
                            CASE WHEN h.is_hopi THEN 1 ELSE 0 END as is_hopi
                        FROM `justdata-ncrc.shared.de_hmda` h
                        WHERE CAST(h.activity_year AS STRING) IN ('{years_list}')
                            AND CAST(h.lei AS STRING) = '{lei}'
                            AND h.geoid5 IN ('{geoid5_list}')
                            {loan_purpose_filter_str}
                            {action_taken_filter}
                            {occupancy_filter}
                            {units_filter}
                            {construction_filter}
                            {reverse_filter}
                    )
                    SELECT 
                        csm.state_name,
                        COUNT(*) as total_loans,
                        SUM(fh.is_lmict) as lmict_loans,
                        SUM(fh.is_lmib) as lmib_loans,
                        SUM(CASE WHEN fh.is_lmib = 1 THEN fh.loan_amount ELSE 0 END) as lmib_amount,
                        SUM(fh.is_mmct) as mmct_loans,
                        SUM(CASE WHEN fh.is_mmct = 1 AND fh.is_lmib = 1 THEN 1 ELSE 0 END) as minb_loans,
                        SUM(CASE WHEN fh.is_asian = 1 THEN 1 ELSE 0 END) as asian_loans,
                        SUM(CASE WHEN fh.is_black = 1 THEN 1 ELSE 0 END) as black_loans,
                        SUM(CASE WHEN fh.is_native_american = 1 THEN 1 ELSE 0 END) as native_american_loans,
                        SUM(CASE WHEN fh.is_hopi = 1 THEN 1 ELSE 0 END) as hopi_loans,
                        SUM(CASE WHEN fh.is_hispanic = 1 THEN 1 ELSE 0 END) as hispanic_loans
                    FROM filtered_hmda fh
                    INNER JOIN county_state_map csm
                        ON fh.geoid5 = csm.geoid5
                    GROUP BY csm.state_name
                    ORDER BY csm.state_name
                    """
                    return query
                
                # Query Bank A - using baseline years for goals
                if acquirer_lei and acquirer_geoids:
                    query = build_county_level_hmda_query(
                        acquirer_lei, acquirer_geoids, baseline_hmda_years, loan_purpose_filter,
                        action_taken, occupancy_type, total_units, construction_method, not_reverse
                    )
                    results = execute_query(client, query)
                    if results:
                        df_a = pd.DataFrame(results)
                        combined_dfs.append(df_a)

                # Query Bank B - using baseline years for goals
                if target_lei and target_geoids:
                    query = build_county_level_hmda_query(
                        target_lei, target_geoids, baseline_hmda_years, loan_purpose_filter,
                        action_taken, occupancy_type, total_units, construction_method, not_reverse
                    )
                    results = execute_query(client, query)
                    if results:
                        df_b = pd.DataFrame(results)
                        combined_dfs.append(df_b)
                
                # Combine and aggregate by state (already aggregated by state in query, just need to sum)
                if combined_dfs:
                    df = pd.concat(combined_dfs, ignore_index=True)
                    if not df.empty and 'state_name' in df.columns:
                        # Sum across both banks for each state
                        agg_dict = {
                            'total_loans': 'sum',
                            'lmict_loans': 'sum',
                            'lmib_loans': 'sum',
                            'lmib_amount': 'sum',
                            'mmct_loans': 'sum',
                            'minb_loans': 'sum',
                            'asian_loans': 'sum',
                            'black_loans': 'sum',
                            'native_american_loans': 'sum',
                            'hopi_loans': 'sum',
                            'hispanic_loans': 'sum'
                        }
                        # Only include columns that exist
                        agg_dict = {k: 'sum' for k in agg_dict.keys() if k in df.columns}
                        if agg_dict:
                            mortgage_goals_data[loan_type] = df.groupby('state_name').agg(agg_dict).reset_index()
                        else:
                            mortgage_goals_data[loan_type] = pd.DataFrame()
                    else:
                        mortgage_goals_data[loan_type] = pd.DataFrame()
                else:
                    mortgage_goals_data[loan_type] = pd.DataFrame()
        
        # Query SB data for SB Goals (aggregated by state, only states with branches)
        update_progress(job_id, {'percent': 93, 'step': 'Querying SB Goals data...', 'done': False, 'error': None})
        sb_goals_data = None
        if (acquirer_sb_id or target_sb_id) and all_geoids:
            # Get state mapping from counties (only states where banks have branches)
            geoid5_list = "', '".join([str(g).zfill(5) for g in all_geoids])
            state_query = f"""
            SELECT DISTINCT 
                LPAD(CAST(geoid5 AS STRING), 5, '0') as geoid5,
                State as state_name
            FROM `justdata-ncrc.shared.cbsa_to_county`
            WHERE LPAD(CAST(geoid5 AS STRING), 5, '0') IN ('{geoid5_list}')
            """
            state_results = execute_query(client, state_query)
            state_map = {row['geoid5']: row.get('state_name', '') for row in state_results} if state_results else {}
            states_with_branches = set(state_map.values())
            
            # Build county-level SB query aggregated by state
            def build_county_level_sb_query(sb_id, geoids, years):
                """Build SB query aggregated by county (GEOID5) then by state"""
                geoid5_list = "', '".join([str(g).zfill(5) for g in geoids])
                years_list = "', '".join([str(y) for y in years])
                
                # Extract respondent ID without prefix
                if '-' in sb_id:
                    respondent_id_no_prefix = sb_id.split('-', 1)[-1]
                else:
                    respondent_id_no_prefix = sb_id
                
                query = f"""
                WITH county_state_map AS (
                    SELECT DISTINCT
                        LPAD(CAST(geoid5 AS STRING), 5, '0') as geoid5,
                        State as state_name
                    FROM `justdata-ncrc.shared.cbsa_to_county`
                    WHERE LPAD(CAST(geoid5 AS STRING), 5, '0') IN ('{geoid5_list}')
                ),
                filtered_sb_data AS (
                    SELECT
                        LPAD(CAST(d.geoid5 AS STRING), 5, '0') as geoid5,
                        COALESCE(d.total_loans, d.num_under_100k + d.num_100k_250k + d.num_250k_1m) as sb_loans_count,
                        -- SB amounts are stored in thousands of dollars, convert to actual dollars
                        (d.amt_under_100k + d.amt_100k_250k + d.amt_250k_1m) * 1000 as sb_loans_amount,
                        -- LMICT: Use pre-computed lmi_tract_loans from summary table
                        COALESCE(d.lmi_tract_loans, 0) as lmict_loans_count,
                        -- Estimate LMICT amount proportionally (lmi_tract_loans / total_loans * total_amount)
                        SAFE_MULTIPLY(
                            SAFE_DIVIDE(COALESCE(d.lmi_tract_loans, 0), NULLIF(COALESCE(d.total_loans, d.num_under_100k + d.num_100k_250k + d.num_250k_1m), 0)),
                            (d.amt_under_100k + d.amt_100k_250k + d.amt_250k_1m) * 1000
                        ) as lmict_loans_amount,
                        COALESCE(d.numsbrev_under_1m, 0) as loans_rev_under_1m,
                        COALESCE(d.amtsbrev_under_1m, 0) * 1000 as amount_rev_under_1m
                    FROM `justdata-ncrc.bizsight.sb_county_summary` d
                    INNER JOIN `justdata-ncrc.bizsight.sb_lenders` l
                        ON d.respondent_id = l.sb_resid
                        AND CAST(d.year AS STRING) = l.sb_year
                    WHERE CAST(d.year AS STRING) IN ('{years_list}')
                        AND LPAD(CAST(d.geoid5 AS STRING), 5, '0') IN ('{geoid5_list}')
                        AND (l.sb_resid = '{respondent_id_no_prefix}' OR l.sb_resid = '{sb_id}')
                )
                SELECT
                    csm.state_name,
                    SUM(fs.sb_loans_count) as sb_loans_total,
                    SUM(fs.lmict_loans_count) as lmict_count,
                    SUM(fs.lmict_loans_amount) as lmict_loans_amount,
                    SUM(fs.loans_rev_under_1m) as loans_rev_under_1m_count,
                    SUM(fs.amount_rev_under_1m) as amount_rev_under_1m
                FROM filtered_sb_data fs
                INNER JOIN county_state_map csm
                    ON fs.geoid5 = csm.geoid5
                GROUP BY csm.state_name
                ORDER BY csm.state_name
                """
                return query
            
            combined_sb_dfs = []

            # Query Bank A - using baseline years for goals
            if acquirer_sb_id and acquirer_geoids:
                query = build_county_level_sb_query(acquirer_sb_id, acquirer_geoids, baseline_sb_years)
                results = execute_query(client, query)
                if results:
                    df_a = pd.DataFrame(results)
                    combined_sb_dfs.append(df_a)

            # Query Bank B - using baseline years for goals
            if target_sb_id and target_geoids:
                query = build_county_level_sb_query(target_sb_id, target_geoids, baseline_sb_years)
                results = execute_query(client, query)
                if results:
                    df_b = pd.DataFrame(results)
                    combined_sb_dfs.append(df_b)
            
            # Combine and aggregate by state
            if combined_sb_dfs:
                df = pd.concat(combined_sb_dfs, ignore_index=True)
                if not df.empty and 'state_name' in df.columns:
                    # Sum across both banks for each state
                    # Column names must match what merger_excel_generator expects
                    agg_dict = {
                        'sb_loans_total': 'sum',
                        'lmict_count': 'sum',
                        'lmict_loans_amount': 'sum',
                        'loans_rev_under_1m_count': 'sum',
                        'amount_rev_under_1m': 'sum'
                    }
                    # Only include columns that exist
                    agg_dict = {k: 'sum' for k in agg_dict.keys() if k in df.columns}
                    if agg_dict:
                        sb_goals_data = df.groupby('state_name').agg(agg_dict).reset_index()
                        # Calculate averages
                        if 'lmict_count' in sb_goals_data.columns and 'lmict_loans_amount' in sb_goals_data.columns:
                            sb_goals_data['avg_sb_lmict_loan_amount'] = sb_goals_data.apply(
                                lambda row: row['lmict_loans_amount'] / row['lmict_count']
                                if pd.notna(row['lmict_count']) and row['lmict_count'] > 0 else 0, axis=1
                            )
                        if 'loans_rev_under_1m_count' in sb_goals_data.columns and 'amount_rev_under_1m' in sb_goals_data.columns:
                            sb_goals_data['avg_loan_amt_rum_sb'] = sb_goals_data.apply(
                                lambda row: row['amount_rev_under_1m'] / row['loans_rev_under_1m_count']
                                if pd.notna(row['loans_rev_under_1m_count']) and row['loans_rev_under_1m_count'] > 0 else 0, axis=1
                            )
                    else:
                        sb_goals_data = pd.DataFrame()
                else:
                    sb_goals_data = pd.DataFrame()
            else:
                sb_goals_data = pd.DataFrame()
        
        update_progress(job_id, {'percent': 95, 'step': 'Generating Excel report...', 'done': False, 'error': None})
        
        # Generate Excel file
        create_merger_excel = _import_local_module('excel_generator', 'create_merger_excel')
        
        # Create filename with shortened acquiring bank name
        import re
        acquirer_name_short = clean_bank_name(acquirer_name)
        # Make filesystem-safe: remove special characters, replace spaces with underscores
        acquirer_name_safe = re.sub(r'[^\w\s-]', '', acquirer_name_short)
        acquirer_name_safe = re.sub(r'[\s-]+', '_', acquirer_name_safe)
        acquirer_name_safe = re.sub(r'__+', '_', acquirer_name_safe).strip('_')
        # Limit length to avoid filesystem issues
        if len(acquirer_name_safe) > 50:
            acquirer_name_safe = acquirer_name_safe[:50]
        
        excel_file = OUTPUT_DIR / f'merger_analysis_{acquirer_name_safe}_{job_id}.xlsx'
        excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Build assessment areas dict with assessment area names and enriched counties
        assessment_areas_dict = {
            'acquirer': {'counties': acquirer_counties_enriched},
            'target': {'counties': target_counties_enriched},
            'assessment_areas': acquirer_aa_with_names + target_aa_with_names
        }
        
        metadata = {
            # Analysis years (for main analysis sheets)
            'hmda_years': hmda_years,
            'sb_years': sb_years,
            # Baseline years (for goal-setting sheets)
            'baseline_hmda_years': baseline_hmda_years,
            'baseline_sb_years': baseline_sb_years,
            'loan_purpose': loan_purpose,
            'action_taken': action_taken,
            'occupancy_type': occupancy_type,
            'total_units': total_units,
            'construction_method': construction_method,
            'not_reverse': not_reverse,
            'acquirer_lei': acquirer_lei,
            'target_lei': target_lei,
            'acquirer_rssd': acquirer_rssd,
            'target_rssd': target_rssd,
            'acquirer_sb_id': acquirer_sb_id,
            'target_sb_id': target_sb_id,
            'acquirer_name': acquirer_name,
            'target_name': target_name,
            'acquirer_name_short': acquirer_name_short,
            'single_bank_mode': single_bank_mode,
            'excel_filename': excel_file.name  # Store the actual filename for later retrieval
        }
        
        # Remove 'State' column from all DataFrames before passing to Excel generator
        def remove_state_column(df):
            """Remove 'State' column from DataFrame if it exists."""
            if df is not None and not df.empty:
                cols_to_drop = [col for col in df.columns if str(col).strip().lower() == 'state']
                if cols_to_drop:
                    df = df.drop(columns=cols_to_drop)
            return df
        
        bank_a_hmda_subject = remove_state_column(bank_a_hmda_subject)
        bank_a_hmda_peer = remove_state_column(bank_a_hmda_peer)
        bank_b_hmda_subject = remove_state_column(bank_b_hmda_subject)
        bank_b_hmda_peer = remove_state_column(bank_b_hmda_peer)
        bank_a_sb_subject = remove_state_column(bank_a_sb_subject)
        bank_a_sb_peer = remove_state_column(bank_a_sb_peer)
        bank_b_sb_subject = remove_state_column(bank_b_sb_subject)
        bank_b_sb_peer = remove_state_column(bank_b_sb_peer)
        bank_a_branch = remove_state_column(bank_a_branch)
        bank_b_branch = remove_state_column(bank_b_branch)
        hhi_df = remove_state_column(hhi_df)
        
        # Store branch details in metadata for Excel generator
        if 'branch_details' not in metadata:
            metadata['branch_details'] = {}
        metadata['branch_details']['bank_a'] = bank_a_branch_details.to_dict('records') if not bank_a_branch_details.empty else []
        metadata['branch_details']['bank_b'] = bank_b_branch_details.to_dict('records') if not bank_b_branch_details.empty else []

        # Diagnostic logging for debugging data issues
        print(f"[DEBUG] Data Summary before Excel generation:")
        print(f"[DEBUG] - Bank A LEI: {acquirer_lei}, Bank B LEI: {target_lei}")
        print(f"[DEBUG] - Bank A SB ID: {acquirer_sb_id}, Bank B SB ID: {target_sb_id}")
        print(f"[DEBUG] - Bank A RSSD: {acquirer_rssd}, Bank B RSSD: {target_rssd}")
        print(f"[DEBUG] - Acquirer geoids count: {len(acquirer_geoids) if acquirer_geoids else 0}")
        print(f"[DEBUG] - Target geoids count: {len(target_geoids) if target_geoids else 0}")
        print(f"[DEBUG] - All geoids count: {len(all_geoids) if all_geoids else 0}")
        print(f"[DEBUG] - Bank A HMDA Subject shape: {bank_a_hmda_subject.shape if not bank_a_hmda_subject.empty else 'EMPTY'}")
        print(f"[DEBUG] - Bank B HMDA Subject shape: {bank_b_hmda_subject.shape if not bank_b_hmda_subject.empty else 'EMPTY'}")
        print(f"[DEBUG] - Bank A SB Subject shape: {bank_a_sb_subject.shape if not bank_a_sb_subject.empty else 'EMPTY'}")
        print(f"[DEBUG] - Bank B SB Subject shape: {bank_b_sb_subject.shape if not bank_b_sb_subject.empty else 'EMPTY'}")
        print(f"[DEBUG] - HHI DataFrame shape: {hhi_df.shape if not hhi_df.empty else 'EMPTY'}")
        print(f"[DEBUG] - SB Goals Data: {sb_goals_data.shape if sb_goals_data is not None and not sb_goals_data.empty else 'NONE/EMPTY'}")
        if sb_goals_data is not None and not sb_goals_data.empty:
            print(f"[DEBUG] - SB Goals States: {list(sb_goals_data['state_name'].unique()) if 'state_name' in sb_goals_data.columns else 'N/A'}")
        print(f"[DEBUG] - Mortgage Goals Data: {list(mortgage_goals_data.keys()) if mortgage_goals_data else 'NONE'}")

        create_merger_excel(
            output_path=excel_file,
            bank_a_name=acquirer_name,
            bank_b_name=target_name,
            bank_a_hmda_subject=bank_a_hmda_subject,
            bank_a_hmda_peer=bank_a_hmda_peer,
            bank_b_hmda_subject=bank_b_hmda_subject,
            bank_b_hmda_peer=bank_b_hmda_peer,
            bank_a_sb_subject=bank_a_sb_subject,
            bank_a_sb_peer=bank_a_sb_peer,
            bank_b_sb_subject=bank_b_sb_subject,
            bank_b_sb_peer=bank_b_sb_peer,
            bank_a_branch=bank_a_branch,
            bank_b_branch=bank_b_branch,
            hhi_data=hhi_df,
            assessment_areas=assessment_areas_dict,
            metadata=metadata,
            mortgage_goals_data=mortgage_goals_data,
            sb_goals_data=sb_goals_data if sb_goals_data is not None and not sb_goals_data.empty else None
        )
        
        # Save metadata
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        # Save raw data for AI analysis (more robust than reading from Excel)
        raw_data_file = OUTPUT_DIR / f'merger_raw_data_{job_id}.json'
        raw_data = {
            'years_analyzed': hmda_years,  # For goals calculator to know data period
            'bank_a_hmda_subject': bank_a_hmda_subject.to_dict('records') if not bank_a_hmda_subject.empty else [],
            'bank_a_hmda_peer': bank_a_hmda_peer.to_dict('records') if not bank_a_hmda_peer.empty else [],
            'bank_b_hmda_subject': bank_b_hmda_subject.to_dict('records') if not bank_b_hmda_subject.empty else [],
            'bank_b_hmda_peer': bank_b_hmda_peer.to_dict('records') if not bank_b_hmda_peer.empty else [],
            'bank_a_sb_subject': bank_a_sb_subject.to_dict('records') if not bank_a_sb_subject.empty else [],
            'bank_a_sb_peer': bank_a_sb_peer.to_dict('records') if not bank_a_sb_peer.empty else [],
            'bank_b_sb_subject': bank_b_sb_subject.to_dict('records') if not bank_b_sb_subject.empty else [],
            'bank_b_sb_peer': bank_b_sb_peer.to_dict('records') if not bank_b_sb_peer.empty else [],
            'bank_a_branch': bank_a_branch.to_dict('records') if not bank_a_branch.empty else [],
            'bank_b_branch': bank_b_branch.to_dict('records') if not bank_b_branch.empty else [],
            'hhi_data': hhi_df.to_dict('records') if not hhi_df.empty else [],
        }
        
        # Convert numpy types to native Python types for JSON serialization
        from justdata.shared.analysis.ai_provider import convert_numpy_types
        raw_data = convert_numpy_types(raw_data)
        
        with open(raw_data_file, 'w') as f:
            json.dump(raw_data, f, indent=2, default=str)

        print(f"  Saved raw data to {raw_data_file}")

        # Upload to GCS for persistence across Cloud Run instances
        try:
            upload_json(f'mergermeter/merger_metadata_{job_id}.json', metadata)
            upload_json(f'mergermeter/merger_raw_data_{job_id}.json', raw_data)
            print(f"[GCS] Uploaded analysis files for job {job_id}")
        except Exception as e:
            print(f"[GCS] Warning: Failed to upload to GCS: {e}")
            # Continue - local files still available for same-instance access

        # "Doing something cool" step before completion
        update_progress(job_id, {'percent': 98, 'step': 'Doing something cool...', 'done': False, 'error': None})
        print("\nDoing something cool...")
        
        update_progress(job_id, {'percent': 100, 'step': 'Analysis complete!', 'done': True, 'error': None})
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        
        # Update progress with error
        update_progress(job_id, {'percent': 0, 'step': 'Error occurred', 'done': True, 'error': error_msg})


@app.route('/api/load-bank-names', methods=['POST'])
def load_bank_names():
    """Load bank names from identifiers (LEI, RSSD, or SB Respondent ID)
    
    Uses LEI number to look up bank name from BigQuery hmda.lenders18 table.
    """
    try:
        data = request.get_json()
        acquirer = data.get('acquirer', {})
        target = data.get('target', {})
        
        result = {
            'success': True,
            'acquirer_name': None,
            'target_name': None
        }
        
        # Get bank name from LEI number using BigQuery
        # PROJECT_ID is imported at module level (line 22/26), no need to import again
        from justdata.shared.utils.bigquery_client import get_bigquery_client

        client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')

        # Look up acquirer bank name using LEI (or use provided name from bulk import)
        if acquirer.get('name'):
            # Use provided name from bulk import
            result['acquirer_name'] = acquirer.get('name').strip()
        elif acquirer.get('lei'):
            acquirer_lei = acquirer.get('lei').strip()
            if len(acquirer_lei) == 20:
                acquirer_name = get_bank_name_from_lei(client, acquirer_lei)
                if acquirer_name:
                    result['acquirer_name'] = acquirer_name
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Could not find bank name for LEI: {acquirer_lei}. Please verify the LEI number is correct.'
                    }), 404
        
        # Look up target bank name using LEI (or use provided name from bulk import)
        if target.get('name'):
            # Use provided name from bulk import
            result['target_name'] = target.get('name').strip()
        elif target.get('lei'):
            target_lei = target.get('lei').strip()
            if len(target_lei) == 20:
                target_name = get_bank_name_from_lei(client, target_lei)
                if target_name:
                    result['target_name'] = target_name
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Could not find bank name for LEI: {target_lei}. Please verify the LEI number is correct.'
                    }), 404
        
        # Validate that at least one bank name was found
        if not result['acquirer_name'] and not result['target_name']:
            return jsonify({
                'success': False,
                'error': 'Please provide LEI numbers for both banks. Bank names are looked up using the LEI number.'
            }), 400
        
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def clean_bank_name(bank_name: str) -> str:
    """
    Clean bank name by removing suffixes and converting to uppercase.
    
    Args:
        bank_name: Raw bank name from database
    
    Returns:
        Cleaned bank name in ALL CAPS with suffixes removed
    """
    if not bank_name:
        return ""
    
    import re
    
    # Remove leading "THE" or "The"
    bank_name = re.sub(r'^THE\s+', '', bank_name, flags=re.IGNORECASE).strip()
    bank_name = re.sub(r'^The\s+', '', bank_name, flags=re.IGNORECASE).strip()
    
    # List of suffixes to remove (case-insensitive)
    # IMPORTANT: Do NOT remove "BANK" if it's part of the actual bank name (e.g., "PNC Bank", "First Bank")
    # Only remove it if it's clearly a suffix after other words (e.g., "Some Name Bank")
    suffixes = [
        r',?\s*NATIONAL\s+ASSOCIATION\s*$',
        r',?\s*National\s+Association\s*$',
        r',?\s*N\.?\s*A\.?\s*$',
        r',?\s*N\.A\.\s*$',
        r',?\s*NA\s*$',
        r',?\s*FEDERAL\s+SAVINGS\s+BANK\s*$',
        r',?\s*Federal\s+Savings\s+Bank\s*$',
        r',?\s*FSB\s*$',
        r',?\s*FEDERAL\s+CREDIT\s+UNION\s*$',
        r',?\s*Federal\s+Credit\s+Union\s*$',
        r',?\s*FCU\s*$',
        r',?\s*STATE\s+BANK\s*$',
        r',?\s*State\s+Bank\s*$',
        r',?\s*SAVINGS\s+BANK\s*$',
        r',?\s*Savings\s+Bank\s*$',
        r',?\s*SAVINGS\s+AND\s+LOAN\s*$',
        r',?\s*Savings\s+and\s+Loan\s*$',
        r',?\s*S&L\s*$',
        r',?\s*INC\.?\s*$',
        r',?\s*LLC\.?\s*$',
        r',?\s*CORPORATION\s*$',
        r',?\s*Corporation\s*$',
        r',?\s*CORP\.?\s*$',
        r',?\s*Corp\.?\s*$',
        r',?\s*COMPANY\s*$',
        r',?\s*Company\s*$',
        r',?\s*CO\.?\s*$',
        r',?\s*Co\.?\s*$',
        # DO NOT remove standalone "BANK" - it's often part of the actual name (e.g., "PNC Bank", "First Bank")
        # Only remove "BANK" if it appears after a comma (e.g., "Some Name, Bank")
        r',\s*BANK\s*$',
        r',?\s*THE\s*$',
        r',?\s*The\s*$',
    ]
    
    # Apply patterns repeatedly until no more matches
    changed = True
    iterations = 0
    max_iterations = 10  # Safety limit
    
    while changed and iterations < max_iterations:
        changed = False
        original_name = bank_name
        
        for pattern in suffixes:
            bank_name = re.sub(pattern, '', bank_name, flags=re.IGNORECASE).strip()
            if bank_name != original_name:
                changed = True
                break
        
        iterations += 1
    
    # Final cleanup: remove trailing commas, spaces, and periods
    bank_name = re.sub(r'[,.\s]+$', '', bank_name).strip()
    
    # Convert to ALL CAPS
    bank_name = bank_name.upper()
    
    return bank_name


def get_bank_name_from_lei(client, lei: str) -> str:
    """
    Get bank name from LEI number by querying BigQuery hmda.lenders18 table.
    Cleans the name by removing suffixes and converting to uppercase.
    
    Args:
        client: BigQuery client
        lei: Legal Entity Identifier (20 characters)
    
    Returns:
        Cleaned bank name in ALL CAPS with suffixes removed, or None if not found
    """
    try:
        query = f"""
        SELECT DISTINCT
            respondent_name
        FROM `{PROJECT_ID}.lendsight.lenders18`
        WHERE lei = '{lei}'
        LIMIT 1
        """
        
        query_job = client.query(query)
        results = list(query_job.result())
        
        if results and results[0].respondent_name:
            raw_name = results[0].respondent_name.strip()
            # Clean the bank name
            return clean_bank_name(raw_name)
        
        return None
    except Exception as e:
        print(f"Error querying bank name for LEI {lei}: {e}")
        return None


@app.route('/api/generate-assessment-areas-from-branches', methods=['POST'])
def generate_assessment_areas_from_branches():
    """Generate assessment areas from branch locations for a bank"""
    print(f"[DEBUG] Route /api/generate-assessment-areas-from-branches CALLED")
    print(f"[DEBUG] Request method: {request.method}")
    print(f"[DEBUG] Request path: {request.path}")
    print(f"[DEBUG] Request URL: {request.url}")
    try:
        data = request.get_json()
        print(f"[DEBUG] Request data: {data}")
        rssd = data.get('rssd', '').strip()
        lei = data.get('lei', '').strip()  # LEI needed for 'loans' method
        bank_type = data.get('bank_type', 'acquirer')  # 'acquirer' or 'target'
        year = int(data.get('year', 2025))
        method = data.get('method', 'all_branches')  # 'all_branches', 'deposits', or 'loans'
        min_share = float(data.get('min_share', 0.01))  # Minimum share threshold (default 1%)
        
        if not rssd:
            return jsonify({
                'success': False,
                'error': 'RSSD number is required to generate assessment areas from branches.'
            }), 400
        
        # For 'loans' method, LEI is required (HMDA uses LEI, not RSSD)
        if method == 'loans' and not lei:
            return jsonify({
                'success': False,
                'error': 'LEI is required for the "lending activity" method. HMDA data uses LEI, not RSSD.'
            }), 400
        
        if method not in ['all_branches', 'deposits', 'loans']:
            return jsonify({
                'success': False,
                'error': f"Invalid method '{method}'. Must be 'all_branches', 'deposits', or 'loans'."
            }), 400
        
        # Import branch assessment area generator - use absolute import to avoid relative import issues
        import sys
        from pathlib import Path
        branch_gen_path = Path(__file__).parent / 'branch_assessment_area_generator.py'
        if branch_gen_path.exists():
            import importlib.util
            spec = importlib.util.spec_from_file_location("branch_assessment_area_generator", branch_gen_path)
            branch_gen_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(branch_gen_module)
            generate_aa_from_branches = branch_gen_module.generate_assessment_areas_from_branches
        else:
            # Fallback: try relative import
            try:
                from .branch_assessment_area_generator import generate_assessment_areas_from_branches as generate_aa_from_branches
            except ImportError:
                from branch_assessment_area_generator import generate_assessment_areas_from_branches as generate_aa_from_branches
        
        # Generate assessment areas using the selected method
        assessment_areas = generate_aa_from_branches(
            rssd=rssd,
            year=year,
            method=method,
            min_share=min_share,
            lei=lei if method == 'loans' else None
        )
        
        if not assessment_areas:
            method_descriptions = {
                'all_branches': 'branches',
                'deposits': 'branch deposits',
                'loans': 'loan applications'
            }
            return jsonify({
                'success': False,
                'error': f'No {method_descriptions.get(method, "data")} found for RSSD {rssd} in year {year}. Please verify the RSSD number and year.'
            }), 404
        
        return jsonify({
            'success': True,
            'assessment_areas': assessment_areas,
            'bank_type': bank_type,
            'method': method,
            'count': len(assessment_areas)
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating assessment areas from branches: {error_details}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@app.route('/api/download-assessment-area-template', methods=['GET'])
def download_assessment_area_template():
    """Download CSV template for assessment areas"""
    import csv
    import io
    from flask import Response
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Assessment Area Name',
        'State Code',
        'County Code',
        'County Name',
        'State Name'
    ])
    
    # Write example rows
    examples = [
        ['Tampa-St. Petersburg-Clearwater FL', '12', '057', 'Hillsborough', 'Florida'],
        ['Tampa-St. Petersburg-Clearwater FL', '12', '103', 'Pinellas', 'Florida'],
        ['Tampa-St. Petersburg-Clearwater FL', '12', '101', 'Pasco', 'Florida'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '42', '101', 'Philadelphia', 'Pennsylvania'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '34', '007', 'Camden', 'New Jersey'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '10', '003', 'New Castle', 'Delaware'],
    ]
    
    for row in examples:
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=assessment_area_template.csv'
        }
    )
    
    return response


@app.route('/api/download-bank-identifiers-template', methods=['GET'])
def download_bank_identifiers_template():
    """Download CSV template for bank identifiers (LEI, RSSD, ResID)"""
    import csv
    import io
    from flask import Response
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Bank Name',
        'LEI',
        'RSSD',
        'ResID'
    ])
    
    # Write example rows (using actual examples that match the recommended format)
    examples = [
        ['1ST MERCHANTS BANK', 'S0Q3AHZRL5K6VQE35M07', '17147', '0000004365'],
        ['1ST', 'WKN6AF1FCL7BBYGTGI83', '785473', '0000785473'],
    ]
    
    for row in examples:
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=bank_identifiers_template.csv'
        }
    )
    
    return response


@app.route('/api/upload-assessment-areas', methods=['POST'])
def upload_assessment_areas():
    """Upload and parse assessment area JSON or CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        bank_type = request.form.get('bank_type', 'acquirer')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check file extension - JSON or CSV allowed
        filename = file.filename.lower()
        is_json = filename.endswith('.json')
        is_csv = filename.endswith('.csv')
        
        if not is_json and not is_csv:
            return jsonify({'success': False, 'error': 'Only JSON or CSV files are supported. Please upload a JSON or CSV file with assessment areas.'}), 400
        
        # Check file size (26KB should be fine, but let's be explicit)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'success': False, 'error': f'File too large ({file_size / 1024:.1f}KB). Maximum size is 10MB.'}), 400
        
        # Parse file (JSON or CSV)
        assessment_areas = []
        
        if is_csv:
            # Parse CSV file
            try:
                import csv
                import io
                
                # Read CSV content
                try:
                    csv_content = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    file.seek(0)
                    csv_content = file.read().decode('utf-8-sig')  # Handle BOM
                
                if not csv_content or not csv_content.strip():
                    return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
                
                # Parse CSV
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                rows = list(csv_reader)
                
                if not rows:
                    return jsonify({'success': False, 'error': 'CSV file contains no data rows'}), 400
                
                # Group rows by Assessment Area Name
                aa_dict = {}
                for row in rows:
                    aa_name = row.get('Assessment Area Name', '').strip() or row.get('Assessment Area', '').strip() or row.get('CBSA Name', '').strip()
                    if not aa_name:
                        continue
                    
                    # Get state and county codes
                    state_code = row.get('State Code', '').strip() or row.get('State FIPS', '').strip()
                    county_code = row.get('County Code', '').strip() or row.get('County FIPS', '').strip()
                    
                    if state_code and county_code:
                        county_dict = {
                            'state_code': state_code.zfill(2),
                            'county_code': county_code.zfill(3)
                        }
                        
                        if aa_name not in aa_dict:
                            aa_dict[aa_name] = []
                        
                        # Check if this county is already in the list
                        county_exists = any(
                            c.get('state_code') == county_dict['state_code'] and 
                            c.get('county_code') == county_dict['county_code']
                            for c in aa_dict[aa_name]
                        )
                        
                        if not county_exists:
                            aa_dict[aa_name].append(county_dict)
                
                # Convert to assessment areas list
                for aa_name, counties in aa_dict.items():
                    if counties:
                        assessment_areas.append({
                            'cbsa_name': aa_name,
                            'counties': counties
                        })
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error parsing CSV file: {error_details}")
                return jsonify({'success': False, 'error': f'Error parsing CSV file: {str(e)}'}), 500
        
        elif is_json:
            # Parse JSON file
            try:
                # Try UTF-8 first, then fall back to other encodings
                try:
                    json_data = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    file.seek(0)
                    json_data = file.read().decode('utf-8-sig')  # Handle BOM
                
                if not json_data or not json_data.strip():
                    return jsonify({'success': False, 'error': 'JSON file is empty'}), 400
                
                data = json.loads(json_data)
                
                # Handle different JSON structures
                if isinstance(data, list):
                    # List of assessment areas
                    for item in data:
                        if isinstance(item, dict):
                            cbsa_name = item.get('cbsa_name') or item.get('name') or item.get('assessment_area') or 'Unknown'
                            counties = item.get('counties') or item.get('county_list') or []
                            
                            # Support optimized formats (in order of preference):
                            # Format 1: Direct GEOID5 list (most efficient - no BigQuery lookup needed)
                            # Format 2: List of county dicts with geoid5, state_code, or county_code
                            # Format 3: List of strings (legacy "County, State" format - requires BigQuery lookup)
                            
                            # Check if there's a direct geoids array (most efficient format)
                            if 'geoids' in item:
                                geoids_list = item.get('geoids', [])
                                if isinstance(geoids_list, list):
                                    # Convert to dict format for consistency
                                    counties = [{'geoid5': str(g).zfill(5)} for g in geoids_list if g]
                            elif isinstance(counties, str):
                                counties = [c.strip() for c in counties.split(',') if c.strip()]
                            elif isinstance(counties, list):
                                # Check if it's a list of dicts with codes or GEOIDs
                                processed_counties = []
                                for county_item in counties:
                                    if isinstance(county_item, dict):
                                        # Optimized format: {"geoid5": "12057"} (preferred)
                                        # Or: {"state_code": "12", "county_code": "057"}
                                        processed_counties.append(county_item)
                                    elif isinstance(county_item, str):
                                        # Check if it's a GEOID5 (5 digits)
                                        if county_item.strip().isdigit() and len(county_item.strip()) == 5:
                                            processed_counties.append({'geoid5': county_item.strip()})
                                        else:
                                            # Legacy format: "County, State" (requires BigQuery lookup)
                                            processed_counties.append(county_item)
                                counties = processed_counties
                            
                            assessment_areas.append({
                                'cbsa_name': cbsa_name,
                                'counties': counties
                            })
                elif isinstance(data, dict):
                    # Single assessment area or nested structure
                    if 'assessment_areas' in data:
                        # Nested structure
                        for aa in data['assessment_areas']:
                            cbsa_name = aa.get('cbsa_name') or aa.get('name') or 'Unknown'
                            counties = aa.get('counties') or aa.get('county_list') or []
                            
                            # Support optimized formats (in order of preference):
                            # Format 1: Direct GEOID5 list (most efficient)
                            # Format 2: List of county dicts with geoid5, state_code, or county_code
                            # Format 3: List of strings (legacy format - requires BigQuery lookup)
                            
                            # Check if there's a direct geoids array (most efficient format)
                            if 'geoids' in aa:
                                geoids_list = aa.get('geoids', [])
                                if isinstance(geoids_list, list):
                                    counties = [{'geoid5': str(g).zfill(5)} for g in geoids_list if g]
                            elif isinstance(counties, str):
                                counties = [c.strip() for c in counties.split(',') if c.strip()]
                            elif isinstance(counties, list):
                                processed_counties = []
                                for county_item in counties:
                                    if isinstance(county_item, dict):
                                        processed_counties.append(county_item)
                                    elif isinstance(county_item, str):
                                        # Check if it's a GEOID5 (5 digits)
                                        if county_item.strip().isdigit() and len(county_item.strip()) == 5:
                                            processed_counties.append({'geoid5': county_item.strip()})
                                        else:
                                            processed_counties.append(county_item)
                                counties = processed_counties
                            
                            assessment_areas.append({
                                'cbsa_name': cbsa_name,
                                'counties': counties
                            })
                    else:
                        # Single assessment area
                        cbsa_name = data.get('cbsa_name') or data.get('name') or 'Unknown'
                        counties = data.get('counties') or data.get('county_list') or []
                        
                        # Support optimized formats (in order of preference):
                        # Format 1: Direct GEOID5 list (most efficient)
                        # Format 2: List of county dicts with geoid5, state_code, or county_code
                        # Format 3: List of strings (legacy format - requires BigQuery lookup)
                        
                        # Check if there's a direct geoids array (most efficient format)
                        if 'geoids' in data:
                            geoids_list = data.get('geoids', [])
                            if isinstance(geoids_list, list):
                                counties = [{'geoid5': str(g).zfill(5)} for g in geoids_list if g]
                        elif isinstance(counties, str):
                            counties = [c.strip() for c in counties.split(',') if c.strip()]
                        elif isinstance(counties, list):
                            processed_counties = []
                            for county_item in counties:
                                if isinstance(county_item, dict):
                                    processed_counties.append(county_item)
                                elif isinstance(county_item, str):
                                    # Check if it's a GEOID5 (5 digits)
                                    if county_item.strip().isdigit() and len(county_item.strip()) == 5:
                                        processed_counties.append({'geoid5': county_item.strip()})
                                    else:
                                        processed_counties.append(county_item)
                            counties = processed_counties
                        
                        assessment_areas.append({
                            'cbsa_name': cbsa_name,
                            'counties': counties
                        })
            except json.JSONDecodeError as e:
                error_msg = f'Invalid JSON format: {str(e)}'
                if hasattr(e, 'lineno') and hasattr(e, 'colno'):
                    error_msg += f' (Line {e.lineno}, Column {e.colno})'
                return jsonify({'success': False, 'error': error_msg}), 500
            except UnicodeDecodeError as e:
                return jsonify({'success': False, 'error': f'File encoding error: {str(e)}. Please ensure the file is UTF-8 encoded.'}), 500
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error parsing JSON file: {error_details}")
                return jsonify({'success': False, 'error': f'Error parsing JSON file: {str(e)}'}), 500
        
        if not assessment_areas:
            return jsonify({'success': False, 'error': 'No assessment areas found in JSON file. Please check the file format.'}), 400
        
        return jsonify({
            'success': True,
            'assessment_areas': assessment_areas,
            'bank_type': bank_type,
            'count': len(assessment_areas)
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in upload_assessment_areas: {error_details}")
        return jsonify({'success': False, 'error': f'Upload error: {str(e)}'}), 500


def get_counties_by_msa_codes(msa_codes: List[str]) -> Dict[str, List[str]]:
    """
    Look up counties for given MSA/CBSA codes from BigQuery.
    
    Args:
        msa_codes: List of MSA/CBSA codes (as strings, e.g., ['14500', '19740', '24540'])
        
    Returns:
        Dictionary mapping MSA code to list of county names in "County, State" format
    """
    if not msa_codes:
        return {}
    
    try:
        from justdata.shared.utils.bigquery_client import get_bigquery_client, execute_query
        
        # Format MSA codes for query
        msa_code_list = ', '.join([f"'{code}'" for code in msa_codes])
        
        query = f"""
        SELECT 
            CAST(cbsa_code AS STRING) as msa_code,
            cbsa as msa_name,
            county_state
                        FROM `justdata-ncrc.shared.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) IN ({msa_code_list})
        ORDER BY msa_code, county_state
        """
        
        client = get_bigquery_client(PROJECT_ID, app_name='MERGERMETER')
        results = execute_query(client, query)

        # Group counties by MSA code
        msa_counties = {}
        for row in results:
            msa_code = str(row.get('msa_code', ''))
            county = row.get('county_state', '')
            msa_name = row.get('msa_name', f'MSA {msa_code}')
            
            if msa_code not in msa_counties:
                msa_counties[msa_code] = {
                    'name': msa_name,
                    'counties': []
                }
            
            if county and county not in msa_counties[msa_code]['counties']:
                msa_counties[msa_code]['counties'].append(county)
        
        return msa_counties
        
    except Exception as e:
        print(f"Error looking up MSA codes: {e}")
        import traceback
        traceback.print_exc()
        return {}


def parse_assessment_areas_from_text(text):
    """
    Parse assessment areas from text content.
    Handles structured formats with MMSAs, state abbreviations, and county lists.
    Also handles MSA numbers and MSA names, looking up counties from BigQuery.
    Returns a list of assessment area dictionaries.
    """
    import re
    get_counties_by_msa_name = _import_local_module('county_mapper', 'get_counties_by_msa_name')
    
    assessment_areas = []
    all_counties = set()  # Track all unique counties found
    
    # State abbreviation to full name mapping
    state_abbrev_map = {
        'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas', 'CA': 'California',
        'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware', 'FL': 'Florida', 'GA': 'Georgia',
        'HI': 'Hawaii', 'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa',
        'KS': 'Kansas', 'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
        'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi', 'MO': 'Missouri',
        'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada', 'NH': 'New Hampshire', 'NJ': 'New Jersey',
        'NM': 'New Mexico', 'NY': 'New York', 'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio',
        'OK': 'Oklahoma', 'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
        'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah', 'VT': 'Vermont',
        'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia', 'WI': 'Wisconsin', 'WY': 'Wyoming',
        'DC': 'District of Columbia'
    }
    
    # Pattern 1: State abbreviation followed by colon and county list
    # Example: "PA: Carbon, Lehigh, Northampton" or "NC: Gaston, Iredell, Mecklenburg, Union"
    state_county_pattern = r'([A-Z]{2}):\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*(?:,\s*[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)*)'
    
    # Pattern 2: County names with "County" suffix
    # Example: "Carbon County" or "Lehigh County"
    county_with_suffix_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County'
    
    # Pattern 3: City names that might be counties (like "Baltimore City")
    city_county_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+City'
    
    # Pattern 4: MSA numbers (e.g., "MSA 14500, 19740, 24540" or "MSA 22660")
    msa_pattern = r'MSA\s+(\d+(?:\s*,\s*\d+)*)'
    
    # Split text into lines for better parsing
    lines = text.split('\n')
    current_cbsa = None
    msa_codes_to_lookup = []  # Collect MSA codes for batch lookup
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Skip header rows and metadata
        if any(skip in line for skip in ['Charter Number', 'Appendix', 'Type of Exam', 'Rating and Assessment']):
            continue
        
        # Look for MSA numbers (e.g., "MSA 14500, 19740, 24540")
        msa_match = re.search(msa_pattern, line, re.IGNORECASE)
        if msa_match:
            msa_numbers_str = msa_match.group(1)
            # Extract all MSA codes
            msa_codes = [code.strip() for code in msa_numbers_str.split(',')]
            msa_codes_to_lookup.extend(msa_codes)
            
            # Try to extract MSA name from the line (text before "MSA")
            msa_name_match = re.match(r'^([^MSA]+?)(?:\s+MSA)', line, re.IGNORECASE)
            if msa_name_match:
                current_cbsa = msa_name_match.group(1).strip()
            else:
                current_cbsa = f"MSA {', '.join(msa_codes)}"
        
        # Look for MMSA/CBSA names (lines that might be MSA names)
        # MSA names often contain hyphens and are followed by state abbreviations
        msa_name_match = re.match(r'^([A-Z][A-Za-z\s\-,]+?)(?:\s+\([0-9\-]+\))?(?:\s+MSA)?$', line)
        if msa_name_match and 'non-metro' not in line.lower() and not msa_match:
            # This might be an MSA name
            potential_msa = msa_name_match.group(1).strip()
            # Check if next lines contain counties
            current_cbsa = potential_msa
        
        # Parse state abbreviation: county list format
        # Example: "PA: Carbon, Lehigh, Northampton" or "NC: Gaston, Iredell"
        state_matches = re.finditer(state_county_pattern, line)
        for match in state_matches:
            state_abbrev = match.group(1)
            counties_str = match.group(2)
            
            # Get full state name
            state_name = state_abbrev_map.get(state_abbrev, state_abbrev)
            
            # Split counties by comma
            county_names = [c.strip() for c in counties_str.split(',')]
            
            for county_name in county_names:
                # Clean up county name (remove trailing periods, etc.)
                county_name = county_name.strip('.,;')
                
                # Check if it already has "County" suffix
                if not county_name.endswith(' County'):
                    # Check if it's a city (like "Baltimore City")
                    if ' City' in county_name:
                        full_county = f"{county_name}, {state_name}"
                    else:
                        full_county = f"{county_name} County, {state_name}"
                else:
                    full_county = f"{county_name}, {state_name}"
                
                all_counties.add(full_county)
                
                # Add to assessment area with current CBSA if available
                if current_cbsa:
                    # Find or create assessment area for this CBSA
                    aa = next((a for a in assessment_areas if a['cbsa_name'] == current_cbsa), None)
                    if not aa:
                        aa = {'cbsa_name': current_cbsa, 'counties': []}
                        assessment_areas.append(aa)
                    if full_county not in aa['counties']:
                        aa['counties'].append(full_county)
        
        # Also look for standalone county patterns with "County" suffix
        county_matches = re.finditer(county_with_suffix_pattern, line)
        for match in county_matches:
            county_name = match.group(1)
            # Try to infer state from context (look for state abbreviations nearby)
            # For now, we'll add it to a general list
            # This is a fallback for counties mentioned without state context
            full_county = f"{county_name} County"
            # We'll try to match these later with state context if available
    
    # Look up counties for MSA codes found
    if msa_codes_to_lookup:
        msa_counties_map = get_counties_by_msa_codes(list(set(msa_codes_to_lookup)))
        
        for msa_code, msa_data in msa_counties_map.items():
            msa_name = msa_data['name']
            counties = msa_data['counties']
            
            # Add to all_counties set
            for county in counties:
                all_counties.add(county)
            
            # Create assessment area for this MSA
            aa = {
                'cbsa_name': msa_name,
                'counties': sorted(counties)
            }
            assessment_areas.append(aa)
    
    # Check if any assessment area names are MSA names and expand them
    # If an assessment area has a name but no counties, try to look it up as an MSA name
    for aa in assessment_areas:
        if aa.get('cbsa_name') and (not aa.get('counties') or len(aa.get('counties', [])) == 0):
            # Try to look up the CBSA name as an MSA name
            counties, _ = get_counties_by_msa_name(aa['cbsa_name'])
            if counties:
                aa['counties'] = sorted(counties)
                for county in counties:
                    all_counties.add(county)
                print(f"  Expanded MSA name '{aa['cbsa_name']}' to {len(counties)} counties")
    
    # If we found counties but no structured assessment areas, create a general one
    if all_counties and not assessment_areas:
        # Group counties by state
        counties_by_state = {}
        for county in all_counties:
            if ', ' in county:
                parts = county.split(', ')
                if len(parts) == 2:
                    county_name, state_name = parts
                    if state_name not in counties_by_state:
                        counties_by_state[state_name] = []
                    counties_by_state[state_name].append(county)
        
        # Create assessment areas by state
        for state_name, counties in counties_by_state.items():
            assessment_areas.append({
                'cbsa_name': f'{state_name} Assessment Area',
                'counties': sorted(counties)
            })
    
    # If still no assessment areas but we have counties, create one general area
    if not assessment_areas and all_counties:
        assessment_areas.append({
            'cbsa_name': 'General Assessment Area',
            'counties': sorted(list(all_counties))
        })
    
    # Remove duplicates and sort
    for aa in assessment_areas:
        aa['counties'] = sorted(list(set(aa['counties'])))
    
    return assessment_areas


@app.route('/api/generate-ai-summary', methods=['POST'])
def generate_ai_summary():
    """Generate AI-powered written summary of merger analysis"""
    try:
        data = request.get_json()
        job_id = data.get('job_id')
        
        if not job_id:
            return jsonify({'success': False, 'error': 'Job ID required'}), 400
        
        # Load report data
        excel_filename = get_excel_filename(job_id)
        excel_file = OUTPUT_DIR / excel_filename
        if not excel_file.exists():
            return jsonify({'success': False, 'error': 'Report file not found'}), 404
        
        # Load metadata - try GCS first, then local file
        import pandas as pd
        metadata = {}
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        gcs_metadata_path = f'mergermeter/merger_metadata_{job_id}.json'

        metadata = download_json(gcs_metadata_path)
        if metadata:
            print(f"  Loaded metadata from GCS")
        elif metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            print(f"  Loaded metadata from local file")

        # Load raw data (preferred - more reliable than reading from Excel)
        # Try GCS first, then local file
        raw_data_file = OUTPUT_DIR / f'merger_raw_data_{job_id}.json'
        gcs_raw_data_path = f'mergermeter/merger_raw_data_{job_id}.json'
        raw_data = {}

        raw_data = download_json(gcs_raw_data_path)
        if raw_data:
            print(f"  Loaded raw data from GCS")
        elif raw_data_file.exists():
            try:
                with open(raw_data_file, 'r') as f:
                    raw_data = json.load(f)
                print(f"  Loaded raw data from local file")
            except Exception as e:
                print(f"  Warning: Could not load raw data file: {e}")
                raw_data = {}
        else:
            print(f"  Warning: Raw data not found (GCS or local)")
        
        # Fallback: Read Excel data if raw data not available
        report_data = {}
        has_raw_data = raw_data and any(v for v in raw_data.values() if v and len(v) > 0)
        
        if not has_raw_data:
            print("  [AI Summary] Raw data not available, falling back to reading Excel file...")
            try:
                excel_file_obj = pd.ExcelFile(excel_file)
                
                # Extract key data sheets
                for sheet_name in excel_file_obj.sheet_names:
                    try:
                        df = pd.read_excel(excel_file_obj, sheet_name=sheet_name)
                        if not df.empty:
                            report_data[sheet_name] = df
                            print(f"  [AI Summary] Loaded {len(df)} rows from Excel sheet: {sheet_name}")
                    except Exception as e:
                        print(f"  [AI Summary] Warning: Could not read sheet {sheet_name}: {e}")
                        continue
            except Exception as e:
                print(f"  [AI Summary] Error reading Excel file: {e}")
        else:
            # Convert raw data to DataFrames for analysis
            print("  [AI Summary] Converting raw data to DataFrames...")
            sheet_name_map = {
                'bank_a_hmda_subject': f"{metadata.get('acquirer_name', 'Bank A')} Mortgage Subject",
                'bank_a_hmda_peer': f"{metadata.get('acquirer_name', 'Bank A')} Mortgage Peer",
                'bank_b_hmda_subject': f"{metadata.get('target_name', 'Bank B')} Mortgage Subject",
                'bank_b_hmda_peer': f"{metadata.get('target_name', 'Bank B')} Mortgage Peer",
                'bank_a_sb_subject': f"{metadata.get('acquirer_name', 'Bank A')} Small Business Subject",
                'bank_a_sb_peer': f"{metadata.get('acquirer_name', 'Bank A')} Small Business Peer",
                'bank_b_sb_subject': f"{metadata.get('target_name', 'Bank B')} Small Business Subject",
                'bank_b_sb_peer': f"{metadata.get('target_name', 'Bank B')} Small Business Peer",
                'bank_a_branch': f"{metadata.get('acquirer_name', 'Bank A')} Branch",
                'bank_b_branch': f"{metadata.get('target_name', 'Bank B')} Branch",
                'hhi_data': 'HHI Analysis'
            }
            
            for key, records in raw_data.items():
                if records and len(records) > 0:
                    try:
                        df = pd.DataFrame(records)
                        if not df.empty:
                            sheet_name = sheet_name_map.get(key, key)
                            report_data[sheet_name] = df
                            print(f"  [AI Summary] Loaded {len(df)} rows from raw data: {key} -> {sheet_name}")
                    except Exception as e:
                        print(f"  [AI Summary] Warning: Could not convert {key} to DataFrame: {e}")
                        import traceback
                        traceback.print_exc()
            
            if not report_data:
                print("  [AI Summary] Warning: No data converted from raw_data, falling back to Excel...")
                try:
                    excel_file_obj = pd.ExcelFile(excel_file)
                    for sheet_name in excel_file_obj.sheet_names:
                        try:
                            df = pd.read_excel(excel_file_obj, sheet_name=sheet_name)
                            if not df.empty:
                                report_data[sheet_name] = df
                        except:
                            continue
                except:
                    pass
        
        # Perform statistical analysis
        analyze_sb_lending_distribution, analyze_branch_distribution, analyze_hhi_concentration = _import_local_module(
            'statistical_analysis',
            'analyze_sb_lending_distribution',
            'analyze_branch_distribution',
            'analyze_hhi_concentration'
        )
        
        # Perform statistical analysis on available data
        print(f"  [AI Summary] Performing statistical analysis on {len(report_data)} data sheets...")
        
        # Analyze Small Business lending
        sb_analysis = {}
        for sheet_name, df in report_data.items():
            if 'small business' in sheet_name.lower() and 'subject' in sheet_name.lower() and not df.empty:
                print(f"  [AI Summary] Analyzing SB lending from: {sheet_name}")
                try:
                    sb_analysis = analyze_sb_lending_distribution(df)
                    break
                except Exception as e:
                    print(f"  [AI Summary] Error analyzing SB lending: {e}")
                    sb_analysis = {'summary': f'Error analyzing data: {str(e)}'}
        
        # Analyze branch distribution
        branch_analysis = {}
        for sheet_name, df in report_data.items():
            if 'branch' in sheet_name.lower() and ('bank a' in sheet_name.lower() or 'bank b' in sheet_name.lower()) and not df.empty:
                print(f"  [AI Summary] Analyzing branch distribution from: {sheet_name}")
                try:
                    branch_analysis = analyze_branch_distribution(df)
                    break
                except Exception as e:
                    print(f"  [AI Summary] Error analyzing branch distribution: {e}")
                    branch_analysis = {'summary': f'Error analyzing data: {str(e)}'}
        
        # Analyze HHI
        hhi_analysis = {}
        if 'HHI Analysis' in report_data and not report_data['HHI Analysis'].empty:
            print(f"  [AI Summary] Analyzing HHI concentration...")
            try:
                hhi_analysis = analyze_hhi_concentration(report_data['HHI Analysis'])
            except Exception as e:
                print(f"  [AI Summary] Error analyzing HHI: {e}")
                hhi_analysis = {'summary': f'Error analyzing data: {str(e)}'}
        
        # Generate AI summary
        from justdata.shared.analysis.ai_provider import AIAnalyzer
        
        try:
            analyzer = AIAnalyzer(ai_provider="claude")
        except Exception as e:
            error_msg = str(e)
            if "No API key found" in error_msg or "API key" in error_msg:
                return jsonify({
                    'success': False, 
                    'error': 'Claude API key not configured. Please set CLAUDE_API_KEY or ANTHROPIC_API_KEY environment variable.',
                    'details': error_msg
                }), 500
            else:
                return jsonify({
                    'success': False,
                    'error': f'Failed to initialize AI analyzer: {error_msg}'
                }), 500
        
        # Get data summaries for all categories - use raw data if available, otherwise use report_data
        mortgage_data_summary = ""
        sb_data_summary = ""
        branch_data_summary = ""
        hhi_data_summary = ""
        
        # Helper function to create comprehensive summary from DataFrame
        def create_comprehensive_summary(df, category_name):
            """Create a detailed summary including key statistics."""
            if df is None or df.empty:
                return f"No {category_name} data available."
            
            summary = f"\n{category_name} Data Summary:\n"
            summary += f"Total rows: {len(df)}\n"
            summary += f"Columns: {', '.join(str(col) for col in df.columns[:15])}\n"
            
            # Add key statistics for numeric columns
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                summary += f"\nKey Statistics:\n"
                for col in numeric_cols[:10]:  # Limit to first 10 numeric columns
                    try:
                        col_sum = df[col].sum()
                        col_mean = df[col].mean()
                        summary += f"  {col}: Sum={col_sum:,.0f}, Mean={col_mean:,.2f}\n"
                    except:
                        pass
            
            # Add sample data
            summary += f"\nSample Data (first 5 rows):\n"
            for idx, row in df.head(5).iterrows():
                row_data = []
                for col in df.columns[:8]:  # First 8 columns
                    val = row[col]
                    if pd.isna(val):
                        val = 'N/A'
                    elif isinstance(val, (int, float)):
                        val = f"{val:,.0f}" if val == int(val) else f"{val:,.2f}"
                    else:
                        val = str(val)[:50]  # Truncate long strings
                    row_data.append(f"{col}={val}")
                summary += f"  Row {idx}: {' | '.join(row_data)}\n"
            
            return summary
        
        # Process mortgage data
        mortgage_dfs = []
        for sheet_name, df in report_data.items():
            sheet_lower = sheet_name.lower()
            if 'mortgage' in sheet_lower and ('subject' in sheet_lower or 'peer' in sheet_lower):
                mortgage_dfs.append((sheet_name, df))
        
        if mortgage_dfs:
            mortgage_data_summary = "\nMortgage Lending Data:\n"
            for sheet_name, df in mortgage_dfs:
                mortgage_data_summary += create_comprehensive_summary(df, sheet_name)
        else:
            mortgage_data_summary = "No Mortgage lending data available."
        
        # Process small business data
        sb_dfs = []
        for sheet_name, df in report_data.items():
            sheet_lower = sheet_name.lower()
            if 'small business' in sheet_lower and ('subject' in sheet_lower or 'peer' in sheet_lower):
                sb_dfs.append((sheet_name, df))
        
        if sb_dfs:
            sb_data_summary = "\nSmall Business Lending Data:\n"
            for sheet_name, df in sb_dfs:
                sb_data_summary += create_comprehensive_summary(df, sheet_name)
        else:
            sb_data_summary = "No Small Business lending data available."
        
        # Process branch data
        branch_dfs = []
        for sheet_name, df in report_data.items():
            sheet_lower = sheet_name.lower()
            if 'branch' in sheet_lower and ('bank a' in sheet_lower or 'bank b' in sheet_lower):
                branch_dfs.append((sheet_name, df))
        
        if branch_dfs:
            branch_data_summary = "\nBranch Network Data:\n"
            for sheet_name, df in branch_dfs:
                branch_data_summary += create_comprehensive_summary(df, sheet_name)
        else:
            branch_data_summary = "No Branch data available."
        
        # Process HHI data
        if 'HHI Analysis' in report_data:
            hhi_data_summary = create_comprehensive_summary(report_data['HHI Analysis'], 'HHI Analysis')
        else:
            hhi_data_summary = "No HHI data available."
        
        # Validate we have data before generating summary
        if not report_data:
            return jsonify({
                'success': False, 
                'error': 'No data available for analysis. Please ensure the analysis completed successfully.'
            }), 400
        
        print(f"  [AI Summary] Preparing AI prompt with data from {len(report_data)} sheets...")
        
        # Identify CBSAs with significant peer disparities
        def identify_disparity_cbsas(subject_df, peer_df, metric_col, threshold=0.10):
            """Identify CBSAs where subject bank differs from peers by threshold or more."""
            disparities = []
            if subject_df.empty or peer_df.empty:
                return disparities
            
            # Group by CBSA
            for cbsa in subject_df.get('cbsa_name', pd.Series()).unique():
                if pd.isna(cbsa):
                    continue
                subject_cbsa = subject_df[subject_df.get('cbsa_name', pd.Series()) == cbsa]
                peer_cbsa = peer_df[peer_df.get('cbsa_name', pd.Series()) == cbsa]
                
                if not subject_cbsa.empty and not peer_cbsa.empty:
                    subject_val = subject_cbsa[metric_col].sum() if metric_col in subject_cbsa.columns else 0
                    peer_val = peer_cbsa[metric_col].sum() if metric_col in peer_cbsa.columns else 0
                    
                    if peer_val > 0:
                        disparity = abs(subject_val - peer_val) / peer_val
                        if disparity >= threshold:
                            disparities.append({
                                'cbsa': str(cbsa),
                                'metric': metric_col,
                                'subject': subject_val,
                                'peer': peer_val,
                                'disparity_pct': disparity * 100
                            })
            return disparities
        
        # Find disparities for each bank
        bank_a_mortgage_disparities = []
        bank_b_mortgage_disparities = []
        bank_a_sb_disparities = []
        bank_b_sb_disparities = []
        
        bank_a_hmda_subject_df = report_data.get(f"{metadata.get('acquirer_name', 'Bank A')} Mortgage Subject", pd.DataFrame())
        bank_a_hmda_peer_df = report_data.get(f"{metadata.get('acquirer_name', 'Bank A')} Mortgage Peer", pd.DataFrame())
        bank_b_hmda_subject_df = report_data.get(f"{metadata.get('target_name', 'Bank B')} Mortgage Subject", pd.DataFrame())
        bank_b_hmda_peer_df = report_data.get(f"{metadata.get('target_name', 'Bank B')} Mortgage Peer", pd.DataFrame())
        
        if not bank_a_hmda_subject_df.empty and not bank_a_hmda_peer_df.empty:
            for metric in ['lmict_percentage', 'lmib_percentage', 'mmct_percentage', 'minb_percentage']:
                if metric in bank_a_hmda_subject_df.columns:
                    bank_a_mortgage_disparities.extend(identify_disparity_cbsas(bank_a_hmda_subject_df, bank_a_hmda_peer_df, metric))
        
        if not bank_b_hmda_subject_df.empty and not bank_b_hmda_peer_df.empty:
            for metric in ['lmict_percentage', 'lmib_percentage', 'mmct_percentage', 'minb_percentage']:
                if metric in bank_b_hmda_subject_df.columns:
                    bank_b_mortgage_disparities.extend(identify_disparity_cbsas(bank_b_hmda_subject_df, bank_b_hmda_peer_df, metric))
        
        # Create concise prompt focusing on disparities
        prompt = f"""You are analyzing a bank merger impact assessment. Generate a CONCISE written summary (600-800 words maximum) that follows this structure:

1. MORTGAGE LENDING ANALYSIS - Analyze mortgage lending data for both banks, comparing subject bank performance to peer banks. Identify patterns, underperformance issues, and areas of concern.

2. SMALL BUSINESS LENDING ANALYSIS - Analyze small business lending data for both banks. Discuss distribution patterns, identify underperformance issues, and highlight CBSAs that need attention based on chi-squared statistical tests.

3. BRANCH NETWORK ANALYSIS - Analyze branch network data for both banks. Discuss branch distribution patterns, identify service gaps, and highlight CBSAs with concerning branch distribution based on chi-squared statistical tests.

4. HHI (MARKET CONCENTRATION) ANALYSIS - Analyze HHI data to identify areas with concerning concentration spikes that would result from the merger. Highlight specific counties where HHI increases are problematic.

BANK INFORMATION:
- Bank A (Acquirer): {metadata.get('acquirer_name', 'Unknown')}
- Bank B (Target): {metadata.get('target_name', 'Unknown')}

MORTGAGE LENDING DATA SUMMARY:
{mortgage_data_summary if mortgage_data_summary else 'No Mortgage lending data available.'}

SMALL BUSINESS LENDING DATA SUMMARY:
{sb_data_summary if sb_data_summary else 'No Small Business lending data available.'}

BRANCH NETWORK DATA SUMMARY:
{branch_data_summary if branch_data_summary else 'No Branch data available.'}

HHI CONCENTRATION DATA SUMMARY:
{hhi_data_summary if hhi_data_summary else 'No HHI data available.'}

STATISTICAL ANALYSIS RESULTS:
- Small Business Lending: {sb_analysis.get('summary', 'No data available')}
- Branch Distribution: {branch_analysis.get('summary', 'No data available')}
- HHI Concentration: {hhi_analysis.get('summary', 'No data available')}

CONCERNING AREAS IDENTIFIED:
- Small Business Lending: {len(sb_analysis.get('concerning_cbsas', []))} CBSAs with distribution concerns
- Branch Distribution: {len(branch_analysis.get('concerning_cbsas', []))} CBSAs with branch distribution concerns
- HHI Concentration: {len(hhi_analysis.get('concerning_counties', []))} counties with high concentration spikes

Write a CONCISE professional summary (600-800 words maximum) in plain text (NOT markdown) that follows this exact structure:

SECTION 1: MORTGAGE LENDING ANALYSIS (150-200 words)
For Bank A ({metadata.get('acquirer_name', 'Unknown')}): Identify the top 3-5 CBSAs where the bank significantly underperforms peers in LMI or minority lending. Specify the CBSA name, the metric (e.g., LMICT%, LMIB%, MMCT%), and the percentage difference from peers. For Bank B ({metadata.get('target_name', 'Unknown')}): Identify the top 3-5 CBSAs with significant peer disparities. Be specific with CBSA names and numbers.

KEY MORTGAGE LENDING DISPARITIES IDENTIFIED:
{chr(10).join([f"- {d.get('cbsa', 'Unknown')}: {d.get('metric', 'metric')} - Subject: {d.get('subject', 0):.1f}%, Peer: {d.get('peer', 0):.1f}%, Difference: {d.get('disparity_pct', 0):.1f}%" for d in sorted(bank_a_mortgage_disparities + bank_b_mortgage_disparities, key=lambda x: x.get('disparity_pct', 0), reverse=True)[:10]]) if (bank_a_mortgage_disparities or bank_b_mortgage_disparities) else "No significant disparities identified."}

SECTION 2: SMALL BUSINESS LENDING ANALYSIS (150-200 words)
For each bank, identify the top 3-5 CBSAs where small business lending significantly lags peer performance. Specify the CBSA names and the nature of the disparity (e.g., "Bank A's LMICT small business lending in Denver MSA is 15% below peer average"). Reference the statistical analysis results.

SECTION 3: BRANCH NETWORK ANALYSIS (150-200 words)
For each bank, identify the top 3-5 CBSAs with concerning branch distribution patterns compared to peers. Specify CBSA names and the specific issue (e.g., "Bank B has 40% fewer branches in LMICT areas in Chicago MSA compared to peer banks"). Reference the statistical analysis results.

SECTION 4: HHI (MARKET CONCENTRATION) ANALYSIS (100-150 words)
Identify the top 5 counties where the merger would create the highest HHI increases or push markets into highly concentrated territory (>2500 HHI). Specify county names and the HHI change. If no HHI data is available, state that clearly.

IMPORTANT: 
- Be concise, specific, and data-driven
- Focus on the most significant disparities only
- Use CBSA names (not codes) and county names
- Include specific percentages and numbers where relevant
- Write in plain text format with clear paragraphs
- Do NOT use markdown formatting (no #, ##, **, etc.)
- Use clear section headings in plain text like "MORTGAGE LENDING ANALYSIS" or "BRANCH NETWORK ANALYSIS"."""
        
        try:
            ai_summary = analyzer._call_ai(prompt, max_tokens=4000, temperature=0.3)
        except Exception as e:
            error_msg = str(e)
            print(f"  [AI Summary] Error calling AI API: {error_msg}")
            return jsonify({
                'success': False,
                'error': f'Failed to generate AI summary: {error_msg}',
                'details': 'This may be due to missing API key, API rate limits, or network issues.'
            }), 500
        
        # Convert numpy types to JSON-serializable Python types
        from justdata.shared.analysis.ai_provider import convert_numpy_types

        return jsonify({
            'success': True,
            'summary': ai_summary,
            'statistical_analysis': convert_numpy_types({
                'sb': sb_analysis,
                'branch': branch_analysis,
                'hhi': hhi_analysis
            })
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/report-data')
def report_data():
    """Return report data from Excel file as JSON for web display"""
    try:
        job_id = request.args.get('job_id')
        if not job_id:
            return jsonify({'success': False, 'error': 'Job ID required'}), 400
        
        # Find the Excel file for this job
        excel_filename = get_excel_filename(job_id)
        excel_file = OUTPUT_DIR / excel_filename
        if not excel_file.exists():
            return jsonify({'success': False, 'error': 'Report file not found. The analysis may not have completed yet.'}), 404
        
        # Read Excel file and convert each sheet to JSON
        import pandas as pd
        from justdata.shared.analysis.ai_provider import convert_numpy_types
        
        report_data = {}
        excel_file_obj = pd.ExcelFile(excel_file)
        
        # Use openpyxl to read Excel with data_only=True to get calculated formula values
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Filter out template sheets that shouldn't appear in the web report
        # These are sheets from the Excel template that contain data from other banks
        excluded_sheet_patterns = [
            'sandy spring',
            'atlantic union',
            'firstbank sb goals',  # Template sheets
            'pnc bank sb goals',   # Template sheets
        ]
        
        def should_exclude_sheet(sheet_name):
            """Check if a sheet should be excluded from the web report"""
            sheet_lower = sheet_name.lower()
            for pattern in excluded_sheet_patterns:
                if pattern in sheet_lower:
                    return True
            return False
        
        # Filter sheet names to exclude template sheets
        valid_sheet_names = [s for s in excel_file_obj.sheet_names if not should_exclude_sheet(s)]
        
        if len(valid_sheet_names) < len(excel_file_obj.sheet_names):
            excluded = [s for s in excel_file_obj.sheet_names if should_exclude_sheet(s)]
            print(f"[DEBUG] Excluding {len(excluded)} template sheets from web report: {excluded}")
        
        for sheet_name in valid_sheet_names:
            # Read with pandas first
            df = pd.read_excel(excel_file_obj, sheet_name=sheet_name)
            
            # Remove "Unnamed" columns - filter out columns that contain "Unnamed" in the name
            unnamed_cols = [col for col in df.columns if 'unnamed' in str(col).lower()]
            if unnamed_cols:
                print(f"[DEBUG] Removing {len(unnamed_cols)} 'Unnamed' columns from sheet '{sheet_name}': {unnamed_cols}")
                df = df.drop(columns=unnamed_cols)
            
            # If we have the openpyxl workbook, try to get actual cell values for deposits column
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Find deposits column - check multiple possible names
                deposits_col_idx = None
                deposits_col_name = None
                for idx, col_name in enumerate(df.columns, 1):
                    col_str = str(col_name).lower()
                    if 'deposit' in col_str:
                        deposits_col_idx = idx
                        deposits_col_name = col_name
                        break
                
                # If deposits column found, read values from openpyxl
                if deposits_col_idx:
                    print(f"[DEBUG] Found deposits column '{deposits_col_name}' at index {deposits_col_idx} in sheet '{sheet_name}'")
                    deposits_values = []
                    for row_idx in range(2, min(ws.max_row + 1, len(df) + 2)):  # Start from row 2 (skip header)
                        cell = ws.cell(row=row_idx, column=deposits_col_idx)
                        cell_value = cell.value
                        
                        # Handle different value types - preserve formatted strings
                        if cell_value is None:
                            deposits_values.append(None)
                        elif isinstance(cell_value, (int, float)):
                            # If it's a number, keep it as is
                            deposits_values.append(cell_value)
                        elif isinstance(cell_value, str):
                            # If it's already a string (formatted), keep it
                            deposits_values.append(cell_value)
                        else:
                            # Try to convert to string first, then handle
                            str_value = str(cell_value)
                            # If it looks like a formatted currency, keep it as string
                            if '$' in str_value or 'M' in str_value:
                                deposits_values.append(str_value)
                            else:
                                # Try to convert to number
                                try:
                                    deposits_values.append(float(str_value.replace('$', '').replace('M', '').replace(',', '').strip()))
                                except:
                                    deposits_values.append(str_value)
                        
                        # Update DataFrame with the actual value (preserve formatted strings)
                        if row_idx - 2 < len(df):
                            df.iloc[row_idx - 2, deposits_col_idx - 1] = deposits_values[-1]
                    
                    print(f"[DEBUG] Deposits values from openpyxl (first 10): {deposits_values[:10]}")
            
            # Also check if pandas already read the deposits column correctly
            deposits_cols = [col for col in df.columns if 'deposit' in str(col).lower()]
            if deposits_cols:
                deposits_col = deposits_cols[0]
                print(f"[DEBUG] Deposits column '{deposits_col}' found in DataFrame")
                print(f"[DEBUG] Sample values (first 10): {df[deposits_col].head(10).tolist()}")
                print(f"[DEBUG] Data types: {df[deposits_col].dtype}")
                print(f"[DEBUG] Non-null count: {df[deposits_col].notna().sum()}")
            
            # Convert DataFrame to list of dictionaries
            # Replace NaN with None for JSON serialization, but preserve strings
            df = df.replace({pd.NA: None, float('nan'): None})
            
            # Ensure deposits column values are preserved as strings if they contain formatting
            for col in df.columns:
                if 'deposit' in str(col).lower():
                    # Convert to string to preserve formatting like "$123.4M"
                    # Handle None/NaN values properly
                    df[col] = df[col].apply(lambda x: str(x) if x is not None and pd.notna(x) else '')
                    # Clean up any 'nan' or 'None' strings that might have been created
                    df[col] = df[col].replace(['nan', 'None', 'NaN'], '')
            
            # Preserve exact column order from Excel
            # Read column order directly from openpyxl worksheet to ensure it matches Excel
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Get headers from first row of Excel (row 1)
                excel_headers = []
                if ws.max_row > 0:
                    for col_idx in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col_idx).value
                        if cell_value is not None:
                            header_str = str(cell_value).strip()
                            # Skip "Unnamed" headers
                            if 'unnamed' not in header_str.lower() and header_str:
                                excel_headers.append(header_str)
                        else:
                            # If Excel header is empty, try to get from DataFrame
                            if col_idx <= len(df.columns):
                                df_col_name = str(df.columns[col_idx - 1])
                                # Skip "Unnamed" headers
                                if 'unnamed' not in df_col_name.lower() and df_col_name:
                                    excel_headers.append(df_col_name)
                    
                    # Reorder DataFrame columns to match Excel order
                    if excel_headers and len(excel_headers) > 0:
                        # Match Excel headers to DataFrame columns (case-insensitive, handle whitespace)
                        ordered_columns = []
                        for excel_header in excel_headers:
                            # Find matching column in DataFrame
                            matching_col = None
                            excel_header_clean = excel_header.strip().lower()
                            for df_col in df.columns:
                                if str(df_col).strip().lower() == excel_header_clean:
                                    matching_col = df_col
                                    break
                            
                            if matching_col and matching_col not in ordered_columns:
                                ordered_columns.append(matching_col)
                        
                        # Add any remaining columns that weren't in Excel headers
                        for col in df.columns:
                            if col not in ordered_columns:
                                ordered_columns.append(col)
                        
                        # Reorder DataFrame
                        df = df[ordered_columns]
                        # Use Excel headers for display (preserve exact Excel header text)
                        df.columns = excel_headers[:len(df.columns)]
            
            # Final filter: remove any remaining "Unnamed" columns from headers
            final_headers = []
            final_data_indices = []
            for idx, header in enumerate(df.columns):
                header_str = str(header).strip()
                if 'unnamed' not in header_str.lower() and header_str:
                    final_headers.append(header_str)
                    final_data_indices.append(idx)
            
            # Filter DataFrame to only include non-Unnamed columns
            if final_data_indices:
                df = df.iloc[:, final_data_indices]
                df.columns = final_headers
            
            records = df.to_dict('records')
            # Convert numpy types
            records = convert_numpy_types(records)
            report_data[sheet_name] = {
                'headers': final_headers if final_headers else list(df.columns),  # Use filtered headers
                'data': records
            }
        
        # Define proper sheet order for MergerMeter reports
        # This ensures sheets appear in a logical order in the web report
        preferred_sheet_order = [
            'Assessment Areas',
            'Mortgage Goals',
            'Bank A Mortgage Subject',
            'Bank A Mortgage Peer',
            'Bank B Mortgage Subject',
            'Bank B Mortgage Peer',
            'Bank A Small Business Subject',
            'Bank A Small Business Peer',
            'Bank B Small Business Subject',
            'Bank B Small Business Peer',
            'Bank A Branch',
            'Bank B Branch',
            'HHI Analysis'
        ]
        
        # Reorder sheets according to preferred order
        ordered_report_data = {}
        ordered_sheet_names = []
        
        # First, add sheets in preferred order
        for sheet_name in preferred_sheet_order:
            # Check for variations (e.g., "PNC Bank Mortgage Subject" vs "Bank A Mortgage Subject")
            matching_sheet = None
            for actual_sheet_name in report_data.keys():
                # Check if sheet name contains key words from preferred name
                preferred_keywords = [word.lower() for word in sheet_name.split() if len(word) > 2]
                actual_keywords = [word.lower() for word in actual_sheet_name.split() if len(word) > 2]
                
                # Count matching keywords
                matches = sum(1 for kw in preferred_keywords if any(akw.startswith(kw) or kw.startswith(akw) for akw in actual_keywords))
                if matches >= 2:  # At least 2 keywords match
                    matching_sheet = actual_sheet_name
                    break
            
            if matching_sheet and matching_sheet in report_data:
                ordered_report_data[matching_sheet] = report_data[matching_sheet]
                ordered_sheet_names.append(matching_sheet)
        
        # Add any remaining sheets that weren't in the preferred order
        for sheet_name in report_data.keys():
            if sheet_name not in ordered_report_data:
                ordered_report_data[sheet_name] = report_data[sheet_name]
                ordered_sheet_names.append(sheet_name)
        
        # Try to load metadata if available - try GCS first, then local file
        metadata = {}
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        gcs_metadata_path = f'mergermeter/merger_metadata_{job_id}.json'

        # Try GCS first (works across Cloud Run instances)
        metadata = download_json(gcs_metadata_path)
        if metadata:
            print(f"[DEBUG] Loaded metadata from GCS for job {job_id}")
        elif metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            print(f"[DEBUG] Loaded metadata from local file for job {job_id}")
        else:
            print(f"[DEBUG] Metadata not found (GCS or local) for job {job_id}")
            metadata = {}

        if metadata:
            print(f"[DEBUG] acquirer_sb_id: '{metadata.get('acquirer_sb_id', 'NOT SET')}'")
            print(f"[DEBUG] target_sb_id: '{metadata.get('target_sb_id', 'NOT SET')}'")
        else:
            print(f"[DEBUG] No metadata available for job {job_id}")
        
        # Load HHI data if available (from raw data or Excel HHI Analysis sheet)
        hhi_data = {}
        raw_data_file = OUTPUT_DIR / f'merger_raw_data_{job_id}.json'
        gcs_raw_data_path = f'mergermeter/merger_raw_data_{job_id}.json'

        # Try to load from GCS first, then local file
        raw_data = download_json(gcs_raw_data_path)
        if raw_data:
            print(f"[HHI] Loaded raw_data from GCS for job {job_id}")
        elif raw_data_file.exists():
            try:
                with open(raw_data_file, 'r') as f:
                    raw_data = json.load(f)
                print(f"[HHI] Loaded raw_data from local file for job {job_id}")
            except Exception as e:
                print(f"Error loading raw data file: {e}")
                raw_data = None

        if raw_data and 'hhi_data' in raw_data and raw_data['hhi_data']:
            try:
                hhi_df = pd.DataFrame(raw_data['hhi_data'])
                if not hhi_df.empty:
                    hhi_data = {
                        'headers': list(hhi_df.columns),
                        'data': convert_numpy_types(hhi_df.to_dict('records'))
                    }
                    print(f"[HHI] Loaded HHI data from raw_data: {len(hhi_df)} rows, {len(hhi_df.columns)} columns")
            except Exception as e:
                print(f"Error processing HHI data from raw data: {e}")
                import traceback
                traceback.print_exc()
        
        # Fallback: Try to load from Excel sheet if raw data not available
        if not hhi_data and 'HHI Analysis' in ordered_report_data:
            hhi_sheet_data = ordered_report_data['HHI Analysis']
            # Check if it's already in the right format
            if isinstance(hhi_sheet_data, dict) and 'headers' in hhi_sheet_data and 'data' in hhi_sheet_data:
                hhi_data = hhi_sheet_data
            elif isinstance(hhi_sheet_data, pd.DataFrame):
                # Convert DataFrame to expected format
                if not hhi_sheet_data.empty:
                    hhi_data = {
                        'headers': list(hhi_sheet_data.columns),
                        'data': convert_numpy_types(hhi_sheet_data.to_dict('records'))
                    }
                    print(f"[HHI] Loaded HHI data from Excel sheet: {len(hhi_sheet_data)} rows")
        
        return jsonify({
            'success': True,
            'report': ordered_report_data,
            'sheet_order': ordered_sheet_names,  # Include order for frontend
            'metadata': metadata,
            'hhi_data': hhi_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def get_excel_filename(job_id: str) -> str:
    """
    Get the Excel filename for a given job_id.
    Tries to get from metadata first (GCS then local), then constructs it if needed.
    """
    import re

    metadata = None
    metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
    gcs_metadata_path = f'mergermeter/merger_metadata_{job_id}.json'

    # Try GCS first (works across Cloud Run instances)
    metadata = download_json(gcs_metadata_path)
    if not metadata and metadata_file.exists():
        # Fall back to local file
        try:
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
        except Exception as e:
            print(f"Error loading local metadata for job {job_id}: {e}")

    if metadata:
        # First try: get the stored filename
        if 'excel_filename' in metadata:
            return metadata['excel_filename']

        # Second try: construct from acquirer_name_short
        if 'acquirer_name_short' in metadata:
            acquirer_name_short = metadata['acquirer_name_short']
            acquirer_name_safe = re.sub(r'[^\w\s-]', '', acquirer_name_short)
            acquirer_name_safe = re.sub(r'[\s-]+', '_', acquirer_name_safe)
            acquirer_name_safe = re.sub(r'__+', '_', acquirer_name_safe).strip('_')
            if len(acquirer_name_safe) > 50:
                acquirer_name_safe = acquirer_name_safe[:50]
            return f'merger_analysis_{acquirer_name_safe}_{job_id}.xlsx'

    # Fallback: old format
    return f'merger_analysis_{job_id}.xlsx'


def generate_filename(metadata, extension='.xlsx'):
    """Generate a filename for downloads with NCRC, MergerMeter, bank names, and timestamp"""
    import re
    from datetime import datetime
    
    # Get bank names from metadata
    bank_a_name = metadata.get('acquirer_name', 'BankA')
    bank_b_name = metadata.get('target_name', 'BankB')
    
    # Clean up names for filename (remove special characters, spaces become underscores, remove commas)
    def clean_name(name):
        # Remove commas
        name = name.replace(',', '')
        # Replace spaces and special characters with underscores
        name = re.sub(r'[^\w\s-]', '', name)
        name = re.sub(r'[\s-]+', '_', name)
        return name
    
    bank_a_clean = clean_name(bank_a_name)
    bank_b_clean = clean_name(bank_b_name)
    
    # Build filename: NCRC_MergerMeter_[BankA]_[BankB]_[timestamp]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'NCRC_MergerMeter_{bank_a_clean}_{bank_b_clean}_{timestamp}{extension}'
    
    # Clean up double underscores
    filename = re.sub(r'__+', '_', filename)
    
    return filename


def download():
    """Download the generated Excel file"""
    try:
        job_id = request.args.get('job_id') or session.get('job_id')
        if not job_id:
            return jsonify({'error': 'Job ID required'}), 400
        
        excel_filename = get_excel_filename(job_id)
        excel_file = OUTPUT_DIR / excel_filename
        if not excel_file.exists():
            return jsonify({'error': 'Report file not found. The analysis may not have completed yet.'}), 404
        
        # Try to load metadata for filename generation - try GCS first, then local
        metadata = {}
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        gcs_metadata_path = f'mergermeter/merger_metadata_{job_id}.json'

        metadata = download_json(gcs_metadata_path)
        if not metadata and metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)

        # Generate filename with NCRC, MergerMeter, bank names, and timestamp
        filename = generate_filename(metadata or {}, '.xlsx')
        
        return send_file(
            str(excel_file),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Register standard routes
register_standard_routes(
    app,
    index_handler=index,
    analyze_handler=analyze,
    progress_handler=progress_handler,
    download_handler=download,
    data_handler=None
)

# Register MergerMeter-specific routes
app.add_url_rule('/report', 'report', report, methods=['GET'])
# Note: /report-data and /api/generate-assessment-areas-from-branches are already registered via @app.route decorators above


# ============================================================================
# GOALS CALCULATOR ROUTES
# ============================================================================

@app.route('/goals-calculator')
def goals_calculator():
    """
    CBA Goals Calculator page - interactive tool for setting lending goals.
    Reads data from the completed analysis and allows users to adjust
    improvement percentages to calculate Community Benefits Agreement targets.
    """
    import pandas as pd
    from justdata.shared.analysis.ai_provider import convert_numpy_types

    job_id = request.args.get('job_id') or session.get('job_id')
    if not job_id:
        return render_template('goals_calculator.html',
                             mortgage_data={},
                             sb_data={},
                             bank_name='No Analysis Found',
                             bank_info={},
                             data_years=3,
                             error='No job ID found. Please run a merger analysis first.',
                             version=__version__)

    # Load metadata - try GCS first (works across Cloud Run instances), then local
    metadata = None
    raw_data = None
    metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
    raw_data_file = OUTPUT_DIR / f'merger_raw_data_{job_id}.json'

    print(f"[Goals Calculator] Looking for job_id={job_id}")

    # Try GCS first (works across container instances)
    gcs_metadata_path = f'mergermeter/merger_metadata_{job_id}.json'
    gcs_raw_data_path = f'mergermeter/merger_raw_data_{job_id}.json'

    metadata = download_json(gcs_metadata_path)
    if metadata:
        print(f"[Goals Calculator] Loaded metadata from GCS: {gcs_metadata_path}")
    else:
        # Fall back to local file (same instance)
        print(f"[Goals Calculator] GCS metadata not found, trying local file: {metadata_file}")
        if metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            print(f"[Goals Calculator] Loaded metadata from local file")
        else:
            print(f"[Goals Calculator] Local metadata file not found")

    if not metadata:
        # Data files not found - neither GCS nor local
        return render_template('goals_calculator.html',
                             mortgage_data={},
                             sb_data={},
                             bank_name='Analysis Data Not Found',
                             bank_info={},
                             data_years=3,
                             error=f'Analysis data not found for job {job_id}. This can happen if Cloud Run served this request on a different container instance and GCS storage is unavailable. Please run a new analysis.',
                             version=__version__)

    bank_name = metadata.get('acquirer_name', 'Combined Entity')
    if metadata.get('target_name'):
        bank_name = f"{metadata.get('acquirer_name', 'Bank A')} + {metadata.get('target_name', 'Bank B')}"

    # Build bank info for display (include all identifiers)
    bank_info = {
        'acquirer': {
            'name': metadata.get('acquirer_name', 'N/A'),
            'city': metadata.get('acquirer_city', ''),
            'state': metadata.get('acquirer_state', ''),
            'totalAssets': metadata.get('acquirer_total_assets', 0),
            'lei': metadata.get('acquirer_lei', ''),
            'rssd': metadata.get('acquirer_rssd', ''),
            'respondentId': metadata.get('acquirer_sb_id', '')
        },
        'target': {
            'name': metadata.get('target_name', 'N/A'),
            'city': metadata.get('target_city', ''),
            'state': metadata.get('target_state', ''),
            'totalAssets': metadata.get('target_total_assets', 0),
            'lei': metadata.get('target_lei', ''),
            'rssd': metadata.get('target_rssd', ''),
            'respondentId': metadata.get('target_sb_id', '')
        }
    }

    # Load raw data if available - try GCS first, then local file
    mortgage_data = {}
    sb_data = {}
    data_years = 3  # Default

    # Try GCS first for raw_data (gcs_raw_data_path defined above)
    raw_data = download_json(gcs_raw_data_path)
    if raw_data:
        print(f"[Goals Calculator] Loaded raw_data from GCS: {gcs_raw_data_path}")
    elif raw_data_file.exists():
        # Fall back to local file
        try:
            with open(raw_data_file, 'r') as f:
                raw_data = json.load(f)
            print(f"[Goals Calculator] Loaded raw_data from local file")
        except Exception as e:
            print(f"[Goals Calculator] Error loading local raw_data: {e}")

    if raw_data:
        try:
            # Extract years from data if available
            years_analyzed = raw_data.get('years_analyzed', [])
            if years_analyzed:
                data_years = len(years_analyzed)

            # Process mortgage data by state (new format with HP/Refi/HI breakdowns)
            mortgage_data = _extract_mortgage_goals_data(raw_data)

            # Process small business data by state (new format with LMICT and <$1M rev)
            sb_data = _extract_sb_goals_data(raw_data)

        except Exception as e:
            print(f"Error processing raw data for goals calculator: {e}")
            import traceback
            traceback.print_exc()

    # If no raw data, try to load from Excel
    if not mortgage_data and not sb_data:
        excel_filename = get_excel_filename(job_id)
        excel_file = OUTPUT_DIR / excel_filename
        if excel_file.exists():
            try:
                mortgage_data, sb_data, data_years = _extract_goals_data_from_excel(excel_file)
            except Exception as e:
                print(f"Error loading Excel for goals calculator: {e}")
                import traceback
                traceback.print_exc()

    # Check if single bank mode
    single_bank_mode = metadata.get('single_bank_mode', False)

    return render_template('goals_calculator.html',
                         mortgage_data=mortgage_data,
                         sb_data=sb_data,
                         bank_name=bank_name,
                         bank_info=bank_info,
                         data_years=data_years,
                         single_bank_mode=single_bank_mode,
                         job_id=job_id,
                         version=__version__,
                         app_name='MergerMeter',
                         breadcrumb_items=[
                             {'name': 'MergerMeter', 'url': '/mergermeter'},
                             {'name': 'CBA Goals Calculator', 'url': '#'}
                         ])


def _extract_mortgage_goals_data(raw_data):
    """
    Extract mortgage metrics by state from raw data for goals calculator.

    New format matches the React design with metrics broken down by loan purpose (HP/Refi/HI):
    {
        'Grand Total': {
            'Loans': { hp: x, refi: y, hi: z },
            '~LMICT': { hp: x, refi: y, hi: z },
            ...
        },
        'Illinois': { ... },
        ...
    }

    Note: HMDA data may be aggregated by CBSA (not state), and may not have loan purpose breakdown.
    In those cases, we aggregate to Grand Total and put all in HP category.
    """
    mortgage_data = {}

    # Get HMDA subject data from both banks
    acquirer_subject = raw_data.get('bank_a_hmda_subject', [])
    target_subject = raw_data.get('bank_b_hmda_subject', [])

    # Debug: Log what data is available
    print(f"[Goals Calculator] HMDA acquirer records: {len(acquirer_subject)}, target records: {len(target_subject)}")
    if acquirer_subject:
        print(f"[Goals Calculator] HMDA data fields: {list(acquirer_subject[0].keys())[:20]}")
        # Log first 3 sample records with TYPE information to see actual values
        for i, sample in enumerate(acquirer_subject[:3]):
            lmib_val = sample.get('lmib_amount')
            loan_cat = sample.get('loan_purpose_cat')
            print(f"[Goals Calculator] Sample HMDA record {i+1}: loan_purpose_cat={loan_cat}, total_loans={sample.get('total_loans')}, lmib_loans={sample.get('lmib_loans')}")

        # Count loan_purpose_cat distribution
        from collections import Counter
        all_records = acquirer_subject + target_subject
        cat_counts = Counter(r.get('loan_purpose_cat') for r in all_records)
        print(f"[Goals Calculator] loan_purpose_cat distribution: {dict(cat_counts)}")

    # CBSA to State mapping (first 2 digits of CBSA = state FIPS)
    # Common state FIPS codes
    state_fips_map = {
        '01': 'Alabama', '02': 'Alaska', '04': 'Arizona', '05': 'Arkansas',
        '06': 'California', '08': 'Colorado', '09': 'Connecticut', '10': 'Delaware',
        '11': 'District of Columbia', '12': 'Florida', '13': 'Georgia', '15': 'Hawaii',
        '16': 'Idaho', '17': 'Illinois', '18': 'Indiana', '19': 'Iowa',
        '20': 'Kansas', '21': 'Kentucky', '22': 'Louisiana', '23': 'Maine',
        '24': 'Maryland', '25': 'Massachusetts', '26': 'Michigan', '27': 'Minnesota',
        '28': 'Mississippi', '29': 'Missouri', '30': 'Montana', '31': 'Nebraska',
        '32': 'Nevada', '33': 'New Hampshire', '34': 'New Jersey', '35': 'New Mexico',
        '36': 'New York', '37': 'North Carolina', '38': 'North Dakota', '39': 'Ohio',
        '40': 'Oklahoma', '41': 'Oregon', '42': 'Pennsylvania', '44': 'Rhode Island',
        '45': 'South Carolina', '46': 'South Dakota', '47': 'Tennessee', '48': 'Texas',
        '49': 'Utah', '50': 'Vermont', '51': 'Virginia', '53': 'Washington',
        '54': 'West Virginia', '55': 'Wisconsin', '56': 'Wyoming', '72': 'Puerto Rico',
    }

    # Helper function to get value with multiple possible field names
    # IMPORTANT: Handles NaN values which pass `!= 0` check but break sums
    import math
    def get_value(record, *field_names):
        for name in field_names:
            val = record.get(name)
            if val is not None and val != 0:
                # Check for NaN - critical because NaN != 0 is True but NaN + x = NaN
                if isinstance(val, float) and math.isnan(val):
                    return 0.0
                return float(val) if isinstance(val, (int, float)) else val
        return 0

    # State code to state name mapping
    state_code_to_name = {
        'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
        'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
        'DC': 'District of Columbia', 'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii',
        'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa',
        'KS': 'Kansas', 'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine',
        'MD': 'Maryland', 'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota',
        'MS': 'Mississippi', 'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska',
        'NV': 'Nevada', 'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico',
        'NY': 'New York', 'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio',
        'OK': 'Oklahoma', 'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island',
        'SC': 'South Carolina', 'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas',
        'UT': 'Utah', 'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington',
        'WV': 'West Virginia', 'WI': 'Wisconsin', 'WY': 'Wyoming', 'PR': 'Puerto Rico',
        'VI': 'Virgin Islands', 'GU': 'Guam', 'AS': 'American Samoa', 'MP': 'Northern Mariana Islands',
    }

    # Helper to determine state from record
    def get_state(record):
        # Try state_code first (from new query format)
        state_code = record.get('state_code')
        if state_code:
            # Normalize to string and pad to 2 digits
            state_code_str = str(state_code).strip().zfill(2)

            # Try numeric FIPS code first (e.g., '01', '06')
            if state_code_str in state_fips_map:
                return state_fips_map[state_code_str]

            # Try 2-letter abbreviation (e.g., 'AL', 'CA')
            state_code_upper = state_code_str.upper()
            if state_code_upper in state_code_to_name:
                return state_code_to_name[state_code_upper]

        # Try direct state fields
        state = (record.get('state_name') or record.get('state') or record.get('State'))
        if state and state != 'Unknown':
            return state

        return None  # Will aggregate to Grand Total only

    # Track lmib_amount values for debugging
    lmib_debug_values = []

    # Initialize empty metric structure
    def empty_metrics():
        return {
            'Loans': {'hp': 0, 'refi': 0, 'hi': 0},
            '~LMICT': {'hp': 0, 'refi': 0, 'hi': 0},
            '~LMIB': {'hp': 0, 'refi': 0, 'hi': 0},
            'LMIB$': {'hp': 0, 'refi': 0, 'hi': 0},
            '~MMCT': {'hp': 0, 'refi': 0, 'hi': 0},
            '~MINB': {'hp': 0, 'refi': 0, 'hi': 0},
            '~Asian': {'hp': 0, 'refi': 0, 'hi': 0},
            '~Black': {'hp': 0, 'refi': 0, 'hi': 0},
            '~Native American': {'hp': 0, 'refi': 0, 'hi': 0},
            '~HoPI': {'hp': 0, 'refi': 0, 'hi': 0},
            '~Hispanic': {'hp': 0, 'refi': 0, 'hi': 0},
        }

    # Process all records - aggregate by state or into Grand Total
    state_data = {}
    grand_total = empty_metrics()

    for record in acquirer_subject + target_subject:
        state = get_state(record)

        # Determine loan purpose from loan_purpose_cat field (from new query format)
        # Falls back to legacy loan_purpose field for older data
        # Categories: hp=Home Purchase, refi=Refinance/Cash-out, hi=Home Improvement, other=Other purposes
        loan_purpose_cat = record.get('loan_purpose_cat')
        if loan_purpose_cat in ('hp', 'refi', 'hi'):
            purpose_key = loan_purpose_cat
        elif loan_purpose_cat == 'other':
            # Skip 'other' purposes - don't include in goals (or include as 'hp' if needed)
            continue  # Skip this record for goals calculation
        else:
            # Legacy fallback - try old loan_purpose field
            # HMDA codes: 1=HP, 2=HI, 31=Refi, 32=Cash-out Refi, 4/5=Other
            loan_purpose = record.get('loan_purpose', record.get('purpose', ''))
            if loan_purpose:
                loan_purpose_str = str(loan_purpose).strip()
                if loan_purpose_str in ('31', '32') or 'refi' in loan_purpose_str.lower():
                    purpose_key = 'refi'
                elif loan_purpose_str == '2' or 'improv' in loan_purpose_str.lower():
                    purpose_key = 'hi'
                elif loan_purpose_str == '1' or 'purchase' in loan_purpose_str.lower():
                    purpose_key = 'hp'
                else:
                    # Other purposes (4, 5) - skip for goals
                    continue
            else:
                # No loan purpose - default to HP
                purpose_key = 'hp'

        # Get metrics from record
        total_loans = get_value(record, 'total_loans', 'Total Loans', 'loans', 'Loans', 'loan_count')
        total_amount = get_value(record, 'total_amount', 'Total Amount', 'amount', 'Amount', 'loan_amount')
        lmict_loans = get_value(record, 'lmict_loans', 'lmict_count', 'lmi_tract_loans', 'LMICT Loans')
        lmib_loans = get_value(record, 'lmib_loans', 'lmib_count', 'lmi_borrower_loans', 'LMIB Loans')
        lmib_amount = get_value(record, 'lmib_amount', 'lmi_borrower_amount', 'LMIB Amount', 'lmib_dollars')

        # Debug: Track lmib_amount values (first 10 non-zero)
        if len(lmib_debug_values) < 10 and (record.get('lmib_amount') is not None or lmib_amount > 0):
            lmib_debug_values.append({
                'raw': record.get('lmib_amount'),
                'extracted': lmib_amount,
                'lmib_loans': record.get('lmib_loans')
            })

        mmct_loans = get_value(record, 'mmct_loans', 'mmct_count', 'minority_tract_loans', 'MMCT Loans')
        minb_loans = get_value(record, 'minb_loans', 'minority_borrower_loans', 'MINB Loans', 'minority_loans')
        asian_loans = get_value(record, 'asian_loans', 'asian_count', 'Asian Loans')
        black_loans = get_value(record, 'black_loans', 'black_count', 'Black Loans', 'african_american_loans')
        native_loans = get_value(record, 'native_american_loans', 'native_loans', 'Native American Loans', 'american_indian_loans')
        hopi_loans = get_value(record, 'hopi_loans', 'pacific_islander_loans', 'HoPI Loans', 'hawaiian_pacific_islander_loans')
        hispanic_loans = get_value(record, 'hispanic_loans', 'hispanic_count', 'Hispanic Loans', 'latino_loans')

        # If we don't have LMIB$ but have LMIB count, estimate the dollar amount
        if lmib_amount == 0 and lmib_loans > 0 and total_loans > 0:
            avg_loan = total_amount / total_loans
            lmib_amount = lmib_loans * avg_loan

        # Aggregate to Grand Total (always)
        grand_total['Loans'][purpose_key] += total_loans
        grand_total['~LMICT'][purpose_key] += lmict_loans
        grand_total['~LMIB'][purpose_key] += lmib_loans
        grand_total['LMIB$'][purpose_key] += lmib_amount
        grand_total['~MMCT'][purpose_key] += mmct_loans
        grand_total['~MINB'][purpose_key] += minb_loans
        grand_total['~Asian'][purpose_key] += asian_loans
        grand_total['~Black'][purpose_key] += black_loans
        grand_total['~Native American'][purpose_key] += native_loans
        grand_total['~HoPI'][purpose_key] += hopi_loans
        grand_total['~Hispanic'][purpose_key] += hispanic_loans

        # Also aggregate by state if we have state info
        if state:
            if state not in state_data:
                state_data[state] = empty_metrics()

            state_data[state]['Loans'][purpose_key] += total_loans
            state_data[state]['~LMICT'][purpose_key] += lmict_loans
            state_data[state]['~LMIB'][purpose_key] += lmib_loans
            state_data[state]['LMIB$'][purpose_key] += lmib_amount
            state_data[state]['~MMCT'][purpose_key] += mmct_loans
            state_data[state]['~MINB'][purpose_key] += minb_loans
            state_data[state]['~Asian'][purpose_key] += asian_loans
            state_data[state]['~Black'][purpose_key] += black_loans
            state_data[state]['~Native American'][purpose_key] += native_loans
            state_data[state]['~HoPI'][purpose_key] += hopi_loans
            state_data[state]['~Hispanic'][purpose_key] += hispanic_loans

    # Build final result with Grand Total first
    mortgage_data['Grand Total'] = grand_total

    # Add state-level data (excluding 'Unknown' which shouldn't exist now)
    for state, metrics in sorted(state_data.items()):
        if state and state != 'Unknown':
            mortgage_data[state] = metrics

    # Debug: Log key metrics
    lmib_total = grand_total['LMIB$']['hp'] + grand_total['LMIB$']['refi'] + grand_total['LMIB$']['hi']
    lmib_count = grand_total['~LMIB']['hp'] + grand_total['~LMIB']['refi'] + grand_total['~LMIB']['hi']
    print(f"[Goals Calculator] Mortgage data summary: {len(mortgage_data)} regions")
    print(f"[Goals Calculator]   - Total Loans: HP={grand_total['Loans']['hp']}, Refi={grand_total['Loans']['refi']}, HI={grand_total['Loans']['hi']}")
    print(f"[Goals Calculator]   - LMIB Count: HP={grand_total['~LMIB']['hp']}, Refi={grand_total['~LMIB']['refi']}, HI={grand_total['~LMIB']['hi']} (total: {lmib_count})")
    print(f"[Goals Calculator]   - LMIB$: HP=${grand_total['LMIB$']['hp']:,.0f}, Refi=${grand_total['LMIB$']['refi']:,.0f}, HI=${grand_total['LMIB$']['hi']:,.0f} (total: ${lmib_total:,.0f})")

    # Debug: Show sample lmib_amount values
    if lmib_debug_values:
        print(f"[Goals Calculator] LMIB$ DEBUG - Sample values (first {len(lmib_debug_values)}):")
        for i, val in enumerate(lmib_debug_values[:5]):
            print(f"[Goals Calculator]   Record {i+1}: raw={val['raw']}, extracted={val['extracted']}, lmib_loans={val['lmib_loans']}")

    return mortgage_data


def _extract_sb_goals_data(raw_data):
    """
    Extract small business metrics by state from raw data for goals calculator.

    New format matches the React design:
    {
        'Grand Total': {
            'SB Loans': 5597,           # Total SB loan count
            '#LMICT': 196,              # Loans in LMI Census Tracts count
            'Avg SB LMICT Loan Amount': 208551,
            'Loans Rev Under $1m': 1953, # Loans to businesses with <$1M revenue
            'Avg Loan Amt for <$1M GAR SB': 188838,
        },
        'New Jersey': { ... },
        ...
    }
    """
    sb_data = {}

    # Get SB subject data from both banks (using actual keys from raw_data)
    acquirer_subject = raw_data.get('bank_a_sb_subject', [])
    target_subject = raw_data.get('bank_b_sb_subject', [])

    # Debug: Log what fields are available
    if acquirer_subject:
        print(f"[Goals Calculator] SB data fields: {list(acquirer_subject[0].keys())[:20]}")
        # Log sample state_code values
        sample_states = [r.get('state_code') for r in acquirer_subject[:5]]
        print(f"[Goals Calculator] SB sample state_codes: {sample_states}")

    # State FIPS code to state name mapping (same as mortgage data)
    state_fips_map = {
        '01': 'Alabama', '02': 'Alaska', '04': 'Arizona', '05': 'Arkansas',
        '06': 'California', '08': 'Colorado', '09': 'Connecticut', '10': 'Delaware',
        '11': 'District of Columbia', '12': 'Florida', '13': 'Georgia', '15': 'Hawaii',
        '16': 'Idaho', '17': 'Illinois', '18': 'Indiana', '19': 'Iowa',
        '20': 'Kansas', '21': 'Kentucky', '22': 'Louisiana', '23': 'Maine',
        '24': 'Maryland', '25': 'Massachusetts', '26': 'Michigan', '27': 'Minnesota',
        '28': 'Mississippi', '29': 'Missouri', '30': 'Montana', '31': 'Nebraska',
        '32': 'Nevada', '33': 'New Hampshire', '34': 'New Jersey', '35': 'New Mexico',
        '36': 'New York', '37': 'North Carolina', '38': 'North Dakota', '39': 'Ohio',
        '40': 'Oklahoma', '41': 'Oregon', '42': 'Pennsylvania', '44': 'Rhode Island',
        '45': 'South Carolina', '46': 'South Dakota', '47': 'Tennessee', '48': 'Texas',
        '49': 'Utah', '50': 'Vermont', '51': 'Virginia', '53': 'Washington',
        '54': 'West Virginia', '55': 'Wisconsin', '56': 'Wyoming', '72': 'Puerto Rico',
    }

    # Helper function to get value with multiple possible field names
    # IMPORTANT: Handles NaN values which pass `!= 0` check but break sums
    import math
    def get_value(record, *field_names):
        for name in field_names:
            val = record.get(name)
            if val is not None and val != 0:
                # Check for NaN - critical because NaN != 0 is True but NaN + x = NaN
                if isinstance(val, float) and math.isnan(val):
                    return 0.0
                return float(val) if isinstance(val, (int, float)) else val
        return 0

    # Helper to determine state from record (same logic as mortgage data)
    def get_state(record):
        # Try state_code first (from query - numeric FIPS code)
        state_code = record.get('state_code')
        if state_code:
            # Normalize to string and pad to 2 digits
            state_code_str = str(state_code).strip().zfill(2)
            if state_code_str in state_fips_map:
                return state_fips_map[state_code_str]

        # Try direct state name fields
        state = (record.get('state_name') or record.get('state') or record.get('State'))
        if state and state != 'Unknown':
            return state

        return None  # Will aggregate to Grand Total only

    # Initialize empty metric structure
    def empty_sb_metrics():
        return {
            'SB Loans': 0,
            '#LMICT': 0,
            'LMICT Total Amount': 0,  # Temp field for calculating average
            'Avg SB LMICT Loan Amount': 0,
            'Loans Rev Under $1m': 0,
            'Rev Under $1m Total Amount': 0,  # Temp field for calculating average
            'Avg Loan Amt for <$1M GAR SB': 0,
        }

    # Process by state
    state_data = {}
    grand_total = empty_sb_metrics()

    for record in acquirer_subject + target_subject:
        state = get_state(record)

        if state not in state_data:
            state_data[state] = empty_sb_metrics()

        # Total SB Loans
        total_loans = get_value(record, 'sb_loans', 'sb_loans_total', 'SB Loans', 'total_loans', 'Total Loans', 'loan_count')
        total_amount = get_value(record, 'sb_amount', 'sb_loans_amount', 'SB Amount', 'total_amount', 'Total Amount')
        state_data[state]['SB Loans'] += total_loans

        # LMICT (Loans in LMI Census Tracts)
        lmict_loans = get_value(record, 'lmict_count', 'lmict_loans', 'sb_lmi_loans', 'LMICT Loans', 'lmi_tract_loans')
        lmict_amount = get_value(record, 'lmict_loans_amount', 'lmict_amount', 'sb_lmi_amount', 'LMICT Amount', 'lmi_tract_amount')
        state_data[state]['#LMICT'] += lmict_loans
        state_data[state]['LMICT Total Amount'] += lmict_amount

        # Loans to businesses with revenue under $1M
        rev_under_1m_loans = get_value(record, 'loans_rev_under_1m_count', 'rev_under_1m_loans', 'small_business_loans', 'gar_under_1m_count', 'Loans Rev Under $1m')
        rev_under_1m_amount = get_value(record, 'amount_rev_under_1m', 'rev_under_1m_amount', 'small_business_amount', 'gar_under_1m_amount', 'Rev Under $1m Amount')

        # If we don't have specific <$1M revenue data, use a portion of total
        if rev_under_1m_loans == 0 and total_loans > 0:
            # Estimate ~35% of loans are to businesses with <$1M revenue (common ratio)
            rev_under_1m_loans = int(total_loans * 0.35)
            rev_under_1m_amount = total_amount * 0.35

        state_data[state]['Loans Rev Under $1m'] += rev_under_1m_loans
        state_data[state]['Rev Under $1m Total Amount'] += rev_under_1m_amount

    # Calculate averages and grand totals
    grand_total = empty_sb_metrics()

    for state, metrics in state_data.items():
        # Calculate averages for this state
        if metrics['#LMICT'] > 0:
            metrics['Avg SB LMICT Loan Amount'] = metrics['LMICT Total Amount'] / metrics['#LMICT']
        if metrics['Loans Rev Under $1m'] > 0:
            metrics['Avg Loan Amt for <$1M GAR SB'] = metrics['Rev Under $1m Total Amount'] / metrics['Loans Rev Under $1m']

        # Aggregate to grand total
        grand_total['SB Loans'] += metrics['SB Loans']
        grand_total['#LMICT'] += metrics['#LMICT']
        grand_total['LMICT Total Amount'] += metrics['LMICT Total Amount']
        grand_total['Loans Rev Under $1m'] += metrics['Loans Rev Under $1m']
        grand_total['Rev Under $1m Total Amount'] += metrics['Rev Under $1m Total Amount']

    # Calculate grand total averages
    if grand_total['#LMICT'] > 0:
        grand_total['Avg SB LMICT Loan Amount'] = grand_total['LMICT Total Amount'] / grand_total['#LMICT']
    if grand_total['Loans Rev Under $1m'] > 0:
        grand_total['Avg Loan Amt for <$1M GAR SB'] = grand_total['Rev Under $1m Total Amount'] / grand_total['Loans Rev Under $1m']

    # Clean up temp fields
    for state in list(state_data.keys()) + ['Grand Total']:
        if state == 'Grand Total':
            data = grand_total
        else:
            data = state_data[state]
        data.pop('LMICT Total Amount', None)
        data.pop('Rev Under $1m Total Amount', None)

    # Build final result with Grand Total first
    sb_data['Grand Total'] = grand_total

    # Add state-level data (excluding None and 'Unknown')
    for state, metrics in sorted(state_data.items(), key=lambda x: (x[0] is None, x[0] or '')):
        if state and state != 'Unknown' and state is not None:
            sb_data[state] = metrics

    print(f"[Goals Calculator] SB data summary: {len(sb_data)} regions, Grand Total SB Loans={grand_total['SB Loans']}")
    print(f"[Goals Calculator] SB states found: {list(sb_data.keys())}")

    return sb_data


def _extract_goals_data_from_excel(excel_file):
    """Extract goals data from Excel file as fallback."""
    import pandas as pd

    mortgage_data = {}
    sb_data = {}
    data_years = 3

    try:
        excel = pd.ExcelFile(excel_file)

        # Look for Mortgage Goals sheet
        mortgage_sheets = [s for s in excel.sheet_names if 'mortgage' in s.lower() and 'goal' in s.lower()]
        if mortgage_sheets:
            df = pd.read_excel(excel, sheet_name=mortgage_sheets[0])
            # Process mortgage data from Excel
            # This is a simplified extraction - adjust based on actual Excel structure
            mortgage_data['Grand Total'] = {
                'hp_loans': df.get('HP Loans', pd.Series([0])).sum() if 'HP Loans' in df.columns else 0,
                'hp_amount': df.get('HP Amount', pd.Series([0])).sum() if 'HP Amount' in df.columns else 0,
                'hp_lmi_loans': 0,
                'hp_lmi_amount': 0,
                'refi_loans': df.get('Refi Loans', pd.Series([0])).sum() if 'Refi Loans' in df.columns else 0,
                'refi_amount': df.get('Refi Amount', pd.Series([0])).sum() if 'Refi Amount' in df.columns else 0,
                'refi_lmi_loans': 0,
                'refi_lmi_amount': 0,
                'hi_loans': df.get('HI Loans', pd.Series([0])).sum() if 'HI Loans' in df.columns else 0,
                'hi_amount': df.get('HI Amount', pd.Series([0])).sum() if 'HI Amount' in df.columns else 0,
                'hi_lmi_loans': 0
            }

        # Look for Small Business sheets
        sb_sheets = [s for s in excel.sheet_names if 'small business' in s.lower() or 'sb' in s.lower()]
        if sb_sheets:
            df = pd.read_excel(excel, sheet_name=sb_sheets[0])
            sb_data['Grand Total'] = {
                'sb_loans': df.get('Total Loans', pd.Series([0])).sum() if 'Total Loans' in df.columns else 0,
                'sb_amount': df.get('Total Amount', pd.Series([0])).sum() if 'Total Amount' in df.columns else 0,
                'sb_lmi_loans': df.get('LMI Loans', pd.Series([0])).sum() if 'LMI Loans' in df.columns else 0,
                'sb_lmi_amount': df.get('LMI Amount', pd.Series([0])).sum() if 'LMI Amount' in df.columns else 0,
                'sb_minority_loans': df.get('Minority Loans', pd.Series([0])).sum() if 'Minority Loans' in df.columns else 0,
                'sb_minority_amount': df.get('Minority Amount', pd.Series([0])).sum() if 'Minority Amount' in df.columns else 0
            }
    except Exception as e:
        print(f"Error extracting goals data from Excel: {e}")

    return mortgage_data, sb_data, data_years


@app.route('/api/export-goals', methods=['POST'])
def export_goals():
    """Export goals calculator configuration and data to Excel."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    try:
        config = json.loads(request.form.get('config', '{}'))
        mortgage_data = json.loads(request.form.get('mortgage_data', '{}'))
        sb_data = json.loads(request.form.get('sb_data', '{}'))
        bank_name = request.form.get('bank_name', 'Combined Entity')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "CBA Goals Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="034ea0", end_color="034ea0", fill_type="solid")
        sb_header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = f"CBA Goals Calculator - {bank_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        # Configuration summary
        ws['A3'] = "Configuration"
        ws['A3'].font = Font(bold=True, size=12)

        ws['A4'] = "Baseline Data Years:"
        ws['B4'] = config.get('dataYears', 3)
        ws['A5'] = "Agreement Length:"
        ws['B5'] = f"{config.get('agreementLength', 5)} years"
        ws['A6'] = "Home Purchase Improvement:"
        ws['B6'] = f"+{config.get('hpPercent', 5)}%"
        ws['A7'] = "Refinance Improvement:"
        ws['B7'] = f"+{config.get('refiPercent', 5)}%"
        ws['A8'] = "Home Improvement:"
        ws['B8'] = f"+{config.get('hiPercent', 5)}%"
        ws['A9'] = "Small Business Improvement:"
        ws['B9'] = f"+{config.get('sbPercent', 5)}%"

        # Mortgage Goals Table
        row = 12
        ws[f'A{row}'] = "Mortgage Lending Goals"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 2
        headers = ['Region', 'Metric', 'Baseline', 'Annual Avg', 'Improvement', f"{config.get('agreementLength', 5)}-Year Goal"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row += 1
        data_years = config.get('dataYears', 3)
        agreement_length = config.get('agreementLength', 5)

        # Add mortgage data
        for region, data in mortgage_data.items():
            metrics = [
                ('HP Loans', data.get('hp_loans', 0), config.get('hpPercent', 5)),
                ('HP Amount', data.get('hp_amount', 0), config.get('hpPercent', 5)),
                ('Refi Loans', data.get('refi_loans', 0), config.get('refiPercent', 5)),
                ('Refi Amount', data.get('refi_amount', 0), config.get('refiPercent', 5)),
                ('HI Loans', data.get('hi_loans', 0), config.get('hiPercent', 5)),
            ]

            for metric_name, baseline, percent in metrics:
                annual = baseline / data_years if data_years > 0 else 0
                goal = ((baseline / data_years) * (1 + percent / 100)) * agreement_length if data_years > 0 else 0

                ws.cell(row=row, column=1, value=region).border = thin_border
                ws.cell(row=row, column=2, value=metric_name).border = thin_border
                ws.cell(row=row, column=3, value=baseline).border = thin_border
                ws.cell(row=row, column=4, value=round(annual)).border = thin_border
                ws.cell(row=row, column=5, value=f"+{percent}%").border = thin_border
                ws.cell(row=row, column=6, value=round(goal)).border = thin_border
                row += 1

        # Small Business Goals Table
        row += 2
        ws[f'A{row}'] = "Small Business Lending Goals"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sb_header_fill
            cell.border = thin_border

        row += 1
        sb_percent = config.get('sbPercent', 5)

        for region, data in sb_data.items():
            metrics = [
                ('SB Loans', data.get('sb_loans', 0)),
                ('SB Amount', data.get('sb_amount', 0)),
                ('SB LMI Loans', data.get('sb_lmi_loans', 0)),
                ('SB Minority Loans', data.get('sb_minority_loans', 0)),
            ]

            for metric_name, baseline in metrics:
                annual = baseline / data_years if data_years > 0 else 0
                goal = ((baseline / data_years) * (1 + sb_percent / 100)) * agreement_length if data_years > 0 else 0

                ws.cell(row=row, column=1, value=region).border = thin_border
                ws.cell(row=row, column=2, value=metric_name).border = thin_border
                ws.cell(row=row, column=3, value=baseline).border = thin_border
                ws.cell(row=row, column=4, value=round(annual)).border = thin_border
                ws.cell(row=row, column=5, value=f"+{sb_percent}%").border = thin_border
                ws.cell(row=row, column=6, value=round(goal)).border = thin_border
                row += 1

        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

        # Save to temp file
        import tempfile
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"NCRC_CBA_Goals_{timestamp}.xlsx"

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            return send_file(
                tmp.name,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/save-goals-config', methods=['POST'])
def save_goals_config():
    """Save goals calculator configuration to session/file."""
    try:
        data = request.get_json()
        config = data.get('config', {})
        bank_name = data.get('bank_name', 'Unknown')

        # Save to session
        session['goals_config'] = config
        session['goals_bank_name'] = bank_name

        # Optionally save to file for persistence
        job_id = session.get('job_id')
        if job_id:
            config_file = OUTPUT_DIR / f'goals_config_{job_id}.json'
            with open(config_file, 'w') as f:
                json.dump({
                    'config': config,
                    'bank_name': bank_name,
                    'saved_at': datetime.now().isoformat()
                }, f, indent=2)

        return jsonify({'success': True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Debug: Verify route is registered
print(f"[DEBUG] Checking if /api/generate-assessment-areas-from-branches is registered...")
route_found = False
for rule in app.url_map.iter_rules():
    if '/api/generate-assessment-areas-from-branches' in rule.rule:
        route_found = True
        print(f"[DEBUG] Found route: {rule.rule} -> {rule.endpoint} [{', '.join(rule.methods)}]")
        break
if not route_found:
    print("[DEBUG] WARNING: Route /api/generate-assessment-areas-from-branches NOT FOUND in registered routes!")
    print("[DEBUG] All /api routes:")
    for rule in app.url_map.iter_rules():
        if '/api/' in rule.rule:
            print(f"  {rule.rule} -> {rule.endpoint} [{', '.join(rule.methods)}]")

