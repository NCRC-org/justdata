"""Mortgage worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.apps.mergermeter.excel.utils import _get_cbsa_name_from_code

logger = logging.getLogger(__name__)


def create_simple_mortgage_sheet(wb, bank_name, subject_data, peer_data,
                                header_fill, header_font, border_thin, denominator_type):
    """Create Mortgage Data sheet as simple table with Grand Total at top."""
    sheet_name = f"{bank_name} Mortgage Data"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed CBSA Code column)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Peer/Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No mortgage data for {bank_name}")
        return
    
    # Metrics in order
    metrics = [
        'Loans', 'LMICT%', 'LMIB%', 'LMIB$', 'MMCT%',
        'MINB%', 'Asian%', 'Black%', 'Native American%', 'HoPI%', 'Hispanic%'
    ]
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    grand_total_agg = subject_data.agg({
        'total_loans': 'sum',
        'lmict_loans': 'sum',
        'lmib_loans': 'sum',
        'lmib_amount': 'sum',
        'mmct_loans': 'sum',
        'minb_loans': 'sum',
        'asian_loans': 'sum',
        'black_loans': 'sum',
        'native_american_loans': 'sum',
        'hopi_loans': 'sum',
        'hispanic_loans': 'sum'
    }).to_dict()
    
    grand_total_loans = grand_total_agg.get('total_loans', 0)
    
    # Calculate peer grand total if available
    peer_grand_total_agg = None
    peer_grand_total_loans = 0
    if peer_data is not None and not peer_data.empty:
        peer_grand_total_agg = peer_data.agg({
            'total_loans': 'sum',
            'lmict_loans': 'sum',
            'lmib_loans': 'sum',
            'lmib_amount': 'sum',
            'mmct_loans': 'sum',
            'minb_loans': 'sum',
            'asian_loans': 'sum',
            'black_loans': 'sum',
            'native_american_loans': 'sum',
            'hopi_loans': 'sum',
            'hispanic_loans': 'sum'
        }).to_dict()
        peer_grand_total_loans = peer_grand_total_agg.get('total_loans', 0)
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'Loans':
            ws.cell(row, 3).value = int(grand_total_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'LMICT%':
            pct = (grand_total_agg.get('lmict_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB%':
            pct = (grand_total_agg.get('lmib_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric == 'LMIB$':
            ws.cell(row, 3).value = float(grand_total_agg.get('lmib_amount', 0))
            ws.cell(row, 3).number_format = '$#,##0'
        elif metric == 'MMCT%':
            pct = (grand_total_agg.get('mmct_loans', 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
            ws.cell(row, 3).value = pct / 100
            ws.cell(row, 3).number_format = '0.00%'
        elif metric.endswith('%'):
            metric_map = {
                'MINB%': 'minb_loans',
                'Asian%': 'asian_loans',
                'Black%': 'black_loans',
                'Native American%': 'native_american_loans',
                'HoPI%': 'hopi_loans',
                'Hispanic%': 'hispanic_loans'
            }
            field = metric_map.get(metric, '')
            if field:
                pct = (grand_total_agg.get(field, 0) / grand_total_loans * 100) if grand_total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
        
        # Peer Grand Total
        if peer_grand_total_agg:
            if metric == 'Loans':
                ws.cell(row, 4).value = int(peer_grand_total_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'LMICT%':
                pct = (peer_grand_total_agg.get('lmict_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB%':
                pct = (peer_grand_total_agg.get('lmib_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric == 'LMIB$':
                ws.cell(row, 4).value = float(peer_grand_total_agg.get('lmib_amount', 0))
                ws.cell(row, 4).number_format = '$#,##0'
            elif metric == 'MMCT%':
                pct = (peer_grand_total_agg.get('mmct_loans', 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                ws.cell(row, 4).value = pct / 100
                ws.cell(row, 4).number_format = '0.00%'
            elif metric.endswith('%'):
                metric_map = {
                    'MINB%': 'minb_loans',
                    'Asian%': 'asian_loans',
                    'Black%': 'black_loans',
                    'Native American%': 'native_american_loans',
                    'HoPI%': 'hopi_loans',
                    'Hispanic%': 'hispanic_loans'
                }
                field = metric_map.get(metric, '')
                if field:
                    pct = (peer_grand_total_agg.get(field, 0) / peer_grand_total_loans * 100) if peer_grand_total_loans > 0 else 0
                    ws.cell(row, 4).value = pct / 100
                    ws.cell(row, 4).number_format = '0.00%'
        
        # Difference for percentages
        if metric.endswith('%') and metric != 'LMIB$':
            ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group by CBSA and aggregate
    grouped = subject_data.groupby('cbsa_code')
    
    # Sort by total loans
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data['total_loans'].sum() if not group_data.empty else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    # Process each CBSA
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Aggregate data - only use total_loans as denominator (no loans_with_demographic_data)
        agg = group_data.agg({
            'total_loans': 'sum',
            'lmict_loans': 'sum',
            'lmib_loans': 'sum',
            'lmib_amount': 'sum',
            'mmct_loans': 'sum',
            'minb_loans': 'sum',
            'asian_loans': 'sum',
            'black_loans': 'sum',
            'native_american_loans': 'sum',
            'hopi_loans': 'sum',
            'hispanic_loans': 'sum'
        }).to_dict()
        
        total_loans = agg.get('total_loans', 0)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        # Always use total_loans as denominator (no loans_with_demographic_data)
        demo_denominator = total_loans
        
        # Write data for each metric
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Subject data (column 3 after removing CBSA Code)
            if metric == 'Loans':
                ws.cell(row, 3).value = int(total_loans)
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'LMICT%':
                pct = (agg.get('lmict_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric == 'LMIB%':
                pct = (agg.get('lmib_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric == 'LMIB$':
                ws.cell(row, 3).value = float(agg.get('lmib_amount', 0))
                ws.cell(row, 3).number_format = '$#,##0'
            elif metric == 'MMCT%':
                pct = (agg.get('mmct_loans', 0) / total_loans * 100) if total_loans > 0 else 0
                ws.cell(row, 3).value = pct / 100
                ws.cell(row, 3).number_format = '0.00%'
            elif metric.endswith('%'):
                # Race/ethnicity percentages
                metric_map = {
                    'MINB%': 'minb_loans',
                    'Asian%': 'asian_loans',
                    'Black%': 'black_loans',
                    'Native American%': 'native_american_loans',
                    'HoPI%': 'hopi_loans',
                    'Hispanic%': 'hispanic_loans'
                }
                field = metric_map.get(metric, '')
                if field:
                    pct = (agg.get(field, 0) / demo_denominator * 100) if demo_denominator > 0 else 0
                    ws.cell(row, 3).value = pct / 100
                    ws.cell(row, 3).number_format = '0.00%'
            
            # Peer data
            if peer_data is not None and not peer_data.empty:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    peer_agg = peer_group.agg({
                        'total_loans': 'sum',
                        'lmict_loans': 'sum',
                        'lmib_loans': 'sum',
                        'lmib_amount': 'sum',
                        'mmct_loans': 'sum',
                        'minb_loans': 'sum',
                        'asian_loans': 'sum',
                        'black_loans': 'sum',
                        'native_american_loans': 'sum',
                        'hopi_loans': 'sum',
                        'hispanic_loans': 'sum'
                    }).to_dict()
                    
                    peer_total = peer_agg.get('total_loans', 0)
                    
                    if metric == 'Loans':
                        ws.cell(row, 4).value = int(peer_total)
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'LMICT%':
                        pct = (peer_agg.get('lmict_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric == 'LMIB%':
                        pct = (peer_agg.get('lmib_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric == 'LMIB$':
                        ws.cell(row, 4).value = float(peer_agg.get('lmib_amount', 0))
                        ws.cell(row, 4).number_format = '$#,##0'
                    elif metric == 'MMCT%':
                        pct = (peer_agg.get('mmct_loans', 0) / peer_total * 100) if peer_total > 0 else 0
                        ws.cell(row, 4).value = pct / 100
                        ws.cell(row, 4).number_format = '0.00%'
                    elif metric.endswith('%'):
                        metric_map = {
                            'MINB%': 'minb_loans',
                            'Asian%': 'asian_loans',
                            'Black%': 'black_loans',
                            'Native American%': 'native_american_loans',
                            'HoPI%': 'hopi_loans',
                            'Hispanic%': 'hispanic_loans'
                        }
                        field = metric_map.get(metric, '')
                        if field:
                            pct = (peer_agg.get(field, 0) / peer_total * 100) if peer_total > 0 else 0
                            ws.cell(row, 4).value = pct / 100
                            ws.cell(row, 4).number_format = '0.00%'
            
            # Difference (for percentages only) - columns shifted: C=Subject, D=Peer, E=Difference
            if metric.endswith('%') and metric != 'LMIB$':
                ws.cell(row, 5).value = f'=IFERROR(C{row}-D{row},0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust column widths (removed CBSA Code column)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


