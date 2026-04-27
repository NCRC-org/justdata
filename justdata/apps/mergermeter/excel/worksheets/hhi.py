"""HHI worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _add_empty_hhi_sheet(wb: Workbook, bank_a_name: str, bank_b_name: str, bank_a_rssd: str, bank_b_rssd: str):
    """Add HHI Analysis sheet with explanatory note when no overlapping counties exist."""

    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))

    # Title style
    title_font = Font(bold=True, size=14, color="034EA0")
    note_font = Font(size=11, italic=True)
    info_font = Font(size=11)

    # Title
    ws.merge_cells('A1:G1')
    cell = ws.cell(1, 1, "HHI Analysis - Deposit Market Concentration")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    # Explanatory note
    ws.merge_cells('A3:G3')
    note_cell = ws.cell(3, 1, "No HHI analysis available - Banks do not have any overlapping counties with branch locations.")
    note_cell.font = Font(size=12, bold=True, color="C00000")  # Red color for emphasis
    note_cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Bank details
    ws.cell(5, 1, "Bank Details:").font = Font(bold=True, size=11)
    ws.cell(6, 1, f"• {bank_a_name}").font = info_font
    ws.cell(6, 3, f"RSSD: {bank_a_rssd if bank_a_rssd else 'N/A'}").font = info_font
    ws.cell(7, 1, f"• {bank_b_name}").font = info_font
    ws.cell(7, 3, f"RSSD: {bank_b_rssd if bank_b_rssd else 'N/A'}").font = info_font

    # Additional explanation
    ws.merge_cells('A9:G9')
    ws.cell(9, 1, "HHI (Herfindahl-Hirschman Index) analysis requires both banks to have branch locations in at least one common county.").font = note_font
    ws.merge_cells('A10:G10')
    ws.cell(10, 1, "Since these banks operate in separate geographic markets, deposit concentration analysis cannot be performed.").font = note_font

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    logger.info(f"[HHI] Created empty HHI sheet with explanatory note")



def _add_hhi_sheet(wb: Workbook, hhi_data: pd.DataFrame, bank_a_name: str, bank_b_name: str):
    """Add HHI Analysis sheet to workbook for counties where both banks have branches."""
    
    # Filter to only counties where both banks have branches
    if 'bank_a_deposits' in hhi_data.columns and 'bank_b_deposits' in hhi_data.columns:
        hhi_filtered = hhi_data[
            (hhi_data['bank_a_deposits'] > 0) & 
            (hhi_data['bank_b_deposits'] > 0)
        ].copy()
    else:
        # If columns don't exist, use all data
        hhi_filtered = hhi_data.copy()
    
    if hhi_filtered.empty:
        print("  No counties with overlapping branches for HHI analysis")
        return
    
    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))
    
    # Header style - NCRC blue (#034ea0)
    header_fill = PatternFill(start_color="034EA0", end_color="034EA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Title
    row = 1
    ws.merge_cells(f'A{row}:{get_column_letter(len(hhi_filtered.columns))}{row}')
    cell = ws.cell(row, 1, f"HHI Analysis - Counties with {bank_a_name} and {bank_b_name} Branches")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")
    row += 2
    
    # Headers
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        cell = ws.cell(row, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row += 1
    
    # Data rows
    for df_row_idx, df_row in hhi_filtered.iterrows():
        for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
            value = df_row[col_name]
            cell = ws.cell(row, col_idx, value)
            
            # Format numbers
            if pd.api.types.is_numeric_dtype(hhi_filtered[col_name]):
                if 'hhi' in col_name.lower():
                    cell.number_format = '#,##0'  # HHI is typically an integer
                elif 'deposits' in col_name.lower() or 'share' in col_name.lower():
                    if 'share' in col_name.lower() or 'percent' in col_name.lower():
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
            
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Adjust column widths
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        col_letter = get_column_letter(col_idx)
        try:
            col_max_len = hhi_filtered[col_name].astype(str).str.len().max() if not hhi_filtered[col_name].empty else 10
        except (ValueError, AttributeError):
            col_max_len = 10
        max_length = max(len(str(col_name)), col_max_len)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    print("  Added HHI Analysis sheet")
