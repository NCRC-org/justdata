"""Small business lending worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.apps.mergermeter.excel.utils import _get_cbsa_name_from_code

logger = logging.getLogger(__name__)


def create_simple_sb_sheet(wb, bank_name, subject_data, peer_data,
                          header_fill, header_font, border_thin):
    """Create Small Business sheet as simple table with Grand Total at top (no State column)."""
    sheet_name = f"{bank_name} SB Lending"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed State and CBSA Code columns)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Peer/Other', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if subject_data is None or subject_data.empty:
        print(f"    No SB data for {bank_name}")
        return
    
    metrics = ['SB Loans', '#LMICT', 'Avg SB LMICT Loan Amount', 'Loans Rev Under $1m', 'Avg Loan Amt for <$1M GAR SB']
    
    # Handle both column names: sb_loans_total (from merger report queries) and sb_loans_count (from MergerMeter queries)
    sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in subject_data.columns else 'sb_loans_count'
    lmict_count_col = 'lmict_count' if 'lmict_count' in subject_data.columns else 'lmict_loans_count'
    loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in subject_data.columns else 'loans_rev_under_1m'
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    grand_total_sb_loans = subject_data[sb_loans_col].sum() if sb_loans_col in subject_data.columns else 0
    grand_total_lmict = subject_data[lmict_count_col].sum() if lmict_count_col in subject_data.columns else 0
    grand_total_rev_under_1m = subject_data[loans_rev_col].sum() if loans_rev_col in subject_data.columns else 0
    
    # Calculate weighted averages for Grand Total
    if 'lmict_loans_amount' in subject_data.columns:
        grand_total_avg_lmict = subject_data['lmict_loans_amount'].sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    elif 'avg_sb_lmict_loan_amount' in subject_data.columns:
        grand_total_avg_lmict = (subject_data[lmict_count_col] * subject_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / grand_total_lmict if grand_total_lmict > 0 else 0
    else:
        grand_total_avg_lmict = 0
    
    if 'amount_rev_under_1m' in subject_data.columns:
        grand_total_avg_rev = subject_data['amount_rev_under_1m'].sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    elif 'avg_loan_amt_rum_sb' in subject_data.columns:
        grand_total_avg_rev = (subject_data[loans_rev_col] * subject_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / grand_total_rev_under_1m if grand_total_rev_under_1m > 0 else 0
    else:
        grand_total_avg_rev = 0
    
    # Calculate peer grand totals if available
    peer_grand_total_sb_loans = 0
    peer_grand_total_lmict = 0
    peer_grand_total_rev_under_1m = 0
    peer_grand_total_avg_lmict = 0
    peer_grand_total_avg_rev = 0
    if peer_data is not None and not peer_data.empty:
        peer_sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in peer_data.columns else 'sb_loans_count'
        peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_data.columns else 'lmict_loans_count'
        peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_data.columns else 'loans_rev_under_1m'
        
        peer_grand_total_sb_loans = peer_data[peer_sb_loans_col].sum() if peer_sb_loans_col in peer_data.columns else 0
        peer_grand_total_lmict = peer_data[peer_lmict_count_col].sum() if peer_lmict_count_col in peer_data.columns else 0
        peer_grand_total_rev_under_1m = peer_data[peer_loans_rev_col].sum() if peer_loans_rev_col in peer_data.columns else 0
        
        if 'lmict_loans_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = peer_data['lmict_loans_amount'].sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        elif 'avg_sb_lmict_loan_amount' in peer_data.columns:
            peer_grand_total_avg_lmict = (peer_data[peer_lmict_count_col] * peer_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_grand_total_lmict if peer_grand_total_lmict > 0 else 0
        else:
            peer_grand_total_avg_lmict = 0
        
        if 'amount_rev_under_1m' in peer_data.columns:
            peer_grand_total_avg_rev = peer_data['amount_rev_under_1m'].sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        elif 'avg_loan_amt_rum_sb' in peer_data.columns:
            peer_grand_total_avg_rev = (peer_data[peer_loans_rev_col] * peer_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_grand_total_rev_under_1m if peer_grand_total_rev_under_1m > 0 else 0
        else:
            peer_grand_total_avg_rev = 0
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        # Subject Grand Total
        if metric == 'SB Loans':
            ws.cell(row, 3).value = int(grand_total_sb_loans)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == '#LMICT':
            ws.cell(row, 3).value = int(grand_total_lmict)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg SB LMICT Loan Amount':
            ws.cell(row, 3).value = float(grand_total_avg_lmict)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Loans Rev Under $1m':
            ws.cell(row, 3).value = int(grand_total_rev_under_1m)
            ws.cell(row, 3).number_format = '#,##0'
        elif metric == 'Avg Loan Amt for <$1M GAR SB':
            ws.cell(row, 3).value = float(grand_total_avg_rev)
            ws.cell(row, 3).number_format = '#,##0'
        
        # Peer Grand Total
        if peer_data is not None and not peer_data.empty:
            if metric == 'SB Loans':
                ws.cell(row, 4).value = int(peer_grand_total_sb_loans)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == '#LMICT':
                ws.cell(row, 4).value = int(peer_grand_total_lmict)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg SB LMICT Loan Amount':
                ws.cell(row, 4).value = float(peer_grand_total_avg_lmict)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Loans Rev Under $1m':
                ws.cell(row, 4).value = int(peer_grand_total_rev_under_1m)
                ws.cell(row, 4).number_format = '#,##0'
            elif metric == 'Avg Loan Amt for <$1M GAR SB':
                ws.cell(row, 4).value = float(peer_grand_total_avg_rev)
                ws.cell(row, 4).number_format = '#,##0'
        
        # Difference formulas
        if metric in ['#LMICT', 'Loans Rev Under $1m']:
            sb_loans_row = row - metrics.index(metric)
            ws.cell(row, 5).value = f'=IFERROR((C{row}/C{sb_loans_row})-(D{row}/D{sb_loans_row}),0)'
            ws.cell(row, 5).number_format = '0.00%'
        elif metric in ['Avg SB LMICT Loan Amount', 'Avg Loan Amt for <$1M GAR SB']:
            ws.cell(row, 5).value = f'=IFERROR((C{row}/D{row})-1,0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group and sort by CBSA
    grouped = subject_data.groupby('cbsa_code')
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data[sb_loans_col].sum() if not group_data.empty and sb_loans_col in group_data.columns else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        # Aggregate - handle both column name formats
        # MergerMeter uses: sb_loans_count, lmict_loans_count, loans_rev_under_1m
        # Merger report uses: sb_loans_total, lmict_count, loans_rev_under_1m_count
        sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in group_data.columns else 'sb_loans_count'
        lmict_count_col = 'lmict_count' if 'lmict_count' in group_data.columns else 'lmict_loans_count'
        loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in group_data.columns else 'loans_rev_under_1m'
        
        agg = {
            'sb_loans_total': group_data[sb_loans_col].sum() if sb_loans_col in group_data.columns else 0,
            'lmict_count': group_data[lmict_count_col].sum() if lmict_count_col in group_data.columns else 0,
            'loans_rev_under_1m_count': group_data[loans_rev_col].sum() if loans_rev_col in group_data.columns else 0,
        }
        
        # Weighted averages - handle different column names
        # MergerMeter uses: lmict_loans_count, lmict_loans_amount, loans_rev_under_1m, amount_rev_under_1m
        # Merger report uses: lmict_count, avg_sb_lmict_loan_amount, loans_rev_under_1m_count, avg_loan_amt_rum_sb
        lmict_count_col = 'lmict_count' if 'lmict_count' in group_data.columns else 'lmict_loans_count'
        loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in group_data.columns else 'loans_rev_under_1m'
        
        if agg['lmict_count'] > 0:
            # Check if we have pre-calculated average or need to calculate from amounts
            if 'avg_sb_lmict_loan_amount' in group_data.columns:
                agg['avg_sb_lmict_loan_amount'] = (group_data[lmict_count_col] * group_data['avg_sb_lmict_loan_amount'].fillna(0)).sum() / agg['lmict_count']
            elif 'lmict_loans_amount' in group_data.columns:
                agg['avg_sb_lmict_loan_amount'] = group_data['lmict_loans_amount'].sum() / agg['lmict_count'] if agg['lmict_count'] > 0 else 0
            else:
                agg['avg_sb_lmict_loan_amount'] = 0
        else:
            agg['avg_sb_lmict_loan_amount'] = 0
        
        if agg['loans_rev_under_1m_count'] > 0:
            # Check if we have pre-calculated average or need to calculate from amounts
            if 'avg_loan_amt_rum_sb' in group_data.columns:
                agg['avg_loan_amt_rum_sb'] = (group_data[loans_rev_col] * group_data['avg_loan_amt_rum_sb'].fillna(0)).sum() / agg['loans_rev_under_1m_count']
            elif 'amount_rev_under_1m' in group_data.columns:
                agg['avg_loan_amt_rum_sb'] = group_data['amount_rev_under_1m'].sum() / agg['loans_rev_under_1m_count'] if agg['loans_rev_under_1m_count'] > 0 else 0
            else:
                agg['avg_loan_amt_rum_sb'] = 0
        else:
            agg['avg_loan_amt_rum_sb'] = 0
        
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Subject data (column 3 after removing State and CBSA Code)
            if metric == 'SB Loans':
                ws.cell(row, 3).value = int(agg['sb_loans_total'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == '#LMICT':
                # Column C: Loan units (count), not percentage
                ws.cell(row, 3).value = int(agg['lmict_count'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Avg SB LMICT Loan Amount':
                ws.cell(row, 3).value = float(agg['avg_sb_lmict_loan_amount'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Loans Rev Under $1m':
                # Column C: Loan units (count), not percentage
                ws.cell(row, 3).value = int(agg['loans_rev_under_1m_count'])
                ws.cell(row, 3).number_format = '#,##0'
            elif metric == 'Avg Loan Amt for <$1M GAR SB':
                ws.cell(row, 3).value = float(agg['avg_loan_amt_rum_sb'])
                ws.cell(row, 3).number_format = '#,##0'
            
            # Peer data
            if peer_data is not None and not peer_data.empty:
                peer_group = peer_data[peer_data['cbsa_code'] == cbsa_code]
                if not peer_group.empty:
                    # Handle both column name formats for peer data too
                    peer_sb_loans_col = 'sb_loans_total' if 'sb_loans_total' in peer_group.columns else 'sb_loans_count'
                    peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                    peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                    
                    peer_agg = {
                        'sb_loans_total': peer_group[peer_sb_loans_col].sum() if peer_sb_loans_col in peer_group.columns else 0,
                        'lmict_count': peer_group[peer_lmict_count_col].sum() if peer_lmict_count_col in peer_group.columns else 0,
                        'loans_rev_under_1m_count': peer_group[peer_loans_rev_col].sum() if peer_loans_rev_col in peer_group.columns else 0,
                    }
                    
                    # Handle different column names for peer data too
                    peer_lmict_count_col = 'lmict_count' if 'lmict_count' in peer_group.columns else 'lmict_loans_count'
                    peer_loans_rev_col = 'loans_rev_under_1m_count' if 'loans_rev_under_1m_count' in peer_group.columns else 'loans_rev_under_1m'
                    
                    if peer_agg['lmict_count'] > 0:
                        if 'avg_sb_lmict_loan_amount' in peer_group.columns:
                            peer_agg['avg_sb_lmict_loan_amount'] = (peer_group[peer_lmict_count_col] * peer_group['avg_sb_lmict_loan_amount'].fillna(0)).sum() / peer_agg['lmict_count']
                        elif 'lmict_loans_amount' in peer_group.columns:
                            peer_agg['avg_sb_lmict_loan_amount'] = peer_group['lmict_loans_amount'].sum() / peer_agg['lmict_count'] if peer_agg['lmict_count'] > 0 else 0
                        else:
                            peer_agg['avg_sb_lmict_loan_amount'] = 0
                    else:
                        peer_agg['avg_sb_lmict_loan_amount'] = 0
                    
                    if peer_agg['loans_rev_under_1m_count'] > 0:
                        if 'avg_loan_amt_rum_sb' in peer_group.columns:
                            peer_agg['avg_loan_amt_rum_sb'] = (peer_group[peer_loans_rev_col] * peer_group['avg_loan_amt_rum_sb'].fillna(0)).sum() / peer_agg['loans_rev_under_1m_count']
                        elif 'amount_rev_under_1m' in peer_group.columns:
                            peer_agg['avg_loan_amt_rum_sb'] = peer_group['amount_rev_under_1m'].sum() / peer_agg['loans_rev_under_1m_count'] if peer_agg['loans_rev_under_1m_count'] > 0 else 0
                        else:
                            peer_agg['avg_loan_amt_rum_sb'] = 0
                    else:
                        peer_agg['avg_loan_amt_rum_sb'] = 0
                    
                    if metric == 'SB Loans':
                        ws.cell(row, 4).value = int(peer_agg['sb_loans_total'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == '#LMICT':
                        # Column D: Loan units (count), not percentage
                        ws.cell(row, 4).value = int(peer_agg['lmict_count'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Avg SB LMICT Loan Amount':
                        ws.cell(row, 4).value = float(peer_agg['avg_sb_lmict_loan_amount'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Loans Rev Under $1m':
                        # Column D: Loan units (count), not percentage
                        ws.cell(row, 4).value = int(peer_agg['loans_rev_under_1m_count'])
                        ws.cell(row, 4).number_format = '#,##0'
                    elif metric == 'Avg Loan Amt for <$1M GAR SB':
                        ws.cell(row, 4).value = float(peer_agg['avg_loan_amt_rum_sb'])
                        ws.cell(row, 4).number_format = '#,##0'
            
            # Difference formulas (columns shifted: C=Subject, D=Peer, E=Difference)
            # For #LMICT and Loans Rev Under $1m: Columns C and D are in loan units (counts),
            # Column E calculates percentage difference: (Subject% - Peer%) where % = metric/SB Loans
            if metric in ['#LMICT', 'Loans Rev Under $1m']:
                sb_loans_row = row - metrics.index(metric)
                ws.cell(row, 5).value = f'=IFERROR((C{row}/C{sb_loans_row})-(D{row}/D{sb_loans_row}),0)'
                ws.cell(row, 5).number_format = '0.00%'
            elif metric in ['Avg SB LMICT Loan Amount', 'Avg Loan Amt for <$1M GAR SB']:
                ws.cell(row, 5).value = f'=IFERROR((C{row}/D{row})-1,0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust widths (removed CBSA Code column)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


