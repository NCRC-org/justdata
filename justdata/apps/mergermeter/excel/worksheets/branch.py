"""Branch worksheet builder."""
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.apps.mergermeter.excel.utils import _get_cbsa_name_from_code

logger = logging.getLogger(__name__)


def create_simple_branch_sheet(wb, bank_name, branch_data,
                               header_fill, header_font, border_thin):
    """Create Branch sheet as simple table with Grand Total at top."""
    sheet_name = f"{bank_name} Branches"
    ws = wb.create_sheet(sheet_name)
    
    # Headers (removed CBSA Code column)
    headers = ['CBSA Name', 'Metric', 'Subject Bank', 'Other/Market', 'Difference']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if branch_data is None or branch_data.empty:
        print(f"    No branch data for {bank_name}")
        return
    
    metrics = ['Branches', 'LMICT', 'MMCT']
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Calculate Grand Total (aggregate across all CBSAs)
    # Branch query returns: total_branches, branches_in_lmict, pct_lmict, branches_in_mmct, pct_mmct
    # Map to expected column names
    total_branches_col = 'total_branches' if 'total_branches' in branch_data.columns else 'subject_total_branches'
    lmict_col = 'branches_in_lmict' if 'branches_in_lmict' in branch_data.columns else 'subject_pct_lmict'
    pct_lmict_col = 'pct_lmict' if 'pct_lmict' in branch_data.columns else 'subject_pct_lmict'
    mmct_col = 'branches_in_mmct' if 'branches_in_mmct' in branch_data.columns else 'subject_pct_mmct'
    pct_mmct_col = 'pct_mmct' if 'pct_mmct' in branch_data.columns else 'subject_pct_mmct'
    
    grand_total_subject = branch_data[total_branches_col].sum() if total_branches_col in branch_data.columns else 0
    market_total_col = 'market_total_branches' if 'market_total_branches' in branch_data.columns else None
    grand_total_other = branch_data[market_total_col].sum() if market_total_col and market_total_col in branch_data.columns else 0
    
    # Calculate LMICT and MMCT counts for Grand Total
    if grand_total_subject > 0:
        if lmict_col in branch_data.columns:
            # Use count column if available
            subject_lmict_count = branch_data[lmict_col].sum()
            subject_lmict_pct = (subject_lmict_count / grand_total_subject * 100) if grand_total_subject > 0 else 0
        elif pct_lmict_col in branch_data.columns:
            # Use weighted average of percentages
            subject_lmict_pct = (branch_data[pct_lmict_col] * branch_data[total_branches_col]).sum() / grand_total_subject
        else:
            subject_lmict_pct = 0
        
        if mmct_col in branch_data.columns:
            # Use count column if available
            subject_mmct_count = branch_data[mmct_col].sum()
            subject_mmct_pct = (subject_mmct_count / grand_total_subject * 100) if grand_total_subject > 0 else 0
        elif pct_mmct_col in branch_data.columns:
            # Use weighted average of percentages
            subject_mmct_pct = (branch_data[pct_mmct_col] * branch_data[total_branches_col]).sum() / grand_total_subject
        else:
            subject_mmct_pct = 0
    else:
        subject_lmict_pct = 0
        subject_mmct_pct = 0
    
    if grand_total_other > 0:
        if 'market_branches_in_lmict' in branch_data.columns:
            other_lmict_count = branch_data['market_branches_in_lmict'].sum()
            other_lmict_pct = (other_lmict_count / grand_total_other * 100) if grand_total_other > 0 else 0
        elif 'market_pct_lmict' in branch_data.columns:
            other_lmict_pct = (branch_data['market_pct_lmict'] * branch_data[market_total_col]).sum() / grand_total_other
        else:
            other_lmict_pct = 0
        
        if 'market_branches_in_mmct' in branch_data.columns:
            other_mmct_count = branch_data['market_branches_in_mmct'].sum()
            other_mmct_pct = (other_mmct_count / grand_total_other * 100) if grand_total_other > 0 else 0
        elif 'market_pct_mmct' in branch_data.columns:
            other_mmct_pct = (branch_data['market_pct_mmct'] * branch_data[market_total_col]).sum() / grand_total_other
        else:
            other_mmct_pct = 0
    else:
        other_lmict_pct = 0
        other_mmct_pct = 0
    
    grand_total_subject_lmict = int(round((subject_lmict_pct / 100.0) * grand_total_subject))
    grand_total_subject_mmct = int(round((subject_mmct_pct / 100.0) * grand_total_subject))
    grand_total_other_lmict = int(round((other_lmict_pct / 100.0) * grand_total_other))
    grand_total_other_mmct = int(round((other_mmct_pct / 100.0) * grand_total_other))
    
    # Write Grand Total section at row 2
    row = 2
    for metric in metrics:
        ws.cell(row, 1).value = 'Grand Total'
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = metric
        
        if metric == 'Branches':
            ws.cell(row, 3).value = int(grand_total_subject)
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = int(grand_total_other)
            ws.cell(row, 4).number_format = '0'
        elif metric == 'LMICT':
            ws.cell(row, 3).value = grand_total_subject_lmict
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = grand_total_other_lmict
            ws.cell(row, 4).number_format = '0'
            branches_row = row - 1
            ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
            ws.cell(row, 5).number_format = '0.00%'
        elif metric == 'MMCT':
            ws.cell(row, 3).value = grand_total_subject_mmct
            ws.cell(row, 3).number_format = '0'
            ws.cell(row, 4).value = grand_total_other_mmct
            ws.cell(row, 4).number_format = '0'
            branches_row = row - 2
            ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
            ws.cell(row, 5).number_format = '0.00%'
        
        row += 1
    
    # Add blank row after Grand Total
    row += 1
    
    # Group and sort by CBSA
    grouped = branch_data.groupby('cbsa_code')
    cbsa_totals = {}
    for cbsa_code, group_data in grouped:
        total = group_data[total_branches_col].sum() if total_branches_col in group_data.columns else 0
        cbsa_totals[cbsa_code] = total
    
    sorted_cbsas = sorted(grouped, key=lambda x: cbsa_totals.get(x[0], 0), reverse=True)
    
    for group_key, group_data in sorted_cbsas:
        cbsa_code = str(group_key)
        
        # Get CBSA name - try from data first, then lookup
        cbsa_name = None
        if 'cbsa_name' in group_data.columns and not group_data.empty:
            cbsa_name_val = group_data['cbsa_name'].iloc[0]
            if cbsa_name_val and str(cbsa_name_val).lower() not in ['nan', 'none', '']:
                cbsa_name = str(cbsa_name_val).strip()
        
        if not cbsa_name:
            cbsa_name = _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
        
        subject_total = group_data[total_branches_col].sum() if total_branches_col in group_data.columns else 0
        other_total = group_data[market_total_col].sum() if market_total_col and market_total_col in group_data.columns else 0
        
        # Calculate LMICT and MMCT counts for subject
        if lmict_col in group_data.columns:
            subject_lmict_count = int(group_data[lmict_col].sum())
        elif pct_lmict_col in group_data.columns:
            subject_lmict_pct_val = group_data[pct_lmict_col].iloc[0] if not group_data.empty else 0
            subject_lmict_count = int(round((subject_lmict_pct_val / 100.0) * subject_total)) if subject_total > 0 else 0
        else:
            subject_lmict_count = 0
        
        if mmct_col in group_data.columns:
            subject_mmct_count = int(group_data[mmct_col].sum())
        elif pct_mmct_col in group_data.columns:
            subject_mmct_pct_val = group_data[pct_mmct_col].iloc[0] if not group_data.empty else 0
            subject_mmct_count = int(round((subject_mmct_pct_val / 100.0) * subject_total)) if subject_total > 0 else 0
        else:
            subject_mmct_count = 0
        
        # Calculate LMICT and MMCT counts for market
        if 'market_branches_in_lmict' in group_data.columns:
            other_lmict_count = int(group_data['market_branches_in_lmict'].sum())
        elif 'market_pct_lmict' in group_data.columns:
            other_lmict_pct_val = group_data['market_pct_lmict'].iloc[0] if not group_data.empty else 0
            other_lmict_count = int(round((other_lmict_pct_val / 100.0) * other_total)) if other_total > 0 else 0
        else:
            other_lmict_count = 0
        
        if 'market_branches_in_mmct' in group_data.columns:
            other_mmct_count = int(group_data['market_branches_in_mmct'].sum())
        elif 'market_pct_mmct' in group_data.columns:
            other_mmct_pct_val = group_data['market_pct_mmct'].iloc[0] if not group_data.empty else 0
            other_mmct_count = int(round((other_mmct_pct_val / 100.0) * other_total)) if other_total > 0 else 0
        else:
            other_mmct_count = 0
        
        for metric in metrics:
            ws.cell(row, 1).value = cbsa_name
            ws.cell(row, 2).value = metric
            
            # Columns shifted: C=Subject, D=Other, E=Difference
            if metric == 'Branches':
                ws.cell(row, 3).value = int(subject_total)
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = int(other_total)
                ws.cell(row, 4).number_format = '0'
            elif metric == 'LMICT':
                ws.cell(row, 3).value = subject_lmict_count
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = other_lmict_count
                ws.cell(row, 4).number_format = '0'
                # Difference: percentage difference (C=Subject, D=Other)
                branches_row = row - 1
                ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
                ws.cell(row, 5).number_format = '0.00%'
            elif metric == 'MMCT':
                ws.cell(row, 3).value = subject_mmct_count
                ws.cell(row, 3).number_format = '0'
                ws.cell(row, 4).value = other_mmct_count
                ws.cell(row, 4).number_format = '0'
                # Difference: percentage difference (C=Subject, D=Other)
                branches_row = row - 2
                ws.cell(row, 5).value = f'=IFERROR(IF(C{branches_row}=0,0,IF(D{branches_row}=0,0,(C{row}/C{branches_row})-(D{row}/D{branches_row}))),0)'
                ws.cell(row, 5).number_format = '0.00%'
            
            row += 1
    
    # Auto-adjust widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created {sheet_name}: {row - 2} data rows")


