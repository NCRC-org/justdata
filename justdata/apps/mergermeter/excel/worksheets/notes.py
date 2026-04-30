"""Notes / cover worksheet builder."""
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def create_simple_notes_sheet(wb, bank_a_name, bank_b_name, years_hmda, years_sb):
    """Create simple Notes/Methodology sheet."""
    ws = wb.create_sheet("Notes")
    
    ws['A1'] = "Data Sources and Timeframes"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    ws.cell(row, 1).value = "1. Mortgage Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from HMDA (Home Mortgage Disclosure Act) data"
    row += 1
    ws.cell(row, 2).value = f"Covers years: {', '.join(years_hmda) if years_hmda else 'N/A'}"
    row += 2
    
    ws.cell(row, 1).value = "2. Small Business Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from Section 1071 Small Business Lending data"
    row += 1
    ws.cell(row, 2).value = f"Covers years: {', '.join(years_sb) if years_sb else 'N/A'}"
    row += 2
    
    ws.cell(row, 1).value = "3. Branch Data:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Sourced from FDIC Summary of Deposits data"
    row += 2
    
    ws.cell(row, 1).value = f"4. Banks Analyzed:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = f"Acquirer: {bank_a_name}"
    row += 1
    ws.cell(row, 2).value = f"Target: {bank_b_name}"
    row += 2
    
    ws.cell(row, 1).value = "Generated:"
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.cell(row, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Auto-adjust width
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    
    print("  Created Notes sheet")


