"""Assessment areas worksheet builder."""
import logging
from collections import defaultdict
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.apps.mergermeter.excel.utils import _get_cbsa_name_from_code

logger = logging.getLogger(__name__)


def create_simple_assessment_areas_sheet(wb, bank_a_name, bank_b_name, assessment_areas_data,
                                        header_fill, header_font, border_thin):
    """Create Assessment Areas sheet as simple side-by-side tables."""
    ws = wb.create_sheet("Assessment Areas")
    
    # Headers
    ws['A1'] = f'{bank_a_name} Assessment Areas'
    ws['A1'].font = Font(bold=True, size=12)
    ws['G1'] = f'{bank_b_name} Assessment Areas'
    ws['G1'].font = Font(bold=True, size=12)
    
    # Column headers for Bank A
    headers_a = ['State', 'Assessment Area', 'CBSA Name', 'CBSA Code', 'County, State', 'GEOID5']
    for col_idx, header in enumerate(headers_a, 1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers for Bank B
    headers_b = ['State', 'Assessment Area', 'CBSA Name', 'CBSA Code', 'County, State', 'GEOID5']
    for col_idx, header in enumerate(headers_b, 7):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get counties and assessment area mappings
    bank_a_key = 'acquirer' if 'acquirer' in assessment_areas_data else 'pnc'
    bank_a_data = assessment_areas_data.get(bank_a_key, {})
    counties_a = bank_a_data.get('mapped_counties', bank_a_data.get('counties', []))
    if not counties_a or not isinstance(counties_a, list):
        if isinstance(assessment_areas_data.get(bank_a_key), list):
            counties_a = assessment_areas_data.get(bank_a_key)
    
    bank_b_key = 'target' if 'target' in assessment_areas_data else 'firstbank'
    bank_b_data = assessment_areas_data.get(bank_b_key, {})
    counties_b = bank_b_data.get('mapped_counties', bank_b_data.get('counties', []))
    if not counties_b or not isinstance(counties_b, list):
        if isinstance(assessment_areas_data.get(bank_b_key), list):
            counties_b = assessment_areas_data.get(bank_b_key)
    
    # Get assessment area names from metadata if available
    # Build maps from GEOID5 to assessment area name for both banks
    bank_a_aa_map = {}  # Map GEOID5 to assessment area name
    bank_b_aa_map = {}  # Map GEOID5 to assessment area name
    
    if 'assessment_areas' in assessment_areas_data:
        aa_list = assessment_areas_data.get('assessment_areas', [])
        # Split into acquirer and target assessment areas
        # We'll match counties to assessment areas by checking if county GEOID5s match
        for aa_info in aa_list:
            aa_name = aa_info.get('name') or aa_info.get('cbsa_name') or aa_info.get('assessment_area') or 'Unnamed Assessment Area'
            counties_list = aa_info.get('counties', [])
            for county_item in counties_list:
                geoid5 = None
                if isinstance(county_item, dict):
                    geoid5 = str(county_item.get('geoid5', '')).zfill(5)
                    if not geoid5 or len(geoid5) != 5:
                        # Try to build from state_code and county_code
                        state_code = county_item.get('state_code') or county_item.get('state_fips')
                        county_code = county_item.get('county_code') or county_item.get('county_fips')
                        if state_code and county_code:
                            geoid5 = str(state_code).zfill(2) + str(county_code).zfill(3)
                
                if geoid5 and len(geoid5) == 5:
                    # Check if this GEOID5 is in bank A or bank B counties
                    for county in counties_a:
                        county_geoid5 = str(county.get('geoid5', '')).zfill(5)
                        if county_geoid5 == geoid5:
                            bank_a_aa_map[geoid5] = aa_name
                            break
                    
                    for county in counties_b:
                        county_geoid5 = str(county.get('geoid5', '')).zfill(5)
                        if county_geoid5 == geoid5:
                            bank_b_aa_map[geoid5] = aa_name
                            break
    
    # Build CBSA name cache
    cbsa_name_cache = {}
    
    # Populate Bank A data
    row = 3
    if counties_a:
        # Group by state and assessment area
        state_groups = defaultdict(lambda: defaultdict(list))
        for county in counties_a:
            state = county.get('state_name', county.get('state', '')) or 'Unknown'
            # Try to get assessment area name from map
            geoid5 = str(county.get('geoid5', '')).zfill(5)
            aa = bank_a_aa_map.get(geoid5) or county.get('assessment_area') or county.get('aa_name') or 'Unnamed Assessment Area'
            state_groups[state][aa].append(county)
        
        for state in sorted(state_groups.keys()):
            for aa in sorted(state_groups[state].keys()):
                for county in state_groups[state][aa]:
                    ws.cell(row, 1).value = state
                    # Use CBSA name as Assessment Area name
                    cbsa_code = str(county.get('cbsa_code', ''))
                    cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
                    ws.cell(row, 2).value = cbsa_name  # Assessment Area = CBSA Name
                    ws.cell(row, 3).value = cbsa_name
                    ws.cell(row, 4).value = cbsa_code
                    county_name = county.get('county_name', '')
                    state_name = county.get('state_name', county.get('state', ''))
                    ws.cell(row, 5).value = f"{county_name}, {state_name}" if county_name and state_name else (county_name or state_name)
                    ws.cell(row, 6).value = str(county.get('geoid5', ''))
                    row += 1
    
    # Populate Bank B data
    row = 3
    if counties_b:
        state_groups = defaultdict(lambda: defaultdict(list))
        for county in counties_b:
            state = county.get('state_name', county.get('state', '')) or 'Unknown'
            # Try to get assessment area name from map
            geoid5 = str(county.get('geoid5', '')).zfill(5)
            aa = bank_b_aa_map.get(geoid5) or county.get('assessment_area') or county.get('aa_name') or 'Unnamed Assessment Area'
            state_groups[state][aa].append(county)
        
        for state in sorted(state_groups.keys()):
            for aa in sorted(state_groups[state].keys()):
                for county in state_groups[state][aa]:
                    ws.cell(row, 7).value = state
                    # Use CBSA name as Assessment Area name
                    cbsa_code = str(county.get('cbsa_code', ''))
                    cbsa_name = county.get('cbsa_name', '') or _get_cbsa_name_from_code(cbsa_code, cbsa_name_cache)
                    ws.cell(row, 8).value = cbsa_name  # Assessment Area = CBSA Name
                    ws.cell(row, 9).value = cbsa_name
                    ws.cell(row, 10).value = cbsa_code
                    county_name = county.get('county_name', '')
                    state_name = county.get('state_name', county.get('state', ''))
                    ws.cell(row, 11).value = f"{county_name}, {state_name}" if county_name and state_name else (county_name or state_name)
                    ws.cell(row, 12).value = str(county.get('geoid5', ''))
                    row += 1
    
    # Auto-adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    print(f"  Created Assessment Areas sheet: {len(counties_a) if counties_a else 0} counties for {bank_a_name}, {len(counties_b) if counties_b else 0} counties for {bank_b_name}")


