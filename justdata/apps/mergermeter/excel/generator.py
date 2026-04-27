"""Mergermeter Excel report generator.

create_merger_excel orchestrates the per-worksheet builders to produce
the merger analysis workbook. The shared generator under
justdata.shared.reporting.merger_excel_generator is used as the
preferred path; this module is the fallback / legacy implementation.
"""
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from justdata.apps.mergermeter.config import PROJECT_ID
from justdata.apps.mergermeter.excel.utils import (
    _get_cbsa_name_from_code,
    _transform_mortgage_goals_data,
)
from justdata.apps.mergermeter.excel.worksheets.assessment_areas import (
    create_simple_assessment_areas_sheet,
)
from justdata.apps.mergermeter.excel.worksheets.branch import create_simple_branch_sheet
from justdata.apps.mergermeter.excel.worksheets.hhi import (
    _add_empty_hhi_sheet,
    _add_hhi_sheet,
)
from justdata.apps.mergermeter.excel.worksheets.mortgage import create_simple_mortgage_sheet
from justdata.apps.mergermeter.excel.worksheets.notes import create_simple_notes_sheet
from justdata.apps.mergermeter.excel.worksheets.sb import create_simple_sb_sheet
from justdata.shared.utils.bigquery_client import execute_query, get_bigquery_client

# Import shared modules
try:
    from justdata.shared.reporting.merger_data_transformer import (
        transform_branch_data,
        transform_mortgage_data,
        transform_sb_data,
    )
    from justdata.shared.reporting.excel import create_merger_excel as shared_create_merger_excel
    SHARED_GENERATOR_AVAILABLE = True
except ImportError as e:
    SHARED_GENERATOR_AVAILABLE = False
    logging.warning(f"Shared generator not available: {e}. Falling back to legacy implementation.")

logger = logging.getLogger(__name__)


def create_merger_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame] = None,
    bank_a_hmda_peer: Optional[pd.DataFrame] = None,
    bank_b_hmda_subject: Optional[pd.DataFrame] = None,
    bank_b_hmda_peer: Optional[pd.DataFrame] = None,
    bank_a_sb_subject: Optional[pd.DataFrame] = None,
    bank_a_sb_peer: Optional[pd.DataFrame] = None,
    bank_b_sb_subject: Optional[pd.DataFrame] = None,
    bank_b_sb_peer: Optional[pd.DataFrame] = None,
    bank_a_branch: Optional[pd.DataFrame] = None,
    bank_b_branch: Optional[pd.DataFrame] = None,
    hhi_data: Optional[pd.DataFrame] = None,
    assessment_areas: Optional[Dict] = None,
    metadata: Optional[Dict] = None,
    mortgage_goals_data: Optional[Dict] = None,
    sb_goals_data: Optional[pd.DataFrame] = None
):
    """
    Create Excel workbook with merger analysis data using shared generator.

    This function:
    1. Transforms query results using shared transformer
    2. Builds assessment_areas_data dictionary
    3. Calls shared Excel generator
    4. Validates the output workbook for data quality issues

    Args:
        output_path: Path to save Excel file
        bank_a_name: Acquirer bank name
        bank_b_name: Target bank name
        All data DataFrames for HMDA, SB, Branch, and HHI
        assessment_areas: Assessment area information (dict with 'acquirer' and 'target' keys)
        metadata: Additional metadata (years, loan purpose, LEI, RSSD, etc.)

    Returns:
        List of validation warning dicts, or empty list if no issues found.
    """
    
    if not SHARED_GENERATOR_AVAILABLE:
        logger.error("Shared generator not available. Cannot create Excel file.")
        raise ImportError("Shared Excel generator modules not found. Please ensure shared.reporting modules are available.")
    
    # Extract metadata
    hmda_years = metadata.get('hmda_years', []) if metadata else []
    sb_years = metadata.get('sb_years', []) if metadata else []
    baseline_hmda_years = metadata.get('baseline_hmda_years', []) if metadata else []
    baseline_sb_years = metadata.get('baseline_sb_years', []) if metadata else []
    bank_a_lei = metadata.get('acquirer_lei', '') if metadata else ''
    bank_b_lei = metadata.get('target_lei', '') if metadata else ''
    bank_a_rssd = metadata.get('acquirer_rssd', '') if metadata else ''
    bank_b_rssd = metadata.get('target_rssd', '') if metadata else ''
    bank_a_sb_id = metadata.get('acquirer_sb_id', '') if metadata else ''
    bank_b_sb_id = metadata.get('target_sb_id', '') if metadata else ''
    
    # Extract filter parameters
    loan_purpose = metadata.get('loan_purpose', '') if metadata else ''
    action_taken = metadata.get('action_taken', '') if metadata else ''
    occupancy_type = metadata.get('occupancy_type', '') if metadata else ''
    total_units = metadata.get('total_units', '') if metadata else ''
    construction_method = metadata.get('construction_method', '') if metadata else ''
    not_reverse = metadata.get('not_reverse', '') if metadata else ''
    
    # Build assessment_areas_data dictionary from assessment_areas
    # Expected format: {'acquirer': {'counties': [...]}, 'target': {'counties': [...]}}
    assessment_areas_data = {}
    if assessment_areas:
        # Get enriched counties from assessment_areas dict
        acquirer_counties = assessment_areas.get('acquirer', {}).get('counties', [])
        target_counties = assessment_areas.get('target', {}).get('counties', [])
        
        assessment_areas_data = {
            'acquirer': {'counties': acquirer_counties},
            'target': {'counties': target_counties}
        }
    else:
        # Empty structure if no assessment areas provided
        assessment_areas_data = {
            'acquirer': {'counties': []},
            'target': {'counties': []}
        }
    
    # Build CBSA name map and required CBSAs from assessment areas
    # Also build reverse map from CBSA names to codes for matching
    cbsa_name_map = {}
    cbsa_name_to_code_map = {}  # Reverse mapping: name -> code
    bank_a_required_cbsas = set()
    bank_b_required_cbsas = set()
    
    def normalize_cbsa_for_matching(code_or_name):
        """Normalize CBSA code or name for matching (handles special cases)."""
        if not code_or_name:
            return None
        s = str(code_or_name).strip()
        # Handle special cases
        if s == '0' or s.lower() == 'nan' or s == '--':
            return None
        # Normalize "Rural *" to "NON-METRO-*"
        if s.lower().startswith('rural '):
            state = s[6:].strip()  # Remove "Rural " prefix
            return f'NON-METRO-{state}'
        # Normalize "State Non-Metro Area" or "State Non-MSA" to "NON-METRO-State"
        if ' non-metro' in s.lower() or ' non-msa' in s.lower():
            # Extract state name (everything before " Non-Metro" or " Non-MSA")
            import re
            match = re.match(r'^(.+?)\s+Non-(?:Metro|MSA)', s, re.IGNORECASE)
            state = match.group(1) if match else s.split()[0]
            return f'NON-METRO-{state}'
        return s
    
    for county in assessment_areas_data.get('acquirer', {}).get('counties', []):
        cbsa_code = str(county.get('cbsa_code', '')).strip()
        cbsa_name = county.get('cbsa_name', '').strip()
        state_name = county.get('state_name', '').strip()

        # Handle rural counties (CBSA code 99999) - these are valid, not invalid
        if cbsa_code == '99999':
            # Add the literal '99999' code since HMDA queries return this for rural areas
            bank_a_required_cbsas.add('99999')
            # Also add state-specific rural identifier for better matching
            if state_name:
                rural_name = f"Rural {state_name}"
                normalized_name = normalize_cbsa_for_matching(rural_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = rural_name
                    cbsa_name_map[normalized_name] = rural_name
            elif cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = cbsa_name
            continue

        # Skip truly invalid CBSA codes (but not 99999 which is valid for rural)
        if cbsa_code in ['0', '--', ''] or cbsa_code.lower() == 'nan':
            # If code is invalid but name exists, use name
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_a_required_cbsas.add(normalized_name)
                    cbsa_name_map[normalized_name] = cbsa_name
            continue
        
        # Normalize code
        normalized_code = normalize_cbsa_for_matching(cbsa_code)
        if normalized_code:
            bank_a_required_cbsas.add(normalized_code)
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                cbsa_name_map[normalized_code] = cbsa_name
                # Also map name to code for reverse lookup
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    cbsa_name_to_code_map[normalized_name] = normalized_code
        
        # Also add by name if code is missing but name exists
        if not normalized_code and cbsa_name and cbsa_name != '0':
            normalized_name = normalize_cbsa_for_matching(cbsa_name)
            if normalized_name:
                bank_a_required_cbsas.add(normalized_name)  # Use name as key if no code
                cbsa_name_map[normalized_name] = cbsa_name
    
    for county in assessment_areas_data.get('target', {}).get('counties', []):
        cbsa_code = str(county.get('cbsa_code', '')).strip()
        cbsa_name = county.get('cbsa_name', '').strip()
        state_name = county.get('state_name', '').strip()

        # Handle rural counties (CBSA code 99999) - these are valid, not invalid
        if cbsa_code == '99999':
            # Add the literal '99999' code since HMDA queries return this for rural areas
            bank_b_required_cbsas.add('99999')
            # Also add state-specific rural identifier for better matching
            if state_name:
                rural_name = f"Rural {state_name}"
                normalized_name = normalize_cbsa_for_matching(rural_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = rural_name
                    cbsa_name_map[normalized_name] = rural_name
            elif cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map['99999'] = cbsa_name
            continue

        # Skip truly invalid CBSA codes (but not 99999 which is valid for rural)
        if cbsa_code in ['0', '--', ''] or cbsa_code.lower() == 'nan':
            # If code is invalid but name exists, use name
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    bank_b_required_cbsas.add(normalized_name)
                    cbsa_name_map[normalized_name] = cbsa_name
            continue
        
        # Normalize code
        normalized_code = normalize_cbsa_for_matching(cbsa_code)
        if normalized_code:
            bank_b_required_cbsas.add(normalized_code)
            if cbsa_name and cbsa_name != '--' and cbsa_name != '0':
                cbsa_name_map[normalized_code] = cbsa_name
                # Also map name to code for reverse lookup
                normalized_name = normalize_cbsa_for_matching(cbsa_name)
                if normalized_name:
                    cbsa_name_to_code_map[normalized_name] = normalized_code
        
        # Also add by name if code is missing but name exists
        if not normalized_code and cbsa_name and cbsa_name != '0':
            normalized_name = normalize_cbsa_for_matching(cbsa_name)
            if normalized_name:
                bank_b_required_cbsas.add(normalized_name)  # Use name as key if no code
                cbsa_name_map[normalized_name] = cbsa_name
    
    # Also extract CBSA names from query results if available
    # Filter out invalid CBSA codes like '0', '99999', etc.
    for df in [bank_a_hmda_subject, bank_a_hmda_peer, bank_b_hmda_subject, bank_b_hmda_peer,
                bank_a_sb_subject, bank_a_sb_peer, bank_b_sb_subject, bank_b_sb_peer,
                bank_a_branch, bank_b_branch]:
        if df is not None and not df.empty and 'cbsa_code' in df.columns and 'cbsa_name' in df.columns:
            for _, row in df.iterrows():
                cbsa_code = str(row.get('cbsa_code', '')).strip()
                cbsa_name = str(row.get('cbsa_name', '')).strip()
                # Filter out invalid CBSA codes
                if cbsa_code and cbsa_code != '--' and cbsa_code.lower() != 'nan' and cbsa_code != '0' and cbsa_code != '99999':
                    if cbsa_name and cbsa_name != '--' and cbsa_name.lower() not in ['nan', 'none', ''] and cbsa_name != '0':
                        cbsa_name_map[cbsa_code] = cbsa_name
    
    logger.info(f"Built CBSA name map with {len(cbsa_name_map)} entries")
    logger.info(f"Bank A required CBSAs: {len(bank_a_required_cbsas)}")
    logger.info(f"Bank B required CBSAs: {len(bank_b_required_cbsas)}")

    # Transform data using shared transformer
    # Ensure DataFrames are not None (use empty DataFrame instead)
    bank_a_mortgage_transformed = transform_mortgage_data(
        subject_df=bank_a_hmda_subject if bank_a_hmda_subject is not None and not bank_a_hmda_subject.empty else pd.DataFrame(),
        peer_df=bank_a_hmda_peer if bank_a_hmda_peer is not None and not bank_a_hmda_peer.empty else pd.DataFrame(),
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )

    # Transform peer data separately for Excel sheet display
    bank_a_mortgage_peer_transformed = transform_mortgage_data(
        subject_df=bank_a_hmda_peer if bank_a_hmda_peer is not None and not bank_a_hmda_peer.empty else pd.DataFrame(),
        peer_df=pd.DataFrame(),  # No peer-of-peer
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )

    bank_b_mortgage_transformed = transform_mortgage_data(
        subject_df=bank_b_hmda_subject if bank_b_hmda_subject is not None and not bank_b_hmda_subject.empty else pd.DataFrame(),
        peer_df=bank_b_hmda_peer if bank_b_hmda_peer is not None and not bank_b_hmda_peer.empty else pd.DataFrame(),
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )

    # Transform peer data separately for Excel sheet display
    bank_b_mortgage_peer_transformed = transform_mortgage_data(
        subject_df=bank_b_hmda_peer if bank_b_hmda_peer is not None and not bank_b_hmda_peer.empty else pd.DataFrame(),
        peer_df=pd.DataFrame(),  # No peer-of-peer
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )
    
    # Rename SB columns to match transformer expectations
    # MergerMeter uses: sb_loans_count, lmict_loans_count, loans_rev_under_1m, lmict_loans_amount, amount_rev_under_1m
    # Transformer expects: sb_loans_total, lmict_count, loans_rev_under_1m_count, avg_sb_lmict_loan_amount, avg_loan_amt_rum_sb
    def rename_sb_columns(df):
        """Rename SB DataFrame columns to match transformer expectations."""
        if df is None or df.empty:
            return pd.DataFrame()  # Return empty DataFrame with correct structure
        
        df = df.copy()
        
        # Log original columns for debugging
        print(f"[DEBUG] rename_sb_columns - Original columns: {list(df.columns)}")
        logger.debug(f"Original SB DataFrame columns: {list(df.columns)}")
        
        # Rename columns - check if they exist first
        rename_map = {}
        if 'sb_loans_count' in df.columns:
            rename_map['sb_loans_count'] = 'sb_loans_total'
            print(f"[DEBUG] Will rename sb_loans_count -> sb_loans_total")
        elif 'sb_loans_total' not in df.columns:
            warning_msg = "Neither sb_loans_count nor sb_loans_total found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if 'lmict_loans_count' in df.columns:
            rename_map['lmict_loans_count'] = 'lmict_count'
            print(f"[DEBUG] Will rename lmict_loans_count -> lmict_count")
        elif 'lmict_count' not in df.columns:
            warning_msg = "Neither lmict_loans_count nor lmict_count found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if 'loans_rev_under_1m' in df.columns:
            rename_map['loans_rev_under_1m'] = 'loans_rev_under_1m_count'
            print(f"[DEBUG] Will rename loans_rev_under_1m -> loans_rev_under_1m_count")
        elif 'loans_rev_under_1m_count' not in df.columns:
            warning_msg = "Neither loans_rev_under_1m nor loans_rev_under_1m_count found in SB DataFrame"
            print(f"[WARNING] {warning_msg}")
            logger.warning(warning_msg)
        
        if rename_map:
            df = df.rename(columns=rename_map)
        
        # Calculate averages from raw amounts and counts (before aggregation)
        # The transformer will aggregate across years, so we calculate per-row averages first
        if 'lmict_loans_amount' in df.columns and 'lmict_count' in df.columns:
            # Calculate avg_sb_lmict_loan_amount per row
            df['avg_sb_lmict_loan_amount'] = df.apply(
                lambda row: row['lmict_loans_amount'] / row['lmict_count'] 
                if pd.notna(row['lmict_count']) and row['lmict_count'] > 0 else None, axis=1
            )
        elif 'avg_sb_lmict_loan_amount' not in df.columns:
            # If we don't have the data to calculate, set to None
            logger.warning("Cannot calculate avg_sb_lmict_loan_amount - missing required columns")
            df['avg_sb_lmict_loan_amount'] = None
        
        if 'amount_rev_under_1m' in df.columns and 'loans_rev_under_1m_count' in df.columns:
            # Calculate avg_loan_amt_rum_sb per row
            df['avg_loan_amt_rum_sb'] = df.apply(
                lambda row: row['amount_rev_under_1m'] / row['loans_rev_under_1m_count']
                if pd.notna(row['loans_rev_under_1m_count']) and row['loans_rev_under_1m_count'] > 0 else None, axis=1
            )
        elif 'avg_loan_amt_rum_sb' not in df.columns:
            # If we don't have the data to calculate, set to None
            logger.warning("Cannot calculate avg_loan_amt_rum_sb - missing required columns")
            df['avg_loan_amt_rum_sb'] = None
        
        print(f"[DEBUG] rename_sb_columns - Final columns: {list(df.columns)}")
        logger.debug(f"Renamed SB DataFrame columns: {list(df.columns)}")
        
        # Verify required columns exist
        required_cols = ['sb_loans_total', 'lmict_count', 'loans_rev_under_1m_count']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            error_msg = f"Missing required columns after rename: {missing_cols}. Available columns: {list(df.columns)}"
            print(f"[ERROR] {error_msg}")
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        return df
    
    try:
        # Log original column names for debugging
        if bank_a_sb_subject is not None and not bank_a_sb_subject.empty:
            print(f"[DEBUG] Bank A SB subject original columns: {list(bank_a_sb_subject.columns)}")
            logger.info(f"Bank A SB subject original columns: {list(bank_a_sb_subject.columns)}")
        if bank_a_sb_peer is not None and not bank_a_sb_peer.empty:
            print(f"[DEBUG] Bank A SB peer original columns: {list(bank_a_sb_peer.columns)}")
            logger.info(f"Bank A SB peer original columns: {list(bank_a_sb_peer.columns)}")
        if bank_b_sb_subject is not None and not bank_b_sb_subject.empty:
            print(f"[DEBUG] Bank B SB subject original columns: {list(bank_b_sb_subject.columns)}")
            logger.info(f"Bank B SB subject original columns: {list(bank_b_sb_subject.columns)}")
        if bank_b_sb_peer is not None and not bank_b_sb_peer.empty:
            print(f"[DEBUG] Bank B SB peer original columns: {list(bank_b_sb_peer.columns)}")
            logger.info(f"Bank B SB peer original columns: {list(bank_b_sb_peer.columns)}")
        
        bank_a_sb_subject_renamed = rename_sb_columns(
            bank_a_sb_subject.copy() if bank_a_sb_subject is not None and not bank_a_sb_subject.empty else pd.DataFrame()
        )
        bank_a_sb_peer_renamed = rename_sb_columns(
            bank_a_sb_peer.copy() if bank_a_sb_peer is not None and not bank_a_sb_peer.empty else pd.DataFrame()
        )
        bank_b_sb_subject_renamed = rename_sb_columns(
            bank_b_sb_subject.copy() if bank_b_sb_subject is not None and not bank_b_sb_subject.empty else pd.DataFrame()
        )
        bank_b_sb_peer_renamed = rename_sb_columns(
            bank_b_sb_peer.copy() if bank_b_sb_peer is not None and not bank_b_sb_peer.empty else pd.DataFrame()
        )
        
        # Log column names after rename for debugging
        if not bank_a_sb_subject_renamed.empty:
            print(f"[DEBUG] Bank A SB subject columns after rename: {list(bank_a_sb_subject_renamed.columns)}")
            logger.info(f"Bank A SB subject columns after rename: {list(bank_a_sb_subject_renamed.columns)}")
        else:
            print(f"[DEBUG] Bank A SB subject DataFrame is empty after rename")
        if not bank_a_sb_peer_renamed.empty:
            print(f"[DEBUG] Bank A SB peer columns after rename: {list(bank_a_sb_peer_renamed.columns)}")
            logger.info(f"Bank A SB peer columns after rename: {list(bank_a_sb_peer_renamed.columns)}")
        else:
            print(f"[DEBUG] Bank A SB peer DataFrame is empty after rename")
        
        bank_a_sb_transformed = transform_sb_data(
            subject_df=bank_a_sb_subject_renamed,
            peer_df=bank_a_sb_peer_renamed,
            cbsa_name_map=cbsa_name_map,
            required_cbsas=bank_a_required_cbsas
        )
        
        bank_b_sb_transformed = transform_sb_data(
            subject_df=bank_b_sb_subject_renamed,
            peer_df=bank_b_sb_peer_renamed,
            cbsa_name_map=cbsa_name_map,
            required_cbsas=bank_b_required_cbsas
        )
    except Exception as e:
        error_msg = f"Error transforming SB data: {e}"
        print(f"[ERROR] {error_msg}")
        print(f"[ERROR] Exception type: {type(e).__name__}")
        
        # Print column information for all SB DataFrames
        print("\n[ERROR] SB DataFrame Column Information:")
        if bank_a_sb_subject is not None:
            print(f"  Bank A SB Subject: {list(bank_a_sb_subject.columns) if not bank_a_sb_subject.empty else 'EMPTY'}")
        else:
            print(f"  Bank A SB Subject: None")
        if bank_a_sb_peer is not None:
            print(f"  Bank A SB Peer: {list(bank_a_sb_peer.columns) if not bank_a_sb_peer.empty else 'EMPTY'}")
        else:
            print(f"  Bank A SB Peer: None")
        if bank_b_sb_subject is not None:
            print(f"  Bank B SB Subject: {list(bank_b_sb_subject.columns) if not bank_b_sb_subject.empty else 'EMPTY'}")
        else:
            print(f"  Bank B SB Subject: None")
        if bank_b_sb_peer is not None:
            print(f"  Bank B SB Peer: {list(bank_b_sb_peer.columns) if not bank_b_sb_peer.empty else 'EMPTY'}")
        else:
            print(f"  Bank B SB Peer: None")
        
        logger.error(error_msg, exc_info=True)
        # Return empty DataFrames if transformation fails
        bank_a_sb_transformed = pd.DataFrame(columns=['assessment_area', 'metric', 'bank_value', 'peer_value', 'difference'])
        bank_b_sb_transformed = pd.DataFrame(columns=['assessment_area', 'metric', 'bank_value', 'peer_value', 'difference'])
        bank_a_sb_peer_renamed = pd.DataFrame()
        bank_b_sb_peer_renamed = pd.DataFrame()
    
    # Transform branch data - need to rename columns to match transformer expectations
    # MergerMeter uses: market_total_branches, market_branches_in_lmict, market_branches_in_mmct
    # Transformer expects: other_total_branches, other_branches_in_lmict, other_branches_in_mmct
    bank_a_branch_for_transform = bank_a_branch.copy() if bank_a_branch is not None and not bank_a_branch.empty else pd.DataFrame()
    if not bank_a_branch_for_transform.empty:
        # Rename market columns to other columns
        rename_map = {
            'market_total_branches': 'other_total_branches',
            'market_branches_in_lmict': 'other_branches_in_lmict',
            'market_branches_in_mmct': 'other_branches_in_mmct'
        }
        bank_a_branch_for_transform = bank_a_branch_for_transform.rename(columns=rename_map)
    
    bank_b_branch_for_transform = bank_b_branch.copy() if bank_b_branch is not None and not bank_b_branch.empty else pd.DataFrame()
    if not bank_b_branch_for_transform.empty:
        # Rename market columns to other columns
        rename_map = {
            'market_total_branches': 'other_total_branches',
            'market_branches_in_lmict': 'other_branches_in_lmict',
            'market_branches_in_mmct': 'other_branches_in_mmct'
        }
        bank_b_branch_for_transform = bank_b_branch_for_transform.rename(columns=rename_map)
    
    bank_a_branch_transformed = transform_branch_data(
        branch_df=bank_a_branch_for_transform,
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_a_required_cbsas
    )
    
    bank_b_branch_transformed = transform_branch_data(
        branch_df=bank_b_branch_for_transform,
        cbsa_name_map=cbsa_name_map,
        required_cbsas=bank_b_required_cbsas
    )
    
    logger.info("Data transformation complete. Calling shared Excel generator...")

    # Transform mortgage goals data to expected format
    transformed_mortgage_goals = _transform_mortgage_goals_data(mortgage_goals_data)
    if transformed_mortgage_goals:
        logger.info(f"Transformed mortgage_goals_data: {len(transformed_mortgage_goals.get('by_state', {}))} states, grand_total has {len(transformed_mortgage_goals.get('grand_total', {}))} loan types")

    # Determine if single bank mode from metadata
    single_bank_mode = metadata.get('single_bank_mode', False) if metadata else False

    # Call shared generator
    shared_create_merger_excel(
        output_path=output_path,
        bank_a_name=bank_a_name,
        bank_b_name=bank_b_name,
        assessment_areas_data=assessment_areas_data,
        mortgage_goals_data=transformed_mortgage_goals,
        sb_goals_data=sb_goals_data if sb_goals_data is not None and not sb_goals_data.empty else None,
        bank_a_mortgage_data=bank_a_mortgage_transformed,
        bank_b_mortgage_data=bank_b_mortgage_transformed,
        bank_a_mortgage_peer_data=bank_a_mortgage_peer_transformed,
        bank_b_mortgage_peer_data=bank_b_mortgage_peer_transformed,
        bank_a_sb_data=bank_a_sb_transformed,
        bank_b_sb_data=bank_b_sb_transformed,
        bank_a_sb_peer_data=bank_a_sb_peer_renamed,
        bank_b_sb_peer_data=bank_b_sb_peer_renamed,
        bank_a_branch_data=bank_a_branch_transformed,
        bank_b_branch_data=bank_b_branch_transformed,
        years_hmda=hmda_years,
        years_sb=sb_years,
        baseline_years_hmda=baseline_hmda_years,
        baseline_years_sb=baseline_sb_years,
        bank_a_lei=bank_a_lei,
        bank_b_lei=bank_b_lei,
        bank_a_rssd=bank_a_rssd,
        bank_b_rssd=bank_b_rssd,
        bank_a_sb_id=bank_a_sb_id,
        bank_b_sb_id=bank_b_sb_id,
        loan_purpose=loan_purpose,
        action_taken=action_taken,
        occupancy_type=occupancy_type,
        total_units=total_units,
        construction_method=construction_method,
        not_reverse=not_reverse,
        single_bank_mode=single_bank_mode
    )
    
    logger.info(f"Excel workbook saved to: {output_path}")
    print(f"\n[OK] Excel workbook saved to: {output_path}")
    
    # Add HHI sheet - always create it, with explanatory note if empty
    try:
        wb = load_workbook(output_path)
        if hhi_data is not None and not hhi_data.empty:
            logger.info(f"[HHI] HHI data provided - Shape: {hhi_data.shape}, Columns: {list(hhi_data.columns)}")
            _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
            logger.info("Added HHI Analysis sheet with data")
        else:
            logger.warning(f"[HHI] No HHI data - creating sheet with explanatory note")
            _add_empty_hhi_sheet(wb, bank_a_name, bank_b_name, bank_a_rssd, bank_b_rssd)
            logger.info("Added HHI Analysis sheet with no-data explanation")

        # Run output validation before final save
        from justdata.apps.mergermeter.output_validator import validate_workbook, add_warnings_sheet
        validation_warnings = validate_workbook(wb)
        add_warnings_sheet(wb, validation_warnings)  # Always add — shows clean message if empty
        if validation_warnings:
            print(f"[VALIDATION] {len(validation_warnings)} data quality issues found:")
            for w in validation_warnings:
                print(f"  [{w['severity'].upper()}] {w['sheet']}: {w['issue']}")
        else:
            print("[VALIDATION] No data quality issues found.")

        wb.save(output_path)
        return validation_warnings
    except Exception as e:
        logger.error(f"Could not add HHI sheet: {e}")
        import traceback
        traceback.print_exc()
        return []


