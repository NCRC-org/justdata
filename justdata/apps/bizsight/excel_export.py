#!/usr/bin/env python3
"""
Excel export functionality for BizSight reports.
Creates Excel files with multiple sheets including Notes, all table tabs, and data.
Each tabbed table gets separate sheets (e.g., "County Summary - Number" and "County Summary - Amount").
"""

import pandas as pd
import os
from typing import Dict, Any, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize sheet name for Excel compatibility.
    
    Excel sheet names cannot contain: \ / ? * [ ]
    And must be <= 31 characters.
    """
    # Remove invalid characters
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncate if too long
    if len(name) > max_length:
        name = name[:max_length]
    
    return name


def save_bizsight_excel_report(analysis_result: Dict[str, Any], output_path: str, metadata: Dict[str, Any] = None):
    """
    Save the BizSight report data to an Excel file with multiple sheets.
    
    Sheets:
    1. Notes (Methods, Definitions, Calculations)
    2. County Summary - Number of Loans (Section 2, Number tab)
    3. County Summary - Amount of Loans (Section 2, Amount tab)
    4. Comparison - Number of Loans (Section 3, Number tab)
    5. Comparison - Amount of Loans (Section 3, Amount tab)
    6. Top Lenders - Number of Loans (Section 4, Number tab)
    7. Top Lenders - Amount of Loans (Section 4, Amount tab)
    8. HHI by Year (Section 5)
    9. Raw Data (if available)
    """
    from openpyxl import Workbook
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    # Get data from analysis_result
    county_summary_table = analysis_result.get('county_summary_table', [])
    comparison_table = analysis_result.get('comparison_table', [])
    top_lenders_table = analysis_result.get('top_lenders_table', [])
    hhi_by_year = analysis_result.get('hhi_by_year', [])
    report_data = analysis_result.get('report_data', {})
    aggregate_df = report_data.get('aggregate_data', pd.DataFrame()) if isinstance(report_data, dict) else pd.DataFrame()
    
    # Get metadata
    county_name = metadata.get('county_name', 'Unknown County') if metadata else 'Unknown County'
    state_name = metadata.get('state_name', '') if metadata else ''
    years = metadata.get('years', [2018, 2019, 2020, 2021, 2022, 2023, 2024]) if metadata else [2018, 2019, 2020, 2021, 2022, 2023, 2024]
    years_str = f"{min(years)}-{max(years)}"
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Get workbook and remove default sheet if it exists
        workbook = writer.book
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Notes sheet (first sheet) with comprehensive methods and definitions
        notes_sheet = workbook.create_sheet(sanitize_sheet_name('Notes'), 0)
        
        # Prepare notes content
        notes_content = []
        
        notes_content.append(('METHODS AND DEFINITIONS', ''))
        notes_content.append(('', ''))
        
        notes_content.append(('Data Sources', ''))
        notes_content.append(('', 'This report uses data from the Community Reinvestment Act (CRA) Small Business Lending data.'))
        notes_content.append(('', 'Data compiled and maintained in NCRC\'s curated BigQuery databases.'))
        notes_content.append(('', 'U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates for demographic context.'))
        notes_content.append(('', 'Census data retrieved from geo.census table in BigQuery.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Report Configuration', ''))
        notes_content.append(('', f'County: {county_name}'))
        if state_name:
            notes_content.append(('', f'State: {state_name}'))
        notes_content.append(('', f'Years: {years_str}'))
        notes_content.append(('', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'))
        notes_content.append(('', ''))
        
        notes_content.append(('Data Limitations', ''))
        notes_content.append(('', 'The data in this report reflects only lending activity reported by financial institutions subject to CRA reporting requirements.'))
        notes_content.append(('', 'It does not include lending by institutions that are exempt from reporting, nor does it include non-traditional lending sources.'))
        notes_content.append(('', 'Additionally, the data may not capture the full extent of small business lending activity, particularly for very small loans.'))
        notes_content.append(('', 'Important Note: In 2020 and 2021, the Paycheck Protection Program (PPP) significantly impacted small business lending patterns.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Income Classifications', ''))
        notes_content.append(('', 'Low Income: ≤50% of Area Median Income (AMI)'))
        notes_content.append(('', 'Moderate Income: 50-80% of AMI'))
        notes_content.append(('', 'Middle Income: 80-120% of AMI'))
        notes_content.append(('', 'Upper Income: >120% of AMI'))
        notes_content.append(('', 'LMI (Low-to-Moderate Income): ≤80% of AMI'))
        notes_content.append(('', ''))
        
        notes_content.append(('Income Group Codes', ''))
        notes_content.append(('', 'Low Income: 001-005 or 101'))
        notes_content.append(('', 'Moderate Income: 006-008 or 102'))
        notes_content.append(('', 'Middle Income: 004, 009-012, or 103'))
        notes_content.append(('', 'Upper Income: 013-014 or 104'))
        notes_content.append(('', 'Unknown: All other codes'))
        notes_content.append(('', ''))
        
        notes_content.append(('Race and Minority Classifications', ''))
        notes_content.append(('', 'Minority percentage is calculated using standard deviation of minority population for the entire county.'))
        notes_content.append(('', 'Categories: Low Minority, Moderate Minority, Middle Minority, Upper Minority'))
        notes_content.append(('', 'Race categories include: White, Black, Hispanic, Asian, and Other'))
        notes_content.append(('', ''))
        
        notes_content.append(('Loan Size Categories', ''))
        notes_content.append(('', 'Under $100K: Loans with original amount < $100,000'))
        notes_content.append(('', '$100K-$250K: Loans with original amount between $100,000 and $250,000'))
        notes_content.append(('', '$250K-$1M: Loans with original amount between $250,000 and $1,000,000'))
        notes_content.append(('', 'Under $1M (SB): Small business loans under $1,000,000'))
        notes_content.append(('', ''))
        
        notes_content.append(('Market Concentration (HHI)', ''))
        notes_content.append(('', 'Herfindahl-Hirschman Index (HHI): Market concentration measured using the sum of squared market shares.'))
        notes_content.append(('', 'HHI scale ranges from 0 to 10,000:'))
        notes_content.append(('', '  - HHI < 1,500: Low concentration (competitive market)'))
        notes_content.append(('', '  - HHI 1,500-2,500: Moderate concentration'))
        notes_content.append(('', '  - HHI > 2,500: High concentration'))
        notes_content.append(('', 'HHI is calculated using the total dollar amount of loans by all lenders in each year.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Calculations', ''))
        notes_content.append(('', 'Percent of loans to income tracts: (Loans in income category / Total loans) × 100'))
        notes_content.append(('', 'Percent of dollars to income tracts: (Dollars in income category / Total dollars) × 100'))
        notes_content.append(('', 'Amounts are shown in thousands of dollars unless otherwise noted.'))
        notes_content.append(('', 'Quartiles for map layers: Data divided into 4 equal groups (Q1 = lowest, Q4 = highest)'))
        notes_content.append(('', ''))
        
        notes_content.append(('Table Descriptions', ''))
        notes_content.append(('', 'Section 1: Geographic Overview - Map and summary statistics with treemap visualizations.'))
        notes_content.append(('', 'Section 2: County Summary - County-level aggregate data for all years (2020-2024).'))
        notes_content.append(('', '  - Number of Loans tab: Shows loan counts by category and year.'))
        notes_content.append(('', '  - Amount of Loans tab: Shows loan amounts by category and year.'))
        notes_content.append(('', 'Section 3: Comparison - Compares county 2024 data with state and national benchmarks.'))
        notes_content.append(('', '  - Number of Loans tab: Loan count comparisons.'))
        notes_content.append(('', '  - Amount of Loans tab: Loan amount comparisons.'))
        notes_content.append(('', 'Section 4: Top Lenders - Top lenders by number of loans for 2024 only.'))
        notes_content.append(('', '  - Number of Loans tab: Loan count metrics by lender.'))
        notes_content.append(('', '  - Amount of Loans tab: Loan amount metrics by lender.'))
        notes_content.append(('', 'Section 5: Market Concentration Trends - HHI by year from 2020 to 2024.'))
        notes_content.append(('', ''))
        
        notes_content.append(('AI-Generated Content Disclosure', ''))
        notes_content.append(('', 'This report includes AI-generated narrative analysis of small business lending data.'))
        notes_content.append(('', 'The AI-generated content provides contextual analysis and interpretation of data patterns.'))
        notes_content.append(('', 'All data and statistics are sourced from the databases described above.'))
        notes_content.append(('', 'AI Model: Narratives generated using Claude AI (Anthropic) and/or OpenAI GPT models.'))
        notes_content.append(('', 'Human Review: AI-generated content has been reviewed for accuracy and appropriateness.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Data Quality Assurance', ''))
        notes_content.append(('', 'All data used in this report has been downloaded, cleaned, and tested by NCRC Research staff.'))
        notes_content.append(('', 'Users can download the complete raw dataset using the Excel Export feature.'))
        notes_content.append(('', 'For official CRA data and documentation, visit the appropriate regulatory agency websites.'))
        notes_content.append(('', ''))
        
        # Write notes content to sheet
        for row_idx, (label, value) in enumerate(notes_content, start=1):
            notes_sheet.cell(row=row_idx, column=1, value=label)
            notes_sheet.cell(row=row_idx, column=2, value=value)
            if label and label.isupper() and not value:  # Headers
                notes_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            elif label and not value:  # Section headers
                notes_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Adjust column widths
        notes_sheet.column_dimensions['A'].width = 30
        notes_sheet.column_dimensions['B'].width = 80
        
        # Section 2: County Summary - Split into Number and Amount tabs
        if county_summary_table:
            county_df = pd.DataFrame(county_summary_table)
            
            # Separate into number and amount rows
            number_rows = []
            amount_rows = []
            
            for row in county_summary_table:
                var_name = row.get('Variable', row.get('variable', ''))
                if var_name:
                    if 'Total Loans' in var_name or ('Loans' in var_name and 'Amount' not in var_name):
                        number_rows.append(row)
                    elif 'Amount' in var_name or 'Total Loan Amount' in var_name:
                        amount_rows.append(row)
            
            # County Summary - Number of Loans
            if number_rows:
                number_df = pd.DataFrame(number_rows)
                number_df.to_excel(writer, sheet_name=sanitize_sheet_name('County Summary - Number'), index=False)
            
            # County Summary - Amount of Loans
            if amount_rows:
                amount_df = pd.DataFrame(amount_rows)
                amount_df.to_excel(writer, sheet_name=sanitize_sheet_name('County Summary - Amount'), index=False)
        
        # Section 3: Comparison Table - Split into Number and Amount tabs
        if comparison_table:
            comparison_df = pd.DataFrame(comparison_table)
            
            # Separate into number and amount rows
            number_rows = []
            amount_rows = []
            
            for row in comparison_table:
                metric = row.get('Metric', row.get('metric', ''))
                if metric:
                    if 'Amount' in metric or metric == 'Total Loan Amount':
                        amount_rows.append(row)
                    else:
                        number_rows.append(row)
            
            # Comparison - Number of Loans
            if number_rows:
                number_df = pd.DataFrame(number_rows)
                number_df.to_excel(writer, sheet_name=sanitize_sheet_name('Comparison - Number'), index=False)
            
            # Comparison - Amount of Loans
            if amount_rows:
                amount_df = pd.DataFrame(amount_rows)
                amount_df.to_excel(writer, sheet_name=sanitize_sheet_name('Comparison - Amount'), index=False)
        
        # Section 4: Top Lenders - Split into Number and Amount tabs
        if top_lenders_table:
            lenders_df = pd.DataFrame(top_lenders_table)
            
            # Determine which columns belong to number vs amount
            number_cols = ['Lender Name', 'Num Total', 'Num Under 100K', 'Num 100K 250K', 'Num 250K 1M', 
                          'Numsb Under 1M', 'Low Income %', 'Moderate Income %', 'Middle Income %', 
                          'Upper Income %', 'Unknown Income %']
            amount_cols = ['Lender Name', 'Amt Total (in $000s)', 'Amt Under 100K', 'Amt 100K 250K', 
                          'Amt 250K 1M', 'Amtsb Under 1M %', 'Low Income Amt %', 'Moderate Income Amt %', 
                          'Middle Income Amt %', 'Upper Income Amt %', 'Unknown Income Amt %']
            
            # Filter to available columns
            available_cols = lenders_df.columns.tolist()
            number_cols = [col for col in number_cols if col in available_cols]
            amount_cols = [col for col in amount_cols if col in available_cols]
            
            # Top Lenders - Number of Loans
            if number_cols:
                number_df = lenders_df[number_cols].copy()
                number_df.to_excel(writer, sheet_name=sanitize_sheet_name('Top Lenders - Number'), index=False)
            
            # Top Lenders - Amount of Loans
            if amount_cols:
                amount_df = lenders_df[amount_cols].copy()
                amount_df.to_excel(writer, sheet_name=sanitize_sheet_name('Top Lenders - Amount'), index=False)
        
        # Section 5: HHI by Year
        if hhi_by_year:
            hhi_df = pd.DataFrame(hhi_by_year)
            if not hhi_df.empty:
                # Reorder columns if they exist
                preferred_order = ['year', 'hhi_value', 'concentration_level']
                columns = [col for col in preferred_order if col in hhi_df.columns]
                columns.extend([col for col in hhi_df.columns if col not in columns])
                hhi_df = hhi_df[columns]
                hhi_df.to_excel(writer, sheet_name=sanitize_sheet_name('HHI by Year'), index=False)
        
        # Raw Data sheet (if available)
        if not aggregate_df.empty:
            aggregate_df.to_excel(writer, sheet_name=sanitize_sheet_name('Raw Data'), index=False)
        
        # Style the header rows for all data sheets
        header_fill = PatternFill(start_color='552D87', end_color='552D87', fill_type='solid')  # NCRC purple
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in writer.sheets:
            if sheet_name == 'Notes':
                continue  # Skip Notes sheet
            
            sheet = writer.sheets[sheet_name]
            
            # Style header row
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Style data rows with alternating colors
            even_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                for cell in row:
                    cell.border = thin_border
                    if row_idx % 2 == 0:  # Even rows
                        cell.fill = even_fill
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_str = str(cell.value)
                            if len(cell_str) > max_length:
                                max_length = len(cell_str)
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            sheet.freeze_panes = 'A2'
