"""
Excel generator for MergerMeter analysis output.

IMPORTANT: This uses the template-based approach to match the original merger report
format exactly, including all formatting, formulas, headers, and structure.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
import pandas as pd
from typing import Dict, Optional, List
from datetime import datetime
import sys
import os

# Import the original merger report Excel generator functions
# Add the merger report utils to the path
USE_ORIGINAL_FUNCTIONS = False  # Initialize to False

# Try to find merger report utils in multiple locations
MERGER_REPORT_UTILS_PATHS = []

# 1. Check environment variable first
merger_base_env = os.getenv('MERGER_REPORT_BASE')
if merger_base_env:
    MERGER_REPORT_UTILS_PATHS.append(Path(merger_base_env) / 'reports' / '_shared' / 'utils')

# 2. Try relative path from workspace root (for GitHub/main branch)
from .config import JUSTDATA_BASE
workspace_root = JUSTDATA_BASE.parent if JUSTDATA_BASE.name == '#JustData_Repo' else JUSTDATA_BASE.parent.parent
relative_merger_path = workspace_root / '1_Merger_Report' / 'reports' / '_shared' / 'utils'
if relative_merger_path.exists():
    MERGER_REPORT_UTILS_PATHS.append(relative_merger_path)

# 3. Fallback to absolute path (for local development)
absolute_merger_path = Path(r'C:\DREAM\1_Merger_Report\reports\_shared\utils')
if absolute_merger_path.exists():
    MERGER_REPORT_UTILS_PATHS.append(absolute_merger_path)

# Try each path until one works
for merger_utils_path in MERGER_REPORT_UTILS_PATHS:
    if merger_utils_path.exists():
        sys.path.insert(0, str(merger_utils_path.parent.parent))
        try:
            from reports._shared.utils.excel_generator import (
                create_goal_setting_excel as create_original_excel,
                populate_mortgage_data_sheet,
                populate_sb_data_sheet,
                populate_branch_data_sheet,
                populate_assessment_areas_sheet,
                populate_notes_sheet,
                update_header_rows
            )
            USE_ORIGINAL_FUNCTIONS = True
            break  # Successfully imported, no need to try other paths
        except (ImportError, Exception) as e:
            # Continue to next path
            continue

if not USE_ORIGINAL_FUNCTIONS:
    # This is expected in GitHub/main branch - the app will use its own Excel generation
    pass  # Will use simplified version without external template support

from .config import TEMPLATE_FILE, CLEAN_TEMPLATE_FILE


def _remove_assessment_area_column_e(excel_path: Path):
    """
    Post-process Excel file to remove columns A (State) and E from the Assessment Areas sheet.
    The original Excel generator adds these columns, so we remove them after generation.
    """
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            # Only process Assessment Areas sheet
            if 'Assessment' in sheet_name and 'Area' in sheet_name:
                ws = wb[sheet_name]
                
                # Remove column A (State) if it exists
                if ws.max_column >= 1:
                    header_a = str(ws.cell(1, 1).value or '').strip().lower()
                    if header_a in ['state', 'state code', 'state name']:
                        print(f"  Removing column A (State) from sheet '{sheet_name}'")
                        ws.delete_cols(1, 1)
                
                # Remove column E (or D if A was removed)
                # After removing A, E becomes D, so check both positions
                if ws.max_column >= 5:
                    # Check if column 5 exists (original E)
                    print(f"  Removing column E from sheet '{sheet_name}'")
                    ws.delete_cols(5, 1)
                elif ws.max_column >= 4:
                    # If only 4 columns exist, column 4 might be the one to remove
                    print(f"  Removing column D (was E) from sheet '{sheet_name}'")
                    ws.delete_cols(4, 1)
        
        wb.save(excel_path)
        print(f"  Post-processed Excel file: removed columns A and E from Assessment Areas sheet")
        
    except Exception as e:
        print(f"  Warning: Could not post-process Excel file to remove Assessment Areas columns: {e}")
        import traceback
        traceback.print_exc()


def create_merger_excel(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame] = None,
    bank_a_hmda_peer: Optional[pd.DataFrame] = None,
    bank_b_hmda_subject: Optional[pd.DataFrame] = None,
    bank_b_hmda_peer: Optional[pd.DataFrame] = None,
    bank_a_sb_subject: Optional[pd.DataFrame] = None,
    bank_a_sb_peer: Optional[pd.DataFrame] = None,
    bank_b_sb_subject: Optional[pd.DataFrame] = None,
    bank_b_sb_peer: Optional[pd.DataFrame] = None,
    bank_a_branch: Optional[pd.DataFrame] = None,
    bank_b_branch: Optional[pd.DataFrame] = None,
    hhi_data: Optional[pd.DataFrame] = None,
    assessment_areas: Optional[Dict] = None,
    metadata: Optional[Dict] = None
):
    """
    Create Excel workbook with merger analysis data using template-based approach.
    
    This matches the original merger report format exactly, including all formatting,
    formulas, headers, and structure.
    
    Args:
        output_path: Path to save Excel file
        bank_a_name: Acquirer bank name
        bank_b_name: Target bank name
        All data DataFrames for HMDA, SB, Branch, and HHI
        assessment_areas: Assessment area information
        metadata: Additional metadata (years, loan purpose, etc.)
    """
    
    # Determine template file to use
    template_file = None
    if TEMPLATE_FILE.exists():
        template_file = TEMPLATE_FILE
        print(f"Using template: {template_file}")
    elif CLEAN_TEMPLATE_FILE.exists():
        template_file = CLEAN_TEMPLATE_FILE
        print(f"Using clean template: {template_file}")
    else:
        print("WARNING: No template file found. Excel will not match original format exactly.")
        print(f"  Looked for: {TEMPLATE_FILE}")
        print(f"  And: {CLEAN_TEMPLATE_FILE}")
        # Fall back to creating from scratch (not ideal)
        _create_excel_from_scratch(
            output_path, bank_a_name, bank_b_name,
            bank_a_hmda_subject, bank_a_hmda_peer,
            bank_b_hmda_subject, bank_b_hmda_peer,
            bank_a_sb_subject, bank_a_sb_peer,
            bank_b_sb_subject, bank_b_sb_peer,
            bank_a_branch, bank_b_branch,
            hhi_data, assessment_areas, metadata
        )
        return
    
    # Use original Excel generator if available
    # Check the global variable (don't modify it in this function)
    use_original = USE_ORIGINAL_FUNCTIONS
    if use_original:
        try:
            # Convert years from metadata
            hmda_years = metadata.get('hmda_years', []) if metadata else []
            sb_years = metadata.get('sb_years', []) if metadata else []
            
            # Convert assessment areas format if needed
            aa_data = assessment_areas or {}
            
            # Call original function
            create_original_excel(
                output_path=output_path,
                bank_a_name=bank_a_name,
                bank_b_name=bank_b_name,
                assessment_areas_data=aa_data,
                mortgage_goals_data=None,  # Not used in MergerMeter
                bank_a_mortgage_data=bank_a_hmda_subject,
                bank_b_mortgage_data=bank_b_hmda_subject,
                bank_a_mortgage_peer_data=bank_a_hmda_peer,
                bank_b_mortgage_peer_data=bank_b_hmda_peer,
                bank_a_sb_data=bank_a_sb_subject,
                bank_b_sb_data=bank_b_sb_subject,
                bank_a_sb_peer_data=bank_a_sb_peer,
                bank_b_sb_peer_data=bank_b_sb_peer,
                bank_a_branch_data=bank_a_branch,
                bank_b_branch_data=bank_b_branch,
                years_hmda=hmda_years,
                years_sb=sb_years,
                template_file=template_file,
                bank_a_lei=metadata.get('acquirer_lei') if metadata else None,
                bank_b_lei=metadata.get('target_lei') if metadata else None,
                bank_a_rssd=metadata.get('acquirer_rssd') if metadata else None,
                bank_b_rssd=metadata.get('target_rssd') if metadata else None,
                bank_a_sb_id=metadata.get('acquirer_sb_id') if metadata else None,
                bank_b_sb_id=metadata.get('target_sb_id') if metadata else None
            )
            
            # Add HHI sheet to the generated workbook
            if hhi_data is not None and not hhi_data.empty:
                wb = load_workbook(output_path)
                _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
                wb.save(output_path)
                print("  Added HHI Analysis sheet")
            
            # Post-process: Remove column E from Assessment Areas sheet
            _remove_assessment_area_column_e(output_path)
            
            return
            
        except Exception as e:
            print(f"Error using original Excel generator: {e}")
            print("Falling back to simplified version")
            use_original = False  # Use local variable instead of modifying global
    
    # Fallback: Load template and populate manually
    if template_file and template_file.exists():
        _create_excel_from_template(
            output_path, template_file, bank_a_name, bank_b_name,
            bank_a_hmda_subject, bank_a_hmda_peer,
            bank_b_hmda_subject, bank_b_hmda_peer,
            bank_a_sb_subject, bank_a_sb_peer,
            bank_b_sb_subject, bank_b_sb_peer,
            bank_a_branch, bank_b_branch,
            hhi_data, assessment_areas, metadata
        )
    else:
        _create_excel_from_scratch(
            output_path, bank_a_name, bank_b_name,
            bank_a_hmda_subject, bank_a_hmda_peer,
            bank_b_hmda_subject, bank_b_hmda_peer,
            bank_a_sb_subject, bank_a_sb_peer,
            bank_b_sb_subject, bank_b_sb_peer,
            bank_a_branch, bank_b_branch,
            hhi_data, assessment_areas, metadata
        )


def _add_hhi_sheet(wb: Workbook, hhi_data: pd.DataFrame, bank_a_name: str, bank_b_name: str):
    """Add HHI Analysis sheet to workbook for counties where both banks have branches."""
    
    # Filter to only counties where both banks have branches
    if 'bank_a_deposits' in hhi_data.columns and 'bank_b_deposits' in hhi_data.columns:
        hhi_filtered = hhi_data[
            (hhi_data['bank_a_deposits'] > 0) & 
            (hhi_data['bank_b_deposits'] > 0)
        ].copy()
    else:
        # If columns don't exist, use all data
        hhi_filtered = hhi_data.copy()
    
    if hhi_filtered.empty:
        print("  No counties with overlapping branches for HHI analysis")
        return
    
    ws = wb.create_sheet("HHI Analysis", len(wb.sheetnames))
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Title
    row = 1
    ws.merge_cells(f'A{row}:{get_column_letter(len(hhi_filtered.columns))}{row}')
    cell = ws.cell(row, 1, f"HHI Analysis - Counties with {bank_a_name} and {bank_b_name} Branches")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")
    row += 2
    
    # Headers
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        cell = ws.cell(row, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row += 1
    
    # Data rows
    for df_row_idx, df_row in hhi_filtered.iterrows():
        for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
            value = df_row[col_name]
            cell = ws.cell(row, col_idx, value)
            
            # Format numbers
            if pd.api.types.is_numeric_dtype(hhi_filtered[col_name]):
                if 'hhi' in col_name.lower():
                    cell.number_format = '#,##0'  # HHI is typically an integer
                elif 'deposits' in col_name.lower() or 'share' in col_name.lower():
                    if 'share' in col_name.lower() or 'percent' in col_name.lower():
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
            
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Adjust column widths
    for col_idx, col_name in enumerate(hhi_filtered.columns, 1):
        col_letter = get_column_letter(col_idx)
        try:
            col_max_len = hhi_filtered[col_name].astype(str).str.len().max() if not hhi_filtered[col_name].empty else 10
        except (ValueError, AttributeError):
            col_max_len = 10
        max_length = max(len(str(col_name)), col_max_len)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def _remove_state_column(excel_path: Path):
    """
    Post-process Excel file to remove column A from sheets where it contains state data,
    and remove column E from the Assessment Areas sheet.
    """
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        # List of US state names and abbreviations for detection
        state_names = {
            'Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado',
            'Connecticut', 'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho',
            'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana',
            'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota',
            'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada',
            'New Hampshire', 'New Jersey', 'New Mexico', 'New York',
            'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon',
            'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota',
            'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington',
            'West Virginia', 'Wisconsin', 'Wyoming', 'District of Columbia'
        }
        state_abbrevs = {
            'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI',
            'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI',
            'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC',
            'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT',
            'VT', 'VA', 'WA', 'WV', 'WI', 'WY', 'DC'
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Check if column A contains state data
            # Look at header row (usually row 1 or 2) and a few data rows
            should_remove_col_a = False
            
            # Check header
            header_cell = ws.cell(1, 1)  # Row 1, Column A
            header_value = str(header_cell.value or '').strip()
            
            if header_value.lower() in ['state', 'state code', 'state name']:
                should_remove_col_a = True
            elif header_value.upper() in state_abbrevs:
                should_remove_col_a = True
            elif header_value in state_names:
                should_remove_col_a = True
            else:
                # Check a few data rows to see if they contain state names/abbrevs
                for row_idx in range(2, min(6, ws.max_row + 1)):  # Check rows 2-5
                    cell_value = str(ws.cell(row_idx, 1).value or '').strip()
                    if cell_value.upper() in state_abbrevs or cell_value in state_names:
                        should_remove_col_a = True
                        break
            
            # Remove column A if it contains state data
            if should_remove_col_a:
                print(f"  Removing column A from sheet '{sheet_name}' (contains state data)")
                ws.delete_cols(1, 1)  # Delete 1 column starting at column 1
            
            # For Assessment Areas sheet, also remove column E
            if 'Assessment' in sheet_name and 'Area' in sheet_name:
                # After removing column A, column E is now column D (if col A was removed) or still E
                # We need to check the current column structure
                current_col_e_idx = 5 if not should_remove_col_a else 4  # If A was removed, E becomes D
                current_col_letter = get_column_letter(current_col_e_idx)
                
                print(f"  Removing column {current_col_letter} (E) from sheet '{sheet_name}'")
                ws.delete_cols(current_col_e_idx, 1)
        
        wb.save(excel_path)
        print(f"  Post-processed Excel file: removed state columns")
        
    except Exception as e:
        print(f"  Warning: Could not post-process Excel file to remove state columns: {e}")
        import traceback
        traceback.print_exc()


def _create_excel_from_template(
    output_path: Path,
    template_file: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame],
    bank_a_hmda_peer: Optional[pd.DataFrame],
    bank_b_hmda_subject: Optional[pd.DataFrame],
    bank_b_hmda_peer: Optional[pd.DataFrame],
    bank_a_sb_subject: Optional[pd.DataFrame],
    bank_a_sb_peer: Optional[pd.DataFrame],
    bank_b_sb_subject: Optional[pd.DataFrame],
    bank_b_sb_peer: Optional[pd.DataFrame],
    bank_a_branch: Optional[pd.DataFrame],
    bank_b_branch: Optional[pd.DataFrame],
    hhi_data: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict],
    metadata: Optional[Dict]
):
    """Load template and populate with data (simplified version)."""
    print("Loading template and populating data...")
    wb = load_workbook(template_file, data_only=False)
    
    # Rename sheets to match bank names
    for sheet_name in list(wb.sheetnames):
        new_name = sheet_name
        if 'PNC Bank' in sheet_name:
            new_name = sheet_name.replace('PNC Bank', bank_a_name)
        elif 'FirstBank' in sheet_name or 'First Bank' in sheet_name:
            new_name = sheet_name.replace('FirstBank', bank_b_name).replace('First Bank', bank_b_name)
        elif 'Bank A' in sheet_name:
            new_name = sheet_name.replace('Bank A', bank_a_name)
        elif 'Bank B' in sheet_name:
            new_name = sheet_name.replace('Bank B', bank_b_name)
        
        if new_name != sheet_name:
            wb[sheet_name].title = new_name
    
    # TODO: Populate data into template sheets
    # This requires the full population functions from the original merger report
    
    # Add HHI sheet
    if hhi_data is not None and not hhi_data.empty:
        _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
    
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


def _create_excel_from_scratch(
    output_path: Path,
    bank_a_name: str,
    bank_b_name: str,
    bank_a_hmda_subject: Optional[pd.DataFrame],
    bank_a_hmda_peer: Optional[pd.DataFrame],
    bank_b_hmda_subject: Optional[pd.DataFrame],
    bank_b_hmda_peer: Optional[pd.DataFrame],
    bank_a_sb_subject: Optional[pd.DataFrame],
    bank_a_sb_peer: Optional[pd.DataFrame],
    bank_b_sb_subject: Optional[pd.DataFrame],
    bank_b_sb_peer: Optional[pd.DataFrame],
    bank_a_branch: Optional[pd.DataFrame],
    bank_b_branch: Optional[pd.DataFrame],
    hhi_data: Optional[pd.DataFrame],
    assessment_areas: Optional[Dict],
    metadata: Optional[Dict]
):
    """Create Excel from scratch (fallback - not ideal, won't match format exactly)."""
    print("WARNING: Creating Excel from scratch - format may not match original exactly")
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Create basic sheets (simplified)
    # Note: This won't match the original format exactly
    
    # Add HHI sheet
    if hhi_data is not None and not hhi_data.empty:
        _add_hhi_sheet(wb, hhi_data, bank_a_name, bank_b_name)
    
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
