#!/usr/bin/env python3
"""
MergerMeter Flask web application - Two-bank merger impact analyzer.
"""

from flask import render_template, request, jsonify, send_file, session, Response
import os
import tempfile
import zipfile
from datetime import datetime
import uuid
import threading
import time
import json
from typing import List, Dict

from shared.web.app_factory import create_app, register_standard_routes
from shared.utils.progress_tracker import get_progress, update_progress, create_progress_tracker
from .config import TEMPLATES_DIR, STATIC_DIR, OUTPUT_DIR, PROJECT_ID
from . import __version__


# Create the Flask app
app = create_app(
    'mergermeter',
    template_folder=TEMPLATES_DIR,
    static_folder=STATIC_DIR
)

# Set maximum file upload size to 10MB (plenty for JSON files)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB


def index():
    """Main page with the analysis form"""
    return render_template('analysis_template.html', version=__version__)


def report():
    """Report display page"""
    # Check if job_id is in URL parameters (fallback if session doesn't persist)
    job_id_from_url = request.args.get('job_id')
    if job_id_from_url and not session.get('job_id'):
        session['job_id'] = job_id_from_url
    return render_template('report_template.html', version=__version__)


def progress_handler(job_id):
    """Progress tracking endpoint using Server-Sent Events"""
    def event_stream():
        last_percent = -1
        while True:
            progress = get_progress(job_id)
            percent = progress.get("percent", 0)
            step = progress.get("step", "Starting...")
            done = progress.get("done", False)
            error = progress.get("error", None)
            if percent != last_percent or done or error:
                yield f"data: {{\"percent\": {percent}, \"step\": \"{step}\", \"done\": {str(done).lower()}, \"error\": {json.dumps(error) if error else 'null'}}}\n\n"
                last_percent = percent
            if done or error:
                break
            time.sleep(0.5)
    return Response(event_stream(), mimetype="text/event-stream")


def analyze():
    """Handle analysis request - returns immediately, runs analysis in background thread"""
    try:
        # Get job ID for progress tracking
        job_id = request.form.get('job_id') or str(uuid.uuid4())
        session['job_id'] = job_id
        
        # Capture all form data BEFORE starting thread (Flask request context is thread-local)
        form_data = {
            'acquirer_lei': request.form.get('acquirer_lei', '').strip(),
            'acquirer_rssd': request.form.get('acquirer_rssd', '').strip(),
            'acquirer_sb_id': request.form.get('acquirer_sb_id', '').strip(),
            'target_lei': request.form.get('target_lei', '').strip(),
            'target_rssd': request.form.get('target_rssd', '').strip(),
            'target_sb_id': request.form.get('target_sb_id', '').strip(),
            'acquirer_name': request.form.get('acquirer_name', 'Bank A').strip(),
            'target_name': request.form.get('target_name', 'Bank B').strip(),
            'acquirer_assessment_areas': request.form.get('acquirer_assessment_areas', '[]'),
            'target_assessment_areas': request.form.get('target_assessment_areas', '[]'),
            'loan_purpose': request.form.get('loan_purpose', ''),
            'hmda_years': request.form.get('hmda_years', '2020,2021,2022,2023,2024'),
            'sb_years': request.form.get('sb_years', '2019,2020,2021,2022,2023'),
            'action_taken': request.form.get('action_taken', '1'),
            'occupancy_type': request.form.get('occupancy_type', '1'),
            'total_units': request.form.get('total_units', '1-4'),
            'construction_method': request.form.get('construction_method', '1'),
            'not_reverse': request.form.get('not_reverse', '1')
        }
        
        # Initialize progress
        update_progress(job_id, {'percent': 0, 'step': 'Initializing analysis...', 'done': False, 'error': None})
        
        # Run analysis in background thread so server can respond to progress requests
        def run_analysis():
            try:
                _perform_analysis(job_id, form_data)
            except Exception as e:
                import traceback
                error_msg = str(e)
                traceback.print_exc()
                update_progress(job_id, {'percent': 0, 'step': 'Error occurred', 'done': True, 'error': error_msg})
        
        thread = threading.Thread(target=run_analysis, daemon=True)
        thread.start()
        
        # Return immediately with job_id
        return jsonify({
            'success': True,
            'job_id': job_id,
            'message': 'Analysis started'
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': error_msg
        }), 500


def _perform_analysis(job_id, form_data):
    """Perform the actual analysis (runs in background thread)"""
    try:
        # Create progress tracker
        progress_tracker = create_progress_tracker(job_id)
        update_progress(job_id, {'percent': 0, 'step': 'Initializing analysis...', 'done': False, 'error': None})
        
        # Get form data from captured dict (handle None values)
        acquirer_lei = (form_data.get('acquirer_lei') or '').strip()
        acquirer_rssd = (form_data.get('acquirer_rssd') or '').strip()
        acquirer_sb_id = (form_data.get('acquirer_sb_id') or '').strip()
        target_lei = (form_data.get('target_lei') or '').strip()
        target_rssd = (form_data.get('target_rssd') or '').strip()
        target_sb_id = (form_data.get('target_sb_id') or '').strip()
        
        # Get bank names (should already be loaded)
        acquirer_name = (form_data.get('acquirer_name') or 'Bank A').strip()
        target_name = (form_data.get('target_name') or 'Bank B').strip()
        
        # Get assessment areas (from form data)
        acquirer_aa_json = form_data.get('acquirer_assessment_areas') or '[]'
        target_aa_json = form_data.get('target_assessment_areas') or '[]'
        
        try:
            acquirer_aa_data = json.loads(acquirer_aa_json) if acquirer_aa_json and isinstance(acquirer_aa_json, str) else []
            target_aa_data = json.loads(target_aa_json) if target_aa_json and isinstance(target_aa_json, str) else []
        except (json.JSONDecodeError, TypeError, AttributeError) as e:
            print(f"Error parsing assessment area JSON: {e}")
            acquirer_aa_data = []
            target_aa_data = []
        
        update_progress(job_id, {'percent': 5, 'step': 'Parsing assessment areas and counties...', 'done': False, 'error': None})
        
        # Extract counties from assessment areas
        # Also handle MSA codes that might be in the counties list
        # IMPORTANT: Preserve assessment area names and ensure counties are properly extracted
        import re
        
        acquirer_counties = []
        acquirer_msa_codes = []
        acquirer_aa_with_names = []  # Keep track of assessment areas with their names
        
        for aa in acquirer_aa_data:
            if isinstance(aa, dict):
                # Try multiple possible keys for assessment area name
                aa_name = aa.get('cbsa_name') or aa.get('name') or aa.get('assessment_area') or aa.get('aa_name') or 'Unnamed Assessment Area'
                counties_list = aa.get('counties') or aa.get('county_list') or aa.get('county') or []
                
                # If counties is a string, try to split it
                if isinstance(counties_list, str):
                    counties_list = [c.strip() for c in counties_list.split(',') if c.strip()]
                
                # Process each county/MSA code in the list
                aa_counties = []
                aa_msa_codes = []
                
                for item in counties_list:
                    if not item:
                        continue
                    
                    # Handle new format: dict with state_code and county_code
                    if isinstance(item, dict):
                        # New format: {"state_code": "12", "county_code": "057"} or {"geoid5": "12057"}
                        aa_counties.append(item)
                        acquirer_counties.append(item)
                    else:
                        # Legacy format: string
                        item_str = str(item).strip()
                        if not item_str:
                            continue
                        
                        # Check if it's an MSA code pattern
                        msa_match = re.search(r'MSA\s+(\d+(?:\s*,\s*\d+)*)', item_str, re.IGNORECASE)
                        if msa_match:
                            msa_numbers = [code.strip() for code in msa_match.group(1).split(',')]
                            aa_msa_codes.extend(msa_numbers)
                            acquirer_msa_codes.extend(msa_numbers)
                        else:
                            aa_counties.append(item_str)
                            acquirer_counties.append(item_str)
                
                # Store assessment area with its counties for reference (even if empty, to track all AAs)
                acquirer_aa_with_names.append({
                    'name': aa_name,
                    'counties': aa_counties,
                    'msa_codes': aa_msa_codes
                })
        
        target_counties = []
        target_msa_codes = []
        target_aa_with_names = []  # Keep track of assessment areas with their names
        
        for aa in target_aa_data:
            if isinstance(aa, dict):
                # Try multiple possible keys for assessment area name
                aa_name = aa.get('cbsa_name') or aa.get('name') or aa.get('assessment_area') or aa.get('aa_name') or 'Unnamed Assessment Area'
                counties_list = aa.get('counties') or aa.get('county_list') or aa.get('county') or []
                
                # If counties is a string, try to split it
                if isinstance(counties_list, str):
                    counties_list = [c.strip() for c in counties_list.split(',') if c.strip()]
                
                # Process each county/MSA code in the list
                aa_counties = []
                aa_msa_codes = []
                
                for item in counties_list:
                    if not item:
                        continue
                    
                    # Handle new format: dict with state_code and county_code
                    if isinstance(item, dict):
                        # New format: {"state_code": "12", "county_code": "057"} or {"geoid5": "12057"}
                        aa_counties.append(item)
                        target_counties.append(item)
                    else:
                        # Legacy format: string
                        item_str = str(item).strip()
                        if not item_str:
                            continue
                        
                        # Check if it's an MSA code pattern
                        msa_match = re.search(r'MSA\s+(\d+(?:\s*,\s*\d+)*)', item_str, re.IGNORECASE)
                        if msa_match:
                            msa_numbers = [code.strip() for code in msa_match.group(1).split(',')]
                            aa_msa_codes.extend(msa_numbers)
                            target_msa_codes.extend(msa_numbers)
                        else:
                            aa_counties.append(item_str)
                            target_counties.append(item_str)
                
                # Store assessment area with its counties for reference (even if empty, to track all AAs)
                target_aa_with_names.append({
                    'name': aa_name,
                    'counties': aa_counties,
                    'msa_codes': aa_msa_codes
                })
        
        # Look up counties for MSA codes
        if acquirer_msa_codes:
            msa_counties_map = get_counties_by_msa_codes(list(set(acquirer_msa_codes)))
            for msa_data in msa_counties_map.values():
                counties_from_msa = msa_data.get('counties', [])
                acquirer_counties.extend(counties_from_msa)
                # Also add to the corresponding assessment areas
                for aa in acquirer_aa_with_names:
                    if aa['msa_codes']:
                        # Check if any of this AA's MSA codes match
                        for msa_code, msa_info in msa_counties_map.items():
                            if msa_code in aa['msa_codes']:
                                aa['counties'].extend(msa_info.get('counties', []))
        
        if target_msa_codes:
            msa_counties_map = get_counties_by_msa_codes(list(set(target_msa_codes)))
            for msa_data in msa_counties_map.values():
                counties_from_msa = msa_data.get('counties', [])
                target_counties.extend(counties_from_msa)
                # Also add to the corresponding assessment areas
                for aa in target_aa_with_names:
                    if aa['msa_codes']:
                        # Check if any of this AA's MSA codes match
                        for msa_code, msa_info in msa_counties_map.items():
                            if msa_code in aa['msa_codes']:
                                aa['counties'].extend(msa_info.get('counties', []))
        
        # Remove duplicates - handle both strings and dicts
        def deduplicate_counties(counties_list):
            """Remove duplicates from a list that may contain strings or dicts"""
            seen = set()
            unique = []
            for county in counties_list:
                if isinstance(county, dict):
                    # Convert dict to hashable tuple for deduplication
                    # Use state_code + county_code or geoid5 as key
                    if 'geoid5' in county:
                        key = ('geoid5', str(county['geoid5']).zfill(5))
                    elif 'state_code' in county and 'county_code' in county:
                        key = ('codes', str(county['state_code']).zfill(2), str(county['county_code']).zfill(3))
                    else:
                        # Fallback: use string representation
                        key = ('dict', str(sorted(county.items())))
                    
                    if key not in seen:
                        seen.add(key)
                        unique.append(county)
                else:
                    # String format - use directly
                    if county not in seen:
                        seen.add(county)
                        unique.append(county)
            return unique
        
        acquirer_counties = deduplicate_counties(acquirer_counties)
        target_counties = deduplicate_counties(target_counties)
        
        # Expand MSA names to counties (if user provided MSA names instead of county lists)
        from .county_mapper import detect_and_expand_msa_names
        
        update_progress(job_id, {'percent': 8, 'step': 'Checking for MSA names and expanding to counties...', 'done': False, 'error': None})
        
        # Create progress callback for MSA expansion
        def msa_progress_callback(progress_data):
            update_progress(job_id, progress_data)
        
        acquirer_counties = detect_and_expand_msa_names(acquirer_counties, progress_callback=msa_progress_callback)
        target_counties = detect_and_expand_msa_names(target_counties, progress_callback=msa_progress_callback)
        
        # Remove duplicates again after expansion
        acquirer_counties = deduplicate_counties(acquirer_counties)
        target_counties = deduplicate_counties(target_counties)
        
        # Get filters from form_data (not request.form - we're in a background thread)
        loan_purpose = form_data.get('loan_purpose') or ''
        hmda_years_str = form_data.get('hmda_years') or '2020,2021,2022,2023,2024'
        sb_years_str = form_data.get('sb_years') or '2019,2020,2021,2022,2023'
        
        # Get HMDA filter values (defaults for multi-select: comma-separated values)
        action_taken = form_data.get('action_taken', '1')
        occupancy_type = form_data.get('occupancy_type', '1')
        total_units = form_data.get('total_units', '1,2,3,4')  # Default to 1-4 units
        construction_method = form_data.get('construction_method', '1')
        not_reverse = form_data.get('not_reverse', '1')
        
        # Parse years (handle None/empty strings)
        if not isinstance(hmda_years_str, str):
            hmda_years_str = '2020,2021,2022,2023,2024'
        if not isinstance(sb_years_str, str):
            sb_years_str = '2019,2020,2021,2022,2023'
        
        hmda_years = [y.strip() for y in hmda_years_str.split(',') if y.strip()]
        sb_years = [y.strip() for y in sb_years_str.split(',') if y.strip()]
        
        update_progress(job_id, {'percent': 10, 'step': 'Mapping counties to GEOIDs...', 'done': False, 'error': None})
        
        # Map counties to GEOIDs
        from .county_mapper import map_counties_to_geoids
        
        acquirer_geoids, acquirer_unmapped = map_counties_to_geoids(acquirer_counties)
        target_geoids, target_unmapped = map_counties_to_geoids(target_counties)
        
        # Combine all GEOIDs for HHI calculation
        all_geoids = list(set(acquirer_geoids + target_geoids))
        
        if not acquirer_geoids and not target_geoids:
            return jsonify({
                'success': False,
                'error': 'No valid counties found in assessment areas. Please check your county names.'
            }), 400
        
        update_progress(job_id, {'percent': 15, 'step': 'Querying HMDA data for Bank A...', 'done': False, 'error': None})
        
        # Query HMDA data
        from .query_builders import (
            build_hmda_subject_query, build_hmda_peer_query,
            build_sb_subject_query, build_sb_peer_query,
            build_branch_query
        )
        from shared.utils.bigquery_client import get_bigquery_client, execute_query
        import pandas as pd
        
        client = get_bigquery_client(PROJECT_ID)
        
        # Bank A HMDA Subject
        bank_a_hmda_subject = pd.DataFrame()
        if acquirer_lei and acquirer_geoids:
            query = build_hmda_subject_query(
                acquirer_lei, acquirer_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_a_hmda_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 25, 'step': 'Querying HMDA peer data for Bank A...', 'done': False, 'error': None})
        
        # Bank A HMDA Peer
        bank_a_hmda_peer = pd.DataFrame()
        if acquirer_lei and acquirer_geoids:
            query = build_hmda_peer_query(
                acquirer_lei, acquirer_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_a_hmda_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 35, 'step': 'Querying HMDA data for Bank B...', 'done': False, 'error': None})
        
        # Bank B HMDA Subject
        bank_b_hmda_subject = pd.DataFrame()
        if target_lei and target_geoids:
            query = build_hmda_subject_query(
                target_lei, target_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_b_hmda_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 45, 'step': 'Querying HMDA peer data for Bank B...', 'done': False, 'error': None})
        
        # Bank B HMDA Peer
        bank_b_hmda_peer = pd.DataFrame()
        if target_lei and target_geoids:
            query = build_hmda_peer_query(
                target_lei, target_geoids, hmda_years, loan_purpose,
                action_taken, occupancy_type, total_units, construction_method, not_reverse
            )
            results = execute_query(client, query)
            if results:
                bank_b_hmda_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 55, 'step': 'Querying Small Business data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Small Business Subject
        bank_a_sb_subject = pd.DataFrame()
        if acquirer_sb_id and acquirer_geoids:
            query = build_sb_subject_query(acquirer_sb_id, acquirer_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_a_sb_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 65, 'step': 'Querying Small Business peer data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Small Business Peer
        bank_a_sb_peer = pd.DataFrame()
        if acquirer_sb_id and acquirer_geoids:
            query = build_sb_peer_query(acquirer_sb_id, acquirer_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_a_sb_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 75, 'step': 'Querying Small Business data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Small Business Subject
        bank_b_sb_subject = pd.DataFrame()
        if target_sb_id and target_geoids:
            query = build_sb_subject_query(target_sb_id, target_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_b_sb_subject = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 85, 'step': 'Querying Small Business peer data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Small Business Peer
        bank_b_sb_peer = pd.DataFrame()
        if target_sb_id and target_geoids:
            query = build_sb_peer_query(target_sb_id, target_geoids, sb_years)
            results = execute_query(client, query)
            if results:
                bank_b_sb_peer = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 87, 'step': 'Querying branch data for Bank A...', 'done': False, 'error': None})
        
        # Bank A Branch Data
        bank_a_branch = pd.DataFrame()
        if acquirer_rssd and acquirer_geoids:
            query = build_branch_query(acquirer_rssd, acquirer_geoids, year=2025)
            results = execute_query(client, query)
            if results:
                bank_a_branch = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 88, 'step': 'Querying branch data for Bank B...', 'done': False, 'error': None})
        
        # Bank B Branch Data
        bank_b_branch = pd.DataFrame()
        if target_rssd and target_geoids:
            query = build_branch_query(target_rssd, target_geoids, year=2025)
            results = execute_query(client, query)
            if results:
                bank_b_branch = pd.DataFrame(results)
        
        update_progress(job_id, {'percent': 90, 'step': 'Calculating HHI...', 'done': False, 'error': None})
        
        # Calculate HHI
        hhi_df = pd.DataFrame()
        if acquirer_rssd and target_rssd and all_geoids:
            from .hhi_calculator import calculate_hhi_by_county
            try:
                hhi_df = calculate_hhi_by_county(
                    county_geoids=all_geoids,
                    acquirer_rssd=acquirer_rssd,
                    target_rssd=target_rssd,
                    year=2025
                )
            except Exception as e:
                print(f"Error calculating HHI: {e}")
                import traceback
                traceback.print_exc()
        
        update_progress(job_id, {'percent': 95, 'step': 'Generating Excel report...', 'done': False, 'error': None})
        
        # Generate Excel file
        from .excel_generator import create_merger_excel
        
        excel_file = OUTPUT_DIR / f'merger_analysis_{job_id}.xlsx'
        excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        assessment_areas_dict = {
            'acquirer': {'counties': acquirer_counties},
            'target': {'counties': target_counties}
        }
        
        metadata = {
            'hmda_years': hmda_years,
            'sb_years': sb_years,
            'loan_purpose': loan_purpose,
            'acquirer_lei': acquirer_lei,
            'target_lei': target_lei,
            'acquirer_rssd': acquirer_rssd,
            'target_rssd': target_rssd,
            'acquirer_sb_id': acquirer_sb_id,
            'target_sb_id': target_sb_id
        }
        
        # Remove 'State' column from all DataFrames before passing to Excel generator
        def remove_state_column(df):
            """Remove 'State' column from DataFrame if it exists."""
            if df is not None and not df.empty:
                cols_to_drop = [col for col in df.columns if str(col).strip().lower() == 'state']
                if cols_to_drop:
                    df = df.drop(columns=cols_to_drop)
            return df
        
        bank_a_hmda_subject = remove_state_column(bank_a_hmda_subject)
        bank_a_hmda_peer = remove_state_column(bank_a_hmda_peer)
        bank_b_hmda_subject = remove_state_column(bank_b_hmda_subject)
        bank_b_hmda_peer = remove_state_column(bank_b_hmda_peer)
        bank_a_sb_subject = remove_state_column(bank_a_sb_subject)
        bank_a_sb_peer = remove_state_column(bank_a_sb_peer)
        bank_b_sb_subject = remove_state_column(bank_b_sb_subject)
        bank_b_sb_peer = remove_state_column(bank_b_sb_peer)
        bank_a_branch = remove_state_column(bank_a_branch)
        bank_b_branch = remove_state_column(bank_b_branch)
        hhi_df = remove_state_column(hhi_df)
        
        create_merger_excel(
            output_path=excel_file,
            bank_a_name=acquirer_name,
            bank_b_name=target_name,
            bank_a_hmda_subject=bank_a_hmda_subject,
            bank_a_hmda_peer=bank_a_hmda_peer,
            bank_b_hmda_subject=bank_b_hmda_subject,
            bank_b_hmda_peer=bank_b_hmda_peer,
            bank_a_sb_subject=bank_a_sb_subject,
            bank_a_sb_peer=bank_a_sb_peer,
            bank_b_sb_subject=bank_b_sb_subject,
            bank_b_sb_peer=bank_b_sb_peer,
            bank_a_branch=bank_a_branch,
            bank_b_branch=bank_b_branch,
            hhi_data=hhi_df,
            assessment_areas=assessment_areas_dict,
            metadata=metadata
        )
        
        # Save metadata
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        update_progress(job_id, {'percent': 100, 'step': 'Analysis complete!', 'done': True, 'error': None})
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        
        # Update progress with error
        update_progress(job_id, {'percent': 0, 'step': 'Error occurred', 'done': True, 'error': error_msg})


@app.route('/api/load-bank-names', methods=['POST'])
def load_bank_names():
    """Load bank names from identifiers (LEI, RSSD, or SB Respondent ID)
    
    Uses LEI number to look up bank name from BigQuery hmda.lenders18 table.
    """
    try:
        data = request.get_json()
        acquirer = data.get('acquirer', {})
        target = data.get('target', {})
        
        result = {
            'success': True,
            'acquirer_name': None,
            'target_name': None
        }
        
        # Get bank name from LEI number using BigQuery
        from shared.utils.bigquery_client import get_bigquery_client
        from .config import PROJECT_ID
        
        client = get_bigquery_client(PROJECT_ID)
        
        # Look up acquirer bank name using LEI (or use provided name from bulk import)
        if acquirer.get('name'):
            # Use provided name from bulk import
            result['acquirer_name'] = acquirer.get('name').strip()
        elif acquirer.get('lei'):
            acquirer_lei = acquirer.get('lei').strip()
            if len(acquirer_lei) == 20:
                acquirer_name = get_bank_name_from_lei(client, acquirer_lei)
                if acquirer_name:
                    result['acquirer_name'] = acquirer_name
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Could not find bank name for LEI: {acquirer_lei}. Please verify the LEI number is correct.'
                    }), 404
        
        # Look up target bank name using LEI (or use provided name from bulk import)
        if target.get('name'):
            # Use provided name from bulk import
            result['target_name'] = target.get('name').strip()
        elif target.get('lei'):
            target_lei = target.get('lei').strip()
            if len(target_lei) == 20:
                target_name = get_bank_name_from_lei(client, target_lei)
                if target_name:
                    result['target_name'] = target_name
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Could not find bank name for LEI: {target_lei}. Please verify the LEI number is correct.'
                    }), 404
        
        # Validate that at least one bank name was found
        if not result['acquirer_name'] and not result['target_name']:
            return jsonify({
                'success': False,
                'error': 'Please provide LEI numbers for both banks. Bank names are looked up using the LEI number.'
            }), 400
        
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def clean_bank_name(bank_name: str) -> str:
    """
    Clean bank name by removing suffixes and converting to uppercase.
    
    Args:
        bank_name: Raw bank name from database
    
    Returns:
        Cleaned bank name in ALL CAPS with suffixes removed
    """
    if not bank_name:
        return ""
    
    import re
    
    # Remove leading "THE" or "The"
    bank_name = re.sub(r'^THE\s+', '', bank_name, flags=re.IGNORECASE).strip()
    bank_name = re.sub(r'^The\s+', '', bank_name, flags=re.IGNORECASE).strip()
    
    # List of suffixes to remove (case-insensitive)
    # IMPORTANT: Do NOT remove "BANK" if it's part of the actual bank name (e.g., "PNC Bank", "First Bank")
    # Only remove it if it's clearly a suffix after other words (e.g., "Some Name Bank")
    suffixes = [
        r',?\s*NATIONAL\s+ASSOCIATION\s*$',
        r',?\s*National\s+Association\s*$',
        r',?\s*N\.?\s*A\.?\s*$',
        r',?\s*N\.A\.\s*$',
        r',?\s*NA\s*$',
        r',?\s*FEDERAL\s+SAVINGS\s+BANK\s*$',
        r',?\s*Federal\s+Savings\s+Bank\s*$',
        r',?\s*FSB\s*$',
        r',?\s*FEDERAL\s+CREDIT\s+UNION\s*$',
        r',?\s*Federal\s+Credit\s+Union\s*$',
        r',?\s*FCU\s*$',
        r',?\s*STATE\s+BANK\s*$',
        r',?\s*State\s+Bank\s*$',
        r',?\s*SAVINGS\s+BANK\s*$',
        r',?\s*Savings\s+Bank\s*$',
        r',?\s*SAVINGS\s+AND\s+LOAN\s*$',
        r',?\s*Savings\s+and\s+Loan\s*$',
        r',?\s*S&L\s*$',
        r',?\s*INC\.?\s*$',
        r',?\s*LLC\.?\s*$',
        r',?\s*CORPORATION\s*$',
        r',?\s*Corporation\s*$',
        r',?\s*CORP\.?\s*$',
        r',?\s*Corp\.?\s*$',
        r',?\s*COMPANY\s*$',
        r',?\s*Company\s*$',
        r',?\s*CO\.?\s*$',
        r',?\s*Co\.?\s*$',
        # DO NOT remove standalone "BANK" - it's often part of the actual name (e.g., "PNC Bank", "First Bank")
        # Only remove "BANK" if it appears after a comma (e.g., "Some Name, Bank")
        r',\s*BANK\s*$',
        r',?\s*THE\s*$',
        r',?\s*The\s*$',
    ]
    
    # Apply patterns repeatedly until no more matches
    changed = True
    iterations = 0
    max_iterations = 10  # Safety limit
    
    while changed and iterations < max_iterations:
        changed = False
        original_name = bank_name
        
        for pattern in suffixes:
            bank_name = re.sub(pattern, '', bank_name, flags=re.IGNORECASE).strip()
            if bank_name != original_name:
                changed = True
                break
        
        iterations += 1
    
    # Final cleanup: remove trailing commas, spaces, and periods
    bank_name = re.sub(r'[,.\s]+$', '', bank_name).strip()
    
    # Convert to ALL CAPS
    bank_name = bank_name.upper()
    
    return bank_name


def get_bank_name_from_lei(client, lei: str) -> str:
    """
    Get bank name from LEI number by querying BigQuery hmda.lenders18 table.
    Cleans the name by removing suffixes and converting to uppercase.
    
    Args:
        client: BigQuery client
        lei: Legal Entity Identifier (20 characters)
    
    Returns:
        Cleaned bank name in ALL CAPS with suffixes removed, or None if not found
    """
    try:
        from .config import PROJECT_ID
        
        query = f"""
        SELECT DISTINCT
            respondent_name
        FROM `{PROJECT_ID}.hmda.lenders18`
        WHERE lei = '{lei}'
        LIMIT 1
        """
        
        query_job = client.query(query)
        results = list(query_job.result())
        
        if results and results[0].respondent_name:
            raw_name = results[0].respondent_name.strip()
            # Clean the bank name
            return clean_bank_name(raw_name)
        
        return None
    except Exception as e:
        print(f"Error querying bank name for LEI {lei}: {e}")
        return None


@app.route('/api/generate-assessment-areas-from-branches', methods=['POST'])
def generate_assessment_areas_from_branches():
    """Generate assessment areas from branch locations for a bank"""
    try:
        data = request.get_json()
        rssd = data.get('rssd', '').strip()
        bank_type = data.get('bank_type', 'acquirer')  # 'acquirer' or 'target'
        year = int(data.get('year', 2025))
        group_by_cbsa = data.get('group_by_cbsa', True)
        min_branches = int(data.get('min_branches', 1))
        
        if not rssd:
            return jsonify({
                'success': False,
                'error': 'RSSD number is required to generate assessment areas from branches.'
            }), 400
        
        from .branch_assessment_area_generator import generate_assessment_areas_from_branches
        
        assessment_areas = generate_assessment_areas_from_branches(
            rssd=rssd,
            year=year,
            group_by_cbsa=group_by_cbsa,
            min_branches_per_county=min_branches
        )
        
        if not assessment_areas:
            return jsonify({
                'success': False,
                'error': f'No branches found for RSSD {rssd} in year {year}. Please verify the RSSD number and year.'
            }), 404
        
        return jsonify({
            'success': True,
            'assessment_areas': assessment_areas,
            'bank_type': bank_type,
            'count': len(assessment_areas)
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating assessment areas from branches: {error_details}")
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500


@app.route('/api/download-assessment-area-template', methods=['GET'])
def download_assessment_area_template():
    """Download CSV template for assessment areas"""
    import csv
    import io
    from flask import Response
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Assessment Area Name',
        'State Code',
        'County Code',
        'County Name',
        'State Name'
    ])
    
    # Write example rows
    examples = [
        ['Tampa-St. Petersburg-Clearwater FL', '12', '057', 'Hillsborough', 'Florida'],
        ['Tampa-St. Petersburg-Clearwater FL', '12', '103', 'Pinellas', 'Florida'],
        ['Tampa-St. Petersburg-Clearwater FL', '12', '101', 'Pasco', 'Florida'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '42', '101', 'Philadelphia', 'Pennsylvania'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '34', '007', 'Camden', 'New Jersey'],
        ['Philadelphia-Camden-Wilmington PA-NJ-DE-MD', '10', '003', 'New Castle', 'Delaware'],
    ]
    
    for row in examples:
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=assessment_area_template.csv'
        }
    )
    
    return response


@app.route('/api/download-bank-identifiers-template', methods=['GET'])
def download_bank_identifiers_template():
    """Download CSV template for bank identifiers (LEI, RSSD, ResID)"""
    import csv
    import io
    from flask import Response
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Bank Name',
        'LEI',
        'RSSD',
        'ResID'
    ])
    
    # Write example rows
    examples = [
        ['PNC BANK', '549300BJX7P13H14EN18', '451965', '123456789'],
        ['FIRSTBANK', '549300ABC123DEF456', '123456', '987654321'],
    ]
    
    for row in examples:
        writer.writerow(row)
    
    # Create response
    output.seek(0)
    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=bank_identifiers_template.csv'
        }
    )
    
    return response


@app.route('/api/upload-assessment-areas', methods=['POST'])
def upload_assessment_areas():
    """Upload and parse assessment area JSON or CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        bank_type = request.form.get('bank_type', 'acquirer')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check file extension - JSON or CSV allowed
        filename = file.filename.lower()
        is_json = filename.endswith('.json')
        is_csv = filename.endswith('.csv')
        
        if not is_json and not is_csv:
            return jsonify({'success': False, 'error': 'Only JSON or CSV files are supported. Please upload a JSON or CSV file with assessment areas.'}), 400
        
        # Check file size (26KB should be fine, but let's be explicit)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if file_size > 10 * 1024 * 1024:  # 10MB limit
            return jsonify({'success': False, 'error': f'File too large ({file_size / 1024:.1f}KB). Maximum size is 10MB.'}), 400
        
        # Parse file (JSON or CSV)
        assessment_areas = []
        
        if is_csv:
            # Parse CSV file
            try:
                import csv
                import io
                
                # Read CSV content
                try:
                    csv_content = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    file.seek(0)
                    csv_content = file.read().decode('utf-8-sig')  # Handle BOM
                
                if not csv_content or not csv_content.strip():
                    return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
                
                # Parse CSV
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                rows = list(csv_reader)
                
                if not rows:
                    return jsonify({'success': False, 'error': 'CSV file contains no data rows'}), 400
                
                # Group rows by Assessment Area Name
                aa_dict = {}
                for row in rows:
                    aa_name = row.get('Assessment Area Name', '').strip() or row.get('Assessment Area', '').strip() or row.get('CBSA Name', '').strip()
                    if not aa_name:
                        continue
                    
                    # Get state and county codes
                    state_code = row.get('State Code', '').strip() or row.get('State FIPS', '').strip()
                    county_code = row.get('County Code', '').strip() or row.get('County FIPS', '').strip()
                    
                    if state_code and county_code:
                        county_dict = {
                            'state_code': state_code.zfill(2),
                            'county_code': county_code.zfill(3)
                        }
                        
                        if aa_name not in aa_dict:
                            aa_dict[aa_name] = []
                        
                        # Check if this county is already in the list
                        county_exists = any(
                            c.get('state_code') == county_dict['state_code'] and 
                            c.get('county_code') == county_dict['county_code']
                            for c in aa_dict[aa_name]
                        )
                        
                        if not county_exists:
                            aa_dict[aa_name].append(county_dict)
                
                # Convert to assessment areas list
                for aa_name, counties in aa_dict.items():
                    if counties:
                        assessment_areas.append({
                            'cbsa_name': aa_name,
                            'counties': counties
                        })
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error parsing CSV file: {error_details}")
                return jsonify({'success': False, 'error': f'Error parsing CSV file: {str(e)}'}), 500
        
        elif is_json:
            # Parse JSON file
            try:
                # Try UTF-8 first, then fall back to other encodings
                try:
                    json_data = file.read().decode('utf-8')
                except UnicodeDecodeError:
                    file.seek(0)
                    json_data = file.read().decode('utf-8-sig')  # Handle BOM
                
                if not json_data or not json_data.strip():
                    return jsonify({'success': False, 'error': 'JSON file is empty'}), 400
                
                data = json.loads(json_data)
                
                # Handle different JSON structures
                if isinstance(data, list):
                    # List of assessment areas
                    for item in data:
                        if isinstance(item, dict):
                            cbsa_name = item.get('cbsa_name') or item.get('name') or item.get('assessment_area') or 'Unknown'
                            counties = item.get('counties') or item.get('county_list') or []
                            
                            # Support new format with state/county codes
                            # Format 1: List of county dicts with state_code and county_code
                            # Format 2: List of strings (legacy "County, State" format)
                            # Format 3: String that can be split
                            
                            if isinstance(counties, str):
                                counties = [c.strip() for c in counties.split(',') if c.strip()]
                            elif isinstance(counties, list):
                                # Check if it's a list of dicts with codes
                                processed_counties = []
                                for county_item in counties:
                                    if isinstance(county_item, dict):
                                        # New format: {"state_code": "12", "county_code": "057"} or {"geoid5": "12057"}
                                        processed_counties.append(county_item)
                                    elif isinstance(county_item, str):
                                        # Legacy format: "County, State"
                                        processed_counties.append(county_item)
                                counties = processed_counties
                            
                            assessment_areas.append({
                                'cbsa_name': cbsa_name,
                                'counties': counties
                            })
                elif isinstance(data, dict):
                    # Single assessment area or nested structure
                    if 'assessment_areas' in data:
                        # Nested structure
                        for aa in data['assessment_areas']:
                            cbsa_name = aa.get('cbsa_name') or aa.get('name') or 'Unknown'
                            counties = aa.get('counties') or aa.get('county_list') or []
                            
                            # Support new format with state/county codes
                            if isinstance(counties, str):
                                counties = [c.strip() for c in counties.split(',') if c.strip()]
                            elif isinstance(counties, list):
                                processed_counties = []
                                for county_item in counties:
                                    if isinstance(county_item, dict):
                                        processed_counties.append(county_item)
                                    elif isinstance(county_item, str):
                                        processed_counties.append(county_item)
                                counties = processed_counties
                            
                            assessment_areas.append({
                                'cbsa_name': cbsa_name,
                                'counties': counties
                            })
                    else:
                        # Single assessment area
                        cbsa_name = data.get('cbsa_name') or data.get('name') or 'Unknown'
                        counties = data.get('counties') or data.get('county_list') or []
                        
                        # Support new format with state/county codes
                        if isinstance(counties, str):
                            counties = [c.strip() for c in counties.split(',') if c.strip()]
                        elif isinstance(counties, list):
                            processed_counties = []
                            for county_item in counties:
                                if isinstance(county_item, dict):
                                    processed_counties.append(county_item)
                                elif isinstance(county_item, str):
                                    processed_counties.append(county_item)
                            counties = processed_counties
                        
                        assessment_areas.append({
                            'cbsa_name': cbsa_name,
                            'counties': counties
                        })
            except json.JSONDecodeError as e:
                error_msg = f'Invalid JSON format: {str(e)}'
                if hasattr(e, 'lineno') and hasattr(e, 'colno'):
                    error_msg += f' (Line {e.lineno}, Column {e.colno})'
                return jsonify({'success': False, 'error': error_msg}), 500
            except UnicodeDecodeError as e:
                return jsonify({'success': False, 'error': f'File encoding error: {str(e)}. Please ensure the file is UTF-8 encoded.'}), 500
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error parsing JSON file: {error_details}")
                return jsonify({'success': False, 'error': f'Error parsing JSON file: {str(e)}'}), 500
        
        if not assessment_areas:
            return jsonify({'success': False, 'error': 'No assessment areas found in JSON file. Please check the file format.'}), 400
        
        return jsonify({
            'success': True,
            'assessment_areas': assessment_areas,
            'bank_type': bank_type,
            'count': len(assessment_areas)
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in upload_assessment_areas: {error_details}")
        return jsonify({'success': False, 'error': f'Upload error: {str(e)}'}), 500


def get_counties_by_msa_codes(msa_codes: List[str]) -> Dict[str, List[str]]:
    """
    Look up counties for given MSA/CBSA codes from BigQuery.
    
    Args:
        msa_codes: List of MSA/CBSA codes (as strings, e.g., ['14500', '19740', '24540'])
        
    Returns:
        Dictionary mapping MSA code to list of county names in "County, State" format
    """
    if not msa_codes:
        return {}
    
    try:
        from shared.utils.bigquery_client import get_bigquery_client, execute_query
        
        # Format MSA codes for query
        msa_code_list = ', '.join([f"'{code}'" for code in msa_codes])
        
        query = f"""
        SELECT 
            CAST(cbsa_code AS STRING) as msa_code,
            cbsa as msa_name,
            county_state
        FROM `hdma1-242116.geo.cbsa_to_county`
        WHERE CAST(cbsa_code AS STRING) IN ({msa_code_list})
        ORDER BY msa_code, county_state
        """
        
        client = get_bigquery_client(PROJECT_ID)
        results = execute_query(client, query)
        
        # Group counties by MSA code
        msa_counties = {}
        for row in results:
            msa_code = str(row.get('msa_code', ''))
            county = row.get('county_state', '')
            msa_name = row.get('msa_name', f'MSA {msa_code}')
            
            if msa_code not in msa_counties:
                msa_counties[msa_code] = {
                    'name': msa_name,
                    'counties': []
                }
            
            if county and county not in msa_counties[msa_code]['counties']:
                msa_counties[msa_code]['counties'].append(county)
        
        return msa_counties
        
    except Exception as e:
        print(f"Error looking up MSA codes: {e}")
        import traceback
        traceback.print_exc()
        return {}


def parse_assessment_areas_from_text(text):
    """
    Parse assessment areas from text content.
    Handles structured formats with MMSAs, state abbreviations, and county lists.
    Also handles MSA numbers and MSA names, looking up counties from BigQuery.
    Returns a list of assessment area dictionaries.
    """
    import re
    from .county_mapper import get_counties_by_msa_name
    
    assessment_areas = []
    all_counties = set()  # Track all unique counties found
    
    # State abbreviation to full name mapping
    state_abbrev_map = {
        'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas', 'CA': 'California',
        'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware', 'FL': 'Florida', 'GA': 'Georgia',
        'HI': 'Hawaii', 'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa',
        'KS': 'Kansas', 'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
        'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi', 'MO': 'Missouri',
        'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada', 'NH': 'New Hampshire', 'NJ': 'New Jersey',
        'NM': 'New Mexico', 'NY': 'New York', 'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio',
        'OK': 'Oklahoma', 'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
        'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah', 'VT': 'Vermont',
        'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia', 'WI': 'Wisconsin', 'WY': 'Wyoming',
        'DC': 'District of Columbia'
    }
    
    # Pattern 1: State abbreviation followed by colon and county list
    # Example: "PA: Carbon, Lehigh, Northampton" or "NC: Gaston, Iredell, Mecklenburg, Union"
    state_county_pattern = r'([A-Z]{2}):\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*(?:,\s*[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)*)'
    
    # Pattern 2: County names with "County" suffix
    # Example: "Carbon County" or "Lehigh County"
    county_with_suffix_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+County'
    
    # Pattern 3: City names that might be counties (like "Baltimore City")
    city_county_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+City'
    
    # Pattern 4: MSA numbers (e.g., "MSA 14500, 19740, 24540" or "MSA 22660")
    msa_pattern = r'MSA\s+(\d+(?:\s*,\s*\d+)*)'
    
    # Split text into lines for better parsing
    lines = text.split('\n')
    current_cbsa = None
    msa_codes_to_lookup = []  # Collect MSA codes for batch lookup
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Skip header rows and metadata
        if any(skip in line for skip in ['Charter Number', 'Appendix', 'Type of Exam', 'Rating and Assessment']):
            continue
        
        # Look for MSA numbers (e.g., "MSA 14500, 19740, 24540")
        msa_match = re.search(msa_pattern, line, re.IGNORECASE)
        if msa_match:
            msa_numbers_str = msa_match.group(1)
            # Extract all MSA codes
            msa_codes = [code.strip() for code in msa_numbers_str.split(',')]
            msa_codes_to_lookup.extend(msa_codes)
            
            # Try to extract MSA name from the line (text before "MSA")
            msa_name_match = re.match(r'^([^MSA]+?)(?:\s+MSA)', line, re.IGNORECASE)
            if msa_name_match:
                current_cbsa = msa_name_match.group(1).strip()
            else:
                current_cbsa = f"MSA {', '.join(msa_codes)}"
        
        # Look for MMSA/CBSA names (lines that might be MSA names)
        # MSA names often contain hyphens and are followed by state abbreviations
        msa_name_match = re.match(r'^([A-Z][A-Za-z\s\-–—,]+?)(?:\s+\([0-9\-]+\))?(?:\s+MSA)?$', line)
        if msa_name_match and 'non-metro' not in line.lower() and not msa_match:
            # This might be an MSA name
            potential_msa = msa_name_match.group(1).strip()
            # Check if next lines contain counties
            current_cbsa = potential_msa
        
        # Parse state abbreviation: county list format
        # Example: "PA: Carbon, Lehigh, Northampton" or "NC: Gaston, Iredell"
        state_matches = re.finditer(state_county_pattern, line)
        for match in state_matches:
            state_abbrev = match.group(1)
            counties_str = match.group(2)
            
            # Get full state name
            state_name = state_abbrev_map.get(state_abbrev, state_abbrev)
            
            # Split counties by comma
            county_names = [c.strip() for c in counties_str.split(',')]
            
            for county_name in county_names:
                # Clean up county name (remove trailing periods, etc.)
                county_name = county_name.strip('.,;')
                
                # Check if it already has "County" suffix
                if not county_name.endswith(' County'):
                    # Check if it's a city (like "Baltimore City")
                    if ' City' in county_name:
                        full_county = f"{county_name}, {state_name}"
                    else:
                        full_county = f"{county_name} County, {state_name}"
                else:
                    full_county = f"{county_name}, {state_name}"
                
                all_counties.add(full_county)
                
                # Add to assessment area with current CBSA if available
                if current_cbsa:
                    # Find or create assessment area for this CBSA
                    aa = next((a for a in assessment_areas if a['cbsa_name'] == current_cbsa), None)
                    if not aa:
                        aa = {'cbsa_name': current_cbsa, 'counties': []}
                        assessment_areas.append(aa)
                    if full_county not in aa['counties']:
                        aa['counties'].append(full_county)
        
        # Also look for standalone county patterns with "County" suffix
        county_matches = re.finditer(county_with_suffix_pattern, line)
        for match in county_matches:
            county_name = match.group(1)
            # Try to infer state from context (look for state abbreviations nearby)
            # For now, we'll add it to a general list
            # This is a fallback for counties mentioned without state context
            full_county = f"{county_name} County"
            # We'll try to match these later with state context if available
    
    # Look up counties for MSA codes found
    if msa_codes_to_lookup:
        msa_counties_map = get_counties_by_msa_codes(list(set(msa_codes_to_lookup)))
        
        for msa_code, msa_data in msa_counties_map.items():
            msa_name = msa_data['name']
            counties = msa_data['counties']
            
            # Add to all_counties set
            for county in counties:
                all_counties.add(county)
            
            # Create assessment area for this MSA
            aa = {
                'cbsa_name': msa_name,
                'counties': sorted(counties)
            }
            assessment_areas.append(aa)
    
    # Check if any assessment area names are MSA names and expand them
    # If an assessment area has a name but no counties, try to look it up as an MSA name
    for aa in assessment_areas:
        if aa.get('cbsa_name') and (not aa.get('counties') or len(aa.get('counties', [])) == 0):
            # Try to look up the CBSA name as an MSA name
            counties, _ = get_counties_by_msa_name(aa['cbsa_name'])
            if counties:
                aa['counties'] = sorted(counties)
                for county in counties:
                    all_counties.add(county)
                print(f"  Expanded MSA name '{aa['cbsa_name']}' to {len(counties)} counties")
    
    # If we found counties but no structured assessment areas, create a general one
    if all_counties and not assessment_areas:
        # Group counties by state
        counties_by_state = {}
        for county in all_counties:
            if ', ' in county:
                parts = county.split(', ')
                if len(parts) == 2:
                    county_name, state_name = parts
                    if state_name not in counties_by_state:
                        counties_by_state[state_name] = []
                    counties_by_state[state_name].append(county)
        
        # Create assessment areas by state
        for state_name, counties in counties_by_state.items():
            assessment_areas.append({
                'cbsa_name': f'{state_name} Assessment Area',
                'counties': sorted(counties)
            })
    
    # If still no assessment areas but we have counties, create one general area
    if not assessment_areas and all_counties:
        assessment_areas.append({
            'cbsa_name': 'General Assessment Area',
            'counties': sorted(list(all_counties))
        })
    
    # Remove duplicates and sort
    for aa in assessment_areas:
        aa['counties'] = sorted(list(set(aa['counties'])))
    
    return assessment_areas


@app.route('/report-data')
def report_data():
    """Return report data from Excel file as JSON for web display"""
    try:
        job_id = request.args.get('job_id')
        if not job_id:
            return jsonify({'success': False, 'error': 'Job ID required'}), 400
        
        # Find the Excel file for this job
        excel_file = OUTPUT_DIR / f'merger_analysis_{job_id}.xlsx'
        if not excel_file.exists():
            return jsonify({'success': False, 'error': 'Report file not found. The analysis may not have completed yet.'}), 404
        
        # Read Excel file and convert each sheet to JSON
        import pandas as pd
        from shared.analysis.ai_provider import convert_numpy_types
        
        report_data = {}
        excel_file_obj = pd.ExcelFile(excel_file)
        
        # Use openpyxl to read Excel with data_only=True to get calculated formula values
        from openpyxl import load_workbook
        wb = load_workbook(excel_file, data_only=True)
        
        # Filter out template sheets that shouldn't appear in the web report
        # These are sheets from the Excel template that contain data from other banks
        excluded_sheet_patterns = [
            'sandy spring',
            'atlantic union',
            'firstbank sb goals',  # Template sheets
            'pnc bank sb goals',   # Template sheets
        ]
        
        def should_exclude_sheet(sheet_name):
            """Check if a sheet should be excluded from the web report"""
            sheet_lower = sheet_name.lower()
            for pattern in excluded_sheet_patterns:
                if pattern in sheet_lower:
                    return True
            return False
        
        # Filter sheet names to exclude template sheets
        valid_sheet_names = [s for s in excel_file_obj.sheet_names if not should_exclude_sheet(s)]
        
        if len(valid_sheet_names) < len(excel_file_obj.sheet_names):
            excluded = [s for s in excel_file_obj.sheet_names if should_exclude_sheet(s)]
            print(f"[DEBUG] Excluding {len(excluded)} template sheets from web report: {excluded}")
        
        for sheet_name in valid_sheet_names:
            # Read with pandas first
            df = pd.read_excel(excel_file_obj, sheet_name=sheet_name)
            
            # Remove "Unnamed" columns - filter out columns that contain "Unnamed" in the name
            unnamed_cols = [col for col in df.columns if 'unnamed' in str(col).lower()]
            if unnamed_cols:
                print(f"[DEBUG] Removing {len(unnamed_cols)} 'Unnamed' columns from sheet '{sheet_name}': {unnamed_cols}")
                df = df.drop(columns=unnamed_cols)
            
            # If we have the openpyxl workbook, try to get actual cell values for deposits column
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Find deposits column - check multiple possible names
                deposits_col_idx = None
                deposits_col_name = None
                for idx, col_name in enumerate(df.columns, 1):
                    col_str = str(col_name).lower()
                    if 'deposit' in col_str:
                        deposits_col_idx = idx
                        deposits_col_name = col_name
                        break
                
                # If deposits column found, read values from openpyxl
                if deposits_col_idx:
                    print(f"[DEBUG] Found deposits column '{deposits_col_name}' at index {deposits_col_idx} in sheet '{sheet_name}'")
                    deposits_values = []
                    for row_idx in range(2, min(ws.max_row + 1, len(df) + 2)):  # Start from row 2 (skip header)
                        cell = ws.cell(row=row_idx, column=deposits_col_idx)
                        cell_value = cell.value
                        
                        # Handle different value types - preserve formatted strings
                        if cell_value is None:
                            deposits_values.append(None)
                        elif isinstance(cell_value, (int, float)):
                            # If it's a number, keep it as is
                            deposits_values.append(cell_value)
                        elif isinstance(cell_value, str):
                            # If it's already a string (formatted), keep it
                            deposits_values.append(cell_value)
                        else:
                            # Try to convert to string first, then handle
                            str_value = str(cell_value)
                            # If it looks like a formatted currency, keep it as string
                            if '$' in str_value or 'M' in str_value:
                                deposits_values.append(str_value)
                            else:
                                # Try to convert to number
                                try:
                                    deposits_values.append(float(str_value.replace('$', '').replace('M', '').replace(',', '').strip()))
                                except:
                                    deposits_values.append(str_value)
                        
                        # Update DataFrame with the actual value (preserve formatted strings)
                        if row_idx - 2 < len(df):
                            df.iloc[row_idx - 2, deposits_col_idx - 1] = deposits_values[-1]
                    
                    print(f"[DEBUG] Deposits values from openpyxl (first 10): {deposits_values[:10]}")
            
            # Also check if pandas already read the deposits column correctly
            deposits_cols = [col for col in df.columns if 'deposit' in str(col).lower()]
            if deposits_cols:
                deposits_col = deposits_cols[0]
                print(f"[DEBUG] Deposits column '{deposits_col}' found in DataFrame")
                print(f"[DEBUG] Sample values (first 10): {df[deposits_col].head(10).tolist()}")
                print(f"[DEBUG] Data types: {df[deposits_col].dtype}")
                print(f"[DEBUG] Non-null count: {df[deposits_col].notna().sum()}")
            
            # Convert DataFrame to list of dictionaries
            # Replace NaN with None for JSON serialization, but preserve strings
            df = df.replace({pd.NA: None, float('nan'): None})
            
            # Ensure deposits column values are preserved as strings if they contain formatting
            for col in df.columns:
                if 'deposit' in str(col).lower():
                    # Convert to string to preserve formatting like "$123.4M"
                    # Handle None/NaN values properly
                    df[col] = df[col].apply(lambda x: str(x) if x is not None and pd.notna(x) else '')
                    # Clean up any 'nan' or 'None' strings that might have been created
                    df[col] = df[col].replace(['nan', 'None', 'NaN'], '')
            
            # Preserve exact column order from Excel
            # Read column order directly from openpyxl worksheet to ensure it matches Excel
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Get headers from first row of Excel (row 1)
                excel_headers = []
                if ws.max_row > 0:
                    for col_idx in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col_idx).value
                        if cell_value is not None:
                            header_str = str(cell_value).strip()
                            # Skip "Unnamed" headers
                            if 'unnamed' not in header_str.lower() and header_str:
                                excel_headers.append(header_str)
                        else:
                            # If Excel header is empty, try to get from DataFrame
                            if col_idx <= len(df.columns):
                                df_col_name = str(df.columns[col_idx - 1])
                                # Skip "Unnamed" headers
                                if 'unnamed' not in df_col_name.lower() and df_col_name:
                                    excel_headers.append(df_col_name)
                    
                    # Reorder DataFrame columns to match Excel order
                    if excel_headers and len(excel_headers) > 0:
                        # Match Excel headers to DataFrame columns (case-insensitive, handle whitespace)
                        ordered_columns = []
                        for excel_header in excel_headers:
                            # Find matching column in DataFrame
                            matching_col = None
                            excel_header_clean = excel_header.strip().lower()
                            for df_col in df.columns:
                                if str(df_col).strip().lower() == excel_header_clean:
                                    matching_col = df_col
                                    break
                            
                            if matching_col and matching_col not in ordered_columns:
                                ordered_columns.append(matching_col)
                        
                        # Add any remaining columns that weren't in Excel headers
                        for col in df.columns:
                            if col not in ordered_columns:
                                ordered_columns.append(col)
                        
                        # Reorder DataFrame
                        df = df[ordered_columns]
                        # Use Excel headers for display (preserve exact Excel header text)
                        df.columns = excel_headers[:len(df.columns)]
            
            # Final filter: remove any remaining "Unnamed" columns from headers
            final_headers = []
            final_data_indices = []
            for idx, header in enumerate(df.columns):
                header_str = str(header).strip()
                if 'unnamed' not in header_str.lower() and header_str:
                    final_headers.append(header_str)
                    final_data_indices.append(idx)
            
            # Filter DataFrame to only include non-Unnamed columns
            if final_data_indices:
                df = df.iloc[:, final_data_indices]
                df.columns = final_headers
            
            records = df.to_dict('records')
            # Convert numpy types
            records = convert_numpy_types(records)
            report_data[sheet_name] = {
                'headers': final_headers if final_headers else list(df.columns),  # Use filtered headers
                'data': records
            }
        
        # Define proper sheet order for MergerMeter reports
        # This ensures sheets appear in a logical order in the web report
        preferred_sheet_order = [
            'Assessment Areas',
            'Mortgage Goals',
            'Bank A Mortgage Subject',
            'Bank A Mortgage Peer',
            'Bank B Mortgage Subject',
            'Bank B Mortgage Peer',
            'Bank A Small Business Subject',
            'Bank A Small Business Peer',
            'Bank B Small Business Subject',
            'Bank B Small Business Peer',
            'Bank A Branch',
            'Bank B Branch',
            'HHI Analysis'
        ]
        
        # Reorder sheets according to preferred order
        ordered_report_data = {}
        ordered_sheet_names = []
        
        # First, add sheets in preferred order
        for sheet_name in preferred_sheet_order:
            # Check for variations (e.g., "PNC Bank Mortgage Subject" vs "Bank A Mortgage Subject")
            matching_sheet = None
            for actual_sheet_name in report_data.keys():
                # Check if sheet name contains key words from preferred name
                preferred_keywords = [word.lower() for word in sheet_name.split() if len(word) > 2]
                actual_keywords = [word.lower() for word in actual_sheet_name.split() if len(word) > 2]
                
                # Count matching keywords
                matches = sum(1 for kw in preferred_keywords if any(akw.startswith(kw) or kw.startswith(akw) for akw in actual_keywords))
                if matches >= 2:  # At least 2 keywords match
                    matching_sheet = actual_sheet_name
                    break
            
            if matching_sheet and matching_sheet in report_data:
                ordered_report_data[matching_sheet] = report_data[matching_sheet]
                ordered_sheet_names.append(matching_sheet)
        
        # Add any remaining sheets that weren't in the preferred order
        for sheet_name in report_data.keys():
            if sheet_name not in ordered_report_data:
                ordered_report_data[sheet_name] = report_data[sheet_name]
                ordered_sheet_names.append(sheet_name)
        
        # Try to load metadata if available
        metadata = {}
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        if metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
        
        return jsonify({
            'success': True,
            'report': ordered_report_data,
            'sheet_order': ordered_sheet_names,  # Include order for frontend
            'metadata': metadata
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def generate_filename(metadata, extension='.xlsx'):
    """Generate a filename for downloads with NCRC, MergerMeter, bank names, and timestamp"""
    import re
    from datetime import datetime
    
    # Get bank names from metadata
    bank_a_name = metadata.get('acquirer_name', 'BankA')
    bank_b_name = metadata.get('target_name', 'BankB')
    
    # Clean up names for filename (remove special characters, spaces become underscores, remove commas)
    def clean_name(name):
        # Remove commas
        name = name.replace(',', '')
        # Replace spaces and special characters with underscores
        name = re.sub(r'[^\w\s-]', '', name)
        name = re.sub(r'[\s-]+', '_', name)
        return name
    
    bank_a_clean = clean_name(bank_a_name)
    bank_b_clean = clean_name(bank_b_name)
    
    # Build filename: NCRC_MergerMeter_[BankA]_[BankB]_[timestamp]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'NCRC_MergerMeter_{bank_a_clean}_{bank_b_clean}_{timestamp}{extension}'
    
    # Clean up double underscores
    filename = re.sub(r'__+', '_', filename)
    
    return filename


def download():
    """Download the generated Excel file"""
    try:
        job_id = request.args.get('job_id') or session.get('job_id')
        if not job_id:
            return jsonify({'error': 'Job ID required'}), 400
        
        excel_file = OUTPUT_DIR / f'merger_analysis_{job_id}.xlsx'
        if not excel_file.exists():
            return jsonify({'error': 'Report file not found. The analysis may not have completed yet.'}), 404
        
        # Try to load metadata for filename generation
        metadata = {}
        metadata_file = OUTPUT_DIR / f'merger_metadata_{job_id}.json'
        if metadata_file.exists():
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
        
        # Generate filename with NCRC, MergerMeter, bank names, and timestamp
        filename = generate_filename(metadata, '.xlsx')
        
        return send_file(
            str(excel_file),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Register standard routes
register_standard_routes(
    app,
    index_handler=index,
    analyze_handler=analyze,
    progress_handler=progress_handler,
    download_handler=download,
    data_handler=None
)

# Register MergerMeter-specific routes
app.add_url_rule('/report', 'report', report, methods=['GET'])
# Note: /report-data is already registered via @app.route decorator above

