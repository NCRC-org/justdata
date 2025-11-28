#!/usr/bin/env python3
"""
Excel export functionality for DataExplorer reports.
Creates Excel files with multiple sheets including Methods, all tables, and data.
"""

import pandas as pd
import os
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    """
    Sanitize sheet name for Excel compatibility.
    
    Excel sheet names cannot contain: \ / ? * [ ]
    And must be <= 31 characters.
    """
    # Remove invalid characters
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncate if too long
    if len(name) > max_length:
        name = name[:max_length]
    
    return name


def save_dataexplorer_excel_report(
    tables: Dict[str, List[Dict[str, Any]]],
    filters: Dict[str, Any],
    output_path: str,
    report_type: str = 'area',  # 'area' or 'lender'
    metadata: Optional[Dict[str, Any]] = None
):
    """
    Save the DataExplorer report data to an Excel file with multiple sheets.
    
    Args:
        tables: Dictionary of table names to table data (list of dicts)
        filters: Dictionary of filters applied to the report
        output_path: Path to save the Excel file
        report_type: Type of report ('area' or 'lender')
        metadata: Optional metadata about the report
    """
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    # Get metadata
    data_type = filters.get('dataType', 'hmda')
    years = filters.get('years', [])
    years_str = f"{min(years)}-{max(years)}" if years else "N/A"
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Get workbook and remove default sheet if it exists
        workbook = writer.book
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create Methods sheet (first sheet) with comprehensive methods and definitions
        notes_sheet = workbook.create_sheet(sanitize_sheet_name('Methods'), 0)
        
        # Prepare notes content
        notes_content = []
        
        notes_content.append(('METHODS AND DEFINITIONS', ''))
        notes_content.append(('', ''))
        
        notes_content.append(('Data Sources', ''))
        if data_type == 'hmda':
            notes_content.append(('', 'This report uses Home Mortgage Disclosure Act (HMDA) data.'))
            notes_content.append(('', 'HMDA data compiled and maintained in NCRC\'s curated BigQuery databases.'))
        elif data_type == 'sb':
            notes_content.append(('', 'This report uses Community Reinvestment Act (CRA) Small Business Lending data.'))
            notes_content.append(('', 'Small Business lending data compiled and maintained in NCRC\'s curated BigQuery databases.'))
        elif data_type == 'branches':
            notes_content.append(('', 'This report uses FDIC Summary of Deposits (SOD) branch data.'))
            notes_content.append(('', 'Branch data compiled and maintained in NCRC\'s curated BigQuery databases.'))
        notes_content.append(('', 'U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates for demographic context.'))
        notes_content.append(('', 'Census data retrieved from geo.census table in BigQuery.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Report Configuration', ''))
        notes_content.append(('', f'Report Type: {report_type.title()} Analysis'))
        notes_content.append(('', f'Data Type: {data_type.upper()}'))
        notes_content.append(('', f'Years: {years_str}'))
        if metadata:
            if 'geography' in metadata:
                notes_content.append(('', f'Geography: {metadata["geography"]}'))
            if 'subject_lender' in metadata:
                notes_content.append(('', f'Subject Lender: {metadata["subject_lender"]}'))
        notes_content.append(('', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'))
        notes_content.append(('', ''))
        
        # Add filter details
        notes_content.append(('Applied Filters', ''))
        if data_type == 'hmda':
            hmda_filters = filters.get('hmdaFilters', {})
            notes_content.append(('', f'Loan Purpose: {", ".join(hmda_filters.get("loanPurpose", []))}'))
            notes_content.append(('', f'Action Taken: {", ".join(hmda_filters.get("actionTaken", []))}'))
            notes_content.append(('', f'Occupancy Type: {", ".join(hmda_filters.get("occupancyType", []))}'))
            notes_content.append(('', f'Total Units: {", ".join(hmda_filters.get("totalUnits", []))}'))
            notes_content.append(('', f'Construction Method: {", ".join(hmda_filters.get("constructionMethod", []))}'))
            notes_content.append(('', f'Exclude Reverse Mortgages: {hmda_filters.get("excludeReverseMortgages", False)}'))
        notes_content.append(('', ''))
        
        notes_content.append(('Data Limitations', ''))
        if data_type == 'hmda':
            notes_content.append(('', 'The data in this report reflects only mortgage lending activity reported by financial institutions subject to HMDA reporting requirements.'))
            notes_content.append(('', 'It does not include lending by institutions that are exempt from reporting.'))
        elif data_type == 'sb':
            notes_content.append(('', 'The data in this report reflects only lending activity reported by financial institutions subject to CRA reporting requirements.'))
            notes_content.append(('', 'It does not include lending by institutions that are exempt from reporting, nor does it include non-traditional lending sources.'))
            notes_content.append(('', 'Important Note: In 2020 and 2021, the Paycheck Protection Program (PPP) significantly impacted small business lending patterns.'))
        elif data_type == 'branches':
            notes_content.append(('', 'The data in this report reflects branch locations and deposits as reported to the FDIC.'))
            notes_content.append(('', 'Data may not reflect recent branch openings or closures.'))
        notes_content.append(('', ''))
        
        if data_type == 'hmda':
            notes_content.append(('HMDA Definitions', ''))
            notes_content.append(('', 'Action Taken Codes:'))
            notes_content.append(('', '  1 = Loan originated'))
            notes_content.append(('', '  2 = Application approved but not accepted'))
            notes_content.append(('', '  3 = Application denied'))
            notes_content.append(('', '  4 = Application withdrawn by applicant'))
            notes_content.append(('', '  5 = File closed for incompleteness'))
            notes_content.append(('', 'Applications (1-5): All loan applications regardless of outcome'))
            notes_content.append(('', 'Denials (3): Only denied applications'))
            notes_content.append(('', ''))
            notes_content.append(('', 'LMI (Low-to-Moderate Income): Census tracts with median family income ≤80% of area median income'))
            notes_content.append(('', 'MMCT (Majority Minority Census Tract): Census tracts with >50% minority population'))
            notes_content.append(('', ''))
        
        if data_type == 'sb':
            notes_content.append(('Small Business Definitions', ''))
            notes_content.append(('', 'Income Classifications:'))
            notes_content.append(('', '  Low Income: ≤50% of Area Median Income (AMI)'))
            notes_content.append(('', '  Moderate Income: 50-80% of AMI'))
            notes_content.append(('', '  Middle Income: 80-120% of AMI'))
            notes_content.append(('', '  Upper Income: >120% of AMI'))
            notes_content.append(('', '  LMI (Low-to-Moderate Income): ≤80% of AMI'))
            notes_content.append(('', ''))
            notes_content.append(('', 'Loan Size Categories:'))
            notes_content.append(('', '  Under $100K: Loans with original amount < $100,000'))
            notes_content.append(('', '  $100K-$250K: Loans with original amount between $100,000 and $250,000'))
            notes_content.append(('', '  $250K-$1M: Loans with original amount between $250,000 and $1,000,000'))
            notes_content.append(('', '  Under $1M: Small business loans under $1,000,000'))
            notes_content.append(('', ''))
        
        if report_type == 'lender':
            notes_content.append(('Peer Comparison', ''))
            notes_content.append(('', 'Peer lenders are identified based on similar lending volume (50%-200% of subject lender volume).'))
            notes_content.append(('', 'Peer statistics represent averages across all peer lenders in the selected geography.'))
            notes_content.append(('', ''))
        
        notes_content.append(('Calculations', ''))
        notes_content.append(('', 'Percentages: (Category value / Total value) × 100'))
        notes_content.append(('', 'Amounts are shown in dollars unless otherwise noted.'))
        notes_content.append(('', 'Year-over-year trends show changes in metrics across selected years.'))
        notes_content.append(('', ''))
        
        notes_content.append(('Table Descriptions', ''))
        for table_name in tables.keys():
            notes_content.append(('', f'{table_name}: See corresponding sheet for detailed data.'))
        notes_content.append(('', ''))
        
        # Write notes content to sheet
        for row_idx, (label, value) in enumerate(notes_content, start=1):
            notes_sheet.cell(row=row_idx, column=1, value=label)
            notes_sheet.cell(row=row_idx, column=2, value=value)
            if label and label.isupper() and not value:  # Headers
                notes_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            elif label and not value:  # Section headers
                notes_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Adjust column widths
        notes_sheet.column_dimensions['A'].width = 30
        notes_sheet.column_dimensions['B'].width = 80
        
        # Create a sheet for each table
        for table_name, table_data in tables.items():
            if table_data:
                df = pd.DataFrame(table_data)
                if not df.empty:
                    sheet_name = sanitize_sheet_name(table_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Style all data sheets
        header_fill = PatternFill(start_color='552D87', end_color='552D87', fill_type='solid')  # NCRC purple
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in writer.sheets:
            if sheet_name == 'Methods':
                continue  # Skip Methods sheet
            
            sheet = writer.sheets[sheet_name]
            
            # Style header row
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Style data rows with alternating colors
            even_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                for cell in row:
                    cell.border = thin_border
                    if row_idx % 2 == 0:  # Even rows
                        cell.fill = even_fill
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_str = str(cell.value)
                            if len(cell_str) > max_length:
                                max_length = len(cell_str)
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            sheet.freeze_panes = 'A2'

